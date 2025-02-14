from django.db.models import Subquery
from rest_framework.decorators import authentication_classes, permission_classes
import random
from random import randint
from django.utils import timezone
from rest_framework.parsers import MultiPartParser
from rest_framework.generics import GenericAPIView
from django.db.models import Avg
import smtplib
import uuid
import bisect 
import pytz
from Tasks.serializers import *
from Tasks.models import *
from Project.models import ProjectMaster
from Project.serializers import *
from django.contrib.auth.hashers import make_password
from .models import Permissions
from rest_framework import views
from django.core.exceptions import ObjectDoesNotExist
import json
from django.db import connection
from django.views.decorators.csrf import csrf_exempt
from rest_framework.authtoken.models import Token
from Users.serializers import UsersSerializer,AttendanceListSerializer,ApiSerializer,GetMenuItemSerializer,RegisterSerializer,GetUserSerializerleavemapping
from re import A
from rest_framework.serializers import Serializer
from CheckTrackPro.serializers import IntermediateGetTrackProResultSerializer,IntermediatePostTrackProResultSerializer
from CheckTrackPro.models import IntermediateTrackProResult
from typing import Mapping
from django.contrib.auth.models import Permission
from django.db.models.fields import files
from django.shortcuts import render, redirect, HttpResponse
from django.http import JsonResponse
from django.contrib.auth import authenticate, login as django_login, logout as django_logout
from django.contrib import messages
# from .forms import CreateUserForm
from django.contrib.auth.decorators import login_required
from django.conf import settings
from rest_framework import permissions
import string
import random
import secrets
from TrackProBackend.settings import EMAIL_HOST_USER
import psycopg2
from django.db.models import Q
from CompanyMaster.models import companyinfo
from rest_framework.decorators import api_view
from rest_framework.views import APIView
from Rules.models import Leaverule,AnnounceMent
from Rules.serializers import leaveruleserializer,annoucementserializer
from .serializers import UserSerializer,BasicInfoSerializer, passwordSerializer, userRegisterSerializer, userUpdateSerializer, RoleSerializer, attendanceserializer, holidaysSerializer,UserSecondarySerializer,leaveserializer,leaveapprovalserializer,leaveMappingserializer,warninglogserializer
from .serializers import DesignationSerializer, LocationSerializer, FinancialYearSerializer, MappingSerializer, GetMappingSerializer,ShiftMasterSerializer,EmployeeShiftDetailsSerializer,ShiftAllotmentSerializer
from .serializers import *
from rest_framework.generics import ListAPIView
from .models import Designation, AttendanceList,Role, Users, ErrorLog, UserToManager, MenuItem, Permissions, Profile,UserSecondaryInfo,Leave,leaveApproval,leaveMapping,adminAttendanceRequest,AttendanceRequest
from Department.models import Department
from Department.serializers import DepartmentSerializer
from .models import FinancialYear, Location, Role, Designation,ApiKey
from rest_framework.authentication import TokenAuthentication
from rest_framework.pagination import PageNumberPagination
from django.shortcuts import render
from django.contrib.auth import authenticate
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.status import (
    HTTP_400_BAD_REQUEST,
    HTTP_404_NOT_FOUND,
    HTTP_200_OK
)
from rest_framework import status
from .custom_functions import *
from django.contrib.auth.signals import user_logged_in
from .static_info import frontUrl,imageUrl
from django.core.mail import send_mail
from psycopg2.extras import RealDictCursor

# emails
from django.template import Context
from django.template.loader import render_to_string, get_template
from django.core.mail import EmailMessage
from django.db.models import Value
from django.db.models.functions import Concat
# from datetime import date, timedelta
from datetime import date,datetime,timedelta,time
import datetime
import arrow
from openpyxl.styles import Alignment,Border, Side,PatternFill,Font
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

from django.db.models import Sum, Value,DecimalField, F, Case, When, Value, IntegerField

import os
import psycopg2
import pandas as pd
import tablib
from tablib import Dataset
from django.utils import timezone

from .models import *
from dateutil.relativedelta import *
from .serializers import monthlydataSerializer
import xlsxwriter
import openpyxl
import calendar
from collections import defaultdict
from calendar import monthrange
from openpyxl import load_workbook
import environ
from django.db.models import Max
from math import ceil
from dateutil import relativedelta
# from cryptography.fernet import Fernet
from .templatetags.encryption_filters import decrypt_parameter,encrypt_parameter

env = environ.Env()
environ.Env.read_env()
from Tasks.views import sendfirebasenotification
from .sendmail import send_async_custom_template_email
from .static_info import adminemail,hremail
from operator import itemgetter
from itertools import chain
from django.db.models.functions import Coalesce
import openai
import requests

from Shift.models import Shiftswap
from Shift.serializers import CustomShiftswapSerializer



def apiheader(companycode):
    current_date =  date.today()
    apikeyobject = ApiKey.objects.filter(isActive = True , company_code = companycode).first()
    if apikeyobject is not None:
        if current_date > apikeyobject.expiry_date:
            return False
        else:
            return True
    return False

# ROLE START------------------------------------------------------------------------------------------



@api_view(['GET'])
def RoleListAPI(request, format=None):
    companycode = request.user.company_code
    # if 'Api-key' in request.headers.keys():
    #     apikeyheader = request.headers['Api-key']
    #     apikeyvalidation = apiheader(apikeyheader)
    #     if apikeyvalidation == True:
    if request.method == 'GET':
        role = Role.objects.filter(company_code=companycode,Active=True)
        serializer = RoleSerializer(role, many=True)
        RoleList=[]

        for i in serializer.data:
            users = Users.objects.filter(RoleID=i['id'],is_active=True,company_code=companycode)
            user_serializer=UserSerializer(users,many=True)
            if len(user_serializer.data) > 0:
                i['Used']=True
            else:
                i['Used']=False
            RoleList.append(i)
            
       
        return Response ({
                            "data": RoleList,
                            "response":{
                                "n" : 1,
                                "msg" : "Roles fetched successfully",
                                "status" : "Success"
                            }
                            })
    #     else:
    #         return Response ({
    #         "data": "",
    #         "response":{
    #             "n" : 0,
    #             "msg" : "Access Denied",
    #             "status" : "Failed"
    #         }
    #         })
    # else:
    #     return Response ({
    #     "data": "",
    #     "response":{
    #         "n" : 0,
    #         "msg" : "Please provide Api-key",
    #         "status" : "Failed"
    #     }
    #     })



@api_view(['POST'])
def addRole(request):
    # if 'Api-key' in request.headers.keys():
    #     apikeyheader = request.headers['Api-key']
    #     apikeyvalidation = apiheader(apikeyheader)
    #     if apikeyvalidation == True:
    user = request.user

    if user.is_admin == True:
        try:
            request_data = request.data.copy()
            role_name = request_data['RoleName'].strip()
            request_data['company_code'] = request.user.company_code
            request_data['Active'] = True
            request_data['RoleName'] = role_name
            serializer = RoleSerializer(data=request_data)
        except Exception as e:
            return Response({'n': 0, 'Msg': 'serializer not accepting data', 'Status': 'Error'})
        else:
            roleObject = Role.objects.filter(RoleName__in = [role_name.strip().capitalize(),role_name.strip(),role_name.title()],company_code = request_data['company_code']).first()
            if roleObject is not None:
                return Response({'n': 0, 'Msg': 'Role already exist', 'Status': 'Failed'})
            if serializer.is_valid():
                data = {}
                serializer.validated_data['CreatedBy'] = request.user
                u = serializer.save()
                return Response({'n': 1, 'Msg': 'Role added successfully', 'Status': 'Success'})
            else:
                return Response({'n': 0, 'Msg': serializer.errors, 'Status': 'Error'})

    else:
        return Response({'Error': 'User has no permission to create'})
    #     else:
    #         return Response ({
    #         "data": "",
    #         "response":{
    #             "n" : 0,
    #             "msg" : "Access Denied",
    #             "status" : "Failed"
    #         }
    #         })
    # else:
    #     return Response ({
    #     "data": "",
    #     "response":{
    #         "n" : 0,
    #         "msg" : "Please provide Api-key",
    #         "status" : "Failed"
    #     }
    #     })
    

# UPDATE ROLE----------------------------------------------------------------------------------------------------------


@api_view(['POST'])
def updateRole(request):
    # if 'Api-key' in request.headers.keys():
    #     apikeyheader = request.headers['Api-key']
    #     apikeyvalidation = apiheader(apikeyheader)
    #     if apikeyvalidation == True:
    data = {'n': '', 'Msg': '', 'Status': ''}
    try:
        request_data = request.data.copy()
        role_name = request_data['RoleName'].strip()
        request_data['company_code'] = request.user.company_code
        request_data['RoleName'] = role_name
        roleID = request.query_params.get('roleID')
        if roleID is None:
            data['n'] = 0
            data['Msg'] = 'role ID is none'
            data['Status'] = "Failed"
        else:
            try:
                object = Role.objects.filter(id=roleID,company_code = request_data['company_code']).first()
            except Exception as e:
                data['n'] = 0
                data['Msg'] = 'ROLE DOES NOT EXIST'
                data['Status'] = "Failed"
            else:
                roleObject = Role.objects.filter(RoleName__in = [role_name.strip().capitalize(),role_name.strip(),role_name.title()],company_code = request_data['company_code']).first()
                if object.RoleName != request_data['RoleName'] and roleObject is not None :
                    return Response({'n': 0, 'Msg': 'Role already exist', 'Status': 'Error'})

                serializer = RoleSerializer(object, data=request_data)
                if serializer.is_valid():
                    serializer.save()
                    return Response({'n': 1, 'Msg': 'Role updated successfully', 'Status': 'Success'})
                else:
                    return Response({'n': 0, 'Msg': serializer.errors, 'Status': 'Error'})
        return Response(data=data)

    except Exception as e:
        data['n'] = 0
        data['Msg'] = 'try method failed'
        data['Status'] = "Failed"
        return Response(data=data)
    #     else:
    #         return Response ({
    #         "data": "",
    #         "response":{
    #             "n" : 0,
    #             "msg" : "Access Denied",
    #             "status" : "Failed"
    #         }
    #         })
    # else:
    #     return Response ({
    #     "data": "",
    #     "response":{
    #         "n" : 0,
    #         "msg" : "Please provide Api-key",
    #         "status" : "Failed"
    #     }
    #     })

# get object in Role---------------------------------------------------------------------------------------------


@api_view(['GET'])
def getRole(request):
    if request.method == 'GET':
        id = request.query_params.get('roleID', None)
        if id is not None:
            i = Role.objects.filter(id=id,company_code=request.user.company_code)
            serializer = RoleSerializer(i, many=True)
            return JsonResponse(serializer.data, safe=False)
            



@api_view(['POST'])
@permission_classes((AllowAny,))
def delete_role(request):
    data={}
    id = request.data.get('id')
    roledata=Role.objects.filter(id=id).first()
    data['Active'] = False
    serializer = RoleSerializer(roledata,data=data,partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({
        "data": serializer.data,
        "response": {
            "n": 1,
            "msg": "deleted successfully",
            "status": "success"
            }
        })
    else:
        return Response({
            "data": serializer.errors,
            "response": {
                "n": 0,
                "msg": "failure",
                "status": "could nt delete"
            }
        })


# ROLE END----------------------------------------------------------------------------------------------------------

# USER MANAGER MAPPING START------------------------------------------------------------------------------------------


@api_view(['GET'])
def mappingListAPI(request, format=None):
    if request.method == 'GET':
        managerlist = []
        mapping = UserToManager.objects.filter(company_code = request.user.company_code)
        if mapping.exists():
            serializer = MappingSerializer(mapping, many=True)
            return Response({"data":serializer.data,"n":1,"Msg":"List fetched successfully","Status":"Success"})
        else:
            return Response({"data":'',"n":0,"Msg":"List not found","Status":"Failed"})




def last_day_of_month(any_day):
    import datetime

    next_month = any_day.replace(day=28) + datetime.timedelta(days=4)
    return next_month - datetime.timedelta(days=next_month.day)


@api_view(['GET'])
def ManagerUserListAPI(request):
    if request.method == 'GET':
        mapping = UserToManager.objects.filter(company_code = request.user.company_code).distinct('ManagerID')
        lst = []
        for m in mapping:
            lst.append({"ManagerID": m.ManagerID_id,
                        "ManagerIDStr": m.ManagerIDStr,
                        "FixedMapping": m.FixedMapping,
                        "Users": []})
        
        for m in UserToManager.objects.filter(company_code = request.user.company_code):
            for l in lst:
                if m.ManagerID_id == l["ManagerID"]:
                    l["Users"].append(
                        {"UserID": m.UserID_id, "UserIDStr": m.UserIDStr})
                    
        return Response({'data':lst,"n":1,"Msg":"List fetched successfully","Status":"Success"})
    

@api_view(['POST'])
def weeklypercentdata(request):
    request_data = request.data.copy()
    today = datetime.date.today()
    year = today.year
    perlist=[]
    labellist = []
    weeklypercentagelist = []
    year, week_num, day_of_week = my_date.isocalendar()
    currentweek = week_num
    week=currentweek

    task_week_obj=TaskMaster.objects.filter(AssignTo=request_data['userID']).order_by('Year','Week').distinct('Week','Year').reverse()[:4]
    if task_week_obj.count() != 0 :
        serializer = YearWeekSerializer(task_week_obj, many=True)
        
        for x in serializer.data:
            
            userobj = IntermediateTrackProResult.objects.filter(EmpID=request_data['userID'],Year=x['Year'],Week=x['Week']).first()
            data={}
            if userobj is not None:
                perlist.append(userobj.TrackProPercent)
                labellist.append("Week " + str(userobj.Week)+ ' ('+str(userobj.Year)+')')
                data['Week']=str(userobj.Week)
                data['Year']=str(userobj.Year)
                data['Percentage']=userobj.TrackProPercent
            else:
                perlist.append(0)
                labellist.append("Week " + str(x['Week'])+ ' ('+str(x['Year'])+')')
                data['Week']=str(x['Week'])
                data['Year']=str(x['Year'])
                data['Percentage']=0
            weeklypercentagelist.append(data)

    else:
        current_date = datetime.datetime.now() + timedelta(weeks=1)
        for i in range(4):
            data={}
            current_date -= timedelta(weeks=1)
            week_number = current_date.strftime("%V")
            year = current_date.year
            perlist.append(0)
            labellist.append("Week " + str(week_number)+ ' ('+str(year)+')')
            data['Week']=str(week_number)
            data['Year']=str(year)
            data['Percentage']=0
            weeklypercentagelist.append(data)

    perlist.reverse()
    labellist.reverse()

    context={
            'percentlist' : perlist,
            'labellist' : labellist,
            'weeklypercentagelist':weeklypercentagelist,
        }
    return Response({'data':context,"n":1,"Msg":"List fetched successfully","Status":"Success"})
   


def fdatelist(month,year):
    day = timedelta(days=1)
    date1 = date(year, month, 1)
    dates = []
    d = date1
    while d.month == month:
        dates.append(d.strftime('%Y-%m-%d'))
        d += day
    return dates

@api_view(['POST'])
@permission_classes((AllowAny,))
def insertattendance(request):
    if request.method == 'POST':
        data={}
        data['employeeId']=request.POST.get('employeeId')
        data['date']=request.POST.get('date')
        data['time']=request.POST.get('time')
        data['created_at']=request.POST.get('created_at')
        data['Week']=request.POST.get('Week')
        data['Month']=request.POST.get('Month')
        data['Year']=request.POST.get('Year')
        data['company_code']=request.user.company_code
        if request.POST.get('emailsent') == "False":
            data['emailsent']=False
            
        already_exist_obj=attendance.objects.filter(date=data['date'],employeeId=data['employeeId']).order_by('time').first()        
        if already_exist_obj is not None:
            serializer=attendanceserializer(already_exist_obj,data=data,partial=True)
            mesg='attendance updated'
        else:
            serializer=attendanceserializer(data=data,partial=True)
            mesg='attendance added'

        if serializer.is_valid():
            serializer.save()
            return Response({'n': 1, 'data':serializer.data ,'Msg': mesg, 'Status': 'succcess'})
        else:
            return Response({'n': 0, 'data':serializer.errors,'Msg': 'failed to add attendance', 'Status': 'Failed'})



@api_view(['POST'])
@permission_classes((AllowAny,))
def updateattendance(request):
    if request.method == 'POST':
        data={}
        data['employeeId']=request.POST.get('employeeId')
        data['date']=request.POST.get('date')
        data['time']=request.POST.get('time')
        if request.POST.get('emailsent') == "False":
            data['emailsent']=False
            
        already_exist_obj=attendance.objects.filter(date=data['date'],employeeId=data['employeeId']).order_by('time').first()        
        if already_exist_obj is not None:
            serializer=attendanceserializer(already_exist_obj,data=data,partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({'n': 1, 'data':serializer.data ,'Msg': 'attendance updated', 'Status': 'succcess'})
            else:
                return Response({'n': 0, 'data':serializer.errors,'Msg': 'failed to update attendance', 'Status': 'Failed'})
        else:
            return Response({'n': 0, 'data':[],'Msg': 'failed to update attendance', 'Status': 'Failed'})


@api_view(['POST'])
@permission_classes((AllowAny,))
def deleteattendance(request):
    if request.method == 'POST':
        data={}
        data['employeeId']=request.POST.get('employeeId')
        data['date']=request.POST.get('date')
        data['time']=request.POST.get('time')
        if request.POST.get('emailsent') == "False":
            data['emailsent']=False
            
        already_exist_obj=attendance.objects.filter(date=data['date'],employeeId=data['employeeId']).order_by('time').first()        
        if already_exist_obj is not None:
            already_exist_obj.delete()

    
            return Response({'n': 1, 'data':[] ,'Msg': 'attendance deleted', 'Status': 'succcess'})

        else:
            return Response({'n': 0, 'data':[],'Msg': 'failed to delete attendance', 'Status': 'Failed'})


@api_view(['GET'])
def attstatisticsapi(request):
    today = datetime.date.today()
    year = today.year
    month = today.month
    weekcount = len(calendar.monthcalendar(year,month))
    companycode = request.user.company_code
    userlist=[]
    weeklist=[]
    weekprcntlist=[]

    # datelist = fdatelist(month,year)

    usercount = Users.objects.filter(is_active=True,company_code=companycode).exclude(Q(employeeId__isnull=True)|Q(employeeId='None')|Q(employeeId=''))
    userser = UsersSerializeronlyattendance(usercount,many=True)
    userlist=list(userser.data)




    totalusercount = len(userlist)

    weeklencount=1
    for i in range(weeklencount, weekcount+1):
        weekdatelist=[]
        weekcalavg = 0
        
        dateobj = attendance.objects.filter(Week = i,Year=year,Month = month).distinct("date")
        weekdatesser  = AttendanceSerializerAttendanceWeekDate(dateobj,many=True)
        weekdatelist=list(weekdatesser.data)
    
        
        for w in  weekdatelist:
            presentcount = 0
            userobjs = attendance.objects.filter(employeeId__in=userlist,Week = i,Year=year,date=w,Month = month).count()
            presentcount=userobjs
            dayaverage  = presentcount/totalusercount
            weekcalavg += dayaverage
        
        
        if i%2 == 0:
            www = (weekcalavg/5)*100
            weekaverage = round(www,2)
        else:
            www2 = (weekcalavg/6)*100
            weekaverage = round(www2,2)
        

        weeklist.append("Week" + str(i))
        weekprcntlist.append(weekaverage)

    context={
        'weekList':weeklist,
        'weekpercentlist':weekprcntlist
    }
   
    return Response({'data':context,"n":1,"Msg":"Lists fetched successfully","Status":"Success"})



@api_view(['POST'])
def teamtrackerdataapi(request):
    if request.method == "POST":
        teamfilter = request.data.get('teamfilter')
        department_id = request.data.get('department_id')
        my_date = datetime.date.today()
        year, week_num, day_of_week = my_date.isocalendar()
        year = year
        week = week_num

        prevweek = week-1
        loguser = request.user.id
        emplist=[]
        empteamdatalist=[]

        companycode = request.user.company_code
        userobj = Users.objects.filter(id=loguser,company_code=companycode,is_active=True).first()
        if userobj is not None:
            userser = UserSerializer(userobj)
            department_allowance=False
            if userser.data['RoleID'].lower() in ['admin','co admin','manager','hr manager','core team']:
                department_allowance=True
                
            department = userser.data['DepartmentID']
            checked_pined_department=ManagerPinedDepartmentMaster.objects.filter(user_id=userser.data['id'],is_active=True,pined=True).first()
            if checked_pined_department is not None:
                department=[checked_pined_department.department_id]
            if userobj.id == 13:
                department=[9,29]
                
            if department_id is not None and department_id !='':
                department=[int(department_id)]
                
                
            if teamfilter == "All":
                empobjects = Users.objects.filter(DepartmentID__in = department,is_active=True,company_code=companycode).exclude(id=loguser).order_by('id')
                allempser =  UsersSerializer(empobjects,many=True)
                for t in  allempser.data:
                    emplist.append(t['id'])
            elif teamfilter == "Present":
                empobjects = Users.objects.filter(DepartmentID__in = department,is_active=True,company_code=companycode).exclude(id=loguser).order_by('id')
                presentempser = UsersSerializer(empobjects,many=True)
                for p in presentempser.data:
                    attendance_id = p['employeeId']
                    attexist = attendance.objects.filter(employeeId=str(attendance_id),date=my_date,company_code=companycode).order_by('time').first()
                    if attexist is not None:
                        emplist.append(p['id'])
            elif teamfilter == "On Leave":
                empobjects = Users.objects.filter(DepartmentID__in = department,is_active=True,company_code=companycode).exclude(id=loguser).order_by('id')
                onleaveempser = UsersSerializer(empobjects,many=True)
                for p in onleaveempser.data:
                    empid = p['id']
                    leaveobj = Leave.objects.filter(employeeId=str(empid),WorkFromHome=False,company_code=companycode,leave_status = "Approved")
                    leaveser = leaveserializer(leaveobj,many=True)
                    for l in leaveser.data:
                        leave_start_date = datetime.datetime.strptime(l['start_date'], '%Y-%m-%d')
                        leave_end_date = datetime.datetime.strptime(l['end_date'], '%Y-%m-%d')
                        check_date = datetime.datetime.strptime(str(my_date), '%Y-%m-%d')
                       
                        result = leave_start_date <= check_date <= leave_end_date
                        if result == True:
                            emplist.append(p['id'])
            else:
                empobjects = Users.objects.filter(DepartmentID__in = department,is_active=True,company_code=companycode).exclude(id=loguser).order_by('id')
                onleaveempser = UsersSerializer(empobjects,many=True)
                for p in onleaveempser.data:
                    empid = p['id']
                    leaveobj = Leave.objects.filter(employeeId=str(empid),WorkFromHome=True,company_code=companycode,leave_status = "Approved")
                    leaveser = leaveserializer(leaveobj,many=True)
                    for l in leaveser.data:
                        leave_start_date = datetime.datetime.strptime(l['start_date'], '%Y-%m-%d')
                        leave_end_date = datetime.datetime.strptime(l['end_date'], '%Y-%m-%d')
                        check_date = datetime.datetime.strptime(str(my_date), '%Y-%m-%d')

                        result = leave_start_date <= check_date <= leave_end_date
                        if result == True:
                            emplist.append(p['id'])
           
            for e in emplist:
                empdict = {}
                empdict['empid'] = e
                userempobj =  Users.objects.filter(id=int(e),company_code=companycode,is_active=True).first()
                if userempobj is not None:
                    taskobj = TaskMaster.objects.filter(AssignTo=e,Active=True,Status = 1,Year=year).first()
                    if taskobj is not None :
                        empdict['Current_task_project'] = taskobj.ProjectName
                        empdict['Current_task_details'] = taskobj.TaskTitle
                    else:
                        empdict['Current_task_project'] = "--"
                        empdict['Current_task_details'] = "--"

                    Taskslist = TaskMaster.objects.filter(AssignTo=e,Week = week,Year=year).order_by('id')
                    taskser =  GetTaskMasterSerializer(Taskslist,many=True)
                    for t in taskser.data:
                        splitdate = str(t['AssignDate'])
                        newdate = splitdate.split("-")[2]+"-"+splitdate.split("-")[1]+"-"+splitdate.split("-")[0]
                        t['notfdate'] = newdate

                    empdict['weektasklist'] = taskser.data

                    attendance_id =userempobj.employeeId
                    username = userempobj.Firstname + " "+ userempobj.Lastname
                    empdict['username'] = username

                    deptid = department
                    for d in deptid:
                        deptobj = Department.objects.filter(id=d,Active=True,company_code=companycode).first()
                        if deptobj is not None:

                            deptname = deptobj.DepartmentName
                        else:
                            deptname = '---'


                    empdict['empdeparment']=deptname
                    designationid = userempobj.DesignationId
                    empdes = Designation.objects.filter(DesignationName=designationid,isactive=True,company_code=companycode).first()
                    if empdes is not None:
                        empdict['empdesignation'] = empdes.DesignationName
                    else:
                        empdict['empdesignation'] = ''

                    trackproscore = IntermediateTrackProResult.objects.filter(Employee_id = int(e),Week = prevweek,Year=year,company_code=companycode).first()
                    if trackproscore is not None:
                        trackproper = trackproscore.TrackProPercent
                        empdict['prevtrackproscore'] = str(trackproper) +" %"
                    else:
                        empdict['prevtrackproscore'] = "--"
                    
                    
                    if attendance_id is not None:
                        get_attendance=get_checkin_checkout(attendance_id)
                        empdict['checkin'] = get_attendance['intime']
                        empdict['checkout'] = get_attendance['outtime']
                    else:
                        empdict['checkin'] = "--:--"
                        empdict['checkout'] = "--:--"



                    userphoto = userempobj.Photo
                    if userphoto is not None and userphoto != "":
                        empdict['userimage'] = imageUrl + "/media/" + str(userphoto)
                    else:
                        empdict['userimage'] = imageUrl + "/static/assets/images/profile.png"
            
                    empteamdatalist.append(empdict)
           
            departmentlist=[]

            departments_obj=Department.objects.filter(Active=True)
            dept_serializer = DepartmentSerializer(departments_obj, many=True)
            for dep in dept_serializer.data:
                if department[0]==dep['id']:
                    departmentlist.append({'department_id':dep['id'],'department_name':dep['DepartmentName'],'user_id':request.user.id,'is_active':True,'pined':True,})
                else:
                    departmentlist.append({'department_id':dep['id'],'department_name':dep['DepartmentName'],'user_id':request.user.id,'is_active':True,'pined':False,})


            return Response({'data':empteamdatalist,'departmentlist':departmentlist,"n":1,"Msg":"List fetched successfully","Status":"Success","department_allowance":department_allowance})
        else:
            return Response({'data':[],'departmentlist':[],"n":0,"Msg":"Employee not found","Status":"Failed","department_allowance":False})





def get_checkin_checkout(employeeId):

    
    current_date = datetime.date.today()
    yesterday_date = current_date - datetime.timedelta(days=1)
    current_datetime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    intime=''
    outtime=''
    intimedate=''
    outtimedate=''
    userobj = Users.objects.filter(employeeId = employeeId).first()
    
    if userobj is not None:
        check_user_type=EmployeeShiftDetails.objects.filter(employeeId=userobj.id,is_active=True).first()   
        
        if check_user_type is not None:
            check_last_checkin = attendance.objects.filter(employeeId=str(employeeId),date=str(current_date),checkout=False).order_by('time').last()
            check_last_checkout=None
            if check_last_checkin is not None:
                check_last_checkout = attendance.objects.filter(employeeId=str(employeeId),time__gt=check_last_checkin.time,date=str(current_date),checkout=True).order_by('time').last()
            else:
                check_last_checkin = attendance.objects.filter(employeeId=str(employeeId),date=str(yesterday_date),checkout=False).order_by('time').last()
                if check_last_checkin is not None:
                    check_last_checkout = attendance.objects.filter(Q(employeeId=str(employeeId),time__gt=check_last_checkin.time,date=str(yesterday_date),checkout=True)|Q(employeeId=str(employeeId),date=str(current_date),checkout=True)).order_by('date','time').last()


            
            if check_last_checkin is not None :
                
                intime = check_last_checkin.time
                intimedate=check_last_checkin.date
                punchout = 0
                get_data=1
                outtime = ''
                outtimedate=''
            else:
                get_data=0
                punchout = 1
                intime = ''
                intimedate=''
                outtime = ''
                outtimedate=''
               
               
            if check_last_checkout is not None:
                
                outtime = check_last_checkout.time
                outtimedate=check_last_checkout.date
                punchout = 1
                get_data=0
            else:
                if check_last_checkin is not None :
                    punchout = 0
                    get_data=1
                    outtime = ''
                    outtimedate=''
                else:
                    punchout = 1
                    get_data=0
                    outtime = ''
                    outtimedate=''
                    

            currentshiftname=''
            currentshiftstarttime=''
            currentshiftendtime=''
            currentshiftstartdate=''
            currentshiftenddate=''
            todays_shiftallotment_objs=ShiftAllotment.objects.filter(attendanceId=employeeId,date=str(current_date),is_active=True)
            todays_shiftallotment_serializers=ShiftAllotmentshiftIdSerializer(todays_shiftallotment_objs,many=True)
            shiftId_list=list(todays_shiftallotment_serializers.data)
            shift_obj=ShiftMaster.objects.filter(id__in=shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime')
            shift_serializer=ShiftMasterSerializer(shift_obj,many=True)
            
            
            # check for todays shift
            todays_runnningshift=gettodaysshift(shift_serializer.data,str(current_date))
            if todays_runnningshift['n'] == 1:
                currentshiftname=todays_runnningshift['data']['shiftname']
                currentshiftstarttime=str(todays_runnningshift['data']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                currentshiftendtime=str(todays_runnningshift['data']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                currentshiftstartdate=str(todays_runnningshift['data']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                currentshiftenddate=str(todays_runnningshift['data']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                
            elif todays_runnningshift['last_runingshift']['shiftstarttime'] !='':
                currentshiftname=todays_runnningshift['last_runingshift']['shiftname']
                currentshiftstarttime=str(todays_runnningshift['last_runingshift']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                currentshiftendtime=str(todays_runnningshift['last_runingshift']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                currentshiftstartdate=str(todays_runnningshift['last_runingshift']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                currentshiftenddate=str(todays_runnningshift['last_runingshift']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
            
            else:
                yesterday_shiftallotment_objs=ShiftAllotment.objects.filter(attendanceId=employeeId,date=str(yesterday_date),is_active=True)
                yesterday_shiftallotment_serializers=ShiftAllotmentshiftIdSerializer(yesterday_shiftallotment_objs,many=True)
                shiftId_list=list(yesterday_shiftallotment_serializers.data)
                shift_obj=ShiftMaster.objects.filter(id__in=shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime')
                yesterday_shift_serializer=ShiftMasterSerializer(shift_obj,many=True)
                    
                
                yesterday_runnningshift=gettodaysshift(yesterday_shift_serializer.data,str(yesterday_date))
                if yesterday_runnningshift['n'] == 1:
                    currentshiftname=yesterday_runnningshift['data']['shiftname']
                    currentshiftstarttime=str(yesterday_runnningshift['data']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                    currentshiftendtime=str(yesterday_runnningshift['data']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                    currentshiftstartdate=str(yesterday_runnningshift['data']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                    currentshiftenddate=str(yesterday_runnningshift['data']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                    
                elif yesterday_runnningshift['last_runingshift']['shiftstarttime'] !='':
                    currentshiftname=yesterday_runnningshift['last_runingshift']['shiftname']
                    currentshiftstarttime=str(yesterday_runnningshift['last_runingshift']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                    currentshiftendtime=str(yesterday_runnningshift['last_runingshift']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                    currentshiftstartdate=str(yesterday_runnningshift['last_runingshift']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                    currentshiftenddate=str(yesterday_runnningshift['last_runingshift']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                else:
                    currentshiftname='General'
                    currentshiftstarttime='07:30:00'
                    currentshiftendtime='18:30:00'
                    currentshiftstartdate=str(current_date)
                    currentshiftenddate=str(current_date)

            shiftdetails={
                        "shiftname":currentshiftname,
                        "shiftstarttime":currentshiftstarttime,
                        "shiftendtime":currentshiftendtime,
                        "shiftstartdate":currentshiftstartdate,
                        "shiftenddate":currentshiftenddate,
                    }
            

            getallattendance = attendance.objects.filter(Q(employeeId=str(employeeId),time__gte=currentshiftstarttime,date=str(currentshiftstartdate))|Q(employeeId=str(employeeId),time__lte=str(current_datetime).split(' ')[1],date=str(current_datetime).split(' ')[0])).order_by('date','time')
            
            # getallattendance = getallattendance.filter(time__gte=currentshiftstarttime).order_by('date','time')
            
            attendance_serializer=attendanceserializer(getallattendance,many=True)
            
            sorted_data = sorted(attendance_serializer.data, key=lambda x: (x['date'], x['time']))
            
            mindate = datetime.datetime.strptime(currentshiftstartdate, '%Y-%m-%d')
            mintime = datetime.datetime.strptime(currentshiftstarttime, '%H:%M:%S').time()

            sorted_data = [entry for entry in sorted_data if (datetime.datetime.strptime(entry['date'],'%Y-%m-%d').date() > mindate.date() or (datetime.datetime.strptime(entry['date'],'%Y-%m-%d').date() == mindate.date() and datetime.datetime.strptime(entry['time'], '%H:%M:%S').time() > mintime))]

            if len(sorted_data) > 0:
                intimedate=sorted_data[0]['date']
                intime=str(sorted_data[0]['time'])
                
            if intimedate !='' and intimedate is not None:
                user_sdt = datetime.datetime.strptime(str(intimedate) + ' ' + str(intime), '%Y-%m-%d %H:%M:%S')
                shif_sdt = datetime.datetime.strptime(str(currentshiftstartdate) + ' ' + str(currentshiftstarttime), '%Y-%m-%d %H:%M:%S')
                if user_sdt < shif_sdt :
                    intimedate=''
                    intime=''
                    outtime=''
                    outtimedate=''
                    punchout = 1
                    get_data=0
                    
            checkin_time = None
            total_working_time = 0
            for index, entry in enumerate(sorted_data):
                if entry['checkout']:
                    if checkin_time:
                        checkout_datetime = datetime.datetime.strptime(entry['date'] + ' ' + entry['time'], '%Y-%m-%d %H:%M:%S')
                        checkin_datetime = datetime.datetime.strptime(checkin_time, '%Y-%m-%d %H:%M:%S')
                        working_time = checkout_datetime - checkin_datetime
                        total_working_time += working_time.total_seconds()
                        checkin_time = None
                elif not entry['checkout']:
                    checkin_time = entry['date'] + ' ' + entry['time']

            # If the last entry is check-in, consider checkout time as current time
            if checkin_time and index == len(sorted_data) - 1:
                checkout_datetime = datetime.datetime.now()
                checkin_datetime = datetime.datetime.strptime(checkin_time, '%Y-%m-%d %H:%M:%S')
                working_time = checkout_datetime - checkin_datetime
                total_working_time += working_time.total_seconds()


            # Convert total_working_time to hours, minutes, and seconds
            hours, remainder = divmod(total_working_time, 3600)
            minutes, seconds = divmod(remainder, 60)

            return {
                'indatetime':str(dd_mm_yyyy(str(intimedate))) + ' ' + str(intime),
                'outdatetime':str(dd_mm_yyyy(str(outtimedate))) + ' ' + str(outtime),
                'total_hrs':str(add_leading_zero(int(hours))) +':' +str(add_leading_zero(int(minutes))) +':'+str(add_leading_zero(int(seconds))),
                'data':get_data,
                'intime':return__timestring__(intime),
                'outtime':return__timestring__(outtime),
                'intimedate':intimedate,
                'outtimedate':outtimedate,
                'punchout':punchout,
                'hours':int(hours),
                'minutes':int(minutes),
                'seconds':int(seconds),
                    "response":{
                        "n" : 1,
                        "msg" : "pass",
                        "status" : "success"
                    }
                    }
            
        else:
            
            check_last_checkin = attendance.objects.filter(employeeId=str(employeeId),date=str(current_date),checkout=False).order_by('time').last()
            check_last_checkout=None
            if check_last_checkin is not None:
                check_last_checkout = attendance.objects.filter(employeeId=str(employeeId),time__gt=check_last_checkin.time,date=str(current_date),checkout=True).order_by('time').last()

            else:
                check_last_checkin = attendance.objects.filter(employeeId=str(employeeId),date=str(yesterday_date),checkout=False).order_by('time').last()
                if check_last_checkin is not None:
                    check_last_checkout = attendance.objects.filter(Q(employeeId=str(employeeId),time__gt=check_last_checkin.time,date=str(yesterday_date),checkout=True)|Q(employeeId=str(employeeId),date=str(current_date),checkout=True)).order_by('date','time').last()

            
            # 1 disable
            # 0 enable
            
            if check_last_checkin is not None :
                
                intime = check_last_checkin.time
                intimedate=check_last_checkin.date
                punchout = 0
                get_data=1
                outtime = ''
                outtimedate=''
            else:
                get_data=0
                punchout = 1
                intime = ''
                intimedate=''
                outtime = ''
                outtimedate=''
               
               
            if check_last_checkout is not None:
                outtime = check_last_checkout.time
                outtimedate=check_last_checkout.date
                punchout = 1
                get_data=0

            else:
                if check_last_checkin is not None :
                    punchout = 0
                    get_data=1
                
                    outtime = ''
                    outtimedate=''
                else:
                    punchout = 1
                    get_data=0
                    outtime = ''
                    outtimedate=''
                    
            


            currentshiftname=''
            currentshiftstarttime=''
            currentshiftendtime=''
            currentshiftstartdate=''
            currentshiftenddate=''
            todays_shiftallotment_objs=ShiftAllotment.objects.filter(attendanceId=employeeId,date=str(current_date),is_active=True)
            todays_shiftallotment_serializers=ShiftAllotmentshiftIdSerializer(todays_shiftallotment_objs,many=True)
            shiftId_list=list(todays_shiftallotment_serializers.data)
            shift_obj=ShiftMaster.objects.filter(id__in=shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime')
            shift_serializer=ShiftMasterSerializer(shift_obj,many=True)
            
            
            # check for todays shift
            todays_runnningshift=gettodaysshift(shift_serializer.data,str(current_date))
            if todays_runnningshift['n'] == 1:
                currentshiftname=todays_runnningshift['data']['shiftname']
                currentshiftstarttime=str(todays_runnningshift['data']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                currentshiftendtime=str(todays_runnningshift['data']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                currentshiftstartdate=str(todays_runnningshift['data']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                currentshiftenddate=str(todays_runnningshift['data']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                
            elif todays_runnningshift['last_runingshift']['shiftstarttime'] !='':
                currentshiftname=todays_runnningshift['last_runingshift']['shiftname']
                currentshiftstarttime=str(todays_runnningshift['last_runingshift']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                currentshiftendtime=str(todays_runnningshift['last_runingshift']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                currentshiftstartdate=str(todays_runnningshift['last_runingshift']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                currentshiftenddate=str(todays_runnningshift['last_runingshift']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
            
            else:
                yesterday_shiftallotment_objs=ShiftAllotment.objects.filter(attendanceId=employeeId,date=str(yesterday_date),is_active=True)
                yesterday_shiftallotment_serializers=ShiftAllotmentshiftIdSerializer(yesterday_shiftallotment_objs,many=True)
                shiftId_list=list(yesterday_shiftallotment_serializers.data)
                shift_obj=ShiftMaster.objects.filter(id__in=shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime')
                yesterday_shift_serializer=ShiftMasterSerializer(shift_obj,many=True)
                    
                
                yesterday_runnningshift=gettodaysshift(yesterday_shift_serializer.data,str(yesterday_date))
                if yesterday_runnningshift['n'] == 1:
                    currentshiftname=yesterday_runnningshift['data']['shiftname']
                    currentshiftstarttime=str(yesterday_runnningshift['data']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                    currentshiftendtime=str(yesterday_runnningshift['data']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                    currentshiftstartdate=str(yesterday_runnningshift['data']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                    currentshiftenddate=str(yesterday_runnningshift['data']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                    
                elif yesterday_runnningshift['last_runingshift']['shiftstarttime'] !='':
                    currentshiftname=yesterday_runnningshift['last_runingshift']['shiftname']
                    currentshiftstarttime=str(yesterday_runnningshift['last_runingshift']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                    currentshiftendtime=str(yesterday_runnningshift['last_runingshift']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                    currentshiftstartdate=str(yesterday_runnningshift['last_runingshift']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                    currentshiftenddate=str(yesterday_runnningshift['last_runingshift']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                else:
                    currentshiftname='General'
                    currentshiftstarttime='07:30:00'
                    currentshiftendtime='18:30:00'
                    currentshiftstartdate=str(current_date)
                    currentshiftenddate=str(current_date)

            shiftdetails={
                        "shiftname":currentshiftname,
                        "shiftstarttime":currentshiftstarttime,
                        "shiftendtime":currentshiftendtime,
                        "shiftstartdate":currentshiftstartdate,
                        "shiftenddate":currentshiftenddate,
                    }
            

            getallattendance = attendance.objects.filter(Q(employeeId=str(employeeId),time__gte=currentshiftstarttime,date=str(currentshiftstartdate))|Q(employeeId=str(employeeId),time__lte=str(current_datetime).split(' ')[1],date=str(current_datetime).split(' ')[0])).order_by('date','time')
            
            # getallattendance = getallattendance.filter(time__gte=currentshiftstarttime).order_by('date','time')
            
            attendance_serializer=attendanceserializer(getallattendance,many=True)
            
            sorted_data = sorted(attendance_serializer.data, key=lambda x: (x['date'], x['time']))
            
            mindate = datetime.datetime.strptime(currentshiftstartdate, '%Y-%m-%d')
            mintime = datetime.datetime.strptime(currentshiftstarttime, '%H:%M:%S').time()

            sorted_data = [entry for entry in sorted_data if (datetime.datetime.strptime(entry['date'],'%Y-%m-%d').date() > mindate.date() or (datetime.datetime.strptime(entry['date'],'%Y-%m-%d').date() == mindate.date() and datetime.datetime.strptime(entry['time'], '%H:%M:%S').time() > mintime))]

            if len(sorted_data) > 0:
                intimedate=sorted_data[0]['date']
                intime=str(sorted_data[0]['time'])
                
            if intimedate !='' and intimedate is not None:
                user_sdt = datetime.datetime.strptime(str(intimedate) + ' ' + str(intime), '%Y-%m-%d %H:%M:%S')
                shif_sdt = datetime.datetime.strptime(str(currentshiftstartdate) + ' ' + str(currentshiftstarttime), '%Y-%m-%d %H:%M:%S')
                if user_sdt < shif_sdt :
                    intimedate=''
                    intime=''
                    outtime=''
                    outtimedate=''
                    punchout = 1
                    get_data=0
                    
            checkin_time = None
            total_working_time = 0
            for index, entry in enumerate(sorted_data):
                if entry['checkout']:
                    if checkin_time:
                        checkout_datetime = datetime.datetime.strptime(entry['date'] + ' ' + entry['time'], '%Y-%m-%d %H:%M:%S')
                        checkin_datetime = datetime.datetime.strptime(checkin_time, '%Y-%m-%d %H:%M:%S')
                        working_time = checkout_datetime - checkin_datetime
                        total_working_time += working_time.total_seconds()
                        checkin_time = None
                elif not entry['checkout']:
                    checkin_time = entry['date'] + ' ' + entry['time']

            # If the last entry is check-in, consider checkout time as current time
            if checkin_time and index == len(sorted_data) - 1:
                checkout_datetime = datetime.datetime.now()
                checkin_datetime = datetime.datetime.strptime(checkin_time, '%Y-%m-%d %H:%M:%S')
                working_time = checkout_datetime - checkin_datetime
                total_working_time += working_time.total_seconds()


            # Convert total_working_time to hours, minutes, and seconds
            hours, remainder = divmod(total_working_time, 3600)
            minutes, seconds = divmod(remainder, 60)

            
            return {
                'indatetime':str(dd_mm_yyyy(str(intimedate))) + ' ' + str(intime),
                'outdatetime':str(dd_mm_yyyy(str(outtimedate))) + ' ' + str(outtime),
                'total_hrs':str(add_leading_zero(int(hours))) +':' +str(add_leading_zero(int(minutes))) +':'+str(add_leading_zero(int(seconds))),
                'data':get_data,
                'intime':return__timestring__(intime),
                'outtime':return__timestring__(outtime),
                'intimedate':intimedate,
                'outtimedate':outtimedate,
                'punchout':punchout,
                'hours':int(hours),
                'minutes':int(minutes),
                'seconds':int(seconds),
                    "response":{
                        "n" : 1,
                        "msg" : "pass",
                        "status" : "success"
                    }
            }
            

    else:
        
        return {
                    'indatetime':'',
                    'outdatetime':'',
                    'data':'',
                    'intime':return__timestring__(''),
                    'outtime':return__timestring__(''),
                    'intimedate':'',
                    'outtimedate':'',
                    'punchout':'',
                    'location':'',
                        "response":{
                            "n" : 0,
                            "msg" : "user not found",
                            "status" : "errror"
                        }
                        }
      
def return__timestring__(string):
    if string  == '' or string == ' ' or string == '--' or string == '-':
        return '--:--'
    else:
        return string

@api_view(['POST'])
def team_tracker_by_department_api(request):
    if request.method == "POST":
        teamfilter = request.data.get('teamfilter')
        my_date = datetime.date.today()
        year, week_num, day_of_week = my_date.isocalendar()
        year = year
        week = week_num
        prevweek = week-1
        loguser = request.user.id
        emplist=[]
        empteamdatalist=[]
        departmentid=request.POST.get('department_id')
        if departmentid is not None and departmentid !='':
            
            companycode = request.user.company_code
            userobj = Users.objects.filter(id=loguser,company_code=companycode,is_active=True).first()
            if userobj is not None:
                userser = UserSerializer(userobj)
                department = [departmentid]
                

                

                if teamfilter == "All":
                    empobjects = Users.objects.filter(DepartmentID__in = department,is_active=True,company_code=companycode).exclude(id=loguser).order_by('id')
                    allempser =  UsersSerializer(empobjects,many=True)
                    for t in  allempser.data:
                        emplist.append(t['id'])
                elif teamfilter == "Present":
                    empobjects = Users.objects.filter(DepartmentID__in = department,is_active=True,company_code=companycode).exclude(id=loguser).order_by('id')
                    presentempser = UsersSerializer(empobjects,many=True)
                    for p in presentempser.data:
                        attendance_id = p['employeeId']
                        attexist = attendance.objects.filter(employeeId=str(attendance_id),date=my_date,company_code=companycode).order_by('time').first()
                        if attexist is not None:
                            emplist.append(p['id'])
                elif teamfilter == "On Leave":
                    empobjects = Users.objects.filter(DepartmentID__in = department,is_active=True,company_code=companycode).exclude(id=loguser).order_by('id')
                    onleaveempser = UsersSerializer(empobjects,many=True)
                    for p in onleaveempser.data:
                        empid = p['id']
                        leaveobj = Leave.objects.filter(employeeId=str(empid),WorkFromHome=False,company_code=companycode,leave_status = "Approved")
                        leaveser = leaveserializer(leaveobj,many=True)
                        for l in leaveser.data:
                            leave_start_date = datetime.datetime.strptime(l['start_date'], '%Y-%m-%d')
                            leave_end_date = datetime.datetime.strptime(l['end_date'], '%Y-%m-%d')
                            check_date = datetime.datetime.strptime(str(my_date), '%Y-%m-%d')
                        
                            result = leave_start_date <= check_date <= leave_end_date
                            if result == True:
                                emplist.append(p['id'])
                else:
                    empobjects = Users.objects.filter(DepartmentID__in = department,is_active=True,company_code=companycode).exclude(id=loguser).order_by('id')
                    onleaveempser = UsersSerializer(empobjects,many=True)
                    for p in onleaveempser.data:
                        empid = p['id']
                        leaveobj = Leave.objects.filter(employeeId=str(empid),WorkFromHome=True,company_code=companycode,leave_status = "Approved")
                        leaveser = leaveserializer(leaveobj,many=True)
                        for l in leaveser.data:
                            leave_start_date = datetime.datetime.strptime(l['start_date'], '%Y-%m-%d')
                            leave_end_date = datetime.datetime.strptime(l['end_date'], '%Y-%m-%d')
                            check_date = datetime.datetime.strptime(str(my_date), '%Y-%m-%d')

                            result = leave_start_date <= check_date <= leave_end_date
                            if result == True:
                                emplist.append(p['id'])

                for e in emplist:
                    empdict = {}
                    userempobj =  Users.objects.filter(id=int(e),company_code=companycode,is_active=True).first()
                    if userempobj is not None:
                        
                        attendance_id =userempobj.employeeId
                        username = userempobj.Firstname + " "+ userempobj.Lastname
                        empdict['username'] = username

                        deptid = department
                        for d in deptid:
                            deptobj = Department.objects.filter(id=d,Active=True,company_code=companycode).first()
                            if deptobj is not None:

                                deptname = deptobj.DepartmentName
                            else:
                                deptname = '---'


                        empdict['empdeparment']=deptname

                        designationid = userempobj.DesignationId
                        empdes = Designation.objects.filter(DesignationName=designationid,isactive=True,company_code=companycode).first()
                        if empdes is not None:
                            empdict['empdesignation'] = empdes.DesignationName
                        else:
                            empdict['empdesignation'] = ''

                        trackproscore = IntermediateTrackProResult.objects.filter(Employee_id = int(e),Week = prevweek,company_code=companycode).first()
                        if trackproscore is not None:
                            trackproper = trackproscore.TrackProPercent
                            empdict['prevtrackproscore'] = str(trackproper) +" %"
                        else:
                            empdict['prevtrackproscore'] = "--"
                        
                    
                        if attendance_id is not None:
                            userattobjcheckin = attendance.objects.filter(employeeId=str(attendance_id),date=my_date,company_code=companycode).order_by('time').first()
                            if userattobjcheckin is not None:
                                checkintime =  userattobjcheckin.time
                                empdict['checkin'] = str(checkintime).split(":")[0]+" : "+str(checkintime).split(":")[1]
                            else:
                                empdict['checkin'] ="--"

                            userattobjcount = attendance.objects.filter(employeeId=str(attendance_id),date=my_date,company_code=companycode).count()
                            if userattobjcount > 1:
                                userattobjcheckout = attendance.objects.filter(employeeId=str(attendance_id),date=my_date,company_code=companycode).order_by('time').last()
                                if userattobjcheckout is not None:
                                    checkouttime = userattobjcheckout.time
                                    empdict['checkout'] = str(checkouttime).split(":")[0]+" : "+str(checkouttime).split(":")[1]
                                else:
                                    empdict['checkout'] = "--"
                            else:
                                empdict['checkout'] = "--"

                        else:
                            empdict['checkin'] = "--"
                            empdict['checkout'] = "--"

                        userphoto = userempobj.Photo
                        if userphoto is not None and userphoto != "":
                            empdict['userimage'] = imageUrl + "/media/" + str(userphoto)
                        else:
                            empdict['userimage'] = imageUrl + "/static/assets/images/profile.png"
                
                        empteamdatalist.append(empdict)
                departmentlist=[]
                departments_obj=ManagerPinedDepartmentMaster.objects.filter(user_id=userser.data['id'],is_active=True)
                if departments_obj.exists():
                    dept_serializer=ManagerPinedDepartmentMasterSerializer(departments_obj,many=True)
                    departmentlist=dept_serializer.data
                else:
                    departments_obj=Department.objects.filter(Active=True)
                    dept_serializer = DepartmentSerializer(departments_obj, many=True)
                    for dep in dept_serializer.data:
                        if department==dep['id']:
                            departmentlist.append({'department_id':dep['id'],'department_name':dep['DepartmentName'],'user_id':request.user.id,'is_active':True,'pined':True,})
                        else:
                            departmentlist.append({'department_id':dep['id'],'department_name':dep['DepartmentName'],'user_id':request.user.id,'is_active':True,'pined':False,})


                return Response({'data':empteamdatalist,'departmentlist':departmentlist,"n":1,"Msg":"List fetched successfully","Status":"Success"})
            else:
                return Response({'data':'','departmentlist':[],"n":0,"Msg":"Employee not found","Status":"Failed"})
            
        else:
            return Response({'data':'','departmentlist':[],"n":0,"Msg":"department not found","Status":"Failed"})

@api_view(['POST'])
def addMapping(request):
    user = request.user
    if user.is_admin == True:
        if type(request.data.get("UserID")) == str:
            UserIDlist = request.data.getlist('UserID')
        else:
            UserIDlist = request.data.get('UserID')

        request_data = request.data.copy()
        com_code = request.user.company_code
        ManagerID = int(request.data.get('ManagerID'))
        FixedMapping = request.data.get('FixedMapping')
        objects_to_delete = UserToManager.objects.filter(
            ManagerID_id=ManagerID,company_code = com_code)
        objects_to_delete.delete()

        M = Users.objects.filter(id=ManagerID,company_code=com_code,is_active=True)
        for mm in M:
            mStr = mm.Firstname + ' ' + mm.Lastname

        for UserID in UserIDlist:

            u = Users.objects.filter(id=int(UserID),company_code=com_code,is_active=True)
            for uu in u:
                ustr = uu.Firstname + ' ' + uu.Lastname
            try:
                
                serializer = MappingSerializer(data={'UserID': int(
                    UserID), 'ManagerID': ManagerID, 'FixedMapping': FixedMapping,'company_code':com_code,'Active':True})
            except Exception as e:
                return Response({'data':'','n':0,'Msg': 'serializer not accepting data','Status':'Failed'})
            else:
                if serializer.is_valid():
                    data = {}
                    serializer.validated_data['UserIDStr'] = ustr
                    serializer.validated_data['ManagerIDStr'] = mStr
                    serializer.validated_data['CreatedBy'] = request.user
                    u = serializer.save()
                    data['data'] = ""
                    data['n'] = 1
                    data['Msg'] = 'Mapping added successfully'
                    data['Status'] = 'Success'
                    data['ID'] = u.id
                else:
                    data['data'] = ""
                    data['n'] = 1
                    data['Msg'] = serializer.errors
                    data['Status'] = 'Success'
        return Response(data=data)
        

    else:
        return Response({'n': 0, 'Msg': 'User has no permission to create','Status':'Failed'})

# UPDATE MAPPING----------------------------------------------------------------------------------------------------------


@api_view(['POST'])
def updateMapping(request):
    data = {'n': '', 'Msg': '', 'Status': ''}
    try:
        request_data = request.data.copy()
        request_data['company_code'] = request.user.company_code
        mappingID = request.query_params.get('mappingID')
        if mappingID is None:
            data['n'] = 0
            data['Msg'] = 'Mapping ID is none'
            data['Status'] = "Failed"
        else:
            try:
                object = UserToManager.objects.get(id=mappingID,company_code = request.user.company_code)
            except Exception as e:
                data['n'] = 0
                data['Msg'] = 'MAPPING DOES NOT EXIST'
                data['Status'] = "Failed"
            else:
                serializer = MappingSerializer(object, data = request_data)
                if serializer.is_valid():
                    serializer.save()
                    data['n'] = 1
                    data['Msg'] = 'update successfull'
                    data['Status'] = "Success"
                else:
                    # data = serializer.errors
                    data['n']=0
                    data['Msg']= serializer.errors
                    data['Status']="Failed"
        return Response(data=data)

    except Exception as e:
        data['n'] = 0
        data['Msg'] = 'try method failed'
        data['Status'] = "Failed"
        return Response(data=data)


@api_view(['GET'])
def mappingJoinQuery(request):
    id = request.query_params.get('ManagerID', None)
    if id is not None:
        users_obj=Users.objects.filter(id=id,is_active=True).first()
        if users_obj is not None:
            userobjects = UserToManager.objects.filter(ManagerID=id)
            Serializer = MappingSerializer(userobjects, many=True)
            return JsonResponse({'n': 1, 'Msg': 'Data fetched successfully', 'Status': 'Success','data':Serializer.data}, safe=False)
        else:
            return Response({'n': 0, 'Msg': 'Team leader is deleted', 'Status': 'Failed'})
            
        # users = Users.objects.filter(id__in =Subquery(userobjects.values('id')) )
    else:
        return Response({'n': 0, 'Msg': 'Manager ID value is None', 'Status': 'Failed'})



# get object in mapping---------------------------------------------------------------------------------------------


@api_view(['GET'])
def getMapping(request):
    if request.method == 'GET':
        id = request.query_params.get('mappingID', None)
        userID = request.query_params.get('userID', None)
        if id is not None:
            i = UserToManager.objects.filter(id=id,company_code = request.user.company_code)
            if i is not None:
                serializer = MappingSerializer(i,many=True)
                return Response({"n":1,"Msg":"List fetched successfully","Status":"Success","data":serializer.data})
            return Response({"n":0,"Msg":"List not found","Status":"Failed","data":''})
        if userID is not None:
            u = UserToManager.objects.filter(UserID=userID,company_code = request.user.company_code)
            if u is not None:
                serializer = MappingSerializer(u,many=True)
                return Response({"n":1,"Msg":"List fetched successfully","Status":"Success","data":serializer.data})
            return Response({"n":0,"Msg":"List not found","Status":"Failed","data":''})


@api_view(['GET'])
@permission_classes((AllowAny,))
def basegetMapping(request):
    if request.method == 'GET':
        userID = request.query_params.get('userID', None)
        if userID is not None and userID !="None":
            u = UserToManager.objects.filter(UserID=userID)
            if u is not None:
                serializer = MappingSerializer(u,many=True)
                return Response({"n":1,"Msg":"List fetched successfully","Status":"Success","data":serializer.data})
            return Response({"n":0,"Msg":"List not found","Status":"Failed","data":''})
        return Response({"n":0,"Msg":"List not found","Status":"Failed","data":''})


@api_view(['GET'])
def getMappingForUpdate(request):
    if request.method == 'GET':
        ManagerID = request.query_params.get('ManagerID', None)
        if ManagerID is not None:
            mappingobj = UserToManager.objects.filter(ManagerID=ManagerID,company_code = request.user.company_code)
            if mappingobj.exists():
                serializer = MappingSerializer(mappingobj, many=True)
                newlist=[]
                for i in serializer.data:
                    i['Photo']=''
                    user_obj=Users.objects.filter(id=i['UserID'],company_code=request.user.company_code,is_active=True).first()
                    user_serializer=UsersSerializer(user_obj)
                    if user_obj is not None:
                        if user_serializer.data['Photo'] is not None and user_serializer.data['Photo'] !='' and user_serializer.data['Photo'] !="None":
                            i['Photo']=user_serializer.data['Photo']
                        else:
                            i['Photo']='/static/assets/images/profile.png'
                            
                        newlist.append(i)
                return Response({"n":1,"Msg":"List fetched successfully","Status":"Success","data":newlist})
            return Response({"n":0,"Msg":"List not found","Status":"Failed","data":''})
        return Response({"n":0,"Msg":"Please provide manager Id","Status":"Failed","data":''})


@api_view(['POST'])
def deleteMappingForManager(request):
    data = {"n": 0, "Msg": "", "Status": ""}
    try:
        ManagerID = request.POST.get('ManagerID', None)
        # userID = request.query_params.get('userID')
        if ManagerID is not None:
            deletemanagerobj = UserToManager.objects.filter(ManagerID=ManagerID,company_code = request.user.company_code)
            if deletemanagerobj is None or not deletemanagerobj:
                data['n'] = 0
                data['Msg'] = 'Mapping does not exists'
                data['Status'] = "Failed"
            else:
                operation = deletemanagerobj.delete()
                # operation = deletemanagerobj.update(Active=False)
                
                if operation:
                    data['n'] = 1
                    data['Msg'] = 'Mapping deleted successfull'
                    data['Status'] = "Success"
                else:
                    data['n'] = 0
                    data['Msg'] = 'Unable to delete mapping'
                    data['Status'] = "Failed"
        else:
            data['n'] = 0
            data['Msg'] = 'Manager id is none'
            data['Status'] = "Failed"
    except Exception as e:
        data['n'] = 0
        data['Msg'] = 'try method failed'
        data['Status'] = "Failed"
    return Response(data=data)




# USER MANAGER MAPPING END---------------------------------------------------------------------------------------------------


# LOCATION START------------------------------------------------------------------------------------------

@api_view(['GET'])
def LocationListAPI(request, format=None):
    if request.method == 'GET':
        loc = Location.objects.filter(Active=True).order_by('id')
        serializer = LocationSerializer(loc, many=True)
        for i in serializer.data:
            user_obj = Users.objects.filter(locationId=i['id'],is_active=True)
            user_serializer=UsersSerializer(user_obj,many=True)
            if len(user_serializer.data) > 0:
                i['Used']=True
            else:
                i['Used']=False
        return Response({'n': 1, 'Msg': 'Location added successfully', 'Status': 'success','data':serializer.data})



@api_view(['POST'])
def addLocation(request):
    user = request.user
    if user.is_admin == True:
        try:
            request_data = request.data.copy()
            LocationName = request_data['LocationName'].strip()
            request_data['LocationName'] = LocationName

            request_data['lattitude'] = request_data['lattitude']
            request_data['longitude'] = request_data['longitude']
            request_data['meter'] = request_data['meter']
            request_data['address'] = request_data['address']
            request_data['company_code'] = request.user.company_code
            request_data['Active'] = True
        except Exception as e:
            return Response({'n': 0, 'Msg': 'Error saving data', 'Status': 'Failed'})
        else:
            locObject = Location.objects.filter(LocationName__in=[LocationName.strip().capitalize(),LocationName.strip(),LocationName.title(),LocationName.lower()],company_code=request.user.company_code,Active=True).first()
            if locObject is None:
                
                serializer = LocationSerializer(data=request_data)
                if serializer.is_valid():
                    data = {}
                    serializer.validated_data['CreatedBy'] = request.user
                    u = serializer.save()
                    return Response({'n': 1, 'Msg': 'Location added successfully', 'Status': 'success'})
                else:
                    return Response({'n': 0, 'Msg': serializer.errors, 'Status': 'Failed'})
            return Response({'n': 0, 'Msg': "Location already exists", 'Status': 'Failed'})   
    else:
        return Response({'n': 0, 'Msg': 'User has no permission to create', 'Status': 'Failed'})

# UPDATE location----------------------------------------------------------------------------------------------------------


@api_view(['POST'])
def updateLocation(request):
    data = {'n': '', 'Msg': '', 'Status': ''}
    try:
        request_data = request.data.copy()
        LocationName = request_data['LocationName'].strip()
        request_data['LocationName'] = LocationName
        request_data['lattitude'] = request_data['lattitude']
        request_data['longitude'] = request_data['longitude']
        request_data['meter'] = request_data['meter']
        request_data['address'] = request_data['address']
        request_data['company_code'] = request.user.company_code
        request_data['Active'] = True
        locationID = request.query_params.get('locationID')
        if locationID is None:
            data['n'] = 0
            data['Msg'] = 'location ID is none'
            data['Status'] = "Failed"
        else:
            try:
                object = Location.objects.filter(id=locationID,company_code = request_data['company_code'],Active=True).first()
            except Exception as e:
                data['n'] = 0
                data['Msg'] = 'LOCATION DOES NOT EXIST'
                data['Status'] = "Failed"
            else:
                locObject = Location.objects.filter(LocationName__in = [LocationName.strip().capitalize(),LocationName.strip(),LocationName.title(),LocationName.lower()],Active= True,company_code = request_data['company_code']).first()
                if object.LocationName != request_data['LocationName'] and locObject is not None :
                    return Response({'n': 0, 'Msg': "Location already exists", 'Status': 'Failed'})
                serializer = LocationSerializer(object, request_data)
                if serializer.is_valid():
                    serializer.save()
                    data['n'] = 1
                    data['Msg'] = 'update successfull'
                    data['Status'] = "Success"
                else:
                    data = serializer.errors
                    return Response({'n': 0, 'Msg': serializer.errors, 'Status': 'Failed'})
        return Response(data=data)

    except Exception as e:
        data['n'] = 0
        data['Msg'] = 'try method failed'
        data['Status'] = "Failed"
        return Response(data=data)

# get object in location---------------------------------------------------------------------------------------------


@api_view(['GET'])
def getLocation(request):
    if request.method == 'GET':
        id = request.query_params.get('locationID', None)
        if id is not None:
            i = Location.objects.filter(id=id,Active = True,company_code = request.user.company_code).first()
            if i is not None:
                serializer = LocationSerializer(i)
                return Response({'n': 1, 'Msg': 'Location fetched successfully', 'Status': 'success','data':serializer.data})
            return Response({'n': 0, 'Msg': 'location not found', 'Status': 'failed','data':''})
        return Response({'n': 0, 'Msg': 'Please provide Id', 'Status': 'failed','data':''})

# LOCATION END--------------------------------------------------------------------------------------------------

# FIN YEAR START--------------------------------------------------------------------------------------------


@api_view(['GET'])
def FinYearListAPI(request, format=None):
    if request.method == 'GET':
        Finyear = FinancialYear.objects.filter(company_code=request.user.company_code)
        serializer = FinancialYearSerializer(Finyear, many=True)
        return Response({'n': 1, 'Msg': 'Financial Year list fetched successfully', 'Status': 'success','data':serializer.data})


@api_view(['POST'])
def addFinYear(request):
    user = request.user
    if user.is_admin == True:
        try:
            request_data = request.data.copy()
            financialYear = request_data['financialYear'].strip()
            request_data['financialYear'] = financialYear
            YearCode = request_data['YearCode'].strip()
            request_data['YearCode'] = YearCode
            request_data['company_code'] = request.user.company_code
            
        except Exception as e:
            return Response({'n': 0, 'Msg': 'serializer not accepting data', 'Status': 'Failed'})
        else:
            finyrObject = FinancialYear.objects.filter(financialYear__in=[financialYear.strip().capitalize(),financialYear.strip(),financialYear.title(),financialYear.lower()],company_code=request.user.company_code,Active=True).first()
            fincodeObject = FinancialYear.objects.filter(YearCode__in=[YearCode.strip().capitalize(),YearCode.strip(),YearCode.title(),YearCode.lower()],company_code=request.user.company_code,Active=True).first()
            if finyrObject is not None:
                return Response({'n': 0, 'Msg': 'Financial year already exists', 'Status': 'Failed'})
            elif fincodeObject is not None:
                return Response({'n': 0, 'Msg': 'Financial year code already exists', 'Status': 'Failed'})
            else:
                serializer = FinancialYearSerializer(data=request_data)
                if serializer.is_valid():
                    serializer.validated_data['CreatedBy'] = request.user
                    u = serializer.save()
                    return Response({'n': 1, 'Msg': 'Financial Year added successfully', 'Status': 'success'})
                else:
                    return Response({'n': 0, 'Msg': serializer.errors, 'Status': 'Failed'})

    else:
        return Response({'Error': 'User has no permission to create'})

# UPDATE FIN YEAR----------------------------------------------------------------------------------------------------------


@api_view(['POST'])
def updateFinYear(request):
    data = {'n': '', 'Msg': '', 'Status': ''}
    try:
        finyearID = request.query_params.get('finyearID')
        request_data = request.data.copy()
        financialYear = request_data['financialYear'].strip()
        request_data['financialYear'] = financialYear
        YearCode =""
        if 'YearCode' in request_data.keys():
            YearCode = request_data['YearCode'].strip()
            request_data['YearCode'] = YearCode
        request_data['company_code'] = request.user.company_code
        if finyearID is None:
            data['n'] = 0
            data['Msg'] = 'financial year ID is none'
            data['Status'] = "Failed"
        else:
            object = FinancialYear.objects.filter(id=finyearID,company_code = request_data['company_code'],Active=True).first()
            if object is None:
                data['n'] = 0
                data['Msg'] = 'YEAR DOES NOT EXIST'
                data['Status'] = "Failed"
            else:
                finyrObject = FinancialYear.objects.filter(financialYear__in=[financialYear.strip().capitalize(),financialYear.strip(),financialYear.title(),financialYear.lower()],company_code=request.user.company_code,Active=True).first()
                fincodeObject = FinancialYear.objects.filter(YearCode__in=[YearCode.strip().capitalize(),YearCode.strip(),YearCode.title(),YearCode.lower()],company_code=request.user.company_code,Active=True).first()
                if object.financialYear != request_data['financialYear'] and finyrObject is not None :
                    return Response({'n': 0, 'Msg': "Finanacial year already exists", 'Status': 'Failed'})
                elif object.YearCode != request_data['YearCode'] and fincodeObject is not None :
                    return Response({'n': 0, 'Msg': "Finanacial year code already exists", 'Status': 'Failed'})
                else:
                    serializer = FinancialYearSerializer(object, data=request_data)
                    if serializer.is_valid():
                        serializer.save()
                        return Response({'n': 1, 'Msg': 'Update Successfull', 'Status': 'Succes'})

                    else:
                        return Response({'n': 0, 'Msg': serializer.errors, 'Status': 'Failed'})
        return Response(data=data)

    except Exception as e:
        data['n'] = 0
        data['Msg'] = 'try method failed'
        data['Status'] = "Failed"
        return Response(data=data)

# get object in Fin Year-------------------------------------------------------------------------------------------


@api_view(['GET'])
def getFinYear(request):
    if request.method == 'GET':
        id = request.query_params.get('finyearID', None)
        if id is not None:
            i = FinancialYear.objects.filter(id=id)
            serializer = FinancialYearSerializer(i, many=True)
            return JsonResponse(serializer.data, safe=False)

# FIN YEAR END----------------------------------------------------------------------------------------------------


# DESIGNATION START-----------------------------------------------------------------------------------------------

@api_view(['GET'])
def DesignationListAPI(request, format=None):
    if request.method == 'GET':
        designationlist=[]
        user = Designation.objects.filter(company_code=request.user.company_code,isactive=True).order_by('id')
        serializer = DesignationSerializer(user, many=True)
        for i in serializer.data:
            user_obj = Users.objects.filter(DesignationId=i['id'],company_code=request.user.company_code,is_active=True).first()
            if user_obj is not None :
                i['Used']=True
            else:
                i['Used']=False
            designationlist.append(i)
        return Response({'n':1,'Msg':'Designation list fetched successfully','Status':'Success','data':designationlist})

# Add designation --------------------------------------------------------------------------------------------


@api_view(['POST'])
def addDesignationAPI(request):
    user = request.user
    if user.is_admin == True:
        try:
            request_data = request.data.copy()
            DesignationName = request_data['DesignationName'].strip()
            request_data['DesignationName'] = DesignationName
            request_data['company_code'] = request.user.company_code
            request_data['isactive'] = True
            serializer = DesignationSerializer(data=request_data)
        except Exception as e:
            return Response({'n': 0, 'Msg': 'serializer not accepting data', 'Status': 'Failed'})
        else:
            desgObject = Designation.objects.filter(DesignationName__in=[DesignationName.strip().capitalize(),DesignationName.strip(),DesignationName.title(),DesignationName.lower()],company_code=request.user.company_code,isactive=True).first()
            if desgObject is None:
                if serializer.is_valid():
                    data = {}
                    u = serializer.save()
                    return Response({'n': 1, 'Msg': 'Designation added successfully', 'Status': 'Success'})
                else:

                    return Response({'n': 0, 'Msg': serializer.errors, 'Status': 'Failed'})
            else:
                return Response({'n': 0, 'Msg': "Designation already exists", 'Status': 'Failed'})

    else:
        return Response({'n': 1, 'Msg': 'User has no permission to create', 'Status': 'Failed'})

# UPDATE DESIGNATION----------------------------------------------------------------------------------------------------------


@api_view(['POST'])
def updateDesignation(request):
    data = {'n': '', 'Msg': '', 'Status': ''}
    try:
        request_data = request.data.copy()
        DesignationName = request_data['DesignationName'].strip()
        request_data['DesignationName'] = DesignationName
        request_data['company_code'] = request.user.company_code
        request_data['isactive'] = True
        designationID = request.query_params.get('designationID')
        if designationID is None:
            data['n'] = 0
            data['Msg'] = 'designation ID is none'
            data['Status'] = "Failed"
        else:
            try:
                object = Designation.objects.filter(id=designationID,isactive=True,company_code=request_data['company_code']).first()
            except Exception as e:
                data['n'] = 0
                data['Msg'] = 'DESIGNATION DOES NOT EXIST'
                data['Status'] = "Failed"
            else:
                desgObject = Designation.objects.filter(DesignationName__in = [DesignationName.strip().capitalize(),DesignationName.strip(),DesignationName.title(),DesignationName.lower()],isactive= True,company_code = request_data['company_code']).first()
                
                if object.DesignationName != request_data['DesignationName'] and desgObject is not None :
                    return Response({'n': 0, 'Msg': "Designation already exists", 'Status': 'Failed'})
                serializer = DesignationSerializer(object, request_data)
                if serializer.is_valid():
                    serializer.save()
                    data['n'] = 1
                    data['Msg'] = 'update successfull'
                    data['Status'] = "Success"
                else:
                    data['n'] = 0
                    data['Msg'] = serializer.errors
                    data['Status'] = "Failed"
        return Response(data=data)

    except Exception as e:
        data['n'] = 0
        data['Msg'] = 'try method failed'
        data['Status'] = "Failed"
        return Response(data=data)

# get object in designation---------------------------------------------------------------------------------------------


@api_view(['GET'])
def getDesignation(request):
    if request.method == 'GET':
        id = request.query_params.get('designationID', None)
        if id is not None:
            i = Designation.objects.filter(id=id,company_code=request.user.company_code).first()
            if i is not None:
                serializer = DesignationSerializer(i)
                return Response({'n': 1, 'Msg': 'Designation fetched successfully', 'Status': 'success','data':serializer.data})
            return Response({'n': 0, 'Msg': 'Designation not found', 'Status': 'failed','data':''})
        return Response({'n': 0, 'Msg': 'Please provide Id', 'Status': 'failed','data':''})
# DESIGNATION END------------------------------------------------------------------------------------------


@api_view(['GET'])
def UserListAPI(request, format=None):

    if request.method == 'GET':
        # user = Users.objects.all().order_by('CreatedOn')
        company_code = request.user.company_code
        user = Users.objects.filter(
            is_active=True,company_code=company_code).order_by('id')
        serializer = UsersSerializer(user, many=True)
        for i in serializer.data:
            i['Fullname'] = str(i['Firstname'] + " " + i['Lastname']).strip()
            if i['DateofJoining'] is not None:
                datestr = i['DateofJoining']
                newdate = datetime.datetime.strptime(datestr,"%Y-%m-%d")
                newmonthdate = newdate.strftime("%d %b %Y")
                i['DateofJoining']=newmonthdate

        return Response({'n':1,'msg':'Employee list fetched successfully','status':'success','data':serializer.data})
   


@api_view(['POST'])
def search_by_name_employee(request, format=None):

    if request.method == 'POST':
        # user = Users.objects.all().order_by('CreatedOn')
        username=request.POST.get('username')
        company_code = request.user.company_code
        user = Users.objects.filter(Q(Firstname__icontains = username,is_active=True,company_code=company_code)|Q(Lastname__icontains = username,is_active=True,company_code=company_code)).order_by('id')[:5]
        serializer = UsersSerializer(user, many=True)
        newlist=[]
        if username !="" and username is not None:
            for i in serializer.data:
                i['Fullname'] = str(i['Firstname'] + " " + i['Lastname']).strip()
                if str(username.lower()) in str(i['Fullname']).lower():
                    newlist.append(i)
        else:
            newlist=serializer.data

        return Response({'n':1,'msg':'Employee list fetched successfully','status':'success','data':newlist})
   

@api_view(['POST'])
def search_emp_by_id(request, format=None):

    if request.method == 'POST':
        # user = Users.objects.all().order_by('CreatedOn')
        managerid=request.POST.get('managerid')
        company_code = request.user.company_code
        user = Users.objects.filter(
            is_active=True,company_code=company_code,id=managerid).first()
        serializer = UsersSerializer(user)


        return Response({'n':1,'msg':'Employee list fetched successfully','status':'success','data':serializer.data})
   





@api_view(['POST'])
def delete_designation(request):
    data={}
    id = request.data.get('id')
    desgdata=Designation.objects.filter(id=id).first()
    data['isactive'] = False
    serializer = DesignationSerializer(desgdata,data=data,partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({
        "data": serializer.data,
        "response": {
            "n": 1,
            "msg": "Designation deleted successfully",
            "status": "success"
            }
        })
    else:
        return Response({
            "data": serializer.errors,
            "response": {
                "n": 0,
                "msg": "unable to delete Designation",
                "status": "warning"
            }
        })

@api_view(['POST'])
def deletelocation(request):
    data={}
    id = request.data.get('id')
    locatgdata=Location.objects.filter(id=id).first()
    data['Active'] = False
    serializer = LocationSerializer(locatgdata,data=data,partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response({
        "data": serializer.data,
        "response": {
            "n": 1,
            "msg": "deleted successfully",
            "status": "success"
            }
        })
    else:
        return Response({
            "data": serializer.errors,
            "response": {
                "n": 0,
                "msg": "unable to delete location",
                "status": "warning"
            }
        })


@api_view(["GET"])
def chkLoggedIn(request):
    if request.user.is_authenticated:
        return Response({'success': 'authenticated'})
    else:
        return Response({'error': 'err'})
    
# LOGIN-------------------------------------------------------------------------------------


@csrf_exempt
@api_view(["POST"])
@permission_classes((AllowAny,))
def login(request):
    username = request.data.get("username")
    password = request.data.get("password")

    uid_obj=Users.objects.filter(email=username,is_active=True).first()
    serializer=UserSerializer(uid_obj)
    
    uid=serializer.data['uid']
    firebasetoken = request.data.get("FirebaseID")
    desktopToken = request.data.get("desktopToken")
   
    if username is None or password is None :
        return Response({'msg': 'Please provide username and password','n':0},
                        status=HTTP_400_BAD_REQUEST)
    
    userexist = Users.objects.filter(
        email=username, is_active=True, is_admin=True,is_blocked=False)
    if userexist is None:
        userdata = Users.objects.filter(
            email=username, is_active=False, is_admin=False,is_blocked=True)
        
        if userdata:
            return Response({'msg': 'This user is blocked','n':0}, status=HTTP_400_BAD_REQUEST)
    companyobj = Users.objects.filter(
        email=username, is_active=True).first()
    if companyobj is None:
        return Response({'msg': 'Invalid Email, Please Try Again','n':0}, status=HTTP_400_BAD_REQUEST)

    user = authenticate(username=uid, password=password)
    
    if not user:
        return Response({'msg': 'Invalid Credentials for this','n':0},
                            status=HTTP_400_BAD_REQUEST)
    if user.is_superuser == True:
        try:
            django_login(request, user)
        except Exception as e:
            print(e)

        token, _ = Token.objects.get_or_create(user=user)
        roleid = user.RoleID

        roleid_ = user.RoleID_id
        roleobj = Role.objects.filter(id=roleid_,Active=True).first()
        rolename = roleobj.RoleName

        PasswordChanged = user.PasswordChanged
        secondary_info = user.secondary_info
        company_code = user.company_code
    
        menuObject = MenuItem.objects.filter(MenuID__in = list(Permissions.objects.filter(RoleID=int(roleid_)).values_list('MenuID', flat=True))).order_by('SortOrder')
        menuSerializer = GetMenuItemSerializer(menuObject,many=True)

        assign_by = False
        usermappingobj=UserToManager.objects.filter(ManagerID=user.id,Active=True,company_code=company_code).first()
        if usermappingobj is not None:
            assign_by = True

        userprofileimage = user.Photo
        if str(userprofileimage) is not None and str(userprofileimage) != "":
            userprf = imageUrl +"/media/" + str(userprofileimage)
        else:
            userprf = imageUrl + "/static/assets/images/profile.png"
        

        logindesig = user.DesignationId_id
        designaobj = Designation.objects.filter(id=logindesig).first()
        if designaobj is not None:
            desname = designaobj.DesignationName
        else:
            desname="---"

        loginlocation = user.locationId_id
        locationobj = Location.objects.filter(id=loginlocation).first()
        if locationobj is not None:
            if locationobj.lattitude is not None and locationobj.longitude is not None and locationobj.meter is not None:
                lattitude = float(locationobj.lattitude)
                longitude = float(locationobj.longitude)
                meter = float(locationobj.meter)
            else:
                lattitude = float(0.0)
                longitude = float(0.0)
                meter = float(0.0)
        else:
            lattitude =float(0.0)
            longitude =float(0.0)
            meter = float(0.0)

        desktoptokenobj = Users.objects.filter(email=username).first()
        desktoptokenobj.desktopToken = desktopToken
        desktoptokenobj.save()

        if 'FirebaseID' in request.data.keys():
            userfireobject = Users.objects.filter(email=username).first()
            userfireobject.FirebaseID = firebasetoken
            userfireobject.save()



        return Response({'token': token.key, 'user': user.id, 'Firstname': user.Firstname,'rolename':rolename,
                        'Lastname': user.Lastname,'Designation':desname,'lattitude':lattitude,'longitude':longitude,'meter':meter,'Menu': menuSerializer.data, 'Roleid': roleid_,'assign_by':assign_by,
                        'PasswordChanged': PasswordChanged, 'email': user.email,'typeofwork':user.typeofwork, 'employeeId':user.employeeId, 'secondary_info':secondary_info,'company_code':company_code,'admin':user.is_admin,'superadmin':user.is_superuser,'masters':user.masters,'rules':user.rules,'userprofile':userprf,'n':1},
                        status=HTTP_200_OK)
    else:

        apiheading = apiheader (user.company_code)
        if apiheading == True:
            try:
                django_login(request, user)
            except Exception as e:
                print(e)
            if not user:
                return Response({'error': 'Invalid Credentials','n':1},
                                status=HTTP_404_NOT_FOUND)
            token, _ = Token.objects.get_or_create(user=user)
            roleid = user.RoleID

            roleid_ = user.RoleID_id
            roleobj = Role.objects.filter(id=roleid_).first()
            rolename = roleobj.RoleName

            PasswordChanged = user.PasswordChanged
            secondary_info = user.secondary_info
            company_code = user.company_code
     
            menuObject = MenuItem.objects.filter(MenuID__in = list(Permissions.objects.filter(RoleID=int(roleid_)).values_list('MenuID', flat=True))).order_by('SortOrder')
            menuSerializer = GetMenuItemSerializer(menuObject,many=True)

            assign_by = False
            usermappingobj=UserToManager.objects.filter(ManagerID=user.id,Active=True,company_code=company_code).first()
            if usermappingobj is not None:
                assign_by = True
                

            usermasterobject = Users.objects.filter(company_code=company_code).first()
            user_rulesobject = Users.objects.filter(company_code=company_code).first()

            comanyinfoobj = companyinfo.objects.filter(companycode=company_code).first()
            if comanyinfoobj is not None:
                companylogo = str(comanyinfoobj.companylogos)
            else:
                companylogo = ""

            userprofileimage = user.Photo
            if str(userprofileimage) is not None and str(userprofileimage) != "":
                userprf = imageUrl +"/media/" + str(userprofileimage)
            else:
                userprf = imageUrl + "/static/assets/images/profile.png"

            logindesig = user.DesignationId_id
            designaobj = Designation.objects.filter(id=logindesig).first()
            if designaobj is not None:
                desname = designaobj.DesignationName
            else:
                desname="---"

            loginlocation = user.locationId_id
            locationobj = Location.objects.filter(id=loginlocation).first()
            if locationobj is not None:
                if locationobj.lattitude is not None and locationobj.longitude is not None and locationobj.meter is not None:
                    lattitude = float(locationobj.lattitude)
                    longitude = float(locationobj.longitude)
                    meter = float(locationobj.meter)
                else:
                    lattitude = float(0.0)
                    longitude = float(0.0)
                    meter = float(0.0)
            else:
                lattitude = float(0.0)
                longitude = float(0.0)
                meter = float(0.0)

            app_version = request.POST.get('app_version')
            unique_device_id = request.POST.get('unique_device_id')
            device_name = request.POST.get('device_name')



            if unique_device_id is not None and unique_device_id !="" and app_version is not None and   app_version !="" and device_name is not None and device_name !="": 
                user_deivice_obj = DeviceVerification.objects.filter(userid=user.id,is_active=True).last()
                if user_deivice_obj is not None:

                    if str(user_deivice_obj.unique_device_id) == str(unique_device_id):
                        user_appversion_obj = DeviceVerification.objects.filter(userid=user.id,app_version=app_version,is_active=True).last()
                        if user_appversion_obj is None:
                            DeviceVerification.objects.create(device_name=device_name,app_version=app_version,unique_device_id=unique_device_id,employee_code=user.uid,userid=user.id)
                            print("device matched logined successfull app version changed")
                        else:
                            print("device matched logined successfull")
                    else:
                        print("device not matched")
                        return Response({'mesg': 'Device does not matched','n':3,'user': user.id,'employeeId':user.employeeId,})
                else:
                    DeviceVerification.objects.create(device_name=device_name,app_version=app_version,unique_device_id=unique_device_id,employee_code=user.uid,userid=user.id)
                    print("user obj in device verification not found created new entry")

            desktoptokenobj = Users.objects.filter(email=username).first()
            desktoptokenobj.desktopToken = desktopToken
            desktoptokenobj.save()

            if 'FirebaseID' in request.data.keys():
                userfireobject = Users.objects.filter(email=username).first()
                userfireobject.FirebaseID = firebasetoken
                userfireobject.save()



            return Response({'token': token.key, 'user': user.id, 'Firstname': user.Firstname,'rolename':rolename,
                            'Lastname': user.Lastname,'Designation':desname,'lattitude':lattitude,'longitude':longitude,'meter':meter, 'Menu': menuSerializer.data, 'Roleid': roleid_,"assign_by":assign_by,
                            'PasswordChanged': PasswordChanged, 'email': user.email,'typeofwork':user.typeofwork, 'employeeId':user.employeeId, 'secondary_info':secondary_info,'company_code':company_code,
                            'superadmin':user.is_superuser,'admin':user.is_admin,'is_staff':user.is_staff,'masters':usermasterobject.masters,'rules':user_rulesobject.rules,'companyimage':companylogo,'userprofile':userprf,'n':1},
                            status=HTTP_200_OK)
        else:
            return Response({'msg':'Access Denieddd','n':0},status=HTTP_404_NOT_FOUND)
















@api_view(["POST"])
def logout(request):
    try:
        request.user.auth_token.delete()
    except (AttributeError, ObjectDoesNotExist):
        pass
    django_logout(request)
    return Response({"n": 1, "success": "Successfully logged out."},
                    status=status.HTTP_200_OK)




@api_view(['POST'])
def addUser(request):
    user = request.user
    if user.is_admin == True:
        try:    
            request_data = request.data.copy()
            userpassword = request.data.get('password')
            request_data['password'] = make_password(userpassword)
            request_data['Password'] = userpassword
            request_data['company_code'] = request.user.company_code
            com_code = request_data['company_code']
          
            uid = EmployeeCode(com_code)
        
            request_data['uid'] = uid
            
            request_data['is_staff'] = True
            request_data['is_active'] = True
            if request.user.is_superuser == True:
                request_data['is_superuser'] = True
            else:
                request_data['is_superuser'] = False
            # request_data['employeeId'] = ""  
            if 'employeeId' in request_data.keys():
                if request_data['employeeId'] !="": 
                    checkempid_obj=Users.objects.filter(employeeId=request_data['employeeId'],company_code=request_data['company_code'],is_active=True).first()
                    if checkempid_obj:
                        return Response({"n": 0, "Msg": "Employee of this Attendance Id Already Exists", "Status": "Failed"})
            serializer = RegisterSerializer(data=request_data, context={'request': request})
        except Exception as e:

            return Response({"n": 0, "Msg": "Exception occured, serializer not accepting data", "Status": "Failed"})
        else:



            if request_data['email'] is not None and request_data['email'] != "":
                useremailobject = Users.objects.filter(company_code=request.user.company_code,is_active=True,email=request_data['email']).exclude(email='').first()
                if useremailobject is not None:
                    return Response({"n": 0, "Msg": "User with this Official Email id already exist", "Status": "Failed"})
            
            if request_data['email'] == '':
                request_data['email']=None
                
            if 'personal_email' in request_data.keys():
                if request_data['personal_email'] is not None and request_data['personal_email'] != "":
                    userpersonalemailobject = Users.objects.filter(personal_email=request_data['personal_email'],company_code=request.user.company_code,is_active=True).exclude(email='').first()
                    if userpersonalemailobject is not None:
                        return Response({"n": 0, "Msg": "User with this Personal Email id already exist", "Status": "Failed"})





            if serializer.is_valid():
                serializer.validated_data['is_admin'] = True
                data = {}
                u = serializer.save()
                data['Status'] = 'Registered successfully'
                data['Email'] = u.email
                data['is_admin'] = u.is_admin
                data['is_staff'] = u.is_staff
                data['Firstname'] = u.Firstname
                data['Lastname'] = u.Lastname
                data['typeofwork']=request.data.get('typeofwork')
                data['personal_email']=u.personal_email

                try:
                    email = request.data.get('email', None)
                    personal_email = request.data.get('personal_email', None)
                    password = request.data.get('password', None)
                    curruser = Users.objects.filter(email=email,company_code=request.user.company_code,is_active=True).annotate(
                        Name=Concat('Firstname', Value(' '), 'Lastname'))
                    
                    Name = serializer.data['Firstname'].capitalize() +" "+ serializer.data['Lastname'].capitalize()
                    designationId = serializer.data['DesignationId']
                    desigobj = Designation.objects.filter(id=designationId,isactive=True).first()
                    designation = desigobj.DesignationName
                    compobj = companyinfo.objects.filter(companycode=serializer.data['company_code'],isactive=True).first()
                    companyName = compobj.companyName
                    d1=datetime.datetime.strptime(serializer.data['DateofJoining'], "%Y-%m-%d")
                    DateofJoining = d1.strftime('%d %B %Y')


                    dicti = {'password': password,
                             'Name': Name, 'email': email,'id':u.id}
                    message = get_template(
                        'useregistermail.html').render(dicti)
                    msg = EmailMessage(
                        'trackpro credentials',
                        message,
                        EMAIL_HOST_USER,
                        [email],
                    )
                  

                    msg.content_subtype = "html"  
                    msg.send()
                except Exception as e:
                    print('exception occured for mail', e)
                return Response({"n": 1, "Msg": "Employee has been added successfully", "Status": "success","data":serializer.data})
            else:
                data = serializer.errors
                return Response({"n": 0, "Msg": data, "Status": data})

    else:
        return Response({"n": 0, "Msg": "User has no permission to create", "Status": "Failed"})





# DELETE USER-------------------------------------------------------------------------------------
@api_view(['GET'])
def userDelete(request):
    data = {"n": 0, "Msg": "", "Status": ""}
    try:
        if request.GET.get('userID') is not None or request.GET.get('userID') !="":
            userID = request.GET.get('userID')
        else:
            userID = request.query_params.get('userID')
        if userID is not None:
            user = Users.objects.filter(id=userID,company_code=request.user.company_code,is_active=True).first()
            if user is None:
                data['n'] = 0
                data['Msg'] = 'USER DOES NOT EXIST'
                data['Status'] = "Failed"
            else:
                operation = user.delete()
                if operation:
                    data['n'] = 1
                    data['Msg'] = 'delete successfull'
                    data['Status'] = "Success"
                else:
                    data['n'] = 0
                    data['Msg'] = 'delete failed'
                    data['Status'] = "Failed"
        else:
            data['n'] = 0
            data['Msg'] = 'user.id is none'
            data['Status'] = "Failed"
    except Exception as e:
        print(e)
        data['n'] = 0
        data['Msg'] = 'try method failed'
        data['Status'] = "Failed"
    return Response(data=data)

# chk how data is coming from front end for foreign keys...accordingly use userserializer or userRegisterSerializer
# UPDATE USER-------------------------------------------------------------------------------------


@api_view(['POST'])
def userUpdate(request):
    data = {'n': '', 'Msg': '', 'Status': ''}
    try:
        request_data = request.data.copy()
        request_data['company_code'] = request.user.company_code
        userID = request.query_params.get('userID')
        if userID is None:
            data['n'] = 0
            data['Msg'] = 'User ID is none'
            data['Status'] = "Failed"
        else:
            try:
                user = Users.objects.filter(id=userID,is_active = True,company_code = request_data['company_code']).first()
            except Exception as e:
                data['n'] = 0
                data['Msg'] = 'USER DOES NOT EXIST'
                data['Status'] = "Failed"
            else:
                request_data = request.data.copy()
                if request.user.is_superuser == True:
                    request_data['is_superuser'] = True
                else:
                    request_data['is_superuser'] = False
                if 'employeeId' in request_data.keys():
                    if request_data['employeeId'] !="" and request_data['employeeId'] is not None:
                        checkempid_obj=Users.objects.filter(employeeId=request_data['employeeId'],company_code=request.user.company_code,is_active=True).exclude(id=userID).first()
                        if checkempid_obj:
                            return Response({"n": 0, "Msg": "Employee of this Attendance Id Already Exists", "Status": "Failed"}) 

                serializer = userUpdateSerializer(user, data = request_data,partial=True)
                if serializer.is_valid():
                    serializer.validated_data['is_admin'] = True
                    serializer.validated_data['UpdatedBy'] = request.user
                    serializer.save()

                    data['n'] = 1
                    data['Msg'] = 'update successfull'
                    data['Status'] = "Success"
                    return Response({"n": 1, "Msg": "Employee has been updated successfully", "Status": data})
                else:
                    data = serializer.errors
                    return Response({"n": 0, "Msg": data, "Status": data})
                    # data['n']=0
                    # data['Msg']= 'update failed serializer invalid'
                    # data['Status']="Failed"
        return Response(data=data)

    except Exception as e:
        print(e)

        data['n'] = 0
        data['Msg'] = 'try method failed'
        data['Status'] = "Failed"
        return Response(data=data)

# get user------------------------------------------


@api_view(['GET'])
@permission_classes((AllowAny,))
def getUser(request):
    if request.method == 'GET':

        deptID = request.query_params.get('departmentID', None)
        userID = request.query_params.get('userID', None)
        if deptID is not None:
            i = Users.objects.filter(DepartmentID=deptID,is_active=True)
            serializer = userUpdateSerializer(i, many=True)
            return JsonResponse(serializer.data, safe=False)
        if userID is not None:
            i = Users.objects.filter(id=userID,is_active=True).first()
            if i is not None:
                serializer = UserSerializer(i)
                
                return Response({'n': 1, 'Msg': 'User fetched successfully', 'Status': 'success','data':serializer.data})
            return Response({'n': 0, 'Msg': 'User not found', 'Status': 'failed','data':''})
        return Response({'n': 0, 'Msg': 'Please provide Id', 'Status': 'failed','data':''})



@api_view(['GET'])
@permission_classes((AllowAny,))
def getUserDetailsForSecondaryInfo(request):
    if request.method == 'GET':
        deptID = request.query_params.get('departmentID', None)
        userID = request.query_params.get('userID', None)
        if deptID is not None:
            i = Users.objects.filter(DepartmentID=deptID,is_active=True)
            serializer = userUpdateSerializer(i, many=True)
            return JsonResponse(serializer.data, safe=False)
        if userID is not None:
            i = Users.objects.filter(id=userID,is_active=True).first()
            if i is not None:
                serializer = UserSerializer(i)
                return Response({'n': 1, 'Msg': 'User fetched successfully', 'Status': 'success','data':serializer.data})
            return Response({'n': 0, 'Msg': 'User not found', 'Status': 'failed','data':''})
        return Response({'n': 0, 'Msg': 'Please provide Id', 'Status': 'failed','data':''})

# BlOCK USER----------------------------------------------------------------------------


@api_view(['GET'])
def userBlock(request):
    data = {"n": 0, "Msg": "", "Status": ""}
    try:
        userID = request.POST.get('userID')
        if userID is not None:
            user = Users.objects.filter(id=userID)
            if user is None:
                data['n'] = 0
                data['Msg'] = 'USER DOES NOT EXIST'
                data['Status'] = "Failed"
            else:
                operation = Users.objects.filter(id=userID).update(
                    is_active=False, is_admin=False,is_blocked=True)
                if operation:
                    data['n'] = 1
                    data['Msg'] = 'User has been deleted successfully'
                    data['Status'] = "Success"
                else:
                    data['n'] = 0
                    data['Msg'] = 'Failed to delete this user'
                    data['Status'] = "Failed"
        else:
            data['n'] = 0
            data['Msg'] = 'user.id is none'
            data['Status'] = "Failed"
    except Exception as e:
        print(e)
        data['n'] = 0
        data['Msg'] = 'try method failed'
        data['Status'] = "Failed"
    return Response(data=data)

# UNBlOCK USER----------------------------------------------------------------------------


@api_view(['GET'])
def userUnblock(request):
    data = {"n": 0, "Msg": "", "Status": ""}
    try:
        userID = request.POST.get('userID')
        if userID is not None:
            user = Users.objects.filter(id=userID)
            if user is None:
                data['n'] = 0
                data['Msg'] = 'USER DOES NOT EXIST'
                data['Status'] = "Failed"
            else:
                operation = Users.objects.filter(id=userID).update(
                    is_active=True, is_admin=True,is_blocked=False)
                if operation:
                    data['n'] = 1
                    data['Msg'] = 'User has been unblocked successfully'
                    data['Status'] = "Success"
                else:
                    data['n'] = 0
                    data['Msg'] = 'Failed to unblock this user'
                    data['Status'] = "Failed"
        else:
            data['n'] = 0
            data['Msg'] = 'user.id is none'
            data['Status'] = "Failed"
    except Exception as e:
        print(e)
        data['n'] = 0
        data['Msg'] = 'try method failed'
        data['Status'] = "Failed"
    return Response(data=data)


# PERMISSION  START-------------------------------------------------------------------------------


@api_view(['GET'])
def PermissionListAPI(request, format=None):
    if request.method == 'GET':
        Perm = Permissions.objects.all()
        serializer = PermissionsSerializer(Perm, many=True)
        return Response(serializer.data)


@api_view(['POST'])
def addPermissions(request):
    user = request.user
    if user.is_admin == True:
        request_data = request.data.copy()
        request_data['company_code'] = request.user.company_code
        permissionobject = Permissions.objects.filter(RoleID = request_data['RoleID']).first()
        if permissionobject is None:
            serializer = PermissionsSerializer(data = request_data)
            if serializer.is_valid():
                serializer.save()
                return Response ({
                            "data": serializer.data,
                            "response":{
                                "n" : 1,
                                "msg" : "Permission added successfully",
                                "status" : "Success"
                            }
                            })
            return Response ({
                            "data": serializer.errors,
                            "response":{
                                "n" : 0,
                                "msg" : "Error adding permissions",
                                "status" : "Failed"
                            }
                            })
        else:
            serializer = PermissionsSerializer(permissionobject,data = request_data)
            if serializer.is_valid():
                serializer.save()
                return Response ({
                            "data": serializer.data,
                            "response":{
                                "n" : 1,
                                "msg" : "Permission updated successfully",
                                "status" : "Success"
                            }
                            })
            return Response ({
                            "data": serializer.errors,
                            "response":{
                                "n" : 0,
                                "msg" : "Error updating permissions",
                                "status" : "Failed"
                            }
                            })  
    else:
        return Response ({
                            "data": '',
                            "response":{
                                "n" : 0,
                                "msg" : "User has no permission to create",
                                "status" : "Failed"
                            }
                            })


@api_view(['GET'])
def deletePermission(request):
    data = {"n": 0, "Msg": "", "Status": ""}
    try:
        RoleID = request.GET.get('RoleID')
        if RoleID is not None:
            Perm = Permissions.objects.filter(RoleID_id=RoleID)
            if Perm is None:
                data['n'] = 0
                data['Msg'] = 'PERMISSION DOES NOT EXIST'
                data['Status'] = "Failed"
            else:
                operation = Perm.delete()
                if operation:
                    data['n'] = 1
                    data['Msg'] = 'delete successful'
                    data['Status'] = "Success"
                else:
                    data['n'] = 0
                    data['Msg'] = 'delete failed'
                    data['Status'] = "Failed"
        else:
            data['n'] = 0
            data['Msg'] = 'permission.id is none'
            data['Status'] = "Failed"
    except Exception as e:
        print(e)
        data['n'] = 0
        data['Msg'] = 'try method failed'
        data['Status'] = "Failed"
    return Response(data=data)


@api_view(['GET'])
def getPermission(request):
    data = {"n": 0, "Msg": "", "Status": ""}
    id = request.query_params.get('RoleID', None)
    companycode = request.user.company_code
    if id is not None:
        i = Permissions.objects.filter(RoleID=id,company_code=companycode)
        serializer = PermissionsSerializer(i, many=True)
        return JsonResponse(serializer.data, safe=False)


@api_view(['GET'])
def MenuitemListAPI(request, format=None):
    if request.method == 'GET':
        company_code = request.user.company_code
        Menu = MenuItem.objects.filter(company_code = company_code).order_by('SortOrder')
        if Menu.exists():
            serializer = GetMenuItemSerializer(Menu, many=True)
            return Response(serializer.data)
        else:
            Menuobject = MenuItem.objects.filter(company_code=None).order_by('SortOrder')
            serializer = GetMenuItemSerializer(Menuobject, many=True)
            return Response(serializer.data)

@api_view(['GET'])
def rolepermissionlist(request, format=None):
    if request.method == 'GET':
        company_code = request.user.company_code
        Perm = Permissions.objects.filter(company_code=company_code).order_by('id')
        serializer = PermissionsSerializer(Perm, many=True )
        rolelist=[]
        for i in serializer.data:
            roleobj = Role.objects.filter(id=i['RoleID'],company_code=company_code,Active=True).first()
            if roleobj is not None:
                i['rolename'] = roleobj.RoleName
                roleused = Users.objects.filter(RoleID=i['RoleID'],company_code=request.user.company_code,is_active=True).first()
                if roleused is not None:
                    i['roleused'] = "True"
                else:
                    i['roleused'] = "False"
                rolelist.append(i)   

        newlist = sorted(rolelist, key=lambda d: d['rolename'])

        return Response(newlist)

      
      
      
def addPermission(RoleId,companycode,menulist):      
        permissioncreate = Permissions.objects.create(
                                                        RoleID_id = RoleId,
                                                        company_code = companycode
                                )
        permissioncreate.MenuID.set(menulist)   

@api_view(['POST'])
def addduplicateroleapi(request,format=None):
    if request.method == 'POST':
        companycode = request.user.company_code
        request_data = request.data.copy()
        rolenamelist=[]

        roleid = request_data['RoleID']
        roleobj = Role.objects.filter(id=roleid).first()
        if roleobj is not None:
            permissionobj = Permissions.objects.filter(RoleID_id = roleid, company_code = companycode ).first()
            Permissionser = PermissionsSerializer(permissionobj)
            for p in [Permissionser.data]:
                menulist = p['MenuID']

            duprolename  = roleobj.RoleName 
            if "-" in duprolename:
                rolesearch = duprolename.split('-')[0]
            else:
                rolesearch = duprolename


            role = Role.objects.filter(company_code=companycode,Active=True,RoleName__startswith = rolesearch)
            serializer = RoleSerializer(role, many=True)
            for r in serializer.data:
                rolenamelist.append(r['RoleName'])

            numberlist=[]
            for k in rolenamelist:
                if "-" in k:
                    numb = k.split('-')[1]
                    numberlist.append(int(numb))

            if len(numberlist) == 0:
                firstdup = 1
                newrolestring = str(rolesearch)+"-"+str(firstdup)
            else:
                highestnumb = max(numberlist)
                nextnumber = int(highestnumb)+1
                newrolestring = str(rolesearch)+"-"+str(nextnumber)
            
            newrolename = newrolestring
            newrole = Role.objects.create(RoleName=newrolename,Active=True,company_code=companycode)
            newroleid = newrole.id
            addPermission (newroleid,companycode,menulist)
            return Response ({"n" : 1,"msg" : "Role and Permissions added successfully","status" : "Success"})
        else:
            return Response ({"n" : 0,"msg" : "Role not found","status" : "failure"})

        

@api_view(['POST'])
def deleteroleapi(request,format=None):
    if request.method == 'POST':
        data={}
        companycode = request.user.company_code
        request_data = request.data.copy()
        roleid = request_data['RoleID']
        roleobj = Role.objects.filter(id=roleid).first()
        data['Active'] = False
        if roleobj is not None:
            serializer = RoleSerializer(roleobj,data=data,partial=True)
            if serializer.is_valid():
                serializer.save()

            Perm = Permissions.objects.filter(RoleID_id=roleid)
            if Perm is None:
                  return Response ({"n" :0,"msg" : "permissions not present","status" : "failure"})
            else:
                operation = Perm.delete()
                if operation:
                    return Response ({"n" : 1,"msg" : "Role and Permissions deleted successfully","status" : "Success"})
                else:
                    return Response ({"n" : 0,"msg" : "Role and Permissions not deleted","status" : "failure"})
@api_view(['POST'])
def editrolenameapi(request,format=None):
    if request.method == 'POST':
        data={}
        companycode = request.user.company_code
        request_data = request.data.copy()
        roleid = request_data['RoleID']
        rolename = request_data['RoleName']
        roleobj = Role.objects.filter(id=roleid).first()
        data['RoleName'] = rolename
        if roleobj is not None:
            roleexist = Role.objects.filter(company_code=companycode,RoleName__in=[rolename.strip().capitalize(),rolename.strip(),rolename.title(),rolename.lower()],Active=True).exclude(id=roleid).first()
            if roleexist is not None:
                return Response ({"n" : 0,"msg" : "Role name already exist","status" : "failure"})
            else:
                serializer = RoleSerializer(roleobj,data=data,partial=True)
                if serializer.is_valid():
                    serializer.save()
                    return Response ({"n" : 1,"msg" : "Role name edited successfully","status" : "Success"})
                else:
                    return Response ({"n" : 0,"msg" : "Could not edit Rolename","status" : "failure"})
        else:
            return Response ({"n" : 0,"msg" : "Role not found","status" : "failure"})




@api_view(['POST'])
def updatemultipleroleapi(request,format=None):
    if request.method == 'POST':
        data={}
        companycode = request.user.company_code
        rolepermissionlist =  json.loads(request.POST.get('permissionlist'))
        for r in rolepermissionlist:
            permissionobject = Permissions.objects.filter(RoleID = r['roleID']).first()
            data['RoleID'] = r['roleID']
            data['MenuID'] =  r['menuid']
            serializer = PermissionsSerializer(permissionobject,data = data,partial=True)
            if serializer.is_valid():
                serializer.save()
            # permissioncreate.MenuID.set(r['menuid'])
        return Response ({"n" : 1,"msg" : "Role name updated successfully","status" : "Success"})






        








       




        











# passsworddd---------------------------------------

@api_view(['POST'])
def passwordUpdate(request):
    data = {'n': '', 'Msg': '', 'Status': ''}
    currentpassword = request.user.password
    try:
        userID = request.query_params.get('userID')
        if userID is None:
            data['n'] = 0
            data['Msg'] = 'User ID is none'
            data['Status'] = "Failed"
            return Response(data=data)
        else:
            email = request.user.email
            oldpassword = request.data['oldpassword']

            uid_obj=Users.objects.filter(email=email,company_code=request.user.company_code,is_active=True).first()
            serializer=UserSerializer(uid_obj)
            
            uid=serializer.data['uid']

            user = authenticate(username=uid, password=oldpassword)
            if user is not None:

                r = {'password': request.data['password']}
                # p = request.data['password']
                serializer = passwordSerializer(user, r)
                # if oldpassword == currentpassword:

                if serializer.is_valid():
                    serializer.validated_data['UpdatedBy'] = request.user
                    p = serializer.validated_data['password']
                    encrypted = make_password(p)
                    serializer.validated_data['password'] = encrypted
                    serializer.validated_data['PasswordChanged'] = True
                    serializer.save()

                    request.user.auth_token.delete()

                    data['n'] = 1
                    data['Msg'] = 'update successfull'
                    data['Status'] = "Success"
                    return Response(data=data)
                else:
                    data = serializer.errors
            else:
            
                data['n'] = 0
                data['Msg'] = 'Old Password does not match'
                data['Status'] = "Failed"
                return Response(data=data)

    except Exception as e:
        print(e)
        data['n'] = 0
        data['Msg'] = 'try method failed'
        data['Status'] = "Failed"
        return Response(data=data)




@csrf_exempt
@api_view(["POST"])
@permission_classes((AllowAny,))
def chkEmailExists(request):
    data = {}
    try:
        email = request.data.get('email', None)
        if email is None or email == '':
            return Response({'n': 0, 'Msg': 'Please enter email', 'Status': 'Failed'})
        if not Users.objects.filter(email=email,is_active=True).first():
            data['n'] = 0
            data['Msg'] = 'No user with this Email found'
            data['Status'] = "Failed"
            return Response(data)
        user_obj = Users.objects.filter(email=email,is_active=True).first()

        token = str(uuid.uuid4())
        try:
            profile_obj = Profile.objects.get(user=user_obj)
        except Exception as e:
            print(e)
            return Response({'n': 0, 'Msg': 'Profile object with this user ID not found', 'Status': 'Failed'})
        else:
            profile_obj.forget_password_token = token
            profile_obj.updatedOn = timezone.now()
            profile_obj.save()

        curruser = Users.objects.filter(email=email,is_active=True).annotate(
            Name=Concat('Firstname', Value(' '), 'Lastname'))
        Name = curruser.first().Name
        token = {'token': token, 'Name': Name,'frontUrl':frontUrl}
        message = get_template('forgot-password-email-template.html').render(token)
        subject = "Forgot Password?"
        msg = EmailMessage(
            subject,
            message,
            EMAIL_HOST_USER,
            [email],
        )
        msg.content_subtype = "html"  # Main content is now text/html
        msg.send()
        data['n'] = 1
        data['Msg'] = 'Email has been sent'
        data['Status'] = "Success"
        return Response(data)
    except Exception as e:
        print(e)
        return Response({'n': 0, 'Msg': 'Email could not be sent', 'Status': 'Failed'})


@csrf_exempt
@api_view(["POST"]) 
@permission_classes((AllowAny,))
def resetPassword(request):
    data = {'n': '', 'Msg': '', 'Status': ''}
    try:
        newpassword = request.data.get('newpassword')
        token = request.data.get('token')
        profile_obj = Profile.objects.get(forget_password_token=token)
        userID = profile_obj.user.id
        
        user = Users.objects.get(id=userID)
        
        user.set_password(newpassword)
        user.save()

    except Exception as e:
      
        data['n'] = 0
        data['Msg'] = 'Exception occured, Please try again'
        data['Status'] = "Failed"
        return Response(data)
    else:
        data['n'] = 1
        data['Msg'] = 'Password successfully changed'
        data['Status'] = "Success"
        return Response(data)


@api_view(["POST"])
@permission_classes((AllowAny,))
def checkIfForgotPasswordToken(request):
    data = {'n': '', 'Msg': '', 'Status': ''}
    try:
        token = request.data.get('token')
        profile_obj = Profile.objects.get(forget_password_token=token)
        time_elapsed = timezone.now()-profile_obj.updatedOn
        # time_elapsed = diff_in_time(profile_obj.updatedOn,datetime.now())
        seconds = time_elapsed.total_seconds()
        # hours = seconds // 3600
        # minutes = (seconds % 3600) // 60
        # seconds = seconds % 60
        if seconds > 1800:
            return Response({'n': "0", 'Msg': 'Time limit exceded', 'Status': ''})
        else:
            return Response({'n': "1", 'Msg': 'Within Time Limit', 'Status': ''})

        # token = request.data.get('token')
        # profile_obj = Profile.objects.get(forget_password_token = token)
        # if profile_obj:
        #     return Response({'n':"1",'Msg':'Token exists','Status':''})
        # else:
        #     return Response({'n':"0",'Msg':'Token does not exist','Status':''})
    except Exception as e:
        return Response({'n': "0", 'Msg': 'exception', 'Status': ''})


def human_format(num):
    num = float('{:.3g}'.format(num))
    magnitude = 0
    while abs(num) >= 1000:
        magnitude += 1
        num /= 1000.0
    return '{}{}'.format('{:f}'.format(num).rstrip('0').rstrip('.'), ['', 'K', 'M', 'B', 'T'][magnitude])


# get dashboard details for employee


@api_view(['POST'])
def dashboarddata(request):
    my_date = datetime.date.today()
    year, week_num, day_of_week = my_date.isocalendar()

    week = week_num
    userID = request.data.get('userID', None)
    userDept = Users.objects.filter(id=userID,company_code=request.user.company_code,is_active=True).values('DepartmentID')
    deptname = Department.objects.filter(id__in=userDept).values('DepartmentName')
    if userDept.exists() and deptname.exists():

        userDept = userDept.first()['DepartmentID']
        deptname = deptname.first()['DepartmentName']
        try:
            # user data
            top = IntermediateTrackProResult.objects.filter(
                Year=year, Week=week).order_by('TrackProPercent').reverse()[:5]
            topserializer = IntermediateGetTrackProResultSerializer(top, many=True)
        
            taskss = TaskMaster.objects.filter(
                AssignTo=userID, Year=year, Week=week).count()
            tasks = human_format(taskss)
            totaltask = TaskMaster.objects.filter(AssignTo=userID).count()
            trackpropercent = IntermediateTrackProResult.objects.filter(
                EmpID=userID, Year=year, Week=week).order_by('TrackProPercent').reverse()
            trackproserializer = IntermediateGetTrackProResultSerializer(
                trackpropercent, many=True)
            trackpropercentavg = IntermediateTrackProResult.objects.filter(
                Employee=userID).aggregate(Avg('TrackProPercent'))
            # trackpro rank
            trackprorank = IntermediateTrackProResult.objects.filter(
                Year=year, Week=week).order_by('TrackProPercent').reverse()
            trackprorankserializer = IntermediateGetTrackProResultSerializer(
                trackprorank, many=True)
            # data for graph

            # department wise
        except Exception as e:
            return Response({"n": 0, "Msg": "no data found", "Status": "Failed"})
        try:
            conn = psycopg2.connect(database=env('DATABASE_NAME'), user= env('DATABASE_USER'),
                                    password=env('DATABASE_PASSWORD'), host=env('DATABASE_HOST'), port=env('DATABASE_PORT'))
            cur = conn.cursor()
        except Exception as e:
            return Response({"n": 0, "Msg": "Could not connect to database", "Status": "Failed"})
        else:
            query = """ SELECT d."DepartmentName",ud."department_id",cw."Week",                        
                        (ROUND( CAST (AVG(cw."TrackProPercent") as NUMERIC),2))
                        FROM "Users_users_DepartmentID" as ud
                        INNER JOIN "Users_users" as u ON
                        u."id" = ud."users_id"
                        INNER JOIN "Department_department" as d ON
                        d."id" = ud."department_id"
                        INNER JOIN "CheckTrackPro_intermediatetrackproresult" as cw ON
                        cw."Employee_id" = u."id" WHERE cw."Year" = {}
                        GROUP BY d."DepartmentName",ud."department_id",cw."Week"
                        ORDER BY ud."department_id",cw."Week" """.format(year)
            cur.execute(query)
            trackpropercent_Dept = cur.fetchall()

            trackpropercent_Dept_dict = {'Department': deptname, 'trackpropercent': [
                s for p, q, r, s in trackpropercent_Dept if q == userDept], 'Weeklist': [r for p, q, r, s in trackpropercent_Dept if q == userDept]}
            # individual
            trackpropercent_individual = []
            for week in trackpropercent_Dept_dict['Weeklist']:
                query = IntermediateTrackProResult.objects.filter(
                    Year=year, EmpID=userID, Week=week).values('TrackProPercent')
                if query:
                    trackpropercent_individual.append(query.first()['TrackProPercent'])
                else:
                    trackpropercent_individual.append(None)

            return Response({"n":1,"Tasks": tasks, "TotalTask": totaltask,
                            "topfive": topserializer.data, "trackpropercent": trackproserializer.data,
                            "trackproavg": trackpropercentavg, "trackprorank": trackprorankserializer.data,
                            "trackpropercent_individual": trackpropercent_individual, 'trackpropercent_Dept': trackpropercent_Dept_dict,
                            })
    else:
        return Response({"n":0,"all_dept_graph_data":"","all_dept_weeklist": [],"trackpropercent_individual": "","trackpropercent_Dept":''})

@api_view(['POST'])
def Admindashboarddata(request):
    my_date = date.today()
    year, week_num, day_of_week = my_date.isocalendar()
    week = week_num
 
    userID = request.data.get('userID', None)
    try:
        tasks = TaskMaster.objects.filter(
            AssignTo=userID, Year=year, Week=week).count()
        totaltask = TaskMaster.objects.filter(AssignTo=userID).count()

        trackpropercent = IntermediateTrackProResult.objects.filter(
            EmpID=userID, Year=year, Week=week).order_by('TrackProPercent').reverse()
        trackproserializer = IntermediateGetTrackProResultSerializer(
            trackpropercent, many=True)
      
        # trackpro rank
        trackprorank = IntermediateTrackProResult.objects.filter(
            Year=year, Week=week).order_by('TrackProPercent').reverse()
        trackprorankserializer = IntermediateGetTrackProResultSerializer(
            trackprorank, many=True)
        # data for graph
        lasttask = TaskMaster.objects.filter(AssignTo=userID).order_by('CreatedOn').last()
        lasttaskser = GetTaskMasterSerializer(lasttask)
        # department wise

        #rank card 

        my_date = datetime.date.today()
        year, week_num, day_of_week = my_date.isocalendar()
        curryear = year
        currentweek = week_num
        prevWeek = currentweek - 1
        com_code = request.user.company_code
        loguserid = request.user.id
        rankdata = []
        rankdatadict = {}
        userprevweekdata = []
        userprevweekdict = {}

        employeeId = request.user.id 
        avaliable = 12
        starting_month = [4,5,6,7,8,9]

        month = date.today()
        current_month = month.month
    
        current_year = month.year
        if 4 <= current_month <= 9:
            monthstr = "April-September"
            pending_leave_countobj = Leave.objects.filter(employeeId=employeeId,WorkFromHome=False,start_date__month__gte=4,start_date__month__lte=9,start_date__year= current_year,Active=True,leave_status='Pending').count()
            approve_leave_countobj = Leave.objects.filter(employeeId=employeeId,WorkFromHome=False,start_date__month__gte=4,start_date__month__lte=9,start_date__year= current_year,Active=True,leave_status='Approved').count()
            rejected_leave_countobj = Leave.objects.filter(employeeId=employeeId,WorkFromHome=False,start_date__month__gte=4,start_date__month__lte=9,start_date__year= current_year,Active=True,leave_status='Rejected').count()
            total_count = Leave.objects.filter(employeeId=employeeId,WorkFromHome=False,start_date__month__gte=4,start_date__month__lte=9,start_date__year= current_year,Active=True).exclude(leave_status='Draft').count()
        else:
            monthstr = "October-March"
            pending_leave_countobj = Leave.objects.filter(employeeId=employeeId,WorkFromHome=False,start_date__month__in=[10,11,12,1,2,3],start_date__year= current_year,Active=True,leave_status='Pending').count()
            approve_leave_countobj = Leave.objects.filter(employeeId=employeeId,WorkFromHome=False,start_date__month__in=[10,11,12,1,2,3],start_date__year= current_year,Active=True,leave_status='Approved').count()
            rejected_leave_countobj = Leave.objects.filter(employeeId=employeeId,WorkFromHome=False,start_date__month__in=[10,11,12,1,2,3],start_date__year= current_year,Active=True,leave_status='Rejected').count()
        
            total_count  = Leave.objects.filter(employeeId=employeeId,WorkFromHome=False,start_date__month__in=[10,11,12,1,2,3],start_date__year= current_year,Active=True).exclude(leave_status='Draft').count()


        prevweekobj = IntermediateTrackProResult.objects.filter(EmpID=loguserid,company_code=com_code,Week=prevWeek,Year=curryear).first()
        if prevweekobj is not None:
            perct = prevweekobj.TrackProPercent 
            if perct is not None:
                userprevweekdict['prevweekperc'] = str(perct) + " %"
            else:
                userprevweekdict['prevweekperc'] = "--"

            if prevweekobj.Rank is not None:
                userprevweekdict['prevweekrank'] = prevweekobj.Rank 
            else:
                userprevweekdict['prevweekrank'] = "--"
        else:
            userprevweekdict['prevweekperc'] = "--"
            userprevweekdict['prevweekrank'] = "--"

        userprevweekdata.append(userprevweekdict)

        context = {
            'userprevweekdata':userprevweekdata,
            'approved_leave_count':approve_leave_countobj,'rejected_leave_count':rejected_leave_countobj,
            'monthstr':monthstr,
            'pending_leave_count':pending_leave_countobj,
            'total_leave_count':total_count,
            "Tasks": tasks, 
            "TotalTask":totaltask,
            "trackpropercent": trackproserializer.data,
            "trackprorank": trackprorankserializer.data,
            "all_dept_graph_data":[],
            "all_dept_weeklist": [],
            "lasttask":lasttaskser.data,
            
        }


    except Exception as e:
        return Response({"n":0,"data":{            
            'userprevweekdata':[],
            'approved_leave_count':0,
            'rejected_leave_count':0,
            'monthstr':'',
            'pending_leave_count':0,
            'total_leave_count':0,
            "Tasks": 0, 
            "TotalTask":0,
            "trackpropercent": [],
            "trackprorank": [],
            "all_dept_graph_data":[],
            "all_dept_weeklist": [],
            "lasttask":[],
            }})
    
    return Response({"n":1,'data':context})
 


 
# @api_view(['POST'])
# def usersDashboard(request):
#     projectId = request.POST.get('projectId')
#     companycode = request.user.company_code
#     companycode = "'"+companycode+"'"
#     try:
#         conn = psycopg2.connect(database=env('DATABASE_NAME'), user= env('DATABASE_USER'),
#                                     password=env('DATABASE_PASSWORD'), host=env('DATABASE_HOST'), port=env('DATABASE_PORT'))
#         cur = conn.cursor()
#     except Exception as e:
#         return Response({"n": 0, "Msg": "Could not connect to database", "Status": "Failed"})
#     else:
#         Project = ""
#         if projectId is not None:
#             Project = 'and tk."Project_id" = '+projectId+''
#         else:
#             Project = ""
#         query = """ SELECT
#                     u."id",u."Firstname",u."Lastname",u."company_code",u."Photo",
#                     tk."Project_id",tk."ProjectName",tk."TaskTitle",
#                     td."Task_id",
#                     tk."Active",
#                     ud."DesignationName",
#                     SUM(EXTRACT(EPOCH FROM (COALESCE(td."EndDate", Current_timestamp) - td."StartDate"))),
#                     u."email"
#                 FROM "Users_users" as u 
#                 LEFT JOIN "Tasks_taskmaster" as tk on tk."AssignTo_id" = u."id" AND tk."Active" = True
#                 LEFT JOIN "Tasks_projecttasks" as td on td."Task_id" = tk."id" """+Project+"""
#                 LEFT JOIN "Users_designation" as ud on ud."id" = u."DesignationId_id"
#                 where u."is_active" = 'True' and u."is_admin" = 'True' and u."company_code" = """+companycode+""" GROUP BY u."id",td."Task_id",u."Firstname",u."Lastname",u."company_code",tk."Project_id",tk."ProjectName",tk."TaskTitle",tk."Active",u."Photo",ud."DesignationName"
#                 ORDER BY tk."ProjectName" """
#         cur.execute(query)
#         usersDash = cur.fetchall()
#         return Response({"UserDashboard": usersDash})



@api_view(['POST'])
def usersDashboard(request):
    projectId = request.POST.get('projectId')
    companycode = request.user.company_code
    departmentId = request.POST.get('DepartmentId')
    Usersobj = Users.objects.filter(is_active=True,company_code=companycode)
    userser = UserSerializer(Usersobj,many=True)
    for u in userser.data:
        if u['DepartmentID'] is not None:
            for ud in u['DepartmentID']:
                u['redepartmentid'] = ud
        else:
            u['redepartmentid'] = u['DepartmentID']

        Tasksobj = TaskMaster.objects.filter(AssignTo=u['id'],Active=True,company_code=companycode).first()
        if Tasksobj is not None:
            taskser = GetTaskMasterSerializer(Tasksobj)
            for t in [taskser.data]:
                u['projectname'] = t['ProjectName']
                u['Projectid'] = t['Project']
                u['taskdetail'] = t['TaskTitle']
                u['status'] = 1
              
                currentzone = pytz.timezone("Asia/Kolkata") 
                currenttime = datetime.datetime.now(currentzone)
                newcurrenttime = currenttime.strftime("%Y-%m-%dT%H:%M:%S.%f%z")

                projecttasktime = ProjectTasks.objects.filter(Task=t['id']).order_by("id")
                if projecttasktime is not None:
                    projectser = ProjectTasksSerializer(projecttasktime,many=True)
                    totaltime=0
                    for o in projectser.data:
                        startstring = o['StartDate']
                        starttime=startstring
                        t1 = datetime.datetime.strptime(starttime, "%Y-%m-%dT%H:%M:%S.%f%z")

                        endstring = o['EndDate']
                        if endstring is not None:
                            endtime = o['EndDate']
                        else:
                            endtime = str(newcurrenttime)
                        t2=datetime.datetime.strptime(endtime, "%Y-%m-%dT%H:%M:%S.%f%z")
                        tdelta=t2-t1
                    
                    
                        if "day" in str(tdelta) or "days" in str(tdelta):
                            daystring = str(tdelta).split(",")[0]
                            noofdays = str(daystring).split(" ")[0]
                            daysmins = int(noofdays)*1440

                            thoursstr =  str(tdelta).split(",")[1]
                            thours = str(thoursstr).split(":")[0]
                            hrs = int(thours)*60
                            tmins = str(thoursstr).split(":")[1]
                            finalmins = int(hrs)+int(tmins)+int(daysmins)
                        else:
                            thours = str(tdelta).split(":")[0]
                            hrs = int(thours)*60
                            tmins = str(tdelta).split(":")[1]
                            finalmins = int(hrs)+int(tmins)
                        totaltime += finalmins
                   

                    totalhours =totaltime
                    hour = int (totalhours) // 60
                    if (len(str(hour)) < 2):
                        hours = "0"+str(hour)
                    else:
                        hours = str(hour)

                    mins = int (totalhours) % 60
                    if (len(str(mins)) < 2):
                        minutes = "0"+str(mins)
                    else:
                        minutes = str(mins)

                    hourstring = str(hours)+":"+str(minutes)+" "+"Hrs"

                    u['taskhours'] = hourstring

        else:
            u['taskhours'] = "--"
            u['projectname'] = "--"
            u['Projectid'] = "--"
            u['taskdetail'] = "--"
            u['status'] = 2
           
    
    userlist = sorted(userser.data, key=lambda x: x['status'])
    return Response({"n": 1, "Msg": "date found successfully", "Status": "Success",'data':userlist}, status=status.HTTP_201_CREATED)
    





    




















# def week_number_of_month(date_value):
#      return (date_value.isocalendar()[1] - date_value.replace(day=1).isocalendar()[1] + 1)

def week_of_month(dt):
    first_day = dt.replace(day=1)

    dom = dt.day
    adjusted_dom = dom + first_day.weekday()

    return int(ceil(adjusted_dom/7.0))



@api_view(['POST'])
def newfiledata(request):
    file = request.FILES['datafile']
    com_code = request.user.company_code
    namefile = file.name
    split_tup = os.path.splitext(namefile)
    file_name = split_tup[0]
    file_extension = split_tup[1]
    

    if file_extension == ".dat" or file_extension == ".txt":
        col_names = ["id", "datetime","a","b", "c", "d"]
        df = pd.read_table(file,sep="\t",names=col_names)
        df.to_excel('data.xlsx')
        filepath = open('data.xlsx',"rb")
        dataset = Dataset()
        imported_data = dataset.load(filepath.read(),format='xlsx')

        importDataList=[]
        for i in imported_data:
            if i[1] is not None:
                importDataList.append(i)

        # com_code = request.user.company_code
        for i in importDataList:
          
            dt=i[2]
            if dt is None:
                return Response({"response":{'n': 0,'msg':'invalid data in upload file','Status': 'warning'}})

            # dt=[int(l) for l in DT.split()]
            Date=dt.split(" ")[0]
            Time=dt.split(" ")[1]
            nYear = str(Date).split("-")[0]
            nMonth = str(Date).split("-")[1]
            nDay = str(Date).split("-")[2]
            date_given = datetime.datetime(year=int(nYear), month=int(nMonth), day=int(nDay)).date()
            nweek = week_of_month(date_given)


            empexist = attendance.objects.filter(employeeId=i[1],date=Date,time=Time,company_code=com_code).order_by('time').first()
            if empexist is None:
                attendance.objects.create(employeeId =i[1],date=Date,time=Time,filename=namefile,company_code=com_code,Week=nweek,Month=nMonth,Year=nYear)

        return Response({"response":{'n': 1,'msg':'filedata saved successfully','Status': 'success'}})
    else:
        return Response({"response":{'n': 0,'msg':'file type is not dat file ','Status': 'warning'}})


@api_view(['POST'])
@permission_classes((AllowAny,))
def appendfiledata(request):
    file = request.FILES['file']
    namefile = file.name
    split_tup = os.path.splitext(namefile)
    file_name = split_tup[0]
    file_extension = split_tup[1]
    
   
    if file_extension == ".dat" or file_extension == ".txt":
        empdata = attendance.objects.filter(filename=namefile)
       
        if empdata:
            empdata.delete()

            col_names = ["id", "datetime","a", "b", "c", "d"]
            df = pd.read_table(file,sep="\t",names=col_names)

            df.to_excel('data.xlsx')

            filepath = open('data.xlsx',"rb")
            dataset = Dataset()
            imported_data = dataset.load(filepath.read(),format='xlsx')

            importDataList=[]
            for i in imported_data:
                if i[1] is not None:
                    importDataList.append(i)
            
            com_code = request.user.company_code 
            for i in importDataList:
                dt=i[2]

                # dt=[int(l) for l in DT.split()]
                Date=dt.split(" ")[0]
                Time=dt.split(" ")[1]
                nYear = Date.split("-")[0]
                nMonth = Date.split("-")[1]
                nDay = Date.split("-")[2]

                date_given = datetime.datetime(year=int(nYear), month=int(nMonth), day=int(nDay)).date()
                nweek = week_of_month(date_given)

                attendance.objects.create(employeeId=i[1],date=Date,time=Time,filename=namefile,company_code=com_code,Week=nweek,Month=nMonth,Year=nYear)

            return Response({"response":{"n": 1 ,"msg" : "File updated successfully","status":"success"}})
        else:
            return Response({"response":{"n": 0 ,"msg" : "File data does not exist","status":"error"}})  
    else:
        return Response({"response":{"n": 0 ,"msg" : "File format is not dat file","status":"error"}})  
        



@api_view(['POST'])
@permission_classes((AllowAny,))
def monthlydata(request):

    listall=[]
    yearlist=[]
    listWO = []
    holidaylist=[]
    emplist=[]
    datetimelist=[] 
    d2=[]
    d6=[]
    list1=[]
    weeklylist=[]

    sdate = request.data.get('sdate')
    edate = request.data.get('edate')
    if sdate not in [None, ""] and edate not in [None, ""]:
        sYear = int(sdate.split("-")[2])
        sMonth = int(sdate.split("-")[1])
        sDay = int(sdate.split("-")[0])
        eYear = int(edate.split("-")[2])
        eMonth = int(edate.split("-")[1])
        eDay = int(edate.split("-")[0])
        csDate = date(sYear, sMonth, sDay)

        cedate = date(eYear, eMonth, eDay) 
        
        delta = cedate - csDate 
        yearlist.append(sYear)
        yearlist.append(eYear)

        allyearList = list(set(yearlist))

        #LIST OF ALL DATES IN DATERANGE
        for i in range(delta.days + 1):
            day = csDate + timedelta(days=i)
            dated= day.strftime("%Y-%m-%d")
            listall.append(dated)   
      
        
        #sat and sunday between daterange
        for i in allyearList:
            dt=date(i,sMonth,1)  
            first_w=dt.isoweekday() 
            if(first_w==7): 
                first_w=0
            saturday2=14-first_w
            dt1=date(i,sMonth,saturday2)
            listWO.append(dt1)
            saturday4=28-first_w  
            dt2=date(i,sMonth,saturday4)
            listWO.append(dt2)


        def allsundays(year):
            d = date(year, 1, 1)                 
            d += timedelta(days = 6 - d.weekday())  
            while d.year == year:
                yield d
                d += timedelta(days = 7)

        for i in allyearList:
            for d in allsundays(i):
                listWO.append(d)
       

        for i in listWO:
            year = i.strftime("%Y")
            month = i.strftime("%m")
            day = i.strftime("%d")
            final_date=year+"-"+month+"-"+day
            weeklylist.append(final_date)
           
        #yearly holiday list
        holidatlist = Holidays.objects.filter(Active=True).order_by('id')
        serializer = holidaysSerializer(holidatlist, many=True)
        for i in serializer.data:
            holiday=i['Date']
            holidaylist.append(holiday)
        
       
        
        #data of employee id
        empdata = Users.objects.exclude(employeeId=None)
        mon_serializer = monthlydataSerializer(empdata,many=True)
        for i in mon_serializer.data:
            emplist.append(i['employeeId'])
       
        # data of employee 
        for p in emplist:
            att_data=attendance.objects.filter(employeeId=p,date__range=(csDate,cedate))#previous month validation
            serializer = attendanceserializer(att_data,many=True)
            data={
            'employeeId':p,
            'records':serializer.data
            }
            datetimelist.append(data)
       
        

        for p in datetimelist:
            A_count=0
            holiday_count=0
            WeeklyOffcount = 0
            Halfdaycount = 0
            for j in listall:
                year, month, day = j.split('-') 
                wname = datetime.date(int(year), int(month), int(day))
                day_name=wname.strftime("%a")

                #first and last time records filter
                loginattObj=attendance.objects.filter(employeeId=p['employeeId'],date=j).order_by('time').first()
                logoutattObj=attendance.objects.filter(employeeId=p['employeeId'],date=j).order_by('time').last()
                if loginattObj is not None:
                    logint=loginattObj.time
                    t1 = datetime.datetime.strptime(logint, "%H:%M:%S")
                
                    logoutt=logoutattObj.time
                    t2 = datetime.datetime.strptime(logoutt, "%H:%M:%S")

                    timeDiff = t2 - t1  #calculated timedifference(Total)
                  
                    time=str(timeDiff)
                    onediff= "1:00:00"
                    one=datetime.datetime.strptime(onediff, "%H:%M:%S")
                    fixtime = "5:30:00"
                    fix=datetime.datetime.strptime(fixtime, "%H:%M:%S")
                    extime= "0:00:00"
                    ex=datetime.datetime.strptime(extime, "%H:%M:%S")
                    tdiff= fix - ex
                    odiff= one - ex

                    halfdaylogin = "11:00:00"
                    halfdaylogintime=datetime.datetime.strptime(halfdaylogin, "%H:%M:%S")

                    halfdaylogout = "17:00:00"
                    halfdaylogouttime=datetime.datetime.strptime(halfdaylogout, "%H:%M:%S")


                   
                    if logint==logoutt:
                        time='00:00:00' 
                        logoutt='-'
                        status = j.split("-")[2]+" "+day_name+" "+"A"
                        hello=status.endswith("A")
                        if hello is True:
                            A_count+=0.5
                        else:
                            A_count+=0
                    elif t1 > halfdaylogintime or t2 < halfdaylogouttime:
                        if timeDiff < odiff:
                            time=str(timeDiff) 
                            logoutt=logoutt
                            status = j.split("-")[2]+" "+day_name+" "+"A"
                            hello=status.endswith("A")
                            if hello is True:
                                A_count+=1
                            else:
                                A_count+=0
                        else:
                            time=str(timeDiff) 
                            logoutt=logoutt
                            status = j.split("-")[2]+" "+day_name+" "+"HF"
                            half=status.endswith("HF")
                            if half is True:
                                Halfdaycount+=1
                            else:
                                Halfdaycount+=0
                    else:
                        status = j.split("-")[2]+" "+day_name+" "+"P"
                    
                    d1={
                        "employeeId":p['employeeId'],
                        "date":status,
                        "inTime":logint,
                        "outTime":logoutt,
                        "Total":time,
                    }
                    d2.append(d1)
                else:
                    if j in weeklylist:
                        logint=""
                        logoutt=""
                        time=""
                        status=j.split("-")[2]+" "+day_name+" "+"WO"
                        woff=status.endswith("WO")
                        if woff is True:
                            WeeklyOffcount+=1
                        else:
                            WeeklyOffcount+=0
                        d1={
                            "employeeId":p['employeeId'],
                            "date":status,
                            "inTime":logint,
                            "outTime":logoutt,
                            "Total":time
                        }
                        d2.append(d1)
                    elif j in holidaylist:
                        logint=""
                        logoutt=""
                        time=""
                        status=j.split("-")[2]+" "+day_name+" "+"H"
                        hello=status.endswith("H")
                        if hello is True:
                            holiday_count+=1
                    
                        d1={
                            "employeeId":p['employeeId'],
                            "date":status,
                            "inTime":logint,
                            "outTime":logoutt,
                            "Total":time
                        }
                        d2.append(d1)
                    else:
                        logint="-"
                        logoutt="-"
                        time="-"
                        status=j.split("-")[2]+" "+day_name+" "+"A"
                        hello=status.endswith("A")
                        if hello is True:
                            A_count+=1
                        else:
                            A_count+=0
                        d1={
                            "employeeId":p['employeeId'],
                            "date":status,
                            "inTime":logint,
                            "outTime":logoutt,
                            "Total":time
                        }
                        d2.append(d1)
        
            d5={
                'employeeId':p['employeeId'],
                'Absent':A_count,
                'Holiday':holiday_count,
                'Weeklyoff':WeeklyOffcount,
                'HalfDay':Halfdaycount,
                'records':dict((v['date'],v) for v in d2).values() 
            }
            d6.append(d5)
            d2.clear()

        if os.path.exists("static/excel/monthlydata.xlsx"):
            os.remove("static/excel/monthlydata.xlsx")
            workbook = xlsxwriter.Workbook('static/excel/monthlydata.xlsx')
            workbook.close()
        else:
          
            workbook = xlsxwriter.Workbook('static/excel/monthlydata.xlsx')
            workbook.close()

        attendancereport(d6)#calling function
        
        return Response({"response":{"data":d6,"url":imageUrl + '/static/excel/monthlydata.xlsx', "n":2 ,"msg" : "Operation successful","status":"success"}})  
    else:
      
        emplist=[]
        datetimelist=[] 
        d2=[]
        d6=[]
        list1=[]
        holidaylist=[]
        empdata = Users.objects.exclude(employeeId=None)
        mon_serializer = monthlydataSerializer(empdata,many=True)
        for i in mon_serializer.data:
            emplist.append(i['employeeId'])

       
        for p in emplist:
            currentMonth = date.today().month
            prevmonth = currentMonth-1
            att_data=attendance.objects.filter(employeeId=p,date__month=prevmonth) #previous month validation
            serializer = attendanceserializer(att_data,many=True)
            data={
            'employeeId':p,
            'records':serializer.data
            }
            datetimelist.append(data)

        month = prevmonth
        year = datetime.datetime.now().year
        number_of_days = calendar.monthrange(year, month)[1]
        first_date = date(year, month, 1)
        last_date = date(year, month, number_of_days)
        delta = last_date - first_date
        dayslist=[(first_date + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(delta.days + 1)]
        

        year=date.today().year 
        month=prevmonth
        dt=date(year,month,1)   # first day of month
        first_w=dt.isoweekday()  # weekday of 1st day of the month
        if(first_w==7): # if it is Sunday return 0 
            first_w=0
        saturday2=14-first_w
        dt1=date(year,month,saturday2)
        list1.append(str(dt1))
        saturday4=28-first_w
        dt1=date(year,month,saturday4)
        list1.append(str(dt1))  
        weeklyofflist=list1
        

        def allsundays(month):
            d = date(year,month, 1) # day 1st of month
            d += timedelta(days = 6 - d.weekday())  # First Sunday
            while d.month == month:
                yield d
                d += timedelta(days = 7)

        for d in allsundays(month):
            d=str(d)
            weeklyofflist.append(d)
        
    

        holidatlist = Holidays.objects.filter(Active=True).order_by('id')
        serializer = holidaysSerializer(holidatlist, many=True)
        for i in serializer.data:
            holiday=i['Date']
            holidaylist.append(holiday)
    
        
        for p in datetimelist:
            WeeklyOffcount=0
            Halfdaycount=0
            A_count=0
            holiday_count=0
            for j in dayslist:
                year, month, day = j.split('-') 
                wname = datetime.date(int(year), int(month), int(day))
                day_name=wname.strftime("%a")
                
                
                #first and last time records filter
                loginattObj=attendance.objects.filter(employeeId=p['employeeId'],date=j).order_by('time').first()
                logoutattObj=attendance.objects.filter(employeeId=p['employeeId'],date=j).order_by('time').last()
                if loginattObj is not None:
                    logint=loginattObj.time
                    t1 = datetime.datetime.strptime(logint, "%H:%M:%S")
                
                    logoutt=logoutattObj.time
                    t2 = datetime.datetime.strptime(logoutt, "%H:%M:%S")

                    timeDiff = t2 - t1  #calculated timedifference(Total)
                  
                    time=str(timeDiff)
                    onediff= "1:00:00"
                    one=datetime.datetime.strptime(onediff, "%H:%M:%S")
                    fixtime = "5:30:00"
                    fix=datetime.datetime.strptime(fixtime, "%H:%M:%S")
                    extime= "0:00:00"
                    ex=datetime.datetime.strptime(extime, "%H:%M:%S")
                    tdiff= fix - ex
                    odiff= one - ex
                   
                    halfdaylogin = "11:00:00"
                    halfdaylogintime=datetime.datetime.strptime(halfdaylogin, "%H:%M:%S")

                    halfdaylogout = "17:00:00"
                    halfdaylogouttime=datetime.datetime.strptime(halfdaylogout, "%H:%M:%S")

                    if logint==logoutt:
                        time='00:00:00' 
                        logoutt='-'
                        status = j.split("-")[2]+" "+day_name+" "+"A"
                        hello=status.endswith("A")
                        if hello is True:
                            A_count+=0.5
                        else:
                            A_count+=0
                    elif t1 > halfdaylogintime or t2 < halfdaylogouttime:
                        if timeDiff < odiff:
                            time=str(timeDiff) 
                            logoutt=logoutt
                            status = j.split("-")[2]+" "+day_name+" "+"A"
                            hello=status.endswith("A")
                            if hello is True:
                                A_count+=1
                            else:
                                A_count+=0
                        else:
                            time=str(timeDiff) 
                            logoutt=logoutt
                            status = j.split("-")[2]+" "+day_name+" "+"HF"
                            half=status.endswith("HF")
                            if half is True:
                                Halfdaycount+=1
                            else:
                                Halfdaycount+=0
                    else:
                        status = j.split("-")[2]+" "+day_name+" "+"P"
                    
                    d1={
                        "employeeId":p['employeeId'],
                        "date":status,
                        "inTime":logint,
                        "outTime":logoutt,
                        "Total":time,
                    }
                    d2.append(d1)
                else:
                    if j in weeklyofflist:
                        logint=""
                        logoutt=""
                        time=""
                        status=j.split("-")[2]+" "+day_name+" "+"WO"
                        woff=status.endswith("WO")
                        if woff is True:
                            WeeklyOffcount+=1
                        else:
                            WeeklyOffcount+=0
                        d1={
                            "employeeId":p['employeeId'],
                            "date":status,
                            "inTime":logint,
                            "outTime":logoutt,
                            "Total":time
                        }
                        d2.append(d1)

                    elif j in holidaylist:
                        logint=""
                        logoutt=""
                        time=""
                        status=j.split("-")[2]+" "+day_name+" "+"H"
                        hello=status.endswith("H")
                        if hello is True:
                            holiday_count+=1
                    
                        d1={
                            "employeeId":p['employeeId'],
                            "date":status,
                            "inTime":logint,
                            "outTime":logoutt,
                            "Total":time
                        }
                        d2.append(d1)
                    else:
                        logint="-"
                        logoutt="-"
                        time="-"
                        status=j.split("-")[2]+" "+day_name+" "+"A"
                        hello=status.endswith("A")
                        if hello is True:
                            A_count+=1
                        else:
                            A_count+=0
                        d1={
                            "employeeId":p['employeeId'],
                            "date":status,
                            "inTime":logint,
                            "outTime":logoutt,
                            "Total":time
                        }
                        d2.append(d1)
            # dict((v['inTime'],v) for v in d2).values()
        
            d5={
                'employeeId':p['employeeId'],
                'Absent':A_count,
                'Holiday':holiday_count,
                'Weeklyoff':WeeklyOffcount,
                'HalfDay':Halfdaycount,
                'records':dict((v['date'],v) for v in d2).values() 
            }
            d6.append(d5)
            d2.clear()

        if os.path.exists("static/excel/monthlydata.xlsx"):
            os.remove("static/excel/monthlydata.xlsx")
            workbook = xlsxwriter.Workbook('static/excel/monthlydata.xlsx')
            workbook.close()
        else:
            workbook = xlsxwriter.Workbook('static/excel/monthlydata.xlsx')
            workbook.close()

        attendancereport(d6)#calling function
        return Response({"response":{"url": imageUrl +'/static/excel/monthlydata.xlsx', "n":1 ,"msg" : "Operation successful","status":"success"}})  
        


    
   

def attendancereport(d6):
    workbook = xlsxwriter.Workbook('static/excel/monthlydata.xlsx')
    wb=load_workbook('static/excel/monthlydata.xlsx')
    printdate=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    sheet=wb.worksheets[0]
    k=6
    z=5

    for p in d6:
        sheet.cell(row=1,column=12).value='Monthly Status Report'
        sheet.cell(row=3,column=1).value='Company:'
        sheet.cell(row=3,column=2).value='Zentro'
        sheet.cell(row=3,column=26).value='printed On:'
        sheet.cell(row=3,column=27).value=printdate

        sheet.cell(row=z,column=1).value='Empcode'
        sheet.cell(row=z,column=2).value=p['employeeId']
        empobj=Users.objects.filter(employeeId=p['employeeId']).first()
        sheet.cell(row=z,column=5).value='Empname'
        sheet.cell(row=z,column=6).value=empobj.Firstname+" "+empobj.Lastname
        sheet.cell(row=z,column=10).value='Absent'
        sheet.cell(row=z,column=11).value=p['Absent']
        sheet.cell(row=z,column=14).value='Holidays'
        sheet.cell(row=z,column=15).value=p['Holiday']
        sheet.cell(row=z,column=18).value='WeeklyOff'
        sheet.cell(row=z,column=19).value=p['Weeklyoff']
        sheet.cell(row=z,column=22).value='HalfDays'
        sheet.cell(row=z,column=23).value=p['HalfDay']
        j=1
        sheet.cell(row=k,column=1).value='Status'
        sheet.cell(row=k+1,column=1).value='InTime'
        sheet.cell(row=k+2,column=1).value='OutTime'
        sheet.cell(row=k+3,column=1).value='Total'
        
        for i in p['records']:
            sheet.cell(row=k,column=j+1).value=i['date']
            sheet.cell(row=k+1,column=j+1).value=i['inTime']
            sheet.cell(row=k+2,column=j+1).value=i['outTime']
            sheet.cell(row=k+3,column=j+1).value=i['Total']
            j+=1       
        k+=8
        z+=8
    wb.save('static/excel/monthlydata.xlsx')

    
    

def weekday_from_date(day, month, year):
    return calendar.day_name[
        datetime.date(day=day, month=month, year=year).weekday()
    ]


@api_view(['POST'])
def add_holidays(request):
    data = {}
    data_Date = request.data.get('Date')
    jquerydate =data_Date
    data['Festival'] = request.data.get('Festival')
    data['company_code'] = request.user.company_code
    Date = str(data_Date)
    nYear = str(Date).split("-")[2]
    nMonth = str(Date).split("-")[1]
    nDay = str(Date).split("-")[0]
    data['Date'] = nYear+"-"+nMonth+"-"+nDay
    data['HolidayYear'] = int(nYear)
   

    holidobj = Holidays.objects.filter(Date=data['Date'],Active=True).first()
    if holidobj is not None:
        return Response({
            "data": '',
            "response": {
                "n": 0,
                "msg": "Holiday for "+jquerydate+"  already exist.",
                "status": "errors"
            }
        })

    date_given = datetime.datetime(year=int(nYear), month=int(nMonth), day=int(nDay)).date()
    nweek = week_of_month(date_given)
    data['Holidayweek_of_month'] = int(nweek)
    data['HolidayMonth'] = int(nMonth)

    serializer = holidaysSerializer(data=data)
    if serializer.is_valid():
        serializer.save()
        return Response({
            "data": serializer.data,
            "response": {
                "n": 1,
                "msg": "Holiday Addedd Successfully.",
                "status": "success"
            }
        })
    else:
        return Response({
            "data": serializer.errors,
            "response": {
                "n": 0,
                "msg": "Failed",
                "status": "errors"
            }
        })

@api_view(['POST'])
def update_holidays(request):
    data = {}
    data['id'] = request.data.get('id')
    holidayobj = Holidays.objects.filter(id=data['id'],Active=True).first()
    if holidayobj is not None:
        data['Date'] = request.data.get('Date')
        jquerydate = data['Date']
        data['Festival'] = request.data.get('Festival')
        data['company_code'] = request.user.company_code

        Date = str(data['Date'])
        nYear = str(Date).split("-")[2]
        nMonth = str(Date).split("-")[1]
        nDay = str(Date).split("-")[0]
        data['HolidayYear'] = int(nYear)

        data['Date'] = nYear+"-"+nMonth+"-"+nDay

        holidobj = Holidays.objects.filter(Date=data['Date'],Active=True).exclude(id=data['id']).first()
        if holidobj is not None:
            return Response({
                "data": '',
                "response": {
                    "n": 0,
                    "msg": "Holiday for "+jquerydate+"  already exist.",
                    "status": "errors"
                }
            })

        date_given = datetime.datetime(year=int(nYear), month=int(nMonth), day=int(nDay)).date()
        nweek = week_of_month(date_given)

        data['Holidayweek_of_month'] = int(nweek)
        data['HolidayMonth'] = int(nMonth)

        serializer = holidaysSerializer(holidayobj,data=data,partial=True)
        if serializer.is_valid():
            serializer.save()
        
            return Response({
                "data": serializer.data,
                "response": {
                    "n": 1,
                    "msg": "Holiday Updated Successfully.",
                    "status": "success"
                }
            })
        else:
            return Response({
                "data": serializer.errors,
                "response": {
                    "n": 0,
                    "msg": "Failed",
                    "status": "errors"
                }
            })
    else:
        return Response({
            "data": '',
            "response": {
                "n": 0,
                "msg": "id not found",
                "status": "errors"
            }
        })


@api_view(['GET'])
@permission_classes((AllowAny,))
def get_holidaydata(request):
    holidatlist = Holidays.objects.filter(Active=True).order_by('id')
    serializer = holidaysSerializer(holidatlist, many=True)
    return Response({
        "data": serializer.data,
        "response": {
            "n": 1,
            "msg": "SUCCESS",
            "status": "success"
        }
    })

@api_view(['GET'])
@permission_classes((AllowAny,))
def get_holidaylistdata(request):
    holidatlist = Holidays.objects.filter(Active=True).order_by('Date')
    serializer = holidaysSerializer(holidatlist, many=True)
    currentMonth = date.today().month
    currentYear = date.today().year
    for i in serializer.data:
        i['olddate'] = i['Date']
        nndatestr = str(i['Date'])
        nYear = nndatestr.split("-")[0]
        nMonth = nndatestr.split("-")[1]
        nDay = nndatestr.split("-")[2]
        day = weekday_from_date(day=int(nDay), month=int(nMonth), year=int(nYear))
        i['Day'] = day
        i['updateolddate'] = nndatestr.split('-')[2]+"-"+ nndatestr.split('-')[1] +"-"+ nndatestr.split('-')[0]
        nmonth_name = calendar.month_abbr[int(nndatestr.split('-')[1])]    
        newdatestr = nndatestr.split('-')[2]+" "+nmonth_name +" "+ nndatestr.split('-')[0]
        i['Date'] = newdatestr

        datemonth = int(nndatestr.split('-')[1])
        dateyear = int(nndatestr.split('-')[0])
        if datemonth == currentMonth and dateyear == currentYear:
            i['currentmonth'] = True
        else:
            i['currentmonth'] = False
            
            
    return Response({
        "data": serializer.data,
        "response": {
            "n": 1,
            "msg": "list fetched successfully",
            "status": "success"
        }
    })



@api_view(['GET'])
@permission_classes((AllowAny,))
def get_holidaybyid(request):
    data = {}
    data['id'] = request.data.get('id')
    holidatlist = Holidays.objects.filter(id=data['id'],Active=True).order_by('id')
    if holidatlist is not None:
        serializer = holidaysSerializer(holidatlist, many=True)
        return Response({
            "data": serializer.data,
            "response": {
                "n": 1,
                "msg": "holiday data fetched successfully.",
                "status": "success"
            }
        })
    else:
        return Response({
            "data": '',
            "response": {
                "n": 0,
                "msg": "id not found",
                "status": "errors"
            }
        })



@api_view(['POST'])
def delete_holidays(request):
    data = {}
    data['id'] = request.data.get('id')
    holidayobj = Holidays.objects.filter(id=data['id'],Active=True).first()
    if holidayobj is not None:
        data['Active'] = False
        serializer = holidaysSerializer(holidayobj,data=data,partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({
                "data": serializer.data,
                "response": {
                    "n": 1,
                    "msg": "Holiday deleted successfully.",
                    "status": "success"
                }
            })
        else:
            return Response({
                "data": serializer.errors,
                "response": {
                    "n": 0,
                    "msg": "Failed",
                    "status": "errors"
                }
            })
    else:
        return Response({
                "data": '',
                "response": {
                    "n": 0,
                    "msg": "id not found",
                    "status": "errors"
                }
            })







@api_view(['GET'])
def usercheckininfo(request):
    data = {}
    userID = request.user.id
    employeeId = request.user.employeeId
    my_date = datetime.date.today()
    checkinallow = False
    requestbtn = True
    locationlist = []
    current_date = datetime.date.today()
    currentzone = pytz.timezone("Asia/Kolkata") 
    currenttime = datetime.datetime.now(currentzone)
    newcurrenttime =currenttime.strftime("%H:%M%p")
    userobj = Users.objects.filter(id=userID,company_code=request.user.company_code,is_active=True).first()
    if userobj is not None:
        leaverequestobj = Leave.objects.filter(employeeId=str(userID),leave_status="Approved",WorkFromHome=False,start_date__lte = my_date,end_date__gte=my_date,Active=True)
        if leaverequestobj:
            requestbtn = False
        else:
            attpresent = attendance.objects.filter(employeeId=(employeeId),date=str(current_date)).order_by('time').first()
            if attpresent is not None:
                requestbtn = True
            else:
                requestbtn = False
      



        locationobjs = Location.objects.filter(Active=True,lattitude__isnull = False,longitude__isnull=False)
        locationser = LocationSerializer(locationobjs,many=True)
        for l in locationser.data:
            locationcoords = {}
            locationcoords['lattitude'] = l['lattitude']
            locationcoords['longitude'] = l['longitude']
            locationcoords['Radius'] = l['meter']

            locationlist.append(locationcoords)



        if userobj.typeofwork is not None:
            TypeOfWork = userobj.typeofwork
            checkinallow = True

            if TypeOfWork == "2" :
                
                workfromhomeobj = Leave.objects.filter(employeeId=str(userID),WorkFromHome=True,leave_status="Approved",start_date__lte = my_date,end_date__gte=my_date,Active=True)
                if workfromhomeobj.exists():
                    if userobj.res_lattitude and userobj.res_longitude is not None:
                        locationcoords = {}
                        locationcoords['lattitude'] = userobj.res_lattitude
                        locationcoords['longitude'] = userobj.res_longitude
                        locationcoords['Radius'] = userobj.res_radius
                        locationlist.append(locationcoords) 
                      
            else:        
                
                if userobj.res_lattitude and userobj.res_longitude is not None:
                    locationcoords = {}
                    locationcoords['lattitude'] = userobj.res_lattitude
                    locationcoords['longitude'] = userobj.res_longitude
                    locationcoords['Radius'] = userobj.res_radius
                    locationlist.append(locationcoords) 
            
            
            
                
        else:
            checkinallow = False

        context = {
            'checkinallowed':checkinallow,
            'requestbtn':requestbtn,
            'remotebtn':True,
            'locationlist':locationlist
        }
        return Response({"n": 1, "Msg": "Task Assigned Successfully", "Status": "Success",'data':context})
    else:
        return Response({"n": 0, "Msg": "userid doesnt exist.", "Status": "Failed",'data':''})

@api_view(['POST'])
def punch_indata(request):
    data = {}
    userid = request.user.id
    data['employeeId'] =request.user.employeeId
    if data['employeeId'] is not None and data['employeeId'] != '':
        if "Remote_Reason" in request.data.keys():
            if request.data.get('Remote_Reason') is not None and request.data.get('Remote_Reason') !='':
                data['Remote_Reason'] = request.data.get('Remote_Reason')
                # return Response ({
                #     "data": [],
                #     "response":{
                #         "n" : 0,
                #         "msg" : "Remote functionality is disabled temporarily for temporary purposes.",
                #         "status" : "errors"
                #     }
                #     })
            
            else:
                data['Remote_Reason'] = None 
        else:
            data['Remote_Reason'] = None
            
        data['remote_latitude'] = request.data.get('latitude')
        data['remote_longitude'] = request.data.get('longitude')
        
        if data['remote_latitude'] is  None or data['remote_latitude'] =="" or data['remote_longitude'] is  None or data['remote_longitude'] =="":
            return Response ({
                "data": [],
                "response":{
                    "n" : 0,
                    "msg" : "Please report to your reporting manager.",
                    "status" : "errors"
                }
                })
        
        data['attendance_type'] = request.data.get('attendance_type')
        
        
        data['date'] = request.data.get('date')
        currentzone = pytz.timezone("Asia/Kolkata") 
        currenttime = datetime.datetime.now(currentzone)
        newcurrenttime =currenttime.strftime("%Y-%m-%dT%H:%M:%S.%f%z")

        newtime = str(newcurrenttime)
        ntime1= newtime.split('T')[1]
        ntimhrs = ntime1.split('.')[0]
        data['time'] = ntimhrs

        Date = str(data['date'])
        nYear = Date.split("-")[0]
        nMonth = Date.split("-")[1]
        nDay = Date.split("-")[2]
        date_given = datetime.datetime(year=int(nYear), month=int(nMonth), day=int(nDay)).date()
        nweek = week_of_month(date_given)

        data['Week'] = nweek
        data['Month'] = int(nMonth)
        data['Year'] = int(nYear)
        data['created_at'] = timezone.now()
        data['company_code'] = request.user.company_code
        serializer = attendanceserializer(data=data)
        if serializer.is_valid():
            serializer.save()
        
            return Response ({
                "data": serializer.data,
                "response":{
                    "n" : 1,
                    "msg" : "Punch in successfully",
                    "status" : "success"
                }
                })
        else:
            return Response ({
                "data": serializer.errors,
                "response":{
                    "n" : 0,
                    "msg" : "Failed",
                    "status" : "errors"
                }
                }) 
    else:
        return Response ({
            "data": '',
            "response":{
                "n" : 0,
                "msg" : "Please Map Attendance Id",
                "status" : "errors"
            }
            }) 


@api_view(['POST'])
def punch_outdata(request):
    data = {}
    data['employeeId'] = request.user.employeeId
    if data['employeeId'] is not None and data['employeeId'] != '':
        if "Remote_Reason" in request.data.keys():
            if request.data.get('Remote_Reason') is not None and request.data.get('Remote_Reason') !='':
                data['Remote_Reason'] = request.data.get('Remote_Reason')
                # return Response ({
                #     "data": [],
                #     "response":{
                #         "n" : 0,
                #         "msg" : "Remote functionality is disabled temporarily for temporary purposes.",
                #         "status" : "errors"
                #         }
                #     })
            
            else:
                data['Remote_Reason'] = None 
        else:
            data['Remote_Reason'] = None
            
        data['remote_latitude'] = request.data.get('latitude')
        data['remote_longitude'] = request.data.get('longitude')
        if data['remote_latitude'] is  None or data['remote_latitude'] =="" or data['remote_longitude'] is  None or data['remote_longitude'] =="":
            return Response ({
                "data": [],
                "response":{
                    "n" : 0,
                    "msg" : "Please report to your reporting manager.",
                    "status" : "errors"
                }
                })
            
        data['attendance_type'] = request.data.get('attendance_type')
        data['checkout'] = True
        
        
        data['date'] = request.data.get('date')
        currentzone = pytz.timezone("Asia/Kolkata") 
        currenttime = datetime.datetime.now(currentzone)
        newcurrenttime =currenttime.strftime("%Y-%m-%dT%H:%M:%S.%f%z")

        newtime = str(newcurrenttime)
        ntime1= newtime.split('T')[1]
        ntimhrs = ntime1.split('.')[0]
        data['time'] = ntimhrs

        Date = str(data['date'])
        nYear = Date.split("-")[0]
        nMonth = Date.split("-")[1]
        nDay = Date.split("-")[2]
        date_given = datetime.datetime(year=int(nYear), month=int(nMonth), day=int(nDay)).date()
        nweek = week_of_month(date_given)

        data['Week'] = nweek
        data['Month'] = int(nMonth)
        data['Year'] = int(nYear)
        data['created_at'] = timezone.now()
        data['company_code'] = request.user.company_code
        serializer = attendanceserializer(data=data)
        if serializer.is_valid():
            serializer.save()
        
            return Response ({
                "data": serializer.data,
                "response":{
                    "n" : 1,
                    "msg" : "Punch out successfully",
                    "status" : "success"
                }
                })
        else:
            return Response ({
                "data": serializer.errors,
                "response":{
                    "n" : 0,
                    "msg" : "Failed",
                    "status" : "errors"
                }
                }) 
    else:
        return Response ({
            "data": '',
            "response":{
                "n" : 0,
                "msg" : "Please Map Attendance Id",
                "status" : "errors"
            }
            }) 



# @api_view(['POST'])
# def punch_outdata(request):
#     data = {}
#     data['employeeId'] = request.user.employeeId
#     data['date'] = request.data.get('date')
    
#     data['remote_latitude'] = request.data.get('latitude')
#     data['remote_longitude'] = request.data.get('longitude')
    
#     currentzone = pytz.timezone("Asia/Kolkata") 
#     currenttime = datetime.datetime.now(currentzone)
#     newcurrenttime =currenttime.strftime("%Y-%m-%dT%H:%M:%S.%f%z")

#     # today730pm = currenttime.replace(hour=19, minute=30, second=0, microsecond=0)

#     # if today730pm > currenttime:
#     newtime = str(newcurrenttime)
#     ntime1= newtime.split('T')[1]
#     ntimhrs = ntime1.split('.')[0]
#     data['time'] = ntimhrs
#     Date = str(data['date'])
#     nYear = Date.split("-")[0]
#     nMonth = Date.split("-")[1]
#     nDay = Date.split("-")[2]
#     date_given = datetime.datetime(year=int(nYear), month=int(nMonth), day=int(nDay)).date()
#     nweek = week_of_month(date_given)

#     data['Week'] = nweek
#     data['Month'] = int(nMonth)
#     data['Year'] = int(nYear)
#     data['company_code'] = request.user.company_code
#     data['checkout'] = True
#     data['created_at'] = timezone.now()
#     serializer = attendanceserializer(data=data)
#     if serializer.is_valid():
#         serializer.save()
#         return Response ({
#             "data": serializer.data,
#             "response":{
#                 "n" : 1,
#                 "msg" : "Check Out Successful",
#                 "status" : "success"
#             }
#             })
#     else:
#         return Response ({
#             "data": serializer.errors,
#             "response":{
#                 "n" : 0,
#                 "msg" : "Failed",
#                 "status" : "errors"
#             }
#             }) 






# @api_view(['GET'])
# def punch_getdata(request):
#     employeeId = request.user.employeeId
#     current_date = datetime.date.today()
#     intime=''
#     outtime=''
#     intimedate=''
#     outtimedate=''
#     userobj = Users.objects.filter(employeeId = employeeId).first()
#     if userobj is not None:
#         userser = UserSerializer(userobj)
#         locationid = userser.data['locationId']
#         if locationid is not None and locationid != "":
#             locationobj = Location.objects.filter(id=int(locationid)).first()
#             if locationobj is not None:
#                 location = locationobj.LocationName
#             else:
#                 location = ""
#         else:
#             location = ""
#     else:
#         location = ""

#     get_empid = attendance.objects.filter(employeeId=(employeeId),date=str(current_date)).order_by('time').first()
#     if get_empid is not None:
#         get_data = 1
#     else: 
#         get_data = 0                  
    
#     lastentry = attendance.objects.filter(employeeId=(employeeId),date=str(current_date)).order_by('time').last()
#     if get_empid is not None and lastentry is not None:
#         if get_empid.time != lastentry.time:
#             intime = get_empid.time
#             outtime = lastentry.time
#             intimedate=get_empid.date
#             outtimedate=lastentry.date
#         else:
#             intime = get_empid.time
#             outtime = ''
#             intimedate=get_empid.date
#             outtimedate=''
            
#     punchoutobj = attendance.objects.filter(employeeId=(employeeId),checkout=True,date=str(current_date)).order_by('time').first()
#     if punchoutobj is not None:
#         punchout = 1
#     else: 
#         punchout = 0
#     return Response ({
#         'indatetime':str(dd_mm_yyyy(str(intimedate))) + ' ' + str(intime),
#         'outdatetime':str(dd_mm_yyyy(str(outtimedate))) + ' ' + str(outtime),
#         'data':get_data,
#         'intime':intime,
#         'outtime':outtime,
#         'intimedate':intimedate,
#         'outtimedate':outtimedate,
#         'punchout':punchout,
#         'location':location,
#             "response":{
#                 "n" : 1,
#                 "msg" : "pass",
#                 "status" : "success"
#             }
#             })



@api_view(['GET'])
@permission_classes((AllowAny,))
def punchout_scheduler(request):
    companycode = "O001"
    my_date = datetime.date.today()

    Date = str(my_date)
    nYear = Date.split("-")[0]
    nMonth = Date.split("-")[1]
    nDay = Date.split("-")[2]
    date_given = datetime.datetime(year=int(nYear), month=int(nMonth), day=int(nDay)).date()
    nweek = week_of_month(date_given)

    Week= nweek
    Month = int(nMonth)
    Year = int(nYear)

    emplist = []
    empobjs = Users.objects.filter(company_code=companycode,typeofwork__in = [1,4],is_blocked=False)
    if empobjs.exists():
        empser =  UserSerializer(empobjs,many=True)
        for u in empser.data:
            emplist.append(int(u['id']))

    leaveobj = Leave.objects.filter(WorkFromHome=True,leave_status="Approved",start_date__lte = my_date,end_date__gte=my_date,Active=True)
    if leaveobj.exists():
        leaveser = leaveserializer(leaveobj,many=True)
        for l in leaveser.data:
            emplist.append(int(l['employeeId']))

    setemplist = list(set(emplist))

    for e in setemplist:
        userobj = Users.objects.filter(id=int(e)).first()
        attendanceid = userobj.employeeId
        if attendanceid is not None:
            punchoutobj = attendance.objects.filter(employeeId=attendanceid,date=str(my_date)).count()
            if punchoutobj == 1:
                leaveexistobj = Leave.objects.filter(employeeId=str(e),WorkFromHome=True,leave_status="Approved",start_date__lte = my_date,end_date__gte=my_date,Active=True).first()
                if leaveexistobj is not None:
                    leavetype = leaveexistobj.leavetype
                    if leavetype == "FirstHalf":
                        checkouttime = "14:00:00"
                    else:
                        checkouttime = "19:30:00"
                
                else:
                    wfhempleave = Leave.objects.filter(employeeId=str(e),WorkFromHome=False,leave_status="Approved",start_date__lte = my_date,end_date__gte=my_date,Active=True).first()
                    if wfhempleave is not None:
                        leavetype = wfhempleave.leavetype
                        if leavetype == "FirstHalf":
                            checkouttime = "14:00:00"
                        else:
                            checkouttime = "19:30:00"
                    else:
                        checkouttime = "19:30:00"

                attendance.objects.create(employeeId=attendanceid,time=checkouttime,date=my_date,company_code= companycode,Week=Week,Month=Month,Year=Year,checkout=True)

    return Response ({
            "data": '',
            "response":{
                "n" : 1,
                "msg" : "Success",
                "status" : "Success"
            }
            }) 


@api_view(['POST'])
def attendancerequestapi(request):
    data = {}
    managerlist = []
    my_date = datetime.date.today()
    data['UserId'] = request.data.get('userid')
    username = request.user.Firstname + " " + request.user.Lastname
    data['AttendanceId'] = request.user.employeeId
    data['Reason'] = request.data.get('attreason')
    reqdate = request.data.get('requestdate')
    data['company_code'] = request.user.company_code
    data['Date'] = datetime.datetime.strptime(reqdate, "%d-%m-%Y").strftime("%Y-%m-%d")

    leaverequestobj = Leave.objects.filter(employeeId=str(data['UserId']),leave_status="Approved",WorkFromHome=False,start_date__lte = my_date,end_date__gte=my_date,Active=True)
    if leaverequestobj:
        return Response ({
                "data": '',
                "response":{
                    "n" : 0,
                    "msg" : "Attendance cannot be applied on Leave ",
                    "status" : "errors"
                }
                }) 
            
    requestobj =  AttendanceRequest.objects.filter(UserId=data['UserId'],Date = data['Date']).first()
    if requestobj is not None:
        return Response ({
                "data": '',
                "response":{
                    "n" : 0,
                    "msg" : "Request Already Exist",
                    "status" : "errors"
                }
                }) 
    
    attexistobj = attendance.objects.filter(employeeId=str(data['AttendanceId']),date=data['Date'],company_code =  data['company_code']).order_by('time').first()
    if attexistobj is None :
        return Response ({
                "data": '',
                "response":{
                    "n" : 0,
                    "msg" : "Attendance Not Found",
                    "status" : "errors"
                }
                }) 
    
    else:
        managerobj = leaveMapping.objects.filter(employeeId = data['UserId'],company_code=request.user.company_code)
        if managerobj:
            serializer = leaveMappingserializer(managerobj,many=True)
            for s in serializer.data:
                managerlist.append(int(s['managerId']))
        data['manager_ids'] = list(set(managerlist))
        manager_list = list(set(managerlist))

        checkintimeobj = attendance.objects.filter(employeeId=data['AttendanceId'],date=data['Date'],company_code =  data['company_code'] ).order_by('time').first()
        if checkintimeobj is not None:
            data['CheckInTime'] = checkintimeobj.time
           
        att_serializer = attendancerequestserializer(data=data)
        if att_serializer.is_valid():
            att_serializer.save()
            if managerlist != []:
                for m in manager_list:
                    TaskNotification.objects.create(
                            NotificationTitle = "Attendance Request",
                            NotificationMsg = "<span class='actionuser'>" + username + "</span> has requested for attendance" ,
                            UserID_id = int(m),
                            NotificationTypeId_id = 6 ,
                            leaveID = 0,
                            created_by = request.user.id,
                            company_code = request.user.company_code,
                        )
            return Response ({
                    "data": '',
                    "response":{
                        "n" : 1,
                        "msg" : "Request Sent Successfully",
                        "status" : "Success"
                    }
                    }) 
        else:
            return Response ({
                "data": att_serializer.errors,
                "response":{
                    "n" : 0,
                    "msg" : "Request Failed",
                    "status" : "errors"
                }
                }) 
   
        

@api_view(['GET'])
def getatt_request(request):
    managerid = request.user.id
    pendingattobj = AttendanceRequest.objects.filter(manager_ids=managerid,Action__isnull=True,Active=True).order_by('-id')
    if pendingattobj:
        pendingserializer = attendancerequestserializer(pendingattobj,many=True)
        for p in pendingserializer.data:
            userobj = Users.objects.filter(id=p['UserId'],is_active=True).first()
            if userobj is not None:
                p['Username'] = userobj.Firstname + " " + userobj.Lastname
            else:
                p['Username'] = "" 
            
            checkintimeobj = attendance.objects.filter(employeeId=p['AttendanceId'],date=p['Date'],company_code =  p['company_code'] ).order_by('time').first()
            if checkintimeobj is not None:
                checkin = checkintimeobj.time

                checkouttimeobj = attendance.objects.filter(employeeId=p['AttendanceId'],date=p['Date'],company_code =  p['company_code'] ).order_by('time').last()
                if checkouttimeobj is not None:
                    checkout = checkouttimeobj.time

                if checkin == checkout:
                    p['UserCheckInTime'] = checkin[:-3]
                    p['UserCheckOutTime'] = ''
                else:
                    p['UserCheckInTime'] = checkin
                    p['UserCheckOutTime'] = checkout
            else:
                p['UserCheckInTime'] = ''
                p['UserCheckOutTime'] = ''
 
            createddate = str(p['Date'])
            p['newdate'] = dd_month_year_format(createddate)
            

            # strtime  = str(datetimestr.split('T')[1])
            # p['newtime'] = str(strtime.split('.')[0])

            managerlist = []
            for mangr in p['manager_ids']:
                managerdict ={}
                managerObj=Users.objects.filter(id=int(mangr),company_code=request.user.company_code,is_active=True).first()
                if managerObj is not None:
                    managerdict['managername'] = managerObj.Firstname + " " + managerObj.Lastname
                    if managerObj.Photo is not None and managerObj.Photo !="":
                        managerdict['managerpic'] = imageUrl +"/media/"+ str(managerObj.Photo)
                    else:
                        managerdict['managerpic'] = imageUrl + "/static/assets/images/profile.png"
                    if p['actiontakenby'] is not None:
                        if mangr == p['actiontakenby']:
                            managerdict['actiontaken'] = True
                        else:
                            managerdict['actiontaken'] = False
                    managerlist.append(managerdict)
                    
                    
                    
            p['managerlist'] = managerlist

        pendinglist = pendingserializer.data
    else:
        pendinglist = []
        

    Approvedattobj = AttendanceRequest.objects.filter(manager_ids=managerid,Action="Approve",Active=True).order_by('-id')
    if Approvedattobj:
        approvedserializer = attendancerequestserializer(Approvedattobj,many=True)
        for a in approvedserializer.data:
            userobj = Users.objects.filter(id=a['UserId'],company_code=request.user.company_code,is_active=True).first()
            if userobj is not None:
                a['Username'] = userobj.Firstname + " " + userobj.Lastname
            else:
                a['Username'] = ""

            checkintimeobj = attendance.objects.filter(employeeId=a['AttendanceId'],date=a['Date'],company_code =  a['company_code'] ).order_by('time').first()
            if checkintimeobj is not None:
                checkin = checkintimeobj.time

                checkouttimeobj = attendance.objects.filter(employeeId=a['AttendanceId'],date=a['Date'],company_code =  a['company_code'] ).order_by('time').last()
                if checkouttimeobj is not None:
                    checkout = checkouttimeobj.time

                if checkin == checkout:
                    a['UserCheckInTime'] = checkin
                    a['UserCheckOutTime'] = '-- : --'
                else:
                    a['UserCheckInTime'] = checkin
                    a['UserCheckOutTime'] = checkout
            else:
                a['UserCheckInTime'] = '-- : --'
                a['UserCheckOutTime'] = '-- : --'

            createddate = str(a['Date'])
            a['newdate'] = dd_month_year_format(createddate)
           
            # strtime  = str(datetimestr.split('T')[1])
            # a['newtime'] = str(strtime.split('.')[0])

            managerlist = []
            for mangr in a['manager_ids']:
                managerdict = {}
                managerObj=Users.objects.filter(id=int(mangr)).first()
                if managerObj is not None:
                    managerdict['managername'] = managerObj.Firstname + " " + managerObj.Lastname
                    if managerObj.Photo is not None and managerObj.Photo !="":
                        managerdict['managerpic'] = imageUrl +"/media/"+ str(managerObj.Photo)
                    else:
                        managerdict['managerpic'] = imageUrl + "/static/assets/images/profile.png"
     
                    
                    if a['actiontakenby'] is not None:
                        if int(mangr) == int(a['actiontakenby']):
                            managerdict['actiontaken'] = True
                        else:
                            managerdict['actiontaken'] = False

                    managerlist.append(managerdict)
            a['managerlist'] = managerlist

        approvedlist = approvedserializer.data
    else:
        approvedlist = []

    rejectedattobj = AttendanceRequest.objects.filter(manager_ids=managerid,Action="Reject",Active=True).order_by('-id')
    if rejectedattobj:
        rejectedserializer = attendancerequestserializer(rejectedattobj,many=True)
        for r in rejectedserializer.data:
            userobj = Users.objects.filter(id=r['UserId'],company_code=request.user.company_code,is_active=True).first()
            if userobj is not None:
                r['Username'] = userobj.Firstname + " " + userobj.Lastname
            else:
                r['Username'] = ""

            checkintimeobj = attendance.objects.filter(employeeId=r['AttendanceId'],date=r['Date'],company_code =  r['company_code'] ).order_by('time').first()
            if checkintimeobj is not None:
                checkin = checkintimeobj.time

                checkouttimeobj = attendance.objects.filter(employeeId=r['AttendanceId'],date=r['Date'],company_code =  r['company_code'] ).order_by('time').last()
                if checkouttimeobj is not None:
                    checkout = checkouttimeobj.time

                if checkin == checkout:
                    r['UserCheckInTime'] = checkin
                    r['UserCheckOutTime'] = '-- : --'
                else:
                    r['UserCheckInTime'] = checkin
                    r['UserCheckOutTime'] = checkout
            else:
                r['UserCheckInTime'] = '-- : --'
                r['UserCheckOutTime'] = '-- : --'

            createddate = str(r['Date'])
            r['newdate'] = dd_month_year_format(createddate)
           
            # strtime  = str(datetimestr.split('T')[1])
            # r['newtime'] = str(strtime.split('.')[0])

            managerlist = []
            for mangr in r['manager_ids']:
                managerdict = {}
                managerObj=Users.objects.filter(id=int(mangr),company_code=request.user.company_code,is_active=True).first()
                if managerObj is not None:
                    
                    managerdict['managername'] = managerObj.Firstname + " " + managerObj.Lastname
                    if managerObj.Photo is not None and managerObj.Photo !="":
                        managerdict['managerpic'] = imageUrl +"/media/"+ str(managerObj.Photo)
                    else:
                        managerdict['managerpic'] = imageUrl + "/static/assets/images/profile.png"

                    
                    if r['actiontakenby'] is not None:
                        if int(mangr) == int(r['actiontakenby']):
                            managerdict['actiontaken'] = True
                        else:
                            managerdict['actiontaken'] = False

                    managerlist.append(managerdict)
            r['managerlist'] = managerlist
        rejectedlist = rejectedserializer.data
    else:
        rejectedlist = []


    cancelattobj = AttendanceRequest.objects.filter(manager_ids=managerid,Active=False).order_by('-id')
    if cancelattobj:
        cancelserializer = attendancerequestserializer(cancelattobj,many=True)
        for c in cancelserializer.data:
            userobj = Users.objects.filter(id=c['UserId'],company_code=request.user.company_code,is_active=True).first()
            if userobj is not None:
                c['Username'] = userobj.Firstname + " " + userobj.Lastname
            else:
                c['Username'] = ""

            checkintimeobj = attendance.objects.filter(employeeId=c['AttendanceId'],date=c['Date'],company_code =  c['company_code'] ).order_by('time').first()
            if checkintimeobj is not None:
                checkin = checkintimeobj.time

                checkouttimeobj = attendance.objects.filter(employeeId=c['AttendanceId'],date=c['Date'],company_code =  c['company_code'] ).order_by('time').last()
                if checkouttimeobj is not None:
                    checkout = checkouttimeobj.time

                if checkin == checkout:
                    c['UserCheckInTime'] = checkin
                    c['UserCheckOutTime'] = '-- : --'
                else:
                    c['UserCheckInTime'] = checkin
                    c['UserCheckOutTime'] = checkout
            else:
                c['UserCheckInTime'] = '-- : --'
                c['UserCheckOutTime'] = '-- : --'

            createddate = str(c['Date'])
            c['newdate'] = dd_month_year_format(createddate)
           
            # strtime  = str(datetimestr.split('T')[1])
            # r['newtime'] = str(strtime.split('.')[0])

            managerlist = []
            for mangr in c['manager_ids']:
                managerdict = {}
                managerObj=Users.objects.filter(id=int(mangr),company_code=request.user.company_code,is_active=True).first()
                if managerObj is not None:
                    
                    managerdict['managername'] = managerObj.Firstname + " " + managerObj.Lastname
                    if managerObj.Photo is not None and managerObj.Photo !="":
                        managerdict['managerpic'] = imageUrl +"/media/"+ str(managerObj.Photo)
                    else:
                        managerdict['managerpic'] = imageUrl + "/static/assets/images/profile.png"

                    
                        
                    if c['actiontakenby'] is not None:
                        if int(mangr) == int(c['actiontakenby']):
                            managerdict['actiontaken'] = True
                        else:
                            managerdict['actiontaken'] = False

                    managerlist.append(managerdict)
            c['managerlist'] = managerlist
        Cancelledlist = cancelserializer.data
    else:
        Cancelledlist = []


    context = {
        "pendinglist" :pendinglist,
        "approvedlist":approvedlist,
        "Rejectedlist":rejectedlist,
        "Cancelledlist":Cancelledlist
    }

    return Response ({
        "data": context,
        "response":{
            "n" : 1,
            "msg" : "Data found successfully",
            "status" : "Success"
        }
        }) 
    

@api_view(['POST'])
def manageratt_request(request):
    data = {}
    requestid = request.data.get('reqid')
    companycode = request.user.company_code
    userid = request.user.id
    data['actiontakenby'] = userid
    data['Action'] = request.data.get('Action')
    usercheckintime = request.data.get('usercheckintime')
    usercheckouttime =  request.data.get('usercheckouttime')
    Managername = request.user.Firstname + " " + request.user.Lastname

    reqobj = AttendanceRequest.objects.filter(id=requestid,Active=True).first()
    if reqobj is not None :
        employeeId = reqobj.UserId
        employee_attid = reqobj.AttendanceId 
        my_date = reqobj.Date
        if data['Action'] == "Approve":
            checkintime = usercheckintime
            checkouttime = usercheckouttime

            reqdate = reqobj.Date
            todaysdate = date.today()

            if reqdate == todaysdate:
                attobj = attendance.objects.filter(employeeId=employee_attid,date=my_date,company_code= companycode).order_by('time').count()
                if attobj > 1 :
                    attfirstobj = attendance.objects.filter(employeeId=employee_attid,date=my_date,company_code= companycode).order_by('time').first()
                    if attfirstobj is not None:
                        if len(checkintime) > 6 :
                            newcheckintime = checkintime 
                        else:
                            newcheckintime = checkintime +":00"
                        
                        attfirstobj.time = newcheckintime
                        attfirstobj.save()

                    if checkouttime != "00:00:00":
                        attlastobj = attendance.objects.filter(employeeId=employee_attid,date=my_date,company_code= companycode).order_by('time').last()
                        if attlastobj is not None:
                            if len(checkouttime) > 6:
                                attlastobj.time = checkouttime
                            else:
                                attlastobj.time = checkouttime +":00" 
                            attlastobj.save()
                if attobj == 1:
                    attfirstobj = attendance.objects.filter(employeeId=employee_attid,date=my_date,company_code= companycode).order_by('time').first()
                    if attfirstobj is not None:
                        if len(checkintime) > 6 :
                            newcheckintime = checkintime 
                        else:
                            newcheckintime = checkintime +":00"
                        attfirstobj.time = newcheckintime
                        attfirstobj.save()
            else:
                attobj = attendance.objects.filter(employeeId=employee_attid,date=my_date,company_code= companycode).order_by('time').count()
                if attobj > 1 :
                    attfirstobj = attendance.objects.filter(employeeId=employee_attid,date=my_date,company_code= companycode).order_by('time').first()
                    if attfirstobj is not None:
                        if len(checkintime) > 6 :
                            newcheckintime = checkintime 
                        else:
                            newcheckintime = checkintime +":00"
                        attfirstobj.time = newcheckintime
                        attfirstobj.save()

                    attlastobj = attendance.objects.filter(employeeId=employee_attid,date=my_date,company_code= companycode).order_by('time').last()
                    if attlastobj is not None:
                        if len(checkouttime) > 6:
                            newcheckouttime = checkouttime                            
                        else:
                            newcheckouttime = checkouttime +":00"
                        attlastobj.time = newcheckouttime
                        attlastobj.save()
                if attobj == 1:
                    attfirstobj = attendance.objects.filter(employeeId=employee_attid,date=my_date,company_code= companycode).order_by('time').first()
                    if attfirstobj is not None:
                        if len(checkintime) > 6 :
                            newcheckintime = checkintime
                        else:
                            newcheckintime = checkintime +":00"
                        attfirstobj.time = newcheckintime
                        attfirstobj.save()    

                    dt = str(my_date)
                    nYear = dt.split("-")[0]
                    nMonth = dt.split("-")[1]
                    nDay = dt.split("-")[2]
                    created_at = timezone.now()
                    date_given = datetime.datetime(year=int(nYear), month=int(nMonth), day=int(nDay)).date()
                    nweek = week_of_month(date_given) 

                    newcheckouttime = checkouttime +":00"
                    attendance.objects.create(employeeId = employee_attid,date = my_date,created_at = created_at,company_code=companycode,time=newcheckouttime,Week=nweek,Month=nMonth,Year=nYear)

            att_serializer =  attendancerequestserializer(reqobj,data=data,partial=True)
            if att_serializer.is_valid():
                att_serializer.save()

                TaskNotification.objects.create(
                    NotificationTitle = "Attendance Request",
                    NotificationMsg ="<span class='actionuser'>" + Managername + "</span> has approved your attendance request" ,
                    UserID_id = int(employeeId),
                    NotificationTypeId_id = 6 ,
                    leaveID = 0,
                    created_by = request.user.id,
                    company_code = request.user.company_code,
                )
                
                return Response ({
                    "data": att_serializer.data,
                    "response":{
                        "n" : 1,
                        "action":"Yes",
                        "msg" : "Attendance Saved successfully",
                        "status" : "Success"
                    }
                    }) 
        else:
            att_serializer =  attendancerequestserializer(reqobj,data=data,partial=True)
            if att_serializer.is_valid():
                att_serializer.save()

                TaskNotification.objects.create(
                            NotificationTitle = "Attendance Request",
                            NotificationMsg = "<span class='actionuser'>" + Managername + "</span> has rejected your attendance request" ,
                            UserID_id = int(employeeId),
                            NotificationTypeId_id = 6 ,
                            leaveID = 0,
                            created_by = request.user.id,
                            company_code = request.user.company_code,
                        )
                return Response ({
                    "data": {},
                    "response":{
                        "n" : 1,
                        "action":"No",
                        "msg" : "Attendance Request Rejected.",
                        "status" : "Success"
                    }
                    }) 
    else:
        reqcancelobj = AttendanceRequest.objects.filter(id=requestid,Active=False).first()
        if reqcancelobj is not None:
            return Response ({
                "data":{},
                "response":{
                    "n" : 0,
                    "msg" : "Attendance Request Cancelled by User",
                    "status" : "errors"
                }
                }) 
        else:
            return Response ({
                "data":{},
                "response":{
                    "n" : 0,
                    "msg" : "Attendance Request not found",
                    "status" : "errors"
                }
                }) 






@api_view(['GET'])
def getemp_requestapi(request):
    empid = request.user.id
    pendingattobj = AttendanceRequest.objects.filter(UserId=empid,Action__isnull=True,Active=True).order_by('-id')
    if pendingattobj:
        pendingserializer = attendancerequestserializer(pendingattobj,many=True)
        for p in pendingserializer.data:
            userobj = Users.objects.filter(id=p['UserId'],company_code=request.user.company_code,is_active=True).first()
            if userobj is not None:
                p['Username'] = userobj.Firstname + " " + userobj.Lastname
            else:
                p['Username'] = "" 
            

            checkintimeobj = attendance.objects.filter(employeeId=p['AttendanceId'],date=p['Date'],company_code =  p['company_code'] ).order_by('time').first()
            if checkintimeobj is not None:
                checkin = checkintimeobj.time

                checkouttimeobj = attendance.objects.filter(employeeId=p['AttendanceId'],date=p['Date'],company_code =  p['company_code'] ).order_by('time').last()
                if checkouttimeobj is not None:
                    checkout = checkouttimeobj.time

                if checkin == checkout:
                    p['UserCheckInTime'] = checkin
                    p['UserCheckOutTime'] = ''
                else:
                    p['UserCheckInTime'] = checkin
                    p['UserCheckOutTime'] = checkout
            else:
                p['UserCheckInTime'] = ''
                p['UserCheckOutTime'] = ''

            datetimestr = str(p['created_at'])
            createddate = str(datetimestr.split('T')[0])
            p['newdate'] = dd_month_year_format(createddate)

            p['requestdate'] = dd_month_year_format(str(p['Date']))

            strtime  = str(datetimestr.split('T')[1])
            p['newtime'] = str(strtime.split('.')[0])

            managerlist = []
            for mangr in p['manager_ids']:
                managerdict ={}
                managerObj=Users.objects.filter(id=int(mangr),company_code=request.user.company_code,is_active=True).first()
                if managerObj is not None:
                    managerdict['managername'] = managerObj.Firstname + " " + managerObj.Lastname
                    if managerObj.Photo is not None and managerObj.Photo !="":
                        managerdict['managerpic'] = imageUrl +"/media/"+ str(managerObj.Photo)
                    else:
                        managerdict['managerpic'] = imageUrl + "/static/assets/images/profile.png"
                else:
                    managerdict['managername']='NA'
                    managerdict['managerpic'] = imageUrl + "/static/assets/images/profile.png"



                if p['actiontakenby'] is not None:
                    if mangr == p['actiontakenby']:
                        managerdict['actiontaken'] = True
                    else:
                        managerdict['actiontaken'] = False
                managerlist.append(managerdict)
            p['managerlist'] = managerlist

        pendinglist = pendingserializer.data
    else:
        pendinglist = []
        

    Approvedattobj = AttendanceRequest.objects.filter(UserId=empid,Action="Approve",Active=True).order_by('-id')
    if Approvedattobj:
        approvedserializer = attendancerequestserializer(Approvedattobj,many=True)
        for a in approvedserializer.data:
            userobj = Users.objects.filter(id=a['UserId'],company_code=request.user.company_code,is_active=True).first()
            if userobj is not None:
                a['Username'] = userobj.Firstname + " " + userobj.Lastname
            else:
                a['Username'] = ""

            checkintimeobj = attendance.objects.filter(employeeId=a['AttendanceId'],date=a['Date'],company_code =  a['company_code'] ).order_by('time').first()
            if checkintimeobj is not None:
                checkin = checkintimeobj.time

                checkouttimeobj = attendance.objects.filter(employeeId=a['AttendanceId'],date=a['Date'],company_code =  a['company_code'] ).order_by('time').last()
                if checkouttimeobj is not None:
                    checkout = checkouttimeobj.time

                if checkin == checkout:
                    a['UserCheckInTime'] = checkin
                    a['UserCheckOutTime'] = '-- : --'
                else:
                    a['UserCheckInTime'] = checkin
                    a['UserCheckOutTime'] = checkout
            else:
                a['UserCheckInTime'] = '-- : --'
                a['UserCheckOutTime'] = '-- : --'

            datetimestr = str(a['created_at'])
            createddate = str(datetimestr.split('T')[0])
            a['newdate'] = dd_month_year_format(createddate)

            a['requestdate'] = dd_month_year_format(str(a['Date']))
           
            strtime  = str(datetimestr.split('T')[1])
            a['newtime'] = str(strtime.split('.')[0])

            managerlist = []
            for mangr in a['manager_ids']:
                managerdict = {}
                managerObj=Users.objects.filter(id=int(mangr),company_code=request.user.company_code,is_active=True).first()
                if managerObj is not None:

                    managerdict['managername'] = managerObj.Firstname + " " + managerObj.Lastname
                    if managerObj.Photo is not None and managerObj.Photo !="":
                        managerdict['managerpic'] = imageUrl +"/media/"+ str(managerObj.Photo)
                    else:
                        managerdict['managerpic'] = imageUrl + "/static/assets/images/profile.png"
                else:
                    managerdict['managername']="NA"
                    managerdict['managerpic'] = imageUrl + "/static/assets/images/profile.png"

                if a['actiontakenby'] is not None:
                    if int(mangr) == int(a['actiontakenby']):
                        managerdict['actiontaken'] = True
                    else:
                        managerdict['actiontaken'] = False

                managerlist.append(managerdict)
            a['managerlist'] = managerlist

        approvedlist = approvedserializer.data
    else:
        approvedlist = []


    rejectedattobj = AttendanceRequest.objects.filter(UserId=empid,Action="Reject",Active=True).order_by('-id')
    if rejectedattobj:
        rejectedserializer = attendancerequestserializer(rejectedattobj,many=True)
        for r in rejectedserializer.data:
            userobj = Users.objects.filter(id=r['UserId'],company_code=request.user.company_code,is_active=True).first()
            if userobj is not None:
                r['Username'] = userobj.Firstname + " " + userobj.Lastname
            else:
                r['Username'] = ""

            checkintimeobj = attendance.objects.filter(employeeId=r['AttendanceId'],date=r['Date'],company_code =  r['company_code'] ).order_by('time').first()
            if checkintimeobj is not None:
                checkin = checkintimeobj.time

                checkouttimeobj = attendance.objects.filter(employeeId=r['AttendanceId'],date=r['Date'],company_code =  r['company_code'] ).order_by('time').last()
                if checkouttimeobj is not None:
                    checkout = checkouttimeobj.time

                if checkin == checkout:
                    r['UserCheckInTime'] = checkin
                    r['UserCheckOutTime'] = '-- : --'
                else:
                    r['UserCheckInTime'] = checkin
                    r['UserCheckOutTime'] = checkout
            else:
                r['UserCheckInTime'] = '-- : --'
                r['UserCheckOutTime'] = '-- : --'

            datetimestr = str(r['created_at'])
            createddate = str(datetimestr.split('T')[0])
            r['newdate'] = dd_month_year_format(createddate)

            r['requestdate'] = dd_month_year_format(str(r['Date']))
           
            strtime  = str(datetimestr.split('T')[1])
            r['newtime'] = str(strtime.split('.')[0])

            managerlist = []
            for mangr in r['manager_ids']:
                managerdict = {}
                managerObj=Users.objects.filter(id=int(mangr),company_code=request.user.company_code,is_active=True).first()
              
                if managerObj is not None:

                    managerdict['managername'] = managerObj.Firstname + " " + managerObj.Lastname
                    if managerObj.Photo is not None and managerObj.Photo !="":
                        managerdict['managerpic'] = imageUrl +"/media/"+ str(managerObj.Photo)
                    else:
                        managerdict['managerpic'] = imageUrl + "/static/assets/images/profile.png"
                else:
                    managerdict['managername']="NA"
                    managerdict['managerpic'] = imageUrl + "/static/assets/images/profile.png"

                if r['actiontakenby'] is not None:
                    if int(mangr) == int(r['actiontakenby']):
                        managerdict['actiontaken'] = True
                    else:
                        managerdict['actiontaken'] = False

                managerlist.append(managerdict)
            r['managerlist'] = managerlist
        rejectedlist = rejectedserializer.data
    else:
        rejectedlist = []

    cancelledobj = AttendanceRequest.objects.filter(UserId=empid,Active=False).order_by('-id')
    if cancelledobj:
        cancelserializer = attendancerequestserializer(cancelledobj,many=True)
        for r in cancelserializer.data:
            userobj = Users.objects.filter(id=r['UserId'],company_code=request.user.company_code,is_active=True).first()
            if userobj is not None:
                r['Username'] = userobj.Firstname + " " + userobj.Lastname
            else:
                r['Username'] = ""

            checkintimeobj = attendance.objects.filter(employeeId=r['AttendanceId'],date=r['Date'],company_code =  r['company_code'] ).order_by('time').first()
            if checkintimeobj is not None:
                checkin = checkintimeobj.time

                checkouttimeobj = attendance.objects.filter(employeeId=r['AttendanceId'],date=r['Date'],company_code =  r['company_code'] ).order_by('time').last()
                if checkouttimeobj is not None:
                    checkout = checkouttimeobj.time

                if checkin == checkout:
                    r['UserCheckInTime'] = checkin
                    r['UserCheckOutTime'] = '-- : --'
                else:
                    r['UserCheckInTime'] = checkin
                    r['UserCheckOutTime'] = checkout
            else:
                r['UserCheckInTime'] = '-- : --'
                r['UserCheckOutTime'] = '-- : --'

            datetimestr = str(r['created_at'])
            createddate = str(datetimestr.split('T')[0])
            r['newdate'] = dd_month_year_format(createddate)

            r['requestdate'] = dd_month_year_format(str(r['Date']))
           
            strtime  = str(datetimestr.split('T')[1])
            r['newtime'] = str(strtime.split('.')[0])

            managerlist = []
            for mangr in r['manager_ids']:
                managerdict = {}
                managerObj=Users.objects.filter(id=int(mangr),company_code=request.user.company_code,is_active=True).first()
              
                if managerObj is not None:

                    managerdict['managername'] = managerObj.Firstname + " " + managerObj.Lastname
                    if managerObj.Photo is not None and managerObj.Photo !="":
                        managerdict['managerpic'] = imageUrl +"/media/"+ str(managerObj.Photo)
                    else:
                        managerdict['managerpic'] = imageUrl + "/static/assets/images/profile.png"
                else:
                    managerdict['managername']="NA"
                    managerdict['managerpic'] = imageUrl + "/static/assets/images/profile.png"

                if r['actiontakenby'] is not None:
                    if int(mangr) == int(r['actiontakenby']):
                        managerdict['actiontaken'] = True
                    else:
                        managerdict['actiontaken'] = False

                managerlist.append(managerdict)
            r['managerlist'] = managerlist
        cancelledlist = cancelserializer.data
    else:
        cancelledlist = []


    context = {
        "pendinglist" :pendinglist,
        "approvedlist":approvedlist,
        "Rejectedlist":rejectedlist,
        "cancelledlist":cancelledlist
    }

    return Response ({
        "data": context,
        "response":{
            "n" : 1,
            "msg" : "Data found successfully",
            "status" : "Success"
        }
        }) 
    

@api_view(['POST'])
def updateattendancerequestreason(request):
    request_id = request.data.get('request_id')
    Reason = request.data.get('Reason')
    if request_id is not None and request_id != "":
        attendancerequestobject = AttendanceRequest.objects.filter(id=request_id,Active=True).first()
        if attendancerequestobject is not None:
            if attendancerequestobject.Action is None:
                attendancerequestobject.Reason = Reason
                attendancerequestobject.save()
                return Response ({
                    "data": [],
                    "response":{
                        "n" : 1,
                        "msg" : "Request has been updated successfully",
                        "status" : "Success"
                    }
                            
                })
            else:
                return Response ({
                    "data": [],
                    "response":{
                        "n" : 0,
                        "msg" : "Request can't be updated ",
                        "status" : "Failed"
                    }
                            
                })
        else:
            return Response ({
                        "data": [],
                        "response":{
                            "n" : 0,
                            "msg" : "Attendance request not found",
                            "status" : "Failed"
                        }
                                
                    })
    else:
        return Response ({
                    "data": [],
                    "response":{
                        "n" : 0,
                        "msg" : "Please provide request id.",
                        "status" : "Failed"
                    }
                            
                })



@api_view(['POST'])
def cancelattendancerequest(request):
    request_id = request.data.get('request_id')
    if request_id is not None and request_id != "":
        attendancerequestobject = AttendanceRequest.objects.filter(id=request_id,Active=True).first()
        if attendancerequestobject is not None:
            if attendancerequestobject.Action is None:
                attendancerequestobject.Active = False
                attendancerequestobject.save()
                return Response ({
                    "data": [],
                    "response":{
                        "n" : 1,
                        "msg" : "Request has been cancelled successfully",
                        "status" : "Success"
                    }
                            
                })
            else:
                return Response ({
                    "data": [],
                    "response":{
                        "n" : 0,
                        "msg" : "Request can't been cancelled ",
                        "status" : "Failed"
                    }
                            
                })
        else:
            return Response ({
                        "data": [],
                        "response":{
                            "n" : 0,
                            "msg" : "Attendance request not found",
                            "status" : "Failed"
                        }
                                
                    })
    else:
        return Response ({
                    "data": [],
                    "response":{
                        "n" : 0,
                        "msg" : "Please provide request id.",
                        "status" : "Failed"
                    }
                            
                })





@api_view(['POST'])
def addSecondaryInfo(request):
    data = {}
    data['userId'] = request.data.get('userId')
    updateObject =  UserSecondaryInfo.objects.filter(userId=data['userId']).first()
    ser = UserSecondarySerializer(updateObject)
    if ser.data['adhaarcardimage'] is None:
        data['adhaarcardimage'] = request.FILES.get('adhaarcardimage')
    elif request.FILES.get('adhaarcardimage') is None:
        pass
    else:
        data['adhaarcardimage'] = request.FILES.get('adhaarcardimage')
    if ser.data['pancardimage'] is None:
        data['pancardimage'] = request.FILES.get('adhaarcardimage')
    elif request.FILES.get('pancardimage') is None:
        pass
    else:
        data['pancardimage'] = request.FILES.get('pancardimage')
    data['permanentaddress'] = request.data.get('permanentaddress')
    data['bloodgroup'] = request.data.get('bloodgroup')
    data['relation1'] = request.data.get('relation1')
    data['relation1number'] = request.data.get('relation1number')
    data['relation2'] = request.data.get('relation2')
    data['relation2number'] = request.data.get('relation2number')
    data['refname1'] = request.data.get('refname1')
    data['refdesg1'] = request.data.get('refdesg1')
    data['refnumber1'] = request.data.get('refnumber1')
    data['refemail1'] = request.data.get('refemail1')
    data['refname2'] = request.data.get('refname2')
    data['refdesg2'] = request.data.get('refdesg2')
    data['refnumber2'] = request.data.get('refnumber2')
    data['refemail2'] = request.data.get('refemail2')
    data['comapanyname'] = request.data.get('comapanyname')
    data['companyaddress'] = request.data.get('companyaddress')
    data['bankname'] = request.data.get('bankname')
    data['ifsccode'] = request.data.get('ifsccode')
    data['accountnumber'] = request.data.get('accountnumber')
    data['confirmaccountnumber'] = request.data.get('confirmaccountnumber')
    data['adhaarcard'] = request.data.get('adhaarcard')
    data['pancard'] = request.data.get('pancard')
    data['company_code'] = request.user.company_code
    data['adhaarcardimage'] = request.FILES.get('adhaarcardimage') 
    data['pancardimage'] = request.FILES.get('pancardimage') 

    userSecondaryObject = UserSecondaryInfo.objects.filter(userId=data['userId']).first()
    if data['accountnumber'] == data['confirmaccountnumber']:

        if userSecondaryObject is None:
            serializer = UserSecondarySerializer(data=data)
           
            if serializer.is_valid():
                serializer.save()
                userobj=Users.objects.filter(id=int(data['userId']),company_code=request.user.company_code,is_active=True).first()
                if userobj is not None:
                    userobj.secondary_info = True
                    userobj.save()
                response_ = {
                    "data": serializer.data,
                    "response":{
                        "n" : 1,
                        "msg" : "Info updated successfully",
                        "status" : "success"
                    }
                    }
                return Response (response_,status=200)
            else:
              
                response_ = {
                    "data": serializer.errors,
                    "response":{
                        "n" : 0,
                        "msg" : "Error updating info",
                        "status" : "Failed"
                    }
                    }
                return Response (response_,status=200)
        else:
            updateserializer = UserSecondarySerializer(userSecondaryObject,data=data)
            if updateserializer.is_valid():
                updateserializer.save()
                userobj=Users.objects.filter(id=int(data['userId']),company_code=request.user.company_code,is_active=True).first()
                if userobj is not None:
                    userobj.secondary_info = True
                    userobj.save()
                response_ = {
                    "data": updateserializer.data,
                    "response":{
                        "n" : 1,
                        "msg" : "Info updated successfully",
                        "status" : "success"
                    }
                    }
                return Response (response_,status=200)
            else:
             
                response_ = {
                    "data": updateserializer.errors,
                    "response":{
                        "n" : 0,
                        "msg" : "Error updating info",
                        "status" : "Failed"
                    }
                    }
                return Response (response_,status=200)
    else:
        response_ = {
            "data": {},
            "response":{
                "n" : 0,
                "msg" : "Account number not matched",
                "status" : "Failed"
            }
            }
        return Response (response_,status=200)        

@api_view(['GET'])
@permission_classes((AllowAny,))
def getSecondaryInfo(request):
    userId = request.GET.get('userId')
    userSecondaryObject = UserSecondaryInfo.objects.filter(userId=userId).order_by('id').last()
    if userSecondaryObject:
        serializer = UserSecondarySerializer(userSecondaryObject)
        response_ = {
            "data": serializer.data,
            "response":{
                "n" : 1,
                "msg" : "User found successfully",
                "status" : "success"
            }
            }
        return Response (response_,status=200)
    else:
        response_ = {
            "data": {},
            "response":{
                "n" : 0,
                "msg" : "User not found",
                "status" : "Failed"
            }
            }
        return Response (response_,status=200)

@api_view(['GET'])
@permission_classes((AllowAny,))
def getSecondaryInfofordocumentverification(request):
    userId = request.GET.get('userId')
    userSecondaryObject = UserSecondaryInfo.objects.filter(userId=userId).order_by('id').last()
    if userSecondaryObject:
        serializer = UserSecondarySerializer(userSecondaryObject)
        data={}
        data=serializer.data
        rcountryobj=Country.objects.filter(id=data['residentialaddresscountry']).first()
        rstateobj=State.objects.filter(id=data['residentialaddressstate']).first()
        rcityobj=Cities.objects.filter(id=data['residentialaddresscity']).first()
        data['rcountryname']="NA"
        data['rstatename']="NA"
        data['rcityname']="NA"
        
        if rcountryobj:
            data['rcountryname']=rcountryobj.name
        if rstateobj:
            data['rstatename']=rstateobj.name
        if rcityobj:
            data['rcityname']=rcityobj.name


        pcountryobj=Country.objects.filter(id=data['permanantaddresscountry']).first()
        pstateobj=State.objects.filter(id=data['permanantaddressstate']).first()
        pcityobj=Cities.objects.filter(id=data['permanantaddresscity']).first()
        data['pcountryname']="NA"
        data['pstatename']="NA"
        data['pcityname']="NA"
        
        if pcountryobj:
            data['pcountryname']=pcountryobj.name
        if pstateobj:
            data['pstatename']=pstateobj.name
        if pcityobj:
            data['pcityname']=pcityobj.name



        response_ = {
            "data": data,
            "response":{
                "n" : 1,
                "msg" : "User found successfully",
                "status" : "success"
            }
            }
        return Response (response_,status=200)
    else:
        response_ = {
            "data": {},
            "response":{
                "n" : 0,
                "msg" : "User not found",
                "status" : "Failed"
            }
            }
        return Response (response_,status=200)



@api_view(['GET'])
@permission_classes((AllowAny,))
def getpreviouscompany(request):
    userid = request.GET.get('userId')
    userpreviouscompanyObject = Previous_Company_Details.objects.filter(userid=userid).order_by('joinDate')
    if userpreviouscompanyObject:
        serializer = Previous_Company_Details_Serializer(userpreviouscompanyObject,many=True)
        for i in serializer.data:

            i['formatedjoinDate']=dd_month_year_format(i['joinDate'])
            i['formatedleaveDate']=dd_month_year_format(i['leaveDate'])
        response_ = {
            "data": serializer.data,
            "response":{
                "n" : 1,
                "msg" : "company found successfully",
                "status" : "success"
            }
            }
        return Response (response_,status=200)
    else:
        response_ = {
            "data": {},
            "response":{
                "n" : 0,
                "msg" : "company not found",
                "status" : "Failed"
            }
            }
        return Response (response_,status=200)


@api_view(['GET'])
@permission_classes((AllowAny,))
def geteducatiopnalqualifications(request):

    userid = request.GET.get('userId')
    usereducationalObject = educational_qualifications.objects.filter(userid=userid).order_by('fromdate')
    if usereducationalObject:
        serializer = educational_qualificationsSerializer(usereducationalObject,many=True)
        for i in serializer.data:
            i['formatedfromdate']=dd_month_year_format(i['fromdate'])
            i['formatedtodate']=dd_month_year_format(i['todate'])
            
        response_ = {
            "data": serializer.data,
            "response":{
                "n" : 1,
                "msg" : "Qualifications found successfully",
                "status" : "success"
            }
            }
        return Response (response_,status=200)
    else:
        response_ = {
            "data": {},
            "response":{
                "n" : 0,
                "msg" : "qualifications not found",
                "status" : "Failed"
            }
            }
        return Response (response_,status=200)




@api_view(['POST'])
@permission_classes((AllowAny,))
def geteducatiopnalqualificationsbyemail(request):
    email=request.POST.get('email')

    userobj=Users.objects.filter(personal_email=email,is_active=True).first()

    if userobj:
        userid = userobj.id
        usereducationalObject = educational_qualifications.objects.filter(userid=userid).order_by('id')
        if usereducationalObject.count() != 0:
            serializer = educational_qualificationsSerializer(usereducationalObject,many=True)
            response_ = {
                "data": serializer.data,
                "response":{
                    "n" : 1,
                    "msg" : "Qualifications found successfully",
                    "status" : "success"
                }
                }
            return Response (response_,status=200)
        else:
            response_ = {
                "data": {},
                "response":{
                    "n" : 0,
                    "msg" : "qualifications not found",
                    "status" : "Failed"
                }
                }
            return Response (response_,status=400)
    else:

        response_ = {
            "data": {},
            "response":{
                "n" : 0,
                "msg" : "User not found.",
                "status" : "Failed"
            }
            }
        return Response (response_,status=400)




@api_view(['POST'])
@permission_classes((AllowAny,))
def addcountry(request):
    statelist=[]
    for i in statelist:
        State.objects.create(statename=i,country=78)
    response_ = {
        "data": {},
        "response":{
            "n" : 0,
            "msg" : "done",
            "status" : "success"
        }
        }
    return Response (response_,status=400)

@api_view(['POST'])
def pindepartment(request):
    department_id=request.POST.get('department_id')
    if department_id is not None and department_id !='':
        department_obj=Department.objects.filter(id=department_id,Active=True).first()
        if department_obj is not None:
            previous_pined=ManagerPinedDepartmentMaster.objects.filter(user_id=request.user.id,is_active=True).first()
            if previous_pined is not None:
                previous_pined_obj=ManagerPinedDepartmentMaster.objects.filter(user_id=request.user.id,is_active=True).update(pined=False)
                new_pined_obj=ManagerPinedDepartmentMaster.objects.filter(user_id=request.user.id,is_active=True,department_id=department_id).first()
                if new_pined_obj is not None:
                    new_pined_obj.update(pined=True)
                else:
                    new_pined_obj=ManagerPinedDepartmentMaster.objects.create(user_id=request.user.id,is_active=True,department_id=department_id,pined=True)
                
                return Response ({"data": {},"response":{"n" : 0,"msg" : "department pined successfully ","status" : "success"}})
            else:
                department_list_obj=Department.objects.filter(Active=True)  
                department_serializer=DepartmentSerializer(department_list_obj,many=True)
                for department in department_serializer.data:
                    create_new=ManagerPinedDepartmentMaster.objects.create(user_id=request.user.id,is_active=True,department_name=department['DepartmentName'],department_id=department['id'])
                    
                previous_pined_obj=ManagerPinedDepartmentMaster.objects.filter(user_id=request.user.id,is_active=True).update(pined=False)
                pin_newly_added=ManagerPinedDepartmentMaster.objects.filter(user_id=request.user.id,is_active=True,department_id=department_id).update(pined=True)
                return Response ({"data": {},"response":{"n" : 0,"msg" : "department pined successfully","status" : "success"}})
        else:
            return Response ({"data": {},"response":{"n" : 1,"msg" : "department not found","status" : "error"}})
    else:
        return Response ({"data": {},"response":{"n" : 1,"msg" : "please provide department id","status" : "error"}})



@api_view(['GET'])
@permission_classes((AllowAny,))
def getecountries(request):
    countries_obj=Country.objects.all()
    serializer=CountrySerializer(countries_obj,many=True)

    response_ = {
        "data": serializer.data,
        "response":{
            "n" : 0,
            "msg" : "done",
            "status" : "success"
        }
        }
    return Response (response_,status=200)

@api_view(['POST'])
@permission_classes((AllowAny,))
def getstatesbycountryid(request):
    countryid=request.POST.get("countryid")
    countries_obj=State.objects.filter(country_id=countryid)
    if countries_obj.count() == 0:
        response_ = {
        "data": {},
        "response":{
            "n" : 0,
            "msg" : "We don't have states releted to this country",
            "status" : "failed"
        }
        }
    else:

        serializer=StateSerializer(countries_obj,many=True)

        response_ = {
            "data": serializer.data,
            "response":{
                "n" : 0,
                "msg" : "State found Successfully",
                "status" : "success"
            }
            }
    return Response (response_,status=200)

@api_view(['POST'])
def getcountrystatebycityid(request):
    cityid = request.POST.get("cityid")

    if cityid is not None and cityid !='' and cityid !='undefined':

        city_obj = Cities.objects.filter(id=cityid).first()
        if city_obj is not None:
            cityser = CitiesSerializer(city_obj)
            citydata = []
            for i in [cityser.data]:
                statename = State.objects.filter(id=i['state']).first()
                countryname = Country.objects.filter(id=i['country']).first()
                i['statename'] = statename.name
                i['countryname'] = countryname.name
                citydata.append(i)

            return Response ({"data":citydata[0],"response":{"n" : 1,"msg" : "data found successfully","status" : "success"}})
        else:
            return Response({"response":{"n": 0 ,"msg" : "city not found","status":"error"}})  
    return Response({"response":{"n": 0 ,"msg" : "city id not found","status":"error"}})  




@api_view(['POST'])
def searchcity(request):

    cityname = request.POST.get("cityname")
    if cityname != "":
        cityobj = Cities.objects.filter(name__icontains=cityname)[:20]
        cityser = CitiesSerializer(cityobj,many=True)   

    else:
        cityobj = Cities.objects.filter(country=101)
        cityser = CitiesSerializer(cityobj,many=True) 


    for i in cityser.data:
        statename_obj=State.objects.filter(id=i['state']).first()
        countryname_obj=Country.objects.filter(id=i['country']).first()
        if statename_obj is not None:
            i['statename']=statename_obj.name
        else:
            i['statename']='---'

        if countryname_obj is not None:
            i['countryname']=countryname_obj.name
        else:
            i['countryname']='---'

    return Response ({"data":cityser.data,"response":{"n" : 1,"msg" : "data found successfully","status" : "success"}})




@api_view(['POST'])
@permission_classes((AllowAny,))
def getcitiesbystateid(request):
    stateid=request.POST.get("stateid")
    cities_obj=Cities.objects.filter(state_id=stateid)
    if cities_obj.count() == 0:
        response_ = {
        "data": {},
        "response":{
            "n" : 0,
            "msg" : "We don't have cities related to this state",
            "status" : "failed"
        }
        }
    else:

        serializer=CitiesSerializer(cities_obj,many=True)

        response_ = {
            "data": serializer.data,
            "response":{
                "n" : 0,
                "msg" : "Cities found Successfully",
                "status" : "success"
            }
            }
    return Response (response_,status=200)



@api_view(['POST'])
def addAttendance(request):
    attendancelist=[]

    com_code = request.user.company_code
    for i in attendancelist:
        Date=str(i['loginTime']).split(" ")[0]
        Time=str(i['loginTime']).split(" ")[1]
      
        empexist = attendance.objects.filter(employeeId=i['employeeId'],date=Date,time=Time).order_by('time').first()
        if empexist is None:
            attendance.objects.create(employeeId =i['employeeId'],date=Date,time=Time,company_code=com_code)
    Response_={
        "data":attendancelist,
        "response":{
            "n":1,
            "Msg":"Data added successfully.",
            "Status":"Success"
        }
    }
    return Response(Response_,status=200)




    
   

@api_view(['POST'])   
def yearlist(request):
    datelist=[]
    yearlist=[]

    empuserID=request.data.get('userID')
    empid=Users.objects.filter(id=empuserID,company_code=request.user.company_code,is_active=True).first()
   
    employee_id=empid.employeeId

    empdate=attendance.objects.filter(employeeId=employee_id)
    serializer=attendanceserializer(empdate,many=True)
    for i in serializer.data:
        dates=i['date']
        datelist.append(dates)
   
    
    datelist = list(set(datelist))
 

    for i in datelist:
        idate = i
        dateyear = datetime.datetime.strptime(idate, "%Y-%m-%d")
        yearlist.append(dateyear.year)
    
    yearlist=list(set(yearlist))
   
   
    return Response ({"data":yearlist,"response":{"n" : 1,"msg" : "SUCCESS","status" : "success"}})

@api_view(['POST'])    
def monthlist(request):
    datelist=[]
    newdatelist=[]
    monthlist=[]
    year=request.data.get('year')
 
    empuserID=request.data.get('userID')
    empid=Users.objects.filter(id=empuserID,company_code=request.user.company_code,is_active=True).first()
    employee_id=empid.employeeId
    empdate=attendance.objects.filter(employeeId=employee_id)
    serializer=attendanceserializer(empdate,many=True)
    for i in serializer.data:
        dates=i['date']
        datelist.append(dates)
 

    for i in datelist:
        idate = i
        dateyear = datetime.datetime.strptime(idate, "%Y-%m-%d")
        lyear= dateyear.year
        lyear=str(lyear)
        if year == lyear:
            newdatelist.append(i)
  
    
    for j in newdatelist:
        idate = j
        datemonth = datetime.datetime.strptime(idate, "%Y-%m-%d")
        monthd=datemonth.month
        monthname=calendar.month_name[monthd]
        monthlist.append(monthname)
    
    monthlist=list(set(monthlist))


    return Response ({"data":monthlist,"response":{"n" : 1,"msg" : "SUCCESS","status" : "success"}})


# =================================================================start custom functions ==================================================================

def calculate_shift_duration(shift_info):
    shift_start = datetime.datetime.strptime(shift_info['shiftstartdate'] + ' ' + shift_info['shiftstarttime'], '%Y-%m-%d %H:%M:%S')
    shift_end = datetime.datetime.strptime(shift_info['shiftenddate'] + ' ' + shift_info['shiftendtime'], '%Y-%m-%d %H:%M:%S')
    duration = shift_end - shift_start
    duration -= timedelta(hours=2)  # Subtracting 2 hours from the duration

    return duration

def convert_datetime(input_datetime_str):
    try:
        # Parse the input datetime string
        input_datetime = datetime.datetime.strptime(input_datetime_str, "%Y-%m-%d %H:%M:%S")
        # Format the datetime object according to the desired output format
        output_datetime_str = input_datetime.strftime("%I:%M %p %d-%m-%Y")

        return output_datetime_str
    except ValueError:
        return "--:--"
    
def convert_to_12_hour_format(time_str):
    # Convert the input time string to a datetime object
    if time_str !='' and time_str is not None:
        
        time_obj = datetime.datetime.strptime(time_str, "%H:%M:%S")

        # Format the time object as "hh:mm am/pm"
        formatted_time = time_obj.strftime("%I:%M %p")

        return formatted_time
    return '--:--'

def get_day(date_str):
    try:
        # Parse the input date string
        date_obj = datetime.datetime.strptime(date_str, "%d-%m-%Y")
        # Get the day of the week (0 = Monday, 1 = Tuesday, ..., 6 = Sunday)
        day_of_week = date_obj.strftime("%A")  # Format the day as a string (e.g., "Monday")

        return day_of_week
    except ValueError:
        return ""

def get_dates_between(start_date, end_date):
    start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
    end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')

    dates = []
    current_date = start_date

    while current_date <= end_date:
        dates.append(current_date.strftime('%Y-%m-%d'))
        current_date += timedelta(days=1)

    return dates
    
def convert_to_dd_month(date_str):
    # Parse the input date string
    date_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d')

    # Format the date as 'DD Month' (e.g., '12 Aug')
    formatted_date = date_obj.strftime('%d %b')

    return formatted_date

def dd_mm_yyyy(input_date):
    # Parse input date string to datetime object
    output_date=''
    if input_date !='':
        
        input_date_object = datetime.datetime.strptime(input_date, "%Y-%m-%d")
        
        # Format the datetime object as a string in the desired format
        output_date = input_date_object.strftime("%d-%m-%Y")
        
        return output_date
    else:
        return output_date
               
        
def hh_mm(input_time):
    # Parse input time string to datetime object
    output_time=''
    if input_time !='':
        
        input_time_object = datetime.datetime.strptime(input_time, "%H:%M:%S")
        
        # Format the datetime object as a string in the desired time format
        output_time = input_time_object.strftime("%H:%M")
        
        return output_time
    else:
        return output_time
        
def conver_created_at_date(datestr):
    datetimestr = str(datestr)
    createddate = str(datetimestr.split('T')[0])
    datestr = dd_month_year_format(createddate)
    return datestr


def convertdate(date):
    if date != "" and date is not None:
        datetime_obj = datetime.datetime.strptime(date, "%Y-%m-%d")
        day = datetime_obj.day
        month = datetime_obj.strftime("%B")
        year = datetime_obj.year

        # convert day to a string with appropriate suffix
        if 4 <= day <= 20 or 24 <= day <= 30:
            suffix = "th"
        else:
            suffix = ["st", "nd", "rd"][day % 10 - 1]

        # format the date in "Dth Month YYYY" format
        formatted_date = f"{day}{suffix} {month} {year}"

        return formatted_date
    return "---"
 
def dateformat_ddmmyy(datestr):
    date_object = datetime.datetime.strptime(datestr, '%Y-%m-%d')
    formatted_date = date_object.strftime('%d-%m-%Y')
    return(formatted_date)

def checkholiday(date):
    holidaylist=[]
    #holiday list
    holidatlist = Holidays.objects.filter(Active=True).order_by('id')
    serializer = holidaysSerializer(holidatlist, many=True)
    for i in serializer.data:
        holiday=i['Date']
        holidaylist.append(holiday) 

    for i in holidaylist:
        if i == str(date):
            return True 

def is_leave_time_valid(leave_date_str):
    current_datetime = datetime.datetime.now()
    leave_date = datetime.datetime.strptime(leave_date_str, '%Y-%m-%d')
    leave_datetime = datetime.datetime(leave_date.year, leave_date.month, leave_date.day, 9, 30)
    if current_datetime > leave_datetime:
        return False
    else:
        return True


def WorkFromHome(leaveid):
    status=""
    leave_obj=Leave.objects.filter(id=leaveid).first()
    if leave_obj.WorkFromHome ==True:
        status="work from home"
    else:
        status="leave"
    return status

def wfhsorter(leaveid):
    status=""
    leave_obj=Leave.objects.filter(id=leaveid).first()
    if leave_obj.WorkFromHome ==True:
        status="WFH"
    else:
        status="Leave"
    return status

def countdays(date1,date2):
    datetime1 = datetime.datetime.strptime(date1, "%Y-%m-%d")
    datetime2 = datetime.datetime.strptime(date2, "%Y-%m-%d")
    timedelta = datetime2 - datetime1
    days = timedelta.days+1
    return days 

def get_date_range(start_date_str, end_date_str):
    start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d").date()
    end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%d").date()
    
    date_range = []
    current_date = start_date
    
    while current_date <= end_date:
        date_range.append(current_date)
        current_date += timedelta(days=1)
    
    return date_range



def calculate_days(date1, date2, date_type):
    date1 = datetime.datetime.strptime(date1, '%Y-%m-%d')
    date2 = datetime.datetime.strptime(date2, '%Y-%m-%d')

    # Calculate the number of days between two dates
    days_difference = (date2 - date1).days
    # Adjust for half day or full day
    if date_type == 'FirstHalf':
        days_difference += 0.5
    elif date_type == 'SecondHalf':
        days_difference += 0.5
    elif date_type == 'Fullday':
        days_difference += 1
    else:
        raise ValueError("Invalid date type. Use 'halfday' or 'fullday'.")

    return days_difference

def dateformat(input_date):
    
    datetime_obj = datetime.datetime.strptime(input_date, "%Y-%m-%d")
    day = datetime_obj.day
    month = datetime_obj.strftime("%B")
    year = datetime_obj.year
    if 4 <= day <= 20 or 24 <= day <= 30:
        suffix = "th"
    else:
        suffix = ["st", "nd", "rd"][day % 10 - 1]
    formatted_date = f"{day}{suffix} {month} {year}"
    return formatted_date
        
def date_handling(date_str1, date_str2):
    date1 = datetime.datetime.strptime(date_str1, "%Y-%m-%d")
    date2 = datetime.datetime.strptime(date_str2, "%Y-%m-%d")

    if date1 == date2:
        return convertdate(date1.strftime("%Y-%m-%d"))  # Return single date

    # Return date range
    return f"{convertdate(date1.strftime('%Y-%m-%d'))} to {convertdate(date2.strftime('%Y-%m-%d'))}"



def convertdate2(input_date):
    try:
        # Parse the input date string into a datetime object
        date_obj = datetime.datetime.strptime(input_date, '%Y-%m-%d')

        # Format the date in the desired output format
        formatted_date = date_obj.strftime('%d %b %y')

        return formatted_date
    except ValueError:
        return " "
    
    
def date_handling2(date_str1, date_str2):
    date1 = datetime.datetime.strptime(date_str1, "%Y-%m-%d")
    date2 = datetime.datetime.strptime(date_str2, "%Y-%m-%d")

    if date1 == date2:
        return convertdate2(date1.strftime("%Y-%m-%d"))  # Return single date

    # Return date range
    return f"{convertdate2(date1.strftime('%Y-%m-%d'))} to {convertdate2(date2.strftime('%Y-%m-%d'))}"


def dd_month_format(input_date):
    # Convert the input date string to a datetime object
    date_obj = datetime.datetime.strptime(input_date, "%Y-%m-%d")

    # Extract day, month, and year
    day = date_obj.strftime("%d")
    month = date_obj.strftime("%b")

    # Determine the day suffix (e.g., "st", "nd", "rd", "th")
    if day.endswith(('11', '12', '13')):
        suffix = "th"
    else:
        last_digit = int(day[-1])
        if last_digit == 1:
            suffix = "st"
        elif last_digit == 2:
            suffix = "nd"
        elif last_digit == 3:
            suffix = "rd"
        else:
            suffix = "th"

    # Format the result
    formatted_date = f"{day}{suffix} {month.upper()}"
    return formatted_date

def dd_month_year_format(input_date):
    # Parse the input date string
    date_obj = datetime.datetime.strptime(input_date, '%Y-%m-%d')
    
    # Define the month abbreviations
    months = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN',
              'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']
    
    # Get the day with an appropriate suffix
    day = date_obj.day
    if 4 <= day <= 20 or 24 <= day <= 30:
        day_suffix = 'th'
    else:
        day_suffix = ['st', 'nd', 'rd'][day % 10 - 1]
    
    # Format the date in the desired format
    formatted_date = f"{day:02d}{day_suffix} {months[date_obj.month - 1]} {date_obj.year}"
    
    return formatted_date





def get_range_of_dates(start_date_str, end_date_str):
    # Convert string dates to datetime objects
    start_date = datetime.datetime.strptime(str(start_date_str), '%Y-%m-%d')
    end_date = datetime.datetime.strptime(str(end_date_str), '%Y-%m-%d')

    # Initialize an empty list to store the dates
    date_list = []

    # Generate the list of dates
    current_date = start_date
    while current_date <= end_date:
        date_list.append(current_date.strftime('%Y-%m-%d'))
        current_date += timedelta(days=1)

    return date_list
# =================================================================end custom functions ==================================================================
       












@api_view(['POST'])
def checkAttendanceData(request):
    data=request.data
    userObj=Users.objects.filter(id=data['UserId'],company_code=request.user.company_code,is_active=True).first()
    if userObj is not None: 
        attendanceObj=attendance.objects.filter(employeeId=userObj.employeeId,date=data['date']).order_by('time').first()
        if attendanceObj is not None:
            lastObj=attendance.objects.filter(employeeId=userObj.employeeId,date=data['date']).exclude(time=attendanceObj.time).order_by('-time').first()
            if lastObj is None:
                Response_={
                    "data":{},
                    "response":{
                        "n":1,
                        "Msg":"data not found.",
                        "Status":"Success"
                    }
                }
                return Response(Response_,status=200)
            loginTime = attendanceObj.time
            logoutTime = lastObj.time
            t1 = datetime.datetime.strptime(loginTime, "%H:%M:%S")
            t2 = datetime.datetime.strptime(logoutTime, "%H:%M:%S")

            diff = t2- t1 
            wh="08:00:00"
            tp="00:00:00"
            ex=datetime.datetime.strptime(wh, "%H:%M:%S")
            
            tp=datetime.datetime.strptime(tp, "%H:%M:%S")
            tdiff=ex-tp
            if diff < tdiff :
            
                Response_={
                    "data":{},
                    "response":{
                        "n":1,
                        "Msg":"data not found.",
                        "Status":"Success"
                    }
                }
                return Response(Response_,status=200)
            Response_={
                "data":{},
                "response":{
                    "n":0,
                    "Msg":"data found successfully.",
                    "Status":"Failed"
                }
            }
            return Response(Response_,status=200)
        Response_={
                "data":{},
                "response":{
                    "n":0,
                    "Msg":"data found successfully.",
                    "Status":"Failed"
                }
            }
        return Response(Response_,status=200)
    Response_={
                "data":{},
                "response":{
                    "n":0,
                    "Msg":"data found successfully.",
                    "Status":"Failed"
                }
            }
    return Response(Response_,status=200)

@api_view(['POST'])
@permission_classes([AllowAny])
def ApiKeygenerate(request):
    N = 60
    res = ''.join(secrets.choice(str('!@#$%^&*') + string.ascii_lowercase + string.digits )
              for i in range(N))
    data = {}
    data['api_key'] = res
    serializer = ApiSerializer(data=data)
    apikeyObject = ApiKey.objects.all()
    if apikeyObject is not None:
        ApiKey.objects.update(isActive=False)
    if serializer.is_valid():
        serializer.save()
        return Response ({
            "data": serializer.data,
            "response":{
                "n" : 1,
                "msg" : "Data found successfully",
                "status" : "success"
            }
            })
    else:
        return Response ({
            "data": serializer.errors,
            "response":{
                "n" : 0,
                "msg" : "Failed",
                "status" : "errors"
            }
            }) 




@api_view(['POST'])
def employee_filter(request):
    data={}
    data['name']=request.POST.get('name')
    data['department']=request.POST.get('department')
    data['designation']=request.POST.get('designation')
    data['location']=request.POST.get('location')
    data['attendance'] = request.POST.get('attendance')
    company_code = request.user.company_code
    allemployeeobject = Users.objects.filter(is_active=True,company_code=company_code)
    firstobject = ""
    if data['department'] != "":
        firstobject = allemployeeobject.filter(DepartmentID=int(data['department']))
    else:
        firstobject = allemployeeobject

    secondobject = ""

    if data['location'] != "":
        secondobject = firstobject.filter(locationId=int(data['location']))
    else:
        secondobject = firstobject

    thirdobject = ""

    if data['designation'] != "":
        thirdobject = secondobject.filter(DesignationId=int(data['designation']))
    else:
        thirdobject = secondobject

    

    fourthobject = ""
    
    if data['attendance'] != "":
        if data['attendance'] == "Yes":
            fourthobject = thirdobject.filter(employeeId__isnull =False)
        else:
            fourthobject = thirdobject.filter(employeeId__isnull =True)
    else:
        fourthobject = thirdobject

    employeesobject = fourthobject


    Serializer=UsersSerializer(employeesobject,many=True)
    newlist=[]
    for user in Serializer.data:
        user['name']=user['Firstname'].lower()  +" "+user['Lastname'].lower()
        
        
        if user['DateofJoining'] is not None:
            stdate = str(user['DateofJoining'])
            startmonth_name = calendar.month_abbr[int(stdate.split('-')[1])]    
            user['DateofJoining'] = stdate.split('-')[2]+" "+startmonth_name+" "+stdate.split('-')[0]
        else:
            user['DateofJoining'] = "--"
        
        
        if data['name'] != "":
            if data['name'].lower() in user['name'].lower():
                secondary_info_obj=UserSecondaryInfo.objects.filter(userId=user['id']).first()
                if secondary_info_obj:
                    secondary_info_obj_serializer=UserSecondarySerializer(secondary_info_obj)
                    user['adharcard']=secondary_info_obj_serializer.data['adhaarcard']
                    user['pancard']=secondary_info_obj_serializer.data['pancard']
                else:
                    user['adharcard']="None"
                    user['pancard']="None"



                newlist.append(user)
            
        else:
            secondary_info_obj=UserSecondaryInfo.objects.filter(userId=user['id']).first()
            if secondary_info_obj:
                secondary_info_obj_serializer=UserSecondarySerializer(secondary_info_obj)
                user['adharcard']=secondary_info_obj_serializer.data['adhaarcard']
                user['pancard']=secondary_info_obj_serializer.data['pancard']
            else:
                user['adharcard']="None"
                user['pancard']="None"
            newlist.append(user)





    return Response({"data": newlist, "response":{"n" : 1,"msg" : "Successful","status" : "success" }})
  
@api_view(['GET'])
def employee_master_emp_list(request, format=None):
    if request.method == 'GET':
        company_code = request.user.company_code
        user = Users.objects.filter(
            is_active=True,company_code=company_code).order_by('Firstname')
        serializer = UsersSerializer(user,many=True)
        for i in serializer.data:
            if i['DateofJoining'] is not None:
                stdate = str(i['DateofJoining'])
                startmonth_name = calendar.month_abbr[int(stdate.split('-')[1])]    
                i['DateofJoining'] = stdate.split('-')[2]+" "+startmonth_name+" "+stdate.split('-')[0]
            else:
                i['DateofJoining'] = ""

        return Response({'n':1,'msg':'Employee list fetched successfully','status':'success','data':serializer.data})


@api_view(['GET'])
def user_emp_list(request, format=None):
    if request.method == 'GET':
        company_code = request.user.company_code
        user = Users.objects.filter(
            is_active=True,company_code=company_code).order_by('id')
        serializer = UsersSerializer(user,many=True)
        userlist=[]
        for i in serializer.data:
            if i['DateofJoining'] is not None:
                stdate = str(i['DateofJoining'])
                startmonth_name = calendar.month_abbr[int(stdate.split('-')[1])]    
                i['DateofJoining'] = stdate.split('-')[2]+" "+startmonth_name+" "+stdate.split('-')[0]
            else:
                i['DateofJoining'] = "--"
    
           
            secondary_info_obj=UserSecondaryInfo.objects.filter(userId=i['id']).first()
            secondary_info_obj_serializer=UserSecondarySerializer(secondary_info_obj)
            if secondary_info_obj:
                i['adharcard']=secondary_info_obj_serializer.data['adhaarcard']
                i['pancard']=secondary_info_obj_serializer.data['pancard']
            else:
                i['adharcard']="Adharcard"
                i['pancard']="Pancard"
                
            allemp = leaveMapping.objects.filter(employeeId=i['id'])        
            filter1= leaveMapping.objects.filter(employeeId=i['id'],position="1",company_code=company_code).first()
            if filter1 is not None: 
                i['manager1']=int(filter1.managerId)
            else:
                i['manager1']="None" 
           

            filter2= leaveMapping.objects.filter(employeeId=i['id'],position="2",company_code=company_code).first()
          

            if filter2 is not None:
              
                i['manager2']=int(filter2.managerId)
            else:
                i['manager2']="None"

            filter3= leaveMapping.objects.filter(employeeId=i['id'],position="3",company_code=company_code).first()
           

            if filter3 is not None:
                i['manager3']=int(filter3.managerId)
            else:
                i['manager3']="None"

            userlist.append(i)

        return Response({'n':1,'msg':'Employee list fetched successfully','status':'success','data':userlist})











@api_view(['GET'])
def locationcountdata(request):
    locationlist=[]
    locationcount=[]
    com_code = request.user.company_code
    locationobj = Location.objects.filter(Active=True,company_code=com_code).order_by('id')
    locser = LocationSerializer(locationobj,many=True)
    for i in locser.data:
        locationlist.append(i['LocationName'])
        userobjexist = Users.objects.filter(locationId=i['id'],company_code=request.user.company_code,is_active=True).count()
        locationcount.append(userobjexist)

    context={
        'locationlist':locationlist,
        'locationcount':locationcount
    }
   
    return Response ({"data":context,"response":{"n" : 1,"msg" : "Mapping applied successfully","status" : "success"}})


@api_view(['GET'])
def allempbdaydata(request):
    todaysdate = date.today()
    companycode = request.user.company_code
    month  = todaysdate.month
    day = todaysdate.day
  
    birthdayobj = Users.objects.filter(BirthDate__month=month,BirthDate__day=day,company_code=companycode,is_active=True).order_by('id')
    bdayuserser = UsersSerializer(birthdayobj,many=True)
    for b in bdayuserser.data:
        if b['Photo'] is not None:
            b['userimage'] = imageUrl + b['Photo']
        else:
            b['userimage'] = imageUrl + "/static/assets/images/profile.png"
    
    announcementobj = AnnounceMent.objects.filter(date=todaysdate,is_active=True,company_code=companycode)
    annser = annoucementserializer(announcementobj,many=True)
       
    context={
        'birthdaylist':bdayuserser.data,
        'announcementlist':annser.data
    }

    return Response ({"data":context,"response":{"n" : 1,"msg" : "list fetched successfully","status" : "success"}})





def month_converter(month):
    month_num = month  # month_num = 4 will work too
    month_name = datetime.datetime(1, int(month_num), 1).strftime("%B")
    return month_name


@api_view(['POST'])
def adminmnholidayapi(request):
    Month = request.data.get('Month')
    companycode = request.user.company_code
    
    if Month is not None:
        monthname = month_converter(Month)
        
        now = datetime.datetime.now()
        # Get the current year
        current_year = now.year
        leavempobj = Leave.objects.filter(start_date__month=Month,start_date__year=str(current_year),leave_status="Approved",company_code=companycode,Active=True).order_by('id')
       
        holidayser = leaveserializer(leavempobj,many=True)
        for i in holidayser.data:
            userid = int(i['employeeId'])
            userobj = Users.objects.filter(id=userid,company_code=request.user.company_code,is_active=True).first()
            if userobj.Photo is not None and userobj.Photo != "":
                i['userimage'] = imageUrl +"/media/"+ str(userobj.Photo)
            else:
                i['userimage'] = imageUrl + "/static/assets/images/profile.png"

            i['Username'] =userobj.Firstname+ " " +userobj.Lastname
            
            startdate = i['start_date']
            enddate = i['end_date']

            if str(startdate) == str(enddate):
                startday = str(startdate).split("-")[2]
                i['daystring'] = str(startday)+" "+str(monthname)
            else:
                startday = str(startdate).split("-")[2]
                endday = str(enddate).split("-")[2]

                i['daystring'] = str(startday)+" "+str(monthname)+" to " +str(endday)+" "+str(monthname)

           

            
        return Response ({"data":holidayser.data,"response":{"n" : 1,"msg" : "Data Found Successfully","status" : "success"}})
    else:
        return Response ({"data":'',"response":{"n" : 0,"msg" : "Month not found","status" : "failed"}})

   
@api_view(['POST'])
def adminmonthlyscoreapi(request):
    if request.method == 'POST':
        Week = request.data.get('Week')
        companycode = request.user.company_code
        year = date.today().year
        scoreobjs = IntermediateTrackProResult.objects.filter(Week=Week,Year=year,company_code=companycode).order_by('-TrackProPercent')
        if scoreobjs is not None:
         
            scoreser = IntermediatePostTrackProResultSerializer(scoreobjs,many=True)
            for s in scoreser.data:
                
                Usersobj = Users.objects.filter(id=s['Employee'],is_active=True,company_code=companycode).first()
             
                if str(Usersobj.Photo) is not None and str(Usersobj.Photo) != "":
                
                    s['userimage'] = imageUrl +"/media/"+ str(Usersobj.Photo)
                else:
                   
                    s['userimage'] = imageUrl + "/static/assets/images/profile.png"

                s['Username'] =Usersobj.Firstname+ " " +Usersobj.Lastname

         
            if len(scoreser.data)>10:
                toplist = scoreser.data[:5]
                for t in toplist:
                    t['trackprostatus'] = "Top"
                underlist = scoreser.data[-5:]
                for un in underlist:
                    un['trackprostatus'] = "Under"
                totallist = toplist + underlist
                context={
                    "data":totallist
                }
            else:
                context={
                    'data':scoreser.data,
                }
            return Response ({"data":context,"response":{"n" : 1,"msg" : "Data Found Successfully","status" : "success"}})
        
        else:
            context={
                        'topperlist':[],
                        'underplist':[]
                    }

            return Response ({"data":context,"response":{"n" : 1,"msg" : "Data Found Successfully","status" : "success"}})
    else:
        return Response ({"data":'',"response":{"n" : 0,"msg" : "Month not found","status" : "failed"}})

      
@api_view(['POST'])
def notfapproveactionapi(request):
    
    if request.method == 'POST':
        notfid = request.data.get('notificationId')
        leaveid = request.data.get('Leaveid')
        com_code = request.user.company_code
        acted_user = request.user.id    

        # TaskNotification.objects.filter(id=notfid).update(action_Taken=True)

        # Leaveobj = Leave.objects.filter(id=leaveid,Active=True).first()
        # if Leaveobj is not None:
        #     appid = Leaveobj.ApplicationId
        #     userid = Leaveobj.employeeId

        #     userobj = Users.objects.filter(id=int(userid),is_active=True).first()
        #     username = userobj.Firstname + " " + userobj.Lastname 

        #     startdate = str(Leaveobj.start_date) 

        #     enddate = str(Leaveobj.end_date)

        #     if startdate == enddate:
        #         notfmsg = "You have approved leave of <span class='actionuser'>" + username + "</span> dated (" +(startdate)+" )"
        #     else:
        #         notfmsg = "You have approved leave of <span class='actionuser'>" + username + "</span> dated (" +startdate+" to "+enddate+" )"


        #     TaskNotification.objects.create(UserID_id=acted_user,company_code=com_code,NotificationTypeId_id=3,NotificationTitle="",NotificationMsg=notfmsg)

        return Response ({"data":'',"response":{"n" : 1,"msg" : "Success","status" : "success"}})
   


       
@api_view(['POST'])
def notfrejectactionapi(request):
    if request.method == 'POST':
        notfid = request.data.get('notificationId')
        leaveid = request.data.get('Leaveid')
        com_code = request.user.company_code
        acted_user = request.user.id  

        manageruserobj = Users.objects.filter(id=int(acted_user),company_code=request.user.company_code,is_active=True).first()
        managername = manageruserobj.Firstname + " " + manageruserobj.Lastname  

        notfobject = TaskNotification.objects.filter(id=notfid).first()
        managerleaveid = notfobject.leaveID
       
        # TaskNotification.objects.filter(leaveID=managerleaveid,To_manager=True).update(action_Taken=True)       

        # 

        #     if startdate == enddate:
        #         notfmsg = "You have rejected leave of " + username + " dated (" +startdate+" )"
        #     else:
        #         notfmsg = "You have rejected leave of " + username + " dated (" +startdate+" to "+enddate+" )"

        #     TaskNotification.objects.create(UserID_id=acted_user,company_code=com_code,NotificationTypeId_id=3,NotificationTitle="",NotificationMsg=notfmsg)

        return Response ({"data":'',"response":{"n" : 1,"msg" : "Success","status" : "success"}})
    

@api_view(['POST'])
@authentication_classes([])
@permission_classes([])
def notfdatecheckcronejob(request):
    notlist = []
    current_date = date.today()
    notficattionobject = TaskNotification.objects.filter(To_manager = False)
    notser = TaskNotificationSerializer(notficattionobject,many=True)
    for i in notser.data:
        createdon = i['CreatedOn'].split("T")[0]
        d1 = datetime.datetime.strptime(createdon, "%Y-%m-%d")
        d2 = datetime.datetime.strptime(str(current_date), "%Y-%m-%d")
        delta = d2 - d1
        difference = delta.days
        if difference > 7:
            notlist.append(i['id'])

    notficattionobject = TaskNotification.objects.filter(id__in=notlist,To_manager = False).update(
        ReadMsg =True
    )
 
    return Response ({"data":'',"response":{"n" : 1,"msg" : "Success","status" : "success"}})

       
@api_view(['POST'])
def notffilterlistapi(request):
    if request.method == 'POST':
        notfid = request.data.get('filtertype')
        my_date = datetime.date.today()
        currentdatestr = str(my_date)
        com_code = request.user.company_code
        acted_user = request.user.id    

        if notfid is not None and notfid != "":
            notfobj = TaskNotification.objects.filter(NotificationTypeId=notfid,company_code=com_code,
            UserID=acted_user).order_by('-id')
        else:
            notfobj = TaskNotification.objects.filter(company_code=com_code,UserID=acted_user).order_by('-id') 

        if notfobj is not None:
            notfser = TaskNotificationSerializer(notfobj,many=True)
            for i in notfser.data:
              
                ndate = str(i['CreatedOn'])
                splitdate = ndate.split("T")[0]
                newdate = splitdate.split("-")[2]+"-"+splitdate.split("-")[1]+"-"+splitdate.split("-")[0]
                i['notfdate'] = newdate

                notftype = i['NotificationTypeId']
                notobj = NotificationTypeMaster.objects.filter(id=notftype).first()
                i['NotificationType'] = notobj.Type

                if i['NotificationTypeId'] == 3 :
                    leaveobject = Leave.objects.filter(id=i['leaveID'],Active=True).first()
                    if leaveobject is not None:
                        leaveapplyser = leaveserializer(leaveobject)
                        leaveapprovalobject = leaveApproval.objects.filter(leave_id=i['leaveID'],managerId = str(acted_user) ).first()
                        if leaveapprovalobject:
                            i['leaveapprovalid'] = leaveapprovalobject.id

                        appdatetime = leaveapplyser.data['created_at']
                        strdatetime = str(appdatetime)
                        appdate = strdatetime.split("T")[0]
                        apptime = strdatetime.split("T")[1]
                        appfinaltime = apptime.split(".")[0]

                        appdatestr = str(appdate)
                        startmonth_name = calendar.month_abbr[int(appdatestr.split('-')[1])]    
                        startdatestr = appdatestr.split('-')[2]+" "+startmonth_name+" "+appdatestr.split('-')[0]
                        i['applieddate'] = startdatestr
                        

                        t_str = str(appfinaltime)
                        t_obj = datetime.datetime.strptime( t_str, '%H:%M:%S')
                        #Time format representation
                        t_am_pm = t_obj.strftime('%I:%M:%S %p')
                        i['appliedtime'] = t_am_pm


                        if currentdatestr == appdatestr:
                            mmn_am_pm = t_obj.strftime('%I:%M:%S %p')
                            i['mobilenotftime'] = str(mmn_am_pm)
                        else:
                            startmonth_name = calendar.month_abbr[int(appdatestr.split('-')[1])]    
                            startdatestr = appdatestr.split('-')[2]+" "+startmonth_name
                            i['mobilenotftime'] = startdatestr

                        leaveser = leaveserializer(leaveobject)
                        for o in [leaveser.data]:
                            startd = str(o['start_date'])
                            startdateday = startd.split("-")[2]
                            startdateyear = startd.split("-")[0]
                            startdatemonth = month_converter(int(startd.split("-")[1]))
                            newstart_date = str(startdateday) +" "+ startdatemonth +" "+str(startdateyear)

                            endd = str(o['end_date'])
                            enddateday = endd.split("-")[2]
                            enddateyear = endd.split("-")[0]
                            enddatemonth = month_converter(int(endd.split("-")[1]))
                            newend_date = str(enddateday) +" "+ enddatemonth +" "+str(enddateyear)

                        i['leavemanager'] = leaveser.data
                        i['newleavestartdate'] = newstart_date
                        i['newleaveenddate'] = newend_date
                        i['leave_employee'] = ""
                    

            return Response ({"data":notfser.data,"response":{"n" : 1,"msg" : "Success","status" : "success"}})
        else:
            return Response ({"data":'',"response":{"n" : 0,"msg" : "Failed","status" : "Failed"}})


       
# @api_view(['GET'])
# def notffiltergetlistapi(request):
#         com_code = request.user.company_code
#         acted_user = request.user.id    
      
#         notfobj = TaskNotification.objects.filter(company_code=com_code,UserID=acted_user).order_by('-id') 
#         if notfobj is not None:
#             notfser = TaskNotificationSerializer(notfobj,many=True)
#             for i in notfser.data:
#                 ndate = str(i['CreatedOn'])
#                 splitdate = ndate.split(" ")[0]
#                 newdate = splitdate.split("-")[2]+"-"+splitdate.split("-")[1]+"-"+splitdate.split("-")[0]
#                 i['notfdate'] = newdate

#                 notftype = i['NotificationTypeId']
#                 notobj = NotificationTypeMaster.objects.filter(id=notftype).first()
#                 i['NotificationType'] = notobj.Type

#             return Response ({"data":notfser.data,"response":{"n" : 1,"msg" : "Success","status" : "success"}})
#         else:
#             return Response ({"data":'',"response":{"n" : 0,"msg" : "Failed","status" : "Failed"}})


       
@api_view(['GET'])
def notftypelist(request):
    com_code = request.user.company_code
    notfobj = NotificationTypeMaster.objects.filter(company_code=com_code).order_by('id')
    notfser = NotificationTypeSerializer(notfobj,many=True)
    return Response ({"data":notfser.data,"response":{"n" : 1,"msg" : "Success","status" : "success"}})

#========================================================================================================


@api_view(['POST'])
@permission_classes((AllowAny,))
def updatePersonalDetailsSecondaryInfo(request):
    id=request.POST.get('id')
    data={}
    data['Firstname']=request.POST.get('Firstname')
    data['Lastname']=request.POST.get('Lastname')
    data['Phone']=request.POST.get('Phone')
    data['BirthDate']=request.POST.get('BirthDate')
    data['alternatemail']=request.POST.get('alternatemail')
    data['email']=request.POST.get('email')
    data['bloodgroup']=request.POST.get('bloodgroup')
    data['maritalstatus']=request.POST.get('maritalstatus')
    data['Addressline2']=request.POST.get('rAddresslinetwo')
    data['Address']=request.POST.get('rAddresslineone')

    data['permanentaddressLine2']=request.POST.get('pAddresslinetwo')
    data['permanentaddress']=request.POST.get('pAddresslineone')

    data['relationname1']=request.POST.get('relationname1')
    data['relation1']=request.POST.get('relation1')
    data['relation1number']=request.POST.get('relation1number')

    data['relationname2']=request.POST.get('relationname2')
    data['relation2']=request.POST.get('relation2')
    data['relation2number']=request.POST.get('relation2number')

    data['Photo'] = request.FILES.get('Photo')
    data['adhaarcard']=request.POST.get('adhaarcard')
    data['pancard']=request.POST.get('pancard')





    data["residentialaddresscountry"]=request.POST.get("residentialaddresscountry")
    data["residentialaddressstate"]=request.POST.get("residentialaddressstate")
    data["residentialaddresscity"]=request.POST.get("residentialaddresscity")
    data["residentialaddresspincode"]=request.POST.get("residentialaddresspincode")

    data["permanantaddresscountry"]=request.POST.get("permanantaddresscountry")
    data["permanantaddressstate"]=request.POST.get("permanantaddressstate")
    data["permanantaddresscity"]=request.POST.get("permanantaddresscity")
    data["permanantaddresspincode"]=request.POST.get("permanantaddresspincode")



    adhar_pic=request.FILES.get('adhaarcardimage')
    pan_pic=request.FILES.get('pancardimage')
    passport_pic=request.FILES.get('passportimage')



    secondaryinfo_obj=UserSecondaryInfo.objects.filter(userId=id).order_by('-id').first()
    if secondaryinfo_obj:
        if secondaryinfo_obj.finalsubmit == True:
            response_={
                'status':0,
                'Msg':'Your form is already submited for verification ,unable to proceed',
                'data':{}
            }
            return Response(response_,status=200)
    if passport_pic is not None:
        data['passportimage']=request.FILES.get('passportimage')


    if adhar_pic is not None:
        data['adhaarcardimage']=request.FILES.get('adhaarcardimage')
    else:
        data['adhaarcardimage']=secondaryinfo_obj.adhaarcardimage

    if pan_pic is not None:
        data['pancardimage']=request.FILES.get('pancardimage')
    else:
        data['pancardimage']=secondaryinfo_obj.pancardimage



    
    if id is not None and id !="":
        if data['email'] is not None and data['email'] !='':
            checkemailobj = Users.objects.filter(email=data['email'],company_code=request.user.company_code,is_active=True).exclude(id=id).first()
            if checkemailobj is not None:
                response_={
                        'status':0,
                        'Msg':'Email Already Exist',
                        'data':{}
                    }
                return Response(response_,status=200)
        mobilenumberobj = Users.objects.filter(Phone=data['Phone'],company_code=request.user.company_code,is_active=True).exclude(id=id).first()
        if mobilenumberobj is not None:
            response_={
                    'status':0,
                    'Msg':'Mobile Number Already Exist',
                    'data':{}
                }
            return Response(response_,status=200)
        

        updatePersonalDetailsobj = Users.objects.filter(id=id,company_code=request.user.company_code,is_active=True).first()
        if data['Photo'] is None or data['Photo'] == "":
            data['Photo'] = updatePersonalDetailsobj.Photo
        serializer = UserSerializer(updatePersonalDetailsobj,data=data,partial=True)
        if serializer.is_valid():
            serializer.save()

        else:
            print("error",serializer.errors)
        updatesecondaryobj = UserSecondaryInfo.objects.filter(userId=id).order_by('id').last()
        if updatesecondaryobj is not None:

            secondaryserializer = UserSecondarySerializer(updatesecondaryobj,data=data,partial=True)
            
            if secondaryserializer.is_valid():

                secondaryserializer.save() 
                response_={
                    'status':1,
                    'Msg':'Details Updated Successfully.',
                    'data':secondaryserializer.data
                }
                return Response(response_,status=200)
            
            else: 
                response_={
                    'status':0,
                    'Msg':'Details Not Updated',
                    'data':secondaryserializer.errors
                }
                return Response(response_,status=200)
        else:

            data['userId'] = id
            secondaryserializer = UserSecondarySerializer(data=data,partial=True)
            if secondaryserializer.is_valid():
                secondaryserializer.save() 
                response_={
                    'status':1,
                    'Msg':'Data Saved Successfully.',
                    'data':secondaryserializer.data
                }
                return Response(response_,status=200)
            else:
                response_={
                    'status':1,
                    'Msg':'Data not Saved .',
                    'data':secondaryserializer.errors
                }
                return Response(response_,status=200)
    else:
        response_={
            'status':0,
            'Msg':'ID not found',
            'data':{}
        }
        return Response(response_,status=200)
   
@api_view(['POST'])  
@permission_classes((AllowAny,))
def updateEmployeDetailsSecondaryInfo(request):

    id=request.POST.get('id')
    data={}

    data['refname1']=request.POST.get('refname1')
    data['refdesg1']=request.POST.get('refdesg1')
    data['refnumber1']=request.POST.get('refnumber1')
    data['refemail1']=request.POST.get('refemail1')
    data['refname2']=request.POST.get('refname2')
    data['refdesg2']=request.POST.get('refdesg2')
    data['refnumber2']=request.POST.get('refnumber2')
    data['refemail2']=request.POST.get('refemail2')




    if id is not None and id !="":
        updateEmployeeobj=UserSecondaryInfo.objects.filter(userId=id).first()
        if id is not None:
            Employeeserializer = UserSecondarySerializer(updateEmployeeobj,data=data,partial=True)
            if Employeeserializer.is_valid():
                Employeeserializer.save()
                response_={
                    'status':1,
                    'Msg':'Details Updated Successfully.',
                    'data':Employeeserializer.data
                }        
                return Response(response_,status=200)
            else:
                response_={
                    'status':0,
                    'Msg':'Details Not Updated',
                    'data':Employeeserializer.errors
                }
                return Response(response_,status=200)
        else:
            response_={
                'status':0,
                'Msg':'ID must be required',
                'data':{}
            }
            return Response(response_,status=200)
    else:
        response_={
            'status':0,
            'Msg':'ID not found',
            'data':{}
        }
        return Response(response_,status=200)
    



@api_view(['POST'])  
@permission_classes((AllowAny,))
def updatecompanydetailssecondaryInfo(request):


    userid=request.POST.getlist("userid")
    companyname=request.POST.getlist("name")
    companyaddress=request.POST.getlist("address")
    designation=request.POST.getlist("designation")
    joinDate=request.POST.getlist("joinDate")
    leaveDate=request.POST.getlist("leaveDate")
    relieving=request.FILES.getlist("relieving")
    salaryslip=request.FILES.getlist("salaryslip")

    for i in range(len(companyname)):
        companay_data = {}
        companay_data['userid'] = userid[i]
        companay_data['companyaddress'] = companyaddress[i]
        companay_data['companyname'] = companyname[i]
        companay_data['designation'] = designation[i]
        companay_data['joinDate'] = joinDate[i]
        companay_data['leaveDate'] = leaveDate[i]
        companay_data['relieving'] = relieving[i]
        companay_data['salaryslip'] = salaryslip[i]
        alredy_exist_obj=Previous_Company_Details.objects.filter(userid=companay_data['userid'],companyname=companay_data['companyname']).first()
        if alredy_exist_obj:
            serializer = Previous_Company_Details_Serializer(alredy_exist_obj,data=companay_data,partial=True)
        else:
            serializer = Previous_Company_Details_Serializer(data=companay_data,partial=True)
        if serializer.is_valid():
            serializer.save()

    response_={
        'status':1,
        'Msg':'Company Details Added Successfully.',
        'data':{},
    }        
    return Response(response_,status=200)



    

@api_view(['POST'])  
@permission_classes((AllowAny,))
def add_educational_details(request):
    if request.method == 'POST':

        userid=request.POST.getlist("user")
        qualificationname=request.POST.getlist("qualificationname")
        obtainmarks=request.POST.getlist("obtainmarks")
        university=request.POST.getlist("university")
        fromdate=request.POST.getlist("fromdate")
        todate=request.POST.getlist("todate")





        checkfinalsubmit=UserSecondaryInfo.objects.filter(userId=userid[0]).first()
        if checkfinalsubmit:

            if checkfinalsubmit.finalsubmit == True:
                response_={
                    'status':400,
                    'Msg':'Your form is already submited for verification ,unable to proceed',
                    'data':{}
                }
                return Response(response_,status=400)


        for i in range(len(userid)):
            user_data = {}

            user_data['userid'] = userid[i]
            user_data['obtain_marks'] = obtainmarks[i]
            user_data['qualification_name'] = qualificationname[i]
            user_data['fromdate'] = fromdate[i]
            user_data['todate'] = todate[i]
            user_data['university'] = university[i]
            for key, value in request.FILES.items():
                if qualificationname[i] == key:
                    user_data['marksheet'] = value


            alredy_exist_obj=educational_qualifications.objects.filter(userid=user_data['userid'],qualification_name=user_data['qualification_name']).first()
            if alredy_exist_obj:
                serializer = educational_qualificationsSerializer(alredy_exist_obj,data=user_data,partial=True)
            else:
                serializer = educational_qualificationsSerializer(data=user_data,partial=True)
            if serializer.is_valid():
                serializer.save()


            

    response_={
        'status':200,
        'Msg':'educational qualifications added',
        'data':{}
    }
    return Response(response_,status=200)
    





@api_view(['POST'])  
@permission_classes((AllowAny,))
def add_user_previous_company(request):
    if request.method == 'POST':

        data={}

        data['userid']=request.POST.get('id')
        data['companyname']=request.POST.get('comapanyname')
        data['companyaddress']=request.POST.get('companyaddress')
        data['designation']=request.POST.get('designation')
        data['joinDate']=request.POST.get('joinDate')
        data['leaveDate']=request.POST.get('leaveDate')
        data['salaryslip']=request.POST.get('salarySlip')
        data['relieving']=request.POST.get('relieving')
        

        alredy_exist_obj=Previous_Company_Details.objects.filter(userid=data['userid'],companyname=data['companyname']).first()
        if alredy_exist_obj:
            serializer = Previous_Company_Details_Serializer(alredy_exist_obj,data=data,partial=True)
        else:
            serializer = Previous_Company_Details_Serializer(data=data,partial=True)
        if serializer.is_valid():
            serializer.save()



            response_={
                'status':200,
                'Msg':'company details saved',
                'data':serializer.data,
                'companyid':serializer.data['id']
            }
            return Response(response_,status=200)
    
        else:
            response_={
                'status':400,
                'Msg':'failed to saved company details ',
                'data':serializer.errors,
                'companyid':""
            }
            return Response(response_,status=400)

    else:
        response_={
            'status':400,
            'Msg':'Method not matched ',
            'data':{},
            'companyid':""
        }
        return Response(response_,status=400)




@api_view(['POST'])  
@permission_classes((AllowAny,))
def edit_user_previous_company(request):
    if request.method == 'POST':

        data={}

        data['id']=request.POST.get('id')
        data['userid']=request.POST.get('userid')
        data['companyname']=request.POST.get('comapanyname')
        data['companyaddress']=request.POST.get('companyaddress')
        data['designation']=request.POST.get('designation')
        data['joinDate']=request.POST.get('joinDate')
        data['leaveDate']=request.POST.get('leaveDate')
        salaryslip=request.POST.get('salarySlip')
        relieving=request.POST.get('relieving')
        if salaryslip is not None:
            data['salaryslip']=request.POST.get('salarySlip')
        if relieving is not None:
            data['relieving']=request.POST.get('relieving')
        
        alredy_exist_obj=Previous_Company_Details.objects.filter(id=data['id']).first()
        if alredy_exist_obj:
            serializer = Previous_Company_Details_Serializer(alredy_exist_obj,data=data,partial=True)
        else:
            serializer = Previous_Company_Details_Serializer(data=data,partial=True)
        if serializer.is_valid():
            serializer.save()



            response_={
                'status':200,
                'Msg':'company details saved',
                'data':serializer.data,
                'companyid':serializer.data['id']
            }
            return Response(response_,status=200)
    
        else:
            response_={
                'status':400,
                'Msg':'failed to saved company details ',
                'data':serializer.errors,
                'companyid':""
            }
            return Response(response_,status=400)

    else:
        response_={
            'status':400,
            'Msg':'Method not matched ',
            'data':{},
            'companyid':""
        }
        return Response(response_,status=400)




@api_view(['POST'])  
@permission_classes((AllowAny,))
def delete_user_previous_company(request):
    if request.method == 'POST':

        data={}

        data['id']=request.POST.get('id')


        
        delete_obj=Previous_Company_Details.objects.filter(id=data['id']).first()
        if delete_obj:
            delete_obj.delete()

        



            response_={
                'status':200,
                'Msg':'company details deleted successfully',
                'data':{'id':data['id']},
            }
            return Response(response_,status=200)
    
        else:
            response_={
                'status':400,
                'Msg':'failed to delete company details ',
                'data':{},
            }
            return Response(response_,status=400)

    else:
        response_={
            'status':400,
            'Msg':'Method not matched ',
            'data':{},
            'companyid':""
        }
        return Response(response_,status=400)




@api_view(['POST'])  
@permission_classes((AllowAny,))
def get_user_previous_company(request):
    if request.method == 'POST':
        data={}
        data['id']=request.POST.get('id')
        exist_obj=Previous_Company_Details.objects.filter(id=data['id']).first()
        if exist_obj:
            serializer = Previous_Company_Details_Serializer(exist_obj)
            response_={
                'status':200,
                'Msg':'company details found',
                'data':serializer.data,
            }
            return Response(response_,status=200)
    
        else:
            response_={
                'status':400,
                'Msg':'company id not found ',
                'data':{},
                'companyid':""
            }
            return Response(response_,status=400)

    else:
        response_={
            'status':400,
            'Msg':'Method not matched ',
            'data':{},
            'companyid':""
        }
        return Response(response_,status=400)



@api_view(['POST']) 
@permission_classes((AllowAny,))
def updateDetailsAndDocument(request):
    requestfiles = request.FILES
    id=request.POST.get('id')
    data={}
    data['bankname']=request.POST.get('bankname')
    data['ifsccode']=request.POST.get('ifsccode')
    data['accountnumber']=request.POST.get('accountnumber')
    data['confirmaccountnumber']=request.POST.get('confirmaccountnumber')
    data['adhaarcard']=request.POST.get('adhaarcard')
    data['pancard']=request.POST.get('pancard')

    data['adhaarcardimage']=request.FILES.get('adhaarcardimage')
    data['pancardimage']=request.FILES.get('pancardimage')
    

    if id is not None and id !="":
        updateDetailsobj = UserSecondaryInfo.objects.filter(userId=id).order_by('id').last()
        if id is not None:
            if 'adhaarcardimage' not in requestfiles.keys():
                data['adhaarcardimage'] = updateDetailsobj.adhaarcardimage
            if 'pancardimage' not in requestfiles.keys():
                data['pancardimage'] = updateDetailsobj.pancardimage
            DetailsAndDocumentserializer = UserSecondarySerializer(updateDetailsobj,data=data,partial=True)
            if DetailsAndDocumentserializer.is_valid():
                    DetailsAndDocumentserializer.save()
                    useroberj = Users.objects.filter(id=id,is_active=True).first()
                    company_code = useroberj.company_code
                    adminobj = Users.objects.filter(company_code = company_code,is_active=True,is_staff=False,is_superuser=False).first()
                   
                    adminid= adminobj.id
                    Name = useroberj.Firstname + " "+useroberj.Lastname
                    data1={}
                    data1['NotificationTitle'] = "Verification"
                    data1['NotificationMsg'] ="<span class='notfempname'>"+ Name+"</span> submitted documents for verification.<a href='/document-verification/"+encrypt_parameter(id)+"'>View Documents </a>"
                    data1['company_code'] = company_code
                    data1['NotificationTypeId'] = 4
                    data1['created_by'] = id
                    data1['UserID'] = adminid
                    data1['leaveID'] = 0
                    serializer = TaskNotificationSerializer(data=data1)
                    if serializer.is_valid():
                        serializer.save()
                        response_={
                            'status':1,
                            'Msg':'Data updated Successfully.',
                            'data':DetailsAndDocumentserializer.data
                        }
                        return Response(response_,status=200)
                    else:
                       
                        response_={
                            'status':0,
                            'Msg':'Failed to  submit documents.',
                            'data': serializer.errors
                        }
                        return Response(response_,status=200)

            else:
                response_={
                    'status':0,
                    'Msg':'Data Not Updated.',
                    'data':DetailsAndDocumentserializer.errors
                }
                return Response(response_,status=200)
        else:
            response_={
                'status':0,
                'Msg':'Id must be Required.',
                'data': {}
            }
            return Response(response_,status=200)
    else:
        response_={
            'status':0,
            'Msg':'ID Not Found.',
            'data':{}
        }
        return Response(response_,status=200)
    
@api_view(['POST'])
@permission_classes((AllowAny,))
def secondaryInfoLinkApi(request):
    data={}
    id=request.POST.get('id')


    data['bankname']=request.POST.get('bankname')
    data['branchname']=request.POST.get('branchname')
    data['accountholdername']=request.POST.get('accountholdername')
    data['ifsccode']=request.POST.get('ifsccode')
    data['accountnumber']=request.POST.get('accountnumber')
    data['confirmaccountnumber']=request.POST.get('confirmaccountnumber')
    data['previous_pf_accountno']=request.POST.get('previouspfaccountno')
    data['esic_number']=request.POST.get('esic')
    data['finalsubmit']=True


    secondaryinfo_obj=UserSecondaryInfo.objects.filter(userId=id).order_by('-id').first()
    if secondaryinfo_obj:

        if secondaryinfo_obj.finalsubmit == True:
            response_={
                'status':0,
                'Msg':'Your form is already submited for verification ,unable to proceed',
                'data':{}
            }
            return Response(response_,status=200)


    data['userId'] = id

    secondaryserializer = UserSecondarySerializer(secondaryinfo_obj,data=data,partial=True)
    if secondaryserializer.is_valid():
        secondaryserializer.save() 
        useroberj = Users.objects.filter(id=id,is_active=True).first()
        company_code = useroberj.company_code
        useremail = useroberj.email
        username = useroberj.Firstname + " "+ useroberj.Lastname

        adminobj = Users.objects.filter(email=adminemail).first()
        adminid = adminobj.id
        Name = useroberj.Firstname + " "+useroberj.Lastname
        data1={}
        data1['NotificationTitle'] = "Verification"
        data1['NotificationMsg'] ="<span class='notfempname'>"+ Name+"</span> submitted documents for verification.<a href='/document-verification/"+encrypt_parameter(id)+"'>View Documents </a>"
        data1['company_code'] = company_code
        data1['NotificationTypeId'] = 4
        data1['created_by'] = id
        data1['UserID'] = adminid
        data1['leaveID']=0
        # data1['action_Taken']=False
        serializer = TaskNotificationSerializer(data=data1)
        if serializer.is_valid():
            serializer.save()
            userobj=Users.objects.filter(id=adminid).first()
            email = userobj.email
            curruser = Users.objects.filter(email=email).annotate(
                Name=Concat('Firstname', Value(' '), 'Lastname'))
            Name = curruser.first().Name.title()
            try:             
                dicti = {'Name': Name, 'employeeName':username,'email': email,'hosturl':frontUrl,'id':id}

                
                message = get_template(
                    'adminmail.html').render(dicti)
                msg = EmailMessage(
                    'Documents Submited',
                    message, 
                    EMAIL_HOST_USER,
                    [email],
                )
                msg.content_subtype = "html"  # Main content is now text/html
                msg.send()
            except Exception as e:
                print('exception WHEN VERIFICATIONl', e)
                    


        

            


            response_={
                'status':1,
                'Msg':'Data Saved Successfully.',
                'data':secondaryserializer.data
            }
            return Response(response_,status=200)
        response_={
                'status':0,
                'Msg':'Data Not Saved.',
                'data':secondaryserializer.errors
            }
        return Response(response_,status=200)
       
@api_view(['POST'])
def acceptDocument(request):
    details=[]
    data1={}



    data1['NotificationTitle'] = "Verification Accepted"
    data1['NotificationMsg'] ="Your Documents Approved Successfully."
    data1['company_code'] = request.user.company_code
    data1['NotificationTypeId'] = 4
    data1['created_by'] = request.user.id
    data1['leaveID'] = 0
    data1['UserID'] = request.POST.get('UserId')
    serializer = TaskNotificationSerializer(data=data1)
    if serializer.is_valid():
        serializer.save()
        

        verifiedobj = Users.objects.filter(id=data1['UserID'],documentverified=False).first()
        if verifiedobj is not None:
            verifiedobj.documentverified = True
            verifiedobj.secondary_info = True
            verifiedobj.reason_of_rejection_documents=""
            verifiedobj.save()
            userobj=Users.objects.filter(id=data1['UserID']).first()
            userserializer= UserSerializer(userobj)
            email = userobj.email
            personal_email = userobj.personal_email
            password = userobj.Password
            curruser = Users.objects.filter(id=data1['UserID']).annotate(
                Name=Concat('Firstname', Value(' '), 'Lastname'))
            Name = curruser.first().Name
            # 
            designation = userserializer.data['DesignationId']
            # desigobj = Designation.objects.filter(id=designationId,isactive=True).first()
            # designation = desigobj.DesignationName
            compobj = companyinfo.objects.filter(companycode=userobj.company_code,isactive=True).first()
            companyName = compobj.companyName
            d1=datetime.datetime.strptime(userserializer.data['DateofJoining'], "%Y-%m-%d")
            
            DateofJoining = d1.strftime('%d %B %Y')
            dicti = {'password': password,
                        'Name': Name, 'email': email,'designation':designation,'companyName':companyName,'DateofJoining':DateofJoining, 'personal_email': personal_email,'id':id}
          
            message = get_template(
                'acceptdocument.html').render(dicti)
            
            msg = EmailMessage(  
                'Congratulations! Your Documents Have Been Successfully Verified - Onboarding Next Steps',
                message,
                EMAIL_HOST_USER,
                [personal_email],
            )
            msg.content_subtype = "html"  # Main content is now text/html
            msg.send()
            notifyupdate = TaskNotification.objects.filter(created_by= data1['UserID'],NotificationTitle="Verification").update(action_Taken=True)
           
            response_={
                'status':1,
                'Msg':'Documents Accepted Successfully.',
                'data':serializer.data
            }
            return Response(response_,status=200)
        else:
            notifyupdate = TaskNotification.objects.filter(created_by= data1['UserID'],NotificationTitle="Verification").update(action_Taken=True)
            
            response_={
                'status':1,
                'Msg':'Documents Accepted Successfully.',
                'data':serializer.data
            }
            return Response(response_,status=200)
        
            
    else:
        response_={
            'status':0,
            'Msg':'Data Not Saved.',
            'data':serializer.errors
        }
        return Response(response_,status=200)
       
@api_view(['POST'])
def rejectedDocument(request): 

    data1={}
    data1['NotificationTitle'] = "Verification"
    resone = request.POST.get('resone')
    if resone !="" and resone is not None:
        data1['NotificationMsg'] = resone + "."
        data1['company_code'] = request.user.company_code
        data1['NotificationTypeId'] = 4
        data1['created_by'] = request.user.id
        data1['UserID'] = request.POST.get('UserId')
        data1['leaveID'] = 0
        serializer = TaskNotificationSerializer(data=data1)
        if serializer.is_valid():
            serializer.save()
            verifiedobj = Users.objects.filter(id=data1['UserID'],documentverified=False).first()
            if verifiedobj is not None:
                verifiedobj.linkdatetime = timezone.now()
                verifiedobj.reason_of_rejection_documents=resone
                verifiedobj.save()
                finalsubmitobj = UserSecondaryInfo.objects.filter(userId=data1['UserID']).update(finalsubmit=False)
                userobj=Users.objects.filter(id=data1['UserID']).first()

                email = userobj.email
                personal_email = userobj.personal_email
                password = userobj.Password
                curruser = Users.objects.filter(id=data1['UserID']).annotate(
                    Name=Concat('Firstname', Value(' '), 'Lastname'))
                Name = curruser.first().Name
                # 
                compobj = companyinfo.objects.filter(companycode=userobj.company_code,isactive=True).first()
                companyName = compobj.companyName

                # 
                try:             

                    dicti = {'password': password,
                                'Name': Name,
                                'hosturl':frontUrl,
                                'email': email,'companyName':companyName,'id':data1['UserID'],'description':data1['NotificationMsg']}
                
                    message = get_template(
                        'userrejectverification.html').render(dicti)
                    
                    msg = EmailMessage(
                        'Onboarding Process Update - Document Verification Delay',
                        message,
                        EMAIL_HOST_USER,
                        [personal_email],
                    )
                    msg.content_subtype = "html"  # Main content is now text/html
                    msg.send()

                except Exception as e:
                    print('exception rejection mail', e)
                    


                




                notifyupdate = TaskNotification.objects.filter(created_by= data1['UserID'],NotificationTitle="Verification").update(action_Taken=True)
                response_={
                    'status':1,
                    'Msg':'Documents Rejected Successfully.',
                    'data':serializer.data
                }
                return Response(response_,status=200)
            notifyupdate = TaskNotification.objects.filter(created_by= data1['UserID'],NotificationTitle="Verification").update(action_Taken=True)
            response_={
                'status':1,
                'Msg':'Documents Rejected Successfully.',
                'data':serializer.data
            }
            return Response(response_,status=200)
        else:
            response_={
                'status':0,
                'Msg':'Data Not Saved.',
                'data':serializer.errors
            }
            return Response(response_,status=200)
        
    else:
            response_={
                'status':0,
                'Msg':'Valid reason required.',
                'data':[]
            }
            return Response(response_,status=200)

@api_view(['POST'])
def admin_attoverview_api(request):
    if request.method == 'POST':
        weekfilter = request.data.get('weekfilter')
        com_code = request.user.company_code

        my_date = datetime.date.today()
        year, week_num, day_of_week = my_date.isocalendar()
        year = year
        userlist=[]

        if weekfilter is not None and weekfilter != "":
            currentweek = weekfilter
        else:
            currentweek = week_num

        year = year
        week = int(currentweek)

        startdate = datetime.date.fromisocalendar(year, week, 1)
        
        dates = []
        for i in range(7):
            day = startdate + datetime.timedelta(days=i)
            dates.append(day)

        userobjs = Users.objects.filter(is_active=True,company_code=com_code).exclude(Q(employeeId__isnull=True)|Q(employeeId='None')|Q(employeeId=''))
        userser = UsersSerializeronlyattendance(userobjs,many=True)
        userlist=list(userser.data)


        useridobjs = Users.objects.filter(is_active=True,company_code=com_code)
        useridser = UsersSerializeronlyid(useridobjs,many=True)
        useridlist=list(useridser.data)

        leavelist =[]
        datelist = []
        latecomerlist =[]
        presentlist=[]
        for d in dates:
            attendance_records = attendance.objects.filter(employeeId__in=userlist,date=d,time__lt=time(10, 0, 0)).distinct('employeeId').count()
            presentlist.append(attendance_records)

            latecomers = attendance.objects.filter(employeeId__in=userlist,date=d,time__gt=time(10, 0, 0),time__lte=time(11, 0, 0)).distinct('employeeId').count()
            latecomerlist.append(latecomers)

            dates = (d).strftime('%b')+ " " + str(d).split("-")[2]
            datelist.append(dates)
       
            leave_count = Leave.objects.filter(employeeId__in=useridlist,leave_status="Approved",start_date__lte=d,end_date__gte=d,Active=True)
            leave_countser = leaveserializer(leave_count,many=True)
            leavelist.append(leave_count.count())


        context = {
            "datelist":datelist,
            "presentlist":presentlist,
            "latecomerlist":latecomerlist,
            "leavelist":leavelist
        }
        return Response({"data":context,"response":{"n" : 1,"msg" : "Success","status" : "success"}})

@api_view(['POST'])
def admin_employee_weekatt_api(request):
    if request.method == 'POST':
        weekfilter = request.POST.get("week")
        employeeuser = request.POST.get("employeeid")
   
        com_code = request.user.company_code
        loggeduser = request.user.id    

        my_date = datetime.date.today()
        year, week_num, day_of_week = my_date.isocalendar()
        curryear = year
        if weekfilter is not None and weekfilter != "":
      
            currentweek = weekfilter
        else:
            currentweek = week_num
      
        userlist=[]

        year = year
        week = int(currentweek)
        
        startdate = datetime.date.fromisocalendar(year, week, 1)
        
        dates = []
        for i in range(7):
            day = startdate + datetime.timedelta(days=i)
            dates.append(day)
        user_emplyeeid = ""
        userobjs = Users.objects.filter(is_active=True,id=employeeuser,company_code=com_code)
        userser = UserSerializer(userobjs,many=True)
        for p in userser.data:
            if p['employeeId'] is not None:
                user_emplyeeid = p['employeeId']

      

       
        presentlist=[]
        for d in dates:
            userpresent=0
            presentobj = attendance.objects.filter(employeeId=user_emplyeeid,date=d)
            if presentobj is not None:
                presentser = attendanceserializer(presentobj,many=True)
                prsnt = presentobj.time
                timestr = str(prsnt)
                timehours = timestr.split(":")[0]
                timemints = timestr.split(":")[1]
                timesecs = timestr.split(":")[2]

                if int(timehours) == 10 and int(timemints) == 0:
                    userpresent += 1
                elif int(timehours) < 10 :
                    userpresent += 1
                else:
                    userpresent += 0
            else:
                userpresent += 0

                    
       
        latecomerlist =[]
        for d in dates:
            userlatecome=0
            for u in userlist:
                presentobj = attendance.objects.filter(employeeId=u,date=d).order_by('time').first()
                if presentobj is not None:
                    prsnt = presentobj.time
                  
                    timestr = str(prsnt)
                    timehours = timestr.split(":")[0]
                    timemints = timestr.split(":")[1]
                    timesecs = timestr.split(":")[2]

                    if int(timehours) == 10 and int(timemints) > 0:
                        userlatecome += 1
                    elif int(timehours) > 10 :
                        userlatecome += 1
                    else:
                        userlatecome += 0
                else:
                    userlatecome += 0
            latecomerlist.append(userlatecome)

                      
        leavelist =[]
        for d in dates:
          
            userleave = 0
            for u in userlist:
                userobjs = Users.objects.filter(employeeId=u).first()
                userid = userobjs.id
            

                leaveobj = Leave.objects.filter(employeeId=userid,leave_status="Approved")
                if leaveobj is not None:
                    leaveser = leaveserializer(leaveobj,many=True)
                    for l in leaveser.data:
                        leave_start_date = datetime.datetime.strptime(l['start_date'], '%Y-%m-%d')
                        leave_end_date = datetime.datetime.strptime(l['end_date'], '%Y-%m-%d')
                        week_data = datetime.datetime.strptime(str(d), '%Y-%m-%d')
                        if l['start_date'] == l['end_date']:
                            if str(d) == l['start_date']:
                             
                                userleave +=1
                            else:
                                userleave +=0
                        else:
                            result = leave_start_date <= week_data <= leave_end_date
                            if result == True:
                           
                                userleave +=1
                            else:
                                userleave +=0
            leavelist.append(userleave)
        context = {
            "presentlist":presentlist,
            "latecomerlist":latecomerlist,
            "leavelist":leavelist
        }
        return Response({"data":context,"response":{"n" : 1,"msg" : "Success","status" : "success"}})
          
def all_dates_current_month(year,month):
    number_of_days = calendar.monthrange(year, month)[1]
    first_date = date(year, month, 1)
    last_date = date(year, month, number_of_days)
    delta = last_date - first_date

    return [(first_date + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(delta.days + 1)]


@api_view(['POST'])
def sch_attendancelistapi(request):
    com_code = request.user.company_code
    my_date = datetime.date.today()
    currmonth = date.today().month
    year, week_num, day_of_week = my_date.isocalendar()
    curryear = year
    currentweek = week_num
    userlist=[]
    startdate = datetime.date.fromisocalendar(curryear, currentweek, 1)
    dates = []
    for i in range(7):
        day = startdate + datetime.timedelta(days=i)
        dates.append(day)
    datelist = dates
    
    userobjs = Users.objects.filter(is_active=True,company_code=com_code)
    userser = UsersSerializeronlyid(userobjs,many=True)
    userlist=list(userser.data)


  

    for u in userlist:
       
        ontimecount  = 0
        latecount = 0
        onleavecount = 0

        
        userobj = Users.objects.filter(id=int(u),is_active=True,company_code=com_code).first()
        if userobj is not None:
            dictFirstname= userobj.Firstname 
            dictLastname= userobj.Lastname
            userser = UserSerializerDesignationId(userobj)
            # for l in [userser.data]:
            ds = userser.data["DesignationId"]

            designation_obj=Designation.objects.filter(id=ds,isactive=True,company_code=com_code).first()
            if designation_obj is not None:
                dictDesignation = designation_obj
            else:
                dictDesignation = None

            dictempcode = userobj.uid
            attendanceid = userobj.employeeId

            if attendanceid is not None and  attendanceid != "":
                for d in datelist :
                    presentobj = attendance.objects.filter(employeeId=str(attendanceid),date=d,time__lte=time(10, 0, 0),company_code=com_code).order_by('time').first()
                    if presentobj is not None:
                        ontimecount += 1
                    lateobj = attendance.objects.filter(employeeId=str(attendanceid),date=d,time__gt=time(10, 0, 0),time__lte=time(11, 0, 0),company_code=com_code).order_by('time').first()
                    if lateobj is not None:
                        latecount += 1

                



                for d in datelist:
                    leaveobj = Leave.objects.filter(employeeId=str(u),leave_status="Approved")
                    if leaveobj is not None:
                        leaveser = leaveserializer(leaveobj,many=True)
                        for l in leaveser.data:
                            leave_start_date = datetime.datetime.strptime(l['start_date'], '%Y-%m-%d')
                            leave_end_date = datetime.datetime.strptime(l['end_date'], '%Y-%m-%d')
                            check_date = datetime.datetime.strptime(str(d), '%Y-%m-%d')
                            if l['start_date'] == l['end_date']:
                                if str(d) == l['start_date']:
                                    onleavecount +=1
                                else:
                                    onleavecount +=0
                            else:
                                result = leave_start_date <= check_date <= leave_end_date
                                if result == True:
                                    onleavecount +=1
                                else:
                                    onleavecount +=0
                    else:
                        onleavecount +=0
            
            

                attobj = AttendanceList.objects.filter(userId=u,Week=currentweek,Month=currmonth).first()
                if attobj is not None :
                    AttendanceList.objects.filter(userId=u,Week=currentweek,Month=currmonth).update(Latecount=latecount,Leavecount=onleavecount,Ontimecount=ontimecount)
                else:
                    AttendanceList.objects.create(userId=u,Week=currentweek,Month=currmonth,Latecount=latecount,Leavecount=onleavecount,Ontimecount=ontimecount,Firstname=dictFirstname,AttendanceId=attendanceid,Lastname=dictLastname,Designation=dictDesignation,company_code=com_code,Empcode=dictempcode)
            else:
                attobj = AttendanceList.objects.filter(userId=u,Week=currentweek,Month=currmonth).first()
                if attobj is not None :
                    AttendanceList.objects.filter(userId=u,Week=currentweek,Month=currmonth).update(Latecount=latecount,Leavecount=onleavecount,Ontimecount=ontimecount)
                else:
                    
                    AttendanceList.objects.create(userId=u,Week=currentweek,Month=currmonth,Latecount=latecount,Leavecount=onleavecount,Ontimecount=ontimecount,Firstname=dictFirstname,Lastname=dictLastname,Designation=dictDesignation,company_code=com_code,Empcode=dictempcode)

    return Response({"data":'',"response":{"n" : 1,"msg" : "Success","status" : "success"}})

@api_view(['POST'])
def admin_getattlist(request):
    if request.method == 'POST':
        listdatafilter = request.data.get('listdatafilter')
        com_code = request.user.company_code

        userdatalist = []
        userlist=[]
        now = datetime.datetime.now()

        # Get the current year
        current_year = now.year
        # Get the current month
        current_month = now.month
        # Get the current week number using isocalendar()
        current_week_number = now.isocalendar()[1]
            



        userobjs = Users.objects.filter(is_active=True,company_code=com_code).order_by("Firstname")
        userser = UsersSerializeronlyid(userobjs,many=True)
        userlist=list(userser.data)

        

        if listdatafilter == "Monthly":
            for u in userlist:
                indvdict = {}
                indvdict['user']=u
                
                userobj = Users.objects.filter(id=int(u),company_code=request.user.company_code,is_active=True).first()
                if userobj is not None:
                    
                    indvdict['empname'] = userobj.Firstname + " " + userobj.Lastname
                    empid = userobj.uid
                    if empid != "" and empid is not None:
                        employeeid = empid
                    elif empid == "" or empid is None:
                        employeeid = "--"
                    else:
                        employeeid = "--"

                    userser = UserSerializer(userobj)
                    if userser.data['DepartmentID'] != []:
                        for ud in userser.data['DepartmentID']:
                            dept = ud
                            deptobj = Department.objects.filter(id=dept).first()
                            empdept = deptobj.DepartmentName
                    else:
                        empdept = "--"
                        
                    AttendanceId=userser.data['employeeId']
                    indvdict['departmentid'] = empdept
                    indvdict['Designation'] =  userser.data["DesignationId"]
                    indvdict['empcode']=employeeid
                    
                    AttendanceId=userser.data['employeeId']
                    
                    if str(userobj.Photo) is not None and str(userobj.Photo) != "":
                        indvdict['userimage'] = imageUrl +"/media/"+ str(userobj.Photo)
                    else:
                        indvdict['userimage'] = imageUrl + "/static/assets/images/profile.png"

                    ontime_obj = attendance.objects.filter(employeeId=AttendanceId,date__year=current_year,date__month=current_month,time__lte=time(10, 0, 0),company_code=com_code).distinct('date')
                    
                    dates_serial=AttendanceSerializerAttendanceWeekDate(ontime_obj,many=True)
                    exclude_dates_list=list(dates_serial.data)
                    
                    ontime_obj = ontime_obj.count()
                    
                    late_obj = attendance.objects.filter(employeeId=AttendanceId,date__year=current_year,date__month=current_month,time__gt=time(10, 0, 0),time__lte=time(11, 0, 0),company_code=com_code).distinct('date').exclude(date__in=exclude_dates_list).count()
                    
                    leave_obj  = Leave.objects.filter(employeeId=u,WorkFromHome=False,start_date__month=current_month,start_date__year= current_year,Active=True,leave_status="Approved").exclude(leave_status='Draft').aggregate(total_days=Coalesce(Sum('number_of_days', output_field=DecimalField()), Value(0, output_field=DecimalField())))



                    indvdict['ontimecount']=ontime_obj
                    indvdict['Latecount']=late_obj
                    indvdict['Leavecount']=leave_obj['total_days']

                    userdatalist.append(indvdict)

        if listdatafilter == "Weekly":

            
            for u in userlist:
                indvdict = {}
                indvdict['user']=u
                userobj = Users.objects.filter(id=int(u),company_code=request.user.company_code,is_active=True).first()
                if userobj is not None:
                    indvdict['empname'] = userobj.Firstname + " " + userobj.Lastname
                    empid = userobj.uid
                    if empid != "" and empid is not None:
                        employeeid = empid
                    elif empid == "" or empid is None:
                        employeeid = "--"
                    else:
                        employeeid = "--"

                    userser = UserSerializer(userobj)
                    if userser.data['DepartmentID'] != []:
                        for ud in userser.data['DepartmentID']:
                            dept = ud
                            deptobj = Department.objects.filter(id=dept).first()
                            empdept = deptobj.DepartmentName
                    else:
                        empdept = "--"
                        
                    AttendanceId=userser.data['employeeId']
                    indvdict['departmentid'] = empdept
                    indvdict['Designation'] =  userser.data["DesignationId"]
                    indvdict['empcode']=employeeid
                    
                    if str(userobj.Photo) is not None and str(userobj.Photo) != "":
                        indvdict['userimage'] = imageUrl +"/media/"+ str(userobj.Photo)
                    else:
                        indvdict['userimage'] = imageUrl + "/static/assets/images/profile.png"
                    

                    ontime_obj = attendance.objects.filter(employeeId=AttendanceId,date__year=current_year,date__month=current_month,date__week=current_week_number,time__lte=time(10, 0, 0),company_code=com_code).distinct('date')

                    dates_serial=AttendanceSerializerAttendanceWeekDate(ontime_obj,many=True)
                    exclude_dates_list=list(dates_serial.data)
                    
                    ontime_obj = ontime_obj.count()

                    late_obj = attendance.objects.filter(employeeId=AttendanceId,date__year=current_year,date__month=current_month,date__week=current_week_number,time__gt=time(10, 0, 0),time__lte=time(11, 0, 0),company_code=com_code).distinct('date').exclude(date__in=exclude_dates_list).count()
                    
                    leave_obj  = Leave.objects.filter(employeeId=u,WorkFromHome=False,start_date__month=current_month,start_date__week= current_week_number,start_date__year= current_year,Active=True,leave_status="Approved").exclude(leave_status='Draft').aggregate(total_days=Coalesce(Sum('number_of_days', output_field=DecimalField()), Value(0, output_field=DecimalField())))
              
                        
                    indvdict['ontimecount']=ontime_obj
                    indvdict['Latecount']=late_obj
                    indvdict['Leavecount']=leave_obj['total_days']
                    userdatalist.append(indvdict)

        if os.path.exists("static/report/Attendancereport.xlsx"):
            os.remove("static/report/Attendancereport.xlsx")
            workbook = xlsxwriter.Workbook('static/report/Attendancereport.xlsx')
            workbook.close()
        else:
            workbook = xlsxwriter.Workbook('static/report/Attendancereport.xlsx')
            workbook.close()
        excelreport(userdatalist)
        downloadurl = imageUrl + "/static/report/Attendancereport.xlsx"

        return Response({"data":userdatalist,"response":{"n" : 1,"msg" : "Success","status" : "success"},"downloadurl":downloadurl})


















@api_view(['POST'])
def admin_attlist(request):
    if request.method == 'POST':
        listdatafilter = request.data.get('listdatafilter')
        
        com_code = request.user.company_code
        my_date = datetime.date.today()
        currmonth = date.today().month
        year, week_num, day_of_week = my_date.isocalendar()
        curryear = year
        currentweek = week_num
        userlist=[]
        userdatalist = []
        userobjs = Users.objects.filter(is_active=True,company_code=com_code)
        user_id_ser = UsersSerializeronlyid(userobjs,many=True)
        userlist=list(user_id_ser.data)
       

        if listdatafilter == "Monthly":
            datelist = all_dates_current_month(curryear,currmonth)
        if listdatafilter == "Weekly":
            startdate = datetime.date.fromisocalendar(curryear, currentweek, 1)
            dates = []
            for i in range(7):
                day = startdate + datetime.timedelta(days=i)
                dates.append(day)
            datelist = dates
            
        if listdatafilter == "Today":
            datelist = []
            datelist.append(my_date)

        for u in userlist:
            userobj = Users.objects.filter(id=int(u),is_active=True).first()
            if userobj is not None:
                userser = UserSerializer(userobj)
                indvuserdict={}
                indvuserdict['user']= u
                indvuserdict['username']= userser.data['Firstname'] + " " + userser.data['Lastname']
                if userser.data['DepartmentID'] != []:
                    for dept in userser.data['DepartmentID']:
                        deptobj = Department.objects.filter(id=dept).first()
                        empdept = deptobj.DepartmentName
                else:
                    empdept = "--"
                indvuserdict['departmentid'] = empdept
                indvuserdict['Designation'] = userser.data["DesignationId"]
                indvuserdict['employeeid'] = userser.data['uid']
                if str(userser.data['Photo']) is not None and str(userser.data['Photo']) != "":
                    indvuserdict['userimage'] = imageUrl +"/media/"+ str(userser.data['Photo'])
                else:
                    indvuserdict['userimage'] = imageUrl + "/static/assets/images/profile.png"

                
                ontimecount  = 0
                latecount = 0
                onleavecount = 0
                attendance_id=userser.data['employeeId']
                if attendance_id is not None and  attendance_id != "":

                    for d in datelist :
                        presentobj = attendance.objects.filter(employeeId=str(attendance_id),date=d,company_code=com_code).order_by('time').first()
                        if presentobj is not None:
                            prsnt = presentobj.time
                            timestr = str(prsnt)
                            timehours = timestr.split(":")[0]
                            timemints = timestr.split(":")[1]

                            if int(timehours) == 10 and int(timemints) == 0:
                                ontimecount += 1
                            elif int(timehours) < 10 :
                                ontimecount += 1
                            else:
                                ontimecount += 0
                        else:
                            ontimecount += 0 
                    indvuserdict['ontimecount'] = ontimecount


                    for d in datelist :
                    
                        lateobj = attendance.objects.filter(employeeId=str(attendance_id),date=d,company_code=com_code).order_by('time').first()
                    
                        if lateobj is not None:
                            latestr = lateobj.time
                            timestr = str(latestr)
                            timehours = timestr.split(":")[0]
                            timemints = timestr.split(":")[1]

                            if int(timehours) == 10 and int(timemints) > 0:
                            
                                latecount += 1
                            elif int(timehours) > 10 :
                            
                                latecount += 1
                            else:
                            
                                latecount += 0
                        else:
                            latecount += 0

                    indvuserdict['latecount'] = latecount


                    for d in datelist:
                        leaveobj = Leave.objects.filter(employeeId=str(u),leave_status="Approved")
                        if leaveobj is not None:
                            leaveser = leaveserializer(leaveobj,many=True)
                            for l in leaveser.data:
                                leave_start_date = datetime.datetime.strptime(l['start_date'], '%Y-%m-%d')
                                leave_end_date = datetime.datetime.strptime(l['end_date'], '%Y-%m-%d')
                                check_date = datetime.datetime.strptime(str(d), '%Y-%m-%d')
                                if l['start_date'] == l['end_date']:
                                    if str(d) == l['start_date']:
                                        onleavecount +=1
                                    else:
                                        onleavecount +=0
                                else:
                                    result = leave_start_date <= check_date <= leave_end_date
                                    if result == True:
                                        onleavecount +=1
                                    else:
                                        onleavecount +=0
                        else:
                            onleavecount +=0
                    indvuserdict['onleavecount'] = onleavecount
                else:
                    indvuserdict['ontimecount'] = 0
                    indvuserdict['onleavecount'] = 0
                    indvuserdict['latecount'] = 0

                userdatalist.append(indvuserdict)
        
        return Response ({"data":userdatalist,"response":{"n" : 1,"msg" : "Success","status" : "success",}})
    else:
        return Response ({"data":'',"response":{"n" : 0,"msg" : "Failed","status" : "Failed"}})
    


def excelreport(data):
   
    workbook = xlsxwriter.Workbook('static/report/Attendancereport.xlsx')
    wb=load_workbook('static/report/Attendancereport.xlsx')
    sheet=wb.worksheets[0]

    sheet.cell(row=1,column=4).value="Attendance List Report"
    
    sheet.cell(row=2,column=1).value="Sr No"
    sheet.cell(row=2,column=2).value="Employee Code"
    sheet.cell(row=2,column=3).value="Employee Name"
    sheet.cell(row=2,column=4).value='Designation'
    sheet.cell(row=2,column=5).value='Department'
    sheet.cell(row=2,column=6).value='On Time'
    sheet.cell(row=2,column=7).value='Late'
    sheet.cell(row=2,column=8).value='Leave'

    k = 3
    counter = 1
    for i in data:
        sheet.cell(row=k,column=1).value=counter
        sheet.cell(row=k,column=2).value=i['empcode']
        sheet.cell(row=k,column=3).value=i['empname']
        sheet.cell(row=k,column=4).value=i['Designation']
        sheet.cell(row=k,column=5).value=i['departmentid']
        sheet.cell(row=k,column=6).value=i['ontimecount']
        sheet.cell(row=k,column=7).value=i['Leavecount']
        sheet.cell(row=k,column=8).value=i['Latecount']
        k+=1
        counter+=1
    
    wb.save('static/report/Attendancereport.xlsx')



from rest_framework import pagination

class CustomPagination(pagination.PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 50
    page_query_param = 'p'

    def get_paginated_response(self,data):
        response = Response({
            'status':"success",
            'count':self.page.paginator.count,
            'next' : self.get_next_link(),
            'previous' : self.get_previous_link(),
            'data':data,
        })
        return response

class attendance_data(GenericAPIView):
    
    pagination_class = CustomPagination

    def post(self, request):
        listdatafilter = request.data.get('listdatafilter')
        com_code = request.user.company_code
        loggeduser = request.user.id    

        my_date = datetime.date.today()
        currmonth = date.today().month
        year, week_num, day_of_week = my_date.isocalendar()
        curryear = year
        currentweek = week_num
        userlist=[]
        userdatalist = []
        Firstname = request.data.get("Firstname")
       
        pagenumber = request.data.get('p')
       
        userobjs = Users.objects.filter(is_active=True,Firstname__icontains=Firstname,company_code=com_code).order_by("Firstname")
        page4 = self.paginate_queryset(userobjs)
      
        alluserser = UserSerializer(page4,many=True)
        # for p in userser.data:
        #     userlist.append(p['id'])

        if listdatafilter == "Monthly":
            datelist = all_dates_current_month(curryear,currmonth)
        if listdatafilter == "Weekly":
            startdate = datetime.date.fromisocalendar(curryear, currentweek, 1)
            dates = []
            for i in range(7):
                day = startdate + datetime.timedelta(days=i)
                dates.append(day)
            datelist = dates
        if listdatafilter == "Today":
            datelist = []
            datelist.append(my_date)
        
        
        for u in alluserser.data:
            indvuserdict={}
            ontimecount  = 0
            latecount = 0
            onleavecount = 0
            indvuserdict['user']= u['id']
            userobj = Users.objects.filter(id=int(u['id'])).first()
            indvuserdict['username']= userobj.Firstname + " " + userobj.Lastname

            userser = UserSerializer(userobj)
        
            for l in [userser.data]:
             
                if l['DepartmentID'] != []:
                    for ud in l['DepartmentID']:
                        dept = ud
                        deptobj = Department.objects.filter(id=dept).first()
                        empdept = deptobj.DepartmentName
                else:
                    empdept = "--"

            indvuserdict['departmentid'] = empdept
            indvuserdict['Designation'] = l["DesignationId"]

            empid = userobj.uid

            if empid != "" and empid is not None:
                employeeid = empid
            elif empid == "" or empid is None:
                employeeid = "--"
            else:
                employeeid = "--"

            indvuserdict['employeeid'] = employeeid

            if str(userobj.Photo) is not None and str(userobj.Photo) != "":
                indvuserdict['userimage'] = imageUrl +"/media/"+ str(userobj.Photo)
            else:
                indvuserdict['userimage'] = imageUrl + "/static/assets/images/profile.png"

            if empid is not None and  empid != "":
               
                for d in datelist :
                    presentobj = attendance.objects.filter(employeeId=str(employeeid),date=d,company_code=com_code).order_by('time').first()
                    if presentobj is not None:
                        prsnt = presentobj.time
                        timestr = str(prsnt)
                        timehours = timestr.split(":")[0]
                        timemints = timestr.split(":")[1]

                        if int(timehours) == 10 and int(timemints) == 0:
                            ontimecount += 1
                        elif int(timehours) < 10 :
                            ontimecount += 1
                        else:
                            ontimecount += 0
                    else:
                        ontimecount += 0 
                indvuserdict['ontimecount'] = ontimecount


                for d in datelist :
                   
                    lateobj = attendance.objects.filter(employeeId=str(employeeid),date=d,company_code=com_code).order_by('time').first()
                   
                    if lateobj is not None:
                        latestr = lateobj.time
                        timestr = str(latestr)
                        timehours = timestr.split(":")[0]
                        timemints = timestr.split(":")[1]

                        if int(timehours) == 10 and int(timemints) > 0:
                          
                            latecount += 1
                        elif int(timehours) > 10 :
                           
                            latecount += 1
                        else:
                          
                            latecount += 0
                    else:
                        latecount += 0

                indvuserdict['latecount'] = latecount


                for d in datelist:
                    leaveobj = Leave.objects.filter(employeeId=str(u['id']),leave_status="Approved")
                    if leaveobj is not None:
                        leaveser = leaveserializer(leaveobj,many=True)
                        for l in leaveser.data:
                            leave_start_date = datetime.datetime.strptime(l['start_date'], '%Y-%m-%d')
                            leave_end_date = datetime.datetime.strptime(l['end_date'], '%Y-%m-%d')
                            check_date = datetime.datetime.strptime(str(d), '%Y-%m-%d')
                            if l['start_date'] == l['end_date']:
                                if str(d) == l['start_date']:
                                    onleavecount +=1
                                else:
                                    onleavecount +=0
                            else:
                                result = leave_start_date <= check_date <= leave_end_date
                                if result == True:
                                    onleavecount +=1
                                else:
                                    onleavecount +=0
                    else:
                        onleavecount +=0
                indvuserdict['onleavecount'] = onleavecount
            else:
                indvuserdict['ontimecount'] = 0
                indvuserdict['onleavecount'] = 0
                indvuserdict['latecount'] = 0

            userdatalist.append(indvuserdict)

        
        
        return self.get_paginated_response(userdatalist)
    

def EmployeeCode(company_code):
    
    firstemplyeestrt = company_code + "/"
    object = Users.objects.filter(company_code = company_code).exclude(uid__isnull=True).last()
    emp = 0
    if object is None:
        emp = 0
        firstSale = firstemplyeestrt +"1001"
        return firstSale
    else:
        emp = str(object.uid)
        place_emp_str =int(str(emp[-4:])) + 1
        newempstr = "%04d" % (place_emp_str)
        newempId = firstemplyeestrt+str(newempstr)
        return newempId

@api_view(['POST'])
def admin_attmodaldata(request):
    if request.method == 'POST':
        weekdatafilter = int(request.data.get('weekdatafilter'))
        com_code = request.user.company_code
        attuserid = request.data.get('useridfilter')
        holidaylist = []
        listWO=[]
        weekdatalist = []
        userdatalist=[]
        userdict={}

        userobj = Users.objects.filter(id=int(attuserid),company_code=com_code).first()
        if userobj is not None:
                
            userser = UserSerializer(userobj)
            userdict['username']= userser.data['Firstname'] + " " + userser.data['Lastname']
            if userser.data['employeeId'] is not None:
                userdict['useremployeeId'] = userser.data['employeeId']
            else:
                userdict['useremployeeId'] = "--"
            if  userser.data['DesignationId'] is not None:
                userdict['userdesignation'] = userser.data['DesignationId']
            else:
                userdict['userdesignation']="--"
                


            if userser.data['Photo'] is not None and userser.data['Photo'] != "":
                userdict['userimage'] = imageUrl + str(userser.data['Photo'])
            else:
                userdict['userimage'] = imageUrl + "/static/assets/images/profile.png"
                    
                    
            userdatalist=userdict




            holidatlist = Holidays.objects.filter(Active=True).order_by('id')
            serializer = holidaysSerializer(holidatlist, many=True)
            
            for i in serializer.data:
                holiday=i['Date']
                convertholiday = datetime.datetime.strptime(holiday,"%Y-%m-%d").date()
                holidaylist.append(convertholiday)
        

            
            my_date = datetime.date.today()
            year, week_num, day_of_week = my_date.isocalendar()
            curryear = year
            year = curryear

            startdate = datetime.date.fromisocalendar(curryear, weekdatafilter, 1)
            weekdates = []
            for r in range(7):
                day = startdate + datetime.timedelta(days=r)
                weekdates.append(day)
        
            
        
            for month in range(1,13):
                dt=date(curryear,month,1)    # first day of month
                first_w=dt.isoweekday()  # weekday of 1st day of the month
                if(first_w==7): # if it is Sunday return 0 
                    first_w=0
                saturday2=14-first_w
                dt1=date(curryear,month,saturday2)
                listWO.append(dt1)
                saturday4=28-first_w  # 4th Saturday 
                dt2=date(curryear,month,saturday4)
                listWO.append(dt2)
        

            def allsundays(year):
                d = date(year, 1, 1)                 
                d += timedelta(days = 6 - d.weekday())  
                while d.year == year:
                    yield d
                    d += timedelta(days = 7)

        

            for d in allsundays(curryear):
                listWO.append(d)

        
        


            for dt in weekdates:
            
                datemonth = datetime.datetime.strptime(str(dt), "%Y-%m-%d")
                monthd=datemonth.month
                monthname=calendar.month_name[monthd]
            
                daystr = str(dt).split("-")[2]
                datestr = str(daystr)+" " +str(monthname)
            

                daydict={}

                daydict['weekdate'] = datestr

                onleavecount=0
                leaveobj = Leave.objects.filter(employeeId=str(attuserid),leave_status="Approved")
                if leaveobj is not None:
                    leaveser = leaveserializer(leaveobj,many=True)
                    for l in leaveser.data:
                        leave_start_date = datetime.datetime.strptime(l['start_date'], '%Y-%m-%d')
                    
                        leave_end_date = datetime.datetime.strptime(l['end_date'], '%Y-%m-%d')
                        check_date = datetime.datetime.strptime(str(dt), '%Y-%m-%d')
                        if l['start_date'] == l['end_date']:
                            if str(d) == l['start_date']:
                                onleavecount +=1
                            else:
                                onleavecount +=0
                        else:
                            result = leave_start_date <= check_date <= leave_end_date
                            if result == True:
                                onleavecount +=1
                            else:
                                onleavecount +=0

                if dt in holidaylist:
                    daydict['daystatus'] = "Holiday"
                elif onleavecount > 0 :
                    daydict['daystatus'] = "Leave"
                elif dt in listWO:
                    daydict['daystatus'] = "WeeklyOff"
                else:
                    daydict['daystatus'] = "Present"
                    empid = Users.objects.filter(id=int(attuserid)).first()
                    if empid.employeeId is not None:
                        attid = empid.employeeId

                        retuen_rdict=get_date_shift_details(str(dt),attid)
                                    
                        daydict['checkintime']=retuen_rdict['inTime']
                        daydict['checkouttime']=retuen_rdict['outTime']
                        daydict['timediff']=retuen_rdict['Total']
                        if daydict['checkintime'] is None or daydict['checkintime'] =="":
                            daydict['checkintime']="--"
                        if daydict['checkouttime'] is None or daydict['checkouttime'] =="":
                            daydict['checkouttime']="--"
                        if daydict['timediff'] is None or daydict['timediff'] =="":
                            daydict['timediff']="--"
                            
                            
                            


                        




                    else:
                        daydict['checkintime']="--"
                        daydict['checkouttime']="--"
                        daydict['timediff']="--"

                weekdatalist.append(daydict)

                context ={
                    'weekdatalist' : weekdatalist,
                    'userdatalist' : userdatalist,
                }

            

            return Response ({"data":context,"response":{"n" : 1,"msg" : "Success","status" : "success"}})
        else:
            return Response ({"data":'',"response":{"n" : 0,"msg" : "Failed","status" : "Failed"}})

    else:
        return Response ({"data":'',"response":{"n" : 0,"msg" : "User not active","status" : "Failed"}})

@api_view(['GET'])
def admintodaysstatus_api(request):
    com_code = request.user.company_code
    my_date = date.today()
    year, week_num, day_of_week = my_date.isocalendar()
    curryear = year
    userlist=[]
    lateemplist=[]
    leaveemplist = []

    userobjs = Users.objects.filter(is_active=True,company_code=com_code)
    userser = UsersSerializeronlyid(userobjs,many=True)
    userlist=list(userser.data)

    attendance_id_obj = Users.objects.filter(is_active=True,company_code=com_code).exclude(Q(employeeId__isnull=True)|Q(employeeId='None')|Q(employeeId=''))
    attendance_id_serializer = UsersSerializeronlyattendance(attendance_id_obj,many=True)
    attendance_id_list=list(attendance_id_serializer.data)

    Total_count = len(userlist)
    att_count = 0
    latecount = 0
    onleavecount = 0
    # present_count = attendance.objects.filter(employeeId__in=attendance_id_list,date=my_date,company_code=com_code).distinct('employeeId').count()
    # late_count = attendance.objects.filter(employeeId__in=attendance_id_list,date=my_date,company_code=com_code,time__gt=time(10, 0, 0),time__lte=time(11, 0, 0)).distinct('employeeId').count()
    # att_count=present_count
    # latecount=late_count
    for u in userlist:
        userinfo={}
        leaveinfo={}
        userobj = Users.objects.filter(id=int(u)).first()
        userser = UserSerializer(userobj)
            
        empid = userobj.employeeId

        if empid is not None and  empid != "":
            presentobj = attendance.objects.filter(employeeId=str(empid),date=my_date,company_code=com_code).order_by('time').first()
            if presentobj is not None:
                att_count += 1
            else:
                att_count += 0 
    


            lateobj = attendance.objects.filter(employeeId=str(empid),date=my_date,company_code=com_code).order_by('time').first()
            if lateobj is not None:
                latestr = lateobj.time
                timestr = str(latestr)
                timehours = timestr.split(":")[0]
                timemints = timestr.split(":")[1]

                if int(timehours) == 10 and int(timemints) > 0:
                    latecount += 1
                    userinfo['empname'] = userobj.Firstname +" "+ userobj.Lastname
                    userinfo['intime']= timestr
                    lateemplist.append(userinfo)
                elif int(timehours) > 10 :
                    latecount += 1
                    userinfo['empname'] = userobj.Firstname +" "+ userobj.Lastname
                    userinfo['intime']= timestr
                    lateemplist.append(userinfo)
                else:
                    latecount += 0
            else:
                latecount += 0

        

            leaveobj = Leave.objects.filter(employeeId=str(u),leave_status="Approved")
            if leaveobj is not None:
                leaveser = leaveserializer(leaveobj,many=True)
                for l in leaveser.data:
                    leave_start_date = datetime.datetime.strptime(l['start_date'], '%Y-%m-%d')
                    leave_end_date = datetime.datetime.strptime(l['end_date'], '%Y-%m-%d')
                    check_date = datetime.datetime.strptime(str(my_date), '%Y-%m-%d')
                    if leave_start_date == leave_end_date:
                        if check_date == leave_start_date:
                            leaveinfo['empname']= userobj.Firstname +" "+ userobj.Lastname
                            leaveinfo['Leavetype']=l['leavetype']
                            leaveemplist.append(leaveinfo)
                            onleavecount += 1
                        else:
                            onleavecount += 0
                    else:
                        result = leave_start_date <= check_date <= leave_end_date
                        if result == True:
                            leaveinfo['empname']= userobj.Firstname +" "+ userobj.Lastname
                            leaveinfo['Leavetype']=l['leavetype']
                            leaveemplist.append(leaveinfo)
                            onleavecount += 1
                        else:
                            onleavecount += 0
            else:
                onleavecount +=0
        
    context={
        'totalcount':Total_count,
        'att_count':att_count,
        'latecount':latecount,
        'lateemplist':lateemplist,
        'leaveemplist':leaveemplist,
        'onleavecount':onleavecount,
    }
    
    return Response ({"data":context,"response":{"n" : 1,"msg" : "Success","status" : "success"}})
    






@api_view(['GET'])
def rankcardapi(request):
    my_date = datetime.date.today()
    year, week_num, day_of_week = my_date.isocalendar()
    curryear = year
    currentweek = week_num
    prevWeek = currentweek - 1
    com_code = request.user.company_code
    loguserid = request.user.id
    rankdata = []
    rankdatadict = {}
    userprevweekdata = []
    userprevweekdict = {}


    prevweekobj = IntermediateTrackProResult.objects.filter(EmpID=loguserid,company_code=com_code,Week=prevWeek,Year=curryear).first()
    if prevweekobj is not None:
        perct = prevweekobj.TrackProPercent 
        if perct is not None:
            userprevweekdict['prevweekperc'] = str(perct) + " %"
        else:
            userprevweekdict['prevweekperc'] = "--"

        if prevweekobj.Rank is not None:
            userprevweekdict['prevweekrank'] = prevweekobj.Rank 
        else:
            userprevweekdict['prevweekrank'] = "--"
    else:
        userprevweekdict['prevweekperc'] = "--"
        userprevweekdict['prevweekrank'] = "--"

    userprevweekdata.append(userprevweekdict)

    context = {
        'userprevweekdata':userprevweekdata,
    }

    return Response ({"data":context,"response":{"n" : 1,"msg" : "Success","status" : "success"}})




API_KEY= "ort1234#"
@csrf_exempt
@api_view(['POST'])
@authentication_classes([])
@permission_classes([])
def EmployeeList(request, format=None):
    # apikeyheader = request.headers['Api-key']
    if 'Api-key' in request.headers.keys():
        if API_KEY == request.headers['Api-key']:
            if request.method == 'POST':
                requestdata = request.POST.copy()
                
                if 'companycode' not in requestdata.keys() or 'empcode' not in requestdata.keys():
                    return Response({'n':0,'msg':'Please Enter required keys','status':'Failed'})
                else:
                    company_code = requestdata['companycode']
                    EmployeeCode = requestdata['empcode']

                    if company_code == "" and EmployeeCode == "":
                        return Response({'n':0,'msg':'Please Enter valid inputs','status':'Failed'})
                    elif company_code == "" and EmployeeCode != "":
                    
                        user = Users.objects.filter(is_active=True,uid=EmployeeCode).order_by('id')
                    elif company_code != "" and EmployeeCode == "":
                    
                        user = Users.objects.filter(is_active=True,company_code=company_code).order_by('id')
                    else:
                    
                        user = Users.objects.filter(is_active=True,uid=EmployeeCode,company_code=company_code).order_by('id')
                
                    employeelist = []
                    if user.exists():
                        serializer = BasicInfoSerializer(user, many=True)
                        for i in serializer.data:
                            
                            if i['DateofJoining'] is not None:
                                datestr = i['DateofJoining']
                                newdate = datetime.datetime.strptime(datestr,"%Y-%m-%d")
                                newmonthdate = newdate.strftime("%d %b %Y")
                                i['DateofJoining']=newmonthdate
                                
                            usersecondaryobject = UserSecondaryInfo.objects.filter(userId=i['id']).first()
                            if usersecondaryobject is not None:
                                usersecser = UserSecondarySerializer(usersecondaryobject)
                                i['Secondary Info'] = usersecser.data
                            else:
                                i['Secondary Info'] = ""

                        return Response({'n':1,'msg':'Employee list fetched successfully','status':'success','data':serializer.data})
                    else:
                        return Response({'n':0,'msg':'Employee not found','status':'Failed'})
        else:
            return Response({'n':0,'msg':'Invalid api-key','status':'Failed'})

    else:
        return Response({'n':0,'msg':'Please Provide api-key','status':'Failed'})
        

























@api_view(['GET'])
def m_notfcount(request, format=None):
    userId = request.user.id
    notficationobjectCount = TaskNotification.objects.filter(UserID = userId ,ReadMsg=False).count()
    return Response({'n':1,'msg':'count fetched successfully','status':'success','data':notficationobjectCount})



@api_view(['GET'])
def onboardinglist(request):
    company_code = request.user.company_code
    serobj = Users.objects.filter(is_active=True,onboarding=False,company_code=company_code).order_by('-id')
    serializer = UsersSerializer(serobj,many=True)
    for i in serializer.data:
        if i['DateofJoining']:
            i["joiningdate"]=convertdate(i['DateofJoining'])
        else:
            i["joiningdate"]="---"

        usersecondaryobj=UserSecondaryInfo.objects.filter(userId=i['id']).first()
        i['reasonofnonverification']=""

        if usersecondaryobj:
            if i['reason_of_rejection_documents'] != None and usersecondaryobj.finalsubmit == False and i['documentverified'] == False:
                i['reasonofnonverification']="Waiting for employee to resubmit form"

            elif i['reason_of_rejection_documents'] == None and usersecondaryobj.finalsubmit == False and i['documentverified'] == False:
                if i['onboarding_get_mail'] == True:
                    i['reasonofnonverification']="Waiting for employee to submit form"
                else:
                    i['reasonofnonverification']="Employee did not get mail"


            elif i['reason_of_rejection_documents'] == None and usersecondaryobj.finalsubmit == True and i['documentverified'] == False:
                i['reasonofnonverification']="Waiting for admin to verify documents"

            elif i['reason_of_rejection_documents'] != None and usersecondaryobj.finalsubmit == True and i['documentverified'] == False:
                i['reasonofnonverification']="Waiting for admin to  reverify documents"
        else:
            
            if i['onboarding_get_mail'] == True:
                i['reasonofnonverification']="Waiting for employee to submit form"
            else:
                i['reasonofnonverification']="Employee did not get mail"

    return Response({'n':1,'msg':'Onboarding data found successfully','status':'success','data':serializer.data})



@api_view(['POST'])
def onboard_login_password(request):
    id = request.POST.get('id')
    data = {}
    data['email'] = request.POST.get('email')

    data['onboarding'] = True
    useremailobj = Users.objects.filter(email=data['email'],company_code=request.user.company_code,is_active=True).exclude(id=id).first()
    if useremailobj is not None:
        return Response({'n':0,'msg':'Email Already Exist.','status':'failed','data':{}})
    else:
        useremailobj = Users.objects.filter(id=id,company_code=request.user.company_code,is_active=True).first()
        data['Password'] = useremailobj.Firstname.capitalize()+'@123'
        data['password'] = make_password(useremailobj.Firstname.capitalize()+'@123') 
        userser = UserSerializer(useremailobj,data=data,partial=True)
        if userser.is_valid():
            userser.save()
            userobj=Users.objects.filter(id=id).first()
            email = userobj.email
            personal_email = userobj.personal_email
            password = userobj.Password
            curruser = Users.objects.filter(email=email,company_code=request.user.company_code,is_active=True).annotate(
                Name=Concat('Firstname', Value(' '), 'Lastname'))
            Name = curruser.first().Name.title()

            dicti = {'password': password,
                        'Name': Name, 'email': email}
          
            message = get_template(
                'useregistermail.html').render(dicti) 
            msg = EmailMessage(
                'TrackPro credentials',
                message,
                EMAIL_HOST_USER,
                [personal_email],
            )
            msg.content_subtype = "html"  # Main content is now text/html
            msg.send()
            return Response({'n':1,'msg':'Onboarded Successfully.','status':'success','data':userser.data})
        else:
            return Response({'n':0,'msg':'Not Onboarded.','status':'failed','data':{}})
            
            

        
   
    
    

# @api_view(['POST'])
# @permission_classes((AllowAny,))
# def uploadattendance(request):
#     attendancedata = request.data.get('attendancedata')
#     companycode = "O001"
#     for i in attendancedata:
#         dt=i['LogDate']
#         Date=dt.split(" ")[0]
#         Time=dt.split(" ")[1]
#         nYear = str(Date).split("-")[2]
#         nMonth = str(Date).split("-")[1]
#         nDay = str(Date).split("-")[0]
#         newDate = str(nYear)+ "-" +str(nMonth)+ "-" +str(nDay)
#         date_given = datetime.datetime(year=int(nYear), month=int(nMonth), day=int(nDay)).date()
#         nweek = week_of_month(date_given)

#         empexist = attendance.objects.filter(employeeId=(i['UserId']),company_code=companycode,deviceId = int(i['DeviceId']),date=newDate,time=Time).order_by('time').first()
        
        
        
        
        
        
        
#         current_date = datetime.datetime.strptime(newDate, '%Y-%m-%d').date()
#         yesterday_date = current_date - datetime.timedelta(days=1)
        
#         check_last_checkin = attendance.objects.filter(employeeId=str(i['UserId']),date=str(current_date),checkout=False).order_by('time').last()
        
#         check_last_checkout=None
#         if check_last_checkin is not None:
#             check_last_checkout = attendance.objects.filter(employeeId=str(i['UserId']),time__gt=check_last_checkin.time,date=str(current_date),checkout=True).order_by('time').last()
#             if check_last_checkout is not None:
#                 print("today allow him to checkin")
#             else: 
#                 print("today allow him to checkout")
#         else:
#             check_last_checkin = attendance.objects.filter(employeeId=str(i['UserId']),date=str(yesterday_date),checkout=False).order_by('time').last()
#             if check_last_checkin is not None:
#                 check_last_checkout = attendance.objects.filter(Q(employeeId=str(i['UserId']),time__gt=check_last_checkin.time,date=str(yesterday_date),checkout=True)|Q(employeeId=str(i['UserId']),date=str(current_date),checkout=True)).order_by('date','time').last()
#                 if check_last_checkout is not None:
#                     print("yesterday allow him to checkin")
#                 else: 
#                     print("yesterday allow him to checkout")
#             else:
#                 print("allow him to checkin")
        
        
#         if check_last_checkin is not None :
#             get_data=1
#         else:
#             get_data=0
            
            
#         if check_last_checkout is not None:
#             get_data=0
#         else:
#             if check_last_checkin is not None :
#                 get_data=1
#             else:
#                 get_data=0


#         # 1 checkout True disable checkin
#         # 0 checkout True enable checkin
        
#         if get_data == 1:
#             checkout=True
#         else:
#             checkout=False
            
#         if empexist is None:
#             attendance.objects.create(employeeId =i['UserId'],date=newDate,company_code=companycode,deviceId = int(i['DeviceId']),time=Time,Week=nweek,Month=nMonth,Year=nYear,checkout=checkout)

#     return Response({"response":{'n': 1,'msg':'file data saved successfully','Status': 'success'}})
   
@api_view(['POST'])
@permission_classes((AllowAny,))
def uploadattendance(request):
    attendancedata = request.data.get('attendancedata')
    company_code = "O001"
    for i in attendancedata:
        Date=i['LogDate'].split(" ")[0]
        Time=i['LogDate'].split(" ")[1]
        NewDate=str(str(Date).split("-")[2])+ "-" +str(str(Date).split("-")[1])+ "-" +str(str(Date).split("-")[0])
        DateTime = NewDate + " " + str(Time)
        date_given = datetime.datetime(year=int(str(str(Date).split("-")[2])), month=int(str(str(Date).split("-")[1])), day=int(str(str(Date).split("-")[0]))).date()
        nweek = week_of_month(date_given)
        shift=get_current_shift_details(str(DateTime),i['UserId'])
        if shift['data'] ==1:
            checkout=True
        else:
            checkout=False
        empexist = attendance.objects.filter(employeeId=(i['UserId']),company_code=company_code,deviceId = int(i['DeviceId']),date=NewDate,time=Time).order_by('time').first()
        if empexist is  None:
            attendance.objects.create(employeeId =i['UserId'],date=NewDate,company_code=company_code,deviceId = int(i['DeviceId']),time=Time,Week=nweek,Month=str(str(Date).split("-")[1]),Year=str(str(Date).split("-")[0]),checkout=checkout)
            
    return Response({"response":{'n': 1,'msg':'file data saved successfully','Status': 'success'}})
   



            

     
    
    
    
    

@api_view(['GET'])
@authentication_classes([])
@permission_classes([])
def late_empscheduler(request):
    today = datetime.date.today()
    year = today.year
    month = today.month
    dateobj = attendance.objects.filter(date=today,emailsent=False).distinct("employeeId")
    print("dateobj",dateobj)
    if dateobj is not None:
        latedatesser  = attendanceserializer(dateobj,many=True)
        for i in latedatesser.data:
            print("i")
            getdata = get_date_shift_details(str(today),i['employeeId'])
            print("shiftlist",getdata['shifts_list'])

            # for s in getdata['shifts_list'] :



    
        return Response({'n':1,'msg':'data found Successfully.','status':'success','data':''})
    else:
        return Response({'n':1,'msg':'No employee found.','status':'success','data':''})
    
    


# @api_view(['GET'])
# @authentication_classes([])
# @permission_classes([])
# def late_empscheduler(request):
#     today = datetime.date.today()
#     year = today.year
#     month = today.month
#     # Year=year,Month = month
#     s1 = "10:00:00"
#     s2 = "11:00:00"


#     nndatestr = str(today)
#     nmonth_name = calendar.month_abbr[int(nndatestr.split('-')[1])]    
#     ndatestr = nndatestr.split('-')[2]+" "+nmonth_name+" "+nndatestr.split('-')[0]
#     STrdate = ndatestr

#     dateobj = attendance.objects.filter(date=today,time__range=(s1,s2),emailsent=False).distinct("employeeId")
#     if dateobj is not None:
#         latedatesser  = attendanceserializer(dateobj,many=True)
#         for i in latedatesser.data:
#             empid = i['employeeId']
#             empobj = Users.objects.filter(employeeId=empid).first()
#             if empobj is not None:
#                 empname = empobj.Firstname +" "+ empobj.Lastname
#                 emailid = empobj.email
#                 arrivalstr = str(i['time'])
            
#                 lateattendancecount = attendance.objects.filter(time__range=(s1, s2),Year=year,Month=month,employeeId=i['employeeId']).distinct('date').count()
#                 dicti = {
#                         "employeename":empname,
#                         "date":STrdate,
#                         "arrivalTime":arrivalstr,
#                         "nooflatemarks":lateattendancecount,
#                         }
#                 message = get_template(
#                     'mails/lateattendance_empemail.html').render(dicti)
                
#                 msg = EmailMessage(
#                     'Reminder: Punctuality at the Office',
#                     message,
#                     EMAIL_HOST_USER,
#                     [emailid],
#                 )
#                 msg.content_subtype = "html"  # Main content is now text/html
#                 # msg.send()
#                 varpp = msg.send()

#                 if varpp == 1:
#                     attendance.objects.filter(date=today,time__range=(s1, s2),employeeId=i['employeeId']).update(emailsent=True)
#         return Response({'n':1,'msg':'data found Successfully.','status':'success','data':latedatesser.data})
#     else:
#         return Response({'n':1,'msg':'No employee found.','status':'success','data':''})
 





@api_view(['GET'])
@authentication_classes([])
@permission_classes([])
def late_manager_scheduler(request):

    todaysdate = date.today()
    current_month = todaysdate.month
    current_year = todaysdate.year
    s1 = "10:00:00"
    s2 = "11:00:00"
    company_code = "O001"

    lst = []
    Usertomanagerobject = UserToManager.objects.filter(company_code = company_code).distinct('ManagerID')
    for m in Usertomanagerobject:
        manageremail = Users.objects.filter(id=m.ManagerID_id).first()
        lst.append({"ManagerID": m.ManagerID_id,
                    "ManagerIDStr": m.ManagerIDStr,
                    "manageremail":manageremail.email,
                    "Users": []})
    for m in UserToManager.objects.filter(company_code = company_code):
        for l in lst:
            if m.ManagerID_id == l["ManagerID"]:
                userobject  = Users.objects.filter(id=m.UserID_id).first()
                employeeId = userobject.employeeId
                attendanceobject = attendance.objects.filter(time__range=(s1, s2),date=todaysdate,employeeId=employeeId).order_by('time').first()
                if attendanceobject is not None:
                # attendanceser = attendanceserializer(attendanceobject,many=True)
                    lateattendancecount = attendance.objects.filter(time__range=(s1, s2),Year=current_year,Month=current_month,employeeId=employeeId).distinct('date').count()

                    l["Users"].append(
                    {"UserID": m.UserID_id, "UserIDStr": m.UserIDStr,"Usertiming":attendanceobject.time,"lateattendancecount":lateattendancecount})
            
    managerlist = []
    for i in lst:
        wholecontent =""
        greetingcontent = "Dear Sir/Madam," +"<br>" +"<span style='padding-left:40px;'>" +" I hope this message finds you well. As part of our commitment to maintaining punctuality in the workplace, I wanted to provide you with a list of employees who have had instance of late attendance today."+"<br>"+"<br>"+"Please find below the list of employees, along with the number of late arrivals for each:"+"<br>"

        footercontent = "<br><br>"+"Regards ,"+"<br>"+ "Zentro Team"
        if i['Users'] != []:
            
            weeklymessage = ""
            message = ""
            headingmessage =""
            newlist = sorted(i["Users"], key=itemgetter('lateattendancecount'),reverse=True)
            i['Users'] = newlist
            weeklymessage += "Late Attendance Report"
            
            for k in i['Users']:
                message += """
                            <tr>
                                <td style="border: 1px solid #dddddd;">"""+ k['UserIDStr'] + """</td>
                                <td style="border: 1px solid #dddddd; text-align: center;">"""+ str(k['Usertiming']) +"""</td>
                                <td style="border: 1px solid #dddddd;">"""+ str(k['lateattendancecount']) +"""</td>
                            </tr>
                        """
            headingmessage += """<table>
                    <tr>
                        <th style="border: 1px solid #dddddd; padding: 4px;">Employee Name</th>
                        <th style="border: 1px solid #dddddd; padding: 4px;">Today's timing</th>
                        <th style="border: 1px solid #dddddd; padding: 4px;">Monthly late Arrivals</th>
                    </tr> 
                    """+message+"""
                    
                    </table>"""
            wholecontent += "<br>" + weeklymessage + headingmessage
            
            managerlist.append(i)

        

            mailcontent = greetingcontent + wholecontent + footercontent
            msg = EmailMessage(
                    'Late attendance Report',
                    mailcontent,
                    EMAIL_HOST_USER,
                    [i['manageremail']],
                    )
            msg.content_subtype = "html"  # Main content is now text/html
            msg.send()
                          
                


    # for i in lst:    
    #     attendanceobject = attendance.objects.filter(time__range=(s1, s2),date=todaysdate)
    #     attendanceser = attendanceserializer(attendanceobject,many=True)
    #     for i in attendanceser.data:
    #         userobject = Users.objects.filter(employeeId= i['employeeId']).first()
    #         usermanagerobject = list(UserToManager.objects.filter(UserID=int(userobject.id)).values_list('ManagerID', flat=True))
    
    return Response({'n':1,'msg':'data found Successfully.','status':'success','data':managerlist})


@api_view(['GET'])
@permission_classes((AllowAny,))
def taskmanagercheck(request):
    userid = request.query_params.get('userID', None)
    # companycode = request.user.company_code
    if userid is not None and userid != "None":
        manobj = UserToManager.objects.filter(ManagerID=userid).first()
        if manobj is not None:
            IsTaskmanager = True
        else:
            IsTaskmanager = False
        leavemanobj = leaveMapping.objects.filter(managerId=userid).first()
        if leavemanobj is not None:
            Isleavemanager = True
        else:
            Isleavemanager = False
        return Response({'n':1,'msg':'data found Successfully.','status':'success','IsTaskmanager':IsTaskmanager,'Isleavemanager':Isleavemanager})
    return Response({'n':0,'msg':'data not found.','status':'failed','IsTaskmanager':False,'Isleavemanager':False})



@api_view(['GET'])
def updateprofile_otpmail(request):
    userid = request.user.id
    userobj = Users.objects.filter(id=userid,company_code=request.user.company_code,is_active=True).first()
    if userobj is not None:
        useremail = userobj.email
        Name = userobj.Firstname +" "+ userobj.Lastname
        otpnumber = randint(100000, 999999)
        existotp =  Checkotp.objects.filter(userid=userobj.id,Active=True)
        if existotp.exists():
            optobject = Checkotp.objects.filter(userid=userobj.id,Active=True).update(Active=False)
            createotp = Checkotp.objects.create(userid=userobj.id,otp=otpnumber)
        else:
            createotp = Checkotp.objects.create(userid=userobj.id,otp=otpnumber)
        dicti = {'otp': otpnumber,'Name': Name}
        message = get_template(
            'otpmail.html').render(dicti)
        msg = EmailMessage(
            'Update Profile - OTP!',
            message,
            EMAIL_HOST_USER,
            [useremail],
        )
        msg.content_subtype = "html"  # Main content is now text/html
        msg.send()
        return Response({"response":{'n': 1,'msg':'An OTP has been sent to '+ useremail,'Status': 'success'}})
    return Response({"response":{'n': 0,'msg':'User not found','Status': 'failed'}})

@api_view(['POST'])
def updatemyprofile(request):

    userid = request.user.id
    otp = request.data.get('otpnumber')
    checkotpobject = Checkotp.objects.filter(userid=userid,Active=True,otp=otp).first()
    if checkotpobject is not None:

        requestdata = request.POST.copy()
        ProfilePicture=request.FILES.get('Photo')
        data=requestdata
        data['userId']=userid
        if ProfilePicture is not None and ProfilePicture !="":
            data['Photo'] = ProfilePicture
        userdict={}
        secondarydict={}

        userobject = Users.objects.filter(is_active=True,id=userid).first()
        if userobject is not None:
            userSerializer=RegisterSerializer(userobject,data=data,partial=True)
            if userSerializer.is_valid():
                userSerializer.save()
                userdict['n']=1
                userdict['data']=userSerializer.data
                userdict['status']='success'
        else:
            userdict['n']=0
            userdict['data']=userSerializer.errors
            userdict['status']='failed'




        secodaryinfoobject = UserSecondaryInfo.objects.filter(userId=userid).first()
        if secodaryinfoobject is not None:
            secodaryserializer = UserSecondarySerializer(secodaryinfoobject,data=data,partial=True)
            if secodaryserializer.is_valid():
                secodaryserializer.save()
                secondarydict['n']=1
                secondarydict['data']=secodaryserializer.data
                secondarydict['status']='success'
                secondarydict['operation']='updated'
                secondarydict['operation']='updatation failed'


        else:

            secodaryserializer = UserSecondarySerializer(data=data,partial=True)
            if secodaryserializer.is_valid():
                secodaryserializer.save()
                secondarydict['n']=1
                secondarydict['data']=secodaryserializer.data
                secondarydict['status']='success'
                secondarydict['operation']='new created'

            else:
                secondarydict['n']=0
                secondarydict['data']=secodaryserializer.errors
                secondarydict['status']='failed'
                secondarydict['operation']='new creatation failed'
                print("secodaryserializer.errors",secodaryserializer.errors)
        
        return Response({"response":{'n': 1,'msg':'profile updated','Status': 'success','data':{'user':userdict,'secondaryinfo':secondarydict}}})

    else:
        return Response({"response":{'n': 0,'msg':'OTP not valid','Status': 'failed'}})

@api_view(['GET'])
def Employee_allinfo(request):
    userid = request.user.id
    documentlist = []
    secondaryinfo = []
    userSecondaryObject = UserSecondaryInfo.objects.filter(userId=userid).order_by('id').first()
    sec_serializer = UserSecondarySerializer(userSecondaryObject)
    for s in [sec_serializer.data]:
        if s['residentialaddresscity'] is not None and s['residentialaddresscity'] !='':
            rcityname_obj=Cities.objects.filter(id=s['residentialaddresscity']).first()
            if rcityname_obj is not None:
                s['residential_cityname']=rcityname_obj.name
            else:
                s['residential_cityname']='---'
        else:
            s['residential_cityname']='---'

        if s['residentialaddressstate'] is not None and s['residentialaddressstate'] !='':
            rstatename_obj=State.objects.filter(id=s['residentialaddressstate']).first()
            if rstatename_obj is not None:
                s['residential_statename']=rstatename_obj.name
            else:
                s['residential_statename']='---'
        else:
            s['residential_statename']='---'

        if s['residentialaddresscountry'] is not None and s['residentialaddresscountry'] !='':
            rcountryname_obj=Country.objects.filter(id=s['residentialaddresscountry']).first()
            if rcountryname_obj is not None:
                s['residential_countryname']=rcountryname_obj.name
            else:
                s['residential_countryname']='---'
        else:
            s['residential_countryname']='---'




        if s['permanantaddresscity'] is not None and s['permanantaddresscity'] !='':
            pcityname_obj=Cities.objects.filter(id=s['permanantaddresscity']).first()
            if pcityname_obj is not None:
                s['permanant_cityname']=pcityname_obj.name
            else:
                s['permanant_cityname']='---'
        else:
            s['permanant_cityname']='---'

        if s['permanantaddressstate'] is not None and s['permanantaddressstate'] !='':
            pstatename_obj=State.objects.filter(id=s['permanantaddressstate']).first()
            if pstatename_obj is not None:
                s['permanant_statename']=pstatename_obj.name
            else:
                s['permanant_statename']='---'
        else:
            s['permanant_statename']='---'

        if s['permanantaddresscountry'] is not None and s['permanantaddresscountry'] !='':
            pcountryname_obj=Country.objects.filter(id=s['permanantaddresscountry']).first()
            if pcountryname_obj is not None:
                s['permanant_countryname']=pcountryname_obj.name
            else:
                s['permanant_countryname']='---'
        else:
            s['permanant_countryname']='---'




        if s['adhaarcardimage'] is not None:
            s['adhaardocid'] = s['adhaarcardimage'].replace("/","")
        else:
            s['adhaardocid'] = ""

        if s['pancardimage'] is not None:
            s['pandocid'] = s['pancardimage'].replace("/","")
        else:
            s['pandocid'] = ""

        if s['passportimage'] is not None:
            s['passportdocid'] = s['passportimage'].replace("/","")
        else:
            s['passportdocid'] = ""

        secondaryinfo.append(s)
        documentlist.append({
            "documentname" : "adhaar",
            "document" : s['adhaarcardimage'],
            "docid" : s['adhaardocid']
        })
        documentlist.append({
            "documentname" : "pancard",
            "document" : s['pancardimage'],
            "docid" : s['pandocid']
        })
        if  s['passportimage'] is not None:
            documentlist.append({
                "documentname" : "passport",
                "document" : s['passportimage'],
                "docid" : s['passportdocid']
            })


   
    usereducationalObject = educational_qualifications.objects.filter(userid=userid).order_by('fromdate')
    edu_serializer = educational_qualificationsSerializer(usereducationalObject,many=True)
    for i in edu_serializer.data:
        if i['marksheet'] is not None:
            i['marksheetdocid'] = i['marksheet'].replace("/","")
        else:
            i['marksheetdocid'] = ""

        documentlist.append({
            "documentname" : i['qualification_name'],
            "document" : i['marksheet'],
            "docid" : i['marksheetdocid']

        })
        i['formatedfromdate']=dd_month_year_format(i['fromdate'])
        i['formatedtodate']=dd_month_year_format(i['todate'])
        if i['marksheet'] is not None:
            i['docid'] = i['marksheet'].replace("/","")




    exist_obj=Previous_Company_Details.objects.filter(userid=userid)
    com_serializer = Previous_Company_Details_Serializer(exist_obj,many=True)
    for i in com_serializer.data:
        i['formatedjoinDate']=dd_month_year_format(i['joinDate'])
        i['formatedleaveDate']=dd_month_year_format(i['leaveDate'])
        if i['salaryslip'] is not None:
            i['docid'] = i['salaryslip'].replace("/","")
        else:
            i['docid'] = ''
        documentlist.append({
            "documentname" : 'salaryslip (' + i['companyname'] +')',
            "document" : i['salaryslip'],
            "docid" :  i['docid']

        })
        if i['relieving'] is not None:
            i['docid'] = i['relieving'].replace("/","")
        else:
            i['docid'] = ''
        documentlist.append({
            "documentname" : 'relieving letter (' + i['companyname']+')',
            "document" : i['relieving'],
            "docid" :  i['docid']

        })

    user_object = Users.objects.filter(is_active=True,id=userid).first()
    user_ser = UserSerializer(user_object)
    
    userdata_i=user_ser.data
    if str(userdata_i['Photo']) is not None and str(userdata_i['Photo']) != "" and str(userdata_i['Photo']) != "None" :
        userdata_i['profileimage'] = imageUrl + str(userdata_i['Photo'])
    else:
        userdata_i['profileimage'] = imageUrl + "/static/assets/images/profile.png"



    context = {
        'userSecondaryObject':secondaryinfo[0],
        'userdata':userdata_i,
        'educationdetails':edu_serializer.data,
        'company_details':com_serializer.data,
        'documentlist':documentlist
    }

    return Response({'n':1,'msg':'success','status':'success','data':context})


@api_view(['GET'])
def Employee_educational_qualification(request):
    userid = request.user.id
    documentlist = []
    secondaryinfo = []

   
    usereducationalObject = educational_qualifications.objects.filter(userid=userid).order_by('fromdate')
    edu_serializer = educational_qualificationsSerializer(usereducationalObject,many=True)
    for i in edu_serializer.data:
        if i['marksheet'] is not None:
            i['marksheetdocid'] = i['marksheet'].replace("/","")
        else:
            i['marksheetdocid'] = ""

        documentlist.append({
            "documentname" : i['qualification_name'],
            "document" : i['marksheet'],
            "docid" : i['marksheetdocid']

        })
        i['formatedfromdate']=dd_month_year_format(i['fromdate'])
        i['formatedtodate']=dd_month_year_format(i['todate'])
        if i['marksheet'] is not None:
            i['docid'] = i['marksheet'].replace("/","")







    context = {
        'educationdetails':edu_serializer.data,
        'documentlist':documentlist
    }

    return Response({'n':1,'msg':'success','status':'success','data':context})

 

@api_view(['POST'])
def Employee_all_info_by_id(request):
    userid = request.POST.get('userid')
    if userid is None or userid == '':
        return Response({'n':0,'msg':'user id not found','status':'error','data':[]})
        
    documentlist = []
    secondaryinfo = []
    userSecondaryObject = UserSecondaryInfo.objects.filter(userId=userid).order_by('id').first()
    sec_serializer = UserSecondarySerializer(userSecondaryObject)
    for s in [sec_serializer.data]:

        if s['adhaarcardimage'] is not None:
            s['adhaardocid'] = s['adhaarcardimage'].replace("/","")
        else:
            s['adhaardocid'] = ""

        if s['pancardimage'] is not None:
            s['pandocid'] = s['pancardimage'].replace("/","")
        else:
            s['pandocid'] = ""

        if s['passportimage'] is not None:
            s['passportdocid'] = s['passportimage'].replace("/","")
        else:
            s['passportdocid'] = ""

        secondaryinfo.append(s)
        documentlist.append({
            "documentname" : "adhaar",
            "document" : s['adhaarcardimage'],
            "docid" : s['adhaardocid']
        })
        documentlist.append({
            "documentname" : "pancard",
            "document" : s['pancardimage'],
            "docid" : s['pandocid']
        })
        if  s['passportimage'] is not None:
            documentlist.append({
                "documentname" : "passport",
                "document" : s['passportimage'],
                "docid" : s['passportdocid']
            })
   
    usereducationalObject = educational_qualifications.objects.filter(userid=userid).order_by('fromdate')
    edu_serializer = educational_qualificationsSerializer(usereducationalObject,many=True)
    for i in edu_serializer.data:
        if i['marksheet'] is not None:
            i['marksheetdocid'] = i['marksheet'].replace("/","")
        else:
            i['marksheetdocid'] = ""

        documentlist.append({
            "documentname" : i['qualification_name'],
            "document" : i['marksheet'],
            "docid" : i['marksheetdocid']

        })
        i['formatedfromdate']=dd_month_year_format(i['fromdate'])
        i['formatedtodate']=dd_month_year_format(i['todate'])
        if i['marksheet'] is not None:
            i['docid'] = i['marksheet'].replace("/","")

    exist_obj=Previous_Company_Details.objects.filter(userid=userid)
    com_serializer = Previous_Company_Details_Serializer(exist_obj,many=True)
    for i in com_serializer.data:
        i['formatedjoinDate']=dd_month_year_format(i['joinDate'])
        i['formatedleaveDate']=dd_month_year_format(i['leaveDate'])
        if i['salaryslip'] is not None:
            i['docid'] = i['salaryslip'].replace("/","")
        else:
            i['docid']=''

        documentlist.append({
            "documentname" : 'salaryslip (' + i['companyname'] +')',
            "document" : i['salaryslip'],
            "docid" : i['docid']

        })
        if i['relieving'] is not None:
            i['docid'] = i['relieving'].replace("/","")
        else:
            i['docid']=''
        documentlist.append({
            "documentname" : 'relieving letter (' + i['companyname']+')',
            "document" : i['relieving'],
            "docid" : i['docid']

        })

    user_object = Users.objects.filter(is_active=True,id=userid).first()
    user_ser = UserAllInfoSerializer(user_object)

    context = {
        'userSecondaryObject':secondaryinfo[0],
        'userdata':user_ser.data,
        'educationdetails':edu_serializer.data,
        'company_details':com_serializer.data,
        'documentlist':documentlist
    }

    return Response({'n':1,'msg':'success','status':'success','data':context})



@api_view(['POST'])
def sendonboardinglink(request):
    data = {'n': '', 'Msg': '', 'Status': ''}

    if request.method == 'POST':
        request_data = request.data.copy()

        current_time =  timezone.localtime(timezone.now())
        formatted_time = current_time.strftime('%Y-%m-%dT%H:%M:%S.%f%z')

        request_data['company_code'] = request.user.company_code
        userID = request.query_params.get('userID')
        if userID is None:
            data['n'] = 0
            data['Msg'] = 'User ID is none'
            data['Status'] = "Failed"
        else:
            user = Users.objects.filter(id=userID,is_active = True,company_code = request_data['company_code']).first()
            if user is None:
                data['n'] = 0
                data['Msg'] = 'USER DOES NOT EXIST'
                data['Status'] = "Failed"

            else:
                request_data = request.data.copy()
                if request.user.is_superuser == True:
                    request_data['is_superuser'] = True
                else:
                    request_data['is_superuser'] = False

                if request_data['employeeId'] !="":
                    checkempid_obj=Users.objects.filter(employeeId=request_data['employeeId'],company_code=request.user.company_code,is_active=True).exclude(id=userID).first()
                    if checkempid_obj:
                        return Response({"n": 0, "Msg": "Employee of this Attendance Id Already Exists", "Status": "Failed"}) 

                serializer = userUpdateSerializer(user, data = request_data,partial=True)
                if serializer.is_valid():
                    serializer.validated_data['is_admin'] = True
                    serializer.validated_data['UpdatedBy'] = request.user
                    serializer.validated_data['linkdatetime'] = formatted_time
                    serializer.save()
                    usersecondaryinfo_obj=UserSecondaryInfo.objects.filter(userId=serializer.data['id']).first()
                    if usersecondaryinfo_obj is None:
                        UserSecondaryInfo.objects.create(userId=serializer.data['id'])

                    data['n'] = 1
                    data['Msg'] = 'update successfull'
                    data['Status'] = "Success"

                    try:
                        personal_email = request.data.get('personal_email', None)
                        password = request.data.get('password', None)
                        Name = serializer.data['Firstname'] +" "+ serializer.data['Lastname']
                        designationId = serializer.data['DesignationId']
                        desigobj = Designation.objects.filter(id=designationId,isactive=True).first()
                        designation = desigobj.DesignationName
                        compobj = companyinfo.objects.filter(companycode=serializer.data['company_code'],isactive=True).first()
                        companyName = compobj.companyName
                        d1=datetime.datetime.strptime(serializer.data['DateofJoining'], "%Y-%m-%d")
                        DateofJoining = d1.strftime('%d %B %Y')

                        dicti = {'password': password,'Name': Name, 'designation':designation,
                                'companyName':companyName,'DateofJoining':DateofJoining,
                                'personal_email': personal_email,'id':serializer.data['id'],
                                'hosturl':frontUrl}


                        message = get_template(
                            'usersecondaryemail.html').render(dicti)
                        msg = EmailMessage(
                            'Onboarding Details and Document Submission - Welcome to Zentro!',
                            message,
                            EMAIL_HOST_USER,
                            [personal_email],
                        )
                        msg.content_subtype = "html"  # Main content is now text/html
                        msg.send()
                        user.onboarding_get_mail = True
                        user.save()
                    except Exception as e:
                        print('exception occured fot mail', e)


                    return Response({"n": 1, "Msg": "Onboarding Details and Document Submission mail sent Successfully", "Status": data})
                else:
                    data = serializer.errors
                    return Response({"n": 0, "Msg": data, "Status": data})

        return Response(data=data)

@api_view(['POST'])
# @permission_classes((AllowAny,))
def update_employee_first_step(request):
    id=request.POST.get('id')
    company_code=request.user.company_code
    data={}
    data['Firstname']=request.POST.get('Firstname')
    data['Lastname']=request.POST.get('Lastname')
    data['Phone']=request.POST.get('Phone')
    data['BirthDate']=request.POST.get('BirthDate')
    data['alternatemail']=request.POST.get('alternatemail')
    data['personal_email']=request.POST.get('personal_email')
    data['bloodgroup']=request.POST.get('bloodgroup')
    data['maritalstatus']=request.POST.get('maritalstatus')
    data['Addressline2']=request.POST.get('rAddresslinetwo')
    data['Address']=request.POST.get('rAddresslineone')
    
    
    data['res_lattitude']=request.POST.get('res_lattitude')
    data['res_longitude']=request.POST.get('res_longitude')
    data['res_radius']=request.POST.get('res_radius')

    data['permanentaddressLine2']=request.POST.get('pAddresslinetwo')
    data['permanentaddress']=request.POST.get('pAddresslineone')

    data['relationname1']=request.POST.get('relationname1')
    data['relation1']=request.POST.get('relation1')
    data['relation1number']=request.POST.get('relation1number')

    data['relationname2']=request.POST.get('relationname2')
    data['relation2']=request.POST.get('relation2')
    data['relation2number']=request.POST.get('relation2number')

    data['adhaarcard']=request.POST.get('adhaarcard')
    data['pancard']=request.POST.get('pancard')
    if request.POST.get("residentialaddresscountry") is not None and request.POST.get("residentialaddresscountry") !="":
        data["residentialaddresscountry"]=request.POST.get("residentialaddresscountry")
    if request.POST.get("residentialaddressstate") is not None and request.POST.get("residentialaddressstate") !="":
        data["residentialaddressstate"]=request.POST.get("residentialaddressstate")
    if request.POST.get("residentialaddresscity") is not None and request.POST.get("residentialaddresscity") !="":    
        data["residentialaddresscity"]=request.POST.get("residentialaddresscity")
    if request.POST.get("residentialaddresspincode") is not None and request.POST.get("residentialaddresspincode") !="":
        data["residentialaddresspincode"]=request.POST.get("residentialaddresspincode")
    if request.POST.get("permanantaddresscountry") is not None and request.POST.get("permanantaddresscountry") !="":
        data["permanantaddresscountry"]=request.POST.get("permanantaddresscountry")
    if request.POST.get("permanantaddressstate") is not None and request.POST.get("permanantaddressstate") !="":
        data["permanantaddressstate"]=request.POST.get("permanantaddressstate")
    if request.POST.get("permanantaddresscity") is not None and request.POST.get("permanantaddresscity") !="":
        data["permanantaddresscity"]=request.POST.get("permanantaddresscity")
    if request.POST.get("permanantaddresspincode") is not None and request.POST.get("permanantaddresspincode") !="":
        data["permanantaddresspincode"]=request.POST.get("permanantaddresspincode")



    adhar_pic=request.POST.get('adhaarbase64')
    pan_pic=request.POST.get('pancardbase64')
    passport_pic=request.POST.get('passportbase64')



    if passport_pic is not None and passport_pic !='':
        data['passportimage']=request.POST.get('passportbase64')


    if adhar_pic is not None and adhar_pic !='':
        data['adhaarcardimage']=request.POST.get('adhaarbase64')


    if pan_pic is not None and pan_pic !='':
        data['pancardimage']=request.POST.get('pancardbase64')

 



    if id is not None and id !="":
        if data['personal_email'] is not None and data['personal_email'] !='':
            checkemailobj = Users.objects.filter(email=data['personal_email'],is_active=True,company_code=company_code).exclude(id=id).first()
            if checkemailobj is not None:
                response_={
                        'status':0,
                        'Msg':'Email Already Exist',
                        'data':{}
                    }
                return Response(response_,status=200)
        if data['Phone'] is not None and data['Phone'] !='':
            mobilenumberobj = Users.objects.filter(Phone=data['Phone'],is_active=True,company_code=company_code).exclude(id=id).first()
            if mobilenumberobj is not None:
                
                response_={
                        'status':0,
                        'Msg':'Mobile Number Already Exist',
                        'data':{}
                    }
                return Response(response_,status=200)
        

        updatePersonalDetailsobj = Users.objects.filter(id=id,is_active=True).first()

        serializer = UserSerializer(updatePersonalDetailsobj,data=data,partial=True)
        if serializer.is_valid():
            serializer.save()
       
        else:
            print("user error",serializer.errors)

        updatesecondaryobj = UserSecondaryInfo.objects.filter(userId=id).order_by('id').last()
        if updatesecondaryobj is not None:

            secondaryserializer = UserSecondarySerializer(updatesecondaryobj,data=data,partial=True)
            if secondaryserializer.is_valid():
                secondaryserializer.save() 
                response_={
                    'status':1,
                    'Msg':'Details Updated Successfully.',
                    'data':secondaryserializer.data
                }
                return Response(response_,status=200)
            else: 
                response_={
                    'status':0,
                    'Msg':'Details Not Updated',
                    'data':secondaryserializer.errors
                }
                return Response(response_,status=200)
        else:

            data['userId'] = id
            secondaryserializer = UserSecondarySerializer(data=data,partial=True)
            if secondaryserializer.is_valid():
                secondaryserializer.save() 
                response_={
                    'status':1,
                    'Msg':'Data Saved Successfully.',
                    'data':secondaryserializer.data
                }
                return Response(response_,status=200)
            else:
                response_={
                    'status':1,
                    'Msg':'Data not Saved.',
                    'data':secondaryserializer.errors
                }
                return Response(response_,status=200)
    else:
        response_={
            'status':0,
            'Msg':'ID not found',
            'data':{}
        }
        return Response(response_,status=200)
   


@api_view(['POST'])  
@permission_classes((AllowAny,))
def update_employee_second_step(request):
    if request.method == 'POST':
        userid=request.POST.getlist("user")
        qualificationname=request.POST.getlist("qualificationname")
        obtainmarks=request.POST.getlist("obtainmarks")
        university=request.POST.getlist("university")
        fromdate=request.POST.getlist("fromdate")
        marksheet=request.POST.getlist("marksheet")
        todate=request.POST.getlist("todate")

        for i in range(len(userid)):
            user_data = {}

            user_data['userid'] = userid[i]
            user_data['obtain_marks'] = obtainmarks[i]
            user_data['qualification_name'] = qualificationname[i]
            user_data['fromdate'] = fromdate[i]
            user_data['todate'] = todate[i]
            user_data['university'] = university[i]
            if marksheet[i] !='' and marksheet[i] is not None:
                user_data['marksheet'] = marksheet[i]


            # for key, value in request.FILES.items():
            #     if qualificationname[i] == key:
            #         user_data['marksheet'] = value




            alredy_exist_obj=educational_qualifications.objects.filter(userid=user_data['userid'],qualification_name=user_data['qualification_name']).first()
            if alredy_exist_obj:
                serializer = educational_qualificationsSerializer(alredy_exist_obj,data=user_data,partial=True)
            else:
                serializer = educational_qualificationsSerializer(data=user_data,partial=True)
            if serializer.is_valid():
                serializer.save()


            

    response_={
        'status':200,
        'Msg':'educational qualifications added',
        'data':{}
    }
    return Response(response_,status=200)
    
    
@api_view(['POST'])
@permission_classes((AllowAny,))
def update_employee_fourth_step(request):
    data={}
    id=request.POST.get('id')


    data['bankname']=request.POST.get('bankname')
    data['branchname']=request.POST.get('branchname')
    data['accountholdername']=request.POST.get('accountholdername')
    data['ifsccode']=request.POST.get('ifsccode')
    data['accountnumber']=request.POST.get('accountnumber')
    data['confirmaccountnumber']=request.POST.get('confirmaccountnumber')
    data['previous_pf_accountno']=request.POST.get('previouspfaccountno')
    data['esic_number']=request.POST.get('esic')
    data['finalsubmit']=True



    secondaryinfo_obj=UserSecondaryInfo.objects.filter(userId=id).order_by('-id').first()
    data['userId'] = id

    secondaryserializer = UserSecondarySerializer(secondaryinfo_obj,data=data,partial=True)
    if secondaryserializer.is_valid():
        secondaryserializer.save() 

        response_={
            'status':1,
            'Msg':'Data Saved Successfully.',
            'data':secondaryserializer.data
        }
        return Response(response_,status=200)
    else:
        response_={
                'status':0,
                'Msg':'Data Not Saved.',
                'data':secondaryserializer.errors
            }
        return Response(response_,status=200)
        


    
@api_view(['POST'])
# @permission_classes((AllowAny,))
def update_employee_fifth_step(request):
    data={}
    id=request.POST.get('id')
    data = request.data.copy()
    userobj=Users.objects.filter(id=id).first()
    if data['employeeId'] !="":
        checkempid_obj=Users.objects.filter(employeeId=data['employeeId'],company_code=request.user.company_code,is_active=True).exclude(id=id).first()
        if checkempid_obj:
            response_={
                    'status':0,
                    'Msg':'Employee of this Attendance Id Already Exists',
                    'data':{}
                }
            return Response(response_,status=400)
        



    if userobj:

        if data['email'] is not None and data['email'] != "":
            useremailobject = Users.objects.filter(company_code=request.user.company_code,is_active=True,email=data['email']).exclude(Q(email='')|Q(id=id)).first()
            if useremailobject is not None:
                return Response({"n": 0, "Msg": "User with this Official Email id already exist", "Status": "Failed"})
        password=request.data.get('password')
        if password is not None and password != "":
            data['Password'] = data['password']
            data['password'] = make_password(data['password'])


        
        Serializer = userUpdateSerializer(userobj,data=data,partial=True)
        if Serializer.is_valid():
            Serializer.save() 
            response_={
                'status':1,
                'Msg':'Data Saved Successfully.',
                'data':Serializer.data
            }
            
            return Response(response_,status=200)
        else:
            response_={
                    'status':0,
                    'Msg':'Data Not Saved.',
                    'data':Serializer.errors
                }
            return Response(response_,status=400)
    else:
        response_={
                'status':0,
                'Msg':'User not found.',
                'data':{}
            }
        return Response(response_,status=400)



def assign_value_based_on_length(length):
    if length >= 27:
        return 5
    if length >= 27:
        return 4.5
    elif length >= 24:
        return 4
    elif length >= 21:
        return 3.5
    elif length >= 18:
        return 3
    elif length >= 15:
        return 2.5
    elif length >= 12:
        return 2
    elif length >= 9:
        return 1.5
    elif length >= 6:
        return 1
    elif length >= 3:
        return 0.5
    else:   
        return 0











from django.db.models import F
@api_view(['POST'])
def getusedashboardata(request):
    Department = request.data.get('department')
    projectmanager =  request.data.get('projectmanager')
    location =  request.data.get('location')
    empworkstatus =  request.data.get('empworkstatus')
    clientside =  request.data.get('clientside')
    project =  request.data.get('project')
    tasktime =  request.data.get('tasktime')
    searchboxquery = request.data.get('searchboxquery')
    taskmanager = request.data.get('taskmanager')
    company_code = request.user.company_code
    current_date = date.today()
   

    userobjlistobject = Users.objects.filter(Q(is_active = True,Firstname__icontains=searchboxquery,company_code=company_code) | Q(is_active = True,Lastname__icontains=searchboxquery,company_code=company_code)).order_by('Firstname')
    firstobj = ""
    if Department is not None and Department != "": 
        firstobj = userobjlistobject.filter(DepartmentID=Department)
    else:
        firstobj = userobjlistobject
    
    secondobj = ""
    if location is not None and location != "": 
        secondobj = firstobj.filter(locationId=location)
    else:
        secondobj = firstobj

    thirdobj = ""
    if empworkstatus is not None and empworkstatus != "":
        thirdobj = secondobj.filter(typeofwork=empworkstatus)
    else:
        thirdobj = secondobj

    fourthobj = ""
    if project is not None and project != "":
        fourthobj = thirdobj.filter(id__in = list(TaskMaster.objects.filter(Project=project,Active=True,Status=1).values_list('AssignTo', flat=True)))
    else:
        fourthobj = thirdobj

    currentzone = pytz.timezone("Asia/Kolkata") 
    currenttime = datetime.datetime.now(currentzone)
    newcurrenttime = currenttime.strftime("%Y-%m-%dT%H:%M:%S.%f%z")

    fifthobj = ""
    if tasktime is not None and tasktime != "":
        if tasktime == "idle":
            userfifthpbject = fourthobj.filter(id__in = list(TaskMaster.objects.filter(AssignDate=current_date,Status__in=[2,3]).values_list('AssignTo', flat=True)))
            attendanceobject = fourthobj.filter(employeeId__in = list(attendance.objects.filter(date=current_date).values_list('employeeId', flat=True)))
            fifthobj = list(chain(userfifthpbject, attendanceobject))
        elif tasktime == "sixhours":
            fifthobj = fourthobj.filter(id__in = list(TaskMaster.objects.filter(Active=True,Status = 1).values_list('AssignTo', flat=True)))
        else:
            fifthobj = fourthobj
    else:
        fifthobj = fourthobj

    sixthobject = ""

    if projectmanager is not None and projectmanager != "":
        sixthobject = fifthobj.filter(id__in = list(TaskMaster.objects.filter(Project__in=list(ProjectMaster.objects.filter(Active=True,ProjectBA=projectmanager).values_list('id',flat=True)),Active=True).values_list('AssignTo', flat=True))) 
    
    else:
        sixthobject = fifthobj

    seventhobject = ""
    if taskmanager is not None and taskmanager != "":
        seventhobject = fourthobj.filter(id__in = list(TaskMaster.objects.filter(Active=True,Status=1,AssignBy=taskmanager).values_list('AssignTo', flat=True)))
    else:
        seventhobject = sixthobject
    
    eigthobject = ""

    if clientside is not None and clientside != "":
        eigthobject = fourthobj.filter(id__in = list(TaskMaster.objects.filter(Active=True,Status=1,AssignTo=clientside).values_list('AssignTo', flat=True)))
    else:
        eigthobject = seventhobject

    finalobject = eigthobject

    totalemployeecount = Users.objects.filter(is_active=True,company_code=company_code).count()
    totalemployeeworking = attendance.objects.filter(date=current_date,company_code=company_code).distinct('employeeId').count()
   

    userlist = []
    serializer = UsersdashboardSerializer(finalobject,many=True)
    
    for f in serializer.data :
        if f['typeofwork'] == "2":
            f['working_from'] = '<i class="fa fa-building-o" aria-hidden="true"></i>'
        elif f['typeofwork'] == "1" or f['typeofwork'] == "4":
            f['working_from'] = '<i class="fa fa-home" aria-hidden="true"></i>'
        else:
            f['working_from'] = ''
        
        if tasktime == "sixhours":
            
            taskobj = TaskMaster.objects.filter(Active=True,Status = 1,AssignTo=f['id']).first()
            taskremark = TaskRemark.objects.filter(task_id =taskobj.id, IsRead = False).first()
            if taskremark is not None:
                f['taskremark'] = False
            else:
                f['taskremark'] = True

            leaveobject = Leave.objects.filter(employeeId=f['id'],leave_status="Approved",start_date__lte=current_date,end_date__gte=current_date).first()
            if leaveobject is not None:
                
                if leaveobject.WorkFromHome == False and leaveobject.leavetype == 'Fullday':
           
                    f['leave_detail'] = 'On leave'
                elif leaveobject.WorkFromHome == False and leaveobject.leavetype == 'FirstHalf':
                    
                    f['leave_detail'] = 'First half leave'
                   
                elif leaveobject.WorkFromHome == False and leaveobject.leavetype == 'SecondHalf':
    
                    f['leave_detail'] = 'Second half leave'
                else:
                    f['leave_detail'] = ''
            else:
                f['leave_detail'] = ''
                
            projecttasks = ProjectTasks.objects.filter(Task = taskobj.id)
            projectser = ProjectTasksSerializer(projecttasks,many=True)
            totaltime=0
            for o in projectser.data:
                startstring = o['StartDate']
                starttime=startstring
                t1 = datetime.datetime.strptime(starttime, "%Y-%m-%dT%H:%M:%S.%f%z")

                endstring = o['EndDate']
                if endstring is not None:
                    endtime = o['EndDate']
                else:
                    endtime = str(newcurrenttime)
                t2=datetime.datetime.strptime(endtime, "%Y-%m-%dT%H:%M:%S.%f%z")
                tdelta=t2-t1
            
            
                if "day" in str(tdelta) or "days" in str(tdelta):
                    daystring = str(tdelta).split(",")[0]
                    noofdays = str(daystring).split(" ")[0]
                    daysmins = int(noofdays)*1440

                    thoursstr =  str(tdelta).split(",")[1]
                    thours = str(thoursstr).split(":")[0]
                    hrs = int(thours)*60
                    tmins = str(thoursstr).split(":")[1]
                    finalmins = int(hrs)+int(tmins)+int(daysmins)
                else:
                    thours = str(tdelta).split(":")[0]
                    hrs = int(thours)*60
                    tmins = str(tdelta).split(":")[1]
                    finalmins = int(hrs)+int(tmins)
                totaltime += finalmins
            
            f['task_detail'] = taskobj.TaskTitle
            totalhours =totaltime
            hour = int (totalhours) // 60
            if (len(str(hour)) < 2):
                hours = "0"+str(hour)
            else:
                hours = str(hour)
            
            mins = int (totalhours) % 60
            if (len(str(mins)) < 2):
                minutes = "0"+str(mins)
            else:
                minutes = str(mins)

            if int(hours) > 6 :
                f['user_status'] = '<i class="fa fa-circle green_status" aria-hidden="true"></i>'
                f['task_time'] = str(hours) + ":" + str(minutes) + " hrs"
                f['project_name'] = taskobj.ProjectName
                f['exceed_time'] = '<i class="fa fa-clock-o" aria-hidden="true"></i>'
                f['sort'] = 2
                userlist.append(f)
                
                
        else:
            taskobj = TaskMaster.objects.filter(Active=True,Status = 1,AssignTo=f['id']).first()


       

                
                   


            
            
            attendance_object_count = attendance.objects.filter(employeeId=f['employeeId'],date=current_date).count()
            if attendance_object_count >= 1 and taskobj is not None:
                f['user_status'] = '<i class="fa fa-circle present_status" aria-hidden="true"></i>'
                f['sort'] = 1
            elif attendance_object_count == 0 and taskobj is not None: 
                f['user_status'] = '<i class="fa fa-circle present_status" aria-hidden="true"></i>'
                f['sort'] = 1
            elif attendance_object_count >=1 and taskobj is None:
                f['user_status'] = '<i class="fa fa-circle green_status" aria-hidden="true"></i>'
                f['sort'] = 2 
            else:
                f['user_status'] = '<i class="fa fa-circle offline_status" aria-hidden="true"></i>'
                f['sort'] = 3
            
            if taskobj is not None:
                taskremark = TaskRemark.objects.filter(task_id =taskobj.id, IsRead = False).first()
                if taskremark is not None:
                    f['taskremark'] = False
                else:
                    f['taskremark'] = True
                    
                f['taskid'] = taskobj.id
                if taskobj.Zone == 1:
                    greenstatusstr = "<img src='/static/Media/taskicons/activegreenpoints.svg' class='activeicons' alt='Paris'>"
                else:
                    greenstatusstr = "<img src='/static/Media/taskicons/nongreen.svg' id='1' class='nonactive' alt='Paris'>"

                if taskobj.Zone == 2:
                    yellowstatusstr = "<img src='/static/Media/taskicons/activeyellowpoints.svg' class='activeicons' alt='Paris'>"
                else:
                    yellowstatusstr = "<img src='/static/Media/taskicons/yellow.svg' id='2' class='nonactive' alt='Paris'>"

                if taskobj.Zone == 3:
                    redstatusstr = "<img src='/static/Media/taskicons/activeredpoints.svg' class='activeicons' alt='Paris' >"
                else:
                    redstatusstr = "<img src='/static/Media/taskicons/red.svg' id='3' class='nonactive' alt='Paris'>"

                if taskobj.Zone == 4:
                    notdonestr = "<img src='/static/Media/taskicons/activenotdonepoints.svg' class='activeicons' alt='Paris'>"
                else:
                    notdonestr = "<img src='/static/Media/taskicons/notdone.svg' id='4' class='nonactive' alt='Paris' >"

                if taskobj.Zone == 5:
                    cancelledstr = "<img src='/static/Media/taskicons/activecancelledpoints.svg' class='activeicons' alt='Paris'>"
                else:
                    cancelledstr = "<img src='/static/Media/taskicons/cancelled.svg' id='5' class='nonactive' alt='Paris'>"

                if taskobj.Zone == 6:
                    rejectedstr = "<img src='/static/Media/taskicons/activerejectpoints.svg' class='activeicons' alt='Paris'>"
                else:
                    rejectedstr = "<img src='/static/Media/taskicons/rejected.svg' id='6' class='nonactive' alt='Paris'>"

              
                f['allzonestr'] = {
                    'greenstatusstr':greenstatusstr,
                    'yellowstatusstr':yellowstatusstr,
                    'redstatusstr':redstatusstr,
                    'notdonestr':notdonestr,
                    'cancelledstr':cancelledstr,
                    'rejectedstr':rejectedstr,
                }
                
                projecttasks = ProjectTasks.objects.filter(Task = taskobj.id)
                projectser = ProjectTasksSerializer(projecttasks,many=True)
                totaltime=0
                for o in projectser.data:
                    startstring = o['StartDate']
                    starttime=startstring
                    t1 = datetime.datetime.strptime(starttime, "%Y-%m-%dT%H:%M:%S.%f%z")

                    endstring = o['EndDate']
                    if endstring is not None:
                        endtime = o['EndDate']
                    else:
                        endtime = str(newcurrenttime)
                    t2=datetime.datetime.strptime(endtime, "%Y-%m-%dT%H:%M:%S.%f%z")
                    tdelta=t2-t1
                
                
                    if "day" in str(tdelta) or "days" in str(tdelta):
                        daystring = str(tdelta).split(",")[0]
                        noofdays = str(daystring).split(" ")[0]
                        daysmins = int(noofdays)*1440

                        thoursstr =  str(tdelta).split(",")[1]
                        thours = str(thoursstr).split(":")[0]
                        hrs = int(thours)*60
                        tmins = str(thoursstr).split(":")[1]
                        finalmins = int(hrs)+int(tmins)+int(daysmins)
                    else:
                        thours = str(tdelta).split(":")[0]
                        hrs = int(thours)*60
                        tmins = str(tdelta).split(":")[1]
                        finalmins = int(hrs)+int(tmins)
                    totaltime += finalmins
                
                f['task_detail'] = taskobj.TaskTitle
                totalhours =totaltime
                hour = int (totalhours) // 60
                if (len(str(hour)) < 2):
                    hours = "0"+str(hour)
                else:
                    hours = str(hour)
                
                mins = int (totalhours) % 60
                if (len(str(mins)) < 2):
                    minutes = "0"+str(mins)
                else:
                    minutes = str(mins)
                
                f['project_name'] = taskobj.ProjectName
                f['task_time'] = str(hours) + ":" + str(minutes) + " hrs"
                if int(hours) > 6 :
                    f['exceed_time'] = '<i class="fa fa-clock-o" aria-hidden="true"></i>'
                else:
                    f['exceed_time'] = ''
                
            else:
                f['task_detail'] = '--'
                f['project_name'] = "--"
                f['task_time'] = "--:--"    
                f['exceed_time'] = ''
                f['taskid'] = ''
               
                f['allzonestr'] = {
                    'greenstatusstr':'',
                    'yellowstatusstr':'',
                    'redstatusstr':'',
                    'notdonestr':'',
                    'cancelledstr':'',
                    'rejectedstr':'',
                }
            
            leaveobject = Leave.objects.filter(employeeId=f['id'],leave_status="Approved",start_date__lte=current_date,end_date__gte=current_date).first()
            if leaveobject is not None:
                
                if leaveobject.WorkFromHome == False and leaveobject.leavetype == 'Fullday':
                    f['user_status'] = '<i class="fa fa-circle leave_status" aria-hidden="true"></i>'

                    f['leave_detail'] = 'On leave'
                elif leaveobject.WorkFromHome == False and leaveobject.leavetype == 'FirstHalf':
                    f['user_status'] = '<i class="fa fa-circle leave_status" aria-hidden="true"></i>'

                    f['leave_detail'] = 'First half leave'
                   
                elif leaveobject.WorkFromHome == False and leaveobject.leavetype == 'SecondHalf':
                    f['user_status'] = '<i class="fa fa-circle leave_status" aria-hidden="true"></i>'

                    f['leave_detail'] = 'Second half leave'
                else:
                    f['leave_detail'] = ''
            else:
                f['leave_detail'] = ''
                
                
            userlist.append(f)     
    
    newlist = sorted(userlist, key=itemgetter('sort')) 
    # for i in serializer.data: sirf order by karna hai abhi
    # split_idx = bisect.bisect(sorted_list, n)
    # under = sorted_list[:split_idx]
    # over = sorted_list[split_idx:]
    # alluserlist = userlist
    # firsthalflist = A[:len(alluserlist)//2]
    # secondhalflist = A[len(alluserlist)//2:]
    midpoint = len(newlist) // 2
    # Split the list into two halves
    first_half = newlist[:midpoint+1]
    second_half = newlist[midpoint+1:]


    return Response ({"data":newlist,"firsthalf":first_half,"secondhalf":second_half,"totalemployeecount":totalemployeecount,"totalemployeeworking":totalemployeeworking,"response":{"n" : 1,"msg" : "SUCCESS","status" : "success"}})


    



    


               

@api_view(['GET'])    
@permission_classes((AllowAny,))
def Attendance_mangreminder(request):
    companycode = "O001"
    attendancereqobj = list(AttendanceRequest.objects.filter(Action__isnull=True,Active=True).distinct('manager_ids').values_list('manager_ids', flat=True))

    managerlist = attendancereqobj
  
    for m in managerlist:
        mangcount = AttendanceRequest.objects.filter(Action__isnull=True,Active=True,manager_ids=m).count()
        managername_obj= Users.objects.filter(id=int(m),is_active=True).first()
        try:             
            dicti = {
                        "requestcount":mangcount,
                    }
            message = get_template(
                'attendancerequestmail.html').render(dicti)
            
            msg = EmailMessage(
                'Urgent: Action Required on Pending Attendance Requests',
                message,
                EMAIL_HOST_USER,
                [managername_obj.email],
            )
            msg.content_subtype = "html"  # Main content is now text/html
            msg.send()
        except Exception as e:
            print('exception occured fot mail', e)

    return Response ({"data":'',"response":{"n" : 1,"msg" : "success","status" : "success"}})


@api_view(['POST'])
def search_by_name_employee(request, format=None):

    if request.method == 'POST':
        username=request.POST.get('username')
        if username != "":
            company_code = request.user.company_code
            user = Users.objects.filter(Q(Firstname__icontains = username,is_active=True,company_code=company_code)|Q(Lastname__icontains = username,is_active=True,company_code=company_code)).order_by('id')[:5]
            serializer = UsersSerializer(user, many=True)
            newlist=[]
            if username !="" and username is not None:
                for i in serializer.data:
                    i['Fullname'] = str(i['Firstname'] + " " + i['Lastname']).strip()
                    if str(username.lower()) in str(i['Fullname']).lower():
                        newlist.append(i)
            else:
                newlist=serializer.data
        else:
            newlist = []

        return Response({'n':1,'msg':'Employee list fetched successfully','status':'success','data':newlist})
   

@api_view(['POST'])
def search_emp_by_id(request, format=None):

    if request.method == 'POST':
        # user = Users.objects.all().order_by('CreatedOn')
        managerid=request.POST.get('managerid')
        company_code = request.user.company_code
        user = Users.objects.filter(
            is_active=True,company_code=company_code,id=managerid).first()
        serializer = UsersSerializer(user)


        return Response({'n':1,'msg':'Employee list fetched successfully','status':'success','data':serializer.data})
   



           
        

@api_view(['GET'])
def checksession(request):
    if request.method == 'GET':
        return Response({'detail': 'valid token.'})

@api_view(['POST'])
@permission_classes((AllowAny,))
def add_device_change_request(request):
    if request.method == 'POST':
        data={}
        data['userid']=request.POST.get('userid')
        data['unique_device_id']=request.POST.get('unique_device_id')
        data['device_name']=request.POST.get('device_name')
        data['remark']=request.POST.get('remark')
        data['app_version']=request.POST.get('app_version')
        userobj=Users.objects.filter(id=data['userid'],is_active=True).first()
        if userobj is not None:
            data['employee_code']=userobj.uid
        else:
            return Response({'n':0,'msg':'user of this user id not found','status':'failed','data':data})
        alredy_exist_obj=DeviceChangeRequest.objects.filter(userid=data['userid'],
                                                            unique_device_id=data['unique_device_id'],
                                                            device_name=data['device_name'],
                                                            app_version=data['app_version'],
                                                            is_active=True,status="Pending").first()
        if alredy_exist_obj is not None:
            return Response({'n':0,'msg':'This device request is already exist','status':'failed','data':data})

        serializer = DeviceChangeRequestSerializer(data=data,partial=False)
        if serializer.is_valid():
            serializer.save()
            return Response({'n':1,'msg':'device change request added','status':'success','data':serializer.data})
        else:
            return Response({'n':0,'msg':'device change request not added','status':'failed','data':serializer.errors})


@api_view(['GET'])
def devicechangerequestslist(request):
    dev_obj=DeviceChangeRequest.objects.filter(is_active=True).order_by('-created_at')
    if dev_obj is not None:
        serializer = DeviceChangeRequestSerializer(dev_obj,many=True)
        for i in serializer.data:
            user_obj=Users.objects.filter(id=i['userid'],is_active=True).first()
            if user_obj:
               i['applicant_name'] = user_obj.Firstname +' '+ user_obj.Lastname
            else:
                i['applicant_name'] = 'NA'

            i['created_at'] = conver_created_at_date(i['created_at'])

        return Response({'n':1,'msg':'device change request found','status':'success','data':serializer.data})
    else:
        return Response({'n':0,'msg':'device change request not found','status':'failed','data':[]})

@api_view(['GET'])
def pendingdevicechangerequestslist(request):
    dev_obj=DeviceChangeRequest.objects.filter(is_active=True,status='Pending').order_by('-created_at')
    if dev_obj is not None:
        serializer = DeviceChangeRequestSerializer(dev_obj,many=True)
        for i in serializer.data:
            user_obj=Users.objects.filter(id=i['userid'],is_active=True).first()
            if user_obj:
               i['applicant_name'] = user_obj.Firstname +' '+ user_obj.Lastname
            else:
                i['applicant_name'] = 'NA'
            i['created_at'] = conver_created_at_date(i['created_at'])

        return Response({'n':1,'msg':'device change request found','status':'success','data':serializer.data})
    else:
        return Response({'n':0,'msg':'device change request not found','status':'failed','data':[]})

@api_view(['GET'])
def approveddevicechangerequestslist(request):
    dev_obj=DeviceChangeRequest.objects.filter(is_active=True,status='Approved').order_by('-created_at')
    if dev_obj is not None:
        serializer = DeviceChangeRequestSerializer(dev_obj,many=True)
        for i in serializer.data:
            user_obj=Users.objects.filter(id=i['userid'],is_active=True).first()
            if user_obj:
               i['applicant_name'] = user_obj.Firstname +' '+ user_obj.Lastname
            else:
                i['applicant_name'] = 'NA'
            i['created_at'] = conver_created_at_date(i['created_at'])

        return Response({'n':1,'msg':'device change request found','status':'success','data':serializer.data})
    else:
        return Response({'n':0,'msg':'device change request not found','status':'failed','data':[]})

@api_view(['GET'])
def rejecteddevicechangerequestslist(request):
    dev_obj=DeviceChangeRequest.objects.filter(is_active=True,status='Rejected').order_by('-created_at')
    if dev_obj is not None:
        serializer = DeviceChangeRequestSerializer(dev_obj,many=True)
        for i in serializer.data:
            user_obj=Users.objects.filter(id=i['userid'],is_active=True).first()
            if user_obj:
               i['applicant_name'] = user_obj.Firstname +' '+ user_obj.Lastname
            else:
                i['applicant_name'] = 'NA'
            i['created_at'] = conver_created_at_date(i['created_at'])

        return Response({'n':1,'msg':'device change request found','status':'success','data':serializer.data})
    else:
        return Response({'n':0,'msg':'device change request not found','status':'failed','data':[]})

@api_view(['POST'])
def approvedevicechangerequest(request):
    data={}
    id = request.data.get('id')
    action_conditional_obj = DeviceChangeRequest.objects.filter(id=id,is_active=True).first()
    if action_conditional_obj is not None:
        if action_conditional_obj.status == "Approved":
            return Response ({"data":{},"response":{"n" : 2,"msg" : "This application is already approved you cannot perform actions","status" : "failed"}})
        elif action_conditional_obj.status == "Rejected":    
            return Response ({"data":{},"response":{"n" : 2,"msg" : "This application is already rejected you cannot perform actions","status" : "failed"}})
    else:
        return Response ({"data":{},"response":{"n" : 2,"msg" : "Application not found","status" : "failed"}})
    
    data['status']="Approved"

    application_obj = DeviceChangeRequest.objects.filter(id=id,is_active=True).first()
    if application_obj:
        # remove_other_device__request_object=DeviceChangeRequest.objects.filter(userid=application_obj.userid,is_active=True).exclude(id=id).update(is_active=False)
        aserializer = DeviceChangeRequestSerializer(application_obj,data=data,partial=True)
        if aserializer.is_valid():
            aserializer.save()
            remove_previous_device_object = DeviceVerification.objects.filter(userid=aserializer.data['userid'],is_active=True).update(is_active=False)
            add_new_device_object = DeviceVerification.objects.create(device_name=aserializer.data['device_name'],app_version=aserializer.data['app_version'],unique_device_id=aserializer.data['unique_device_id'],employee_code=aserializer.data['employee_code'],userid=aserializer.data['userid'])
            user_obj=Users.objects.filter(id=aserializer.data['userid'],is_active=True).first()
            if user_obj is not None:
                data_dict = {
                            "employeename":user_obj.Firstname +' '+user_obj.Lastname,
                            "device_name":aserializer.data['device_name']
                        }
                send_async_custom_template_email(
                    'Approval of Device Change Request',
                    data_dict,
                    "no-reply@onerooftech.com",
                    [str(user_obj.email)],
                    "device_change_request_mails/devicechangerequestapprovelmail.html"
                )
                return Response ({"data":aserializer.data,"response":{"n" : 1,"msg" : "Device change request is approved","status" : "success"}})

            else:
                return Response ({"data":{},"response":{"n" : 2,"msg" : "user not found","status" : "error"}})
        else:
            return Response ({"data":{},"response":{"n" : 2,"msg" : "serializer error " + aserializer.errors,"status" : "error"}})
    else:

        return Response ({"data":{},"response":{"n" : 2,"msg" : "application not found","status" : "error"}})

@api_view(['POST'])
def reject_devicechangerequest(request):
    data={}
    id = request.data.get('id')
    action_conditional_obj = DeviceChangeRequest.objects.filter(id=id,is_active=True).first()
    if action_conditional_obj is not None:
        if action_conditional_obj.status == "Approved":
            return Response ({"data":{},"response":{"n" : 2,"msg" : "This application is already approved you cannot perform actions","status" : "failed"}})
        elif action_conditional_obj.status == "Rejected":    
            return Response ({"data":{},"response":{"n" : 2,"msg" : "This application is already rejected you cannot perform actions","status" : "failed"}})
    else:
        return Response ({"data":{},"response":{"n" : 2,"msg" : "Application with this application id  not found","status" : "failed"}})
    
    data['status']="Rejected"
    data['rejection_reason']=request.data.get('rejection_reason')

    application_obj = DeviceChangeRequest.objects.filter(id=id,is_active=True).first()
    if application_obj:
        aserializer = DeviceChangeRequestSerializer(application_obj,data=data,partial=True)
        if aserializer.is_valid():
            aserializer.save()
            user_obj=Users.objects.filter(id=aserializer.data['userid'],is_active=True).first()
            if user_obj is not None:
                data_dict = {
                            "employeename":user_obj.Firstname +' '+user_obj.Lastname,
                            "device_name":aserializer.data['device_name'],
                            "rejection_reason":aserializer.data['rejection_reason']
                        }
                send_async_custom_template_email(
                    ' Device Change Request Rejection',
                    data_dict,
                    "no-reply@onerooftech.com",
                    [str(user_obj.email)],
                    "device_change_request_mails/devicechangerequestrejectedmail.html"
                )
                return Response ({"data":aserializer.data,"response":{"n" : 1,"msg" : "Device change request is approved","status" : "success"}})
            else:
                return Response ({"data":{},"response":{"n" : 2,"msg" : "user not found","status" : "error"}})
        else:
            return Response ({"data":{},"response":{"n" : 2,"msg" : "serializer error " + aserializer.errors,"status" : "error"}})
    else:
        return Response ({"data":{},"response":{"n" : 2,"msg" : "application not found","status" : "error"}})






























@api_view(['POST'])
def add_shift(request):
    if request.method == 'POST':
        data={}
        data['shiftname']=request.POST.get('shiftname')
        data['intime']=request.POST.get('intime')
        data['outtime']=request.POST.get('outtime')
        alredyexist_name=ShiftMaster.objects.filter(shiftname=data['shiftname'],is_active=True).first()
        if alredyexist_name is not None:
            return Response({'n':0,'msg':'Shift with this Name is already exist.','status':'failed','data':data})
        alredyexist_time=ShiftMaster.objects.filter(intime=data['intime'],outtime=data['outtime'],is_active=True).exclude(intime='00:00',outtime='00:00').first()
        if alredyexist_time is not None:
            return Response({'n':0,'msg':'Shift with this timing  is already exist.','status':'failed','data':data})
        
        serializer = ShiftMasterSerializer(data=data,partial=False)
        if serializer.is_valid():
            serializer.save()
            return Response({'n':1,'msg':'New shift added','status':'success','data':serializer.data})
        else:
            return Response({'n':0,'msg':'New shift not added','status':'failed','data':serializer.errors})




@api_view(['GET'])
def get_all_shifts(request):
    shift_obj=ShiftMaster.objects.filter(is_active=True)
    if shift_obj is not None:
        serializer = ShiftMasterSerializer(shift_obj,many=True)
        return Response({'n':1,'msg':' Shift found','status':'success','data':serializer.data})
    else:
        return Response({'n':0,'msg':' Shift not added','status':'failed','data':[]})




@api_view(['POST'])
def update_shift(request):
    if request.method == 'POST':
        data={}
        data['id']=request.POST.get('id')
        data['shiftname']=request.POST.get('shiftname')
        data['intime']=request.POST.get('intime')
        data['outtime']=request.POST.get('outtime')
        shift_obj=ShiftMaster.objects.filter(id=data['id'],is_active=True).first()
        if shift_obj:
            already_obj=ShiftMaster.objects.filter(is_active=True,shiftname=data['shiftname']).exclude(id=data['id']).first()
            if already_obj:
                return Response({'n':0,'msg':'Shift name already exist','status':'failed','data':[]})

            alredyexist_time=ShiftMaster.objects.filter(intime=data['intime'],outtime=data['outtime'],is_active=True).exclude(id=data['id']).first()
            if alredyexist_time is not None:
                return Response({'n':0,'msg':'Shift with this timing  is already exist.','status':'failed','data':data}) 
                       
            serializer = ShiftMasterSerializer(shift_obj,data=data,partial=False)
            if serializer.is_valid():
                serializer.save()
                return Response({'n':1,'msg':'Shift updated successfully','status':'success','data':serializer.data})
            else:
                return Response({'n':0,'msg':'Unable to update shift','status':'failed','data':serializer.errors})
        return Response({'n':0,'msg':' Shift not found','status':'failed','data':[]})





@api_view(['POST'])
def delete_shift(request):
    if request.method == 'POST':
        data={}
        data['id']=request.POST.get('id')
        data['is_active']=False
        shift_obj=ShiftMaster.objects.filter(id=data['id'],is_active=True).first()
        if shift_obj:
            serializer = ShiftMasterSerializer(shift_obj,data=data,partial=False)
            if serializer.is_valid():
                serializer.save()
                return Response({'n':1,'msg':'Shift deleted successfully','status':'success','data':serializer.data})
            else:
                return Response({'n':0,'msg':' Unable to delete shift','status':'failed','data':serializer.errors})
        return Response({'n':0,'msg':' Shift not found','status':'failed','data':[]})





@api_view(['POST'])
def get_shift_details(request):
    if request.method == 'POST':
        data={}
        data['id']=request.POST.get('id')
        shift_obj=ShiftMaster.objects.filter(id=data['id'],is_active=True).first()
        if shift_obj:
            serializer = ShiftMasterSerializer(shift_obj)

            return Response({'n':1,'msg':'Shift found successfully','status':'success','data':serializer.data})
        return Response({'n':0,'msg':' Shift not found','status':'failed','data':[]})



@api_view(['GET'])
def get_all_empshiftsdetails(request):
    shift_obj=EmployeeShiftDetails.objects.filter(is_active=True)
    if shift_obj is not None:
        serializer = EmployeeShiftDetailsSerializer(shift_obj,many=True)
        newlist=[]
        for i in serializer.data:
            user_obj=Users.objects.filter(id=i['employeeId'],is_active=True).first()
            if user_obj is not None:
                i['employee_name']=user_obj.Firstname +" "+user_obj.Lastname
                newlist.append(i)

        return Response({'n':1,'msg':'EMP Shift found','status':'success','data':serializer.data})
    else:
        return Response({'n':0,'msg':'EMP Shift not added','status':'failed','data':[]})

@api_view(['POST'])
def add_empshiftdetails(request):
    if request.method == 'POST':
        data={}
        data['employee_name']=request.POST.get('employee_name')
        data['employeeId']=request.POST.get('employeeId')
        already_exist=EmployeeShiftDetails.objects.filter(employeeId=data['employeeId'],is_active=True).first()
        if already_exist is not None:
            return Response({'n':0,'msg':'employee already exists in list','status':'failed','data':[]})


        serializer = EmployeeShiftDetailsSerializer(data=data,partial=False)
        if serializer.is_valid():
            serializer.save()
            return Response({'n':1,'msg':'New employee shift details added','status':'success','data':serializer.data})
        else:
            return Response({'n':0,'msg':'New employee shift details not added','status':'failed','data':serializer.errors})


@api_view(['POST'])
def delete_empshiftdetails(request):
    if request.method == 'POST':
        data={}
        data['id']=request.POST.get('id')
        data['is_active']=False
        shift_obj=EmployeeShiftDetails.objects.filter(id=data['id'],is_active=True).first()
        if shift_obj:
            serializer = EmployeeShiftDetailsSerializer(shift_obj,data=data,partial=False)
            if serializer.is_valid():
                serializer.save()
                return Response({'n':1,'msg':'employee Shift deleted successfully','status':'success','data':serializer.data})
            else:
                return Response({'n':0,'msg':' Unable to delete employee shift','status':'failed','data':serializer.errors})
        return Response({'n':0,'msg':' employee Shift not found','status':'failed','data':[]})



@api_view(['POST'])
def add_empshiftallotment(request):
    if request.method == 'POST':
        data=request.POST.copy()

        dates=get_range_of_dates(data["start_date"],data["end_date"])
        allotedshiftes=[]
        nonallotedshiftes=[]
        for i in dates:
            data['date']=i
            data['is_active']=True
            user_obj=Users.objects.filter(id=data['employeeId']).first()
            if user_obj is not None:
                if user_obj.employeeId is not None and user_obj.employeeId !="":
                    data['attendanceId']=user_obj.employeeId
                else:
                    return Response({'n':0,'msg':'User attendance id not found','status':'failed','data':[]})
            else:
                return Response({'n':0,'msg':'User not found','status':'failed','data':[]})
            
            already_exist=ShiftAllotment.objects.filter(employeeId=data['employeeId'],is_active=True,date=i,shiftId=data['shiftId']).first()
            if already_exist is not None:
                continue
                # return Response({'n':0,'msg':'employee shift already exist for this day','status':'failed','data':[]})
            else:
                serializer = ShiftAllotmentSerializer(data=data,partial=False)
            if serializer.is_valid():
                serializer.save()
                allotedshiftes.append(serializer.data)
            else:
                nonallotedshiftes.append(i)
        return Response({'n':1,'msg':'employee shift alloted successfully','status':'success','data':{"allotedshiftes":allotedshiftes,"nonallotedshiftes":nonallotedshiftes}})
    else:
        return Response({'n':0,'msg':'employee shift  not alloted','status':'failed','data':[]})
















def is_valid_date_yyyy_mm_dd(date_string):
    try:
        # Attempt to parse the date string
        datetime.datetime.strptime(date_string, '%Y-%m-%d %H:%M:%S')
        return True
    except ValueError:
        # Raised if the format is incorrect or an invalid date
        return False





@api_view(['POST'])
def bulkuploadshiftallotment(request):
    fileerrorlist=[]
    successlist=[]
    allotedshiftes=[]
    nonallotedshiftes=[]

    if 'file' in request.FILES and request.FILES['file'] is not None and request.FILES['file'] != "":
        file = request.FILES['file']

        dataset = Dataset()
        namefile = file.name
        split_tup = os.path.splitext(namefile)
        file_extension = split_tup[1]
        company_code=request.user.company_code
    
        if file_extension == ".xls" or file_extension == ".xlsx":
        
            imported_data = dataset.load(file.read(), format='xlsx')
            count=0
            for i in imported_data:
               
                count+=1
                if i[0] is not None and i[0] !="":
                    data={}
                    data['employee_name'] = str(i[0]).lower()
                    employeeobj=EmployeeShiftDetails.objects.filter(employee_name__icontains = data['employee_name'],is_active=True).first()
                    if employeeobj is not None:
                        user_obj = Users.objects.filter(id = employeeobj.employeeId,is_active=True,company_code=company_code).first()
                        if user_obj is not None:
                            if user_obj.employeeId !="" and user_obj.employeeId is not None:
                                data['employeeId']=user_obj.id
                                data['attendanceId']=user_obj.employeeId
                                if i[1] is not None and i[1] !="":

                                    data['shift_name'] = str(i[1]).lower()
                                    shift_obj=ShiftMaster.objects.filter(shiftname__icontains = data['shift_name'],is_active=True).first()
                                    if shift_obj is not None:
                                        data['shiftId']=shift_obj.id
                                        if i[2] is not None and i[2] != "":
                                            if is_valid_date_yyyy_mm_dd(str(i[2])):
                                                dt_object = datetime.datetime.strptime(str(i[2]), '%Y-%m-%d %H:%M:%S')
                                                # 01/02/20203
                                                # Format the datetime object as a date string
                                                date_string = dt_object.strftime('%Y-%m-%d')
                                                data['date'] = date_string


                                                successlist.append(data)

                                            else:
                                                entrydict={
                                                    "remark":"The date is not in valid date format YYYY-MM-DD",
                                                    "Entry":i,
                                                    "srno":count,
                                                }
                                                fileerrorlist.append(entrydict)
                                                continue

                                        else:
                                            entrydict={
                                                "remark":"date is required",
                                                "Entry":i,
                                                "srno":count,
                                            }
                                            fileerrorlist.append(entrydict)
                                            continue
                                    else:
                                        entrydict={
                                            "remark":"shift with this shift "+ str(data['shift_name'])+ " name not found",
                                            "Entry":i,
                                            "srno":count,
                                        }
                                        fileerrorlist.append(entrydict)
                                        continue
                                else:
                                    entrydict={
                                        "remark":"shift name is required",
                                        "Entry":i,
                                        "srno":count,
                                    }
                                    fileerrorlist.append(entrydict)
                                    continue
                            else:
                                entrydict={
                                    "remark":str(data['employee_name']) +" employee does not have attendance id",
                                    "Entry":i,
                                    "srno":count,
                                }
                                fileerrorlist.append(entrydict)
                                continue
                        else:
                            entrydict={
                                "remark":"This employee is no longer working",
                                "Entry":i,
                                "srno":count,
                            }
                            fileerrorlist.append(entrydict)
                            continue
                    else:
                        entrydict={
                            "remark":"employee with this "+ str(data['employee_name']) +" name dosen't exists",
                            "Entry":i,
                            "srno":count,
                        }
                        fileerrorlist.append(entrydict)
                        continue

                # else:
                #     entrydict={
                #         "remark":"employee name is required",
                #         "Entry":i,
                #         "srno":count,
                #     }
                #     fileerrorlist.append(entrydict)
                #     continue




            if len(fileerrorlist) == 0:
                for i in successlist:
                    already_exist=ShiftAllotment.objects.filter(employeeId=i['employeeId'],is_active=True,date=i['date'],shiftId=i['shiftId']).first()
                    if already_exist is not None:
                        serializer = ShiftAllotmentSerializer(already_exist,data=i,partial=False)
                    else:
                        serializer = ShiftAllotmentSerializer(data=i,partial=False)
                    if serializer.is_valid():
                        serializer.save()
                        allotedshiftes.append(serializer.data)
                    else:                     
                        nonallotedshiftes.append(i)

                return Response({"response":{"n": 1 ,"data":{"errorlist":fileerrorlist,'successlist':successlist,"nonallotedshiftes":nonallotedshiftes,'allotedshiftes':allotedshiftes},"msg" : "File uploaded","status":"success"}})  
            else:
                if len(fileerrorlist) > 1:
                    extra_string=str(fileerrorlist[0]['remark'] +" and " + str(len(fileerrorlist)-1) +" more. ")
                else:
                    extra_string=str(fileerrorlist[0]['remark'])


                return Response({"response":{"n": 0 ,"data":{"errorlist":fileerrorlist,'successlist':successlist,"nonallotedshiftes":nonallotedshiftes,'allotedshiftes':allotedshiftes},"msg" : "Unable to allot shift ,Reason : "+extra_string ,"status":"error"}})  


        else:
            return Response({"response":{"n": 0 ,"msg" : "File format is not excel file","status":"error"}})  

    else:
        return Response({"response":{"n": 0 ,"msg" : "  excel file is requird","status":"error"}})  




@api_view(['POST'])
def delete_empshiftallotment(request):
    if request.method == 'POST':
        data={}
        data['id']=request.POST.get('id')
        data['is_active']=False
        shift_obj=ShiftAllotment.objects.filter(id=data['id'],is_active=True).first()
        if shift_obj:
            serializer = ShiftAllotmentSerializer(shift_obj,data=data,partial=False)
            if serializer.is_valid():
                serializer.save()
                return Response({'n':1,'msg':'Employee alloted shift deleted successfully','status':'success','data':serializer.data})
            else:
                return Response({'n':0,'msg':' Unable to delete employee alloted shift','status':'failed','data':serializer.errors})
        return Response({'n':0,'msg':' employee Shift not found','status':'failed','data':[]})



@api_view(['GET'])
def get_all_alloted_shifts(request):
    shift_obj=ShiftAllotment.objects.filter(is_active=True)
    if shift_obj is not None:
        serializer = ShiftAllotmentSerializer(shift_obj,many=True)
        newlist=serializer.data
        for i in newlist:
            user_obj=Users.objects.filter(id=i['employeeId'],is_active=True).first()
            if user_obj is not None:
                i['employee_name']=user_obj.Firstname +" "+user_obj.Lastname
            i['date']=convertdate(i['date'])
        return Response({'n':1,'msg':'EMP Shift found','status':'success','data':serializer.data})
    else:
        return Response({'n':0,'msg':'EMP Shift not added','status':'failed','data':[]})



@api_view(['POST'])
def alloted_shifts_by_date(request):
    context = []
    search_date=request.POST.get('search_date')
    if search_date is not None and search_date != "":
        shiftsobj = ShiftMaster.objects.filter(is_active=True).order_by('intime')
        if shiftsobj.exists():
            shiftser = ShiftMasterSerializer(shiftsobj,many=True)
            for s in shiftser.data:
                shift_obj=ShiftAllotment.objects.filter(shiftId=s['id'],is_active=True,date=search_date)
                if shift_obj is not None:
                    serializer = ShiftAllotmentSerializer(shift_obj,many=True)
                    newlist=serializer.data
                    for i in newlist:
                        user_obj=Users.objects.filter(id=i['employeeId'],is_active=True).first()
                        if user_obj is not None:
                            userser = UserSerializer(user_obj)
                            locationid = userser.data['locationId']
                            if locationid is not None and locationid != "":
                                locationobj = Location.objects.filter(id=int(locationid)).first()
                                if locationobj is not None:
                                    i['location'] = locationobj.LocationName
                                else:
                                    i['location'] = ""
                            else:
                                i['location'] = ""
                            i['employee_name']=user_obj.Firstname +" "+user_obj.Lastname
                        i['date']=convertdate(i['date'])
                    data={
                        
                        'shiftname':s['shiftname'],
                        'shifttime':s['intime']+"-"+s['outtime'],
                        'employees':serializer.data,
                    }
                    context.append(data)
            return Response({"data": context,"response":{'n':1,'msg':'EMP Shift found successfully','status':'success'}})
        else:
            return Response({"data": '',"response":{'n':0,'msg':'shifts are not defined','status':'Failed'}})
    else:
        return Response({"data": '',"response":{'n':0,'msg':'Please enter correct date format','status':'Failed'}})


@api_view(['GET'])
def get_all_shifts_employees(request):
    shift_obj=EmployeeShiftDetails.objects.filter(is_active=True).order_by('employee_name')
    if shift_obj is not None:
        serializer = EmployeeShiftDetailsSerializer(shift_obj,many=True)
        newlist=serializer.data
        for i in newlist:
            user_obj=Users.objects.filter(id=i['employeeId'],is_active=True).first()
            if user_obj is not None:
                i['employee_name']=user_obj.Firstname +" "+user_obj.Lastname

        
        return Response({'n':1,'msg':'EMP Shift found','status':'success','data':newlist})
    else:
        return Response({'n':0,'msg':'EMP Shift not added','status':'failed','data':[]})






class paginationshiftallotmentlist(GenericAPIView):

    pagination_class = CustomPagination

    def post(self,request):
        shiftallotmentname = request.POST.get('shiftallotmentname')
        employee_name = request.POST.get('employee_name')
        searchdate = request.POST.get('searchdate')


        # Assuming today is the current date
        today = datetime.date.today()

        shiftallotment = ShiftAllotment.objects.filter(is_active=True).annotate(
            is_today=Case(
                When(date=today, then=Value(1)),
                default=Value(0),
                output_field=IntegerField(),
            )
        ).order_by('-is_today', 'date')


        # shiftallotment = ShiftAllotment.objects.filter(is_active=True).order_by('date')
        if shiftallotmentname is not None and shiftallotmentname !='':
            shiftallotment=shiftallotment.filter(shift_name__icontains=shiftallotmentname)
        if employee_name is not None and employee_name !='':
            shiftallotment=shiftallotment.filter(employee_name__icontains=employee_name)
        if searchdate is not None and searchdate !='':
            shiftallotment=shiftallotment.filter(date=searchdate)
            
        if shiftallotment is not None:
            page4 = self.paginate_queryset(shiftallotment)
            serializer = ShiftAllotmentSerializer(page4, many=True)
            # grouped_shifts = group_shifts(ShiftAllotmentSerializer(page4, many=True).data)
            # serializer = ShiftAllotmentSerializer(grouped_shifts, many=True)

            newlist=serializer.data 
            for i in newlist:
                
                user_obj=Users.objects.filter(id=i['employeeId'],is_active=True).first()
                if user_obj is not None:
                    i['employee_name']=user_obj.Firstname +" "+user_obj.Lastname
                i['date']=convertdate(i['date'])
            return self.get_paginated_response(newlist)
        else:
            return Response({'n':0,'Msg':'No data found','Status':'Failed','data':''})
        


@api_view(['POST'])
def dashboardcalender(request):
    company_code=request.user.company_code
    year = int(request.POST.get('year'))
    month = int(request.POST.get('month'))
    userId=request.user.id
    # Get the first day of the month
    first_day = datetime.datetime(year, month, 1)
    # Find the first Saturday of the month
    first_saturday = first_day + timedelta(days=(calendar.SATURDAY - first_day.weekday() + 7) % 7)

    # Calculate the second Saturday and fourth Saturday
    second_saturday = first_saturday + timedelta(weeks=1)
    fourth_saturday = first_saturday + timedelta(weeks=3)

    # Get the calendar for the given month and year
    cal = calendar.monthcalendar(year, month)

    # Create a list to store the schedule
    schedule = []

    # Iterate through each week in the calendar
    for week in cal:
        # Iterate through each day in the week
        for day in week:
            # Check if the day is not zero (zero represents a day outside the month)
            if day != 0:
                # Create a date string in the "yyyy-mm-dd" format
                date_string = f"{year:04d}-{month:02d}-{day:02d}"
                # Check if the day is a Sunday or the second or fourth Saturday
                holiday_obj=Holidays.objects.filter(HolidayYear=year,Date=date_string,HolidayMonth=month,company_code=company_code,Active=True).first()
                if holiday_obj is not None:
                    schedule.append({'day'+str(day):{'date': date_string, 'reason': holiday_obj.Festival, 'class': 'Holiday'}})
                else:
                    if calendar.weekday(year, month, day) == calendar.SUNDAY or (date_string == second_saturday.strftime("%Y-%m-%d")) or (date_string == fourth_saturday.strftime("%Y-%m-%d")):
                        schedule.append({'day'+str(day):{'date': date_string, 'reason': 'Weekly off', 'class': 'weeklyoff'}})
                    else:
                        leave_obj=Leave.objects.filter(employeeId=userId,leave_status="Approved",Active=True,start_date__lte=date_string,end_date__gte=date_string).first()
                        if leave_obj is not None:
                            if leave_obj.WorkFromHome==False:
                                if leave_obj.leavetype=="Fullday":
                                    schedule.append({'day'+str(day):{'date': date_string, 'reason': 'Leave', 'class': 'leavefullday'}})
                                elif leave_obj.leavetype=="SecondHalf":
                                    schedule.append({'day'+str(day):{'date': date_string, 'reason': 'Leave', 'class': 'leavesecondhalf'}})
                                elif leave_obj.leavetype=="FirstHalf":
                                    schedule.append({'day'+str(day):{'date': date_string, 'reason': 'Leave', 'class': 'leavefirsthalf'}})    
                            else:
                                if leave_obj.leavetype=="Fullday":
                                    schedule.append({'day'+str(day):{'date': date_string, 'reason': 'Working from home', 'class': 'workingformhomefullday'}})
                                elif leave_obj.leavetype=="SecondHalf":
                                    schedule.append({'day'+str(day):{'date': date_string, 'reason': 'Working from home', 'class': 'workingformhomesecondhalf'}})
                                elif leave_obj.leavetype=="FirstHalf":
                                    schedule.append({'day'+str(day):{'date': date_string, 'reason': 'Working from home', 'class': 'workingformhomefirsthalf'}})  
                        else:
                            schedule.append({'day'+str(day):{'date': date_string, 'reason': 'Working', 'class': 'workingday'}})

    return Response({'n': 1, 'data': schedule, 'Msg': 'dashboard calendar', 'Status': 'success'})


@api_view(['POST'])
@permission_classes((AllowAny,))
def monthly_attendance_report_shedular(request):
    # Get the current date
    current_date = datetime.datetime.now()
    # Calculate the first day of the current month
    first_day_of_current_month = datetime.datetime(current_date.year, current_date.month, 1)
    # Calculate the last day of the last month
    last_day_of_last_month = first_day_of_current_month - timedelta(days=1)
    # Get the month number and year of the last month
    last_month_number = last_day_of_last_month.month
    last_month_year = last_day_of_last_month.year
    
    last_month_name = last_day_of_last_month.strftime("%B")
    
    number_of_days = calendar.monthrange(last_month_year, last_month_number)[1]
    first_date = date(last_month_year, last_month_number, 1)
    last_date = date(last_month_year, last_month_number, number_of_days)
    delta = last_date - first_date
    dayslist=[(first_date + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(delta.days + 1)]
    
    
    employee_list_obj=Users.objects.filter(is_active=True).exclude(Q(employeeId__isnull=True)|Q(employeeId='None')|Q(employeeId=''))
    employee_serializer=UserSerializer(employee_list_obj,many=True)
    All_attendance_report=[]
    
    for emp in employee_serializer.data:
        exclude_obj_list=[]
        list1=[]
        weeklyofflist=[]
        holidaylist=[]

        monthly_total_days=0
        monthly_working_days=len(dayslist)
        monthly_worked_days=0
        monthly_late_marks=0
        monthly_leaves=0
        
        
        #2nd 4th saturday of month
        year=last_month_year 
        month=last_month_number
        dt=date(year,month,1)   # first day of month
        first_w=dt.isoweekday()  # weekday of 1st day of the month
        if(first_w==7): # if it is Sunday return 0 
            first_w=0
        saturday2=14-first_w
        dt1=date(year,last_month_number,saturday2)
        list1.append(str(dt1))
        saturday4=28-first_w
        dt1=date(year,last_month_number,saturday4)
        list1.append(str(dt1))  
        weeklyofflist=list1
        
        
        
        # all sundays of the month
        def allsundays(month):
            d = date(year,month, 1) # day 1st of month
            d += timedelta(days = 6 - d.weekday())  # First Sunday
            while d.month == month:
                yield d
                d += timedelta(days = 7)

        for d in allsundays(month):
            d=str(d)
            weeklyofflist.append(d)
        

        holidatlist = Holidays.objects.filter(Active=True).order_by('id')
        serializer = holidaysSerializer(holidatlist, many=True)
        for h in serializer.data:
            holiday=h['Date']
            holidaylist.append(holiday) 

        def get_dates_between(start_date, end_date):
            start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')

            dates = []
            current_date = start_date

            while current_date <= end_date:
                dates.append(current_date.strftime('%Y-%m-%d'))
                current_date += timedelta(days=1)

            return dates
    
        leave_objs = Leave.objects.filter(employeeId=emp['id'],leave_status="Approved",Active=True)
        leavesserializer=leaveserializer(leave_objs,many=True)
        workfromhomelist=[]
        normalleave=[]

        for l in leavesserializer.data:
            if l['WorkFromHome']==True:
                datelist=get_dates_between(l['start_date'],l['end_date'],)
                for d in datelist:
                    if d not in workfromhomelist:
                        workfromhomelist.append(d)
            else:
                normalleave.append(l)
                
        halfdayleave=[]
        fulldayleave=[]

        for leave in normalleave:
            if leave['leavetype'] != "Fullday":
                datelist=get_dates_between(leave['start_date'],leave['end_date'],)
                for d in datelist:
                    if d not in halfdayleave:
                        halfdayleave.append(d)
            else:
                datelist=get_dates_between(leave['start_date'],leave['end_date'],)
                for d in datelist:
                    if d not in fulldayleave:
                        fulldayleave.append(d)
        
        date_dict=[]
        for j in dayslist:
            monthly_total_days=monthly_total_days+1

            shift={}
            shift['shiftname']='General'
            shift['intime']='09:30'
            shift['outtime']='18:30'
            shift['swap']=''
            
            alloted_shift_obj=ShiftAllotment.objects.filter(employeeId=emp['id'],attendanceId=emp['employeeId'],date=j,is_active=True).first()
            if alloted_shift_obj is not None:
                ShiftAllotment_serializer=ShiftAllotmentSerializer(alloted_shift_obj)
                if ShiftAllotment_serializer.data['shiftId'] is not None and ShiftAllotment_serializer.data['shiftId'] !='':
                    shift_obj=ShiftMaster.objects.filter(id=ShiftAllotment_serializer.data['shiftId'],is_active=True).first()
                    if shift_obj is not None:
                        shift_obj_serializer=ShiftMasterSerializer(shift_obj)
                        shift['shiftname']=shift_obj_serializer.data['shiftname']
                        shift['intime']=shift_obj_serializer.data['intime']
                        shift['outtime']=shift_obj_serializer.data['outtime']
                        if ShiftAllotment_serializer.data['is_swaped'] == True:
                            check_swap_details=Shiftswap.objects.filter(id=ShiftAllotment_serializer.data['swap_request_id']).first()
                            if check_swap_details is not None:
                                Shiftswap_serializer=CustomShiftswapSerializer(check_swap_details)
                                if ShiftAllotment_serializer.data['swapper'] == True:
                                    shift['swap']='Swaped with '+ str(Shiftswap_serializer.data['second_employee_name'])
                                elif ShiftAllotment_serializer.data['swapper'] == False:
                                    shift['swap']='Swaped with '+ str(Shiftswap_serializer.data['first_employee_name'])
                                else:
                                    shift['swap']=''
                            else:
                                shift['swap']=''
                        else:
                            shift['swap']=''
                            
            shift_starting_time = datetime.datetime.strptime(str(shift['intime']), "%H:%M")
            shift_ending_time = datetime.datetime.strptime(str(shift['outtime']), "%H:%M")


            def get_nearest_login_logout_times(employee_id, date,shift_start_time,shift_end_time,exclude_obj_list):
                fnextday = datetime.datetime.strptime(date , "%Y-%m-%d").date() + timedelta(days=1)  
                  
                start_time = datetime.datetime.strptime(str(shift_start_time), "%H:%M:%S")
                end_time = datetime.datetime.strptime(str(shift_end_time), "%H:%M:%S")
                
                start_date=date
                end_date=date
                
                if end_time < start_time:
                    end_date = datetime.datetime.strptime(date , "%Y-%m-%d").date() + timedelta(days=1) 

                start_datetime = datetime.datetime.strptime(start_date + " " + shift_start_time, "%Y-%m-%d %H:%M:%S")
                shift_allotment_obj_next_day=ShiftAllotment.objects.filter(date=fnextday,is_active=True,employeeId=emp['id']).first()
                nextdayshift={}
                
                
                if shift_allotment_obj_next_day is not None:
                    shift_obj_next_day=ShiftMaster.objects.filter(id=shift_allotment_obj_next_day.shiftId,is_active=True).first()
                    if shift_obj_next_day is not None:
                        nextdayshift['inTime']=shift_obj_next_day.intime
                        nextdayshift['outTime']=shift_obj_next_day.outtime
                        nextdayshift['shiftname']=shift_obj_next_day.shiftname
                    else:
                        nextdayshift['shiftname']='General '
                        nextdayshift['inTime']='09:30'
                        nextdayshift['outTime']='18:30'
                else:
                    nextdayshift['shiftname']='General '
                    nextdayshift['inTime']='09:30'
                    nextdayshift['outTime']='18:30'
                    
                    
                    
                nextdayshift_start_time = datetime.datetime.strptime(str(nextdayshift['inTime']),"%H:%M")
                nextday_datetime = datetime.datetime.strptime(str(fnextday) + " " + str(nextdayshift_start_time).split(" ")[1], "%Y-%m-%d %H:%M:%S")
                start_datetime_before_2_hours = start_datetime - timedelta(hours=2)
                nextday_start_datetime_before_2_hours=nextday_datetime - timedelta(hours=2)
                

                start_date_before_2_hours = start_datetime_before_2_hours.strftime("%Y-%m-%d")
                start_time_before_2_hours = start_datetime_before_2_hours.strftime("%H:%M:%S")
                nextday_date_before_2_hours=nextday_start_datetime_before_2_hours.strftime("%Y-%m-%d")

                
                # Get the login time entry nearest to the shift starting time
                if shift_start_time > shift_end_time:
                    login_obj=attendance.objects.filter(employeeId=employee_id,date=str(start_date_before_2_hours),time__gte=start_time_before_2_hours).order_by('time').first()
                    if login_obj is None:
                        login_obj=attendance.objects.filter(Q(employeeId=employee_id,date=str(start_date_before_2_hours),time__gte=start_time_before_2_hours)|Q(employeeId=employee_id,date=nextday_date_before_2_hours,time__lte=datetime.datetime.strptime(shift_end_time, "%H:%M:%S").time())).order_by('time').first()
                else:
                    login_obj=attendance.objects.filter(employeeId=employee_id,date=str(start_date_before_2_hours),time__gte=start_time_before_2_hours).order_by('time').first()
                         
                # login_obj = attendance.objects.filter(employeeId=employee_id,date=start_date_before_2_hours, time__gte=start_time_before_2_hours).order_by('time').first()
                nearest_login_time = login_obj if login_obj else None
                nextday_date_before_2_hours = nextday_start_datetime_before_2_hours.strftime("%Y-%m-%d")
                nextday_time_before_2_hours = nextday_start_datetime_before_2_hours.strftime("%H:%M:%S")
                # Get the logout time entry nearest to the shift ending time
                logout_obj = attendance.objects.filter(employeeId=employee_id,date=nextday_date_before_2_hours, time__lt=nextday_time_before_2_hours).exclude(id__in=exclude_obj_list).order_by('time').last()
                extra_days=''
                if logout_obj is None:
                    logout_obj = attendance.objects.filter(employeeId=employee_id,date=start_date, time__gt=start_time_before_2_hours).exclude(id__in=exclude_obj_list).order_by('time').last()

                if logout_obj is not None:
                    if str(start_date_before_2_hours) != str(logout_obj.date):
                        extra_days='+1'

                    
                    
                nearest_logout_time = logout_obj if logout_obj else None

                return nearest_login_time, nearest_logout_time, extra_days

            nearest_login_time, nearest_logout_time ,extra_days= get_nearest_login_logout_times(emp['employeeId'], j,str(shift_starting_time).split(" ")[1],str(shift_ending_time).split(" ")[1],exclude_obj_list)

            if j in fulldayleave:

                monthly_leaves=monthly_leaves+1
                logint=""
                logoutt=""
                total_time=""

                d1={
                    "employeeId":emp['employeeId'],
                    "att_status":"Leave",
                    "fulldate":convertdate2(str(j)),
                    "inTime":logint,
                    "outTime":logoutt,
                    "Total":total_time,
                    "shift":shift,
                }
                date_dict.append(d1)
                
            elif j in halfdayleave:

                monthly_leaves=monthly_leaves+0.5
                if nearest_login_time is not None:
                    
                    monthly_worked_days=0.5 + monthly_worked_days
                    
                    logint=nearest_login_time.time
                    login_date=nearest_login_time.date
                    exclude_obj_list.append(nearest_login_time.id)    

                    if nearest_logout_time is not None:
                        logoutt=nearest_logout_time.time
                        logout_date=nearest_logout_time.date
                        exclude_obj_list.append(nearest_logout_time.id)    

                    else:
                        logoutt=logint
                        logout_date=login_date

                    login_datetime = datetime.datetime.strptime(str(login_date) + " " + str(logint), "%Y-%m-%d %H:%M:%S")
                    logout_datetime = datetime.datetime.strptime(str(logout_date) + " " + str(logoutt), "%Y-%m-%d %H:%M:%S")
                    timeDiff = logout_datetime - login_datetime
                    total_time = str(timeDiff)

                    if logint==logoutt: 

                        total_time='00:00:00' 
                        logoutt=''
                        logout_date=j


                    
                    
                else:
                    logint=""
                    logoutt=""
                    total_time=""
                    login_date=j





                d1={
                    "employeeId":emp['employeeId'],
                    "att_status":"Half day",
                    "fulldate":convertdate2(str(j)),
                    "inTime":logint,
                    "outTime":logoutt,
                    "Total":total_time,
                    "shift":shift,
                    "extra_days":extra_days,
                    "inTime_date":convertdate2(str(login_date)),
                    "outTime_date":convertdate2(str(logout_date)),
                }
                date_dict.append(d1)

            elif j in workfromhomelist:
                if nearest_login_time is not None:
                    logint=nearest_login_time.time
                    login_date=nearest_login_time.date
                    exclude_obj_list.append(nearest_login_time.id)    


                    monthly_worked_days=1 + monthly_worked_days
                    if nearest_logout_time is not None:
                        logoutt=nearest_logout_time.time
                        logout_date=nearest_logout_time.date
                        exclude_obj_list.append(nearest_logout_time.id)    

                    else:
                        logoutt=logint
                        logout_date=login_date

                    login_datetime = datetime.datetime.strptime(str(login_date) + " " + str(logint), "%Y-%m-%d %H:%M:%S")
                    logout_datetime = datetime.datetime.strptime(str(logout_date) + " " + str(logoutt), "%Y-%m-%d %H:%M:%S")
                    timeDiff = logout_datetime - login_datetime
                    total_time = str(timeDiff)
                    
                else:
                    logint=""
                    logoutt=""
                    total_time=""
                    login_date=j
                    logout_date=j
                d1={
                    "employeeId":emp['employeeId'],
                    "att_status":"Work From Home",
                    "fulldate":convertdate2(str(j)),
                    "inTime":logint,
                    "outTime":logoutt,
                    "Total":total_time,
                    "shift":shift,
                    "extra_days":extra_days,
                    "inTime_date":convertdate2(str(login_date)),
                    "outTime_date":convertdate2(str(logout_date)),
                    
                    
                    

                }
                date_dict.append(d1)

            elif nearest_login_time is not None:
                logint=nearest_login_time.time
                login_date=nearest_login_time.date
                exclude_obj_list.append(nearest_login_time.id)    
                monthly_worked_days=1 + monthly_worked_days

                if nearest_logout_time is not None:
                    logoutt=nearest_logout_time.time
                    logout_date=nearest_logout_time.date
                    exclude_obj_list.append(nearest_logout_time.id)    

                else:
                    logoutt=logint
                    logout_date=login_date

                login_datetime = datetime.datetime.strptime(str(login_date) + " " + str(logint), "%Y-%m-%d %H:%M:%S")
                logout_datetime = datetime.datetime.strptime(str(logout_date) + " " + str(logoutt), "%Y-%m-%d %H:%M:%S")
                timeDiff = logout_datetime - login_datetime
                total_time = str(timeDiff)

                if logint==logoutt: 

                    total_time='00:00:00' 
                    logoutt=''
                    logout_date=j
                    att_status = "Working"


                d1={ 
                    "employeeId":emp['employeeId'],
                    "att_status":att_status,
                    "fulldate":convertdate2(str(j)),
                    "inTime":logint,
                    "outTime":logoutt,
                    "extra_days":extra_days,
                    "inTime_date":convertdate2(str(login_date)),
                    "outTime_date":convertdate2(str(logout_date)),
                    "Total":total_time,
                    "shift":shift,

                }
                date_dict.append(d1)

            elif j in weeklyofflist:
                logint=""
                logoutt=""
                total_time=""
                monthly_working_days=monthly_working_days-1
                d1={
                    "employeeId":emp['employeeId'],
                    "att_status":"Weekly OFF",
                    "fulldate":convertdate2(str(j)),
                    "inTime":logint,
                    "outTime":logoutt,
                    "Total":total_time,
                    "shift":shift,


                }
                date_dict.append(d1)

            elif j in holidaylist:
                monthly_working_days=monthly_working_days-1
                
                logint=""
                logoutt=""
                total_time=""
                reason='Holiday'
                Holidays_obj=Holidays.objects.filter(Date=j,Active=True).first()
                if Holidays_obj:
                    reason=str(Holidays_obj.Festival)
 
            
                d1={
                    "employeeId":emp['employeeId'],
                    "att_status":"Holiday",
                    "fulldate":convertdate2(str(j)),
                    "inTime":logint,
                    "outTime":logoutt,
                    "Total":total_time,
                    "reason":reason,
                    "shift":shift,

                }
                date_dict.append(d1)

            else:
                logint=""
                logoutt=""
                total_time=""

                d1={
                    "employeeId":emp['employeeId'],
                    "att_status":"Not Found",
                    "fulldate":convertdate2(str(j)),
                    "inTime":logint,
                    "outTime":logoutt,
                    "Total":total_time,
                    "shift":shift,
                }
                date_dict.append(d1)

            if d1['inTime'] !='' :
               
                time_1 =  datetime.datetime.strptime(str(d1['shift']['intime']), "%H:%M")
                time_2 =  time_1 + datetime.timedelta(minutes=30)
                time_3 =  time_2 + datetime.timedelta(minutes=60)
                if time_1 >= time_2 and time_1 < time_3:
                    monthly_late_marks=monthly_late_marks+1



        attendance_report_dict={'employeeId':emp['id'],'Employee_Name': str(emp['Firstname']) + ' ' + str(emp['Lastname']),
                                    'month':last_month_name,
                                    'year':last_month_year,
                                    
                                    'monthly_total_days':   monthly_total_days,
                                    'monthly_working_days':  monthly_working_days,
                                    'monthly_worked_days':monthly_worked_days,
                                    'monthly_late_marks':monthly_late_marks,
                                    'monthly_leaves':monthly_leaves,
                                    'report':date_dict,

        }
        # All_attendance_report.append(attendance_report_dict)
        send_async_custom_template_email(
            'Monthly Attendance Report',
            attendance_report_dict,
            EMAIL_HOST_USER,
            [emp['email']],
            "mails/attendence_report.html"
        )


    return Response({'n': 1, 'data': All_attendance_report, 'Msg': 'monthly attendance report', 'Status': 'success'})




@api_view(['POST'])
def add_employeetype(request):
    data={}
    data['employeetype']=request.POST.get("employeetype")
    employeetype_obj=employeetypemaster.objects.filter(employeetype=data['employeetype'],is_active=True).first()
    if employeetype_obj is None:
        serializer = employeetypemasterserializer(employeetype_obj,data=data)
        if serializer.is_valid():
            serializer.save()
            return Response({'n':1,'msg':'Employee type added successfully','status':'success','data':serializer.data})
        else:
            return Response({'n':0,'msg':'error occurs in serializer','status':'error','data':[]})
    else:
        return Response({'n':0,'msg':'employee type already exists','status':'error','data':[]})
    

    
@api_view(['GET'])
def get_all_employeetype(request):
    employeetype_obj=employeetypemaster.objects.filter(is_active=True)
    if employeetype_obj is not None:
        serializer = employeetypemasterserializer(employeetype_obj,many=True)
        return Response({'n':1,'msg':'employee types found','status':'success','data':serializer.data})
    else:
        return Response({'n':0,'msg':'employee type not found','status':'error','data':[]})
    
    
    
@api_view(['POST'])
def update_employeetype(request):
    data={}
    data['employeetype']=request.POST.get("employeetype")
    data['id']=request.POST.get("id")
    employeetype_obj=employeetypemaster.objects.filter(employeetype=data['employeetype'],is_active=True).exclude(id=data['id']).first()
    if employeetype_obj is None:
        update_obj=employeetypemaster.objects.filter(id=data['id'],is_active=True).first()
        if update_obj is not None:
            serializer = employeetypemasterserializer(update_obj,data=data)
            if serializer.is_valid():
                serializer.save()
                return Response({'n':1,'msg':'Employee type updated successfully','status':'success','data':serializer.data})
            else:
                return Response({'n':0,'msg':'error occurs in serializer','status':'error','data':[]})
        else:
            return Response({'n':0,'msg':'employee type not found','status':'error','data':[]})
    else:
        return Response({'n':0,'msg':'employee type already exists','status':'error','data':[]})
    
    
@api_view(['POST'])
def delete_employeetype(request):
    data={}
    data['id']=request.POST.get("id")
    delete_obj=employeetypemaster.objects.filter(id=data['id'],is_active=True).first()
    if delete_obj is not None:
        userobj=Users.objects.filter(employeetype=delete_obj.id,is_active=True).first()
        if userobj is None:
            
            data['is_active']=False
            serializer = employeetypemasterserializer(delete_obj,data=data,partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({'n':1,'msg':'Employee type deleted successfully','status':'success','data':serializer.data})
            else:
                return Response({'n':0,'msg':'error occurs in serializer','status':'error','data':[]})
        else:
            return Response({'n':0,'msg':'This employee type is mapped to employees ,please unmapped this employee type from all employees to delete ','status':'error','data':[]})
    else:
        return Response({'n':0,'msg':'employee type not found','status':'error','data':[]})

    
@api_view(['POST'])
def get_employeetype_data(request):
    data={}
    data['id']=request.POST.get("id")
    emptype_obj=employeetypemaster.objects.filter(id=data['id'],is_active=True).first()
    if emptype_obj is not None:
        serializer = employeetypemasterserializer(emptype_obj)
        return Response({'n':1,'msg':'Employee type found successfully','status':'success','data':serializer.data})
    else:
        return Response({'n':0,'msg':'employee type not found','status':'error','data':[]})
    
    
  

@api_view(['POST'])    
@permission_classes((AllowAny,))
def employee_calendar(request):
    
    
    
    month=int(request.data.get("month"))
    year=int(request.data.get("year"))
    
    UserId=request.data.get("UserId")
    user_obj=Users.objects.filter(id=UserId,is_active=True).exclude(Q(employeeId__isnull=True)|Q(employeeId='None')|Q(employeeId='')).first()

               
    weeklyofflist=[]
    holidaylist=[]
    list1=[]
    #2nd 4th saturday of month
    year=year 
    dt=date(year,month,1)   # first day of month
    first_w=dt.isoweekday()  # weekday of 1st day of the month
    if(first_w==7): # if it is Sunday return 0 
        first_w=0
    saturday2=14-first_w
    dt1=date(year,month,saturday2)
    list1.append(str(dt1))
    saturday4=28-first_w
    dt1=date(year,month,saturday4)
    list1.append(str(dt1))  
    weeklyofflist=list1
        
    # all sundays of the month
    def allsundays(month):
        d = date(year,month, 1) # day 1st of month
        d += timedelta(days = 6 - d.weekday())  # First Sunday
        while d.month == month:
            yield d
            d += timedelta(days = 7)

    for d in allsundays(month):
        d=str(d)
        weeklyofflist.append(d)
    

 
    
    
    
    office_location=Location.objects.filter(Active=True).exclude(Q(lattitude='',meter='',longitude='') |Q(lattitude=None,meter=None,longitude=None)|Q(lattitude__isnull=True,meter__isnull=True,longitude__isnull=True))
    location_serializer=LocationSerializer(office_location,many=True)
    
    if user_obj is not None:
        employee_serializer=UsersSerializer(user_obj)
        employeeId=user_obj.employeeId
        if month !='' and year !=''  and  month is not None and year is not None:
            leave_objs = Leave.objects.filter(employeeId=UserId,WorkFromHome=False,leave_status="Approved",start_date__month__gte=month,start_date__month__lte=month,start_date__year= year,Active=True).exclude(leave_status='Draft')
            leavesserializer=leaveserializer(leave_objs,many=True)
            wfh_objs = Leave.objects.filter(employeeId=UserId,WorkFromHome=True,leave_status="Approved",start_date__month__gte=month,start_date__month__lte=month,start_date__year= year,Active=True).exclude(leave_status='Draft')
            wfhsserializer=leaveserializer(wfh_objs,many=True)
            cal = calendar.monthcalendar(year, month)
            dates_list = [
                    f"{year}-{month:02d}-{day:02d}"
                    for week in cal
                    for day in week
                    if day != 0
                ]    
            attendance_report=[]    
            for datestr in dates_list:
                current_date_shift_list=[]
                new_current_date_shift_list=[]
                return_dict={}
                
                shiftdate = datetime.datetime.strptime(datestr, '%Y-%m-%d').date()
                tomarrow_date = shiftdate + datetime.timedelta(days=1)
                yesterday_date = shiftdate - datetime.timedelta(days=1)
                current_shiftallotment_objs=ShiftAllotment.objects.filter(attendanceId=employeeId,date=str(datestr),is_active=True)
                current_shiftallotment_serializers=ShiftAllotmentshiftIdSerializer(current_shiftallotment_objs,many=True)
                current_shiftId_list=list(current_shiftallotment_serializers.data)
                current_shift_obj=ShiftMaster.objects.filter(id__in=current_shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime')
                current_shift_serializer=ShiftMasterSerializer(current_shift_obj,many=True)
                current_shiftlist=current_shift_serializer.data
                
                tomarrow_shiftallotment_objs=ShiftAllotment.objects.filter(attendanceId=employeeId,date=str(tomarrow_date),is_active=True)
                tomarrow_shiftallotment_serializers=ShiftAllotmentshiftIdSerializer(tomarrow_shiftallotment_objs,many=True)
                tomarrow_shiftId_list=list(tomarrow_shiftallotment_serializers.data)
                tomarrow_shift_obj=ShiftMaster.objects.filter(id__in=tomarrow_shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime')
                tomarrow_shift_serializer=ShiftMasterSerializer(tomarrow_shift_obj,many=True)
                tomarrow_shiftlist=tomarrow_shift_serializer.data

                count=1
                
                for shift in current_shiftlist:
                    start_time = datetime.datetime.strptime(str(shiftdate) +' '+ shift['intime'], '%Y-%m-%d %H:%M')
                    start_time_before_2hrs = start_time - timedelta(hours=2)
                    
                    if shift['intime'] > shift['outtime']:
                        shift_end_date = shiftdate + timedelta(days=1)
                    else:
                        shift_end_date = shiftdate
                        
                    end_time = datetime.datetime.strptime(str(shift_end_date) +' '+ shift['outtime'], '%Y-%m-%d %H:%M')
                    check_login_till=''
                    if len(current_shiftlist) >= count+1:
                   
                        check_next_shift_in = datetime.datetime.strptime(str(shiftdate) +' '+ current_shiftlist[count]['intime'], '%Y-%m-%d %H:%M')
                        check_login_till = check_next_shift_in - timedelta(hours=2)

                        
                    elif check_login_till=='':
                          
                        if len(tomarrow_shiftlist) >= 1:
                         
                            check_next_shift_in = datetime.datetime.strptime(str(tomarrow_date) +' '+ tomarrow_shiftlist[0]['intime'], '%Y-%m-%d %H:%M')
                            check_login_till = check_next_shift_in - timedelta(hours=2)
                        else:
                      
                            if_no_shift_check_shift=ShiftAllotment.objects.filter(attendanceId=employeeId,date=str(tomarrow_date),is_active=True)
                            tomerrow_shiftlist=ShiftAllotmentshiftIdSerializer(if_no_shift_check_shift,many=True)
                            next_date_shiftId_list=list(tomerrow_shiftlist.data)
                            check_weekly_off=ShiftMaster.objects.filter(id__in=next_date_shiftId_list,intime='00:00',outtime='00:00',is_active=True).order_by('intime').first()
                            if check_weekly_off is not None:
                              
                                if shift['outtime'] < shift['intime']:
                         
                                    check_current_shift_out = datetime.datetime.strptime(str(tomarrow_date) +' '+ shift['outtime'], '%Y-%m-%d %H:%M')
                                    check_current_shift_out_time = check_current_shift_out + timedelta(hours=2)
                                    given_datetime = datetime.datetime.strptime(str(check_current_shift_out_time), '%Y-%m-%d %H:%M:%S')

                                    truncated_datetime_str = given_datetime.strftime('%Y-%m-%d %H:%M')

                                    
                                    shift_end_date_time=str(truncated_datetime_str)
                                else:
                                    shift_end_date_time=str(datestr) +' '+ '23:59'
                                    
                                check_login_till = datetime.datetime.strptime(shift_end_date_time, '%Y-%m-%d %H:%M')
                            else:
                            
                                check_login_till = datetime.datetime.strptime(str(tomarrow_date) +' '+ '07:30', '%Y-%m-%d %H:%M')
                                
               
                    shift_data={
                        'shiftname':shift['shiftname'],
                        'indatetime':str(start_time_before_2hrs),
                        'outdatetime':str(check_login_till),
                        'intime':shift['intime'],
                        'outtime':shift['outtime'],
                    }
                    current_date_shift_list.append(shift_data)
                    new_current_date_shift_list.append(shift_data)
                    count+=1

                if len(current_shiftlist) == 0:       
               
                    if_no_shift_check_shift=ShiftAllotment.objects.filter(attendanceId=employeeId,date=str(datestr),is_active=True)
                    todays_shiftlist=ShiftAllotmentshiftIdSerializer(if_no_shift_check_shift,many=True)
                    toadys_shiftId_list=list(todays_shiftlist.data)
                    check_weekly_off=ShiftMaster.objects.filter(id__in=toadys_shiftId_list,intime='00:00',outtime='00:00',is_active=True).order_by('intime').first()
                    if check_weekly_off is not None:
   
                        get_previous_day_shift=ShiftAllotment.objects.filter(attendanceId=employeeId,date=str(yesterday_date),is_active=True)
                        previous_day_shiftlist=ShiftAllotmentshiftIdSerializer(get_previous_day_shift,many=True)
                        previous_shiftId_list=list(previous_day_shiftlist.data)
                        get_previous_last_shift=ShiftMaster.objects.filter(id__in=previous_shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime').last()
                        
                        if get_previous_last_shift is not None:
                            if get_previous_last_shift.outtime < get_previous_last_shift.intime:
                                shift_end_date_time=str(datestr) +' '+ str(get_previous_last_shift.outtime)
                            else:
                                shift_end_date_time=str(yesterday_date) +' '+ '23:59'
                                    
                            check_previous_shift_out = datetime.datetime.strptime(str(shift_end_date_time), '%Y-%m-%d %H:%M')
                            previous_shift_in_time = check_previous_shift_out + timedelta(hours=2)  
                        else:
            
                            
                            previous_shift_in_time= str(datestr)+' 00:00:00'
                        # checking previous  shift out time must be currrent shift in time in weekly off    # 
                        get_net_day_shift=ShiftAllotment.objects.filter(attendanceId=employeeId,date=str(tomarrow_date),is_active=True)
                        nest_day_shiftlist=ShiftAllotmentshiftIdSerializer(get_net_day_shift,many=True)
                        next_date_shiftId_list=list(nest_day_shiftlist.data)
                        get_next_day_first_shift=ShiftMaster.objects.filter(id__in=next_date_shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime').first()
                        if get_next_day_first_shift is not None:
                            check_next_shift_in = datetime.datetime.strptime(str(tomarrow_date) +' '+ str(get_next_day_first_shift.intime), '%Y-%m-%d %H:%M')
                            next_shift_in_time = check_next_shift_in - timedelta(hours=2)
                            shift_data={
                                'shiftname':str(check_weekly_off.shiftname),
                                'indatetime':str(previous_shift_in_time),
                                'outdatetime':str(next_shift_in_time),
                                'intime':check_weekly_off.intime,
                                'outtime':check_weekly_off.outtime,
                            }
                            current_date_shift_list.append(shift_data)
                            new_current_date_shift_list.append(shift_data)
                            
                        else:
                            shift_data={
                                'shiftname':str(check_weekly_off.shiftname),
                                'indatetime':str(previous_shift_in_time),
                                'outdatetime':str(datestr)+' 23:59:59',
                                'intime':check_weekly_off.intime,
                                'outtime':check_weekly_off.outtime,
                            }
                            current_date_shift_list.append(shift_data)
                            new_current_date_shift_list.append(shift_data)

                    else:
                        shift_data={
                            'shiftname':'General',
                            'indatetime':str(datestr)+' 07:30:00',
                            'outdatetime':str(tomarrow_date)+' 07:30:00',
                            'intime':'09:30',
                            'outtime':'18:30',
                        }
                        current_date_shift_list.append(shift_data)
                        new_current_date_shift_list.append(shift_data)

                        
                current_date_first_in_datetime=''
                current_date_last_out_datetime=''
                extra_days=''
                total_working_hrs=''
                for s in new_current_date_shift_list:
                    intime=''
                    intimedate=''
                    
                    
                    getallattendance = attendance.objects.filter(Q(employeeId=str(employeeId),time__gte=s['indatetime'].split(' ')[1],date=s['indatetime'].split(' ')[0])|Q(employeeId=str(employeeId),time__lte=str(s['outdatetime']).split(' ')[1],date=str(s['outdatetime']).split(' ')[0])).order_by('date','time')
                                        
                    attendance_serializer=attendanceserializer(getallattendance,many=True)
                
                    
                    sorted_data = sorted(attendance_serializer.data, key=lambda x: (x['date'], x['time']))
                    
                    mindatetime = datetime.datetime.strptime(s['indatetime'], '%Y-%m-%d %H:%M:%S')
                    maxdatetime = datetime.datetime.strptime(s['outdatetime'], '%Y-%m-%d %H:%M:%S')
                
                    sorted_data = [entry for entry in sorted_data if (mindatetime <= datetime.datetime.strptime(entry['date'] +' '+entry['time'],'%Y-%m-%d %H:%M:%S') <= maxdatetime)]
                    if len(sorted_data) > 0:
                        intimedate=sorted_data[0]['date']
                        intime=str(sorted_data[0]['time'])
                        
                    if intimedate !='' and intimedate is not None:
                        user_sdt = datetime.datetime.strptime(str(intimedate) + ' ' + str(intime), '%Y-%m-%d %H:%M:%S')
                        shif_sdt = datetime.datetime.strptime(s['indatetime'].split(' ')[0] + ' ' + s['indatetime'].split(' ')[1], '%Y-%m-%d %H:%M:%S')
                        if user_sdt < shif_sdt :
                            intimedate=''
                            intime=''

                            
                    checkin_time = None
                    total_working_time = 0
                    attendance_log=[]
                    attendance_history=[]

                    for index, entry in enumerate(sorted_data):

                        if entry['checkout']:
                            if entry['deviceId'] is not None and entry['deviceId'] !='':
                                attendance_type="Machine Checkout"
                                attendance_type_resaon=""
                                if entry['deviceId'] == 20:
                                    remote_map_location=create_google_maps_url(18.566064883698555, 73.77574703366226)
                                    remote_map_name='Zentro Pune Office'
                                elif entry['deviceId'] == 19:
                                    remote_map_location=create_google_maps_url(19.1764898841813, 72.95988765068624)
                                    remote_map_name='Zentro Mumbai Office'
                                else:
                                    remote_map_location=''
                                    remote_map_name=''
                            elif entry['Remote_Reason'] is not None and entry['Remote_Reason'] !='':
                                attendance_type="Remote Checkout"
                                attendance_type_resaon=entry['Remote_Reason']
                                remote_map_location=create_google_maps_url(entry['remote_latitude'],entry['remote_longitude'])
                                remote_map_name=get_location_name(entry['remote_latitude'],entry['remote_longitude'])

                            else:
                                attendance_type="Web Checkout"
                                attendance_type_resaon=''
                                
                                if entry['attendance_type'] !='' and  entry['attendance_type'] is not None:
                                    attendance_type = entry['attendance_type'] +' '+ 'Checkout'
                                    
                                remote_map_location=''
                                            
                                if entry['remote_latitude'] is not None and  entry['remote_latitude'] !='' and entry['remote_longitude'] is not None and entry['remote_longitude'] !='':

                                    for location in location_serializer.data:
                                        within_radius = is_within_radius(float(entry['remote_latitude']), float(entry['remote_longitude']), float(location['lattitude']), float(location['longitude']), float(location['meter']))
                                        if within_radius:
                                            remote_map_location=create_google_maps_url(entry['remote_latitude'],entry['remote_longitude'])
                                            remote_map_name='Zentro '+str(location['LocationName']) +' Office'
                                    if remote_map_location == '' or  remote_map_name =='' or remote_map_location is None or  remote_map_name  is  None:
                                        remote_map_location=create_google_maps_url(entry['remote_latitude'],entry['remote_longitude'])
                                        remote_map_name=get_location_name(entry['remote_latitude'],entry['remote_longitude'])
                                else:
                                    remote_map_location=''
                                    remote_map_name=''                        

                            attendance_history.append({'Status':'Check-Out','datetime':entry['date']+' '+entry['time'],'attendance_type':attendance_type,'attendance_type_resaon':attendance_type_resaon,'remote_map_location':remote_map_location,'remote_map_name':remote_map_name})
                                
                            attendance_log.append({'checkout':entry['date']+' '+entry['time']})
                            
                            
                            if checkin_time:
                                checkout_datetime = datetime.datetime.strptime(entry['date'] + ' ' + entry['time'], '%Y-%m-%d %H:%M:%S')
                                checkin_datetime = datetime.datetime.strptime(checkin_time, '%Y-%m-%d %H:%M:%S')
                                working_time = checkout_datetime - checkin_datetime
                                total_working_time += working_time.total_seconds()
                                checkin_time = None
                                
                        elif not entry['checkout']:
                            remote_map_location=''
                            remote_map_name='' 
                            if entry['deviceId'] is not None and entry['deviceId'] !='':
                                attendance_type="Machine Checkin"
                                attendance_type_resaon=''
                                if entry['deviceId'] == 20:
                                    remote_map_location=create_google_maps_url(18.566064883698555, 73.77574703366226)
                                    remote_map_name='Zentro Pune Office'
                                elif entry['deviceId'] == 19:
                                    remote_map_location=create_google_maps_url(19.1764898841813, 72.95988765068624)
                                    remote_map_name='Zentro Mumbai Office'
                                else:
                                    remote_map_location=''
                                    remote_map_name=''
                                    
                            elif entry['Remote_Reason'] is not None and entry['Remote_Reason'] !='':
                                attendance_type="Remote Checkin"
                                attendance_type_resaon=entry['Remote_Reason']
                                remote_map_location=create_google_maps_url(entry['remote_latitude'],entry['remote_longitude'])
                                remote_map_name=get_location_name(entry['remote_latitude'],entry['remote_longitude'])
                            else:
                                attendance_type="Web Checkin"
                                attendance_type_resaon='' 
                                if entry['attendance_type'] !='' and  entry['attendance_type'] is not None:
                                    attendance_type = entry['attendance_type'] +' '+ 'Checkin'
                                remote_map_location=''
                                
                                if entry['remote_latitude'] is not None and  entry['remote_latitude'] !='' and entry['remote_longitude'] is not None and entry['remote_longitude'] !='':
                
                                    for location in location_serializer.data:
                                        within_radius = is_within_radius(float(entry['remote_latitude']), float(entry['remote_longitude']), float(location['lattitude']), float(location['longitude']), float(location['meter']))
                                        if within_radius:
                                            remote_map_location=create_google_maps_url(entry['remote_latitude'],entry['remote_longitude'])
                                            remote_map_name='Zentro '+str(location['LocationName']) +' Office'
                                    if remote_map_location == '' or  remote_map_name =='' or remote_map_location is None or  remote_map_name  is  None:
                                        remote_map_location=create_google_maps_url(entry['remote_latitude'],entry['remote_longitude'])
                                        remote_map_name=get_location_name(entry['remote_latitude'],entry['remote_longitude'])
                                else:
                                    remote_map_location=''
                                    remote_map_name='' 
                                 
                            checkin_time = entry['date'] + ' ' + entry['time']
                            attendance_log.append({'checkin':entry['date']+' '+entry['time']})
                            attendance_history.append({'Status':'Check-In','datetime':entry['date']+' '+entry['time'],'attendance_type':attendance_type,'attendance_type_resaon':attendance_type_resaon,'remote_map_location':remote_map_location,'remote_map_name':remote_map_name})
                    # If the last entry is check-in, consider checkout time as current shift end  time
                    if checkin_time and index == len(sorted_data) - 1:
                        
                        # check if the previous entry is also checkin or not if exist get the  difference betwnn  the current checkin  and  previous  checkin

                        if int(int(index)-1) >=0:
                            if sorted_data[index-1]['checkout']==False:
                                previous_checkin_date_time=sorted_data[index-1]['date']+ ' '+sorted_data[index-1]['time']
                                checkout_datetime = datetime.datetime.strptime(previous_checkin_date_time, '%Y-%m-%d %H:%M:%S')
                                checkin_datetime = datetime.datetime.strptime(checkin_time, '%Y-%m-%d %H:%M:%S')
                                working_time = checkin_datetime-checkout_datetime 
                                total_working_time += working_time.total_seconds()


                    # Convert total_working_time to hours, minutes, and seconds
                    hours, remainder = divmod(total_working_time, 3600)
                    minutes, seconds = divmod(remainder, 60)
                    # print("attendance_log",attendance_log)

                    s['total_hrs']=str(add_leading_zero(int(hours))) +':' +str(add_leading_zero(int(minutes))) +':'+str(add_leading_zero(int(seconds)))
                    
                    if total_working_hrs == '':
                        total_working_hrs=s['total_hrs']
                    else:
                        time_str_1 = total_working_hrs
                        time_str_2 = s['total_hrs']

                        # Convert time strings to timedelta objects and add them
                        total_time_delta = timedelta(hours=int(time_str_1[:2]), minutes=int(time_str_1[3:5]), seconds=int(time_str_1[6:])) + \
                                        timedelta(hours=int(time_str_2[:2]), minutes=int(time_str_2[3:5]), seconds=int(time_str_2[6:]))
                        total_working_hrs=str(total_time_delta)
                        
                    s['attendance_log']=attendance_log
                    s['attendance_history']=attendance_history
                    if len(attendance_log) !=0:
                        infirst_key = next(iter(attendance_log[0]))
                        outfirst_key = next(iter(attendance_log[-1]))

                        s['usersintime']=attendance_log[0][infirst_key]
                        s['usersouttime']=attendance_log[-1][outfirst_key]
                        if s['usersouttime'] == s['usersintime']:
                            s['usersouttime']= ''
                            
                    else:
                        s['usersintime']=''
                        s['usersouttime']=''

    
                    
                    if current_date_first_in_datetime =='':
                        if s['usersintime'] !='':
                            current_date_first_in_datetime=s['usersintime']
                    elif s['usersintime'] !='':
                        if datetime.datetime.strptime(current_date_first_in_datetime, '%Y-%m-%d %H:%M:%S') > datetime.datetime.strptime(s['usersintime'], '%Y-%m-%d %H:%M:%S'):
                            current_date_first_in_datetime=s['usersintime']

                    if current_date_last_out_datetime =='':
                        if s['usersouttime'] !='':
                            current_date_last_out_datetime=s['usersouttime']
                    elif s['usersouttime'] !='':
                        if datetime.datetime.strptime(current_date_last_out_datetime, '%Y-%m-%d %H:%M:%S') < datetime.datetime.strptime(s['usersouttime'], '%Y-%m-%d %H:%M:%S'):
                            current_date_last_out_datetime=s['usersouttime']
                    

                        
                    if current_date_last_out_datetime != '' and current_date_first_in_datetime !='':
                        if current_date_last_out_datetime.split(' ')[0] != current_date_first_in_datetime.split(' ')[0]:
                            current_date_last_out_datetime1 = datetime.datetime.strptime(str(current_date_last_out_datetime).split(' ')[0], '%Y-%m-%d')
                            current_date_first_in_datetime1 = datetime.datetime.strptime(str(current_date_first_in_datetime).split(' ')[0], '%Y-%m-%d')
                            extradaysdiff = current_date_last_out_datetime1 - current_date_first_in_datetime1
                            if extradaysdiff.days > 0:
                                extra_days='+'+str(extradaysdiff.days)
                        
                        
                return_dict['employeeId']=employeeId
                return_dict['fulldate']=datestr
                return_dict['shifts_list']=current_date_shift_list
                return_dict['shift']=new_current_date_shift_list[0]
                
                if len(new_current_date_shift_list) > 1:
                    shift_name=new_current_date_shift_list[0]['shiftname']
                    return_dict['shift']['shiftname'] = str(shift_name) +  ' <span class="text-danger"> +1 </span>'
                    new_current_date_shift_list[0]['swap']=''

                else:

                    new_current_date_shift_list[0]['swap']=''
                    
                
                if current_date_first_in_datetime !='':
                    return_dict['inTime']=str(current_date_first_in_datetime).split(' ')[1]
                    return_dict['inTime_date']=convertdate2(str(str(current_date_first_in_datetime).split(' ')[0]))
                else:
                    return_dict['inTime']=''
                    return_dict['inTime_date']=''
                    
                if current_date_last_out_datetime !='':
                    return_dict['outTime']=str(current_date_last_out_datetime).split(' ')[1]
                    return_dict['outTime_date']=convertdate2(str(str(current_date_last_out_datetime).split(' ')[0]))
                else:
                    return_dict['outTime']=''
                    return_dict['outTime_date']=''
                    
                return_dict['extra_days']=extra_days
                return_dict['Total']=total_working_hrs
                if return_dict['inTime'] !='':
                    return_dict['att_status']="P"
                else:
                    return_dict['att_status']="A"
            
                foundleaves = [
                    leave for leave in leavesserializer.data if datetime.datetime.strptime(leave['start_date'], '%Y-%m-%d').date() <= datetime.datetime.strptime(datestr, '%Y-%m-%d').date() <= datetime.datetime.strptime(leave['end_date'], '%Y-%m-%d').date()
                ]
                
                foundwfhs = [
                    wfh for wfh in wfhsserializer.data if datetime.datetime.strptime(wfh['start_date'], '%Y-%m-%d').date() <= datetime.datetime.strptime(datestr, '%Y-%m-%d').date() <= datetime.datetime.strptime(wfh['end_date'], '%Y-%m-%d').date()
                ]
                
                if foundwfhs:
                    return_dict['att_status']="WFH"
                if foundleaves:
                    return_dict['att_status']="L"
                if return_dict['fulldate'] in weeklyofflist:
                    return_dict['att_status']="WO"

                #holiday list
                holidatlist = Holidays.objects.filter(Active=True).order_by('id')
                serializer = holidaysSerializer(holidatlist, many=True)
                for i in serializer.data:
                    if return_dict['fulldate']==i['Date']:
                        return_dict['att_status']="H"
                        return_dict['reason']=i['Festival']


                attendance_report.append(return_dict)

            return Response ({  "data":attendance_report,   
                                "Photo": '',
                                "Designation": employee_serializer.data['DesignationId'],
                                "Name": employee_serializer.data['Firstname']+" "+employee_serializer.data['Lastname'],
                                "EmployeeId": employee_serializer.data['employeeId'],
                                "totalhours": 0,
                                "avghour": 0,
                                "year_total_leaves": 0,
                                "year_total_days": 0,
                                "response":{"n" : 1,"msg" : "success ","status" : "success"}})
            
        return Response ({      "data":[],   
                                "Photo": "",
                                "Designation":'',
                                "Name": '',
                                "EmployeeId": '',
                                "totalhours": 0,
                                "avghour": 0,
                                "year_total_leaves": 0,
                                "year_total_days": 0,
                                "response":{"n" : 0,"msg" : "month and year required ","status" : "failed"}})
        
    return Response ({          "data":[],   
                                "Photo": "",
                                "Designation":'',
                                "Name": '',
                                "EmployeeId": '',
                                "totalhours": 0,
                                "avghour": 0,
                                "year_total_leaves": 0,
                                "year_total_days": 0,
                                "response":{"n" : 0,"msg" : "user not found ","status" : "failed"}})



    
@api_view(['POST'])
def Addtyperules(request):
    data={}
    data['TypeId']= request.POST.get("TypeId")
    data['Shifts'] = request.POST.getlist("Shifts")
    data['CompOff'] = request.POST.get("CompOff")
    data['CompOffTime'] = request.POST.get('CompOffTime')
    data['CompOffValidity'] = request.POST.get('CompOffValidity')
    
    rulestype_obj = TypeRules.objects.filter(TypeId=data['TypeId'],is_active=True).first()
    if rulestype_obj is None:
        serializer = TypeRulesserializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response({'n':1,'msg':'Rules added successfully','status':'success','data':serializer.data})
        else:
            return Response({'n':0,'msg':'error occurs in serializer','status':'error','data':[]})
    else:
        return Response({'n':0,'msg':'Rules for type already exists','status':'error','data':[]})



@api_view(['GET'])
def gettyperules(request):
    rulestypeobj=TypeRules.objects.filter(is_active=True)
    if rulestypeobj is not None:
        serializer = TypeRulesserializer(rulestypeobj,many=True)
        for i in serializer.data:
            shiftlist=[]
            typeobj = employeetypemaster.objects.filter(id=i['TypeId']).first()
            i['Typename'] = typeobj.employeetype

            for s in i['Shifts']:
                shiftobj = ShiftMaster.objects.filter(id=s).first()
                shiftlist.append(shiftobj.shiftname)
            i['shiftlist'] = shiftlist

        return Response({'n':1,'msg':'rules of types found successfully!','status':'success','data':serializer.data})
    else:
        return Response({'n':0,'msg':'rules of types not found','status':'error','data':[]})


@api_view(['POST'])
def Updatetyperules(request):
    data={}

    rulestypeid =  request.POST.get("id")
    data['TypeId']= request.POST.get("TypeId")
    data['Shifts'] = json.loads(request.POST.get("Shifts"))
    data['CompOff'] = request.POST.get("CompOff")
    data['CompOffTime'] = request.POST.get('compofftime')
    data['CompOffValidity'] = request.POST.get('compoffvalidity')

    
    rulestype_obj = TypeRules.objects.filter(id = rulestypeid,is_active=True).first()
    if rulestype_obj  is not None:
        serializer = TypeRulesserializer(rulestype_obj,data=data,partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({'n':1,'msg':'Rules Updated successfully','status':'success','data':serializer.data})
        else:
            first_key, first_value = next(iter(serializer.errors.items()))
            return Response({'n':0,'msg':first_value[0],'status':'error','data':[]})
    else:
        return Response({'n':0,'msg':'Rules for type not found','status':'error','data':[]})
   
@api_view(['POST'])
def delete_employee_type_rules(request):
    data={}
    rulestypeid =  request.POST.get("id")
    data['is_active'] = False
    rulestype_obj = TypeRules.objects.filter(id = rulestypeid,is_active=True).first()
    if rulestype_obj  is not None:
        serializer = TypeRulesserializer(rulestype_obj,data=data,partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({'n':1,'msg':'Rules deleted successfully','status':'success','data':serializer.data})
        else:
            first_key, first_value = next(iter(serializer.errors.items()))
            return Response({'n':0,'msg':first_value[0],'status':'error','data':[]})
    else:
        return Response({'n':0,'msg':'Rules for type not found','status':'error','data':[]})
   
        
def get_event_by_time(events, start_time, end_time):
    for event in events:
        event_start = event.get("start")
        event_end = event.get("end")
        if event_start == start_time and event_end == end_time:
            return event
    return None
        
@csrf_exempt
@api_view(['POST'])
def getshiftevents(request):
    month=request.POST.get("month")
    year=request.POST.get("year")
    dayslist=json.loads(request.data.get('dayslist'))
    
    l1l2emplst_obj=EmployeeShiftDetails.objects.filter(is_active=True)
    l1l2serializers = L1L2Serializer(l1l2emplst_obj,many=True)
    team_list=list(l1l2serializers.data)
    
    Users_l1l2emplst_obj=Users.objects.filter(id__in=team_list,is_active=True).exclude(Q(employeeId__isnull=True)|Q(employeeId='None')|Q(employeeId=''))
    Users_l1l2serializers = UsersSerializeronlyattendance(Users_l1l2emplst_obj,many=True)
    Userlist=list(Users_l1l2serializers.data)
    



    if dayslist is not None  and dayslist !='' and dayslist !=[]:
        eventslist=[]
        holidatlist = Holidays.objects.filter(company_code=request.user.company_code,Active=True,Date__month=month,Date__year=year).order_by('id')
        serializer = holidaysSerializer(holidatlist, many=True)
        
        for h in serializer.data:
            event =  {
                        'title': "Holiday : " +h['Festival'], 
                        'start': h['Date'],
                        'end': h['Date'],
                        'color':"#00a306",
                        'textColor':"#ffffff",
                        'uniqueText':"All Day",
                        'date':h['Date'],
                        'shiftid':'',
                        
                    }
            eventslist.append(event)
        shifts_obj=ShiftMaster.objects.filter(is_active=True).exclude(intime='00:00',outtime='00:00')   
        shift_serializers=ShiftMasterSerializer(shifts_obj,many=True)   

        for date in dayslist:
            color_list=["#F4FCFF","#F4FCF5","#FFF7F4","#F9F4FF","#DEFCFC","#ffd3a9","#b5f3a4","#97E9D0","#b5f3a4","#FCE0A7","#FD4D4F","#FCE0A7", "#eccf00ab", "#7b00ec5e","#76a0e6"]
            textcolorlist=["#33BFFF","#29CC39","#FF6633","#8833FF","#3998A2","#c86b32","#369d18","#066c71","#0bc38c",'#cba250',"#5d0606","#9b6e13","#000","#000","#13459a"]

            color_counter=0

                

            
            for s in shift_serializers.data:
                s['bgcolor']=color_list[color_counter]
                s['textcolor']=textcolorlist[color_counter]
                color_counter=color_counter+1
                
                
                if s['intime']> s['outtime']:
                    date_obj = datetime.datetime.strptime(date, "%Y-%m-%d")
                    next_date_obj = date_obj + timedelta(days=1)
                    next_date_str = next_date_obj.strftime("%Y-%m-%d")
                    esddate=next_date_str
                else:
                    esddate=date
                    
                event =  {
                            'title': '', 
                            'start': date+'T'+s['intime']+':00',
                            'end': esddate+'T'+s['outtime']+':00',
                            'color':s['bgcolor'],
                            'textColor':s['textcolor'],
                            'uniqueText':" " +s['shiftname'].capitalize()  ,
                            'date':date,
                            'shiftid':s['id'],
                        }
                eventslist.append(event)
                
            for attendanceId in Userlist:
                user_obj=Users.objects.filter(employeeId=attendanceId,is_active=True).first()
                leave_objs = Leave.objects.filter(employeeId=user_obj.id,WorkFromHome=False,leave_status="Approved",start_date__month__gte=month,start_date__month__lte=month,start_date__year= year,Active=True).exclude(leave_status='Draft')
                leavesserializer=leaveserializer(leave_objs,many=True)
                wfh_objs = Leave.objects.filter(employeeId=user_obj.id,WorkFromHome=True,leave_status="Approved",start_date__month__gte=month,start_date__month__lte=month,start_date__year= year,Active=True).exclude(leave_status='Draft')
                wfhsserializer=leaveserializer(wfh_objs,many=True)
            
                new_current_date_shift_list=[]
                return_dict={}
                
                shiftdate = datetime.datetime.strptime(date, '%Y-%m-%d').date()
                tomarrow_date = shiftdate + datetime.timedelta(days=1)
                yesterday_date = shiftdate - datetime.timedelta(days=1)
                current_shiftallotment_objs=ShiftAllotment.objects.filter(attendanceId=attendanceId,date=str(date),is_active=True)
                current_shiftallotment_serializers=ShiftAllotmentshiftIdSerializer(current_shiftallotment_objs,many=True)
                current_shiftId_list=list(current_shiftallotment_serializers.data)
                current_shift_obj=ShiftMaster.objects.filter(id__in=current_shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime')
                current_shift_serializer=ShiftMasterSerializer(current_shift_obj,many=True)
                current_shiftlist=current_shift_serializer.data
                
                tomarrow_shiftallotment_objs=ShiftAllotment.objects.filter(attendanceId=attendanceId,date=str(tomarrow_date),is_active=True)
                tomarrow_shiftallotment_serializers=ShiftAllotmentshiftIdSerializer(tomarrow_shiftallotment_objs,many=True)
                tomarrow_shiftId_list=list(tomarrow_shiftallotment_serializers.data)
                tomarrow_shift_obj=ShiftMaster.objects.filter(id__in=tomarrow_shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime')
                tomarrow_shift_serializer=ShiftMasterSerializer(tomarrow_shift_obj,many=True)
                tomarrow_shiftlist=tomarrow_shift_serializer.data
                count=1
                
                for shift in current_shiftlist:
                    start_time = datetime.datetime.strptime(str(shiftdate) +' '+ shift['intime'], '%Y-%m-%d %H:%M')
                    start_time_before_2hrs = start_time - timedelta(hours=2)
                    
                    if shift['intime'] > shift['outtime']:
                        shift_end_date = shiftdate + timedelta(days=1)
                    else:
                        shift_end_date = shiftdate
                        
                    check_login_till=''
                    if len(current_shiftlist) >= count+1:
                    
                        check_next_shift_in = datetime.datetime.strptime(str(shiftdate) +' '+ current_shiftlist[count]['intime'], '%Y-%m-%d %H:%M')
                        check_login_till = check_next_shift_in - timedelta(hours=2)

                        
                    elif check_login_till=='':
                            
                        if len(tomarrow_shiftlist) >= 1:
                            
                            check_next_shift_in = datetime.datetime.strptime(str(tomarrow_date) +' '+ tomarrow_shiftlist[0]['intime'], '%Y-%m-%d %H:%M')
                            check_login_till = check_next_shift_in - timedelta(hours=2)
                        else:
                        
                            if_no_shift_check_shift=ShiftAllotment.objects.filter(attendanceId=attendanceId,date=str(tomarrow_date),is_active=True)
                            tomerrow_shiftlist=ShiftAllotmentshiftIdSerializer(if_no_shift_check_shift,many=True)
                            next_date_shiftId_list=list(tomerrow_shiftlist.data)
                            check_weekly_off=ShiftMaster.objects.filter(id__in=next_date_shiftId_list,intime='00:00',outtime='00:00',is_active=True).order_by('intime').first()
                            if check_weekly_off is not None:
                                
                                if shift['outtime'] < shift['intime']:
                            
                                    check_current_shift_out = datetime.datetime.strptime(str(tomarrow_date) +' '+ shift['outtime'], '%Y-%m-%d %H:%M')
                                    check_current_shift_out_time = check_current_shift_out + timedelta(hours=2)
                                    given_datetime = datetime.datetime.strptime(str(check_current_shift_out_time), '%Y-%m-%d %H:%M:%S')

                                    truncated_datetime_str = given_datetime.strftime('%Y-%m-%d %H:%M')

                                    
                                    shift_end_date_time=str(truncated_datetime_str)
                                else:
                                    shift_end_date_time=str(date) +' '+ '23:59'
                                    
                                check_login_till = datetime.datetime.strptime(shift_end_date_time, '%Y-%m-%d %H:%M')
                            else:
                            
                                check_login_till = datetime.datetime.strptime(str(tomarrow_date) +' '+ '07:30', '%Y-%m-%d %H:%M')
                    if shift['outtime'] < shift['intime']:
                        eddate=tomarrow_date
                    else:
                        eddate=date
                        
                    shift_data={
                        'shiftname':shift['shiftname'],
                        'indatetime':str(start_time_before_2hrs),
                        'outdatetime':str(check_login_till),
                        'intime':shift['intime'],
                        'outtime':shift['outtime'],
                        'in_date_T_time':str(date)+'T'+shift['intime']+':00',
                        'out_date_T_time':str(eddate)+'T'+shift['outtime']+':00',
                    }
                    
                    new_current_date_shift_list.append(shift_data)
                    count+=1

                if len(current_shiftlist) == 0:       
                
                    if_no_shift_check_shift=ShiftAllotment.objects.filter(attendanceId=attendanceId,date=str(date),is_active=True)
                    todays_shiftlist=ShiftAllotmentshiftIdSerializer(if_no_shift_check_shift,many=True)
                    toadys_shiftId_list=list(todays_shiftlist.data)
                    check_weekly_off=ShiftMaster.objects.filter(id__in=toadys_shiftId_list,intime='00:00',outtime='00:00',is_active=True).order_by('intime').first()
                    if check_weekly_off is not None:
                        get_previous_day_shift=ShiftAllotment.objects.filter(attendanceId=attendanceId,date=str(yesterday_date),is_active=True)
                        previous_day_shiftlist=ShiftAllotmentshiftIdSerializer(get_previous_day_shift,many=True)
                        previous_shiftId_list=list(previous_day_shiftlist.data)
                        get_previous_last_shift=ShiftMaster.objects.filter(id__in=previous_shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime').last()
                        if get_previous_last_shift is not None:
                            if get_previous_last_shift.outtime < get_previous_last_shift.intime:
                                shift_end_date_time=str(date) +' '+ str(get_previous_last_shift.outtime)
                            else:
                                shift_end_date_time=str(yesterday_date) +' '+ '23:59'
                                    
                            check_previous_shift_out = datetime.datetime.strptime(str(shift_end_date_time), '%Y-%m-%d %H:%M')
                            previous_shift_in_time = check_previous_shift_out + timedelta(hours=2)  
                        else:
                            previous_shift_in_time= str(date)+' 00:00:00'
                        get_net_day_shift=ShiftAllotment.objects.filter(attendanceId=attendanceId,date=str(tomarrow_date),is_active=True)
                        nest_day_shiftlist=ShiftAllotmentshiftIdSerializer(get_net_day_shift,many=True)
                        next_date_shiftId_list=list(nest_day_shiftlist.data)
                        get_next_day_first_shift=ShiftMaster.objects.filter(id__in=next_date_shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime').first()
                        if get_next_day_first_shift is not None:
                            check_next_shift_in = datetime.datetime.strptime(str(tomarrow_date) +' '+ str(get_next_day_first_shift.intime), '%Y-%m-%d %H:%M')
                            next_shift_in_time = check_next_shift_in - timedelta(hours=2)
                            shift_data={
                                'shiftname':str(check_weekly_off.shiftname),
                                'indatetime':str(previous_shift_in_time),
                                'outdatetime':str(next_shift_in_time),
                                'intime':check_weekly_off.intime,
                                'outtime':check_weekly_off.outtime,
                                'in_date_T_time':str(date),
                                'out_date_T_time':str(tomarrow_date),
                            }
                            new_current_date_shift_list.append(shift_data)
                            
                        else:
                            shift_data={
                                'shiftname':str(check_weekly_off.shiftname),
                                'indatetime':str(previous_shift_in_time),
                                'outdatetime':str(date)+' 23:59:59',
                                'intime':check_weekly_off.intime,
                                'outtime':check_weekly_off.outtime,
                                'in_date_T_time':str(date),
                                'out_date_T_time':str(tomarrow_date),
                            }
                            new_current_date_shift_list.append(shift_data)

                    else:
                        shift_data={
                            'shiftname':'General',
                            'indatetime':str(date)+' 07:30:00',
                            'outdatetime':str(tomarrow_date)+' 07:30:00',
                            'intime':'09:30',
                            'outtime':'18:30',
                            'in_date_T_time':str(date)+'T09:30:00',
                            'out_date_T_time':str(date)+'T18:30:00',
                        }
                        new_current_date_shift_list.append(shift_data)

                        
                current_date_first_in_datetime=''
                current_date_last_out_datetime=''
                for s in new_current_date_shift_list:
                    intime=''
                    intimedate=''
                    
                    
                    getallattendance = attendance.objects.filter(Q(employeeId=str(attendanceId),time__gte=s['indatetime'].split(' ')[1],date=s['indatetime'].split(' ')[0])|Q(employeeId=str(attendanceId),time__lte=str(s['outdatetime']).split(' ')[1],date=str(s['outdatetime']).split(' ')[0])).order_by('date','time')
                    
                    
                    attendance_serializer=attendanceserializer(getallattendance,many=True)
                
                    sorted_data = sorted(attendance_serializer.data, key=lambda x: (x['date'], x['time']))
                    
                    mindatetime = datetime.datetime.strptime(s['indatetime'], '%Y-%m-%d %H:%M:%S')
                    maxdatetime = datetime.datetime.strptime(s['outdatetime'], '%Y-%m-%d %H:%M:%S')
                
                    sorted_data = [entry for entry in sorted_data if (mindatetime <= datetime.datetime.strptime(entry['date'] +' '+entry['time'],'%Y-%m-%d %H:%M:%S') <= maxdatetime)]
                    if len(sorted_data) > 0:
                        intimedate=sorted_data[0]['date']
                        intime=str(sorted_data[0]['time'])
                        
                    if intimedate !='' and intimedate is not None:
                        user_sdt = datetime.datetime.strptime(str(intimedate) + ' ' + str(intime), '%Y-%m-%d %H:%M:%S')
                        shif_sdt = datetime.datetime.strptime(s['indatetime'].split(' ')[0] + ' ' + s['indatetime'].split(' ')[1], '%Y-%m-%d %H:%M:%S')
                        if user_sdt < shif_sdt :
                            intimedate=''
                            intime=''

                            
                    checkin_time = None
                    total_working_time = 0
                    attendance_log=[]

                    for index, entry in enumerate(sorted_data):

                        if entry['checkout']:
                            if entry['deviceId'] is not None and entry['deviceId'] !='':
                                attendance_type="Machine Checkout"
                                attendance_type_resaon=""
                            elif entry['Remote_Reason'] is not None and entry['Remote_Reason'] !='':
                                attendance_type="Remote Checkout"
                                attendance_type_resaon=entry['Remote_Reason']
                            else:
                                attendance_type="Online Checkout"
                                attendance_type_resaon=''                                

                                
                            attendance_log.append({'checkout':entry['date']+' '+entry['time']})
                            if checkin_time:
                                checkout_datetime = datetime.datetime.strptime(entry['date'] + ' ' + entry['time'], '%Y-%m-%d %H:%M:%S')
                                checkin_datetime = datetime.datetime.strptime(checkin_time, '%Y-%m-%d %H:%M:%S')
                                working_time = checkout_datetime - checkin_datetime
                                total_working_time += working_time.total_seconds()
                                checkin_time = None
                                
                        elif not entry['checkout']:
                            if entry['deviceId'] is not None and entry['deviceId'] !='':
                                attendance_type="Machine Checkin"
                                attendance_type_resaon=''
                            elif entry['Remote_Reason'] is not None and entry['Remote_Reason'] !='':
                                attendance_type="Remote Checkin"
                                attendance_type_resaon=entry['Remote_Reason']
                            else:
                                attendance_type="Online Checkout"
                                attendance_type_resaon=''  

                            checkin_time = entry['date'] + ' ' + entry['time']
                            attendance_log.append({'checkin':entry['date']+' '+entry['time']})
                    # If the last entry is check-in, consider checkout time as current shift end  time
                    if checkin_time and index == len(sorted_data) - 1:
                        
                        # check if the previous entry is also checkin or not if exist get the  difference betwnn  the current checkin  and  previous  checkin

                        if int(int(index)-1) >=0:
                            if sorted_data[index-1]['checkout']==False:
                                previous_checkin_date_time=sorted_data[index-1]['date']+ ' '+sorted_data[index-1]['time']
                                checkout_datetime = datetime.datetime.strptime(previous_checkin_date_time, '%Y-%m-%d %H:%M:%S')
                                checkin_datetime = datetime.datetime.strptime(checkin_time, '%Y-%m-%d %H:%M:%S')
                                working_time = checkin_datetime-checkout_datetime 
                                total_working_time += working_time.total_seconds()


                    # Convert total_working_time to hours, minutes, and seconds
                    hours, remainder = divmod(total_working_time, 3600)
                    minutes, seconds = divmod(remainder, 60)
                    s['total_hrs']=str(add_leading_zero(int(hours))) +':' +str(add_leading_zero(int(minutes))) +':'+str(add_leading_zero(int(seconds)))




  
                        
                    s['attendance_log']=attendance_log
                    if len(attendance_log) !=0:
                        
                        infirst_key = next(iter(attendance_log[0]))
                        outfirst_key = next(iter(attendance_log[-1]))
                        
                        s['usersintime']=attendance_log[0][infirst_key]
                        s['usersouttime']=attendance_log[-1][outfirst_key]
                        if s['usersouttime'] == s['usersintime']:
                            s['usersouttime']= ''
                            
                    else:
                        s['usersintime']=''
                        s['usersouttime']=''


                    
                    if current_date_first_in_datetime =='':
                        if s['usersintime'] !='':
                            current_date_first_in_datetime=s['usersintime']
                    elif s['usersintime'] !='':
                        if datetime.datetime.strptime(current_date_first_in_datetime, '%Y-%m-%d %H:%M:%S') > datetime.datetime.strptime(s['usersintime'], '%Y-%m-%d %H:%M:%S'):
                            current_date_first_in_datetime=s['usersintime']

                    if current_date_last_out_datetime =='':
                        if s['usersouttime'] !='':
                            current_date_last_out_datetime=s['usersouttime']
                    elif s['usersouttime'] !='':
                        if datetime.datetime.strptime(current_date_last_out_datetime, '%Y-%m-%d %H:%M:%S') < datetime.datetime.strptime(s['usersouttime'], '%Y-%m-%d %H:%M:%S'):
                            current_date_last_out_datetime=s['usersouttime']

                        

                        
                
                return_dict['shift']=new_current_date_shift_list[0]


                if current_date_first_in_datetime !='':
                    return_dict['checkintime']=str(current_date_first_in_datetime).split(' ')[1]
                else:
                    return_dict['checkintime']='--:--'

                    
                    
                if current_date_last_out_datetime !='':
                    return_dict['checkouttime']=str(current_date_last_out_datetime).split(' ')[1]
                else:
                    return_dict['checkouttime']='--:--'

                    
                    
                
                
                if return_dict['checkintime'] !='' and return_dict['checkintime'] !='--:--':
                    return_dict['att_status']="P"
                    check_indatetime = datetime.datetime.strptime(return_dict['shift']['indatetime'], "%Y-%m-%d %H:%M:%S")
                    check_usersintime = datetime.datetime.strptime(return_dict['shift']['usersintime'], "%Y-%m-%d %H:%M:%S")
                    grace_period = timedelta(hours=float(2.5))
                    allowed_time = check_indatetime + grace_period
                    if check_usersintime >= allowed_time:
                        return_dict['att_status']="late"
                else:
                    return_dict['att_status']="A"

                foundleaves = [
                    leave for leave in leavesserializer.data if datetime.datetime.strptime(leave['start_date'], '%Y-%m-%d').date() <= datetime.datetime.strptime(date, '%Y-%m-%d').date() <= datetime.datetime.strptime(leave['end_date'], '%Y-%m-%d').date()
                ]
                
                foundwfhs = [
                    wfh for wfh in wfhsserializer.data if datetime.datetime.strptime(wfh['start_date'], '%Y-%m-%d').date() <= datetime.datetime.strptime(date, '%Y-%m-%d').date() <= datetime.datetime.strptime(wfh['end_date'], '%Y-%m-%d').date()
                ]
                
                if foundwfhs:
                    return_dict['att_status']="WFH"
                if foundleaves:
                    return_dict['att_status']="L"
                    

                    
                    
                if return_dict['att_status']=='P':
                    imgstr='<img width="12px" height="12px" src="'+frontUrl+'static/assets/images/success.png">'
                elif return_dict['att_status']=='L':
                    imgstr='<img width="12px" height="12px" src="'+frontUrl+'static/assets/images/error.png">'
                elif return_dict['att_status']=='WFH':
                    imgstr='<img width="12px" height="12px" src="'+frontUrl+'static/assets/images/success.png">'
                elif return_dict['att_status']=='late':
                    imgstr='<img width="12px" height="12px" src="'+frontUrl+'static/assets/images/late.png">'
                else:
                    imgstr='<img width="12px" height="12px" src="'+frontUrl+'static/assets/images/error.png">'

                employeestr=imgstr +' '  + str(user_obj.Firstname +' '+user_obj.Lastname) 

                event = get_event_by_time(eventslist, return_dict['shift']['in_date_T_time'], return_dict['shift']['out_date_T_time'])
                if event:
                    event['title']=event['title']+'<br>'+employeestr
                # else:
                    # event =  {
                    #             'title': employeestr, 
                    #             'start': return_dict['shift']['in_date_T_time'],
                    #             'end': return_dict['shift']['out_date_T_time'],
                    #             'color':'#00a306',
                    #             'textColor':'#33BFFF',
                    #             'uniqueText':" " +return_dict['shift']['shiftname'].capitalize()  
                    #         }
                    # eventslist.append(event)


        return Response({'n':1,'msg':'list found successfully','status':'success','data':eventslist,'month':month,'attendance_report':'',})
    return Response({'n':0,'msg':'please give date','status':'failed','data':[]})
        
        
  

@csrf_exempt
@api_view(['POST'])
def testgetshiftevents(request):
    data={}
    data['month']=request.POST.get("month")
    data['year']=request.POST.get("year")
    if data['year'] is not None  and data['year'] !=''  and data['month'] is not None and data['month'] !='':
        
        try:
            conn = psycopg2.connect(database=env('DATABASE_NAME'), user= env('DATABASE_USER'),
                                    password=env('DATABASE_PASSWORD'), host=env('DATABASE_HOST'), port=env('DATABASE_PORT'), cursor_factory=RealDictCursor)
            cur = conn.cursor()
        except Exception as e:
            return Response({"n": 0, "Msg": "Could not connect to database", "Status": "Failed"})
        else:
            query = """ select u.id,u.company_code,
                        array (select concat('{ 
                                            "id":' ,sm.id,',
                                            "shiftname":"',sm."shiftname",'",
                                            "intime":"',sm.intime,'",
                                            "outtime":"',sm.outtime,'",	
                                            }')
                            from public."Users_shiftmaster" as sm where sm."is_active" = 'True'

                            order by sm.id desc 
                            ) as shift_master_data,
                        array (select concat('{ 
                                            "id":' ,sa.id,',
                                            "employee_name":"',sa.employee_name,'",
                                            "employeeId":"',sa."employeeId",'",
                                            "attendanceId":"',sa."attendanceId",'",
                                            "date":"',sa.date,'",
                                            "shift_name":"',sa.shift_name,'",					
                                            "shiftId":"',sa."shiftId",'",	
                                            "is_swaped":"',sa.is_swaped,'",
                                            "swaped_date":"',sa.swaped_date,'",
                                            "swaped_employeeId":"',sa."swaped_employeeId",'",
                                            "swaped_employee_name":"',sa.swaped_employee_name,'",
                                            "swaped_resaon":"',sa."swaped_resaon",'",
                                            "swaped_shiftId":"',sa."swaped_shiftId",'",
                                            "swaped_shift_name":"',sa.swaped_shift_name,'",
                                            }')
                            from public."Users_shiftallotment" as sa where sa."is_active" = true  order by sa.id desc 
                            ) as shiftallotment_data,
                            
                        array (select concat('{ 
                                            "id":' ,ul.id,',
                                            "employeeId":"',ul."employeeId",'",
                                            "start_date":"',ul.start_date,'",
                                            "end_date":"',ul.end_date,'",					
                                            "leavetype":"',ul."leavetype",'",	
                                            "WorkFromHome":"',ul."WorkFromHome",'",
                                            }')
                            from public."Users_leave" as ul where ul."Active" = true and ul."leave_status" = 'Approved' 

                            order by ul.id desc 
                            ) as leave_data,
                        array (select concat('{ 
                                            "id":' ,ua.id,',
                                            "employeeId":"',ua."employeeId",'",
                                            "date":"',ua.date,'",
                                            "time":"',ua.time,'",					
                                            "Month":"',ua."Month",'",	
                                            "Year":"',ua."Year",'",
                                            }')
                            from public."Users_attendance" as ua where ua."Month" = '2' and ua."Year" = '2024'

                            order by ua.id desc 
                            ) as attendance_data

                        from public."Users_users" as u

                        where u."is_active" = 'True' and u."company_code" = 'O001' limit 1 """
                        
            cur.execute(query)
            alldata = cur.fetchall()  

            shiftmaster_data_list = []
            shiftallotment_data_list = []
            leave_data_list = []
            attendance_data_list = []

            for i in alldata:
                

                for p in i['shift_master_data']:
                    shiftmaster_data_list.append(eval(p))   
                i['shift_master_data'] = shiftmaster_data_list

                for p in i['shiftallotment_data']:
                    
                   
                    shiftallotment_data_list.append(eval(p))   

                for s in shiftallotment_data_list:
                    if s['is_swaped'] == "t":
                        s['is_swaped'] = True
                    else:
                        s['is_swaped'] = False

                
                i['shiftallotment_data'] = shiftallotment_data_list
                
                for p in i['leave_data']:
                    
                    leave_data_list.append(eval(p))  

                for s in leave_data_list:
                    if s['WorkFromHome'] == "t":
                        s['WorkFromHome'] = True
                    else:
                        s['WorkFromHome'] = False 
                
                i['leave_data'] = leave_data_list

                for p in i['attendance_data']:
                    attendance_data_list.append(eval(p))   
                
                i['attendance_data'] = attendance_data_list
                
            
            last_day = calendar.monthrange(int(data['year']), int(data['month']))[1]
            date_list = [datetime.datetime(int(data['year']), int(data['month']), day).date() for day in range(1, last_day + 1)]

            

            
            
            for date in date_list:
                for shift in alldata[0]['shift_master_data'] :
                    for shiftallotment in alldata[0]['shiftallotment_data']:     
                        if str(shiftallotment['shiftId']) == str(shift['id']) and str(shiftallotment['date']) == str(shift['date']):
                            print('shiftallotment',shiftallotment['employee_name'],shiftallotment['date'])
                        
                        
                
            return Response({'n':1,'msg':'Data sound successfully','status':'success','data':alldata})

        
        
        
@api_view(['POST'])
@permission_classes((AllowAny,))
def attendanceexcelreport(request):

    holidaylist=[]

    d2=[]
    d6=[]

    year_input = request.data.get('year')
    month_input = request.data.get('month')
    search_employees=json.loads(request.data.get('search_employees'))
    
    if year_input not in [None, ""] and month_input not in [None, ""]:

      
 
        
        def get_date_list(year, month):
            # Get the last day of the month
            _, last_day = calendar.monthrange(year, month)
            
            # Generate a list of dates in the format 'yyyy-mm-dd'
            date_list = [f"{year}-{month:02d}-{day:02d}" for day in range(1, last_day + 1)]
            
            return date_list
        

        # Example usage:
       
        datelist = get_date_list(int(year_input), int(month_input))



        weeklyofflist=[]
        list1=[]
        #2nd 4th saturday of month
        dt=date(int(year_input), int(month_input),1)   # first day of month
        first_w=dt.isoweekday()  # weekday of 1st day of the month
        if(first_w==7): # if it is Sunday return 0 
            first_w=0
        saturday2=14-first_w
        dt1=date(int(year_input), int(month_input),saturday2)
        list1.append(str(dt1))
        saturday4=28-first_w
        dt1=date(int(year_input), int(month_input),saturday4)
        list1.append(str(dt1))  
        weeklyofflist=list1
            
        # all sundays of the month
        def allsundays(month_input):
            d = date(int(year_input), int(month_input), 1) # day 1st of month
            d += timedelta(days = 6 - d.weekday())  # First Sunday
            while d.month == month_input:
                yield d
                d += timedelta(days = 7)

        for d in allsundays(month_input):
            d=str(d)
            weeklyofflist.append(d)


           
        #yearly holiday list
        holidatlist = Holidays.objects.filter(Active=True).order_by('id')
        serializer = holidaysSerializer(holidatlist, many=True)
        for i in serializer.data:
            holiday=i['Date']
            holidaylist.append(holiday)
        
    

        empdata = Users.objects.filter(id__in=search_employees).exclude(Q(employeeId__isnull=True)|Q(employeeId='None')|Q(employeeId='')).distinct('id')
        mon_serializer = UsersSerializer(empdata,many=True)

       
        
        exclude_obj_list=[]
        for p in mon_serializer.data:
            leave_count=0
            holiday_count=0
            totalworkedhrs=timedelta()
            
            
            
                
            def get_dates_between(start_date, end_date):
                start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
                end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')

                dates = []
                current_date = start_date

                while current_date <= end_date:
                    dates.append(current_date.strftime('%Y-%m-%d'))
                    current_date += timedelta(days=1)

                return dates
        
            leave_objs = Leave.objects.filter(employeeId=p['id'],leave_status="Approved",Active=True)
            leavesserializer=leaveserializer(leave_objs,many=True)
            workfromhomelist=[]
            normalleave=[]

            for l in leavesserializer.data:
                if l['WorkFromHome']==True:
                    leavedatelist=get_dates_between(l['start_date'],l['end_date'],)
                    for d in leavedatelist:
                        if d not in workfromhomelist:
                            workfromhomelist.append(d)
                else:
                    normalleave.append(l)
                
            halfdayleave=[]
            fulldayleave=[]

            for leave in normalleave:
                if leave['leavetype'] != "Fullday":
                    leavedatelist=get_dates_between(leave['start_date'],leave['end_date'],)
                    for d in leavedatelist:
                        if d not in halfdayleave:
                            halfdayleave.append(d)
                else:
                    leavedatelist=get_dates_between(leave['start_date'],leave['end_date'],)
                    for d in leavedatelist:
                        if d not in fulldayleave:
                            fulldayleave.append(d)
        
            
            
            
            
            
            
            presentcount=0
            absentcount=0

            for j in datelist:
                year, month, day = j.split('-') 
                wname = datetime.date(int(year), int(month), int(day))
                day_name=wname.strftime("%A")
                       

                shift={}
                shift['shiftname']='General'
                shift['intime']='09:30'
                shift['outtime']='18:30'
                shift['swap']=''
                            
                            
                alloted_shift_obj=ShiftAllotment.objects.filter(employeeId=p['id'],attendanceId=p['employeeId'],date=j,is_active=True).first()
                if alloted_shift_obj is not None:
                    ShiftAllotment_serializer=ShiftAllotmentSerializer(alloted_shift_obj)
                    if ShiftAllotment_serializer.data['shiftId'] is not None and ShiftAllotment_serializer.data['shiftId'] !='':
                        shift_obj=ShiftMaster.objects.filter(id=ShiftAllotment_serializer.data['shiftId'],is_active=True).first()
                        if shift_obj is not None:
                            shift_obj_serializer=ShiftMasterSerializer(shift_obj)
                            shift['shiftname']=shift_obj_serializer.data['shiftname']
                            shift['intime']=shift_obj_serializer.data['intime']
                            shift['outtime']=shift_obj_serializer.data['outtime']
                            if ShiftAllotment_serializer.data['is_swaped'] == True:
                                check_swap_details=Shiftswap.objects.filter(id=ShiftAllotment_serializer.data['swap_request_id']).first()
                                if check_swap_details is not None:
                                    Shiftswap_serializer=CustomShiftswapSerializer(check_swap_details)
                                    if ShiftAllotment_serializer.data['swapper'] == True:
                                        shift['swap']='Swaped with '+ str(Shiftswap_serializer.data['second_employee_name'])
                                    elif ShiftAllotment_serializer.data['swapper'] == False:
                                        shift['swap']='Swaped with '+ str(Shiftswap_serializer.data['first_employee_name'])
                                    else:
                                        shift['swap']=''
                                else:
                                    shift['swap']=''
                            else:
                                shift['swap']=''
                                
                shift_starting_time = datetime.datetime.strptime(str(shift['intime']), "%H:%M")
                shift_ending_time = datetime.datetime.strptime(str(shift['outtime']), "%H:%M")








                def get_nearest_login_logout_times(employee_id, date,shift_start_time,shift_end_time,exclude_obj_list):
                    fnextday = datetime.datetime.strptime(date , "%Y-%m-%d").date() + timedelta(days=1)  
                    
                    start_time = datetime.datetime.strptime(str(shift_start_time), "%H:%M:%S")
                    end_time = datetime.datetime.strptime(str(shift_end_time), "%H:%M:%S")
                    
                    start_date=date
                    end_date=date
                    
                    if end_time < start_time:
                        end_date = datetime.datetime.strptime(date , "%Y-%m-%d").date() + timedelta(days=1) 

                    start_datetime = datetime.datetime.strptime(start_date + " " + shift_start_time, "%Y-%m-%d %H:%M:%S")
                    shift_allotment_obj_next_day=ShiftAllotment.objects.filter(date=fnextday,is_active=True,employeeId=p['id']).first()
                    nextdayshift={}
                    
                    
                    if shift_allotment_obj_next_day is not None:
                        shift_obj_next_day=ShiftMaster.objects.filter(id=shift_allotment_obj_next_day.shiftId,is_active=True).first()
                        if shift_obj_next_day is not None:
                            nextdayshift['inTime']=shift_obj_next_day.intime
                            nextdayshift['outTime']=shift_obj_next_day.outtime
                            nextdayshift['shiftname']=shift_obj_next_day.shiftname
                        else:
                            nextdayshift['shiftname']='General '
                            nextdayshift['inTime']='09:30'
                            nextdayshift['outTime']='18:30'
                    else:
                        nextdayshift['shiftname']='General '
                        nextdayshift['inTime']='09:30'
                        nextdayshift['outTime']='18:30'
                        
                        
                        
                    nextdayshift_start_time = datetime.datetime.strptime(str(nextdayshift['inTime']),"%H:%M")
                    nextday_datetime = datetime.datetime.strptime(str(fnextday) + " " + str(nextdayshift_start_time).split(" ")[1], "%Y-%m-%d %H:%M:%S")
                    start_datetime_before_2_hours = start_datetime - timedelta(hours=2)
                    nextday_start_datetime_before_2_hours=nextday_datetime - timedelta(hours=2)
                    
                    # print('start_datetime_before_2_hours',start_datetime_before_2_hours ,'--',nextday_start_datetime_before_2_hours)
                    start_date_before_2_hours = start_datetime_before_2_hours.strftime("%Y-%m-%d")
                    start_time_before_2_hours = start_datetime_before_2_hours.strftime("%H:%M:%S")
                    nextday_date_before_2_hours=nextday_start_datetime_before_2_hours.strftime("%Y-%m-%d")

                    
                    # Get the login time entry nearest to the shift starting time
                    if shift_start_time > shift_end_time:
                        login_obj=attendance.objects.filter(employeeId=employee_id,date=str(start_date_before_2_hours),time__gte=start_time_before_2_hours).order_by('time').first()
                        if login_obj is None:
                            login_obj=attendance.objects.filter(Q(employeeId=employee_id,date=str(start_date_before_2_hours),time__gte=start_time_before_2_hours)|Q(employeeId=employee_id,date=nextday_date_before_2_hours,time__lte=datetime.datetime.strptime(shift_end_time, "%H:%M:%S").time())).order_by('time').first()
                    else:
                        login_obj=attendance.objects.filter(employeeId=employee_id,date=str(start_date_before_2_hours),time__gte=start_time_before_2_hours).order_by('time').first()
                            
                    # login_obj = attendance.objects.filter(employeeId=employee_id,date=start_date_before_2_hours, time__gte=start_time_before_2_hours).order_by('time').first()
                    nearest_login_time = login_obj if login_obj else None
                    nextday_date_before_2_hours = nextday_start_datetime_before_2_hours.strftime("%Y-%m-%d")
                    nextday_time_before_2_hours = nextday_start_datetime_before_2_hours.strftime("%H:%M:%S")
                    # Get the logout time entry nearest to the shift ending time
                    logout_obj = attendance.objects.filter(employeeId=employee_id,date=nextday_date_before_2_hours, time__lt=nextday_time_before_2_hours).exclude(id__in=exclude_obj_list).order_by('time').last()
                    extra_days=''
                    if logout_obj is None:
                        logout_obj = attendance.objects.filter(employeeId=employee_id,date=start_date, time__gt=start_time_before_2_hours).exclude(id__in=exclude_obj_list).order_by('time').last()

                    if logout_obj is not None:
                        if str(start_date_before_2_hours) != str(logout_obj.date):
                            extra_days='+1'

                        
                        
                    nearest_logout_time = logout_obj if logout_obj else None

                    return nearest_login_time, nearest_logout_time, extra_days

                nearest_login_time, nearest_logout_time ,extra_days= get_nearest_login_logout_times(p['employeeId'], j,str(shift_starting_time).split(" ")[1],str(shift_ending_time).split(" ")[1],exclude_obj_list)









                if i in weeklyofflist:
                    if nearest_login_time is not None:
                        presentcount+=1
                        logint=nearest_login_time.time
                        login_date=nearest_login_time.date
                        exclude_obj_list.append(nearest_login_time.id)    

                        if nearest_logout_time is not None:
                            logoutt=nearest_logout_time.time
                            logout_date=nearest_logout_time.date
                            exclude_obj_list.append(nearest_logout_time.id)    

                        else:
                            logoutt=logint
                            logout_date=login_date

                        login_datetime = datetime.datetime.strptime(str(login_date) + " " + str(logint), "%Y-%m-%d %H:%M:%S")
                        logout_datetime = datetime.datetime.strptime(str(logout_date) + " " + str(logoutt), "%Y-%m-%d %H:%M:%S")
                        timeDiff = logout_datetime - login_datetime
                        total_time = str(timeDiff)

                        if logint==logoutt: 

                            total_time='00:00:00' 
                            logoutt=''
                        logint=convert_to_12_hour_format(logint)
                        logoutt=convert_to_12_hour_format(logoutt)
                        
                        if extra_days !='':
                            logint=str(logint) +" ("+ convertdate2(str(login_date))+")"
                            logoutt=str(logoutt)+ " ("+convertdate2(str(logout_date))+")"
                        sub_status='P'   
                    else:
                        logint="--:--"
                        logoutt="--:--"
                        total_time=""
                        sub_status='Weekly Off'   
                        
                    d1={
                        "employeeId":p['employeeId'],
                        "date":j.split("-")[2]+"-"+j.split("-")[1] +"-"+j.split("-")[0] ,
                        "inTime":logint,
                        "outTime":logoutt,
                        "Total":total_time,
                        "Day":day_name,
                        "shift":str(shift['shiftname']) ,
                        # "shift":str(shift['shiftname']) +' : ' +str(shift['intime'])+'-'+str(shift['outtime'])
                        'status':'',
                        'reason': "",
                        'sub_status':sub_status,
                        
                    }

                    d2.append(d1)

                elif j in fulldayleave:
                    leave_count+=1
                    logint="--:--"
                    logoutt="--:--"
                    total_time=""

                    d1={
                        "employeeId":p['employeeId'],
                        "date":j.split("-")[2]+"-"+j.split("-")[1] +"-"+j.split("-")[0] ,
                        "inTime":logint,
                        "outTime":logoutt,
                        "Total":total_time,
                        "Day":day_name,
                        "shift":str(shift['shiftname']) ,
                        # "shift":str(shift['shiftname']) +' : ' +str(shift['intime'])+'-'+str(shift['outtime'])
                        'status':'',
                        'reason': "Full Day Leave",
                        'sub_status':'Fullday Leave',
                        
                    }

                    d2.append(d1)

                elif j in halfdayleave:
                    leave_count+=0.5
                    if nearest_login_time is not None:
                        presentcount+=0.5
                        logint=nearest_login_time.time
                        login_date=nearest_login_time.date
                        exclude_obj_list.append(nearest_login_time.id)    

                        if nearest_logout_time is not None:
                            logoutt=nearest_logout_time.time
                            logout_date=nearest_logout_time.date
                            exclude_obj_list.append(nearest_logout_time.id)    

                        else:
                            logoutt=logint
                            logout_date=login_date

                        login_datetime = datetime.datetime.strptime(str(login_date) + " " + str(logint), "%Y-%m-%d %H:%M:%S")
                        logout_datetime = datetime.datetime.strptime(str(logout_date) + " " + str(logoutt), "%Y-%m-%d %H:%M:%S")
                        timeDiff = logout_datetime - login_datetime
                        total_time = str(timeDiff)

                        if logint==logoutt: 

                            total_time='00:00:00' 
                            logoutt=''
                        logint=convert_to_12_hour_format(logint)
                        logoutt=convert_to_12_hour_format(logoutt)
                        if extra_days !='':
                            logint=str(logint) +" ("+ convertdate2(str(login_date))+")"
                            logoutt=str(logoutt)+ " ("+convertdate2(str(logout_date))+")"

                        
                        
                    else:
                        logint="--:--"
                        logoutt="--:--"
                        total_time=""





                    d1={

                        "employeeId":p['employeeId'],
                        "date":j.split("-")[2]+"-"+j.split("-")[1] +"-"+j.split("-")[0],
                        "inTime":logint,
                        "outTime":logoutt,
                        "Total":total_time,
                        "Day":day_name,
                        "shift":str(shift['shiftname']) ,
                        'status':'',
                        'reason': "Half Day Leave",
                        'sub_status':'Halfday Leave',
                        
                    }

                    d2.append(d1)

                elif j in workfromhomelist:
                    if nearest_login_time is not None:
                        presentcount+=1
                        logint=nearest_login_time.time
                        login_date=nearest_login_time.date
                        exclude_obj_list.append(nearest_login_time.id)    

                        if nearest_logout_time is not None:
                            logoutt=nearest_logout_time.time
                            logout_date=nearest_logout_time.date
                            exclude_obj_list.append(nearest_logout_time.id)    

                        else:
                            logoutt=logint
                            logout_date=login_date

                        login_datetime = datetime.datetime.strptime(str(login_date) + " " + str(logint), "%Y-%m-%d %H:%M:%S")
                        logout_datetime = datetime.datetime.strptime(str(logout_date) + " " + str(logoutt), "%Y-%m-%d %H:%M:%S")
                        timeDiff = logout_datetime - login_datetime
                        total_time = str(timeDiff)
                        logint=convert_to_12_hour_format(logint)
                        logoutt=convert_to_12_hour_format(logoutt)
                        if extra_days !='':
                            logint=str(logint) +" ("+ convertdate2(str(login_date))+")"
                            logoutt=str(logoutt)+ " ("+convertdate2(str(logout_date))+")"
                        sub_status="P"
                    else:
                        logint="--:--"
                        logoutt="--:--"
                        sub_status=""
                        total_time=""

                    d1={

                        "employeeId":p['employeeId'],
                        "date":j.split("-")[2]+"-"+j.split("-")[1] +"-"+j.split("-")[0] ,
                        "inTime":logint,
                        "outTime":logoutt,
                        "Total":total_time,
                        "Day":day_name,
                        "shift":str(shift['shiftname']) ,
                        
                        'status':'',
                        'reason':'',
                        'sub_status':sub_status,
                        
                    }

                    d2.append(d1)

                elif nearest_login_time is not None:
                    presentcount+=1
                    logint=nearest_login_time.time
                    login_date=nearest_login_time.date
                    exclude_obj_list.append(nearest_login_time.id)   
                    
                    if nearest_logout_time is not None:
                        logoutt=nearest_logout_time.time
                        logout_date=nearest_logout_time.date
                        exclude_obj_list.append(nearest_logout_time.id)    
                    else:
                        logoutt=nearest_login_time.time
                        logout_date=login_date

                  
                    login_datetime = datetime.datetime.strptime(str(login_date) + " " + str(logint), "%Y-%m-%d %H:%M:%S")
                    logout_datetime = datetime.datetime.strptime(str(logout_date) + " " + str(logoutt), "%Y-%m-%d %H:%M:%S")
                    timeDiff = logout_datetime - login_datetime
                    totalworkedhrs += timeDiff
                    total_time = str(timeDiff)


                    if logint==logoutt: 
                        total_time='00:00:00' 
                        logoutt=''
                        logout_date=login_date
  
                        
                    logint=convert_to_12_hour_format(logint)
                    logoutt=convert_to_12_hour_format(logoutt)
                    
                    if extra_days !='':
                        logint=str(logint) +" ("+ convertdate2(str(login_date))+")"
                        logoutt=str(logoutt)+ " ("+convertdate2(str(logout_date))+")"

 
                        
              






                    d1={

                        
                        "employeeId":p['employeeId'],
                        "date":j.split("-")[2]+"-"+j.split("-")[1] +"-"+j.split("-")[0] ,
                        "inTime":logint,
                        "outTime":logoutt,
                        "Total":total_time,
                        "Day":day_name,
                        "shift":str(shift['shiftname']) ,
                        'status':'',
                        'reason':'',
                        'sub_status':'P',

                        
                    }
                    d2.append(d1)
                else:

                    if j in holidaylist:
                        holiday_count+=1
                        logint="--:--"
                        logoutt="--:--"
                        total_time=""

                    
                        d1={

                            "employeeId":p['employeeId'],
                            "date":j.split("-")[2]+"-"+j.split("-")[1] +"-"+j.split("-")[0] ,
                            "inTime":logint,
                            "outTime":logoutt,
                            "Total":total_time,
                            "Day":day_name,
                            "shift":str(shift['shiftname']) ,
                            
                            'status':'',
                            'reason':'Holiday',
                            'sub_status':'Holiday',
                        }
                        d2.append(d1)
                    else:
                        
                        absentcount+=1
                        logint="--:--"
                        logoutt="--:--"
                        total_time=""

                        d1={
                            "employeeId":p['employeeId'],
                            "date":j.split("-")[2]+"-"+j.split("-")[1] +"-"+j.split("-")[0] ,
                            "inTime":logint,
                            "outTime":logoutt,
                            "Total":total_time,
                            "Day":day_name,
                            "shift":str(shift['shiftname']) ,
                            'status':'',
                            'reason':'',
                            'sub_status':'-',
                            
                        }
                        d2.append(d1)
                
                    
                    
                    
            hours, remainder = divmod(totalworkedhrs.seconds, 3600)
            d5={ 
                'employeeId':p['employeeId'],
                'Leave':leave_count,
                'Holiday':holiday_count,
                'totalworkedhrs':str(totalworkedhrs.days) +' days ,'+str(hours)+ ' hours',
                'records':dict((v['date'],v) for v in d2).values() ,
                'present':presentcount,
                'absent':absentcount,
            }
            d6.append(d5)
            d2.clear()

        if os.path.exists("static/excel/monthlydata.xlsx"):
            os.remove("static/excel/monthlydata.xlsx")
            workbook = xlsxwriter.Workbook('static/excel/monthlydata.xlsx')
            workbook.close()
        else:
            workbook = xlsxwriter.Workbook('static/excel/monthlydata.xlsx')
            workbook.close()
            
            
        if len(mon_serializer.data) >1:
            Allshiftattendancereport(d6)
        else:
            shiftattendancereport(d6)
            
            
        return Response({"response":{"data":d6,"url":imageUrl + '/static/excel/monthlydata.xlsx', "n":2 ,"msg" : "Excel report generated successfully ","status":"success"}})  
    else:
        return Response({"response":{"url": '', "n":0 ,"msg" : "Operation failed please provide month and year","status":"error"}})  
        

        
        
        


def shiftattendancereport(d6):
    workbook = xlsxwriter.Workbook('static/excel/monthlydata.xlsx')
    wb=load_workbook('static/excel/monthlydata.xlsx')
    printdate=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    sheet=wb.worksheets[0]
    k=11
    z=5

    for p in d6:
        sheet.cell(row=1,column=12).value='Monthly Status Report'
        sheet.cell(row=3,column=1).value='Company:'
        sheet.cell(row=3,column=2).value='Zentro'
        sheet.cell(row=3,column=26).value='printed On:'
        sheet.cell(row=3,column=27).value=printdate

        sheet.cell(row=z,column=1).value='Empcode'
        sheet.cell(row=z,column=2).value=p['employeeId']
        empobj=Users.objects.filter(employeeId=p['employeeId']).first()
        sheet.cell(row=z+1,column=1).value='Employee Name'
        sheet.cell(row=z+1,column=2).value=empobj.Firstname+" "+empobj.Lastname
        sheet.cell(row=z+2,column=1).value='Leave'
        sheet.cell(row=z+2,column=2).value=p['Leave']
        sheet.cell(row=z+3,column=1).value='Holidays'
        sheet.cell(row=z+3,column=2).value=p['Holiday']
        sheet.cell(row=z+4,column=1).value='Total Working HRS'
        sheet.cell(row=z+4,column=2).value=p['totalworkedhrs']

        for row in range(z, z+5):
            sheet.cell(row=row,column=1).font = openpyxl.styles.Font(bold=True)



        j=1
        sheet.cell(row=k,column=1).value='Date'
        sheet.cell(row=k,column=2).value='Day'
        sheet.cell(row=k,column=3).value='In Time'
        sheet.cell(row=k,column=4).value='Out Time'
        sheet.cell(row=k,column=5).value='Status'
        sheet.cell(row=k,column=6).value='Reason'
        sheet.cell(row=k,column=7).value='Shift'
        for col in range(1, 8):
            sheet.cell(row=k,column=col).font = openpyxl.styles.Font(bold=True)
        
        for i in p['records']:
            sheet.cell(row=k+1,column=j).value=i['date']
            sheet.cell(row=k+1,column=j+1).value=i['Day']
            if 'shifts_list' in i:
                if len(i['shifts_list'])>1:
                    for s in i['shifts_list']:
                        sheet.cell(row=k+1,column=j+2).value=s['usersintime']
                        sheet.cell(row=k+1,column=j+3).value=s['usersouttime']
                        sheet.cell(row=k+1,column=j+4).value=i['status']
                        sheet.cell(row=k+1,column=j+5).value=i['reason']
                        sheet.cell(row=k+1,column=j+6).value=s['shiftname']
                        sheet.cell(row=k+1,column=j+7).value=i['shiftswap']
                        k+=1
                else:
                    sheet.cell(row=k+1,column=j+2).value=i['inTime']
                    sheet.cell(row=k+1,column=j+3).value=i['outTime']
                    sheet.cell(row=k+1,column=j+4).value=i['status']
                    sheet.cell(row=k+1,column=j+5).value=i['reason']
                    sheet.cell(row=k+1,column=j+6).value=i['shift']['shiftname']
                    if 'shiftswap' in i:
                        sheet.cell(row=k+1,column=j+7).value=i['shiftswap']   
                    k+=1 
            else:
                sheet.cell(row=k+1,column=j+2).value=i['inTime']
                sheet.cell(row=k+1,column=j+3).value=i['outTime']
                sheet.cell(row=k+1,column=j+4).value=i['status']
                sheet.cell(row=k+1,column=j+5).value=i['reason']
                sheet.cell(row=k+1,column=j+6).value=i['shift']
                if 'shiftswap' in i:
                    sheet.cell(row=k+1,column=j+7).value=i['shiftswap']   
                k+=1 
            # Add borders to the cells
            for col in range(1, 8):

                for row in range(k, k + 2):
                    cell = sheet.cell(row=row, column=j + col - 1)
                    cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                        right=openpyxl.styles.Side(style='thin'),
                                                        top=openpyxl.styles.Side(style='thin'),
                                                        bottom=openpyxl.styles.Side(style='thin'))

                  
        k+=8
        z=k-6
    wb.save('static/excel/monthlydata.xlsx')

def Allshiftattendancereport(d6):
    workbook = xlsxwriter.Workbook('static/excel/monthlydata.xlsx')
    wb=load_workbook('static/excel/monthlydata.xlsx')
    printdate=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if len(wb.sheetnames) < 2:
        wb.create_sheet(title='Sheet2')  # Add a new sheet named 'Sheet2'
        wb.save('static/excel/monthlydata.xlsx')
    sheet=wb.worksheets[0]
    sheet1=wb.worksheets[1]
    
    
    
    
    ko=11
    zo=5

    for q in d6:
        sheet1.cell(row=1,column=12).value='Monthly Status Report'
        sheet1.cell(row=3,column=1).value='Company:'
        sheet1.cell(row=3,column=2).value='Zentro'
        sheet1.cell(row=3,column=26).value='printed On:'
        sheet1.cell(row=3,column=27).value=printdate

        sheet1.cell(row=zo,column=1).value='Empcode'
        sheet1.cell(row=zo,column=2).value=q['employeeId']
        empobj=Users.objects.filter(employeeId=q['employeeId']).first()
        sheet1.cell(row=zo+1,column=1).value='Employee Name'
        sheet1.cell(row=zo+1,column=2).value=empobj.Firstname+" "+empobj.Lastname
        sheet1.cell(row=zo+2,column=1).value='Leave'
        sheet1.cell(row=zo+2,column=2).value=q['Leave']
        sheet1.cell(row=zo+3,column=1).value='Holidays'
        sheet1.cell(row=zo+3,column=2).value=q['Holiday']
        sheet1.cell(row=zo+4,column=1).value='Total Working HRS'
        sheet1.cell(row=zo+4,column=2).value=q['totalworkedhrs']

        for row in range(zo, zo+5):
            sheet1.cell(row=row,column=1).font = openpyxl.styles.Font(bold=True)



        jo=1
        sheet1.cell(row=ko,column=1).value='Date'
        sheet1.cell(row=ko,column=2).value='Day'
        sheet1.cell(row=ko,column=3).value='In Time'
        sheet1.cell(row=ko,column=4).value='Out Time'
        sheet1.cell(row=ko,column=5).value='Status'
        sheet1.cell(row=ko,column=6).value='Reason'
        sheet1.cell(row=ko,column=7).value='Shift'
        for col in range(1, 8):
            sheet1.cell(row=ko,column=col).font = openpyxl.styles.Font(bold=True)

        for i in q['records']:
            sheet1.cell(row=ko+1,column=jo).value=i['date']
            sheet1.cell(row=ko+1,column=jo+1).value=i['Day']
            if len(i['shifts_list'])>1:
                for s in i['shifts_list']:
                    sheet1.cell(row=ko+1,column=jo+2).value=s['usersintime']
                    sheet1.cell(row=ko+1,column=jo+3).value=s['usersouttime']
                    sheet1.cell(row=ko+1,column=jo+4).value=i['status']
                    sheet1.cell(row=ko+1,column=jo+5).value=i['reason']
                    sheet1.cell(row=ko+1,column=jo+6).value=s['shiftname']
                    sheet1.cell(row=ko+1,column=jo+7).value=i['shiftswap']
                    ko+=1
            else:
                sheet1.cell(row=ko+1,column=jo+2).value=i['inTime']
                sheet1.cell(row=ko+1,column=jo+3).value=i['outTime']
                sheet1.cell(row=ko+1,column=jo+4).value=i['status']
                sheet1.cell(row=ko+1,column=jo+5).value=i['reason']
                sheet1.cell(row=ko+1,column=jo+6).value=i['shift']['shiftname']
                sheet1.cell(row=ko+1,column=jo+7).value=i['shiftswap']   
                ko+=1 
            # Add borders to the cells
            for col in range(1, 8):
                for row in range(ko, ko + 2):
                    cell = sheet1.cell(row=row, column=jo + col - 1)
                    cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                        right=openpyxl.styles.Side(style='thin'),
                                                        top=openpyxl.styles.Side(style='thin'),
                                                        bottom=openpyxl.styles.Side(style='thin'))

                  
        ko+=8
        zo=ko-6
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    k=6
    r=6
    z=5
    rowcounter=6
    sheet.cell(row=1,column=12).value='Monthly Status Report'
    sheet.cell(row=3,column=1).value='Company:'
    sheet.cell(row=3,column=2).value='Zentro'
    sheet.cell(row=3,column=26).value='printed On:'
    sheet.cell(row=3,column=27).value=printdate

    sheet.cell(row=k,column=1).value='Date'
    sheet.cell(row=k,column=2).value='Day'
    sheet.cell(row=k,column=1).font = openpyxl.styles.Font(bold=True)
    sheet.cell(row=k,column=2).font = openpyxl.styles.Font(bold=True)
    department_name=''
    department_cell_count=1
    colcounter=2
    counter=2            

    color_list = [ "FFA090E6", "000000", "00ADFF","0400FF", "FF0000", "FF00E0", "FFFFCCAA", "FFFF7777", "FFFFCCAA", "FFAACCFF", "FF5ECC5E", "FFA090E6"]
    colorintervale=0
    current_col = counter + 1  # Track the current column for merging

    for p in d6:

        empobj=Users.objects.filter(employeeId=p['employeeId']).first()
        user_serializer = UsersSerializer(empobj)
        if user_serializer.data['DepartmentID'] !=[]:
            
            if department_name != str(user_serializer.data['DepartmentID'][0]):
                department_cell_count=1

                colorintervale += 1
                department_color = color_list[colorintervale]
                sheet.cell(row=z, column=counter+1).fill = PatternFill(start_color=department_color, end_color=department_color, fill_type="solid")
                sheet.cell(row=z, column=counter+1).font = Font(color="FFFFFF",bold=True)
                sheet.cell(row=z, column=counter+1).value = str(user_serializer.data['DepartmentID'][0])
                department_name = str(user_serializer.data['DepartmentID'][0])                
                current_col += 1

            elif department_name == str(user_serializer.data['DepartmentID'][0]):
                department_cell_count+=1
                department_color = color_list[colorintervale]
                sheet.cell(row=z, column=counter+1).fill = PatternFill(start_color=department_color, end_color=department_color, fill_type="solid")
                sheet.cell(row=z, column=counter+1).font = Font(color="FFFFFF",bold=True)




        #         sheet.cell(row=z, column=counter+1).value =  str(user_serializer.data['DepartmentID'][0])
        #         sheet.merge_cells(start_row=z, start_column=counter-1, end_row=z, end_column=counter+1)
        #         merged_cell = sheet.cell(row=z, column=counter+1)
        #         merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                sheet.merge_cells(start_row=z, start_column=current_col, end_row=z, end_column=current_col)
                merged_cell = sheet.cell(row=z, column=current_col)
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
    
            else :
                colorintervale += 1
                department_cell_count=1
                department_color = color_list[colorintervale]
                sheet.cell(row=z, column=counter+1).fill = PatternFill(start_color=department_color, end_color=department_color, fill_type="solid")
                sheet.cell(row=z, column=counter+1).font = Font(color="FFFFFF",bold=True)

                sheet.cell(row=z, column=counter+1).value = str(user_serializer.data['DepartmentID'][0])
                department_name = str(user_serializer.data['DepartmentID'][0])
                
        else:
            department_cell_count=1

            colorintervale += 1
            department_color = color_list[colorintervale]
            sheet.cell(row=z, column=counter+1).fill = PatternFill(start_color=department_color, end_color=department_color, fill_type="solid")
            sheet.cell(row=z, column=counter+1).font = Font(color="FFFFFF",bold=True)
            sheet.cell(row=z, column=counter+1).value = ''
            department_name = ''

        

        
        
        sheet.cell(row=k,column=counter+1).value=empobj.Firstname+" " +empobj.Lastname
        sheet.cell(row=k,column=counter+1).font = openpyxl.styles.Font(bold=True)

        sheet.cell(row=39,column=1).value='Total Billable Days'
        sheet.cell(row=39,column=counter+1).value=p['present']
        sheet.cell(row=40,column=1).value='Absent Days'
        sheet.cell(row=40,column=counter+1).value=p['absent']

        j=1
        r=6
         
        rowcounter=6
        for i in p['records']:
            
            sheet.cell(row=r+1,column=j).value=i['date']
            sheet.cell(row=r+1,column=j+1).value=i['Day']
            sheet.cell(row=r+1,column=counter+1).value=i['sub_status']

                    
            # Add borders to the cells
            # for col in range(1, 8):



            r+=1  
            rowcounter+=1
        
        
        counter+=1
        

    for row in range(5, 38):
        for col in range(1, counter + 1):
            cell = sheet.cell(row=row, column=col)
            cell.border = Border(left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin'))
 
    wb.save('static/excel/monthlydata.xlsx')


def attendance_calender_report(d6):
    workbook = xlsxwriter.Workbook('static/excel/monthlydata.xlsx')
    wb=load_workbook('static/excel/monthlydata.xlsx')
    printdate=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    sheet=wb.worksheets[0]
    k=11
    z=5

    for p in d6:
        sheet.cell(row=1,column=12).value='Monthly Status Report'
        sheet.cell(row=3,column=1).value='Company:'
        sheet.cell(row=3,column=2).value='Zentro'
        sheet.cell(row=3,column=26).value='printed On:'
        sheet.cell(row=3,column=27).value=printdate

        sheet.cell(row=z,column=1).value='Empcode'
        sheet.cell(row=z,column=2).value=p['employeeId']
        empobj=Users.objects.filter(employeeId=p['employeeId']).first()
        sheet.cell(row=z+1,column=1).value='Employee Name'
        sheet.cell(row=z+1,column=2).value=empobj.Firstname+" "+empobj.Lastname
        sheet.cell(row=z+2,column=1).value='Leave'
        sheet.cell(row=z+2,column=2).value=p['Leave']
        sheet.cell(row=z+3,column=1).value='Holidays'
        sheet.cell(row=z+3,column=2).value=p['Holiday']
        sheet.cell(row=z+4,column=1).value='Total Working HRS'
        sheet.cell(row=z+4,column=2).value=p['totalworkedhrs']

        for row in range(z, z+5):
            sheet.cell(row=row,column=1).font = openpyxl.styles.Font(bold=True)



        j=1
        sheet.cell(row=k,column=1).value='Date'
        sheet.cell(row=k,column=2).value='Day'
        sheet.cell(row=k,column=3).value='In Time'
        sheet.cell(row=k,column=4).value='Out Time'
        sheet.cell(row=k,column=5).value='Status'
        sheet.cell(row=k,column=6).value='Reason'
        for col in range(1, 7):
            sheet.cell(row=k,column=col).font = openpyxl.styles.Font(bold=True)

        for i in p['records']:
            sheet.cell(row=k+1,column=j).value=i['date']
            sheet.cell(row=k+1,column=j+1).value=i['Day']
            sheet.cell(row=k+1,column=j+2).value=i['inTime']
            sheet.cell(row=k+1,column=j+3).value=i['outTime']
            sheet.cell(row=k+1,column=j+4).value=i['status']
            sheet.cell(row=k+1,column=j+5).value=i['reason']


            # Add borders to the cells
            for col in range(1, 7):

                for row in range(k, k + 2):
                    cell = sheet.cell(row=row, column=j + col - 1)
                    cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'),
                                                        right=openpyxl.styles.Side(style='thin'),
                                                        top=openpyxl.styles.Side(style='thin'),
                                                        bottom=openpyxl.styles.Side(style='thin'))
            k+=1 

                  
        k+=8
        z=k-6
    wb.save('static/excel/monthlydata.xlsx')

    


@api_view(['POST'])
def getemployeeallotedshift(request):
    
    employee=request.POST.get('employee')
    shiftdate=request.POST.get('shiftdate')

    
    shiftallotmentobj=ShiftAllotment.objects.filter(is_active=True,date=shiftdate,employeeId=employee)
    if shiftallotmentobj is not None:
        serializer = ShiftAllotmentSerializer(shiftallotmentobj,many=True)
        return Response({'n':1,'msg':'Shift allotment found successfully!','status':'success','data':serializer.data})
    else:
        return Response({'n':0,'msg':' Alloted shift not found','status':'error','data':[]})
    

@api_view(['POST'])
def swapshift(request):
    
    employee1=request.POST.get('employee1')
    employee2=request.POST.get('employee2')
    
    shiftdate1=request.POST.get('shiftdate1')
    shiftdate2=request.POST.get('shiftdate2')
    
    shift1=request.POST.get('shift1')
    shift2=request.POST.get('shift2')
    

    
    shiftallotmentobj1=ShiftAllotment.objects.filter(is_active=True,date=shiftdate1,employeeId=employee1,shiftId=shift1).first()
    if shiftallotmentobj1 is not None:
        serializer1 = ShiftAllotmentSerializer(shiftallotmentobj1)
        shiftallotmentobj2=ShiftAllotment.objects.filter(is_active=True,date=shiftdate2,employeeId=employee2,shiftId=shift2).first()
        if shiftallotmentobj2 is not None:
            serializer2 = ShiftAllotmentSerializer(shiftallotmentobj2)
            firstempshiftname=serializer1.data['shift_name']
            firstempshiftid=serializer1.data['shiftId']
            firstempswaped_employeeId =serializer1.data['employeeId']
            firstempswaped_employee_name =serializer1.data['employee_name']
            firstempswaped_date =serializer1.data['date']
            firstempswaped_shiftId =serializer1.data['shiftId']
            firstempswaped_shift_name =serializer1.data['shift_name']
            
            
            
            
            shiftallotmentobj1.shift_name=serializer2.data['shift_name']
            shiftallotmentobj1.shiftId=serializer2.data['shiftId']
            shiftallotmentobj1.swaped_resaon =''
            shiftallotmentobj1.swaped_employeeId =serializer2.data['employeeId']
            shiftallotmentobj1.swaped_employee_name =serializer2.data['employee_name']
            shiftallotmentobj1.swaped_date =serializer2.data['date']
            shiftallotmentobj1.swaped_shiftId =serializer2.data['shiftId']
            shiftallotmentobj1.swaped_shift_name =serializer2.data['shift_name']
            shiftallotmentobj1.is_swaped=True
            shiftallotmentobj1.save()
            
            shiftallotmentobj2.shift_name=firstempshiftname
            shiftallotmentobj2.shiftId=firstempshiftid
            shiftallotmentobj2.swaped_resaon=''
            shiftallotmentobj2.swaped_employeeId=firstempswaped_employeeId
            shiftallotmentobj2.swaped_employee_name=firstempswaped_employee_name
            shiftallotmentobj2.swaped_date=firstempswaped_date
            shiftallotmentobj2.swaped_shiftId=firstempswaped_shiftId
            shiftallotmentobj2.swaped_shift_name=firstempswaped_shift_name
            shiftallotmentobj2.is_swaped=True
            shiftallotmentobj2.save()
            
            
            return Response({'n':1,'msg':'Shift Swaped Successfully','status':'success','data':[]})
            
        else:
            return Response({'n':0,'msg':'Second employee alloted shift not found','status':'error','data':[]})
    else:
        return Response({'n':0,'msg':'First employee alloted shift not found','status':'error','data':[]})
    
    
    
    

@api_view(['GET'])
def getdevicelist(request):

    user_deivice_obj = DeviceVerification.objects.filter(is_active=True).distinct('userid')
    seri=DeviceVerificationSerializer(user_deivice_obj,many=True)
    nowlist=[]
    for i in seri.data:
        userobj=Users.objects.filter(id=i['userid']).first()
        if userobj is not None:
            i['username']=userobj.Firstname + ' ' + userobj.Lastname
        else:
            i['username']=''
            
            

        nowlist.append(i)
    return Response ({
            "data": nowlist,
            "response":{
                "n" : 0,
                "msg" : "success",
                "status" : "success"
            }
            }) 
    


            

@api_view(['GET'])
def punch_getdata(request):
    employeeId = request.user.employeeId
    current_date = datetime.date.today()
    yesterday_date = current_date - datetime.timedelta(days=1)
    current_datetime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    intime=''
    outtime=''
    intimedate=''
    outtimedate=''
    userobj = Users.objects.filter(employeeId = employeeId).first()
    
    
    if userobj is not None:
        userser = UserSerializer(userobj)
        locationid = userser.data['locationId']
        if locationid is not None and locationid != "":
            locationobj = Location.objects.filter(id=int(locationid)).first()
            if locationobj is not None:
                location = locationobj.LocationName
            else:
                location = ""
        else:
            location = ""
    else:
        location = ""
    if userobj is not None:       
        check_user_type=userobj.employeetype
        if check_user_type is not None:
            check_last_checkin = attendance.objects.filter(employeeId=str(employeeId),date=str(current_date),checkout=False).order_by('time').last()
            check_last_checkout=None
            if check_last_checkin is not None:
                check_last_checkout = attendance.objects.filter(employeeId=str(employeeId),time__gt=check_last_checkin.time,date=str(current_date),checkout=True).order_by('time').last()
                if check_last_checkout is not None:
                    print("today allow him to checkin")
                else: 
                    print("today allow him to checkout")
            else:
                check_last_checkin = attendance.objects.filter(employeeId=str(employeeId),date=str(yesterday_date),checkout=False).order_by('time').last()
                if check_last_checkin is not None:
                    check_last_checkout = attendance.objects.filter(Q(employeeId=str(employeeId),time__gt=check_last_checkin.time,date=str(yesterday_date),checkout=True)|Q(employeeId=str(employeeId),date=str(current_date),checkout=True)).order_by('date','time').last()
                    if check_last_checkout is not None:
                        print("yesterday allow him to checkin")
                    else: 
                        print("yesterday allow him to checkout")
                else:
                    print("allow him to checkin")
            
            # 1 disable
            # 0 enable
            
            if check_last_checkin is not None :
                
                intime = check_last_checkin.time
                intimedate=check_last_checkin.date
                punchout = 0
                get_data=1
                outtime = ''
                outtimedate=''
            else:
                get_data=0
                punchout = 1
                intime = ''
                intimedate=''
                outtime = ''
                outtimedate=''
               
               
            if check_last_checkout is not None:
                
                outtime = check_last_checkout.time
                outtimedate=check_last_checkout.date
                punchout = 1
                get_data=0
                # intime = ''
                # intimedate=''
            else:
                if check_last_checkin is not None :
                    punchout = 0
                    get_data=1
                
                    outtime = ''
                    outtimedate=''
                else:
                    punchout = 1
                    get_data=0
                    outtime = ''
                    outtimedate=''
                    

            currentshiftname=''
            currentshiftstarttime=''
            currentshiftendtime=''
            currentshiftstartdate=''
            currentshiftenddate=''
            todays_shiftallotment_objs=ShiftAllotment.objects.filter(attendanceId=employeeId,date=str(current_date),is_active=True)
            todays_shiftallotment_serializers=ShiftAllotmentshiftIdSerializer(todays_shiftallotment_objs,many=True)
            shiftId_list=list(todays_shiftallotment_serializers.data)
            shift_obj=ShiftMaster.objects.filter(id__in=shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime')
            shift_serializer=ShiftMasterSerializer(shift_obj,many=True)
            
            
            # check for todays shift
            todays_runnningshift=gettodaysshift(shift_serializer.data,str(current_date))
            if todays_runnningshift['n'] == 1:
                currentshiftname=todays_runnningshift['data']['shiftname']
                currentshiftstarttime=str(todays_runnningshift['data']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                currentshiftendtime=str(todays_runnningshift['data']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                currentshiftstartdate=str(todays_runnningshift['data']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                currentshiftenddate=str(todays_runnningshift['data']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                
            elif todays_runnningshift['last_runingshift']['shiftstarttime'] !='':
                currentshiftname=todays_runnningshift['last_runingshift']['shiftname']
                currentshiftstarttime=str(todays_runnningshift['last_runingshift']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                currentshiftendtime=str(todays_runnningshift['last_runingshift']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                currentshiftstartdate=str(todays_runnningshift['last_runingshift']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                currentshiftenddate=str(todays_runnningshift['last_runingshift']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
            
            else:
                yesterday_shiftallotment_objs=ShiftAllotment.objects.filter(attendanceId=employeeId,date=str(yesterday_date),is_active=True)
                yesterday_shiftallotment_serializers=ShiftAllotmentshiftIdSerializer(yesterday_shiftallotment_objs,many=True)
                shiftId_list=list(yesterday_shiftallotment_serializers.data)
                shift_obj=ShiftMaster.objects.filter(id__in=shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime')
                yesterday_shift_serializer=ShiftMasterSerializer(shift_obj,many=True)
                    
                
                yesterday_runnningshift=gettodaysshift(yesterday_shift_serializer.data,str(yesterday_date))
                if yesterday_runnningshift['n'] == 1:
                    currentshiftname=yesterday_runnningshift['data']['shiftname']
                    currentshiftstarttime=str(yesterday_runnningshift['data']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                    currentshiftendtime=str(yesterday_runnningshift['data']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                    currentshiftstartdate=str(yesterday_runnningshift['data']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                    currentshiftenddate=str(yesterday_runnningshift['data']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                    
                elif yesterday_runnningshift['last_runingshift']['shiftstarttime'] !='':
                    currentshiftname=yesterday_runnningshift['last_runingshift']['shiftname']
                    currentshiftstarttime=str(yesterday_runnningshift['last_runingshift']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                    currentshiftendtime=str(yesterday_runnningshift['last_runingshift']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                    currentshiftstartdate=str(yesterday_runnningshift['last_runingshift']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                    currentshiftenddate=str(yesterday_runnningshift['last_runingshift']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                else:
                    currentshiftname='General'
                    currentshiftstarttime='07:30:00'
                    currentshiftendtime='18:30:00'
                    currentshiftstartdate=str(current_date)
                    currentshiftenddate=str(current_date)

            shiftdetails={
                        "shiftname":currentshiftname,
                        "shiftstarttime":currentshiftstarttime,
                        "shiftendtime":currentshiftendtime,
                        "shiftstartdate":currentshiftstartdate,
                        "shiftenddate":currentshiftenddate,
                    }
            

            getallattendance = attendance.objects.filter(Q(employeeId=str(employeeId),time__gte=currentshiftstarttime,date=str(currentshiftstartdate))|Q(employeeId=str(employeeId),time__lte=str(current_datetime).split(' ')[1],date=str(current_datetime).split(' ')[0])).order_by('date','time')
            
            # getallattendance = getallattendance.filter(time__gte=currentshiftstarttime).order_by('date','time')
            
            attendance_serializer=attendanceserializer(getallattendance,many=True)
            
            sorted_data = sorted(attendance_serializer.data, key=lambda x: (x['date'], x['time']))
            
            mindate = datetime.datetime.strptime(currentshiftstartdate, '%Y-%m-%d')
            mintime = datetime.datetime.strptime(currentshiftstarttime, '%H:%M:%S').time()

            sorted_data = [entry for entry in sorted_data if (datetime.datetime.strptime(entry['date'],'%Y-%m-%d').date() > mindate.date() or (datetime.datetime.strptime(entry['date'],'%Y-%m-%d').date() == mindate.date() and datetime.datetime.strptime(entry['time'], '%H:%M:%S').time() > mintime))]

            if len(sorted_data) > 0:
                intimedate=sorted_data[0]['date']
                intime=str(sorted_data[0]['time'])
                
            if intimedate !='' and intimedate is not None:
                user_sdt = datetime.datetime.strptime(str(intimedate) + ' ' + str(intime), '%Y-%m-%d %H:%M:%S')
                shif_sdt = datetime.datetime.strptime(str(currentshiftstartdate) + ' ' + str(currentshiftstarttime), '%Y-%m-%d %H:%M:%S')
                if user_sdt < shif_sdt :
                    intimedate=''
                    intime=''
                    outtime=''
                    outtimedate=''
                    punchout = 1
                    get_data=0
                    
            checkin_time = None
            total_working_time = 0
            for index, entry in enumerate(sorted_data):
                if entry['checkout']:
                    if checkin_time:
                        checkout_datetime = datetime.datetime.strptime(entry['date'] + ' ' + entry['time'], '%Y-%m-%d %H:%M:%S')
                        checkin_datetime = datetime.datetime.strptime(checkin_time, '%Y-%m-%d %H:%M:%S')
                        working_time = checkout_datetime - checkin_datetime
                        total_working_time += working_time.total_seconds()
                        checkin_time = None
                elif not entry['checkout']:
                    checkin_time = entry['date'] + ' ' + entry['time']


            # If the last entry is check-in, consider checkout time as current time
            if checkin_time and index == len(sorted_data) - 1:
                print("yes")
                checkout_datetime = datetime.datetime.now()
                checkin_datetime = datetime.datetime.strptime(checkin_time, '%Y-%m-%d %H:%M:%S')
                working_time = checkout_datetime - checkin_datetime
                total_working_time += working_time.total_seconds()


            # Convert total_working_time to hours, minutes, and seconds
            hours, remainder = divmod(total_working_time, 3600)
            minutes, seconds = divmod(remainder, 60)


            
            # remove this code when mutiple checkin and checkout is approved
            # start-------------------------------
            # if intime !='':
            #     get_data=1
            #     if outtime !='':
            #         punchout=1
            # else:
            #     get_data=0
            #     if outtime !='':
            #         punchout=1
            # end-------------------------------
            
            
            current_shift_start_datetime = datetime.datetime.strptime(shiftdetails['shiftstartdate'] + ' ' + shiftdetails['shiftstarttime'], '%Y-%m-%d %H:%M:%S')
            current_shift_end_datetime = datetime.datetime.strptime(shiftdetails['shiftenddate'] + ' ' + shiftdetails['shiftendtime'], '%Y-%m-%d %H:%M:%S')
            shift_total_working_hrs = current_shift_end_datetime - current_shift_start_datetime
            shift_total_working_hrs -= timedelta(hours=2)  # Subtracting 2 hours from the duration
            print("check_user_type",check_user_type.id)
            rules_obj=TypeRules.objects.filter(TypeId=check_user_type.id,is_active=True).first()

                
                
            total_hrs=str(add_leading_zero(int(hours))) +':' +str(add_leading_zero(int(minutes))) +':'+str(add_leading_zero(int(seconds)))
            print("shift_total_working_hrs",shift_total_working_hrs)   
            print("total_hrs",total_hrs)
            if get_data == 1:
                if total_hrs >=str(shift_total_working_hrs):
                    if rules_obj is not None:
                        shift_total_compoff_hrs = shift_total_working_hrs + timedelta(hours=int(rules_obj.CompOffTime)) 
                        if total_hrs > str(shift_total_compoff_hrs):
                            print("compoff is granted")
                        else:
                            print("not fitted for compoff")
                    else:
                        print("no rules found")
                else:
                    print("working hrs is not filled")
            else:
                print("not checked in")
                    
           
            return Response ({
                'indatetime':str(dd_mm_yyyy(str(intimedate))) + ' ' + str(intime),
                'outdatetime':str(dd_mm_yyyy(str(outtimedate))) + ' ' + str(outtime),
                'total_hrs':total_hrs,
                'shift_total_working_hrs':str(shift_total_working_hrs),
                'data':get_data,
                'intime':intime,
                'outtime':outtime,
                'intimedate':intimedate,
                'outtimedate':outtimedate,
                'punchout':punchout,
                'location':location,
                'hours':int(hours),
                'minutes':int(minutes),
                'seconds':int(seconds),
                    "response":{
                        "n" : 1,
                        "msg" : "pass",
                        "status" : "success"
                    }
                    })
            
        else:
            check_last_checkin = attendance.objects.filter(employeeId=str(employeeId),date=str(current_date),checkout=False).order_by('time').last()
            check_last_checkout=None
            if check_last_checkin is not None:
                check_last_checkout = attendance.objects.filter(employeeId=str(employeeId),time__gt=check_last_checkin.time,date=str(current_date),checkout=True).order_by('time').last()
                if check_last_checkout is not None:
                    print("today allow him to checkin")
                else: 
                    print("today allow him to checkout")
            else:
                check_last_checkin = attendance.objects.filter(employeeId=str(employeeId),date=str(yesterday_date),checkout=False).order_by('time').last()
                if check_last_checkin is not None:
                    check_last_checkout = attendance.objects.filter(Q(employeeId=str(employeeId),time__gt=check_last_checkin.time,date=str(yesterday_date),checkout=True)|Q(employeeId=str(employeeId),date=str(current_date),checkout=True)).order_by('date','time').last()
                    if check_last_checkout is not None:
                        print("yesterday allow him to checkin")
                    else: 
                        print("yesterday allow him to checkout")
                else:
                    print("allow him to checkin")
            
            # 1 disable
            # 0 enable
            
            if check_last_checkin is not None :
                
                intime = check_last_checkin.time
                intimedate=check_last_checkin.date
                punchout = 0
                get_data=1
                outtime = ''
                outtimedate=''
            else:
                get_data=0
                punchout = 1
                intime = ''
                intimedate=''
                outtime = ''
                outtimedate=''
               
               
            # print("get_data1",get_data) 
            if check_last_checkout is not None:
                outtime = check_last_checkout.time
                outtimedate=check_last_checkout.date
                punchout = 1
                get_data=0
                # intime = ''
                # intimedate=''
            else:
                if check_last_checkin is not None :
                    punchout = 0
                    get_data=1
                
                    outtime = ''
                    outtimedate=''
                else:
                    punchout = 1
                    get_data=0
                    outtime = ''
                    outtimedate=''
                    
            


            currentshiftname=''
            currentshiftstarttime=''
            currentshiftendtime=''
            currentshiftstartdate=''
            currentshiftenddate=''
            todays_shiftallotment_objs=ShiftAllotment.objects.filter(attendanceId=employeeId,date=str(current_date),is_active=True)
            todays_shiftallotment_serializers=ShiftAllotmentshiftIdSerializer(todays_shiftallotment_objs,many=True)
            shiftId_list=list(todays_shiftallotment_serializers.data)
            shift_obj=ShiftMaster.objects.filter(id__in=shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime')
            shift_serializer=ShiftMasterSerializer(shift_obj,many=True)
            
            
            # check for todays shift
            todays_runnningshift=gettodaysshift(shift_serializer.data,str(current_date))
            if todays_runnningshift['n'] == 1:
                currentshiftname=todays_runnningshift['data']['shiftname']
                currentshiftstarttime=str(todays_runnningshift['data']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                currentshiftendtime=str(todays_runnningshift['data']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                currentshiftstartdate=str(todays_runnningshift['data']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                currentshiftenddate=str(todays_runnningshift['data']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                
            elif todays_runnningshift['last_runingshift']['shiftstarttime'] !='':
                currentshiftname=todays_runnningshift['last_runingshift']['shiftname']
                currentshiftstarttime=str(todays_runnningshift['last_runingshift']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                currentshiftendtime=str(todays_runnningshift['last_runingshift']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                currentshiftstartdate=str(todays_runnningshift['last_runingshift']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                currentshiftenddate=str(todays_runnningshift['last_runingshift']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
            
            else:
                yesterday_shiftallotment_objs=ShiftAllotment.objects.filter(attendanceId=employeeId,date=str(yesterday_date),is_active=True)
                yesterday_shiftallotment_serializers=ShiftAllotmentshiftIdSerializer(yesterday_shiftallotment_objs,many=True)
                shiftId_list=list(yesterday_shiftallotment_serializers.data)
                shift_obj=ShiftMaster.objects.filter(id__in=shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime')
                yesterday_shift_serializer=ShiftMasterSerializer(shift_obj,many=True)
                    
                
                yesterday_runnningshift=gettodaysshift(yesterday_shift_serializer.data,str(yesterday_date))
                if yesterday_runnningshift['n'] == 1:
                    currentshiftname=yesterday_runnningshift['data']['shiftname']
                    currentshiftstarttime=str(yesterday_runnningshift['data']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                    currentshiftendtime=str(yesterday_runnningshift['data']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                    currentshiftstartdate=str(yesterday_runnningshift['data']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                    currentshiftenddate=str(yesterday_runnningshift['data']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                    
                elif yesterday_runnningshift['last_runingshift']['shiftstarttime'] !='':
                    currentshiftname=yesterday_runnningshift['last_runingshift']['shiftname']
                    currentshiftstarttime=str(yesterday_runnningshift['last_runingshift']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                    currentshiftendtime=str(yesterday_runnningshift['last_runingshift']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                    currentshiftstartdate=str(yesterday_runnningshift['last_runingshift']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                    currentshiftenddate=str(yesterday_runnningshift['last_runingshift']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                else:
                    currentshiftname='General'
                    currentshiftstarttime='07:30:00'
                    currentshiftendtime='18:30:00'
                    currentshiftstartdate=str(current_date)
                    currentshiftenddate=str(current_date)

            shiftdetails={
                        "shiftname":currentshiftname,
                        "shiftstarttime":currentshiftstarttime,
                        "shiftendtime":currentshiftendtime,
                        "shiftstartdate":currentshiftstartdate,
                        "shiftenddate":currentshiftenddate,
                    }
            

            getallattendance = attendance.objects.filter(Q(employeeId=str(employeeId),time__gte=currentshiftstarttime,date=str(currentshiftstartdate))|Q(employeeId=str(employeeId),time__lte=str(current_datetime).split(' ')[1],date=str(current_datetime).split(' ')[0])).order_by('date','time')
            
            # getallattendance = getallattendance.filter(time__gte=currentshiftstarttime).order_by('date','time')
            
            attendance_serializer=attendanceserializer(getallattendance,many=True)
            
            sorted_data = sorted(attendance_serializer.data, key=lambda x: (x['date'], x['time']))
            
            mindate = datetime.datetime.strptime(currentshiftstartdate, '%Y-%m-%d')
            mintime = datetime.datetime.strptime(currentshiftstarttime, '%H:%M:%S').time()

            sorted_data = [entry for entry in sorted_data if (datetime.datetime.strptime(entry['date'],'%Y-%m-%d').date() > mindate.date() or (datetime.datetime.strptime(entry['date'],'%Y-%m-%d').date() == mindate.date() and datetime.datetime.strptime(entry['time'], '%H:%M:%S').time() > mintime))]

            if len(sorted_data) > 0:
                intimedate=sorted_data[0]['date']
                intime=str(sorted_data[0]['time'])
                
            if intimedate !='' and intimedate is not None:
                user_sdt = datetime.datetime.strptime(str(intimedate) + ' ' + str(intime), '%Y-%m-%d %H:%M:%S')
                shif_sdt = datetime.datetime.strptime(str(currentshiftstartdate) + ' ' + str(currentshiftstarttime), '%Y-%m-%d %H:%M:%S')
                if user_sdt < shif_sdt :
                    intimedate=''
                    intime=''
                    outtime=''
                    outtimedate=''
                    punchout = 1
                    get_data=0
                    
            checkin_time = None
            total_working_time = 0
            for index, entry in enumerate(sorted_data):
                if entry['checkout']:
                    if checkin_time:
                        checkout_datetime = datetime.datetime.strptime(entry['date'] + ' ' + entry['time'], '%Y-%m-%d %H:%M:%S')
                        checkin_datetime = datetime.datetime.strptime(checkin_time, '%Y-%m-%d %H:%M:%S')
                        working_time = checkout_datetime - checkin_datetime
                        total_working_time += working_time.total_seconds()
                        checkin_time = None
                elif not entry['checkout']:
                    checkin_time = entry['date'] + ' ' + entry['time']

            # If the last entry is check-in, consider checkout time as current time
            if checkin_time and index == len(sorted_data) - 1:
                checkout_datetime = datetime.datetime.now()
                checkin_datetime = datetime.datetime.strptime(checkin_time, '%Y-%m-%d %H:%M:%S')
                working_time = checkout_datetime - checkin_datetime
                total_working_time += working_time.total_seconds()


            # Convert total_working_time to hours, minutes, and seconds
            hours, remainder = divmod(total_working_time, 3600)
            minutes, seconds = divmod(remainder, 60)


            
            # 1 checkout True disable checkin
            # 0 checkout True enable checkin
            
            
            # remove this code when mutiple checkin and checkout is approved
            # start-------------------------------
            # if intime !='':
            #     get_data=1
            #     if outtime !='':
            #         punchout=1
            # else:
            #     get_data=0
            #     if outtime !='':
            #         punchout=1
            # end-------------------------------

                    
                    
            return Response ({
                'indatetime':str(dd_mm_yyyy(str(intimedate))) + ' ' + str(intime),
                'outdatetime':str(dd_mm_yyyy(str(outtimedate))) + ' ' + str(outtime),
                'total_hrs':str(add_leading_zero(int(hours))) +':' +str(add_leading_zero(int(minutes))) +':'+str(add_leading_zero(int(seconds))),
                'data':get_data,
                'intime':intime,
                'outtime':outtime,
                'intimedate':intimedate,
                'outtimedate':outtimedate,
                'punchout':punchout,
                'location':location,
                'hours':int(hours),
                'minutes':int(minutes),
                'seconds':int(seconds),
                    "response":{
                        "n" : 1,
                        "msg" : "pass",
                        "status" : "success"
                    }
            })
            

    else:
        
        return Response ({
                    'indatetime':'',
                    'outdatetime':'',
                    'data':'',
                    'intime':'',
                    'outtime':'',
                    'intimedate':'',
                    'outtimedate':'',
                    'punchout':'',
                    'location':'',
                        "response":{
                            "n" : 0,
                            "msg" : "user not found",
                            "status" : "errror"
                        }
                        })
      


@api_view(['POST'])    
@permission_classes((AllowAny,))
def emp_monthly_shift_details(request):
    list1=[]
    weeklyofflist=[]
    holidaylist=[]
    d2=[]
    d6=[]
    statuslist=[]

    month=request.data.get("month")
    month=int(month)
    year=request.data.get("year")
    empuserID=request.data.get('UserId')
    if empuserID is not None and empuserID !="":
        empid=Users.objects.filter(id=empuserID).first()
        if empid is None:
            return Response ({"data":[],"response":{"n" : 0,"msg" : "user  not found ","status" : "failed"}})
            
        employee_id=empid.employeeId 
        month_number=month
        year=int(year)
        month = month_number
        year = year
        number_of_days = calendar.monthrange(year, month)[1]
        first_date = date(year, month, 1)
        last_date = date(year, month, number_of_days)
        delta = last_date - first_date
        dayslist=[(first_date + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(delta.days + 1)]

                

        #2nd 4th saturday of month
        year=year 
        month=month_number
        dt=date(year,month,1)   # first day of month
        first_w=dt.isoweekday()  # weekday of 1st day of the month
        if(first_w==7): # if it is Sunday return 0 
            first_w=0
        saturday2=14-first_w
        dt1=date(year,month,saturday2)
        list1.append(str(dt1))
        saturday4=28-first_w
        dt1=date(year,month,saturday4)
        list1.append(str(dt1))  
        weeklyofflist=list1
            
        # all sundays of the month
        def allsundays(month):
            d = date(year,month, 1) # day 1st of month
            d += timedelta(days = 6 - d.weekday())  # First Sunday
            while d.month == month:
                yield d
                d += timedelta(days = 7)

        for d in allsundays(month):
            d=str(d)
            weeklyofflist.append(d)
        

        #holiday list
        holidatlist = Holidays.objects.filter(Active=True).order_by('id')
        serializer = holidaysSerializer(holidatlist, many=True)
        for i in serializer.data:
            holiday=i['Date']
            holidaylist.append(holiday) 

        def get_dates_between(start_date, end_date):
            start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d')
            end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d')

            dates = []
            current_date = start_date

            while current_date <= end_date:
                dates.append(current_date.strftime('%Y-%m-%d'))
                current_date += timedelta(days=1)

            return dates
        
        leave_objs = Leave.objects.filter(employeeId=empuserID,leave_status="Approved",Active=True)
        leavesserializer=leaveserializer(leave_objs,many=True)
        workfromhomelist=[]
        normalleave=[]

        for i in leavesserializer.data:
            if i['WorkFromHome']==True:
                datelist=get_dates_between(i['start_date'],i['end_date'],)
                for d in datelist:
                    if d not in workfromhomelist:
                        workfromhomelist.append(d)
            else:
                normalleave.append(i)



        
        halfdayleave=[]
        fulldayleave=[]

        for leave in normalleave:
            if leave['leavetype'] != "Fullday":
                datelist=get_dates_between(leave['start_date'],leave['end_date'],)
                for d in datelist:
                    if d not in halfdayleave:
                        halfdayleave.append(d)
            else:
                datelist=get_dates_between(leave['start_date'],leave['end_date'],)
                for d in datelist:
                    if d not in fulldayleave:
                        fulldayleave.append(d)
        



        # added
        att_month = attendance.objects.filter(employeeId=employee_id,date__month=month_number,date__year=year).order_by('time').first()
        if att_month is not None:
            
            att_month = attendance.objects.filter(employeeId=employee_id,date__month=month_number,date__year=year)
            serializer = attendanceserializer(att_month,many=True)
            A_count=0
            holiday_count=0

            exclude_obj_list=[]
            for j in dayslist:
                year, month, day = j.split('-') 
                wname = datetime.date(int(year), int(month), int(day))
                day_name=wname.strftime("%a")
                shift={}
                shift['shiftname']='No Shift'
                shift['intime']='09:30'
                shift['outtime']='18:30'
                shift['swap']=''
                alloted_shift_obj=ShiftAllotment.objects.filter(employeeId=empuserID,attendanceId=employee_id,date=j,is_active=True).first()
                if alloted_shift_obj is not None:
                    ShiftAllotment_serializer=ShiftAllotmentSerializer(alloted_shift_obj)
                    if ShiftAllotment_serializer.data['shiftId'] is not None and ShiftAllotment_serializer.data['shiftId'] !='':
                        shift_obj=ShiftMaster.objects.filter(id=ShiftAllotment_serializer.data['shiftId'],is_active=True).first()
                        if shift_obj is not None:
                            shift_obj_serializer=ShiftMasterSerializer(shift_obj)
                            shift['shiftname']=shift_obj_serializer.data['shiftname']
                            shift['intime']=shift_obj_serializer.data['intime']
                            shift['outtime']=shift_obj_serializer.data['outtime']
                            if ShiftAllotment_serializer.data['is_swaped'] == True:
                                check_swap_details=Shiftswap.objects.filter(id=ShiftAllotment_serializer.data['swap_request_id']).first()
                                if check_swap_details is not None:
                                    Shiftswap_serializer=CustomShiftswapSerializer(check_swap_details)
                                    if ShiftAllotment_serializer.data['swapper'] == True:
                                        shift['swap']='Swaped with '+ str(Shiftswap_serializer.data['second_employee_name'])
                                    elif ShiftAllotment_serializer.data['swapper'] == False:
                                        shift['swap']='Swaped with '+ str(Shiftswap_serializer.data['first_employee_name'])
                                    else:
                                        shift['swap']=''
                                else:
                                    shift['swap']=''
                            else:
                                shift['swap']=''
                            

                nextday = datetime.datetime.strptime(j , "%Y-%m-%d").date() + timedelta(days=1)            
                shift_starting_time = datetime.datetime.strptime(str(shift['intime']), "%H:%M")
                shift_ending_time = datetime.datetime.strptime(str(shift['outtime']), "%H:%M")


                def get_nearest_login_logout_times(employee_id, date,shift_start_time,shift_end_time,exclude_obj_list):
                    fnextday = datetime.datetime.strptime(date , "%Y-%m-%d").date() + timedelta(days=1)    
                    start_time = datetime.datetime.strptime(str(shift_start_time), "%H:%M:%S")
                    end_time = datetime.datetime.strptime(str(shift_end_time), "%H:%M:%S")
                    start_date=date
                    end_date=date
                    if end_time < start_time:
                        end_date = datetime.datetime.strptime(date , "%Y-%m-%d").date() + timedelta(days=1) 

                    start_datetime = datetime.datetime.strptime(start_date + " " + shift_start_time, "%Y-%m-%d %H:%M:%S")
                    shift_allotment_obj_next_day=ShiftAllotment.objects.filter(date=fnextday,is_active=True,employeeId=empid.id).first()
                    nextdayshift={}
                    
                    
                    if shift_allotment_obj_next_day is not None:
                        shift_obj_next_day=ShiftMaster.objects.filter(id=shift_allotment_obj_next_day.shiftId,is_active=True).first()
                        if shift_obj_next_day is not None:
                            nextdayshift['inTime']=shift_obj_next_day.intime
                            nextdayshift['outTime']=shift_obj_next_day.outtime
                            nextdayshift['shiftname']=shift_obj_next_day.shiftname
                        else:
                            nextdayshift['shiftname']='No Shift'
                            nextdayshift['inTime']='09:30'
                            nextdayshift['outTime']='18:30'
                            
                            
                    else:
                        nextdayshift['shiftname']='No Shift'
                        nextdayshift['inTime']='09:30'
                        nextdayshift['outTime']='18:30'
                        
                        
                        
                    nextdayshift_start_time = datetime.datetime.strptime(str(nextdayshift['inTime']),"%H:%M")
                    nextday_datetime = datetime.datetime.strptime(str(fnextday) + " " + str(nextdayshift_start_time).split(" ")[1], "%Y-%m-%d %H:%M:%S")
                    start_datetime_before_2_hours = start_datetime - timedelta(hours=2)
                    nextday_start_datetime_before_2_hours=nextday_datetime - timedelta(hours=2)
                    
                    # print('start_datetime_before_2_hours',start_datetime_before_2_hours ,'--',nextday_start_datetime_before_2_hours)
                    start_date_before_2_hours = start_datetime_before_2_hours.strftime("%Y-%m-%d")
                    start_time_before_2_hours = start_datetime_before_2_hours.strftime("%H:%M:%S")
                    nextday_date_before_2_hours=nextday_start_datetime_before_2_hours.strftime("%Y-%m-%d")
                    
                    if shift_start_time > shift_end_time:
                        login_obj=attendance.objects.filter(employeeId=employee_id,date=str(start_date_before_2_hours),time__gte=start_time_before_2_hours).order_by('time').first()
                        if login_obj is None:
                            login_obj=attendance.objects.filter(Q(employeeId=employee_id,date=str(start_date_before_2_hours),time__gte=start_time_before_2_hours)|Q(employeeId=employee_id,date=nextday_date_before_2_hours,time__lte=datetime.datetime.strptime(shift_end_time, "%H:%M:%S").time())).order_by('time').first()
                    else:
                        login_obj=attendance.objects.filter(employeeId=employee_id,date=str(start_date_before_2_hours),time__gte=start_time_before_2_hours).order_by('time').first()
                            
                    # Get the login time entry nearest to the shift starting time
                    # login_obj = attendance.objects.filter(employeeId=employee_id,date=start_date_before_2_hours, time__gte=start_time_before_2_hours).order_by('time').first()
                    nearest_login_time = login_obj if login_obj else None
                    nextday_date_before_2_hours = nextday_start_datetime_before_2_hours.strftime("%Y-%m-%d")
                    nextday_time_before_2_hours = nextday_start_datetime_before_2_hours.strftime("%H:%M:%S")
                    # Get the logout time entry nearest to the shift ending time
                    logout_obj = attendance.objects.filter(employeeId=employee_id,date=nextday_date_before_2_hours, time__lt=nextday_time_before_2_hours).exclude(id__in=exclude_obj_list).order_by('time').last()
                    extra_days=''
                    if logout_obj is None:
                        logout_obj = attendance.objects.filter(employeeId=employee_id,date=start_date, time__gt=start_time_before_2_hours).exclude(id__in=exclude_obj_list).order_by('time').last()

                    if logout_obj is not None:
                        if str(start_date_before_2_hours) != str(logout_obj.date):
                            extra_days='+1'

                        
                        
                    nearest_logout_time = logout_obj if logout_obj else None

                    return nearest_login_time, nearest_logout_time, extra_days



                    
                    
                nearest_login_time, nearest_logout_time ,extra_days= get_nearest_login_logout_times(employee_id, j,str(shift_starting_time).split(" ")[1],str(shift_ending_time).split(" ")[1],exclude_obj_list)
                


                
                if j in fulldayleave:
                    
                    logint=""
                    logoutt=""
                    total_time=""
                    status=j.split("-")[2]+" "+day_name+" "+"L"
                    hello=status.endswith("L") 
                    if hello is True:
                        A_count+=1
                    else:
                        A_count+=0
                    d1={
                        "employeeId":employee_id,
                        "status":status,
                        "att_status":"L",
                        "fulldate":j,
                        "inTime":logint,
                        "outTime":logoutt,
                        "Total":total_time,
                        "shift":shift,
                        "extra_days":'',
                        "inTime_date":'',
                        "outTime_date":'',
                    }

                    d2.append(d1)

                elif j in halfdayleave:
                    
                    if nearest_login_time is not None:

                        logint=nearest_login_time.time
                        login_date=nearest_login_time.date
                        exclude_obj_list.append(nearest_login_time.id)    

                        if nearest_logout_time is not None:
                            logoutt=nearest_logout_time.time
                            logout_date=nearest_logout_time.date
                            exclude_obj_list.append(nearest_logout_time.id)    

                        else:
                            logoutt=logint
                            logout_date=login_date

                        login_datetime = datetime.datetime.strptime(str(login_date) + " " + str(logint), "%Y-%m-%d %H:%M:%S")
                        logout_datetime = datetime.datetime.strptime(str(logout_date) + " " + str(logoutt), "%Y-%m-%d %H:%M:%S")
                        timeDiff = logout_datetime - login_datetime
                        total_time = str(timeDiff)

                        if logint==logoutt: 

                            total_time='00:00:00' 
                            logoutt='----'
                            logout_date=j
                            status=j.split("-")[2]+" "+day_name+" "+"HF"
                        else:
                            status=j.split("-")[2]+" "+day_name+" "+"HF"

                        
                        
                    else:
                        logint="--"
                        logoutt="--"
                        total_time="--"
                        login_date=j
                        status=j.split("-")[2]+" "+day_name+" "+"HF"





                    d1={
                        "employeeId":employee_id,
                        "status":status,
                        "att_status":"HF",
                        "fulldate":j,
                        "inTime":logint,
                        "outTime":logoutt,
                        "Total":total_time,
                        "shift":shift,
                        "extra_days":extra_days,
                        "inTime_date":convertdate2(str(login_date)),
                        "outTime_date":convertdate2(str(logout_date)),
                    }

                    d2.append(d1)

                elif j in workfromhomelist:
                    if nearest_login_time is not None:
                        logint=nearest_login_time.time
                        login_date=nearest_login_time.date
                        exclude_obj_list.append(nearest_login_time.id)    

                        if nearest_logout_time is not None:
                            logoutt=nearest_logout_time.time
                            logout_date=nearest_logout_time.date
                            exclude_obj_list.append(nearest_logout_time.id)    

                        else:
                            logoutt=logint
                            logout_date=login_date

                        login_datetime = datetime.datetime.strptime(str(login_date) + " " + str(logint), "%Y-%m-%d %H:%M:%S")
                        logout_datetime = datetime.datetime.strptime(str(logout_date) + " " + str(logoutt), "%Y-%m-%d %H:%M:%S")
                        timeDiff = logout_datetime - login_datetime
                        total_time = str(timeDiff)
                        
                    else:
                        logint="--"
                        logoutt="--"
                        total_time="--"
                        login_date=j
                        logout_date=j
                    d1={
                        "employeeId":employee_id,
                        "status":status,
                        "att_status":"WFH",
                        "fulldate":j,
                        "inTime":logint,
                        "outTime":logoutt,
                        "Total":total_time,
                        "shift":shift,
                        "extra_days":extra_days,
                        "inTime_date":convertdate2(str(login_date)),
                        "outTime_date":convertdate2(str(logout_date)),
                        
                        
                        

                    }

                    d2.append(d1)

                elif nearest_login_time is not None:
                    logint=nearest_login_time.time
                    login_date=nearest_login_time.date
                    exclude_obj_list.append(nearest_login_time.id)    

                    if nearest_logout_time is not None:
                        logoutt=nearest_logout_time.time
                        logout_date=nearest_logout_time.date
                        exclude_obj_list.append(nearest_logout_time.id)    

                    else:
                        logoutt=logint
                        logout_date=login_date

                    login_datetime = datetime.datetime.strptime(str(login_date) + " " + str(logint), "%Y-%m-%d %H:%M:%S")
                    logout_datetime = datetime.datetime.strptime(str(logout_date) + " " + str(logoutt), "%Y-%m-%d %H:%M:%S")
                    timeDiff = logout_datetime - login_datetime
                    total_time = str(timeDiff)

                    if logint==logoutt: 

                        total_time='00:00:00' 
                        logoutt='----'
                        logout_date=j
                        status = j.split("-")[2]+" "+day_name+" "+"P"
                        att_status = "P"
                    else:
                        status = j.split("-")[2]+" "+day_name+" "+"P"
                        att_status = "P"
                        login_date=j

                    d1={ 
                        "employeeId":employee_id,
                        "status":status,
                        "att_status":att_status,
                        "fulldate":j,
                        "inTime":logint,
                        "outTime":logoutt,
                        "extra_days":extra_days,
                        "inTime_date":convertdate2(str(login_date)),
                        "outTime_date":convertdate2(str(logout_date)),
                        "Total":total_time,
                        "shift":shift,

                    }

                    d2.append(d1)
                    
                elif j in weeklyofflist:
                    logint=""
                    logoutt=""
                    total_time=""
                    status=j.split("-")[2]+" "+day_name+" "+"WO"

                    d1={
                        "employeeId":employee_id,
                        "status":status,
                        "att_status":"WO",
                        "fulldate":j,
                        "inTime":logint,
                        "outTime":logoutt,
                        "Total":total_time,
                        "shift":shift,
                    }

                    d2.append(d1)

                elif j in holidaylist:
                    logint=""
                    logoutt=""
                    total_time=""
                    status=j.split("-")[2]+" "+day_name+" "+"H"
                    hello=status.endswith("H")
                    reason='Holiday'
                    Holidays_obj=Holidays.objects.filter(Date=j,Active=True).first()
                    if Holidays_obj:
                        reason=str(Holidays_obj.Festival)
                    if hello is True:
                        holiday_count+=1
                
                    d1={
                        "employeeId":employee_id,
                        "status":status,
                        "att_status":"H",
                        "fulldate":j,
                        "inTime":logint,
                        "outTime":logoutt,
                        "Total":total_time,
                        "reason":reason,
                        "shift":shift,
                        "extra_days":'',
                        "inTime_date":'',
                        "outTime_date":'',

                    }
                    d2.append(d1)

                else:
                    logint="-"
                    logoutt="-"
                    total_time="-"
                    status=j.split("-")[2]+" "+day_name+" "+"A"

                    d1={
                        "employeeId":employee_id,
                        "status":status,
                        "att_status":"A",
                        "fulldate":j,
                        "inTime":logint,
                        "outTime":logoutt,
                        "Total":total_time,
                        "shift":shift,
                        "extra_days":'',
                        "inTime_date":'',
                        "outTime_date":'',
                    }
                    d2.append(d1)


            d5={
                'employeeId':employee_id,
                'Absent':A_count,
                'Holiday':holiday_count,
                'records':dict((v['status'],v) for v in d2).values() 
            }

            d6.append(d5)
            d2.clear()


            
            for p in d6:
                for i in p['records']:
                    statuslist.append(i)



            return Response ({"data":statuslist,"response":{"n" : 1,"msg" : "SUCCESS","status" : "success"}})


        else:
            A_count=0
            holiday_count=0
            
            for j in dayslist:
            
                year, month, day = j.split('-') 
                wname = datetime.date(int(year), int(month), int(day))
                day_name=wname.strftime("%a")
                
                #first and last time records filter
                loginattObj=attendance.objects.filter(employeeId=employee_id,date=j).order_by('time').first()
            
                logoutattObj=attendance.objects.filter(employeeId=employee_id,date=j).order_by('time').last()
                
                






                shift={}
                shift['shiftname']='No Shift'
                shift['intime']='00'
                shift['outtime']='00'
                shift['swap']=''

                alloted_shift_obj=ShiftAllotment.objects.filter(employeeId=empuserID,attendanceId=employee_id,date=j,is_active=True).first()
                if alloted_shift_obj is not None:
                    ShiftAllotment_serializer=ShiftAllotmentSerializer(alloted_shift_obj)
                    if ShiftAllotment_serializer.data['shiftId'] is not None and ShiftAllotment_serializer.data['shiftId'] !='':
                        shift_obj=ShiftMaster.objects.filter(id=ShiftAllotment_serializer.data['shiftId'],is_active=True).first()
                        if shift_obj is not None:
                            shift_obj_serializer=ShiftMasterSerializer(shift_obj)
                            shift['shiftname']=shift_obj_serializer.data['shiftname']
                            shift['intime']=shift_obj_serializer.data['intime']
                            shift['outtime']=shift_obj_serializer.data['outtime']

                            if ShiftAllotment_serializer.data['is_swaped'] == True:
                                check_swap_details=Shiftswap.objects.filter(id=ShiftAllotment_serializer.data['swap_request_id']).first()
                                if check_swap_details is not None:
                                    Shiftswap_serializer=CustomShiftswapSerializer(check_swap_details)
                                    if ShiftAllotment_serializer.data['swapper'] == True:
                                        shift['swap']='Swaped with '+ str(Shiftswap_serializer.data['second_employee_name'])
                                    elif ShiftAllotment_serializer.data['swapper'] == False:
                                        shift['swap']='Swaped with '+ str(Shiftswap_serializer.data['first_employee_name'])
                                    else:
                                        shift['swap']=''
                                else:
                                    shift['swap']=''
                            else:
                                shift['swap']=''





                if j in fulldayleave:
                    
                    logint=""
                    logoutt=""
                    total_time=""
                    status=j.split("-")[2]+" "+day_name+" "+"L"
                    hello=status.endswith("L") 
                    if hello is True:
                        A_count+=1
                    else:
                        A_count+=0
                    d1={
                        "employeeId":employee_id,
                        "status":status,
                        "att_status":"L",
                        "fulldate":j,
                        "inTime":logint,
                        "outTime":logoutt,
                        "Total":total_time,
                        "shift":shift,
                        "extra_days":'',
                        "inTime_date":'',
                        "outTime_date":'',

                    }

                    d2.append(d1)

                elif j in halfdayleave:
                    
                    if loginattObj is not None:
                        logint=loginattObj.time
                        t1 = datetime.datetime.strptime(logint, "%H:%M:%S")
                        logoutt=logoutattObj.time
                        t2 = datetime.datetime.strptime(logoutt, "%H:%M:%S")
                        timeDiff = t2 - t1 
                        total_time=str(timeDiff)
                        status=j.split("-")[2]+" "+day_name+" "+"HF"
                    else:
                        logint="--"
                        logoutt="--"
                        total_time="--"
                        status=j.split("-")[2]+" "+day_name+" "+"HF"
                    hello=status.endswith("HF") 
                    if hello is True:
                        A_count+=0.5
                    else:
                        A_count+=0
                    d1={
                        "employeeId":employee_id,
                        "status":status,
                        "att_status":"HF",
                        "fulldate":j,
                        "inTime":logint,
                        "shift":shift,
                        "outTime":logoutt,
                        "Total":total_time,
                        "extra_days":'',
                        "inTime_date":'',
                        "outTime_date":'',
                    }

                    d2.append(d1)

                elif j in workfromhomelist:
                    if loginattObj is not None:
                        logint=loginattObj.time
                        t1 = datetime.datetime.strptime(logint, "%H:%M:%S")
                        logoutt=logoutattObj.time
                        t2 = datetime.datetime.strptime(logoutt, "%H:%M:%S")
                        timeDiff = t2 - t1 
                        total_time=str(timeDiff)
                        status=j.split("-")[2]+" "+day_name+" "+"WFH"
                    else:
                        logint="--"
                        logoutt="--"
                        total_time="--"
                        status=j.split("-")[2]+" "+day_name+" "+"WFH"

                    d1={
                        "employeeId":employee_id,
                        "status":status,
                        "att_status":"WFH",
                        "fulldate":j,
                        "shift":shift,
                        "inTime":logint,
                        "outTime":logoutt,
                        "Total":total_time,
                        "extra_days":'',
                        "inTime_date":'',
                        "outTime_date":'',
                    }

                    d2.append(d1)

                
                elif loginattObj is None:
                
                    
                    if j in weeklyofflist:
                    
                        logint=""
                        logoutt=""
                        total_time=""
                        status=j.split("-")[2]+" "+day_name+" "+"WO"
                        
                        d1={
                            "employeeId":employee_id,
                            "status":status,
                            "att_status":"WO",
                            "fulldate":j,
                            "inTime":logint,
                            "shift":shift,
                            "outTime":logoutt,
                            "Total":total_time,
                            "extra_days":'',
                            "inTime_date":'',
                            "outTime_date":'',

                        }
                        d2.append(d1)

                    elif j in holidaylist:
                    
                        logint=""
                        logoutt=""
                        total_time=""
                        status=j.split("-")[2]+" "+day_name+" "+"H"
                        hello=status.endswith("H")
                        if hello is True:
                            holiday_count+=1
                        reason='Holiday'
                        Holidays_obj=Holidays.objects.filter(Date=j,Active=True).first()
                        if Holidays_obj:
                            reason=str(Holidays_obj.Festival)
                        if hello is True:
                            holiday_count+=1
                            
                        d1={
                            "employeeId":employee_id,
                            "status":status,
                            "att_status":"H",
                            "fulldate":j,
                            "inTime":logint,
                            "outTime":logoutt,
                            "Total":total_time,
                            "shift":shift,
                            "reason":reason,
                            "extra_days":'',
                            "inTime_date":'',
                            "outTime_date":'',
                        }
                        d2.append(d1) 
                    else:
                    
                        logint=""
                        logoutt=""
                        total_time=""
                        status=j.split("-")[2]+" "+day_name+" "+"NF"
                        att_status = "NF"
                    
                        d1={
                            "employeeId":employee_id,
                            "status":status,
                            "att_status":att_status,
                            "fulldate":j,
                            "inTime":logint,
                            "outTime":logoutt,
                            "shift":shift,
                            "Total":total_time,
                            "extra_days":'',
                            "inTime_date":'',
                            "outTime_date":'',
                        }
                        d2.append(d1) 


            d5={
                'employeeId':employee_id,
                'Absent':A_count,
                'Holiday':holiday_count,
                'records':dict((v['status'],v) for v in d2).values() 
            }
            d6.append(d5)
            d2.clear()
        





            
            for p in d6:
                for i in p['records']:
                    statuslist.append(i)



                    
            return Response ({"data":statuslist,"response":{"n" : 1,"msg" : "SUCCESS","status" : "success"}})

    return Response ({"data":[],"response":{"n" : 0,"msg" : "user id not found ","status" : "failed"}})

            
def add_leading_zero(number):
    if 0 <= number <= 9:
        return f'0{number}'
    else:
        return str(number)
     
        
def gettodaysshift(shift_data,current_date):
    
    current_time_string = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    current_time = datetime.datetime.strptime(current_time_string, '%Y-%m-%d %H:%M:%S')

    # current_time = datetime.datetime.strptime('2024-03-06 02:36', '%Y-%m-%d %H:%M')
    
    current_date = datetime.datetime.strptime(current_date, '%Y-%m-%d').date()
    # print("\ncurrent_time",current_time,current_date) 
    
    last_runingshift={"shiftname":'','shiftstarttime':'','shiftendtime':''}
    runingshift={'n':0,'data':{"shiftname":'','shiftstarttime':'','shiftendtime':''},'last_runingshift':last_runingshift}
    
    for shift in shift_data:
        start_time = datetime.datetime.strptime(str(current_date) +' '+ shift['intime'], '%Y-%m-%d %H:%M')
        start_time_before_2hrs = start_time - timedelta(hours=2)
        if shift['intime'] > shift['outtime']:
            shift_end_date = current_date + timedelta(days=1)
        else:
            shift_end_date = current_date
        end_time = datetime.datetime.strptime(str(shift_end_date) +' '+ shift['outtime'], '%Y-%m-%d %H:%M')
        
        
        if start_time_before_2hrs < current_time:
            
            last_runingshift={"shiftname":shift['shiftname'],'shiftstarttime':start_time_before_2hrs,'shiftendtime':end_time}
            # print("forloop ---",last_runingshift)
        
        if start_time_before_2hrs <= current_time <= end_time:
            runingshift={'n':1,'data':{"shiftname":shift['shiftname'],'shiftstarttime':start_time_before_2hrs,'shiftendtime':end_time},'last_runingshift':last_runingshift}
            
            

    runingshift['last_runingshift']=last_runingshift   
    

    return runingshift



@api_view(['POST'])    
@permission_classes((AllowAny,))
def employee_monthly_shift_details(request):
    
    
    
    month=int(request.data.get("month"))
    year=int(request.data.get("year"))
    
    UserId=request.data.get("UserId")
    user_obj=Users.objects.filter(id=UserId,is_active=True).exclude(Q(employeeId__isnull=True)|Q(employeeId='None')|Q(employeeId='')).first()
    
    
    office_location=Location.objects.filter(Active=True).exclude(Q(lattitude='',meter='',longitude='') |Q(lattitude=None,meter=None,longitude=None)|Q(lattitude__isnull=True,meter__isnull=True,longitude__isnull=True))
    location_serializer=LocationSerializer(office_location,many=True)
    
    if user_obj is not None:
        employeeId=user_obj.employeeId
        if month !='' and year !=''  and  month is not None and year is not None:
            leave_objs = Leave.objects.filter(employeeId=UserId,WorkFromHome=False,leave_status="Approved",start_date__month__gte=month,start_date__month__lte=month,start_date__year= year,Active=True).exclude(leave_status='Draft')
            leavesserializer=leaveserializer(leave_objs,many=True)
            wfh_objs = Leave.objects.filter(employeeId=UserId,WorkFromHome=True,leave_status="Approved",start_date__month__gte=month,start_date__month__lte=month,start_date__year= year,Active=True).exclude(leave_status='Draft')
            wfhsserializer=leaveserializer(wfh_objs,many=True)
            cal = calendar.monthcalendar(year, month)
            dates_list = [
                    f"{year}-{month:02d}-{day:02d}"
                    for week in cal
                    for day in week
                    if day != 0
                ]    
            attendance_report=[]    
            for date in dates_list:
                current_date_shift_list=[]
                new_current_date_shift_list=[]
                return_dict={}
                
                shiftdate = datetime.datetime.strptime(date, '%Y-%m-%d').date()
                tomarrow_date = shiftdate + datetime.timedelta(days=1)
                yesterday_date = shiftdate - datetime.timedelta(days=1)
                current_shiftallotment_objs=ShiftAllotment.objects.filter(attendanceId=employeeId,date=str(date),is_active=True)
                current_shiftallotment_serializers=ShiftAllotmentshiftIdSerializer(current_shiftallotment_objs,many=True)
                current_shiftId_list=list(current_shiftallotment_serializers.data)
                current_shift_obj=ShiftMaster.objects.filter(id__in=current_shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime')
                current_shift_serializer=ShiftMasterSerializer(current_shift_obj,many=True)
                current_shiftlist=current_shift_serializer.data
                
                tomarrow_shiftallotment_objs=ShiftAllotment.objects.filter(attendanceId=employeeId,date=str(tomarrow_date),is_active=True)
                tomarrow_shiftallotment_serializers=ShiftAllotmentshiftIdSerializer(tomarrow_shiftallotment_objs,many=True)
                tomarrow_shiftId_list=list(tomarrow_shiftallotment_serializers.data)
                tomarrow_shift_obj=ShiftMaster.objects.filter(id__in=tomarrow_shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime')
                tomarrow_shift_serializer=ShiftMasterSerializer(tomarrow_shift_obj,many=True)
                tomarrow_shiftlist=tomarrow_shift_serializer.data

                count=1
                
                for shift in current_shiftlist:
                    start_time = datetime.datetime.strptime(str(shiftdate) +' '+ shift['intime'], '%Y-%m-%d %H:%M')
                    start_time_before_2hrs = start_time - timedelta(hours=2)
                    
                    if shift['intime'] > shift['outtime']:
                        shift_end_date = shiftdate + timedelta(days=1)
                    else:
                        shift_end_date = shiftdate
                        
                    end_time = datetime.datetime.strptime(str(shift_end_date) +' '+ shift['outtime'], '%Y-%m-%d %H:%M')
                    check_login_till=''
                    if len(current_shiftlist) >= count+1:
                   
                        check_next_shift_in = datetime.datetime.strptime(str(shiftdate) +' '+ current_shiftlist[count]['intime'], '%Y-%m-%d %H:%M')
                        check_login_till = check_next_shift_in - timedelta(hours=2)

                        
                    elif check_login_till=='':
                          
                        if len(tomarrow_shiftlist) >= 1:
                         
                            check_next_shift_in = datetime.datetime.strptime(str(tomarrow_date) +' '+ tomarrow_shiftlist[0]['intime'], '%Y-%m-%d %H:%M')
                            check_login_till = check_next_shift_in - timedelta(hours=2)
                        else:
                      
                            if_no_shift_check_shift=ShiftAllotment.objects.filter(attendanceId=employeeId,date=str(tomarrow_date),is_active=True)
                            tomerrow_shiftlist=ShiftAllotmentshiftIdSerializer(if_no_shift_check_shift,many=True)
                            next_date_shiftId_list=list(tomerrow_shiftlist.data)
                            check_weekly_off=ShiftMaster.objects.filter(id__in=next_date_shiftId_list,intime='00:00',outtime='00:00',is_active=True).order_by('intime').first()
                            if check_weekly_off is not None:
                              
                                if shift['outtime'] < shift['intime']:
                         
                                    check_current_shift_out = datetime.datetime.strptime(str(tomarrow_date) +' '+ shift['outtime'], '%Y-%m-%d %H:%M')
                                    check_current_shift_out_time = check_current_shift_out + timedelta(hours=2)
                                    given_datetime = datetime.datetime.strptime(str(check_current_shift_out_time), '%Y-%m-%d %H:%M:%S')

                                    truncated_datetime_str = given_datetime.strftime('%Y-%m-%d %H:%M')

                                    
                                    shift_end_date_time=str(truncated_datetime_str)
                                else:
                                    shift_end_date_time=str(date) +' '+ '23:59'
                                    
                                check_login_till = datetime.datetime.strptime(shift_end_date_time, '%Y-%m-%d %H:%M')
                            else:
                            
                                check_login_till = datetime.datetime.strptime(str(tomarrow_date) +' '+ '07:30', '%Y-%m-%d %H:%M')
                                
                    # print("shift "+str(count),shift['shiftname'],start_time_before_2hrs,check_login_till)
                    shift_data={
                        'shiftname':shift['shiftname'],
                        'indatetime':str(start_time_before_2hrs),
                        'outdatetime':str(check_login_till),
                        'intime':shift['intime'],
                        'outtime':shift['outtime'],
                    }
                    current_date_shift_list.append(shift_data)
                    new_current_date_shift_list.append(shift_data)
                    count+=1

                if len(current_shiftlist) == 0:       
               
                    if_no_shift_check_shift=ShiftAllotment.objects.filter(attendanceId=employeeId,date=str(date),is_active=True)
                    todays_shiftlist=ShiftAllotmentshiftIdSerializer(if_no_shift_check_shift,many=True)
                    toadys_shiftId_list=list(todays_shiftlist.data)
                    check_weekly_off=ShiftMaster.objects.filter(id__in=toadys_shiftId_list,intime='00:00',outtime='00:00',is_active=True).order_by('intime').first()
                    if check_weekly_off is not None:
   
                        get_previous_day_shift=ShiftAllotment.objects.filter(attendanceId=employeeId,date=str(yesterday_date),is_active=True)
                        previous_day_shiftlist=ShiftAllotmentshiftIdSerializer(get_previous_day_shift,many=True)
                        previous_shiftId_list=list(previous_day_shiftlist.data)
                        get_previous_last_shift=ShiftMaster.objects.filter(id__in=previous_shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime').last()
                        
                        if get_previous_last_shift is not None:
                            if get_previous_last_shift.outtime < get_previous_last_shift.intime:
                                shift_end_date_time=str(date) +' '+ str(get_previous_last_shift.outtime)
                            else:
                                shift_end_date_time=str(yesterday_date) +' '+ '23:59'
                                    
                            check_previous_shift_out = datetime.datetime.strptime(str(shift_end_date_time), '%Y-%m-%d %H:%M')
                            previous_shift_in_time = check_previous_shift_out + timedelta(hours=2)  
                        else:
            
                            
                            previous_shift_in_time= str(date)+' 00:00:00'
                        # checking previous  shift out time must be currrent shift in time in weekly off    # 
                        get_net_day_shift=ShiftAllotment.objects.filter(attendanceId=employeeId,date=str(tomarrow_date),is_active=True)
                        nest_day_shiftlist=ShiftAllotmentshiftIdSerializer(get_net_day_shift,many=True)
                        next_date_shiftId_list=list(nest_day_shiftlist.data)
                        get_next_day_first_shift=ShiftMaster.objects.filter(id__in=next_date_shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime').first()
                        if get_next_day_first_shift is not None:
                            check_next_shift_in = datetime.datetime.strptime(str(tomarrow_date) +' '+ str(get_next_day_first_shift.intime), '%Y-%m-%d %H:%M')
                            next_shift_in_time = check_next_shift_in - timedelta(hours=2)
                            shift_data={
                                'shiftname':str(check_weekly_off.shiftname),
                                'indatetime':str(previous_shift_in_time),
                                'outdatetime':str(next_shift_in_time),
                                'intime':check_weekly_off.intime,
                                'outtime':check_weekly_off.outtime,
                            }
                            current_date_shift_list.append(shift_data)
                            new_current_date_shift_list.append(shift_data)
                            
                        else:
                            shift_data={
                                'shiftname':str(check_weekly_off.shiftname),
                                'indatetime':str(previous_shift_in_time),
                                'outdatetime':str(date)+' 23:59:59',
                                'intime':check_weekly_off.intime,
                                'outtime':check_weekly_off.outtime,
                            }
                            current_date_shift_list.append(shift_data)
                            new_current_date_shift_list.append(shift_data)

                    else:
                        # print("No shift",'General',str(date)+' 07:30:00',str(tomarrow_date)+' 07:30:00')
                        shift_data={
                            'shiftname':'General',
                            'indatetime':str(date)+' 07:30:00',
                            'outdatetime':str(tomarrow_date)+' 07:30:00',
                            'intime':'09:30',
                            'outtime':'18:30',
                        }
                        current_date_shift_list.append(shift_data)
                        new_current_date_shift_list.append(shift_data)

                        
                current_date_first_in_datetime=''
                current_date_last_out_datetime=''
                extra_days=''
                total_working_hrs=''
                for s in new_current_date_shift_list:
                    intime=''
                    intimedate=''
                    
                    
                    getallattendance = attendance.objects.filter(Q(employeeId=str(employeeId),time__gte=s['indatetime'].split(' ')[1],date=s['indatetime'].split(' ')[0])|Q(employeeId=str(employeeId),time__lte=str(s['outdatetime']).split(' ')[1],date=str(s['outdatetime']).split(' ')[0])).order_by('date','time')
                                        
                    attendance_serializer=attendanceserializer(getallattendance,many=True)
                
                    
                    sorted_data = sorted(attendance_serializer.data, key=lambda x: (x['date'], x['time']))
                    
                    mindatetime = datetime.datetime.strptime(s['indatetime'], '%Y-%m-%d %H:%M:%S')
                    maxdatetime = datetime.datetime.strptime(s['outdatetime'], '%Y-%m-%d %H:%M:%S')
                
                    sorted_data = [entry for entry in sorted_data if (mindatetime <= datetime.datetime.strptime(entry['date'] +' '+entry['time'],'%Y-%m-%d %H:%M:%S') <= maxdatetime)]
                    if len(sorted_data) > 0:
                        intimedate=sorted_data[0]['date']
                        intime=str(sorted_data[0]['time'])
                        
                    if intimedate !='' and intimedate is not None:
                        user_sdt = datetime.datetime.strptime(str(intimedate) + ' ' + str(intime), '%Y-%m-%d %H:%M:%S')
                        shif_sdt = datetime.datetime.strptime(s['indatetime'].split(' ')[0] + ' ' + s['indatetime'].split(' ')[1], '%Y-%m-%d %H:%M:%S')
                        if user_sdt < shif_sdt :
                            intimedate=''
                            intime=''

                            
                    checkin_time = None
                    total_working_time = 0
                    attendance_log=[]
                    attendance_history=[]

                    for index, entry in enumerate(sorted_data):

                        if entry['checkout']:
                            if entry['deviceId'] is not None and entry['deviceId'] !='':
                                attendance_type="Machine Checkout"
                                attendance_type_resaon=""
                                if entry['deviceId'] == 20:
                                    remote_map_location=create_google_maps_url(18.566064883698555, 73.77574703366226)
                                    remote_map_name='Zentro Pune Office'
                                elif entry['deviceId'] == 19:
                                    remote_map_location=create_google_maps_url(19.1764898841813, 72.95988765068624)
                                    remote_map_name='Zentro Mumbai Office'
                                else:
                                    remote_map_location=''
                                    remote_map_name=''
                            elif entry['Remote_Reason'] is not None and entry['Remote_Reason'] !='':
                                attendance_type="Remote Checkout"
                                attendance_type_resaon=entry['Remote_Reason']
                                remote_map_location=create_google_maps_url(entry['remote_latitude'],entry['remote_longitude'])
                                remote_map_name=get_location_name(entry['remote_latitude'],entry['remote_longitude'])

                            else:
                                attendance_type="Web Checkout"
                                attendance_type_resaon=''
                                
                                if entry['attendance_type'] !='' and  entry['attendance_type'] is not None:
                                    attendance_type = entry['attendance_type'] +' '+ 'Checkout'
                                    
                                remote_map_location=''            
                                if entry['remote_latitude'] is not None and  entry['remote_latitude'] !='' and entry['remote_longitude'] is not None and entry['remote_longitude'] !='':

                                    for location in location_serializer.data:
                                        within_radius = is_within_radius(float(entry['remote_latitude']), float(entry['remote_longitude']), float(location['lattitude']), float(location['longitude']), float(location['meter']))
                                        if within_radius:
                                            remote_map_location=create_google_maps_url(entry['remote_latitude'],entry['remote_longitude'])
                                            remote_map_name='ORT '+str(location['LocationName']) +' Office'
                                    if remote_map_location == '' or  remote_map_name =='' or remote_map_location is None or  remote_map_name  is  None:
                                        remote_map_location=create_google_maps_url(entry['remote_latitude'],entry['remote_longitude'])
                                        remote_map_name=get_location_name(entry['remote_latitude'],entry['remote_longitude'])
                                else:
                                    remote_map_location=''
                                    remote_map_name=''                        

                            attendance_history.append({'Status':'Check-Out','datetime':entry['date']+' '+entry['time'],'attendance_type':attendance_type,'attendance_type_resaon':attendance_type_resaon,'remote_map_location':remote_map_location,'remote_map_name':remote_map_name})
                                
                            attendance_log.append({'checkout':entry['date']+' '+entry['time']})
                            
                            
                            if checkin_time:
                                checkout_datetime = datetime.datetime.strptime(entry['date'] + ' ' + entry['time'], '%Y-%m-%d %H:%M:%S')
                                checkin_datetime = datetime.datetime.strptime(checkin_time, '%Y-%m-%d %H:%M:%S')
                                working_time = checkout_datetime - checkin_datetime
                                total_working_time += working_time.total_seconds()
                                checkin_time = None
                                
                        elif not entry['checkout']:
                            remote_map_location=''
                            remote_map_name='' 
                            if entry['deviceId'] is not None and entry['deviceId'] !='':
                                attendance_type="Machine Checkin"
                                attendance_type_resaon=''
                                if entry['deviceId'] == 20:
                                    remote_map_location=create_google_maps_url(18.566064883698555, 73.77574703366226)
                                    remote_map_name='Zentro Pune Office'
                                elif entry['deviceId'] == 19:
                                    remote_map_location=create_google_maps_url(19.1764898841813, 72.95988765068624)
                                    remote_map_name='Zentro Mumbai Office'
                                else:
                                    remote_map_location=''
                                    remote_map_name=''
                                    
                            elif entry['Remote_Reason'] is not None and entry['Remote_Reason'] !='':
                                attendance_type="Remote Checkin"
                                attendance_type_resaon=entry['Remote_Reason']
                                remote_map_location=create_google_maps_url(entry['remote_latitude'],entry['remote_longitude'])
                                remote_map_name=get_location_name(entry['remote_latitude'],entry['remote_longitude'])
                            else:
                                attendance_type="Web Checkin"
                                attendance_type_resaon='' 
                                if entry['attendance_type'] !='' and  entry['attendance_type'] is not None:
                                    attendance_type = entry['attendance_type'] +' '+ 'Checkin'
                                remote_map_location=''
                                
                                if entry['remote_latitude'] is not None and  entry['remote_latitude'] !='' and entry['remote_longitude'] is not None and entry['remote_longitude'] !='':
                
                                    for location in location_serializer.data:
                                        within_radius = is_within_radius(float(entry['remote_latitude']), float(entry['remote_longitude']), float(location['lattitude']), float(location['longitude']), float(location['meter']))
                                        if within_radius:
                                            remote_map_location=create_google_maps_url(entry['remote_latitude'],entry['remote_longitude'])
                                            remote_map_name='Zentro '+str(location['LocationName']) +' Office'
                                    if remote_map_location == '' or  remote_map_name =='' or remote_map_location is None or  remote_map_name  is  None:
                                        remote_map_location=create_google_maps_url(entry['remote_latitude'],entry['remote_longitude'])
                                        remote_map_name=get_location_name(entry['remote_latitude'],entry['remote_longitude'])
                                else:
                                    remote_map_location=''
                                    remote_map_name='' 
                                 
                            checkin_time = entry['date'] + ' ' + entry['time']
                            attendance_log.append({'checkin':entry['date']+' '+entry['time']})
                            attendance_history.append({'Status':'Check-In','datetime':entry['date']+' '+entry['time'],'attendance_type':attendance_type,'attendance_type_resaon':attendance_type_resaon,'remote_map_location':remote_map_location,'remote_map_name':remote_map_name})
                    # If the last entry is check-in, consider checkout time as current shift end  time
                    if checkin_time and index == len(sorted_data) - 1:
                        
                        # check if the previous entry is also checkin or not if exist get the  difference betwnn  the current checkin  and  previous  checkin

                        if int(int(index)-1) >=0:
                            if sorted_data[index-1]['checkout']==False:
                                previous_checkin_date_time=sorted_data[index-1]['date']+ ' '+sorted_data[index-1]['time']
                                checkout_datetime = datetime.datetime.strptime(previous_checkin_date_time, '%Y-%m-%d %H:%M:%S')
                                checkin_datetime = datetime.datetime.strptime(checkin_time, '%Y-%m-%d %H:%M:%S')
                                working_time = checkin_datetime-checkout_datetime 
                                total_working_time += working_time.total_seconds()


                    # Convert total_working_time to hours, minutes, and seconds
                    hours, remainder = divmod(total_working_time, 3600)
                    minutes, seconds = divmod(remainder, 60)
                    # print("attendance_log",attendance_log)
                    s['total_hrs']=str(add_leading_zero(int(hours))) +':' +str(add_leading_zero(int(minutes))) +':'+str(add_leading_zero(int(seconds)))

                    if total_working_hrs == '':
       
                        total_working_hrs=s['total_hrs']
                    else:
                        time_str_1 = total_working_hrs
                        time_str_2 = s['total_hrs']



                        # Convert time strings to timedelta objects and add them
                        total_time_delta = timedelta(hours=int(time_str_1[:2]), minutes=int(time_str_1[3:5]), seconds=int(time_str_1[6:])) + \
                                        timedelta(hours=int(time_str_2[:2]), minutes=int(time_str_2[3:5]), seconds=int(time_str_2[6:]))
                        total_working_hrs=str(total_time_delta)
                        
                    s['attendance_log']=attendance_log
                    s['attendance_history']=attendance_history
                    if len(attendance_log) !=0:
                        infirst_key = next(iter(attendance_log[0]))
                        outfirst_key = next(iter(attendance_log[-1]))

                        s['usersintime']=attendance_log[0][infirst_key]
                        s['usersouttime']=attendance_log[-1][outfirst_key]
                        if s['usersouttime'] == s['usersintime']:
                            s['usersouttime']= ''
                            
                    else:
                        s['usersintime']=''
                        s['usersouttime']=''

    
                    
                    if current_date_first_in_datetime =='':
                        if s['usersintime'] !='':
                            current_date_first_in_datetime=s['usersintime']
                    elif s['usersintime'] !='':
                        if datetime.datetime.strptime(current_date_first_in_datetime, '%Y-%m-%d %H:%M:%S') > datetime.datetime.strptime(s['usersintime'], '%Y-%m-%d %H:%M:%S'):
                            current_date_first_in_datetime=s['usersintime']

                    if current_date_last_out_datetime =='':
                        if s['usersouttime'] !='':
                            current_date_last_out_datetime=s['usersouttime']
                    elif s['usersouttime'] !='':
                        if datetime.datetime.strptime(current_date_last_out_datetime, '%Y-%m-%d %H:%M:%S') < datetime.datetime.strptime(s['usersouttime'], '%Y-%m-%d %H:%M:%S'):
                            current_date_last_out_datetime=s['usersouttime']
                    
                    
                    
                    
                    # if current_date_last_out_datetime != '' and current_date_first_in_datetime !='':
                    #     if current_date_last_out_datetime.split(' ')[0] != current_date_first_in_datetime.split(' ')[0]:
                    #         current_date_last_out_datetime = datetime.datetime.strptime(current_date_last_out_datetime, '%Y-%m-%d %H:%M:%S')
                    #         current_date_first_in_datetime = datetime.datetime.strptime(current_date_first_in_datetime, '%Y-%m-%d %H:%M:%S')
                    #         extradaysdiff = current_date_last_out_datetime - current_date_first_in_datetime
                    #         if extradaysdiff.days >0:
                    #             extra_days='+'+str(extradaysdiff.days)
                        
                    if current_date_last_out_datetime != '' and current_date_first_in_datetime !='':
                        if current_date_last_out_datetime.split(' ')[0] != current_date_first_in_datetime.split(' ')[0]:
                            current_date_last_out_datetime1 = datetime.datetime.strptime(str(current_date_last_out_datetime).split(' ')[0], '%Y-%m-%d')
                            current_date_first_in_datetime1 = datetime.datetime.strptime(str(current_date_first_in_datetime).split(' ')[0], '%Y-%m-%d')
                            extradaysdiff = current_date_last_out_datetime1 - current_date_first_in_datetime1
                            if extradaysdiff.days > 0:
                                extra_days='+'+str(extradaysdiff.days)
                        
                        
                return_dict['employeeId']=employeeId
                return_dict['fulldate']=date
                return_dict['shifts_list']=current_date_shift_list

                return_dict['shift']=new_current_date_shift_list[0]
                
                if len(new_current_date_shift_list) > 1:
                    shift_name=new_current_date_shift_list[0]['shiftname']
                    return_dict['shift']['shiftname'] = str(shift_name) +  ' <span class="text-danger"> +1 </span>'
                    new_current_date_shift_list[0]['swap']=''

                else:

                    new_current_date_shift_list[0]['swap']=''
                    
                
                if current_date_first_in_datetime !='':
                    return_dict['inTime']=str(current_date_first_in_datetime).split(' ')[1]
                    return_dict['inTime_date']=convertdate2(str(str(current_date_first_in_datetime).split(' ')[0]))
                else:
                    return_dict['inTime']=''
                    return_dict['inTime_date']=''
                    
                if current_date_last_out_datetime !='':
                    return_dict['outTime']=str(current_date_last_out_datetime).split(' ')[1]
                    return_dict['outTime_date']=convertdate2(str(str(current_date_last_out_datetime).split(' ')[0]))
                else:
                    return_dict['outTime']=''
                    return_dict['outTime_date']=''
                    
                return_dict['extra_days']=extra_days
                return_dict['Total']=total_working_hrs
                if return_dict['inTime'] !='':
                    return_dict['att_status']="P"
                else:
                    return_dict['att_status']="A"
            
                foundleaves = [
                    leave for leave in leavesserializer.data if datetime.datetime.strptime(leave['start_date'], '%Y-%m-%d').date() <= datetime.datetime.strptime(date, '%Y-%m-%d').date() <= datetime.datetime.strptime(leave['end_date'], '%Y-%m-%d').date()
                ]
                
                foundwfhs = [
                    wfh for wfh in wfhsserializer.data if datetime.datetime.strptime(wfh['start_date'], '%Y-%m-%d').date() <= datetime.datetime.strptime(date, '%Y-%m-%d').date() <= datetime.datetime.strptime(wfh['end_date'], '%Y-%m-%d').date()
                ]
                
                if foundwfhs:
                    return_dict['att_status']="WFH"
                if foundleaves:
                    return_dict['att_status']="L"
   
                    
                attendance_report.append(return_dict)

            return Response ({"data":attendance_report,"response":{"n" : 1,"msg" : "success ","status" : "success"}})
        return Response ({"data":[],"response":{"n" : 0,"msg" : "month and year required ","status" : "failed"}})
    return Response ({"data":[],"response":{"n" : 0,"msg" : "user not found ","status" : "failed"}})




@api_view(['POST'])    
@permission_classes((AllowAny,))
def getmailsubjectdata(request):
    data={}
    empid = request.POST.get('empid')
    mailtype = request.POST.get('mailtype')
    mailtypee = mailtype.trim()
    empdataobj = warninglog.objects.filter(mailTo=empid,mailType=mailtypee)
    if empdataobj is None:
        typekey = mailtypee + "-1"
    else:
        mailtypecount = warninglog.objects.filter(mailTo=empid,mailType=mailtypee).count()
        nexttype = mailtypecount + 1
        typekey = mailtypee + "-" + nexttype
    return Response ({"data":typekey,"response":{"n" : 1,"msg" : "success ","status" : "success"}})


@api_view(['POST'])    
@permission_classes((AllowAny,))
def getmailhistorydata(request):
    data={}
    empid = request.POST.get('empid')
    empdataobj = warninglog.objects.filter(mailTo=empid).order_by('id')
    empser = warninglogserializer(empdataobj,many=True)
    return Response ({"data":empser.data,"response":{"n" : 1,"msg" : "success ","status" : "success"}})



@api_view(['POST'])    
@permission_classes((AllowAny,))
def sendwarningmail(request):
    data={}
    currentzone = pytz.timezone("Asia/Kolkata") 
    currenttime = datetime.datetime.now(currentzone)
    newcurrenttime =currenttime.strftime("%Y-%m-%dT%H:%M:%S.%f%z")

    newtime = str(newcurrenttime)
    ntime1= newtime.split('T')[1]
    ntimhrs = ntime1.split('.')[0]
    data['mail_time'] = ntimhrs
    data['mail_date'] =  date.today()
    data['mailType'] = request.POST.get('mailType')
    data['mailsubject'] = request.POST.get('mailsubject')
    data['mailcontent'] = request.POST.get('mailcontent')
    data['mailFrom'] = request.user.id
    data['mailTo'] = request.POST.get('empid')
    data['CreatedBy'] = request.user.id

    title = data['mailsubject']

    userobj = Users.objects.filter(id=int(data['mailTo'])).first()
    if userobj is not None:
        useremail = userobj.email
        empname =  userobj.Firstname + " "+ userobj.Lastname
    else:
        return Response ({"data":[],"response":{"n" : 0,"msg" : "user not found","status" : "failed"}})

    managerobj  = Users.objects.filter(id=int(data['mailFrom'])).first()
    if managerobj is not None:
        manageremail = managerobj.email
        managername =  managerobj.Firstname + " "+ managerobj.Lastname
    else:
        return Response ({"data":[],"response":{"n" : 0,"msg" : "manager not found","status" : "failed"}})

    if data['mailType'] == "Concern":
        ccmails = []
    elif data['mailType'] == "Escalation":
        ccmails = []
    else:
        ccmails = [adminemail,hremail]
 
    try:             
        dicti = {'empname': empname, 'managerName':managername,'content':data['mailcontent'],'title':title}

        message = get_template(
            'warningmail.html').render(dicti)
        msg = EmailMessage(
            data['mailsubject'],
            message, 
            EMAIL_HOST_USER,
            [useremail],
            ccmails,
        )
        msg.content_subtype = "html"  # Main content is now text/html
        msg.send()

        serializer = warninglogserializer(data=data)
        if serializer.is_valid():
            serializer.save()

    except Exception as e:
        print('exception while sending mail', e)

    return Response ({"data":'',"response":{"n" : 1,"msg" : "success ","status" : "success"}})


@api_view(['POST'])    
@permission_classes((AllowAny,))
def chatgptconversion(request):
    openai.api_key = "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"

    message = request.POST.get('message')
    prompt = message
    model = "gpt-3.5-turbo-instruct"
    response = openai.Completion.create(engine=model, prompt=prompt, max_tokens=500)

    generated_text = response.choices[0].text
    print(generated_text)
    return Response ({"data":generated_text,"response":{"n" : 1,"msg" : "success ","status" : "success"}})


@api_view(['POST'])
@permission_classes((AllowAny,))
def shiftexcelreport(request):

    d6=[]
    
    year_input = int(request.data.get('year'))
    month_input = int(request.data.get('month'))
    search_employees=json.loads(request.data.get('search_employees'))
    if year_input not in [None, ""] and month_input not in [None, ""]:
        #data of employee id
        shift_emp_obj=EmployeeShiftDetails.objects.filter(employeeId__in=search_employees,is_active=True)
        l1l2emplist=L1L2Serializer(shift_emp_obj,many=True)
        empdata = Users.objects.filter(id__in=list(l1l2emplist.data)).exclude(Q(employeeId__isnull=True)|Q(employeeId='None')|Q(employeeId='')).order_by('DepartmentID')
        mon_serializer = UsersSerializer(empdata,many=True)

       
        
        for p in mon_serializer.data:
            leave_count=0
            holiday_count=0
            totalworkedhrs=timedelta()
            presentcount=0
            absentcount=0  
            user_obj=Users.objects.filter(id=p['id'],is_active=True).exclude(Q(employeeId__isnull=True)|Q(employeeId='None')|Q(employeeId='')).first()
            if user_obj is not None:
                employeeId=user_obj.employeeId
                if month_input !='' and year_input !=''  and  month_input is not None and year_input is not None:
                    leave_objs = Leave.objects.filter(employeeId=p['id'],WorkFromHome=False,leave_status="Approved",start_date__month__gte=month_input,start_date__month__lte=month_input,start_date__year= year_input,Active=True).exclude(leave_status='Draft')
                    leavesserializer=leaveserializer(leave_objs,many=True)
                    wfh_objs = Leave.objects.filter(employeeId=p['id'],WorkFromHome=True,leave_status="Approved",start_date__month__gte=month_input,start_date__month__lte=month_input,start_date__year= year_input,Active=True).exclude(leave_status='Draft')
                    wfhsserializer=leaveserializer(wfh_objs,many=True)
                    cal = calendar.monthcalendar(year_input, month_input)
                    dates_list = [
                            f"{year_input}-{month_input:02d}-{day:02d}"
                            for week in cal
                            for day in week
                            if day != 0
                        ]    
                    attendance_report=[]    
                    for date in dates_list:
                        current_date_shift_list=[]
                        return_dict={}
                        shiftdate = datetime.datetime.strptime(date, '%Y-%m-%d').date()
                        tomarrow_date = shiftdate + datetime.timedelta(days=1)
                        yesterday_date = shiftdate - datetime.timedelta(days=1)
                        current_shiftallotment_objs=ShiftAllotment.objects.filter(attendanceId=employeeId,date=str(date),is_active=True)
                        current_shiftallotment_serializers=ShiftAllotmentshiftIdSerializer(current_shiftallotment_objs,many=True)
                        current_shiftId_list=list(current_shiftallotment_serializers.data)
                        current_shift_obj=ShiftMaster.objects.filter(id__in=current_shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime')
                        current_shift_serializer=ShiftMasterSerializer(current_shift_obj,many=True)
                        current_shiftlist=current_shift_serializer.data
                        
                        tomarrow_shiftallotment_objs=ShiftAllotment.objects.filter(attendanceId=employeeId,date=str(tomarrow_date),is_active=True)
                        tomarrow_shiftallotment_serializers=ShiftAllotmentshiftIdSerializer(tomarrow_shiftallotment_objs,many=True)
                        tomarrow_shiftId_list=list(tomarrow_shiftallotment_serializers.data)
                        tomarrow_shift_obj=ShiftMaster.objects.filter(id__in=tomarrow_shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime')
                        tomarrow_shift_serializer=ShiftMasterSerializer(tomarrow_shift_obj,many=True)
                        tomarrow_shiftlist=tomarrow_shift_serializer.data

                        count=1
                        
                        for shift in current_shiftlist:
                            start_time = datetime.datetime.strptime(str(shiftdate) +' '+ shift['intime'], '%Y-%m-%d %H:%M')
                            start_time_before_2hrs = start_time - timedelta(hours=2)
                            
                            if shift['intime'] > shift['outtime']:
                                shift_end_date = shiftdate + timedelta(days=1)
                            else:
                                shift_end_date = shiftdate
                                
                            end_time = datetime.datetime.strptime(str(shift_end_date) +' '+ shift['outtime'], '%Y-%m-%d %H:%M')
                            check_login_till=''
                            if len(current_shiftlist) >= count+1:
                        
                                check_next_shift_in = datetime.datetime.strptime(str(shiftdate) +' '+ current_shiftlist[count]['intime'], '%Y-%m-%d %H:%M')
                                check_login_till = check_next_shift_in - timedelta(hours=2)

                                
                            elif check_login_till=='':
                                
                                if len(tomarrow_shiftlist) >= 1:
                                
                                    check_next_shift_in = datetime.datetime.strptime(str(tomarrow_date) +' '+ tomarrow_shiftlist[0]['intime'], '%Y-%m-%d %H:%M')
                                    check_login_till = check_next_shift_in - timedelta(hours=2)
                                else:
                                    if_no_shift_check_shift=ShiftAllotment.objects.filter(attendanceId=employeeId,date=str(tomarrow_date),is_active=True)
                                    tomerrow_shiftlist=ShiftAllotmentshiftIdSerializer(if_no_shift_check_shift,many=True)
                                    next_date_shiftId_list=list(tomerrow_shiftlist.data)
                                    check_weekly_off=ShiftMaster.objects.filter(id__in=next_date_shiftId_list,intime='00:00',outtime='00:00',is_active=True).order_by('intime').first()
                                    if check_weekly_off is not None:
                                    
                                        if shift['outtime'] < shift['intime']:
                                
                                            check_current_shift_out = datetime.datetime.strptime(str(tomarrow_date) +' '+ shift['outtime'], '%Y-%m-%d %H:%M')
                                            check_current_shift_out_time = check_current_shift_out + timedelta(hours=2)
                                            given_datetime = datetime.datetime.strptime(str(check_current_shift_out_time), '%Y-%m-%d %H:%M:%S')

                                            truncated_datetime_str = given_datetime.strftime('%Y-%m-%d %H:%M')

                                            
                                            shift_end_date_time=str(truncated_datetime_str)
                                        else:
                                            shift_end_date_time=str(date) +' '+ '23:59'
                                            
                                        check_login_till = datetime.datetime.strptime(shift_end_date_time, '%Y-%m-%d %H:%M')
                                    else:
                                    
                                        check_login_till = datetime.datetime.strptime(str(tomarrow_date) +' '+ '07:30', '%Y-%m-%d %H:%M')
                                        
                            # print("shift "+str(count),shift['shiftname'],start_time_before_2hrs,check_login_till)
                            shift_data={
                                'shiftname':shift['shiftname'],
                                'indatetime':str(start_time_before_2hrs),
                                'outdatetime':str(check_login_till),
                                'intime':shift['intime'],
                                'outtime':shift['outtime'],
                            }
                            current_date_shift_list.append(shift_data)
                            count+=1

                        if len(current_shiftlist) == 0:       
                    
                            if_no_shift_check_shift=ShiftAllotment.objects.filter(attendanceId=employeeId,date=str(date),is_active=True)
                            todays_shiftlist=ShiftAllotmentshiftIdSerializer(if_no_shift_check_shift,many=True)
                            toadys_shiftId_list=list(todays_shiftlist.data)
                            check_weekly_off=ShiftMaster.objects.filter(id__in=toadys_shiftId_list,intime='00:00',outtime='00:00',is_active=True).order_by('intime').first()
                            if check_weekly_off is not None:
        
                                get_previous_day_shift=ShiftAllotment.objects.filter(attendanceId=employeeId,date=str(yesterday_date),is_active=True)
                                previous_day_shiftlist=ShiftAllotmentshiftIdSerializer(get_previous_day_shift,many=True)
                                previous_shiftId_list=list(previous_day_shiftlist.data)
                                get_previous_last_shift=ShiftMaster.objects.filter(id__in=previous_shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime').last()
                                
                                if get_previous_last_shift is not None:
                                    if get_previous_last_shift.outtime < get_previous_last_shift.intime:
                                        shift_end_date_time=str(date) +' '+ str(get_previous_last_shift.outtime)
                                    else:
                                        shift_end_date_time=str(yesterday_date) +' '+ '23:59'
                                            
                                    check_previous_shift_out = datetime.datetime.strptime(str(shift_end_date_time), '%Y-%m-%d %H:%M')
                                    previous_shift_in_time = check_previous_shift_out + timedelta(hours=2)  
                                else:
                    
                                    
                                    previous_shift_in_time= str(date)+' 00:00:00'
                                # checking previous  shift out time must be currrent shift in time in weekly off    # 
                                get_net_day_shift=ShiftAllotment.objects.filter(attendanceId=employeeId,date=str(tomarrow_date),is_active=True)
                                nest_day_shiftlist=ShiftAllotmentshiftIdSerializer(get_net_day_shift,many=True)
                                next_date_shiftId_list=list(nest_day_shiftlist.data)
                                get_next_day_first_shift=ShiftMaster.objects.filter(id__in=next_date_shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime').first()
                                if get_next_day_first_shift is not None:
                                    check_next_shift_in = datetime.datetime.strptime(str(tomarrow_date) +' '+ str(get_next_day_first_shift.intime), '%Y-%m-%d %H:%M')
                                    next_shift_in_time = check_next_shift_in - timedelta(hours=2)
                                    shift_data={
                                        'shiftname':str(check_weekly_off.shiftname),
                                        'indatetime':str(previous_shift_in_time),
                                        'outdatetime':str(next_shift_in_time),
                                        'intime':check_weekly_off.intime,
                                        'outtime':check_weekly_off.outtime,
                                    }
                                    current_date_shift_list.append(shift_data)
                                else:
                                    shift_data={
                                        'shiftname':str(check_weekly_off.shiftname),
                                        'indatetime':str(previous_shift_in_time),
                                        'outdatetime':str(date)+' 23:59:59',
                                        'intime':check_weekly_off.intime,
                                        'outtime':check_weekly_off.outtime,
                                    }
                                    current_date_shift_list.append(shift_data)
                            else:
                        
                                shift_data={
                                    'shiftname':'General',
                                    'indatetime':str(date)+' 07:30:00',
                                    'outdatetime':str(tomarrow_date)+' 07:30:00',
                                    'intime':'09:30',
                                    'outtime':'18:30',
                                }
                                current_date_shift_list.append(shift_data)
                                
                                
                        # print("current_date_shift_list",current_date_shift_list)
                        current_date_first_in_datetime=''
                        current_date_last_out_datetime=''
                        extra_days=''
                        total_working_hrs=''
                        
                        for s in current_date_shift_list:
                            intime=''
                            intimedate=''
                            
                            getallattendance = attendance.objects.filter(Q(employeeId=str(employeeId),time__gte=s['indatetime'].split(' ')[1],date=s['indatetime'].split(' ')[0])|Q(employeeId=str(employeeId),time__lte=str(s['outdatetime']).split(' ')[1],date=str(s['outdatetime']).split(' ')[0])).order_by('date','time')
                            
                            # getallattendance = getallattendance.filter(time__gte=currentshiftstarttime).order_by('date','time')
                            
                            attendance_serializer=attendanceserializer(getallattendance,many=True)
                        
                            
                            sorted_data = sorted(attendance_serializer.data, key=lambda x: (x['date'], x['time']))
                            
                            mindatetime = datetime.datetime.strptime(s['indatetime'], '%Y-%m-%d %H:%M:%S')
                            maxdatetime = datetime.datetime.strptime(s['outdatetime'], '%Y-%m-%d %H:%M:%S')
                        
                            sorted_data = [entry for entry in sorted_data if (mindatetime <= datetime.datetime.strptime(entry['date'] +' '+entry['time'],'%Y-%m-%d %H:%M:%S') <= maxdatetime)]

                            if len(sorted_data) > 0:
                                intimedate=sorted_data[0]['date']
                                intime=str(sorted_data[0]['time'])
                                
                            if intimedate !='' and intimedate is not None:
                                user_sdt = datetime.datetime.strptime(str(intimedate) + ' ' + str(intime), '%Y-%m-%d %H:%M:%S')
                                shif_sdt = datetime.datetime.strptime(s['indatetime'].split(' ')[0] + ' ' + s['indatetime'].split(' ')[1], '%Y-%m-%d %H:%M:%S')
                                if user_sdt < shif_sdt :
                                    intimedate=''
                                    intime=''

                                    
                            checkin_time = None
                            total_working_time = 0
                            attendance_log=[]
                            for index, entry in enumerate(sorted_data):
                                if entry['checkout']:
                                    attendance_log.append({'checkout':entry['date']+' '+entry['time']})
                                    if checkin_time:
                                        checkout_datetime = datetime.datetime.strptime(entry['date'] + ' ' + entry['time'], '%Y-%m-%d %H:%M:%S')
                                        checkin_datetime = datetime.datetime.strptime(checkin_time, '%Y-%m-%d %H:%M:%S')
                                        working_time = checkout_datetime - checkin_datetime
                                        total_working_time += working_time.total_seconds()
                                        checkin_time = None
                                elif not entry['checkout']:
                                    checkin_time = entry['date'] + ' ' + entry['time']
                                    attendance_log.append({'checkin':entry['date']+' '+entry['time']})

                            # If the last entry is check-in, consider checkout time as current shift end  time
                            if checkin_time and index == len(sorted_data) - 1:
                                
                                # check if the previous entry is also checkin or not if exist get the  difference betwnn  the current checkin  and  previous  checkin

                                if int(int(index)-1) >=0:
                                    if sorted_data[index-1]['checkout']==False:
                                        previous_checkin_date_time=sorted_data[index-1]['date']+ ' '+sorted_data[index-1]['time']
                                        checkout_datetime = datetime.datetime.strptime(previous_checkin_date_time, '%Y-%m-%d %H:%M:%S')
                                        checkin_datetime = datetime.datetime.strptime(checkin_time, '%Y-%m-%d %H:%M:%S')
                                        working_time = checkin_datetime-checkout_datetime 
                                        total_working_time += working_time.total_seconds()


                            # Convert total_working_time to hours, minutes, and seconds
                            hours, remainder = divmod(total_working_time, 3600)
                            minutes, seconds = divmod(remainder, 60)
                            # print("attendance_log",attendance_log)
                            s['total_hrs']=str(add_leading_zero(int(hours))) +':' +str(add_leading_zero(int(minutes))) +':'+str(add_leading_zero(int(seconds)))
                            def parse_timedelta(time_str):
                                hours, minutes, seconds = map(int, time_str.split(':'))
                                return timedelta(hours=hours, minutes=minutes, seconds=seconds)

                            dum_thrs =  parse_timedelta(s['total_hrs'])

                            # Add the timedelta objects
                            dum_thrs += totalworkedhrs
                            totalworkedhrs=dum_thrs
          
                            if total_working_hrs == '':
                                total_working_hrs=s['total_hrs']
                            else:
                                time_str_1 = total_working_hrs
                                time_str_2 = s['total_hrs']



                                # Convert time strings to timedelta objects and add them
                                total_time_delta = timedelta(hours=int(time_str_1[:2]), minutes=int(time_str_1[3:5]), seconds=int(time_str_1[6:])) + \
                                                timedelta(hours=int(time_str_2[:2]), minutes=int(time_str_2[3:5]), seconds=int(time_str_2[6:]))
                                total_working_hrs=str(total_time_delta)
                                
                            s['attendance_log']=attendance_log
                            if len(attendance_log) !=0:
                                infirst_key = next(iter(attendance_log[0]))
                                outfirst_key = next(iter(attendance_log[-1]))

                                s['usersintime']=attendance_log[0][infirst_key]
                                s['usersouttime']=attendance_log[-1][outfirst_key]
                                if s['usersouttime'] == s['usersintime']:
                                    s['usersouttime']= ''
                                    
                            else:
                                s['usersintime']=''
                                s['usersouttime']=''

            
                            
                            if current_date_first_in_datetime =='':
                                if s['usersintime'] !='':
                                    current_date_first_in_datetime=s['usersintime']
                            elif s['usersintime'] !='':
                                if datetime.datetime.strptime(current_date_first_in_datetime, '%Y-%m-%d %H:%M:%S') > datetime.datetime.strptime(s['usersintime'], '%Y-%m-%d %H:%M:%S'):
                                    current_date_first_in_datetime=s['usersintime']

                            if current_date_last_out_datetime =='':
                                if s['usersouttime'] !='':
                                    current_date_last_out_datetime=s['usersouttime']
                            elif s['usersouttime'] !='':
                                if datetime.datetime.strptime(current_date_last_out_datetime, '%Y-%m-%d %H:%M:%S') < datetime.datetime.strptime(s['usersouttime'], '%Y-%m-%d %H:%M:%S'):
                                    current_date_last_out_datetime=s['usersouttime']
                        
                            s['usersintime']=convert_datetime(s['usersintime'])
                            s['usersouttime']=convert_datetime(s['usersouttime'])
                            if current_date_last_out_datetime != '' and current_date_first_in_datetime !='':
                                if current_date_last_out_datetime.split(' ')[0] != current_date_first_in_datetime.split(' ')[0]:
                                    current_date_last_out_datetime1 = datetime.datetime.strptime(str(current_date_last_out_datetime).split(' ')[0], '%Y-%m-%d')
                                    current_date_first_in_datetime1 = datetime.datetime.strptime(str(current_date_first_in_datetime).split(' ')[0], '%Y-%m-%d')
                                    extradaysdiff = current_date_last_out_datetime1 - current_date_first_in_datetime1
                                    if extradaysdiff.days > 0:
                                        extra_days='+'+str(extradaysdiff.days)
                                
                                
                        
                        return_dict['employeeId']=employeeId
                        
                        
                        if len(current_date_shift_list) > 1:
                            return_dict['shift']=current_date_shift_list[0]
                            return_dict['shift']['shiftname'] = current_date_shift_list[0]['shiftname'] 
                            current_date_shift_list[0]['swap']=''
                        else:
                            return_dict['shift']=current_date_shift_list[0]
                            current_date_shift_list[0]['swap']=''

                        return_dict['shifts_list']=current_date_shift_list
                        
                        if current_date_first_in_datetime !='':
                            return_dict['inTime']=str(current_date_first_in_datetime).split(' ')[1]
                            return_dict['inTime_date']=convertdate2(str(str(current_date_first_in_datetime).split(' ')[0]))
                        else:
                            return_dict['inTime']=''
                            return_dict['inTime_date']=''
                            
                        if current_date_last_out_datetime !='':
                            return_dict['outTime']=str(current_date_last_out_datetime).split(' ')[1]
                            return_dict['outTime_date']=convertdate2(str(str(current_date_last_out_datetime).split(' ')[0]))
                        else:
                            
                            return_dict['outTime']=''
                            return_dict['outTime_date']=''
                        return_dict["status"]= ""
                        return_dict["sub_status"]= ''

                        return_dict['extra_days']=extra_days
                        return_dict['Total']=total_working_hrs
                        if return_dict['inTime'] !='':
                            return_dict['att_status']="P"
                            return_dict["status"]= "Working From Office"
                            presentcount+=1
                            return_dict["sub_status"]= 'P'

                        else:
                            return_dict['att_status']="A"
                            absentcount+=1
                            
                        foundleaves = [
                            leave for leave in leavesserializer.data if datetime.datetime.strptime(leave['start_date'], '%Y-%m-%d').date() <= datetime.datetime.strptime(date, '%Y-%m-%d').date() <= datetime.datetime.strptime(leave['end_date'], '%Y-%m-%d').date()
                        ]
                        
                        foundwfhs = [
                            wfh for wfh in wfhsserializer.data if datetime.datetime.strptime(wfh['start_date'], '%Y-%m-%d').date() <= datetime.datetime.strptime(date, '%Y-%m-%d').date() <= datetime.datetime.strptime(wfh['end_date'], '%Y-%m-%d').date()
                        ]
                        
                        if foundwfhs:
                            return_dict["status"]= "Working From Home"
                            return_dict['att_status']="WFH"
                        if foundleaves:
                            leave_count+=1
                            return_dict['att_status']="L"
                            return_dict["status"]= "Leave"

                        return_dict['inTime']=convert_to_12_hour_format(return_dict['inTime'])
                        return_dict['outTime']=convert_to_12_hour_format(return_dict['outTime'])
                        
                        check_shift_swap=ShiftAllotment.objects.filter(attendanceId=return_dict['employeeId'],date=date,is_active=True,is_swaped=True).first()
                        if check_shift_swap is not None:
                            return_dict["shiftswap"] ='Swaped with ' + str(check_shift_swap.swaped_employee_name)
                            check_swap_details=Shiftswap.objects.filter(id=check_shift_swap.swap_request_id).first()
                            if check_swap_details is not None:
                                Shiftswap_serializer=CustomShiftswapSerializer(check_swap_details)
                                if check_shift_swap.swapper == True:
                                    return_dict["shiftswap"]='Swaped with '+ str(Shiftswap_serializer.data['second_employee_name'])
                                elif check_shift_swap.swapper == False:
                                    return_dict["shiftswap"]='Swaped with '+ str(Shiftswap_serializer.data['first_employee_name'])
                                else:
                                    return_dict["shiftswap"]=''
                            else:
                                return_dict["shiftswap"]=''




                        else:
                            return_dict["shiftswap"]= ""
   
                        return_dict['date']=dateformat_ddmmyy(date)
                        return_dict["Day"]= get_day(return_dict["date"])
                        return_dict["reason"]= ""
                        
                        if return_dict['shift']['shiftname'].lower() == 'weeklyoff' or return_dict['shift']['shiftname'].lower() == 'weekly off' or return_dict['shift']['shiftname'].lower()=='weekly of':
                            return_dict["sub_status"]= 'Weekly Off'
                        attendance_report.append(return_dict)
                

                        # print("p['employeeId'] d2",p['employeeId'],attendance_report)
                        hours, remainder = divmod(totalworkedhrs.seconds, 3600)
                    d5={ 
                        'employeeId':p['employeeId'],
                        'Leave':leave_count,
                        'Holiday':holiday_count,
                        'totalworkedhrs':str(totalworkedhrs.days) +' days ,'+str(hours)+ ' hours',
                        'records':dict((v['date'],v) for v in attendance_report).values() ,
                        'present':presentcount,
                        'absent':absentcount,
                    }
                    d6.append(d5)
                    attendance_report.clear()

        if os.path.exists("static/excel/monthlydata.xlsx"):
            os.remove("static/excel/monthlydata.xlsx")
            workbook = xlsxwriter.Workbook('static/excel/monthlydata.xlsx')
            workbook.close()
        else:
            workbook = xlsxwriter.Workbook('static/excel/monthlydata.xlsx')
            workbook.close()
            
            
        if len(mon_serializer.data) >1:
            Allshiftattendancereport(d6)
        else:
            shiftattendancereport(d6)
        
        return Response({"response":{"data":d6,"url":imageUrl + '/static/excel/monthlydata.xlsx', "n":2 ,"msg" : "Excel report generated successfully ","status":"success"}})  
    else:
        return Response({"response":{"url": '', "n":0 ,"msg" : "Operation failed please provide month and year","status":"error"}})  
        



        

def create_google_maps_url(latitude, longitude):
    base_url = "https://www.google.com/maps/search/?api=1&query={},{}"
    return base_url.format(latitude, longitude)


def get_location_name(latitude, longitude):
    base_url = "https://nominatim.openstreetmap.org/reverse"
    params = {
        "format": "json",
        "lat": latitude,
        "lon": longitude,
    }
    response = requests.get(base_url, params=params)
    if response.status_code == 200:
        data = response.json()
        display_name = data.get('display_name', "Location not found")
        return display_name
    else:
        return "Failed to retrieve location"
    
    
import math
def haversine(lat1, lon1, lat2, lon2):
    """
    Calculate the great circle distance between two points 
    on the earth (specified in decimal degrees)
    """
    # Convert decimal degrees to radians
    lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])

    # Haversine formula
    dlon = lon2 - lon1
    dlat = lat2 - lat1
    a = math.sin(dlat/2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    r = 6371  # Radius of Earth in kilometers
    return r * c * 1000  # Distance in meters

def is_within_radius(lat1, lon1, lat2, lon2, radius):
    """
    Check if a point (lat1, lon1) is within a certain radius (in meters)
    of another point (lat2, lon2)
    """
    distance = haversine(lat1, lon1, lat2, lon2)
    return distance <= radius


@api_view(['POST'])    
@permission_classes((AllowAny,))
def get_the_user_device_location(request):
    print("request",request.POST)
    latitude=request.POST.get('latitude')
    longitude=request.POST.get('longitude')
    remote_map_name=''
    remote_map_location=''
    
    if longitude is not None and longitude !="" and longitude is not None and longitude !="":
        office_location=Location.objects.filter(Active=True).exclude(Q(lattitude='',meter='',longitude='') |Q(lattitude=None,meter=None,longitude=None)|Q(lattitude__isnull=True,meter__isnull=True,longitude__isnull=True))
        location_serializer=LocationSerializer(office_location,many=True)
        
        for location in location_serializer.data:
            within_radius = is_within_radius(float(latitude), float(longitude), float(location['lattitude']), float(location['longitude']), float(location['Radius']))
            if within_radius:
                remote_map_location=create_google_maps_url(latitude,longitude)
                remote_map_name='Zentro '+str(location['LocationName']) +' Office'
                
        
        if remote_map_name is not None and remote_map_name !='':
                
            return Response ({"data":{'location_name':remote_map_name,'location_url':remote_map_location},"response":{"n" : 1,"msg" : "employee is in office location ","status" : "success"}})
        else:
            remote_map_name=get_location_name(latitude,longitude)
            remote_map_location=create_google_maps_url(latitude,longitude)
            return Response ({"data":{'location_name':remote_map_name,'location_url':remote_map_location},"response":{"n" : 0,"msg" : "employee is not in office location ","status" : "error"}})
    else:
        remote_map_name=''
        return Response ({"data":{'location_name':remote_map_name,'location_url':remote_map_location},"response":{"n" : 0,"msg" : "lattitude and longitude not found","status" : "error"}})
    
@api_view(['POST'])
@permission_classes((AllowAny,))
def get_current_shift(request):
    employeeId = request.POST.get('UserId')
    current_date = datetime.date.today()
    yesterday_date = current_date - datetime.timedelta(days=1)
    current_datetime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    intime=''
    outtime=''
    intimedate=''
    outtimedate=''
    userobj = Users.objects.filter(employeeId=employeeId,is_active=True).first()

    
    if userobj is not None:   
        check_user_type=userobj.employeetype
        check_last_checkin = attendance.objects.filter(employeeId=str(employeeId),date=str(current_date),checkout=False).order_by('time').last()
        check_last_checkout=None
        if check_last_checkin is not None:
            check_last_checkout = attendance.objects.filter(employeeId=str(employeeId),time__gt=check_last_checkin.time,date=str(current_date),checkout=True).order_by('time').last()
            if check_last_checkout is not None:
                print("today allow him to checkin")
            else: 
                print("today allow him to checkout")
        else:
            check_last_checkin = attendance.objects.filter(employeeId=str(employeeId),date=str(yesterday_date),checkout=False).order_by('time').last()
            if check_last_checkin is not None:
                check_last_checkout = attendance.objects.filter(Q(employeeId=str(employeeId),time__gt=check_last_checkin.time,date=str(yesterday_date),checkout=True)|Q(employeeId=str(employeeId),date=str(current_date),checkout=True)).order_by('date','time').last()
                if check_last_checkout is not None:
                    print("yesterday allow him to checkin")
                else: 
                    print("yesterday allow him to checkout")
            else:
                print("allow him to checkin")

        
        if check_last_checkin is not None :
            
            intime = check_last_checkin.time
            intimedate=check_last_checkin.date
            punchout = 0
            get_data=1
            outtime = ''
            outtimedate=''
        else:
            get_data=0
            punchout = 1
            intime = ''
            intimedate=''
            outtime = ''
            outtimedate=''
            
            
        if check_last_checkout is not None:
            
            outtime = check_last_checkout.time
            outtimedate=check_last_checkout.date
            punchout = 1
            get_data=0
            # intime = ''
            # intimedate=''
        else:
            if check_last_checkin is not None :
                punchout = 0
                get_data=1
                outtime = ''
                outtimedate=''
            else:
                punchout = 1
                get_data=0
                outtime = ''
                outtimedate=''
                

        currentshiftname=''
        currentshiftstarttime=''
        currentshiftendtime=''
        currentshiftstartdate=''
        currentshiftenddate=''
        todays_shiftallotment_objs=ShiftAllotment.objects.filter(attendanceId=employeeId,date=str(current_date),is_active=True)
        todays_shiftallotment_serializers=ShiftAllotmentshiftIdSerializer(todays_shiftallotment_objs,many=True)
        shiftId_list=list(todays_shiftallotment_serializers.data)
        shift_obj=ShiftMaster.objects.filter(id__in=shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime')
        shift_serializer=ShiftMasterSerializer(shift_obj,many=True)
        
        
        # check for todays shift
        todays_runnningshift=gettodaysshift(shift_serializer.data,str(current_date))
        if todays_runnningshift['n'] == 1:
            currentshiftname=todays_runnningshift['data']['shiftname']
            currentshiftstarttime=str(todays_runnningshift['data']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
            currentshiftendtime=str(todays_runnningshift['data']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
            currentshiftstartdate=str(todays_runnningshift['data']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
            currentshiftenddate=str(todays_runnningshift['data']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
            
        elif todays_runnningshift['last_runingshift']['shiftstarttime'] !='':
            currentshiftname=todays_runnningshift['last_runingshift']['shiftname']
            currentshiftstarttime=str(todays_runnningshift['last_runingshift']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
            currentshiftendtime=str(todays_runnningshift['last_runingshift']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
            currentshiftstartdate=str(todays_runnningshift['last_runingshift']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
            currentshiftenddate=str(todays_runnningshift['last_runingshift']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
        
        else:
            yesterday_shiftallotment_objs=ShiftAllotment.objects.filter(attendanceId=employeeId,date=str(yesterday_date),is_active=True)
            yesterday_shiftallotment_serializers=ShiftAllotmentshiftIdSerializer(yesterday_shiftallotment_objs,many=True)
            shiftId_list=list(yesterday_shiftallotment_serializers.data)
            shift_obj=ShiftMaster.objects.filter(id__in=shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime')
            yesterday_shift_serializer=ShiftMasterSerializer(shift_obj,many=True)
                
            
            yesterday_runnningshift=gettodaysshift(yesterday_shift_serializer.data,str(yesterday_date))
            if yesterday_runnningshift['n'] == 1:
                currentshiftname=yesterday_runnningshift['data']['shiftname']
                currentshiftstarttime=str(yesterday_runnningshift['data']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                currentshiftendtime=str(yesterday_runnningshift['data']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                currentshiftstartdate=str(yesterday_runnningshift['data']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                currentshiftenddate=str(yesterday_runnningshift['data']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                
            elif yesterday_runnningshift['last_runingshift']['shiftstarttime'] !='':
                currentshiftname=yesterday_runnningshift['last_runingshift']['shiftname']
                currentshiftstarttime=str(yesterday_runnningshift['last_runingshift']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                currentshiftendtime=str(yesterday_runnningshift['last_runingshift']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[1]+':00'
                currentshiftstartdate=str(yesterday_runnningshift['last_runingshift']['shiftstarttime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
                currentshiftenddate=str(yesterday_runnningshift['last_runingshift']['shiftendtime'].strftime('%Y-%m-%d %H:%M')).split(' ')[0]
            else:
                currentshiftname='General'
                currentshiftstarttime='07:30:00'
                currentshiftendtime='18:30:00'
                currentshiftstartdate=str(current_date)
                currentshiftenddate=str(current_date)

        shiftdetails={
                    "shiftname":currentshiftname,
                    "shiftstarttime":currentshiftstarttime,
                    "shiftendtime":currentshiftendtime,
                    "shiftstartdate":currentshiftstartdate,
                    "shiftenddate":currentshiftenddate,
                }
        

        getallattendance = attendance.objects.filter(Q(employeeId=str(employeeId),time__gte=currentshiftstarttime,date=str(currentshiftstartdate))|Q(employeeId=str(employeeId),time__lte=str(current_datetime).split(' ')[1],date=str(current_datetime).split(' ')[0])).order_by('date','time')
        
        # getallattendance = getallattendance.filter(time__gte=currentshiftstarttime).order_by('date','time')
        
        attendance_serializer=attendanceserializer(getallattendance,many=True)
        
        sorted_data = sorted(attendance_serializer.data, key=lambda x: (x['date'], x['time']))
        
        mindate = datetime.datetime.strptime(currentshiftstartdate, '%Y-%m-%d')
        mintime = datetime.datetime.strptime(currentshiftstarttime, '%H:%M:%S').time()

        sorted_data = [entry for entry in sorted_data if (datetime.datetime.strptime(entry['date'],'%Y-%m-%d').date() > mindate.date() or (datetime.datetime.strptime(entry['date'],'%Y-%m-%d').date() == mindate.date() and datetime.datetime.strptime(entry['time'], '%H:%M:%S').time() > mintime))]

        if len(sorted_data) > 0:
            intimedate=sorted_data[0]['date']
            intime=str(sorted_data[0]['time'])
            
        if intimedate !='' and intimedate is not None:
            user_sdt = datetime.datetime.strptime(str(intimedate) + ' ' + str(intime), '%Y-%m-%d %H:%M:%S')
            shif_sdt = datetime.datetime.strptime(str(currentshiftstartdate) + ' ' + str(currentshiftstarttime), '%Y-%m-%d %H:%M:%S')
            if user_sdt < shif_sdt :
                intimedate=''
                intime=''
                outtime=''
                outtimedate=''

                
        checkin_time = None
        total_working_time = 0
        for index, entry in enumerate(sorted_data):
            if entry['checkout']:
                if checkin_time:
                    checkout_datetime = datetime.datetime.strptime(entry['date'] + ' ' + entry['time'], '%Y-%m-%d %H:%M:%S')
                    checkin_datetime = datetime.datetime.strptime(checkin_time, '%Y-%m-%d %H:%M:%S')
                    working_time = checkout_datetime - checkin_datetime
                    total_working_time += working_time.total_seconds()
                    checkin_time = None
            elif not entry['checkout']:
                checkin_time = entry['date'] + ' ' + entry['time']

        # If the last entry is check-in, consider checkout time as current time
        if checkin_time and index == len(sorted_data) - 1:
            checkout_datetime = datetime.datetime.now()
            checkin_datetime = datetime.datetime.strptime(checkin_time, '%Y-%m-%d %H:%M:%S')
            working_time = checkout_datetime - checkin_datetime
            total_working_time += working_time.total_seconds()


        # Convert total_working_time to hours, minutes, and seconds
        hours, remainder = divmod(total_working_time, 3600)
        minutes, seconds = divmod(remainder, 60)

        current_shift_start_datetime = datetime.datetime.strptime(shiftdetails['shiftstartdate'] + ' ' + shiftdetails['shiftstarttime'], '%Y-%m-%d %H:%M:%S')
        current_shift_end_datetime = datetime.datetime.strptime(shiftdetails['shiftenddate'] + ' ' + shiftdetails['shiftendtime'], '%Y-%m-%d %H:%M:%S')
        shift_total_working_hrs = current_shift_end_datetime - current_shift_start_datetime
        shift_total_working_hrs -= timedelta(hours=2) 
        current_shift_start_datetime += timedelta(hours=2)

        if intime is not None and intime !='' and intimedate is not None and intimedate !='':
            current_user_checkin_datetime = datetime.datetime.strptime(intimedate + ' ' + intime, '%Y-%m-%d %H:%M:%S')
            if current_user_checkin_datetime > (current_shift_start_datetime + timedelta(minutes=30)):
                LateMark=True
            else:
                LateMark=False    
            present=True
        else:
            present=False
            LateMark=False  

        current_shift_details={
            'shiftname':shiftdetails['shiftname'],
            'current_shift_start_datetime':current_shift_start_datetime,
            'current_shift_end_datetime':current_shift_end_datetime,
        }
        total_hrs=str(add_leading_zero(int(hours))) +':' +str(add_leading_zero(int(minutes))) +':'+str(add_leading_zero(int(seconds)))
        return Response ({
            'total_hrs':total_hrs,
            'in_time':intime,
            'out_time':outtime,
            'in_date':intimedate,
            'out_date':outtimedate,
            'shift_details':current_shift_details,
            'present':present,
            'LateMark':LateMark,
            "response":{
                "n" : 1,
                "msg" : "pass",
                "status" : "success"
            }
            })
        

    else:
        
        return Response ({

                    'total_hrs':'',
                    'in_time':'',
                    'out_time':'',
                    'in_date':'',
                    'out_date':'',
                    'shift_details':{},
                    'present':False,
                    'LateMark':False,
                    "response":{
                        "n" : 0,
                        "msg" : "user not found",
                        "status" : "errror"
                    }
                })
      
@api_view(['POST'])
@permission_classes((AllowAny,))
def compoff_shedular(request):
    current_date = datetime.datetime.now()
    date_after_60_days = current_date + timedelta(days=60)
    date_after_60_days=date_after_60_days.strftime('%Y-%m-%d')
    users_obj=Users.objects.filter(is_active=True).exclude(Q(employeeId__isnull=True)|Q(employeeId='None')|Q(employeeId='')).order_by('Firstname','Lastname')
    usersserializer=UsersSerializer(users_obj,many=True)
    for user in usersserializer.data:

        attendanceId=user['employeeId']
        check_eligiblity=compoff_eligiblity(attendanceId)

        if check_eligiblity['present']==True:
            if check_eligiblity['over_time']==True:
                notificationmsg='Hi <span class="actionuser">'+str(user['Firstname']) + ' ' +str(user['Lastname'])+'</span> are you still working.'
                TaskNotification.objects.create(UserID_id=user['id'],To_manager=False,leaveID=0,company_code=user['company_code'],NotificationTypeId_id=7,NotificationTitle="Working Status",NotificationMsg=notificationmsg)
                
            if check_eligiblity['eligible']==True:
                compoff_date=datetime.datetime.strptime(str(check_eligiblity['shift_details']['current_shift_start_datetime']).split(' ')[0] + ' ' + '05:30:00', '%Y-%m-%d %H:%M:%S')
                check_already_exist_compoff=EligibleCompOffMaster.objects.filter(user_id=user['id'],date=compoff_date,shift_name=check_eligiblity['shift_details']['shiftname']).first()
                if check_already_exist_compoff is None:

                    compoff_data={}
                    compoff_data['user_id']=user['id']
                    compoff_data['user_name']=str(user['Firstname']) + ' ' +str(user['Lastname'])
                    compoff_data['date']=compoff_date
                    compoff_data['working_hrs']=check_eligiblity['total_hrs']
                    compoff_data['shift_name']=check_eligiblity['shift_details']['shiftname']
                    compoff_data['valid_date']=datetime.datetime.strptime(str(date_after_60_days) + ' ' + '05:30:00', '%Y-%m-%d %H:%M:%S')

                    compoff_serializer=save_compoff_serializers(data=compoff_data)
                    if compoff_serializer.is_valid():
                        compoff_serializer.save()
                        notificationmsg='Hi <span class="actionuser">'+str(user['Firstname']) + ' ' +str(user['Lastname'])+'</span> you are eligible for compoff'
                        TaskNotification.objects.create(UserID_id=user['id'],To_manager=False,leaveID=0,company_code=user['company_code'],NotificationTypeId_id=8,NotificationTitle="CompOff",NotificationMsg=notificationmsg)
                        managers_obj=leaveMapping.objects.filter(employeeId=user['id']).distinct("managerId")
                        managers_serializers=leaveMappingserializer(managers_obj,many=True)
                        for manager in managers_serializers.data:
                            managernotificationmsg='Hi <span class="actionuser">'+str(user['Firstname']) + ' ' +str(user['Lastname'])+'</span> is eligible for compoff'
                            TaskNotification.objects.create(UserID_id=manager['managerId'],To_manager=False,leaveID=0,company_code=user['company_code'],NotificationTypeId_id=8,NotificationTitle="CompOff",NotificationMsg=managernotificationmsg)
                            CompoffApproval.objects.create(user_id=user['id'],manager_id=manager['managerId'],eligible_compoff_id=compoff_serializer.data['id'])
            
            



        


    return Response ({"response":{"n" : 1,"msg" : "SUCCESS","status" : "success"}})

def two_months_from_now():
    return datetime.now() + timedelta(days=60)

@api_view(['POST'])    
def get_user_available_compoff(request):
    user_id = request.user.id
    current_date = datetime.datetime.now()
    formatted_date = current_date.strftime('%Y-%m-%d')
    available_compoff = EligibleCompOffMaster.objects.filter(user_id=user_id,claim=False,valid_date__gte=formatted_date,is_active=True).order_by('date')
    if available_compoff.exists():
        compoff_serializer = compoff_serializers(available_compoff,many=True)
        return Response ({"data":compoff_serializer.data,"response":{"n" : 1,"msg" : "available comp-offs found ","status" : "success"}})
    else:
        return Response ({"data":[],"response":{"n" : 0,"msg" : "No available comp-offs found ","status" : "error"}})

@api_view(['POST'])    
def get_user_pending_compoff(request):
    user_id = request.user.id
    current_date = datetime.datetime.now()
    formatted_date = current_date.strftime('%Y-%m-%d')
    pending_compoff = ClaimedCompOffMaster.objects.filter(user_id=user_id,valid_date__gte=formatted_date,is_active=True,status='Pending',claim_date__gte=formatted_date).order_by('-claim_date')
    if pending_compoff.exists():
        compoff_serializer = claimed_compoff_serializers(pending_compoff,many=True)
        compoffs=compoff_serializer.data
        for compoff in compoffs:
            compoff_approvals_obj=CompoffApproval.objects.filter(claimed_compoff_id=compoff['id'])
            compoff_managers_serializer=compoff_approval_serializers(compoff_approvals_obj,many=True)
            compoff['managers']=compoff_managers_serializer.data
            editable_obj=compoff_approvals_obj.exclude(status=None).first()
            if editable_obj is not None:
                compoff['editable']=False
            else:
                compoff['editable']=True

            claim_date_obj = datetime.datetime.strptime(compoff['claim_date'], "%d %b %Y").strftime("%Y-%m-%d")
            if claim_date_obj > formatted_date:     
                compoff['withdrawable']=True
            else:
                compoff['withdrawable']=False


        return Response ({"data":compoff_serializer.data,"response":{"n" : 1,"msg" : "pending comp-offs found ","status" : "success"}})
    else:
        return Response ({"data":[],"response":{"n" : 0,"msg" : "No pending comp-offs found ","status" : "error"}})

@api_view(['POST'])    
def get_user_approved_compoff(request):
    user_id = request.user.id
    current_date = datetime.datetime.now()
    formatted_date = current_date.strftime('%Y-%m-%d')
    approved_compoff = ClaimedCompOffMaster.objects.filter(user_id=user_id,is_active=True,status='Approved').order_by('-claim_date')
    if approved_compoff.exists():
        compoff_serializer = claimed_compoff_serializers(approved_compoff,many=True)
        compoffs=compoff_serializer.data
        for compoff in compoffs:
            compoff_approvals_obj=CompoffApproval.objects.filter(claimed_compoff_id=compoff['id'])
            compoff_managers_serializer=compoff_approval_serializers(compoff_approvals_obj,many=True)
            compoff['managers']=compoff_managers_serializer.data

            claim_date_obj = datetime.datetime.strptime(compoff['claim_date'], "%d %b %Y").strftime("%Y-%m-%d")
            if claim_date_obj > formatted_date:     
                compoff['withdrawable']=True
            else:
                compoff['withdrawable']=False
        return Response ({"data":compoff_serializer.data,"response":{"n" : 1,"msg" : "approved comp-offs found ","status" : "success"}})
    else:
        return Response ({"data":[],"response":{"n" : 0,"msg" : "No approved comp-offs found ","status" : "error"}})

@api_view(['POST'])    
def get_user_withdraw_compoff(request):
    user_id = request.user.id
    current_date = datetime.datetime.now()
    formatted_date = current_date.strftime('%Y-%m-%d')
    withdraw_compoff = ClaimedCompOffMaster.objects.filter(user_id=user_id,is_active=True,status='Withdraw').order_by('-claim_date')
    if withdraw_compoff.exists():
        compoff_serializer = claimed_compoff_serializers(withdraw_compoff,many=True)
        compoffs=compoff_serializer.data
        for compoff in compoffs:
            compoff_approvals_obj=CompoffApproval.objects.filter(claimed_compoff_id=compoff['id'])
            compoff_managers_serializer=compoff_approval_serializers(compoff_approvals_obj,many=True)
            compoff['managers']=compoff_managers_serializer.data

            claim_date_obj = datetime.datetime.strptime(compoff['claim_date'], "%d %b %Y").strftime("%Y-%m-%d")
            if claim_date_obj > formatted_date:     
                compoff['withdrawable']=True
            else:
                compoff['withdrawable']=False
        return Response ({"data":compoff_serializer.data,"response":{"n" : 1,"msg" : "withdraw comp-offs found ","status" : "success"}})
    else:
        return Response ({"data":[],"response":{"n" : 0,"msg" : "No withdraw comp-offs found ","status" : "error"}})

@api_view(['POST'])    
def get_user_rejected_compoff(request):
    user_id = request.user.id
    rejected_compoff = ClaimedCompOffMaster.objects.filter(user_id=user_id,is_active=True,status='Rejected').order_by('-claim_date')
    if rejected_compoff.exists():
        compoff_serializer = claimed_compoff_serializers(rejected_compoff,many=True)
        compoffs=compoff_serializer.data
        for compoff in compoffs:
            compoff_approvals_obj=CompoffApproval.objects.filter(claimed_compoff_id=compoff['id'])
            compoff_managers_serializer=compoff_approval_serializers(compoff_approvals_obj,many=True)
            compoff['managers']=compoff_managers_serializer.data
            compoff_approvals_rejected_obj=CompoffApproval.objects.filter(claimed_compoff_id=compoff['id'],status=False).first()
            if compoff_approvals_rejected_obj is not None:
                compoff['reason']=compoff_approvals_rejected_obj.rejected_reason
            else:
                compoff['reason']='Not Available'

        return Response ({"data":compoff_serializer.data,"response":{"n" : 1,"msg" : "rejected comp-offs found ","status" : "success"}})
    else:
        return Response ({"data":[],"response":{"n" : 0,"msg" : "No rejected comp-offs found ","status" : "error"}})

@api_view(['POST'])    
def get_user_expired_compoff(request):
    user_id = request.user.id
    current_date = datetime.datetime.now()
    formatted_date = current_date.strftime('%Y-%m-%d')
    applied_expired_compoff = ClaimedCompOffMaster.objects.filter(user_id=user_id,claim_date__lt=formatted_date,is_active=True).exclude(Q(status='Approved')|Q(status='Rejected')).order_by('-claim_date')


    unapplied_expired_compoff = EligibleCompOffMaster.objects.filter(user_id=user_id,valid_date__lt=formatted_date,is_active=True).order_by('-date')
    applied_serializer=claimed_eligiliblecompoff_id_serializer(applied_expired_compoff,many=True)
    unapplied_expired_compoff.exclude(id__in=list(applied_serializer.data))#exclude the claimed compoffs
    
    compoffs=[]

    applied_compoff_serializer = claimed_compoff_serializers(applied_expired_compoff,many=True)
    for compoff in applied_compoff_serializer.data:
        compoff_approvals_obj=CompoffApproval.objects.filter(claimed_compoff_id=compoff['id'])
        compoff_managers_serializer=compoff_approval_serializers(compoff_approvals_obj,many=True)
        compoff['managers']=compoff_managers_serializer.data
        compoffs.append(compoff)

    unapplied_compoff_serializer = compoff_serializers(unapplied_expired_compoff,many=True)
    for compoff in unapplied_compoff_serializer.data:
        compoff_approvals_obj=CompoffApproval.objects.filter(eligible_compoff_id=compoff['id'])
        compoff_managers_serializer=compoff_approval_serializers(compoff_approvals_obj,many=True)
        compoff['managers']=compoff_managers_serializer.data
        compoffs.append(compoff)




    if len(compoffs)!=0:
        return Response ({"data":compoffs,"response":{"n" : 1,"msg" : "expired comp-offs found ","status" : "success"}})
    else:
        return Response ({"data":[],"response":{"n" : 0,"msg" : "No expired comp-offs found ","status" : "error"}})

@api_view(['POST'])    
def get_user_reschedule_compoff(request):
    user_id = request.user.id
    reschedule_compoff = ClaimedCompOffMaster.objects.filter(user_id=user_id,is_active=True,status='Reschedule').order_by('-claim_date')
    if reschedule_compoff.exists():
        compoff_serializer = claimed_compoff_serializers(reschedule_compoff,many=True)
        compoffs=compoff_serializer.data
        for compoff in compoffs:
            compoff_approvals_obj=CompoffApproval.objects.filter(claimed_compoff_id=compoff['id'])
            compoff_managers_serializer=compoff_approval_serializers(compoff_approvals_obj,many=True)
            compoff['managers']=compoff_managers_serializer.data
            compoff_approvals_reschedule_obj=CompoffApproval.objects.filter(claimed_compoff_id=compoff['id'],reschedule=True).first()
            if compoff_approvals_reschedule_obj is not None:
                compoff['reason']=compoff_approvals_reschedule_obj.reschedule_reason
            else:
                compoff['reason']='Not Available'

        return Response ({"data":compoff_serializer.data,"response":{"n" : 1,"msg" : "reschedule comp-offs found ","status" : "success"}})
    else:
        return Response ({"data":[],"response":{"n" : 0,"msg" : "No reschedule comp-offs found ","status" : "error"}})



@api_view(['POST'])    
def get_manager_pending_compoff_requests(request):
    user_id = request.user.id
    current_date = datetime.datetime.now()
    formatted_date = current_date.strftime('%Y-%m-%d')
    compoff_approval_obj=CompoffApproval.objects.filter(manager_id=user_id,status=None).exclude(reschedule=True).distinct('claimed_compoff_id')
    compoff_ids_serializer=compoff_id_serializer(compoff_approval_obj,many=True)
    pending_compoff = ClaimedCompOffMaster.objects.filter(id__in=list(compoff_ids_serializer.data),claim_date__gte=formatted_date,is_active=True,status='Pending').order_by('-claim_date')

    if pending_compoff.exists():
        compoff_serializer = claimed_compoff_serializers(pending_compoff,many=True)
        compoffs=compoff_serializer.data
        for compoff in compoffs:
            earlier_reschedule_compoff_approval=CompoffApproval.objects.filter(claimed_compoff_id=compoff['earlier_reschedule_id'],reschedule=True).first()
            earlier_reschedule_compoff_approval_serializer=compoff_approval_serializers(earlier_reschedule_compoff_approval)
            print("earlier_reschedule", compoff['earlier_reschedule'])

            if earlier_reschedule_compoff_approval is not None:
                compoff['reason']='- '+earlier_reschedule_compoff_approval_serializer.data['manager_name'] +': '+earlier_reschedule_compoff_approval.reschedule_reason
            else:
                compoff['reason']='Earlier reschedule'

            compoff_approvals_obj=CompoffApproval.objects.filter(claimed_compoff_id=compoff['id'])
            compoff_managers_serializer=compoff_approval_serializers(compoff_approvals_obj,many=True)
            compoff['managers']=compoff_managers_serializer.data
            
        return Response ({"data":compoff_serializer.data,"response":{"n" : 1,"msg" : "pending comp-offs found ","status" : "success"}})
    else:
        return Response ({"data":[],"response":{"n" : 0,"msg" : "No pending comp-offs found ","status" : "error"}})


@api_view(['POST'])    
def get_manager_approved_compoff_requests(request):
    user_id = request.user.id
    compoff_approval_obj=CompoffApproval.objects.filter(manager_id=user_id,status=True).distinct('claimed_compoff_id')
    compoff_ids_serializer=compoff_id_serializer(compoff_approval_obj,many=True)
    approved_compoff = ClaimedCompOffMaster.objects.filter(id__in=list(compoff_ids_serializer.data),is_active=True).exclude(Q(status='Rejected')|Q(status='Reschedule')|Q(status='Withdraw')).order_by('-claim_date')

    if approved_compoff.exists():
        compoff_serializer = claimed_compoff_serializers(approved_compoff,many=True)
        compoffs=compoff_serializer.data
        for compoff in compoffs:
            compoff_approvals_obj=CompoffApproval.objects.filter(claimed_compoff_id=compoff['id'])
            compoff_managers_serializer=compoff_approval_serializers(compoff_approvals_obj,many=True)
            compoff['managers']=compoff_managers_serializer.data
            
        return Response ({"data":compoff_serializer.data,"response":{"n" : 1,"msg" : "approved comp-offs found ","status" : "success"}})
    else:
        return Response ({"data":[],"response":{"n" : 0,"msg" : "No approved comp-offs found ","status" : "error"}})

@api_view(['POST'])    
def get_manager_rejected_compoff_requests(request):
    user_id = request.user.id
    compoff_approval_obj=CompoffApproval.objects.filter(manager_id=user_id).distinct('claimed_compoff_id')
    compoff_ids_serializer=compoff_id_serializer(compoff_approval_obj,many=True)

    rejected_compoff = ClaimedCompOffMaster.objects.filter(id__in=list(compoff_ids_serializer.data),is_active=True,status='Rejected').order_by('-claim_date')
    if rejected_compoff.exists():
        compoff_serializer = claimed_compoff_serializers(rejected_compoff,many=True)
        compoffs=compoff_serializer.data
        for compoff in compoffs:
            compoff_approvals_obj=CompoffApproval.objects.filter(claimed_compoff_id=compoff['id'])
            compoff_managers_serializer=compoff_approval_serializers(compoff_approvals_obj,many=True)
            compoff['managers']=compoff_managers_serializer.data
            compoff_approvals_rejected_obj=CompoffApproval.objects.filter(claimed_compoff_id=compoff['id'],status=False).first()
            if compoff_approvals_rejected_obj is not None:
                compoff['reason']=compoff_approvals_rejected_obj.rejected_reason
            else:
                compoff['reason']='Not Available'



        return Response ({"data":compoff_serializer.data,"response":{"n" : 1,"msg" : "rejected comp-offs found ","status" : "success"}})
    else:
        return Response ({"data":[],"response":{"n" : 0,"msg" : "No rejected comp-offs found ","status" : "error"}})

@api_view(['POST'])    
def get_manager_expired_compoff_requests(request):
    user_id = request.user.id
    current_date = datetime.datetime.now()
    formatted_date = current_date.strftime('%Y-%m-%d')
    compoff_approval_obj=CompoffApproval.objects.filter(manager_id=user_id,status=None).distinct('claimed_compoff_id')
    compoff_ids_serializer=compoff_id_serializer(compoff_approval_obj,many=True)

    expired_compoff = ClaimedCompOffMaster.objects.filter(id__in=list(compoff_ids_serializer.data),is_active=True,claim_date__lt=formatted_date,status='Pending').order_by('-claim_date')
    if expired_compoff.exists():
        compoff_serializer = claimed_compoff_serializers(expired_compoff,many=True)
        compoffs=compoff_serializer.data
        for compoff in compoffs:
            compoff_approvals_obj=CompoffApproval.objects.filter(claimed_compoff_id=compoff['id'])
            compoff_managers_serializer=compoff_approval_serializers(compoff_approvals_obj,many=True)
            compoff['managers']=compoff_managers_serializer.data
            
        return Response ({"data":compoff_serializer.data,"response":{"n" : 1,"msg" : "expired comp-offs found ","status" : "success"}})
    else:
        return Response ({"data":[],"response":{"n" : 0,"msg" : "No expired comp-offs found ","status" : "error"}})

@api_view(['POST'])    
def get_manager_withdraw_compoff_requests(request):
    user_id = request.user.id
    compoff_approval_obj=CompoffApproval.objects.filter(manager_id=user_id).distinct('claimed_compoff_id')
    compoff_ids_serializer=compoff_id_serializer(compoff_approval_obj,many=True)
    withdraw_compoff = ClaimedCompOffMaster.objects.filter(id__in=list(compoff_ids_serializer.data),is_active=True,status='Withdraw').order_by('-claim_date')

    if withdraw_compoff.exists():
        compoff_serializer = claimed_compoff_serializers(withdraw_compoff,many=True)
        compoffs=compoff_serializer.data
        for compoff in compoffs:
            compoff_approvals_obj=CompoffApproval.objects.filter(claimed_compoff_id=compoff['id'])
            compoff_managers_serializer=compoff_approval_serializers(compoff_approvals_obj,many=True)
            compoff['managers']=compoff_managers_serializer.data
            
        return Response ({"data":compoff_serializer.data,"response":{"n" : 1,"msg" : "withdraw comp-offs found ","status" : "success"}})
    else:
        return Response ({"data":[],"response":{"n" : 0,"msg" : "No withdraw comp-offs found ","status" : "error"}})

@api_view(['POST'])    
def get_manager_reschedule_compoff_requests(request):
    user_id = request.user.id
    compoff_approval_obj=CompoffApproval.objects.filter(manager_id=user_id).distinct('claimed_compoff_id')
    compoff_ids_serializer=compoff_id_serializer(compoff_approval_obj,many=True)
    reschedule_compoff = ClaimedCompOffMaster.objects.filter(id__in=list(compoff_ids_serializer.data),is_active=True,status='Reschedule').order_by('-claim_date')

    if reschedule_compoff.exists():
        compoff_serializer = claimed_compoff_serializers(reschedule_compoff,many=True)
        compoffs=compoff_serializer.data
        for compoff in compoffs:
            compoff_approvals_obj=CompoffApproval.objects.filter(claimed_compoff_id=compoff['id'])
            compoff_managers_serializer=compoff_approval_serializers(compoff_approvals_obj,many=True)
            compoff['managers']=compoff_managers_serializer.data
            compoff_approvals_rejected_obj=CompoffApproval.objects.filter(claimed_compoff_id=compoff['id'],reschedule=True).first()
            if compoff_approvals_rejected_obj is not None:
                compoff['reason']=compoff_approvals_rejected_obj.reschedule_reason
            else:
                compoff['reason']='Not Available'


        return Response ({"data":compoff_serializer.data,"response":{"n" : 1,"msg" : "reschedule comp-offs found ","status" : "success"}})
    else:
        return Response ({"data":[],"response":{"n" : 0,"msg" : "No reschedule comp-offs found ","status" : "error"}})



@api_view(['POST'])    
def approve_compoff_requests(request):
    user_id = request.user.id
    compoff_id=request.POST.get('id')
    company_code=request.user.company_code
    if compoff_id is not None and compoff_id !='':
        compoff_approval_obj=CompoffApproval.objects.filter(manager_id=user_id,claimed_compoff_id=compoff_id).first()
        if compoff_approval_obj is not None:
            compoff_approval_serializer=compoff_approval_serializers(compoff_approval_obj)
            compoff_obj = ClaimedCompOffMaster.objects.filter(id=compoff_approval_serializer.data['claimed_compoff_id'],is_active=True).first()
            if compoff_obj is not None:
                compoff_serializer=claimed_compoff_serializers(compoff_obj)
                if compoff_obj.status == 'Pending':
                    compoff_approval_obj.status=True
                    compoff_approval_obj.save() #approved the manager status
                    check_all_approvals=CompoffApproval.objects.filter(claimed_compoff_id=compoff_id,status=None).first()  
                    
                    if check_all_approvals is None: #check any pending approvals
                        compoff_obj.status='Approved'
                        compoff_obj.save()
                        notificationmsg = "Your Compoff request dated on <span class='notfappid'>" +compoff_serializer.data['claim_date']+"</span> has been <span class='rejectedmsg'> Approved </span>"
                        TaskNotification.objects.create(UserID_id=compoff_obj.user_id,company_code=company_code,NotificationTypeId_id=8,NotificationTitle="Comp-Off Approved",NotificationMsg=notificationmsg,leaveID=0)
                    
                    else:
                        
                        notificationmsg = "Your Compoff request dated on <span class='notfappid'>" +compoff_serializer.data['claim_date']+"</span> has been partially approved by <span class='rejectedmsg'>" +compoff_approval_serializer.data['manager_name'] +" </span>"
                        TaskNotification.objects.create(UserID_id=compoff_obj.user_id,company_code=company_code,NotificationTypeId_id=8,NotificationTitle="Comp-Off Approved",NotificationMsg=notificationmsg,leaveID=0)
                    
                    notfmsg = "You have approved compoff of <span class='actionuser'>" + compoff_serializer.data['user_name'] + "</span> dated <span class='notfleavedate'>" +compoff_serializer.data['claim_date']+" </span>"
                    TaskNotification.objects.create(UserID_id=int(request.user.id),To_manager=True,company_code=company_code,action_Taken=None,NotificationTypeId_id=8,NotificationTitle="Comp-Off",leaveID=0,NotificationMsg=notfmsg)


                    compoff_managers=CompoffApproval.objects.filter(claimed_compoff_id=compoff_id).exclude(manager_id=user_id).distinct('manager_id')  

                    managers_serializer=compoff_approval_serializers(compoff_managers,many=True)
                    for manager in managers_serializer.data:
                        manager_msg = compoff_approval_serializer.data['manager_name'] +" has approved compoff request of <span class='actionuser'>" + compoff_serializer.data['user_name'] + "</span> dated on <span class='notfleavedate'>" +compoff_serializer.data['claim_date']+" </span>"
                        TaskNotification.objects.create(UserID_id=manager['manager_id'],To_manager=True,company_code=company_code,action_Taken=None,NotificationTypeId_id=8,NotificationTitle="Comp-Off",leaveID=0,NotificationMsg=manager_msg)





                    return Response ({"data":[],"response":{"n" : 1,"msg" : "comp-off approved successfully","status" : "success"}})
                elif compoff_obj.status == 'Rejected':
                    compoff_rejected_obj=CompoffApproval.objects.filter(claimed_compoff_id=compoff_id,status=False).first()
                    if compoff_rejected_obj is not None:
                        rejected_compoff_serializer=compoff_approval_serializers(compoff_rejected_obj)
                        mesg='by '+ rejected_compoff_serializer.data['manager_name']
                    else:
                        mesg=''
                    return Response ({"data":[],"response":{"n" : 0,"msg" : "comp-off is already rejected "+mesg,"status" : "error"}})
                elif compoff_obj.status == 'Reschedule':
                    compoff_reschedule_obj=CompoffApproval.objects.filter(claimed_compoff_id=compoff_id,reschedule=True).first()
                    if compoff_reschedule_obj is not None:
                        reschedule_compoff_serializer=compoff_approval_serializers(compoff_reschedule_obj)
                        mesg='by '+ reschedule_compoff_serializer.data['manager_name']
                    else:
                        mesg=''
                    return Response ({"data":[],"response":{"n" : 0,"msg" : "comp-off is already reschedule "+mesg,"status" : "error"}})
                else:
                    compoff_approval_obj.status=True
                    compoff_approval_obj.save() #approved the manager status
                    return Response ({"data":[],"response":{"n" : 0,"msg" : "comp-off is already approved ","status" : "error"}})
            else:
                return Response ({"data":[],"response":{"n" : 0,"msg" : "No comp-off found ","status" : "error"}})
        else:
            return Response ({"data":[],"response":{"n" : 0,"msg" : "Sorry you dont have access to approve this compoff","status" : "error"}})
    else:
        return Response ({"data":[],"response":{"n" : 0,"msg" : "Please provide comp-offs id ","status" : "error"}})

@api_view(['POST'])    
def reject_compoff_requests(request):
    user_id = request.user.id
    company_code = request.user.company_code

    compoff_id=request.POST.get('id')
    if compoff_id is not None and compoff_id !='':
        compoff_approval_obj=CompoffApproval.objects.filter(manager_id=user_id,claimed_compoff_id=compoff_id).first()
        if compoff_approval_obj is not None:
            compoff_approval_serializer=compoff_approval_serializers(compoff_approval_obj)
            compoff_obj = ClaimedCompOffMaster.objects.filter(id=compoff_approval_obj.claimed_compoff_id,is_active=True).first()
            if compoff_obj is not None:
                compoff_serializer=claimed_compoff_serializers(compoff_obj)
                if compoff_obj.status == 'Pending':
                    reason=request.POST.get('reason')
                    if reason is not None and reason !='':
                        compoff_approval_obj.status=False
                        compoff_approval_obj.rejected_reason=reason
                        compoff_approval_obj.save() #reject the manager status
                        compoff_obj.status='Rejected'
                        compoff_obj.save()



                        notificationmsg = "Your Compoff request dated on <span class='notfappid'>" +compoff_serializer.data['claim_date']+"</span> has been rejected by <span class='rejectedmsg'>" +compoff_approval_serializer.data['manager_name'] +" </span>"
                        TaskNotification.objects.create(UserID_id=compoff_obj.user_id,company_code=company_code,NotificationTypeId_id=8,NotificationTitle="Comp-Off Rejected",NotificationMsg=notificationmsg,leaveID=0)
                    
                        notfmsg = "You have rejected compoff of <span class='actionuser'>" + compoff_serializer.data['user_name'] + "</span> dated <span class='notfleavedate'>" +compoff_serializer.data['claim_date']+" </span>"
                        TaskNotification.objects.create(UserID_id=int(request.user.id),To_manager=True,company_code=company_code,action_Taken=None,NotificationTypeId_id=8,NotificationTitle="Comp-Off",leaveID=0,NotificationMsg=notfmsg)


                        compoff_managers=CompoffApproval.objects.filter(claimed_compoff_id=compoff_id).exclude(manager_id=user_id).distinct('manager_id')  
                        
                        managers_serializer=compoff_approval_serializers(compoff_managers,many=True)
                        for manager in managers_serializer.data:
                            manager_msg = compoff_approval_serializer.data['manager_name'] +" has rejected compoff request of <span class='actionuser'>" + compoff_serializer.data['user_name'] + "</span> dated on <span class='notfleavedate'>" +compoff_serializer.data['claim_date']+" </span>"
                            TaskNotification.objects.create(UserID_id=manager['manager_id'],To_manager=True,company_code=company_code,action_Taken=None,NotificationTypeId_id=8,NotificationTitle="Comp-Off",leaveID=0,NotificationMsg=manager_msg)


















                        return Response ({"data":[],"response":{"n" : 1,"msg" : "comp-off rejected successfully","status" : "success"}})
                    else:
                        return Response ({"data":[],"response":{"n" : 0,"msg" : "Please provide rejection reason","status" : "error"}})
                elif compoff_obj.status =='Rejected':
                    compoff_rejected_obj=CompoffApproval.objects.filter(claimed_compoff_id=compoff_id,status=False).first()
                    if compoff_rejected_obj is not None:
                        rejected_compoff_serializer=compoff_approval_serializers(compoff_rejected_obj)
                        mesg='by '+ rejected_compoff_serializer.data['manager_name']
                    else:
                        mesg=''
                    return Response ({"data":[],"response":{"n" : 0,"msg" : "comp-off is already rejected "+mesg,"status" : "error"}})
                elif compoff_obj.status == 'Reschedule':
                    compoff_reschedule_obj=CompoffApproval.objects.filter(claimed_compoff_id=compoff_id,reschedule=True).first()
                    if compoff_reschedule_obj is not None:
                        reschedule_compoff_serializer=compoff_approval_serializers(compoff_reschedule_obj)
                        mesg='by '+ reschedule_compoff_serializer.data['manager_name']
                    else:
                        mesg=''
                    return Response ({"data":[],"response":{"n" : 0,"msg" : "comp-off is already reschedule "+mesg,"status" : "error"}})
                
                else:
                    compoff_approval_obj.status=False
                    compoff_approval_obj.save() #approved the manager status
                    return Response ({"data":[],"response":{"n" : 0,"msg" : "comp-off is already rejected by you ","status" : "error"}})
            else:
                return Response ({"data":[],"response":{"n" : 0,"msg" : "No comp-off found ","status" : "error"}})
        else:
            return Response ({"data":[],"response":{"n" : 0,"msg" : "Sorry you dont have access to reject this compoff","status" : "error"}})
    else:
        return Response ({"data":[],"response":{"n" : 0,"msg" : "Please provide comp-offs id ","status" : "error"}})

@api_view(['POST'])    
def reschedule_compoff_requests(request):
    user_id = request.user.id
    compoff_id=request.POST.get('id')
    company_code=request.user.company_code
    if compoff_id is not None and compoff_id !='':
        compoff_approval_obj=CompoffApproval.objects.filter(manager_id=user_id,claimed_compoff_id=compoff_id).first()
        if compoff_approval_obj is not None:
            compoff_approval_serializer=compoff_approval_serializers(compoff_approval_obj)
            compoff_obj = ClaimedCompOffMaster.objects.filter(id=compoff_approval_obj.claimed_compoff_id,is_active=True).first()
            if compoff_obj is not None:
                compoff_serializer=claimed_compoff_serializers(compoff_obj)
                if compoff_obj.status == 'Pending':
                    reason=request.POST.get('reason')
                    if reason is not None and reason !='':
                        compoff_approval_obj.reschedule=True
                        compoff_approval_obj.reschedule_reason=reason
                        compoff_approval_obj.save() #reschedule the manager status
                        compoff_obj.status='Reschedule'
                        compoff_obj.save()
                        EligibleCompOffMaster.objects.filter(id=compoff_obj.eligible_compoff_id,is_active=True).update(claim=False,reschedule=True)





                        notificationmsg = "Your Compoff request dated on <span class='notfappid'>" +compoff_serializer.data['claim_date']+"</span> has been reschedule by <span class='rejectedmsg'>" +compoff_approval_serializer.data['manager_name'] +" </span>"
                        TaskNotification.objects.create(UserID_id=compoff_obj.user_id,company_code=company_code,NotificationTypeId_id=8,NotificationTitle="Comp-Off Reschedule",NotificationMsg=notificationmsg,leaveID=0)
                    
                        notfmsg = "You have reschedule compoff of <span class='actionuser'>" + compoff_serializer.data['user_name'] + "</span> dated <span class='notfleavedate'>" +compoff_serializer.data['claim_date']+" </span>"
                        TaskNotification.objects.create(UserID_id=int(request.user.id),To_manager=True,company_code=company_code,action_Taken=None,NotificationTypeId_id=8,NotificationTitle="Comp-Off",leaveID=0,NotificationMsg=notfmsg)


                        compoff_managers=CompoffApproval.objects.filter(claimed_compoff_id=compoff_id).exclude(manager_id=user_id).distinct('manager_id')  
                        
                        managers_serializer=compoff_approval_serializers(compoff_managers,many=True)
                        for manager in managers_serializer.data:
                            manager_msg = compoff_approval_serializer.data['manager_name'] +" has reschedule compoff request of <span class='actionuser'>" + compoff_serializer.data['user_name'] + "</span> dated on <span class='notfleavedate'>" +compoff_serializer.data['claim_date']+" </span>"
                            TaskNotification.objects.create(UserID_id=manager['manager_id'],To_manager=True,company_code=company_code,action_Taken=None,NotificationTypeId_id=8,NotificationTitle="Comp-Off",leaveID=0,NotificationMsg=manager_msg)









                        return Response ({"data":[],"response":{"n" : 1,"msg" : "Record updated successfully","status" : "success"}})
                    else:
                        return Response ({"data":[],"response":{"n" : 0,"msg" : "Please provide reschedule reason","status" : "error"}})
                elif compoff_obj.status =='Rejected':
                    compoff_rejected_obj=CompoffApproval.objects.filter(claimed_compoff_id=compoff_id,status=False).first()
                    if compoff_rejected_obj is not None:
                        rejected_compoff_serializer=compoff_approval_serializers(compoff_rejected_obj)
                        mesg='by '+ rejected_compoff_serializer.data['manager_name']
                    else:
                        mesg=''
                    return Response ({"data":[],"response":{"n" : 0,"msg" : "comp-off is already rejected "+mesg,"status" : "error"}})
                elif compoff_obj.status == 'Reschedule':
                    compoff_reschedule_obj=CompoffApproval.objects.filter(claimed_compoff_id=compoff_id,reschedule=True).first()
                    if compoff_reschedule_obj is not None:
                        reschedule_compoff_serializer=compoff_approval_serializers(compoff_reschedule_obj)
                        mesg='by '+ reschedule_compoff_serializer.data['manager_name']
                    else:
                        mesg=''
                    return Response ({"data":[],"response":{"n" : 0,"msg" : "comp-off is already reschedule "+mesg,"status" : "error"}})
                else:
                    return Response ({"data":[],"response":{"n" : 0,"msg" : "comp-off journy is completed unable to perform actions","status" : "error"}})
            else:
                return Response ({"data":[],"response":{"n" : 0,"msg" : "No comp-off found ","status" : "error"}})
        else:
            return Response ({"data":[],"response":{"n" : 0,"msg" : "Sorry you dont have access to reschedule this compoff","status" : "error"}})
    else:
        return Response ({"data":[],"response":{"n" : 0,"msg" : "Please provide comp-offs id ","status" : "error"}})

@api_view(['POST'])    
def claim_compoff(request):
    user_id = request.user.id
    company_code=request.user.company_code
    current_date = datetime.datetime.now()
    formatted_date = current_date.strftime('%Y-%m-%d')
    compoff_obj = EligibleCompOffMaster.objects.filter(user_id=user_id,id=request.POST.get('id'),claim=False,is_active=True,valid_date__gte=formatted_date).first()
    data={}
    if compoff_obj is not None:
        claim_date=request.POST.get('claim_date')
        if claim_date is not None and claim_date !='':
            claim_date_str = datetime.datetime.strptime(claim_date, '%Y-%m-%d').date()
            today = datetime.datetime.today().date()
            if claim_date_str <= today:
                return Response ({"data":[],"response":{"n" : 0,"msg" : "Claim date must be a future date.","status" : "warning"}})

            check_holiday=Holidays.objects.filter(Date=claim_date,company_code=company_code,Active=True).first()
            if check_holiday is None:
                if is_weeklyoff(claim_date) != None:
                    return Response ({"data":[],"response":{"n" : 0,"msg" : "Applied compoff is on weekly off","status" : "warning"}})
                check_leave_exist=Leave.objects.filter(employeeId=user_id,Active=True,WorkFromHome=False,leave_status="Approved",start_date__lte=claim_date,end_date__gte=claim_date).exclude(leave_status='Draft').first()
                if check_leave_exist is not None:
                    return Response ({"data":[],"response":{"n" : 0,"msg" : "You already have approved leave on this date","status" : "warning"}})
                data['claim_date']= datetime.datetime.strptime(claim_date + ' ' + '05:30:00', '%Y-%m-%d %H:%M:%S')

                check_existing_compoff=ClaimedCompOffMaster.objects.filter(user_id=user_id,claim_date=data['claim_date'],is_active=True).exclude(status="Withdrawn").first()
                if check_existing_compoff is not None:
                    return Response ({"data":[],"response":{"n" : 0,"msg" : "You already have claimed compoff on this date","status" : "warning"}})



                eligible_compoff_serializer = save_compoff_serializers(compoff_obj)

                compoff_obj.claim=True
                compoff_obj.save()

                data['user_id']=eligible_compoff_serializer.data['user_id']
                data['user_name']=eligible_compoff_serializer.data['user_name']
                data['date']=eligible_compoff_serializer.data['date']
                data['valid_date']=eligible_compoff_serializer.data['valid_date']
                data['working_hrs']=eligible_compoff_serializer.data['working_hrs']
                data['eligible_compoff_id']=eligible_compoff_serializer.data['id']
                data['shift_name']=eligible_compoff_serializer.data['shift_name']
                if eligible_compoff_serializer.data['reschedule']:
                    earlier_claimed_compoff_obj=ClaimedCompOffMaster.objects.filter(eligible_compoff_id=eligible_compoff_serializer.data['id'],status='Reschedule').last()
                    if earlier_claimed_compoff_obj is not None:
                        data['earlier_reschedule']=True
                        data['earlier_reschedule_id']=earlier_claimed_compoff_obj.id
                    


                claimed_compoff_serializer = save_claimed_compoff_serializers(data=data,partial=True)

                if claimed_compoff_serializer.is_valid():
                    claimed_compoff_serializer.save()
                    managers_obj=leaveMapping.objects.filter(employeeId=user_id).distinct("managerId")
                    managers_serializers=leaveMappingserializer(managers_obj,many=True)
                    notificationmsg = "<span class='notfempname'>" + claimed_compoff_serializer.data['user_name'] + "</span> applied for comp-off"
                    Team_members=[adminemail,hremail]

                    for manager in managers_serializers.data:
                        manager_obj = Users.objects.filter(id=int(manager['managerId']),is_active=True).first()
                        if manager_obj is not None:
                            Team_members.append(manager_obj.email)
                            TaskNotification.objects.create(UserID_id=manager['managerId'],To_manager=True,leaveID=0,company_code=company_code,NotificationTypeId_id=8,NotificationTitle="CompOff",NotificationMsg=notificationmsg)

                            CompoffApproval.objects.create(claimed_compoff_id=claimed_compoff_serializer.data['id'],eligible_compoff_id=eligible_compoff_serializer.data['id'],user_id=user_id,manager_id=manager['managerId'])

                    
                    # data_dict = {
                    #             "user_name":str(request.user.Firstname) +' '+str(request.user.Lastname),
                    #             "generated_compoff_date":convertdate(str(eligible_compoff_serializer.data['date'].split('T')[0])),
                    #             "claimed_compoff_date":convertdate(claim_date),
                    # }


                    # send_async_custom_template_email(
                    #    'Compensatory Off Request from ' +str(request.user.Firstname) +' '+str(request.user.Lastname),
                    #     data_dict,
                    #     EMAIL_HOST_USER,
                    #     Team_members,
                    #     "compoff/apply_compoff_mail.html"
                    # )

                    return Response ({"data":claimed_compoff_serializer.data,"response":{"n" : 1,"msg" : "Comp-off claimed successfully","status" : "success"}})
                else:
                    first_key, first_value = next(iter(claimed_compoff_serializer.errors.items()))
                    return Response ({"data":[],"response":{"n" : 0,"msg" : first_key[0] +' '+ first_value[0],"status" : "error"}})
            else:
                return Response ({"data":[],"response":{"n" : 0,"msg" : "Holiday exists on claimed compoff date","status" : "error"}})
        else:
            return Response ({"data":[],"response":{"n" : 0,"msg" : "Please provide claim date","status" : "error"}})

    else:
        return Response ({"data":[],"response":{"n" : 0,"msg" : "This compoff is no more available","status" : "error"}})

@api_view(['POST'])    
def change_claim_compoff_date(request):
    user_id = request.user.id
    company_code=request.user.company_code
    current_date = datetime.datetime.now()
    formatted_date = current_date.strftime('%Y-%m-%d')
    compoff_obj = ClaimedCompOffMaster.objects.filter(user_id=user_id,id=request.POST.get('id'),is_active=True,valid_date__gte=formatted_date).first()
    data={}
    if compoff_obj is not None:
        claim_date=request.POST.get('claim_date')
        if claim_date is not None and claim_date !='':
            check_holiday=Holidays.objects.filter(Date=claim_date,company_code=company_code,Active=True).first()
            if check_holiday is None:
                if is_weeklyoff(claim_date) != None:
                    return Response ({"data":[],"response":{"n" : 0,"msg" : "Applied compoff is on weekly off","status" : "warning"}})
                check_leave_exist=Leave.objects.filter(employeeId=user_id,Active=True,WorkFromHome=False,leave_status="Approved",start_date__lte=claim_date,end_date__gte=claim_date).exclude(leave_status='Draft').first()
                if check_leave_exist is not None:
                    return Response ({"data":[],"response":{"n" : 0,"msg" : "You already have approved leave on this date","status" : "warning"}})

                data['claim']=True
                data['claim_date']= datetime.datetime.strptime(claim_date + ' ' + '05:30:00', '%Y-%m-%d %H:%M:%S')
                
                compoff_serializer = save_claimed_compoff_serializers(compoff_obj,data=data,partial=True)
                if compoff_serializer.is_valid():
                    compoff_serializer.save()
                    managers_obj=CompoffApproval.objects.filter(claimed_compoff_id=compoff_serializer.data['id'])
                    managers_serializer=compoff_approval_serializers(managers_obj,many=True)
                    notificationmsg = "<span class='notfempname'>" + compoff_serializer.data['user_name'] + "</span> has changed their comp-off date"
                    Team_members=[adminemail,hremail]

                    for manager in managers_serializer.data:
                        TaskNotification.objects.create(UserID_id=manager['manager_id'],To_manager=True,leaveID=0,company_code=company_code,NotificationTypeId_id=8,NotificationTitle="CompOff",NotificationMsg=notificationmsg)
                        manager_obj = Users.objects.filter(id=int(manager['manager_id']),is_active=True).first()
                        if manager_obj is not None:
                            Team_members.append(manager_obj.email)
                    data_dict = {
                                "user_name":str(request.user.Firstname) +' '+str(request.user.Lastname),
                                "generated_compoff_date":convertdate(str(compoff_serializer.data['date'].split('T')[0])),
                                "claimed_compoff_date":convertdate(claim_date),
                    }


                    send_async_custom_template_email(
                       'Compensatory Off Date Change Request from ' +str(request.user.Firstname) +' '+str(request.user.Lastname),
                        data_dict,
                        EMAIL_HOST_USER,
                        Team_members,
                        "compoff/compoff_date_change_mail.html"
                    )
                    return Response ({"data":compoff_serializer.data,"response":{"n" : 1,"msg" : "Comp-off date changed successfully","status" : "success"}})
                else:
                    first_key, first_value = next(iter(compoff_serializer.errors.items()))
                    return Response ({"data":[],"response":{"n" : 0,"msg" : first_key[0] +' '+ first_value[0],"status" : "error"}})
            else:
                return Response ({"data":[],"response":{"n" : 0,"msg" : "Holiday exists on claimed compoff date","status" : "error"}})
        else:
            return Response ({"data":[],"response":{"n" : 0,"msg" : "Please provide claim date","status" : "error"}})

    else:
        return Response ({"data":[],"response":{"n" : 0,"msg" : "This compoff is no more available","status" : "error"}})
   
@api_view(['POST'])    
def withdraw_compoff(request):
    user_id = request.user.id
    company_code=request.user.company_code
    current_date = datetime.datetime.now()
    formatted_date = current_date.strftime('%Y-%m-%d')
    compoff_obj = ClaimedCompOffMaster.objects.filter(user_id=user_id,id=request.POST.get('id'),is_active=True).first()
    data={}
    if compoff_obj is not None:
        if compoff_obj.status =='Pending' or compoff_obj.status =='Approved':
            print("compoff_obj.claim_date0",str(compoff_obj.claim_date).split(' ')[0])
            claim_date_str = datetime.datetime.strptime(str(compoff_obj.claim_date).split(' ')[0], '%Y-%m-%d').date()
            today = datetime.datetime.today().date()
            if claim_date_str <= today:
                return Response ({"data":[],"response":{"n" : 0,"msg" : "Withdrawn time has passed","status" : "error"}})
            data['status']='Withdraw'

            claimed_compoff_serializer = save_claimed_compoff_serializers(compoff_obj,data=data,partial=True)
            if claimed_compoff_serializer.is_valid():
                claimed_compoff_serializer.save()
                eligible_compoff_obj=EligibleCompOffMaster.objects.filter(id=compoff_obj.eligible_compoff_id,is_active=True).first()
                if eligible_compoff_obj is not None:
                    eligible_compoff_obj.claim=False
                    eligible_compoff_obj.save() 

                managers_obj=CompoffApproval.objects.filter(claimed_compoff_id=claimed_compoff_serializer.data['id']).distinct("manager_id")
                managers_serializers=compoff_approval_serializers(managers_obj,many=True)
                notificationmsg = "<span class='notfempname'>" + claimed_compoff_serializer.data['user_name'] + "</span> has withdrawn their comp-off"
                Team_members=[adminemail,hremail]
                for manager in managers_serializers.data:
                    manager_obj = Users.objects.filter(id=int(manager['manager_id']),is_active=True).first()
                    if manager_obj is not None:
                        Team_members.append(manager_obj.email)
                        TaskNotification.objects.create(UserID_id=manager['manager_id'],To_manager=True,leaveID=0,company_code=company_code,NotificationTypeId_id=8,NotificationTitle="CompOff",NotificationMsg=notificationmsg)

                
                data_dict = {
                            "user_name":str(request.user.Firstname) +' '+str(request.user.Lastname),
                            "generated_compoff_date":convertdate(str(claimed_compoff_serializer.data['date'].split('T')[0])),
                            "claimed_compoff_date":convertdate(str(claimed_compoff_serializer.data['claim_date'].split('T')[0])),
                }


                send_async_custom_template_email(
                    str(request.user.Firstname) +' '+str(request.user.Lastname) +' has withdrawn their Compensatory Off Request',
                    data_dict,
                    EMAIL_HOST_USER,
                    Team_members,
                    "compoff/withdrawn_compoff_mail.html"
                )

                return Response ({"data":claimed_compoff_serializer.data,"response":{"n" : 1,"msg" : "Comp-off withdrawn successfully","status" : "success"}})
            else:
                first_key, first_value = next(iter(claimed_compoff_serializer.errors.items()))
                return Response ({"data":[],"response":{"n" : 0,"msg" : first_key[0] +' '+ first_value[0],"status" : "error"}})

        else:
            return Response ({"data":[],"response":{"n" : 0,"msg" : "This Compoff cannot be withdrawn","status" : "error"}})

    else:
        return Response ({"data":[],"response":{"n" : 0,"msg" : "This compoff is no more available","status" : "error"}})
  



