from rest_framework.decorators import authentication_classes, permission_classes
import json
from django.views.decorators.csrf import csrf_exempt
from Users.serializers import *
from CheckTrackPro.serializers import IntermediateGetTrackProResultSerializer
from CheckTrackPro.models import IntermediateTrackProResult
from Rules.models import Leaverule
from Rules.serializers import leaveruleserializer
from Users.serializers import RoleIdSerializer,UserSerializer, attendanceserializer, holidaysSerializer,UserSecondarySerializer,leaveserializer,leaveapprovalserializer,leaveMappingserializer,L1L2Serializer,UsersSerializeronlyid,UsersSerializeronlyattendance
from Users.serializers import MappingSerializer,EmployeeShiftDetailsSerializer,ShiftAllotmentSerializer,ShiftMasterSerializer,CompOffGrantedMasterSerializer
from .serializers import *
from Tasks.serializers import *
from Tasks.models import *
from .models import *
from TrackProBackend.settings import EMAIL_HOST_USER
from Users.models import Role, Users,UserToManager,UserSecondaryInfo,Leave,leaveApproval,leaveMapping,adminAttendanceRequest,EmployeeShiftDetails
from Users.views import create_google_maps_url,get_location_name
from django.http import JsonResponse
from rest_framework.decorators import api_view
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.status import (
    HTTP_400_BAD_REQUEST,
    HTTP_404_NOT_FOUND,
    HTTP_200_OK
)
from rest_framework import status
from rest_framework.decorators import api_view
# from CompanyMaster.views import imageUrl
from django.template.loader import  get_template
from django.core.mail import EmailMessage

import arrow
from dateutil.relativedelta import *
import xlsxwriter
import calendar
import environ
env = environ.Env()
environ.Env.read_env()
from collections import defaultdict
from openpyxl import load_workbook
from Tasks.views import sendfirebasenotification,senddesktopnotf
from Users.sendmail import send_async_custom_template_email
from Users.static_info import adminemail,hremail,frontUrl,imageUrl
from rest_framework import pagination
from rest_framework.pagination import PageNumberPagination
from rest_framework.generics import GenericAPIView
from rest_framework.response import Response
from Users.custom_functions import *
import os
from datetime import date,datetime,timedelta,time
from django.db.models import F, Sum, Value, FloatField, Case, When, Q



class CustomPagination(pagination.PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 50
    page_query_param = 'p'

    def get_paginated_response(self,data):
        response = Response({
            'n':1,
            'status':"success",
            'count':self.page.paginator.count,
            'next' : self.get_next_link(),
            'previous' : self.get_previous_link(),
            'data':data,
        })
        return response
   
@api_view(['POST'])    
def leaveapi(request):
    data={}
    empuserID=request.data.get('UserId')
    empid=Users.objects.filter(id=empuserID).first()
    if empid is not None:
        employee_id=empid.id
    
        unique_id = empid.uid
        todays_date = date.today()
        year=todays_date.year

        user_wfh_id = request.user.typeofwork

        data['employeeId']=employee_id
        data['start_date']=request.data.get('start_date')
        data['end_date']=request.data.get('end_date')
        data['leavetype'] = request.data.get('leavetype')
        data['reason'] = request.data.get('reason')
        data['leave_status'] = request.data.get('leave_status')
        data['company_code'] = request.user.company_code
        data['WorkFromHome'] = request.POST.get('WorkFromHome')
        if request.FILES.get('attachment') is not None and request.FILES.get('attachment') !='': 
            data['attachment'] = request.FILES.get('attachment') 
        if data['WorkFromHome'] =="True":
            data['WorkFromHome']=True
        else:
            data['WorkFromHome']=False        
        data['number_of_days']=calculate_days(data['start_date'],data['end_date'],data['leavetype'])




        

        
        
        if is_weeklyoff(data['start_date']) != None:
            return Response ({"data":[],"response":{"n" : 0,"msg" : "Start date should not on holidays","status" : "warning"}})
        elif checkholiday(data['start_date']) != None:
            return Response ({"data":[],"response":{"n" : 0,"msg" : "Start date should not on holidays","status" : "warning"}})


        if(data['start_date']==data['end_date']):
            weekoff_obj=is_weeklyoff(data['start_date'])
            if weekoff_obj !=None:
                return Response(weekoff_obj)
            if checkholiday(data['start_date']) != None:
                return Response ({"data":[],"response":{"n" : 0,"msg" : "Applied leaves dates are on holidays","status" : "warning"}})

        else:
            checkvaliddate=[]
            daterangelist=date_range_list(data['start_date'],data['end_date'])
            for i in daterangelist:
                if is_weeklyoff(str(i)) != None:
                    checkvaliddate.append(False)
                elif checkholiday(i) != None:
                    checkvaliddate.append(False)
                else: 
                    checkvaliddate.append(True)
            if sum(checkvaliddate) <1:
                return Response ({"data":[],"response":{"n" : 0,"msg" : "Applied leaves dates are on holidays","status" : "warning"}})







        def check_leaves(sdate,edate,wfh):
            applied_leaves_dates=date_range_list(sdate,edate)
            existing_leaves_dates=[]
            
            if wfh==True:
                recordexist = Leave.objects.filter(employeeId=employee_id,Active=True)
            else:
                recordexist = Leave.objects.filter(employeeId=employee_id,Active=True,WorkFromHome=False)

            recordserializer=leaveserializer(recordexist,many=True)
            for i in recordserializer.data:
                if i['leave_status'] !="Draft" and i['leave_status'] !="Withdraw" :
                    exist_date_obj=date_range_list(i['start_date'],i['end_date']) 
                    for j in exist_date_obj:
                        existing_leaves_dates.append(j)

            set1 = set(applied_leaves_dates)
            set2 = set(existing_leaves_dates)
            common_dates = set1.intersection(set2)
            for common_date in common_dates:
                if common_date is not None:
                    return ({"data":[],"response":{"n" : 0,"msg" : "Request of date "+dateformat_ddmmyy(str(common_date))+" Already exist","status" : "warning"}})

        output_check=check_leaves(data['start_date'],data['end_date'],data['WorkFromHome'])
        if output_check != None:
            return Response (output_check)
        
        mappingexist = leaveMapping.objects.filter(employeeId= data['employeeId'],company_code= data['company_code']).first()
        if mappingexist is None:
            return Response ({"data":[],"response":{"n" : 0,"msg" : "You don't have mapped managers to inform please connect with respective managers","status" : "warning"}})
        
        if user_wfh_id == "1" and data['WorkFromHome'] == True:
            return Response ({"data":[],"response":{"n" : 0,"msg" : "You can't apply for Work for home","status" : "warning"}})
        

        if data['leave_status'] == "Normal":
            data['leave_status'] ="Pending"

            leavedata = Leave.objects.filter(employeeId=employee_id).order_by('-id').exclude(ApplicationId=None)
            leaves_obj_serializer=leaveserializer(leavedata,many=True)
                
            appnumberlist=[]
            for y in leaves_obj_serializer.data:
                number_obj = y['ApplicationId'].split("/")[3]
                appnumberlist.append(int(number_obj))
            if len(appnumberlist) < 1:
                Idnumber = "1"
                application_ID = str(year)+"/"+unique_id+"/"+Idnumber
                data['ApplicationId']=application_ID
            else:
                maxnumber=int(max(appnumberlist))

                maxnumber=maxnumber+1

                Idnumber = str(maxnumber)
                application_ID =str(year)+"/"+unique_id+"/"+Idnumber
                data['ApplicationId']=application_ID

        lserializer=leaveserializer(data=data)
        if lserializer.is_valid():
            lserializer.save()
            newleavelist = []
            for l in [lserializer.data]:
                lstdate = str(l['start_date'])
                lsmonth_name = calendar.month_abbr[int(lstdate.split('-')[1])]    
                lsdatestr = lstdate.split('-')[2]+" "+lsmonth_name
                l['strstartdate'] = lsdatestr
                

                ledate = str(l['end_date'])
                lemonth_name = calendar.month_abbr[int(ledate.split('-')[1])]    
                ledatestr = ledate.split('-')[2]+" "+lemonth_name
                l['strenddate'] = ledatestr
                newleavelist.append(l)

            leaveid = lserializer.data['id']
            Idapplication =lserializer.data['ApplicationId']
            leavestatus =lserializer.data['leave_status']
            workfromhome=lserializer.data['WorkFromHome']
            if workfromhome==True:
                object_string="Work From Home"
            else:
                object_string="Leave"

            managerobj = leaveMapping.objects.filter(employeeId=empuserID,company_code=lserializer.data['company_code'])
            managerserializer=leaveMappingserializer(managerobj,many=True) 
            managerlist=[]
            for m in managerserializer.data:            
                if m['managerId'] not in managerlist:
                    managerlist.append(m['managerId'])

            empname = Users.objects.filter(id=empuserID,is_active=True).first()
            userempname = empname.Firstname +" " + empname.Lastname

            if workfromhome == True:
                notificationmsg = "<span class='notfempname'>" + userempname + "</span> Applied for Work From Home"
            else:
                notificationmsg = "<span class='notfempname'>" + userempname + "</span> Applied for leave"
                

            if workfromhome == True:
                firebasenotfmsg =   userempname + " Applied for Work From Home"
            else:
                firebasenotfmsg =  userempname + " Applied for leave"

            if leavestatus == "Pending":
                for manager_obj in managerlist:
                    leaveApproval.objects.create(employeeId=empuserID,leave_id=leaveid,managerId=manager_obj,ApplicationId=Idapplication,company_code=data['company_code'])

                    TaskNotification.objects.create(UserID_id=manager_obj,To_manager=True,leaveID=leaveid,company_code=data['company_code'],NotificationTypeId_id=3,NotificationTitle="Leave",NotificationMsg=notificationmsg)

                    manageridobj = Users.objects.filter(id=manager_obj).first()
                    if manageridobj is not None:
                        firebasemsg =  firebasenotfmsg
                        fcmtoken = manageridobj.FirebaseID
                        notftype = "Leave"
                        fcmleaveid = leaveid
                        fcmtomanager = True
                        
                        desktoptoken = manageridobj.desktopToken 
                        # desktopnotf = senddesktopnotf(desktoptoken,firebasemsg)
                        
                        if fcmtoken is not None and fcmtoken != "":
                            firebasenotification = ""
                            # firebasenotification = sendfirebasenotification(fcmtoken,firebasemsg,notftype,fcmleaveid,fcmtomanager)
                        else:
                            firebasenotification = ""

                ad_mail = adminemail
                adminobj = Users.objects.filter(email=ad_mail).first()
                if adminobj is not None:
                    adminid = adminobj.id
                else:
                    adminid = ''

                TaskNotification.objects.create(UserID_id=int(adminid),leaveID=leaveid,company_code=data['company_code'],NotificationTypeId_id=3,NotificationTitle="Leave",NotificationMsg=notificationmsg)
            
                userempname=userempname.title()
        
                Team_members=[adminemail,hremail]
                leave_approval_object=leaveApproval.objects.filter(employeeId=data['employeeId'],leave_id=leaveid,company_code=data['company_code'])
                leave_approval_serializer = leaveapprovalserializer(leave_approval_object,many=True)
                for manager in leave_approval_serializer.data:
                    m_obj = Users.objects.filter(id=int(manager['managerId']),is_active=True).first()
                    if m_obj is not None:
                        Team_members.append(m_obj.email)

                try:             
                    dicti = {
                                "employeename":userempname,
                                "dates":date_handling(lserializer.data['start_date'],lserializer.data['end_date']),
                                "startdate":convertdate(lserializer.data['start_date']),
                                "enddate":convertdate(lserializer.data['end_date']),
                                "reason":lserializer.data['reason'],
                                "type":wfhsorter(lserializer.data['id'])
                            }
                    message = get_template(
                        'leaveremindermail.html').render(dicti)
                    
                    msg = EmailMessage(
                        str(userempname) +' has applied for '+ str(WorkFromHome(lserializer.data['id'])),
                        message,
                        EMAIL_HOST_USER,
                        Team_members,
                    )
                    msg.content_subtype = "html"  # Main content is now text/html
                    msg.send()
                except Exception as e:
                    print('exception occured fot mail', e)

            
            
            return Response ({"data":newleavelist,"response":{"n" : 1,"msg" :str(object_string)+" Request Submitted Successfully.","status" : "success"}})
        else: 
            return Response ({"data":lserializer.errors,"response":{"n" : 0,"msg" :"Failed","status" : "warning"}})
    return Response ({"data":{},"response":{"n" : 0,"msg" :"user id is invalid","status" : "warning"}})




@api_view(['POST'])    
def apply_leave_api(request):
    data={}
    UserId=request.user.id
    user_obj=Users.objects.filter(id=UserId,is_active=True).first()
    if user_obj is not None:
        userid=user_obj.id
        print("rr",request.data)
        EmployeeId=userid
        company_code=request.user.company_code
        data['leavetype']='Fullday'
        data['employeeId']=EmployeeId
        data['company_code'] = company_code

        AttachmentStatus=False
        if request.FILES.get('attachment') is not None and request.FILES.get('attachment') !='': 
            data['attachment'] = request.FILES.get('attachment') 
            AttachmentStatus=True


        StartDate=request.data.get('start_date')
        EndDate=request.data.get('end_date')
        StartDayLeaveType=request.data.get('StartDayLeaveType')
        EndDayLeaveType=request.data.get('EndDayLeaveType')
        LeaveTypeId=request.data.get('LeaveTypeId')
        data['number_of_days']=calculate_total_leave_days(StartDate, StartDayLeaveType, EndDate, EndDayLeaveType)
        data['leave_status']=request.data.get('leavestatus')
        data['start_date']=StartDate
        data['end_date']=EndDate
        data['startdayleavetype']=StartDayLeaveType
        data['enddayleavetype']=EndDayLeaveType
        data['LeaveTypeId']=LeaveTypeId
        data['reason']=request.data.get('reason')


        validate_leave_rule=validate_leave_rules(EmployeeId,company_code,LeaveTypeId,StartDate,EndDate,StartDayLeaveType,EndDayLeaveType,AttachmentStatus)
        if validate_leave_rule['n']==1:
            ApplicationId=generate_leave_applicationid(EmployeeId)
            if ApplicationId is not None and ApplicationId!='':
                data['ApplicationId']=ApplicationId
            else:
                return Response ({"data":{},"response":{"n" : 0,"msg" :"user is invalid","status" : "error"}})

            print("data",data)  

            # return Response ({"data":{},"response":{"n" : 0,"msg" :"user is invalid","status" : "error"}})

            # return Response ({"data":[],"response":validate_leave_rule})
            managers_objs=leaveMapping.objects.filter(employeeId=EmployeeId,company_code=company_code)
            if managers_objs.exists():
                manager_serializer=leaveMappingserializer(managers_objs,many=True)
                serializer=leaveserializer(data=data)
                if serializer.is_valid():
                    serializer.save()

                    userempname = user_obj.Firstname +" " + user_obj.Lastname
                    userempname=userempname.title()

                    notificationmsg = "<span class='notfempname'>" + userempname + "</span> Applied for leave"
                    firebasenotfmsg =  userempname + " Applied for leave"

                



                    if serializer.data['leave_status'] == "Pending":
                        Team_members=[adminemail,hremail]

                        for manager in manager_serializer.data:
                            leaveApproval.objects.create(employeeId=EmployeeId,leave_id=serializer.data['id'],managerId=manager['managerId'],ApplicationId=serializer.data['ApplicationId'],company_code=company_code)
                            TaskNotification.objects.create(UserID_id=manager['managerId'],To_manager=True,leaveID=serializer.data['id'],company_code=company_code,NotificationTypeId_id=3,NotificationTitle="Leave",NotificationMsg=notificationmsg)

                            manageridobj = Users.objects.filter(id=manager['managerId']).first()
                            if manageridobj is not None:
                                Team_members.append(manageridobj.email)
                            #     fcmtoken = manageridobj.FirebaseID
                            #     desktopnotf = senddesktopnotf(manageridobj.desktopToken,firebasenotfmsg)
                            #     firebasenotification = sendfirebasenotification(fcmtoken,firebasenotfmsg,"Leave",serializer.data['id'],True)

                        if adminemail is not None and adminemail !='':
                            adminobj = Users.objects.filter(email=adminemail).first()
                            if adminobj is not None:
                                TaskNotification.objects.create(UserID_id=int(adminobj.id),leaveID=serializer.data['id'],company_code=company_code,NotificationTypeId_id=3,NotificationTitle="Leave",NotificationMsg=notificationmsg)

                        try:             
                            dicti = {
                                        "employeename":userempname,
                                        "dates":date_handling(serializer.data['start_date'],serializer.data['end_date']),
                                        "startdate":convertdate(serializer.data['start_date']),
                                        "enddate":convertdate(serializer.data['end_date']),
                                        "reason":serializer.data['reason'],
                                        "type":'Leave'
                                    }
                            message = get_template(
                                'leaveremindermail.html').render(dicti)
                            
                            msg = EmailMessage(
                                str(userempname) +' has applied for leave',
                                message,
                                EMAIL_HOST_USER,
                                Team_members,
                            )
                            msg.content_subtype = "html"  # Main content is now text/html
                            msg.send()
                        except Exception as e:
                            print('exception occured fot mail', e)
                        return Response ({"data":[],"response":{"n" : 1,"msg" :"Leave Request Submited Successfully.","status" : "success"}})
                    else:
                        return Response ({"data":[],"response":{"n" : 1,"msg" :"Leave Request Saved Successfully.","status" : "success"}})

                else:
                    first_key, first_value = next(iter(serializer.errors.items()))
                    return Response ({"data":[],"response":{"n" : 0,"msg" : first_key[0] +' '+ first_value[0],"status" : "error"}})
            else: 
                return Response ({"data":[],"response":{"n" : 0,"msg" :"You don't have a mapped manager. Please contact the admin.","status" : "error"}})
        else:
            return Response ({"data":[],"response":validate_leave_rule})
    else:

        return Response ({"data":{},"response":{"n" : 0,"msg" :"user id is invalid","status" : "error"}})

@api_view(['POST'])    
def get_employee_leave_balance_api(request):
    EmployeeId=request.user.id
    company_code=request.user.company_code
    LeaveTypeId=request.data.get('LeaveTypeId')
    Year=date.today().year
    PacketLeaveCount=0


    employee_obj = Users.objects.filter(id=EmployeeId, company_code=company_code, is_active=True).first()
    if not employee_obj:
        return Response({'n': 0, 'msg': 'User not found.', 'status': 'error', 'data': {}})


    packet_mapping_obj=PacketEmployeeMapping.objects.filter(EmployeeId=EmployeeId,company_code=company_code,is_active=True).first()
    if packet_mapping_obj is None:
        return Response({'n': 0, 'msg': 'The employee is not assigned to any packet.', 'status': 'error', 'data': {}})

    packet_obj=PacketMaster.objects.filter(id=packet_mapping_obj.PacketId,is_active=True,company_code=company_code).first()
    if packet_obj is None:
        return Response({'n': 0, 'msg': 'The assigned packet for the employee is no longer available.', 'status': 'error', 'data': {}})

    PacketId=packet_obj.id



    leave_type_obj = LeaveTypeMaster.objects.filter(id=LeaveTypeId, company_code=company_code, is_active=True).first()
    if not leave_type_obj:
        return Response({'n': 0, 'msg': 'The selected leave type is no longer available.', 'status': 'error', 'data': {}})

    LeaveTypeId = leave_type_obj.id

    packet_rules_obj=PacketLeaveRules.objects.filter(company_code=company_code,PacketId=PacketId,LeaveTypeId=LeaveTypeId).first()
    if packet_rules_obj is not None:
        PacketLeaveCount=int(packet_rules_obj.LeaveCount)


    carry_forwarded_leave_counts_obj = CarryForwardedLeave.objects.filter(
        Year=Year, EmployeeId=EmployeeId, LeaveTypeId=LeaveTypeId, company_code=company_code, is_active=True
    ).first()
    
    carry_forwarded_leave_counts = carry_forwarded_leave_counts_obj.LeaveCount if carry_forwarded_leave_counts_obj else 0

    
    alloted_packet_leave_for_this_year = round(float(PacketLeaveCount) + float(carry_forwarded_leave_counts), 2)

     
    leaves = Leave.objects.filter(
        Q(employeeId=EmployeeId),
        Q(company_code=company_code),
        Q(LeaveTypeId=LeaveTypeId),  # Ensure leave type matches
        Q(start_date__year=Year) | 
        Q(end_date__year=Year) | 
        (Q(start_date__lte=date(Year, 12, 31)) & Q(end_date__gte=date(Year, 1, 1))),
        Active=True
    )

     
    leave_days_count = leaves.annotate(
        leave_day_count=Case(
            When(startdayleavetype="FullDay", enddayleavetype="FullDay", then=F('number_of_days')),
            When(startdayleavetype="FirstHalf", enddayleavetype="FirstHalf", then=F('number_of_days') - Value(0.5)),
            When(startdayleavetype="SecondHalf", enddayleavetype="SecondHalf", then=F('number_of_days') - Value(0.5)),
            When(startdayleavetype="FirstHalf", enddayleavetype="SecondHalf", then=F('number_of_days') - Value(1.0)),
            When(startdayleavetype="SecondHalf", enddayleavetype="FirstHalf", then=F('number_of_days') - Value(1.0)),
            When(startdayleavetype="FullDay", enddayleavetype="FirstHalf", then=F('number_of_days') - Value(0.5)),
            When(startdayleavetype="FullDay", enddayleavetype="SecondHalf", then=F('number_of_days') - Value(0.5)),
            default=F('number_of_days'),
            output_field=FloatField()
        )
    ).aggregate(total_leave_days=Sum('leave_day_count'))

     
    leave_taken_by_employee = leave_days_count['total_leave_days'] or 0

     
    leave_balance = round(alloted_packet_leave_for_this_year - leave_taken_by_employee, 2)
    show_leave_balance = max(leave_balance, 0)   

    data = {
        'show_leave_balance': show_leave_balance,
        'leave_balance': leave_balance,
        'leave_taken_by_employee': leave_taken_by_employee,
        'carry_forwarded_leave': carry_forwarded_leave_counts,
        'alloted_packet_leave_for_this_year': alloted_packet_leave_for_this_year,
    }

    return Response({'n': 1, 'msg': 'Leave balance found successfully.', 'status': 'success', 'data': data})

@api_view(['POST'])    
def calculate_total_leave_days_api(request):
    print("startdate",request.data)
    
    startdate=request.data.get('start_date')
    enddate=request.data.get('end_date')
    enddayleavetype=request.data.get('enddayleavetype')
    startdayleavetype=request.data.get('startdayleavetype')
    if startdate is not None and startdate !='':
        if enddate is not None and enddate !='':
            if enddayleavetype is not None and enddayleavetype !='':
                if startdayleavetype is not None and startdayleavetype !='':

                    start_date = datetime.strptime(startdate, '%Y-%m-%d').date()
                    end_date = datetime.strptime(enddate, '%Y-%m-%d').date()

                    
                    total_days = (end_date - start_date).days + 1   

                    
                    if startdayleavetype in ["FirstHalf", "SecondHalf"]:
                        total_days -= 0.5   

                    if enddayleavetype in ["FirstHalf", "SecondHalf"]:
                        total_days -= 0.5   
                    return Response({'n': 1, 'msg': 'Leave total days calculated successfully.', 'status': 'success', 'data': {'total_days':total_days}})

                else:
                    return Response({'n': 0, 'msg': 'Leave total days not calculated.4', 'status': 'error', 'data': {'total_days':0}})

            else:
                return Response({'n': 0, 'msg': 'Leave total days not calculated.3', 'status': 'error', 'data': {'total_days':0}})
        else:
            return Response({'n': 0, 'msg': 'Leave total days not calculated.2', 'status': 'error', 'data': {'total_days':0}})

    else:
        return Response({'n': 0, 'msg': 'Leave total days not calculated.1', 'status': 'error', 'data': {'total_days':0}})











@api_view(['GET'])
def todays_applications_list(request):

    logined_user=request.user.id
    company_code=request.user.company_code
    if request.method == 'GET':

        
        role_obj = Role.objects.filter(Q(RoleName__iexact="admin") | Q(RoleName__iexact="core team"),company_code=company_code,Active=True)
        role_serializer=RoleIdSerializer(role_obj,many=True)
        distinct_ids = {item['id'] for item in role_serializer.data}

        coreteam1=Users.objects.filter(RoleID__in=distinct_ids)

        mapped_managers_obj=leaveMapping.objects.filter(company_code=company_code).distinct("managerId")
        managers_serializer=leaveMappingserializer(mapped_managers_obj,many=True)
        distinct_mids = {item['managerId'] for item in managers_serializer.data}
        coreteam2=Users.objects.filter(id__in=distinct_mids)

        combined_coreteam = coreteam1 | coreteam2
        logined_user_obj=combined_coreteam.filter(id=logined_user).first()



        leavelist=[]
        superlist=[]
        exist_checkinglist=[]
        searchdate = datetime.now().strftime("%Y-%m-%d")

        application_obj= Leave.objects.filter(Active=True,company_code=company_code,start_date__lte=searchdate,end_date__gte=searchdate).exclude(Q(leave_status="Rejected")|Q(leave_status="Withdraw")|Q(leave_status="Draft")).order_by('start_date','leavetype')
        serializer=leaveserializer(application_obj,many=True)





        for i in serializer.data:
            
            leave_appprovel_obj=leaveApproval.objects.filter(employeeId=i['employeeId'],leave_id=i['id'],company_code=company_code)
            leave_appprovel_serializer=leaveapprovalserializer(leave_appprovel_obj,many=True)

            employee=Users.objects.filter(id=i['employeeId']).first()
            employee_serializer=UserSerializer(employee)
            i['Firstname']=employee_serializer.data['Firstname']
            i['Lastname']=employee_serializer.data['Lastname']
            start_date =  arrow.get(i['start_date'])
            end_date =  arrow.get(i['end_date'])
            delta =(start_date-end_date)
            i['days']=abs(delta.days)+1
            pstart_date = i['start_date'] 
            pendingstart_date = pstart_date.split('-')
            pstart_datestr = '-'.join(reversed(pendingstart_date))
            i['start_date'] = pstart_datestr
            pend_date = i['end_date'] 
            psplitend_date = pend_date.split('-')
            psplitend_datestr = '-'.join(reversed(psplitend_date))
            i['end_date'] = psplitend_datestr
            
            created = i['created_at']
            datet = created.split('T')[0]
            i['created_at']= dateformat(datet)
            

                    
            i['start_date']=ddmonthyy(i['start_date'])
            i['end_date']=ddmonthyy(i['end_date'])
            
            if i['start_date']==i['end_date']:
                i['leavedates']="<p class='leave-date'>"+i['start_date'] +"</p><div class='leave-date'></div>"
            else:
                i['leavedates']= "<div class='leave-date'>"+i['start_date'] +" to </div> <div class='leave-date'>" +i['end_date']+"</div>"

            if i['leavetype']=="Fullday":

                if i['days'] > 1:
                    i['leaveduration']="Total " + str(i['days']) + " Days"
                else:
                    i['leaveduration']="Full Day"
            else:
                if i['leavetype'] =="SecondHalf":
                    i['leaveduration']="Second Half"
                if i['leavetype'] =="FirstHalf":
                    i['leaveduration']="First Half"


            if i['WorkFromHome']==True:
                if logined_user_obj is None:
                    i['reason']="This employee is working from home. "
                i['applicationtype']=" Work From Home"
            else: 
                if logined_user_obj is None:
                    i['reason']="This employee is on leave."
                i['applicationtype']="Leave"

                    
            dummy_idlist=[]   
            count=1         
            managerlist_unique=[]    

            for j in leave_appprovel_serializer.data:

                managers=Users.objects.filter(id=j['managerId']).first()
                manager_serializer=UserSerializer(managers)
                j['position']=count
                j['Photo']=manager_serializer.data['Photo']
                j['Firstname']=manager_serializer.data['Firstname'] 
                j['Lastname']=manager_serializer.data['Lastname']
                count+=1
                if j['managerId'] not in dummy_idlist:
                    dummy_idlist.append(j['managerId'])
                    managerlist_unique.append(j)
                
                
            i['managerList']=managerlist_unique

            todaydate = str(date.today())
 

                
            delta1 = end_date - start_date   
            for da in range(delta1.days + 1): 
                day1 = start_date + timedelta(days=da) 
                new_date=day1.format('YYYY-MM-DD')
                if  todaydate == new_date:
                    leavelist.append(i)
                    
        return Response ({"data":leavelist,"response":{"n" : 1,"msg" : "SUCCESS","status" : "success"}})

@api_view(['GET'])
def UserleaveListAPI(request):


    employee_id=request.user.id
    if request.method == 'GET':
        
        def dateformat2(date_str):
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            formatted_date = date_obj.strftime("%d-%m-%Y")
            return(formatted_date)

        #draft
        user = Leave.objects.filter(employeeId=employee_id,leave_status="Draft",Active=True).order_by('-id','start_date','leavetype').distinct("id")
        dserializer = leaveserializer(user, many=True)
        for i in dserializer.data:
            created = i['created_at']
            datet = created.split('T')[0]
            i['dd_month_sdate'] = dd_month_year_format(str(i['start_date']))
            i['dd_month_edate'] = dd_month_year_format(str(i['end_date']))
            
            i['total_days']=countdays(i['start_date'],i['end_date'])
            i['created_at']= dd_month_year_format(datet)

            i['fstart_date'] = dateformat2(i['start_date'])
            i['fend_date'] =dateformat2(i['end_date'])

            i['end_date'] = dateformat(i['end_date'])
            i['start_date'] = dateformat(i['start_date'])

            i['managerstatus']=[]

    


        user = Leave.objects.filter(employeeId=employee_id,leave_status="Pending",Active=True).order_by('-id').distinct("id")
        pserializer = leaveserializer(user, many=True)
        for i in pserializer.data:
            created = i['created_at']
            datet = created.split('T')[0]
            i['dd_month_sdate'] = dd_month_year_format(str(i['start_date']))
            i['dd_month_edate'] = dd_month_year_format(str(i['end_date']))
            Withraw_status  = is_leave_time_valid(i['start_date'])
            i['Withraw_status']=Withraw_status
            i['total_days']=countdays(i['start_date'],i['end_date'])
            i['created_at']= dd_month_year_format(datet)
            i['fstart_date'] = dateformat2(i['start_date'])
            i['fend_date'] =dateformat2(i['end_date'])
            i['start_date'] = dateformat(i['start_date'])
            i['end_date'] = dateformat(i['end_date'])

            prevempmanager = leaveApproval.objects.filter(leave_id=i['id'],approvedBy=True).last()
            mappingobj=leaveApproval.objects.filter(leave_id=i['id']).distinct("managerId")
            elserializer=leaveapprovalserializer(mappingobj,many=True)
            emplist=[]
            editable_status=[]
            pcount=1
            for t in elserializer.data:
                leaveobjone=leaveApproval.objects.filter(managerId=t['managerId'],leave_id=i['id']).first()
                userObj=Users.objects.filter(id=int(t['managerId'])).first()
                userprofileimage = userObj.Photo
                if str(userprofileimage) is not None and str(userprofileimage) != "":
                    userprf = imageUrl +"/media/" + str(userprofileimage)
                else:
                    userprf = imageUrl + "/static/assets/images/profile.png"
                    
                if leaveobjone is not None:
                    editable_status.append(leaveobjone.approvedBy)

                    if leaveobjone.approvedBy == True:
                        data={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":leaveobjone.comment,
                            "managerid":leaveobjone.managerId,
                            "managerimage":userprf,
                            "position":pcount,
                            "status":"Approved",
                        }
                        emplist.append(data)
                    elif leaveobjone.rejectedBy == True:
                        data={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":leaveobjone.comment,
                            "managerid":leaveobjone.managerId,
                            "position":pcount,
                            "status":"Rejected",
                            "managerimage":userprf,
                        }
                        emplist.append(data)

                    else:
                        data={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":"",
                            "managerid":leaveobjone.managerId,
                            "status":"Pending",
                            "position":pcount,
                            "managerimage":userprf,
                        }
                        emplist.append(data)


                pcount+=1
            i['managerstatus']=emplist
            

                

            operationsonit=sum(editable_status)
            if operationsonit>0:
                i['editable']=False
            else:
                i['editable']=True


        #approved
        emplist=[]
        userdata= Leave.objects.filter(employeeId=employee_id,leave_status="Approved",Active=True).order_by('-id').distinct("id")
        appserializer = leaveserializer(userdata, many=True)
        for i in appserializer.data:
            created = i['created_at']
            datet = created.split('T')[0] 
            i['dd_month_sdate'] = dd_month_year_format(str(i['start_date']))
            i['dd_month_edate'] = dd_month_year_format(str(i['end_date']))
            Withraw_status  = is_leave_time_valid(i['start_date'])
            i['Withraw_status']=Withraw_status
            i['total_days']=countdays(i['start_date'],i['end_date'])
            i['created_at']= dd_month_year_format(datet)
            i['fstart_date'] =i['start_date']
            i['fend_date'] =i['end_date']
            i['start_date'] = dateformat(i['start_date'])
            i['end_date'] = dateformat(i['end_date'])


            mappingobj=leaveApproval.objects.filter(leave_id=i['id']).distinct("managerId")
            elserializer=leaveapprovalserializer(mappingobj,many=True)
            emplist=[]
            acount=1
            for t in elserializer.data:
                leaveobjone=leaveApproval.objects.filter(managerId=t['managerId'],leave_id=i['id']).first()
              
                userObj=Users.objects.filter(id=int(t['managerId'])).first()
                userprofileimage = userObj.Photo
                if str(userprofileimage) is not None and str(userprofileimage) != "":
                    userprf = imageUrl +"/media/" + str(userprofileimage)
                else:
                    userprf = imageUrl + "/static/assets/images/profile.png"
                if leaveobjone is not None:
                    if leaveobjone.approvedBy == True:
                        data={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":leaveobjone.comment,
                            "managerid":leaveobjone.managerId,
                            "managerimage":userprf,
                            "status":"Approved",
                            "position":acount,    
                        }
                        emplist.append(data)
                    elif leaveobjone.rejectedBy == True:
                        data={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":leaveobjone.comment,
                            "managerid":leaveobjone.managerId,
                            "managerimage":userprf,
                            "status":"Rejected",
                            "position":acount,     
                        }
                        emplist.append(data)
                    else:
                        data={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":"",
                            "managerid":leaveobjone.managerId,
                            "managerimage":userprf ,
                            "status":"Pending",
                            "position":acount, 
                        }
                        emplist.append(data)

                acount+=1
            i['managerstatus']=emplist

           

        #rejected
        rejemplist=[]
        userdata= Leave.objects.filter(employeeId=employee_id,leave_status="Rejected",Active=True).order_by('-id').distinct("id")
        rejserializer = leaveserializer(userdata, many=True)
        for i in rejserializer.data:
            created = i['created_at']
            datet = created.split('T')[0]
            i['dd_month_sdate'] = dd_month_year_format(str(i['start_date']))
            i['dd_month_edate'] = dd_month_year_format(str(i['end_date']))
            i['total_days']=countdays(i['start_date'],i['end_date'])
            i['created_at']= dd_month_year_format(datet)
            i['fstart_date'] =i['start_date']
            i['fend_date'] =i['end_date']
            i['start_date'] = dateformat(i['start_date'])
            i['end_date'] = dateformat(i['end_date'])

            mappingobj=leaveApproval.objects.filter(leave_id=i['id']).distinct("managerId")
            elserializer=leaveapprovalserializer(mappingobj,many=True)
            rejemplist=[]
            rcount=1
            managercount=mappingobj.count()
            rejectedcount=0
            for t in elserializer.data:
   
                leaveobjone=leaveApproval.objects.filter(managerId=t['managerId'],leave_id=i['id']).first()
              
                userObj=Users.objects.filter(id=int(t['managerId'])).first()
                userprofileimage = userObj.Photo
                if str(userprofileimage) is not None and str(userprofileimage) != "":
                    userprf = imageUrl +"/media/" + str(userprofileimage)
                else:
                    userprf = imageUrl + "/static/assets/images/profile.png"
                if leaveobjone is not None:
                    if leaveobjone.approvedBy == True:
                        data={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":leaveobjone.comment,
                            "managerid":leaveobjone.managerId,
                            "managerimage":userprf,
                            "position":rcount,
                            "status":"Approved",  
                        }
                        rejemplist.append(data)
                    elif leaveobjone.rejectedBy == True:
                        rejectedcount+=1
                        data={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":leaveobjone.comment,
                            "managerid":leaveobjone.managerId,
                            "managerimage":userprf,
                            "position":rcount,
                            "status":"Rejected",    
                        }
                        rejemplist.append(data)
                    else:
                        data={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":"",
                            "managerid":leaveobjone.managerId,
                            "managerimage":userprf,
                            "position":rcount,
                            "status":"Pending",    
                        }
                        rejemplist.append(data)

                rcount+=1
            if rejectedcount == managercount:
                i['universel_comment']="No response to your application"
            else:
                i['universel_comment']=""
            i['managerstatus']=rejemplist
          





        #withdrawed
        userdata= Leave.objects.filter(employeeId=employee_id,leave_status="Withdraw",Active=True).order_by('-id').distinct("id")
        withrawserializer = leaveserializer(userdata, many=True)
        for i in withrawserializer.data:
            created = i['created_at']
            datet = created.split('T')[0]
            i['dd_month_sdate'] = dd_month_year_format(str(i['start_date']))
            i['dd_month_edate'] = dd_month_year_format(str(i['end_date']))
            i['total_days']=countdays(i['start_date'],i['end_date'])
            i['created_at']= dd_month_year_format(datet)
            i['fstart_date'] =i['start_date']
            i['fend_date'] =i['end_date']
            i['start_date'] = dateformat(i['start_date'])
            i['end_date'] = dateformat(i['end_date'])

            mappingobj=leaveApproval.objects.filter(leave_id=i['id']).distinct("managerId")
            elserializer=leaveapprovalserializer(mappingobj,many=True)
            withdraw_emplist=[]
            wcount=1
            for t in elserializer.data:
   
                leaveobjone=leaveApproval.objects.filter(managerId=t['managerId'],leave_id=i['id']).first()
              
                userObj=Users.objects.filter(id=int(t['managerId'])).first()
                userprofileimage = userObj.Photo
                if str(userprofileimage) is not None and str(userprofileimage) != "":
                    userprf = imageUrl +"/media/" + str(userprofileimage)
                else:
                    userprf = imageUrl + "/static/assets/images/profile.png"
                if leaveobjone is not None:
                    if leaveobjone.approvedBy == True:
                        data={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":leaveobjone.comment,
                            "managerid":leaveobjone.managerId,
                            "managerimage":userprf,
                            "status":"Approved",
                            "position":wcount,  
                        }
                        withdraw_emplist.append(data)
                    elif leaveobjone.rejectedBy == True:
                     
                        data={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":leaveobjone.comment,
                            "managerid":leaveobjone.managerId,
                            "managerimage":userprf,
                            "status":"Rejected",
                            "position":wcount,  
                        }
                        withdraw_emplist.append(data)
                    else:
                        data={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":"",
                            "managerid":leaveobjone.managerId,
                            "managerimage":userprf,
                            "status":"Pending",
                            "position":wcount,  
                        }
                        withdraw_emplist.append(data)

                wcount+=1    
            i['managerstatus']=withdraw_emplist
          
 

        leavelist ={
            "pending_list":pserializer.data,
            "Approved_list":appserializer.data,
            "Rejected_list":rejserializer.data,
            "Draft_list":dserializer.data,
            "Withdraw_list":withrawserializer.data,
            }
        
        return Response ({"data":leavelist,"response":{"n" : 1,"msg" : "SUCCESS","status" : "success"}})
    
@api_view(['POST'])
def FilterleaveListAPI(request):
  
    dictlist=[]
    rejectedlist=[]
    loginuserID=request.user.id 
    current_date = datetime.now().strftime('%Y-%m-%d')
    month=request.POST.get('month')
    year=request.POST.get('year')

    leaveapp = leaveApproval.objects.filter(managerId=loginuserID,approvedBy=False,rejectedBy=False).order_by('-leave_id', '-id').distinct("leave_id")
    pendingSerializer = leaveapprovalserializer(leaveapp,many=True)
    
    mid_pending_list=[]
    
    for i in pendingSerializer.data:
    
        leavedata = Leave.objects.filter(id=i['leave_id'],start_date__gte=current_date,Active=True,).exclude(leave_status="Withdraw")
        if leavedata is not None:
            if month is not None and month  != "":
                leavedata=leavedata.filter(start_date__month=month)
            if year is not None and year  != "":
                leavedata=leavedata.filter(start_date__year=year)
            leavedata=leavedata.first()
            lSerializer = leaveserializer(leavedata)
            if leavedata is not None:
                i['attachment'] = lSerializer.data['attachment']

                i['dd_month_sdate'] = dd_month_year_format(str(lSerializer.data['start_date']))
                i['dd_month_edate'] = dd_month_year_format(str(lSerializer.data['end_date']))
                i['start_date'] = lSerializer.data['start_date']
                i['end_date'] = lSerializer.data['end_date']
                datedata = lSerializer.data['created_at']
                i['created_at']=datedata.split('T')[0]
                i['reason']= lSerializer.data['reason']
                i['ApplicationId'] = lSerializer.data['ApplicationId']
                i['leavetype'] = lSerializer.data['leavetype']
                i['leave_status'] = lSerializer.data['leave_status']
                i['action_taken'] = False
                i['currentuser'] = loginuserID
                userObj=Users.objects.filter(id=int(lSerializer.data['employeeId'])).first()
                user_leave_serializer=UserSerializer(userObj)
                i['applicant_name'] = userObj.Firstname +" "+ userObj.Lastname
                i['department'] = user_leave_serializer.data['DepartmentID']

                if lSerializer.data['WorkFromHome'] == True:
                    i['ApplicationType']='Work From Home'
                else:
                    i['ApplicationType']='Leave'


                start_date =  arrow.get(i['start_date'])
                end_date =  arrow.get(i['end_date'])
                delta =(start_date-end_date)
                tdays=abs(delta.days)+1
                i['total_days']=tdays
                i['created_at']=dd_month_year_format(i['created_at'])
                i['start_date']=convertdate(i['start_date'])
                i['end_date']=convertdate(i['end_date'])

                manager = leaveMapping.objects.filter(employeeId=lSerializer.data['employeeId']).distinct("managerId")
                serializer=leaveMappingserializer(manager,many=True)
                dictlist=[]
                for c in serializer.data:
                    position = c['position']
                    managerdata = Users.objects.filter(id=int(c['managerId'])).first()
                    managerstr = managerdata.Firstname +" "+ managerdata.Lastname
                    dict1 = {"name":managerstr,
                        "position":position,
                        "manageruserid":c['managerId']}
                    dictlist.append(dict1)
                i['managerlist']=dictlist
                
                if lSerializer.data['leavetype'] == 'Fullday':
                    i['leavetype']="Full Day"
                elif lSerializer.data['leavetype'] =="SecondHalf":
                    i['leavetype']="Second Half "
                elif lSerializer.data['leavetype'] =="FirstHalf":
                    i['leavetype']="First Half "


                if(i['start_date'] == i['end_date']):
                    i['leaveduration']=i['start_date']
                    
                else:
                    i['leaveduration']=i['start_date'] +" to "+ i['end_date']


                filterleaveobj=leaveApproval.objects.filter(leave_id=i['leave_id']).first()
                previousManager=leaveApproval.objects.filter(leave_id=i['leave_id'],rejectedBy=True).exclude(managerId=loginuserID).order_by('-id').first()
                approveManager = leaveApproval.objects.filter(leave_id=i['leave_id'],approvedBy=True).exclude(managerId=loginuserID).order_by('-id').first()
            
                mappingobj=leaveApproval.objects.filter(leave_id=i['leave_id'])
                elserializer=leaveapprovalserializer(mappingobj,many=True)
                llist=[]
                pcount=1    

                for t in elserializer.data:
                    leaveobjone=leaveApproval.objects.filter(managerId=t['managerId'],leave_id=i['leave_id']).first()
                    userObj=Users.objects.filter(id=int(t['managerId'])).first()
                    
                    if userObj.Photo is not None and userObj.Photo !="":
                        managerpic = imageUrl +"/media/"+ str(userObj.Photo)
                    else:
                        managerpic = imageUrl + "/static/assets/images/profile.png"
                    if leaveobjone is not None:
                        if leaveobjone.approvedBy == True:
                        
                            data={
                                "applicationId":i['ApplicationId'],
                                "approvedBy":leaveobjone.approvedBy,
                                "name":userObj.Firstname+" "+userObj.Lastname,
                                "rejectedBy":leaveobjone.rejectedBy,
                                "comment":leaveobjone.comment,
                                "managerid":leaveobjone.managerId,
                                "status":"Approved",
                                "Photo":managerpic,
                                "position":pcount,
                                
                            }
                            llist.append(data)
                        elif leaveobjone.rejectedBy == True:
                        
                            data={
                                "applicationId":i['ApplicationId'],
                                "approvedBy":leaveobjone.approvedBy,
                                "name":userObj.Firstname+" "+userObj.Lastname,
                                "rejectedBy":leaveobjone.rejectedBy,
                                "comment":leaveobjone.comment,
                                "managerid":leaveobjone.managerId,
                                "status":"Rejected",
                                "Photo":managerpic,
                                "position":pcount,
                            }
                            llist.append(data)
                        else:
                            data={
                                "applicationId":i['ApplicationId'],
                                "approvedBy":leaveobjone.approvedBy,
                                "name":userObj.Firstname+" "+userObj.Lastname,
                                "rejectedBy":leaveobjone.rejectedBy,
                                "comment":"",
                                "managerid":leaveobjone.managerId,
                                "status":"Pending",
                                "Photo":managerpic,
                                "position":pcount,
                            }
                            llist.append(data)

                    pcount+=1
                i['allleavedata']=llist
                if previousManager is not None:
                    i['passedBy']=previousManager.managerId
                    userObj=Users.objects.filter(id=previousManager.managerId).first()
                    i['passedByName'] =userObj.Firstname+" "+userObj.Lastname

                else:
                    i['passedBy']="-----------"
                    i['passedByName'] ="--"

                if approveManager is not None:
                    i['approvedpassBy']=approveManager.managerId
                    userObj=Users.objects.filter(id=approveManager.managerId).first()
                    i['apprvpassedByName'] =userObj.Firstname+" "+userObj.Lastname

                else:
                    i['approvedpassBy']="-----------"
                    i['apprvpassedByName'] ="--"
                
                
                
                if lSerializer.data['leave_status'] !="Withdraw":
                    mid_pending_list.append(i)
            
            
           
    pendinglist=[] 
    for i in mid_pending_list:
        i['discard']=False 
        if i['allleavedata']: 
            for j in i['allleavedata']:
                if j['rejectedBy'] == True:
                    i['discard']=True
            
             
    for i in mid_pending_list:
        if i['discard']!=True:
            if i not in pendinglist:
                pendinglist.append(i)



                 
    apprvedleavelist=[]
    leaveapp = leaveApproval.objects.filter(managerId=loginuserID,approvedBy=True).order_by('-leave_id', '-id').distinct("leave_id")
    appSerializer = leaveapprovalserializer(leaveapp,many=True)
    for i in appSerializer.data:


        leavedata = Leave.objects.filter(Q(id=i['leave_id']),Q(Active=True),Q(leave_status="Approved")|Q(leave_status="Pending"))
        if month is not None and month  != "":
            leavedata=leavedata.filter(start_date__month=month)
        if year is not None and year  != "":
            leavedata=leavedata.filter(start_date__year=year)
        leavedata=leavedata.first()

        lSerializer = leaveserializer(leavedata)
        if leavedata is not None:
            i['dd_month_sdate'] = dd_month_year_format(str(lSerializer.data['start_date']))
            i['dd_month_edate'] = dd_month_year_format(str(lSerializer.data['end_date']))
            i['leave_status'] = lSerializer.data['leave_status']
            i['action_taken'] = True

          

            
            i['attachment'] = lSerializer.data['attachment']

            i['start_date'] = lSerializer.data['start_date']
            i['end_date'] = lSerializer.data['end_date']
            datedata = lSerializer.data['created_at']

            if current_date <= i['start_date']:
                i['allow_reject'] = True
                i['allow_approve'] = True
            else:
                i['allow_approve'] = False
                i['allow_reject'] = False
                
            i['created_at']=datedata.split('T')[0]
            i['reason']= lSerializer.data['reason']
            i['ApplicationId'] = lSerializer.data['ApplicationId']
            userObj=Users.objects.filter(id=int(lSerializer.data['employeeId'])).first()
            user_leave_serializer=UserSerializer(userObj)
            i['applicant_name'] = userObj.Firstname +" "+ userObj.Lastname
            i['department'] = user_leave_serializer.data['DepartmentID']

            start_date =  arrow.get(i['start_date'])
            end_date =  arrow.get(i['end_date'])
            delta =(start_date-end_date)
            tdays=abs(delta.days)+1
            i['total_days']=tdays
            i['created_at']=dd_month_year_format(i['created_at'])
            i['start_date']=convertdate(i['start_date'])
            i['end_date']=convertdate(i['end_date'])

            if lSerializer.data['WorkFromHome'] == True:
                i['ApplicationType']='Work From Home'
            else:
                i['ApplicationType']='Leave'
                
            if lSerializer.data['leavetype'] == 'Fullday':
                i['leavetype']="Full Day"
            elif lSerializer.data['leavetype'] =="SecondHalf":
                i['leavetype']="Second Half "
            elif lSerializer.data['leavetype'] =="FirstHalf":
                i['leavetype']="First Half "

            if(i['start_date'] == i['end_date']):
                i['leaveduration']=i['start_date']
                
            else:
                i['leaveduration']=i['start_date'] +" to "+ i['end_date']

            mappingobj=leaveApproval.objects.filter(leave_id=i['leave_id']).distinct("managerId")
            elserializer=leaveapprovalserializer(mappingobj,many=True)
            llist=[]
            acount=1    

            for t in elserializer.data:
                leaveobjone=leaveApproval.objects.filter(managerId=t['managerId'],leave_id=i['leave_id']).first()
                userObj=Users.objects.filter(id=int(t['managerId'])).first()
                if userObj.Photo is not None and userObj.Photo !="":
                    managerpic = imageUrl +"/media/"+ str(userObj.Photo)
                else:
                    managerpic = imageUrl + "/static/assets/images/profile.png"
                if leaveobjone is not None:
                    if leaveobjone.approvedBy == True: 
                    
                        data={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":leaveobjone.comment,
                            "managerid":leaveobjone.managerId,
                            "Photo":managerpic,
                            "position":acount,
                            "status":"Approved",
                        }
                        llist.append(data)
                    elif leaveobjone.rejectedBy == True:
                    
                        data={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":leaveobjone.comment,
                            "managerid":leaveobjone.managerId,
                            "Photo":managerpic,
                            "position":acount,
                            "status":"Rejected",
                        }
                        llist.append(data)
                    else:
                        data={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":"",
                            "managerid":leaveobjone.managerId,
                            "Photo":managerpic,
                            "position":acount,
                            "status":"Pending",
                        }
                        llist.append(data)


            acount+=1
            i['allleavedata']=llist

            apprvedleavelist.append(i)



    leaveapp = leaveApproval.objects.filter(managerId=loginuserID).order_by('-leave_id', '-id').distinct("leave_id")
    rejectedSerializer = leaveapprovalserializer(leaveapp,many=True)
    rej_duplicatelist=[]           

    for i in rejectedSerializer.data: 
        i['discard']=False

        leave_id = i['leave_id']
        leaves_obj = leaveApproval.objects.filter(leave_id=leave_id)
        leaves_obj_Serializer = leaveapprovalserializer(leaves_obj,many=True)
        for l in leaves_obj_Serializer.data:
            if l['rejectedBy']==True:
                i['discard']=True

        if i['discard']==True:

            if i['leave_id'] not in rej_duplicatelist:





                leavedata_rej = Leave.objects.filter(id=i['leave_id'],Active=True)
                if month is not None and month  != "":
                    leavedata_rej=leavedata_rej.filter(start_date__month=month)
                if year is not None and year  != "":
                    leavedata_rej=leavedata_rej.filter(start_date__year=year)
                leavedata_rej=leavedata_rej.first()

                lSerializer = leaveserializer(leavedata_rej)
                if leavedata_rej is not None:
                    rej_duplicatelist.append(i['leave_id'])

                    i['leave_status'] = lSerializer.data['leave_status']
                    i['action_taken'] = True
                    i['attachment'] = lSerializer.data['attachment']

                    i['dd_month_sdate'] = dd_month_year_format(str(lSerializer.data['start_date']))
                    i['dd_month_edate'] = dd_month_year_format(str(lSerializer.data['end_date']))
                    i['start_date'] = lSerializer.data['start_date'] 
                    i['end_date'] = lSerializer.data['end_date']
                    datedata = lSerializer.data['created_at']
                    i['created_at']=datedata.split('T')[0]
                    i['reason']= lSerializer.data['reason']
                    i['ApplicationId'] = lSerializer.data['ApplicationId']
                    userObj=Users.objects.filter(id=int(lSerializer.data['employeeId'])).first()
                    user_leave_serializer=UserSerializer(userObj)
                    i['applicant_name'] = userObj.Firstname +" "+ userObj.Lastname
                    i['department'] = user_leave_serializer.data['DepartmentID']
                    if current_date <= i['start_date']:
                        i['allow_reject'] = True
                        i['allow_approve'] = True
                    else:
                        i['allow_approve'] = False
                        i['allow_reject'] = False
                    
                    start_date =  arrow.get(i['start_date'])
                    end_date =  arrow.get(i['end_date'])
                    delta =(start_date-end_date)
                    tdays=abs(delta.days)+1
                    i['total_days']=tdays
                    i['created_at']=dd_month_year_format(i['created_at'])
                    i['start_date']=convertdate(i['start_date'])
                    i['end_date']=convertdate(i['end_date'])

                    if lSerializer.data['WorkFromHome'] == True:
                        i['ApplicationType']='Work From Home'
                    else:
                        i['ApplicationType']='Leave'


                    if lSerializer.data['leavetype'] == 'Fullday':
                        i['leavetype']="Full Day"
                    elif lSerializer.data['leavetype'] =="SecondHalf":
                        i['leavetype']="Second Half "
                    elif lSerializer.data['leavetype'] =="FirstHalf":
                        i['leavetype']="First Half "

                    if(i['start_date'] == i['end_date']):
                        i['leaveduration']=i['start_date']
                        
                    else:
                        i['leaveduration']=i['start_date'] +" to "+ i['end_date']

                    mappingobj=leaveApproval.objects.filter(leave_id=i['leave_id']).distinct("managerId")
                    elserializer=leaveapprovalserializer(mappingobj,many=True)
                    llist=[]
                    rcount=1    

                    for t in elserializer.data:
                        leaveobjone=leaveApproval.objects.filter(managerId=t['managerId'],leave_id=i['leave_id']).first()
                        userObj=Users.objects.filter(id=int(t['managerId'])).first()
                        if userObj.Photo is not None and userObj.Photo !="":
                            managerpic = imageUrl +"/media/"+ str(userObj.Photo)
                        else:
                            managerpic = imageUrl + "/static/assets/images/profile.png"
                        if leaveobjone is not None:
                            if leaveobjone.approvedBy == True: 
                            
                                data={
                                    "applicationId":i['ApplicationId'],
                                    "approvedBy":leaveobjone.approvedBy,
                                    "name":userObj.Firstname+" "+userObj.Lastname,
                                    "rejectedBy":leaveobjone.rejectedBy,
                                    "comment":leaveobjone.comment,
                                    "managerid":leaveobjone.managerId,
                                    "position":rcount,
                                    "Photo":managerpic,
                                    "status":"Approved",
                                }
                                llist.append(data)
                            elif leaveobjone.rejectedBy == True:
                            
                                data={
                                    "applicationId":i['ApplicationId'],
                                    "approvedBy":leaveobjone.approvedBy,
                                    "name":userObj.Firstname+" "+userObj.Lastname,
                                    "rejectedBy":leaveobjone.rejectedBy,
                                    "comment":leaveobjone.comment,
                                    "managerid":leaveobjone.managerId,
                                    "position":rcount,
                                    "Photo":managerpic,
                                    "status":"Rejected",
                                }
                                llist.append(data)
                            else:
                                data={
                                    "applicationId":i['ApplicationId'],
                                    "approvedBy":leaveobjone.approvedBy,
                                    "name":userObj.Firstname+" "+userObj.Lastname,
                                    "rejectedBy":leaveobjone.rejectedBy,
                                    "comment":"",
                                    "managerid":leaveobjone.managerId,
                                    "position":rcount,
                                    "Photo":managerpic,
                                    "status":"Pending",
                                }
                                llist.append(data)


                    rcount+=1
                    i['allleavedata']=llist
                    rejectedlist.append(i)



            






    mid_expired_list=[]

    leaveexpired = leaveApproval.objects.filter(managerId=loginuserID,approvedBy=False,rejectedBy=False).order_by('-leave_id', '-id').distinct("leave_id")
    expiredSerializer = leaveapprovalserializer(leaveexpired,many=True)
    for i in expiredSerializer.data:
        leavedata = Leave.objects.filter(id=i['leave_id'],start_date__lt=current_date,Active=True,).exclude(Q(leave_status="Rejected")|Q(leave_status="Withdrawn"))
        if month is not None and month  != "":
            leavedata=leavedata.filter(start_date__month=month)
        if year is not None and year  != "":
            leavedata=leavedata.filter(start_date__year=year)
        leavedata=leavedata.first()

        lSerializer = leaveserializer(leavedata)
        
        
        if leavedata is not None:
            
          
            i['attachment'] = lSerializer.data['attachment']

        
            i['dd_month_sdate'] = dd_month_year_format(str(lSerializer.data['start_date']))
            i['dd_month_edate'] = dd_month_year_format(str(lSerializer.data['end_date']))
            i['start_date'] = lSerializer.data['start_date']
            i['end_date'] = lSerializer.data['end_date']
            datedata = lSerializer.data['created_at']
            i['created_at']=datedata.split('T')[0]
            i['reason']= lSerializer.data['reason']
            i['ApplicationId'] = lSerializer.data['ApplicationId']
            i['leavetype'] = lSerializer.data['leavetype']
            i['leave_status'] = lSerializer.data['leave_status']
            i['action_taken'] = False
            i['currentuser'] = loginuserID
            userObj=Users.objects.filter(id=int(lSerializer.data['employeeId'])).first()
            user_leave_serializer=UserSerializer(userObj)
            i['applicant_name'] = userObj.Firstname +" "+ userObj.Lastname
            i['department'] = user_leave_serializer.data['DepartmentID']

            if lSerializer.data['WorkFromHome'] == True:
                i['ApplicationType']='Work From Home'
            else:
                i['ApplicationType']='Leave'







            start_date =  arrow.get(i['start_date'])
            end_date =  arrow.get(i['end_date'])
            delta =(start_date-end_date)
            tdays=abs(delta.days)+1
            i['total_days']=tdays
            i['created_at']=dd_month_year_format(i['created_at'])
            i['start_date']=convertdate(i['start_date'])
            i['end_date']=convertdate(i['end_date'])

            manager = leaveMapping.objects.filter(employeeId=lSerializer.data['employeeId']).distinct("managerId")
            serializer=leaveMappingserializer(manager,many=True)
            dictlist=[]
            for c in serializer.data:
                position = c['position']
                managerdata = Users.objects.filter(id=int(c['managerId'])).first()
                managerstr = managerdata.Firstname +" "+ managerdata.Lastname
                dict1 = {"name":managerstr,
                    "position":position,
                    "manageruserid":c['managerId']}
                dictlist.append(dict1)
            i['managerlist']=dictlist
            
            if lSerializer.data['leavetype'] == 'Fullday':
                i['leavetype']="Full Day"
            elif lSerializer.data['leavetype'] =="SecondHalf":
                i['leavetype']="Second Half "
            elif lSerializer.data['leavetype'] =="FirstHalf":
                i['leavetype']="First Half "


            if(i['start_date'] == i['end_date']):
                i['leaveduration']=i['start_date']
                
            else:
                i['leaveduration']=i['start_date'] +" to "+ i['end_date']



            mappingobj=leaveApproval.objects.filter(leave_id=i['leave_id'])
            elserializer=leaveapprovalserializer(mappingobj,many=True)
            expired_llist=[]
            pcount=1    

            for t in elserializer.data:
                leaveobjone=leaveApproval.objects.filter(managerId=t['managerId'],leave_id=i['leave_id']).first()
                userObj=Users.objects.filter(id=int(t['managerId'])).first()
                
                if userObj.Photo is not None and userObj.Photo !="":
                    managerpic = imageUrl +"/media/"+ str(userObj.Photo)
                else:
                    managerpic = imageUrl + "/static/assets/images/profile.png"
                if leaveobjone is not None:
                    if leaveobjone.approvedBy == True:
                    
                        data={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":leaveobjone.comment,
                            "managerid":leaveobjone.managerId,
                            "status":"Approved",
                            "Photo":managerpic,
                            "position":pcount,
                            
                        }
                        expired_llist.append(data)
                    elif leaveobjone.rejectedBy == True:
                    
                        data={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":leaveobjone.comment,
                            "managerid":leaveobjone.managerId,
                            "status":"Rejected",
                            "Photo":managerpic,
                            "position":pcount,
                        }
                        expired_llist.append(data)
                    else:
                        data={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":"",
                            "managerid":leaveobjone.managerId,
                            "status":"Pending",
                            "Photo":managerpic,
                            "position":pcount,
                        }
                        expired_llist.append(data)

                pcount+=1
            i['allleavedata']=expired_llist
            
            
            
            
            if lSerializer.data['leave_status'] !="Withdraw":
                mid_expired_list.append(i)
        
        
        
        
        
            
            
            

            
            
            
            
    expiredlist=[] 
    for i in mid_expired_list:
        i['discard']=False 
        if i['allleavedata']: 
            for j in i['allleavedata']:
                if j['rejectedBy'] == True:
                    i['discard']=True
            
             
    for i in mid_expired_list:
        if i['discard']!=True:
            if i not in expiredlist:
                expiredlist.append(i)



    # withdraw list



    leavewithdraw = leaveApproval.objects.filter(managerId=loginuserID,).exclude(rejectedBy=True).order_by('-leave_id', '-id').distinct("leave_id")
    withdrawSerializer = leaveapprovalserializer(leavewithdraw,many=True)
    
    mid_withdraw_list=[]
    
    for i in withdrawSerializer.data:
        leavedata = Leave.objects.filter(id=i['leave_id'],Active=True,leave_status="Withdraw")
        if month is not None and month  != "":
            leavedata=leavedata.filter(start_date__month=month)
        if year is not None and year  != "":
            leavedata=leavedata.filter(start_date__year=year)
        leavedata=leavedata.first()

        lSerializer = leaveserializer(leavedata)
        if leavedata is not None:
            i['dd_month_sdate'] = dd_month_year_format(str(lSerializer.data['start_date']))
            i['dd_month_edate'] = dd_month_year_format(str(lSerializer.data['end_date']))
            i['start_date'] = lSerializer.data['start_date']
            i['end_date'] = lSerializer.data['end_date']
            datedata = lSerializer.data['created_at']
            i['leave_status'] = lSerializer.data['leave_status']
            i['action_taken'] = True
            i['attachment'] = lSerializer.data['attachment']
            i['created_at']=datedata.split('T')[0]
            i['reason']= lSerializer.data['reason']
            i['ApplicationId'] = lSerializer.data['ApplicationId']
            i['leavetype'] = lSerializer.data['leavetype']
            i['currentuser'] = loginuserID
            userObj=Users.objects.filter(id=int(lSerializer.data['employeeId'])).first()
            user_leave_serializer=UserSerializer(userObj)
            i['applicant_name'] = userObj.Firstname +" "+ userObj.Lastname
            i['department'] = user_leave_serializer.data['DepartmentID']

            if lSerializer.data['WorkFromHome'] == True:
                i['ApplicationType']='Work From Home'
            else:
                i['ApplicationType']='Leave'


            start_date =  arrow.get(i['start_date'])
            end_date =  arrow.get(i['end_date'])
            delta =(start_date-end_date)
            tdays=abs(delta.days)+1
            i['total_days']=tdays
            i['created_at']=dd_month_year_format(i['created_at'])
            i['start_date']=convertdate(i['start_date'])
            i['end_date']=convertdate(i['end_date'])

            manager = leaveMapping.objects.filter(employeeId=lSerializer.data['employeeId']) 
            serializer=leaveMappingserializer(manager,many=True)
            dictlist=[]
            for c in serializer.data:
                position = c['position']
                managerdata = Users.objects.filter(id=int(c['managerId'])).first()
                managerstr = managerdata.Firstname +" "+ managerdata.Lastname
                dict1 = {"name":managerstr,
                    "position":position,
                    "manageruserid":c['managerId']}
                dictlist.append(dict1)
            i['managerlist']=dictlist
            
            if lSerializer.data['leavetype'] == 'Fullday':
                i['leavetype']="Full Day"
            elif lSerializer.data['leavetype'] =="SecondHalf":
                i['leavetype']="Second Half "
            elif lSerializer.data['leavetype'] =="FirstHalf":
                i['leavetype']="First Half "


            if(i['start_date'] == i['end_date']):
                i['leaveduration']=i['start_date']
                
            else:
                i['leaveduration']=i['start_date'] +" to "+ i['end_date']


            filterleaveobj=leaveApproval.objects.filter(leave_id=i['leave_id']).first()
            previousManager=leaveApproval.objects.filter(leave_id=i['leave_id'],rejectedBy=True).exclude(managerId=loginuserID).order_by('-id').first()
            approveManager = leaveApproval.objects.filter(leave_id=i['leave_id'],approvedBy=True).exclude(managerId=loginuserID).order_by('-id').first()
        
            mappingobj=leaveApproval.objects.filter(leave_id=i['leave_id']).distinct("managerId")
            elserializer=leaveapprovalserializer(mappingobj,many=True)
            llist=[]
            wcount=1    

            for t in elserializer.data:
            
                leaveobjone=leaveApproval.objects.filter(managerId=t['managerId'],leave_id=i['leave_id']).first()
                userObj=Users.objects.filter(id=int(t['managerId'])).first()
                
                if userObj.Photo is not None and userObj.Photo !="":
                    managerpic = imageUrl +"/media/"+ str(userObj.Photo)
                else:
                    managerpic = imageUrl + "/static/assets/images/profile.png"
                if leaveobjone is not None:
                    if leaveobjone.approvedBy == True:
                    
                        data={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":leaveobjone.comment,
                            "managerid":leaveobjone.managerId,
                            "status":"Approved",
                            "Photo":managerpic,
                            "position":wcount,
                            
                        }
                        llist.append(data)
                    elif leaveobjone.rejectedBy == True:

                        data={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":leaveobjone.comment,
                            "managerid":leaveobjone.managerId,
                            "status":"Rejected",
                            "Photo":managerpic,
                            "position":wcount,
                        }
                        llist.append(data)
                    else:
                        data={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":"",
                            "managerid":leaveobjone.managerId,
                            "status":"Pending",
                            "Photo":managerpic,
                            "position":wcount,
                        }
                        llist.append(data)

                wcount+=1
            i['allleavedata']=llist
            
            
            if previousManager is not None:
                i['passedBy']=previousManager.managerId
                userObj=Users.objects.filter(id=previousManager.managerId).first()
                i['passedByName'] =userObj.Firstname+" "+userObj.Lastname

            else:
                i['passedBy']="-----------"
                i['passedByName'] ="--"

            if approveManager is not None:
                i['approvedpassBy']=approveManager.managerId
                userObj=Users.objects.filter(id=approveManager.managerId).first()
                i['apprvpassedByName'] =userObj.Firstname+" "+userObj.Lastname

            else:
                i['approvedpassBy']="-----------"
                i['apprvpassedByName'] ="--"
            
            
            
            if lSerializer.data['leave_status'] =="Withdraw":
                mid_withdraw_list.append(i)
           
           
           
    withdrawlist=[] 
    for i in mid_withdraw_list:
        i['discard']=False 
        if i['allleavedata']: 
            for j in i['allleavedata']:
                if j['rejectedBy'] == True:
                    i['discard']=True
            
             
    for i in mid_withdraw_list:
        if i['discard']!=True:
            if i not in withdrawlist:
                withdrawlist.append(i)








    employee_name=request.POST.get('name')
    ApplicationType=request.POST.get('ApplicationType')
    if request.POST.get('department_value') != "" and request.POST.get('department_value') is not None:


        department_value=[int(request.POST.get('department_value'))]
    else:
        department_value=[]




    pendingleaves=[]
    Approvedleaves=[]
    rejectedleaves=[]
    expiredleaves=[]
    withdrawleaves=[]

    def filterdef(applicationlist, employee_name, ApplicationType, department):
        newlist = []

        for leave in applicationlist:
            applicant_name = leave.get('applicant_name', '').lower()
            leave_application_type = leave.get('ApplicationType', '').lower()
            leave_department = leave.get('department', '')
            if employee_name and ApplicationType and department:
                if employee_name.lower() in applicant_name and ApplicationType.lower() in leave_application_type and department == leave_department:
                    newlist.append(leave)
            elif employee_name and ApplicationType:
                if employee_name.lower() in applicant_name and ApplicationType.lower() in leave_application_type:
                    newlist.append(leave)
            elif employee_name and department:
                if employee_name.lower() in applicant_name and department == leave_department:
                    newlist.append(leave)
            elif ApplicationType and department:
                if ApplicationType.lower() in leave_application_type and department == leave_department:
                    newlist.append(leave)
            elif employee_name:
                if employee_name.lower() in applicant_name:
                    newlist.append(leave)
            elif ApplicationType:
                if ApplicationType.lower() in leave_application_type:
                    newlist.append(leave)
            elif department:
                if department == leave_department:
                    newlist.append(leave)
            else:
                newlist.append(leave)

        return newlist



    pendingleaves=filterdef(pendinglist,employee_name,ApplicationType,department_value)
    Approvedleaves=filterdef(apprvedleavelist,employee_name,ApplicationType,department_value)
    rejectedleaves=filterdef(rejectedlist,employee_name,ApplicationType,department_value)
    expiredleaves=filterdef(expiredlist,employee_name,ApplicationType,department_value)
    withdrawleaves=filterdef(withdrawlist,employee_name,ApplicationType,department_value)

    allapplicationslist = pendingleaves+expiredleaves+Approvedleaves+rejectedleaves+withdrawleaves
    return Response ({"allapplicationslist":allapplicationslist,"withdrawlist":withdrawleaves,"expiredlist":expiredleaves,"pendinglist":pendingleaves,"approvedlist":Approvedleaves,"rejectedlist":rejectedleaves,"response":{"n" : 1,"msg" : "SUCCESS","status" : "success"}})
   
@api_view(['POST'])
def withdraw_application(request):
    data={}
    data['id']=request.data.get('id')
    leave_id = data['id']
    company_code=request.user.company_code
    leavedata = Leave.objects.filter(id=data['id']).first()
    data['leave_status']='Withdraw'
    check_access=is_leave_time_valid(str(leavedata.start_date))
    if check_access:
        serializer = leaveserializer(leavedata,data=data,partial=True)
        if serializer.is_valid():
            serializer.save()
            employee_obj = Users.objects.filter(id=int(serializer.data['employeeId']),is_active=True).first()
            empname=employee_obj.Firstname + " " + employee_obj.Lastname
            empname=empname.title()

            workfromhomest = serializer.data['WorkFromHome']

            dates=date_handling(str(serializer.data['start_date']),str(serializer.data['end_date']))
            current_date = datetime.now().strftime('%Y-%m-%d')

            if workfromhomest == True:
                firebasenotfmsg =   empname + " has Withdrawn Work From Home Application"
                notfmsg = "<span class='actionuser'>" +empname + "</span> has Withdrawn Work From Home Application"
            else:
                firebasenotfmsg =  empname + " has Withdrawn Leave Application"
                notfmsg = "<span class='actionuser'>" + empname + "</span> has Withdrawn Work From Home Application"
        
            Team_members=[adminemail,hremail]

            leave_approval_object=leaveApproval.objects.filter(employeeId=serializer.data['employeeId'],leave_id=serializer.data['id'],company_code=serializer.data['company_code'])

            leave_approval_serializer = leaveapprovalserializer(leave_approval_object,many=True)

            TaskNotification.objects.filter(leaveID=leave_id,To_manager=True,action_Taken=False).update(action_Taken=True)

            for manager in leave_approval_serializer.data:
                m_obj = Users.objects.filter(id=int(manager['managerId']),is_active=True).first()
                if m_obj is not None:
                    Team_members.append(m_obj.email)

                    TaskNotification.objects.create(UserID_id=int(manager['managerId']),To_manager=True,leaveID=leave_id,company_code=company_code,NotificationTypeId_id=3,NotificationTitle="Leave",NotificationMsg=notfmsg,action_Taken=None)

                    manageridobj = Users.objects.filter(id=int(manager['managerId'])).first()
                    if manageridobj is not None:
                        firebasemsg =  firebasenotfmsg
                        fcmtoken = manageridobj.FirebaseID
                        notftype = "Leave"
                        fcmleaveid = leave_id
                        fcmtomanager = True
                        desktoptoken = manageridobj.desktopToken 
                        # desktopnotf = senddesktopnotf(desktoptoken,firebasemsg)
                        if fcmtoken is not None and fcmtoken != "":
                            firebasenotification = ""
                            # firebasenotification = sendfirebasenotification(fcmtoken,firebasemsg,notftype,fcmleaveid,fcmtomanager)
                        else:
                            firebasenotification = ""

                         
                
            try:             
                dicti = {
                    "applicationid":serializer.data['ApplicationId'],
                    "employeename":empname,
                    "dates":dates,
                    "withdrawn_date":convertdate(str(current_date)),
                    "startdate":convertdate(serializer.data['start_date']),
                    "enddate":convertdate(serializer.data['end_date']),
                    "type":str(wfhsorter(serializer.data['id'])),
                    "reason":serializer.data['reason'],

                }
                
                message = get_template(
                    'withdraw_application.html').render(dicti)
                msg = EmailMessage(
                    str(empname) +'  has withdrawn '+ str(WorkFromHome(serializer.data['id'])) +' application ',
                    message,
                    EMAIL_HOST_USER,
                    Team_members,
                )
                msg.content_subtype = "html"  # Main content is now text/html
                msg.send()
            except Exception as e:
                print('e2xception occured fot mail', e)
                
            return Response ({"data":serializer.data,"response":{"n" : 1,"msg" : "withdrew successfully","status" : "success"}})
        else:
            return Response ({"data":{},"response":{"n" : 0,"msg" : serializer.errors,"status" : "error"}})
    else:
        return Response ({"data":{},"response":{"n" : 0,"msg" : "The withdrawal time has passed. You cannot withdraw your application","status" : "error"}})
    
@api_view(['POST'])
def deleteleaveAPI(request):
    data={}
    data['id']=request.data.get('id')
   
    leavedata = Leave.objects.filter(id=data['id']).first()
    data['Active']=False
    serializer = leaveserializer(leavedata,data=data,partial=True)
    if serializer.is_valid():
        serializer.save()
        Leaveobjs = leaveApproval.objects.filter(leave_id=data['id'])
        if Leaveobjs.exists():
           Leaveobjs.delete()
        TaskNotification.objects.filter(leaveID = data['id']).delete()
        return Response ({"data":serializer.data,"response":{"n" : 1,"msg" : "Deleted successfully","status" : "success"}})
    else:
        return Response ({"data":serializer.errors,"response":{"n" : 0,"msg" : "failed","status" : "failed"}})

@api_view(['GET'])
def getbyidleaveAPI(request,id):
    leavedata = Leave.objects.filter(id=id,Active=True)
    if leavedata:
        serializer=leaveserializer(leavedata,many=True)
        for s in serializer.data:
            
            userid = s['employeeId']
            userobj = Users.objects.filter(id=int(userid),is_active=True).first()
            
            if userobj is not None:
                s['username'] = userobj.Firstname + " " + userobj.Lastname
                if str(userobj.Photo) is not None and str(userobj.Photo) != "":
                    s['userimage'] = imageUrl +"/media/"+ str(userobj.Photo)
                else:
                    s['userimage'] = imageUrl + "/static/assets/images/profile.png"
            else:
                s['username'] = "--"
                s['userimage'] = imageUrl + "/static/assets/images/profile.png"

            leaveactionobj = leaveApproval.objects.filter(leave_id=s['id'],employeeId=str(userid))
            if leaveactionobj is not None:
                leaveactionser = leaveapprovalserializer(leaveactionobj,many=True)
                managerlist = []
                for l in leaveactionser.data:
                    managerobj={}
                    managerobj['managerId'] = l['managerId']
                    manager = Users.objects.filter(id=int(l['managerId'])).first()
                    managerobj['managername'] = manager.Firstname+" "+manager.Lastname
                    if l['approvedBy'] == True:
                        managerobj['action'] = "Approved"
                        managerobj['reason'] = ""
                    elif l['rejectedBy'] == True:
                        managerobj['action'] = "Rejected"
                        managerobj['reason'] = l['comment']
                    else:
                        managerobj['action'] = "Pending"
                        managerobj['reason'] = ""

                    managerlist.append(managerobj)

                s['manageraction'] = managerlist
            else:
                s['manageraction'] = []

        return Response ({"data":serializer.data,"response":{"n" : 1,"msg" : "SUCCESS","status" : "success"}})
    else:
        return Response ({"data":'',"response":{"n" : 0,"msg" : "leave not found","status" : "failed"}})

@api_view(['POST'])
def updatebyidleaveAPI(request,id):
    data={}
    leavedata = Leave.objects.filter(id=id,Active=True).first()
    data['start_date']=request.data.get('start_date')
    data['end_date']=request.data.get('end_date')
    data['reason'] = request.data.get('reason')
    data['leavetype'] = request.data.get('leavetype')
    
    if request.FILES.get('attachment') is not None and request.FILES.get('attachment') !='': 
        data['attachment'] = request.FILES.get('attachment')
        
    data['number_of_days']=calculate_days(data['start_date'],data['end_date'],data['leavetype'])

    serializer = leaveserializer(leavedata,data=data,partial=True)

    


    
    if is_weeklyoff(data['start_date']) != None:
        return Response ({"data":[],"response":{"n" : 0,"msg" : "Start date should not on holidays","status" : "warning"}})
    elif checkholiday(data['start_date']) != None:
        return Response ({"data":[],"response":{"n" : 0,"msg" : "Start date should not on holidays","status" : "warning"}})


    if(data['start_date']==data['end_date']):
        weekoff_obj=is_weeklyoff(data['start_date'])
        if weekoff_obj !=None:
            return Response(weekoff_obj)
        if checkholiday(data['start_date']) != None:
            return Response ({"data":[],"response":{"n" : 0,"msg" : "Applied leaves dates are on holidays","status" : "warning"}})

    else:
        checkvaliddate=[]
        daterangelist=date_range_list(data['start_date'],data['end_date'])
        for i in daterangelist:
            if is_weeklyoff(str(i)) != None:
                checkvaliddate.append(False)
            elif checkholiday(i) != None:
                checkvaliddate.append(False)
            else: 
                checkvaliddate.append(True)
        if sum(checkvaliddate) <1:
            return Response ({"data":[],"response":{"n" : 0,"msg" : "Applied leaves dates are on holidays","status" : "warning"}})



    def check_leaves(sdate,edate,wfh,empid):
        applied_leaves_dates=date_range_list(sdate,edate)
        existing_leaves_dates=[]

        if wfh==True:
            recordexist = Leave.objects.filter(employeeId=empid,Active=True).exclude(id=id)
        else:
            recordexist = Leave.objects.filter(employeeId=empid,Active=True,WorkFromHome=False).exclude(id=id)

        recordserializer=leaveserializer(recordexist,many=True) 
        for i in recordserializer.data:
            if i['leave_status'] !="Draft" and i['leave_status'] !="Withdraw":
                exist_date_obj=date_range_list(i['start_date'],i['end_date']) 
                for j in exist_date_obj:
                    existing_leaves_dates.append(j)
        set1 = set(applied_leaves_dates)
        set2 = set(existing_leaves_dates)
        common_dates = set1.intersection(set2)
        for common_date in common_dates:
            if common_date is not None:
                return ({"data":[],"response":{"n" : 0,"msg" : "Request of date "+dateformat_ddmmyy(str(common_date))+" Already exist","status" : "warning"}})
            







    output_check=check_leaves(data['start_date'],data['end_date'],leavedata.WorkFromHome,leavedata.employeeId) 
    if output_check != None:
        return Response (output_check)
    

    if serializer.is_valid():
        serializer.save()

        userempname=""
        userempname_obj=Users.objects.filter(id=serializer.data['employeeId']).first()
        if userempname_obj:
            userempname=str(userempname_obj.Firstname) +" "+ str(userempname_obj.Lastname)
        userempname=userempname.title()

        Team_members=[adminemail,hremail]
        leave_approval_object=leaveApproval.objects.filter(employeeId=serializer.data['employeeId'],leave_id=serializer.data['id'],company_code=serializer.data['company_code'])
        leave_approval_serializer = leaveapprovalserializer(leave_approval_object,many=True)
        for manager in leave_approval_serializer.data:
            m_obj = Users.objects.filter(id=int(manager['managerId']),is_active=True).first()
            if m_obj is not None:
                Team_members.append(m_obj.email)
        try:             
            dicti = {
                        "employeename":userempname,
                        "dates":date_handling(serializer.data['start_date'],serializer.data['end_date']),
                        "startdate":convertdate(serializer.data['start_date']),
                        "enddate":convertdate(serializer.data['end_date']),
                        "reason":serializer.data['reason'],
                        "type":wfhsorter(serializer.data['id']),
                        "applicationtype":str(WorkFromHome(serializer.data['id'])),
                        "applicationid":serializer.data['ApplicationId'],

                    }
            message = get_template(
                'applicationupdatedmail.html').render(dicti)
            
            msg = EmailMessage(
                str(userempname) +' has updated their '+ str(WorkFromHome(serializer.data['id'])) + " application",
                message,
                EMAIL_HOST_USER,
                Team_members,
            )
            msg.content_subtype = "html"  # Main content is now text/html
            msg.send()
        except Exception as e:
            print('exception occured  mail', e)

        leavedeatil = []
        for s in [serializer.data]:
            stdate = str(s['start_date'])
            smonth_name = calendar.month_abbr[int(stdate.split('-')[1])]    
            sdatestr = stdate.split('-')[2]+" "+smonth_name
            s['strstartdate'] = sdatestr

            edate = str(s['end_date'])
            emonth_name = calendar.month_abbr[int(edate.split('-')[1])]    
            edatestr = edate.split('-')[2]+" "+emonth_name
            s['strenddate'] = edatestr
            leavedeatil.append(s)
        
        return Response ({"data":leavedeatil,"response":{"n" : 1,"msg" : "Application updated successfully","status" : "success"}})
    else:
        return Response ({"data":serializer.errors,"response":{"n" : 0,"msg" : "Failed to Update","status" : "warning"}})




@api_view(['POST'])
def update_leave_api(request):
    data={}
    id=request.data.get('id')
    userempname=""
    userempname_obj=Users.objects.filter(id=str(request.user.id)).first()
    if userempname_obj:
        userempname=str(userempname_obj.Firstname) +" "+ str(userempname_obj.Lastname)
    userempname=userempname.title()
    print("request.data",request.data)
    EmployeeId=request.user.id
    company_code=request.user.company_code
    leavedata = Leave.objects.filter(id=id,Active=True).first()

    if leavedata is not None:
        AttachmentStatus=False
        if request.FILES.get('attachment') is not None and request.FILES.get('attachment') !='': 
            data['attachment'] = request.FILES.get('attachment') 
            AttachmentStatus=True


        StartDate=request.data.get('start_date')
        EndDate=request.data.get('end_date')
        StartDayLeaveType=request.data.get('startdayleavetype')
        EndDayLeaveType=request.data.get('enddayleavetype')
        LeaveTypeId=request.data.get('LeaveTypeId')
        data['number_of_days']=calculate_total_leave_days(StartDate, StartDayLeaveType, EndDate, EndDayLeaveType)
        data['leave_status']=request.data.get('leavestatus')
        data['start_date']=StartDate
        data['end_date']=EndDate
        data['startdayleavetype']=StartDayLeaveType
        data['enddayleavetype']=EndDayLeaveType
        data['LeaveTypeId']=LeaveTypeId
        data['reason']=request.data.get('reason')
        serializer = leaveserializer(leavedata,data=data,partial=True)
        if serializer.is_valid():
            serializer.save()



            Team_members=[adminemail,hremail]
            leave_approval_object=leaveApproval.objects.filter(employeeId=EmployeeId,leave_id=serializer.data['id'],company_code=company_code)
            leave_approval_serializer = leaveapprovalserializer(leave_approval_object,many=True)
            for manager in leave_approval_serializer.data:
                m_obj = Users.objects.filter(id=int(manager['managerId']),is_active=True).first()
                if m_obj is not None:
                    Team_members.append(m_obj.email)

            try:             
                dicti = {
                            "employeename":userempname,
                            "dates":date_handling(serializer.data['start_date'],serializer.data['end_date']),
                            "startdate":convertdate(serializer.data['start_date']),
                            "enddate":convertdate(serializer.data['end_date']),
                            "reason":serializer.data['reason'],
                            "type":wfhsorter(serializer.data['id']),
                            "applicationtype":str(WorkFromHome(serializer.data['id'])),
                            "applicationid":serializer.data['ApplicationId'],

                        }
                message = get_template(
                    'applicationupdatedmail.html').render(dicti)
                
                msg = EmailMessage(
                    str(userempname) +' has updated their '+ str(WorkFromHome(serializer.data['id'])) + " application",
                    message,
                    EMAIL_HOST_USER,
                    Team_members,
                )
                msg.content_subtype = "html"  # Main content is now text/html
                msg.send()
            except Exception as e:
                print('exception occured  mail', e)


            
            return Response ({"data":{},"response":{"n" : 1,"msg" : "Application updated successfully","status" : "success"}})
        else:
            return Response ({"data":serializer.errors,"response":{"n" : 0,"msg" : "Failed to Update","status" : "warning"}})
    else:
        return Response ({"data":serializer.errors,"response":{"n" : 0,"msg" : "Failed to Update","status" : "warning"}})

@api_view(['POST'])
def updatedrafyleaveAPI(request,id):
    data={}
    leavedata = Leave.objects.filter(id=id,Active=True).first()
    data['start_date']=request.data.get('start_date')
    data['end_date']=request.data.get('end_date')
    data['reason'] = request.data.get('reason')
    data['leavetype']= request.data.get('leavetype')
    data['leave_status']= request.data.get('leavestatus')
    
    if request.FILES.get('attachment') is not None and request.FILES.get('attachment') !='': 
        data['attachment'] = request.FILES.get('attachment')
        
    data['number_of_days']=calculate_days(data['start_date'],data['end_date'],data['leavetype'])

    serializer = leaveserializer(leavedata,data=data,partial=True)


    
    
    if is_weeklyoff(data['start_date']) != None:
        return Response ({"data":'',"response":{"n" : 0,"msg" : "Start date should not on holidays","status" : "warning"}})
    elif checkholiday(data['start_date']) != None:
        return Response ({"data":'',"response":{"n" : 0,"msg" : "Start date should not on holidays","status" : "warning"}})


    if(data['start_date']==data['end_date']):
        weekoff_obj=is_weeklyoff(data['start_date'])
        if weekoff_obj !=None:
            return Response(weekoff_obj)
        if checkholiday(data['start_date']) != None:
            return Response ({"data":'',"response":{"n" : 0,"msg" : "Applied leaves dates are on holidays","status" : "warning"}})

    else:
        checkvaliddate=[]
        daterangelist=date_range_list(data['start_date'],data['end_date'])
        for i in daterangelist:
            if is_weeklyoff(str(i)) != None:
                checkvaliddate.append(False)
            elif checkholiday(i) != None:
                checkvaliddate.append(False)
            else: 
                checkvaliddate.append(True)
        if sum(checkvaliddate) <1:
            return Response ({"data":'',"response":{"n" : 0,"msg" : "Applied leaves dates are on holidays","status" : "warning"}})



    def check_leaves(sdate,edate,wfh,empid):
        applied_leaves_dates=date_range_list(sdate,edate)
        existing_leaves_dates=[]
        if wfh==True:
            recordexist = Leave.objects.filter(employeeId=empid,Active=True).exclude(id=id)
        else:
            recordexist = Leave.objects.filter(employeeId=empid,Active=True,WorkFromHome=False).exclude(id=id)
        recordserializer=leaveserializer(recordexist,many=True) 
        for i in recordserializer.data:
            if i['leave_status'] !="Draft" and i['leave_status'] !="Withdraw":
                exist_date_obj=date_range_list(i['start_date'],i['end_date']) 
                for j in exist_date_obj:
                    existing_leaves_dates.append(j)
        set1 = set(applied_leaves_dates)
        set2 = set(existing_leaves_dates)
        common_dates = set1.intersection(set2)
        for common_date in common_dates:
            if common_date is not None:
                return ({"data":'',"response":{"n" : 0,"msg" : "Request of date "+dateformat_ddmmyy(str(common_date))+" Already exist","status" : "warning"}})
            



    output_check=check_leaves(data['start_date'],data['end_date'],leavedata.WorkFromHome,leavedata.employeeId) 
    if output_check != None:
        return Response (output_check)
    

    if serializer.is_valid():
        serializer.save()
        leavedeatil = []
        for s in [serializer.data]:
            stdate = str(s['start_date'])
            smonth_name = calendar.month_abbr[int(stdate.split('-')[1])]    
            sdatestr = stdate.split('-')[2]+" "+smonth_name
            s['strstartdate'] = sdatestr

            edate = str(s['end_date'])
            emonth_name = calendar.month_abbr[int(edate.split('-')[1])]    
            edatestr = edate.split('-')[2]+" "+emonth_name
            s['strenddate'] = edatestr
            leavedeatil.append(s)
            

        return Response ({"data":leavedeatil,"response":{"n" : 1,"msg" : "Application updated successfully","status" : "success"}})
    else:
        return Response ({"data":serializer.errors,"response":{"n" : 0,"msg" : "serializer error occurs","status" : "failed"}})

@api_view(['POST'])
def leave_mapping_emp_filter(request, format=None):
    if request.method == 'POST':
        department=request.POST.get("department")
        name=request.POST.get("name")
        company_code = request.user.company_code

        if department !="" and department is not None:
            user = Users.objects.filter(Firstname__icontains=name,DepartmentID=department,is_active=True,company_code=company_code).order_by('id')
        else:
            user = Users.objects.filter(Firstname__icontains=name,is_active=True,company_code=company_code).order_by('id')
        
        
        serializer = UsersSerializer(user,many=True)
        userlist=[]
        for i in serializer.data:
                
            allemp = leaveMapping.objects.filter(employeeId=i['id'])
            leavemapingserializer = leaveMappingserializer(allemp,many=True)
            count =0

            filter1= leaveMapping.objects.filter(employeeId=i['id'],position="1",company_code=company_code).first()
       
            if filter1 is not None: 
                i['manager1']=int(filter1.managerId)
            else:
                i['manager1']="None" 
           

            filter2= leaveMapping.objects.filter(employeeId=i['id'],position="2",company_code=company_code).first()
          

            if filter2 is not None:
              
                i['manager2']=int(filter2.managerId)
            else:
                i['manager2']="None"

            filter3= leaveMapping.objects.filter(employeeId=i['id'],position="3",company_code=company_code).first()
           

            if filter3 is not None:
                i['manager3']=int(filter3.managerId)
            else:
                i['manager3']="None"


        return Response({'n':1,'count':len(userlist),'msg':'Employee list fetched successfully','status':'success','data':serializer.data})

@api_view(['POST'])
@permission_classes((AllowAny,))
def leave_history(request):
    LeaveId = request.POST.get('LeaveId')
    # EmployeeId = request.POST.get('EmployeeId')
    context = {}

    currentLeaveobj =  Leave.objects.filter(id=LeaveId,Active=True).first()
    if currentLeaveobj is not None:
        currentleaveser = leaveserializer(currentLeaveobj)
      
        context={
            'currentleavedata' :currentleaveser.data,
        }
        return Response({'n':1,'msg':'data found Successfully.','status':'success','data':context})
    else:
        return Response({'n':0,'msg':'Leave not found','status':'failed','data':"Leave Not Found"})
     
@api_view(['POST'])
def statusleaveAPI(request):
  
    data={}
    empuserID=request.user.id
    empid=Users.objects.filter(id=empuserID).first()
    employee_id=request.user.id
    data['company_code'] = request.user.company_code
    unique_id = empid.uid
    todays_date = date.today()
    year=todays_date.year
    data['id']=request.data.get('id')
  
    leavedata = Leave.objects.filter(employeeId=employee_id).order_by('-id').exclude(ApplicationId=None)
    leaves_obj_serializer=leaveserializer(leavedata,many=True)




    appnumberlist=[]
    for y in leaves_obj_serializer.data:
        number_obj = y['ApplicationId'].split("/")[3]
        appnumberlist.append(int(number_obj))


    if len(appnumberlist) < 1:
        Idnumber = "1"
        application_ID = str(year)+"/"+unique_id+"/"+Idnumber
        data['ApplicationId']=application_ID
    else:
        maxnumber=int(max(appnumberlist))
        maxnumber=int(maxnumber)+1

        
        Idnumber = str(maxnumber)
        application_ID =str(year)+"/"+unique_id+"/"+Idnumber
        data['ApplicationId']=application_ID

 


    leavedata = Leave.objects.filter(id=data['id'],Active=True).first()
    if leavedata.leave_status != "Draft":
        return Response ({"data":[],"response":{"n" : 0,"msg" : "Already applied application","status" : "warning"}})
    
    data['leave_status']="Pending"
    serializer = leaveserializer(leavedata,data=data,partial=True)
    startdate=str(leavedata.start_date)
    enddate=str(leavedata.end_date)
    specified_date = datetime.strptime(startdate, '%Y-%m-%d').date()
    today_date = date.today()

    is_greater = specified_date < today_date
    if is_greater:
        return Response ({"data":[],"response":{"n" : 0,"msg" : "Sorry you cannot apply past days applications","status" : "error"}})

       

    





    if is_weeklyoff(startdate) != None:
        return Response ({"data":[],"response":{"n" : 0,"msg" : "Start date should not on holidays","status" : "warning"}})
    elif checkholiday(startdate) != None:
        return Response ({"data":[],"response":{"n" : 0,"msg" : "Start date should not on holidays","status" : "warning"}})


    if(startdate==enddate):
        weekoff_obj=is_weeklyoff(startdate)
        if weekoff_obj !=None:
            return Response(weekoff_obj)
        if checkholiday(startdate) != None:
            return Response ({"data":[],"response":{"n" : 0,"msg" : "Applied application date are on holidays","status" : "warning"}})

    else:
        checkvaliddate=[]
        daterangelist=date_range_list(startdate,enddate)

        for i in daterangelist:
            if is_weeklyoff(str(i)) != None:
                checkvaliddate.append(False)
            elif checkholiday(i) != None:
                checkvaliddate.append(False)
            else: 
                checkvaliddate.append(True)

        if sum(checkvaliddate) <1:
            return Response ({"data":[],"response":{"n" : 0,"msg" : "Applied application dates are on holidays","status" : "warning"}})



    def check_leaves(sdate,edate,wfh,empid,id):
        applied_leaves_dates=date_range_list(sdate,edate)
        existing_leaves_dates=[]
        if wfh==True:
            recordexist = Leave.objects.filter(employeeId=empid,Active=True).exclude(id=id)
        else:
            recordexist = Leave.objects.filter(employeeId=empid,Active=True,WorkFromHome=False).exclude(id=id)

        recordserializer=leaveserializer(recordexist,many=True) 
        for i in recordserializer.data:
            if i['leave_status'] !="Draft" and i['leave_status'] !="Withdraw":
                exist_date_obj=date_range_list(i['start_date'],i['end_date']) 
                for j in exist_date_obj:
                    existing_leaves_dates.append(j)
        set1 = set(applied_leaves_dates)
        set2 = set(existing_leaves_dates)
        common_dates = set1.intersection(set2)
        for common_date in common_dates:
            if common_date is not None:
                return ({"data":[],"response":{"n" : 0,"msg" : "Request of date "+dateformat_ddmmyy(str(common_date))+" Already exist","status" : "warning"}})



    print("startdate,enddate,leavedata.WorkFromHome,leavedata.employeeId,data['id']",startdate,enddate,leavedata.WorkFromHome,leavedata.employeeId,data['id'])
    output_check=check_leaves(startdate,enddate,leavedata.WorkFromHome,leavedata.employeeId,data['id']) 
    if output_check != None:
        return Response (output_check)
    

   
    if serializer.is_valid():
        serializer.save()
        newleavelist = []
        for l in [serializer.data]:
            lstdate = str(l['start_date'])
            lsmonth_name = calendar.month_abbr[int(lstdate.split('-')[1])]    
            lsdatestr = lstdate.split('-')[2]+" "+lsmonth_name
            l['strstartdate'] = lsdatestr
            

            ledate = str(l['end_date'])
            lemonth_name = calendar.month_abbr[int(ledate.split('-')[1])]    
            ledatestr = ledate.split('-')[2]+" "+lemonth_name
            l['strenddate'] = ledatestr
            newleavelist.append(l)

        leaveid = serializer.data['id']
        Idapplication =serializer.data['ApplicationId']
        leavestatus =serializer.data['leave_status']
        workfromhomest=serializer.data['WorkFromHome']

        managerobj = leaveMapping.objects.filter(employeeId=empuserID)
        managerserializer=leaveMappingserializer(managerobj,many=True) 
        managerlist=[]

        empname = Users.objects.filter(id=empuserID,is_active=True).first()
        userempname = empname.Firstname + " " + empname.Lastname

        if workfromhomest == True:
            notificationmsg = "<span class='notfempname'>" + userempname + "</span> Applied for Work From Home"
        else:
            notificationmsg = "<span class='notfempname'>" + userempname + "</span> Applied for leave"

        if workfromhomest == True:
            firebasenotfmsg =   userempname + " Applied for Work From Home"
        else:
            firebasenotfmsg =  userempname + " Applied for leave"

        for m in managerserializer.data:            
            if m['managerId'] not in managerlist:
                managerlist.append(m['managerId'])
        if serializer.data['leave_status'] == "Pending":
            for manager_obj in managerlist:
                already_exist=leaveApproval.objects.filter(employeeId=empuserID,leave_id=serializer.data['id'],managerId=manager_obj,company_code=data['company_code']).first()
                if already_exist:
                    leaveApproval.objects.filter(employeeId=empuserID,leave_id=serializer.data['id'],managerId=manager_obj,company_code=data['company_code']).update(ApplicationId= serializer.data['ApplicationId'])
                else:
                    leaveApproval.objects.create(employeeId=empuserID,leave_id=serializer.data['id'],managerId=manager_obj,ApplicationId= serializer.data['ApplicationId'],company_code=data['company_code'])

                TaskNotification.objects.create(UserID_id=manager_obj,To_manager=True,leaveID=leaveid,company_code=data['company_code'],NotificationTypeId_id=3,NotificationTitle="Leave",NotificationMsg=notificationmsg)

                #firebasenotf for managers
                manageridobj = Users.objects.filter(id=manager_obj).first()
                if manageridobj is not None:
                    firebasemsg =  firebasenotfmsg
                    fcmtoken = manageridobj.FirebaseID
                    notftype = "Leave" 
                    fcmleaveid = leaveid
                    fcmtomanager = True
                    
                    desktoptoken = manageridobj.desktopToken 
                    # desktopnotf = senddesktopnotf(desktoptoken,firebasemsg)
                    
                    if fcmtoken is not None and fcmtoken != "":
                        firebasenotification = ""
                        # firebasenotification = sendfirebasenotification(fcmtoken,firebasemsg,notftype,fcmleaveid,fcmtomanager)
                    else:
                        firebasenotification = ""


            userempname=userempname.title()

            

         
            Team_members=[adminemail,hremail]

            leave_approval_object=leaveApproval.objects.filter(employeeId=empuserID,leave_id=leaveid,company_code=data['company_code'])
            leave_approval_serializer = leaveapprovalserializer(leave_approval_object,many=True)
            for manager in leave_approval_serializer.data:
                m_obj = Users.objects.filter(id=int(manager['managerId']),is_active=True).first()
                if m_obj is not None:
                    Team_members.append(m_obj.email)
            try:             
                dicti = {
                            "employeename":userempname,
                            "dates":date_handling(serializer.data['start_date'],serializer.data['end_date']),

                            "startdate":convertdate(serializer.data['start_date']),
                            "enddate":convertdate(serializer.data['end_date']),
                            "reason":serializer.data['reason'],
                            "type":wfhsorter(serializer.data['id'])
                        }
                message = get_template(
                    'leaveremindermail.html').render(dicti)
                msg = EmailMessage(
                    str(userempname) +' has applied for '+ str(WorkFromHome(serializer.data['id'])),
                    message,
                    EMAIL_HOST_USER,
                    Team_members,
                )
                msg.content_subtype = "html"  # Main content is now text/html
                msg.send()
            except Exception as e:
                print('exception occured fot mail', e)
                
                
        return Response ({"data":newleavelist,"response":{"n" : 1,"msg" :  str(WorkFromHome(serializer.data['id']))+" Applied Successfully.","status" : "success"}})
    else:
        return Response ({"data":serializer.errors,"response":{"n" : 0,"msg" : "Couldn't apply " + str(WorkFromHome(serializer.data['id'])),"status" : "failed"}})

@api_view(['POST'])
def leavemappingAPI(request):
    com_code = request.user.company_code
    ManagerIDlist = request.POST.getlist('managerids')
    employeeList = request.data.getlist('employeeId')
    
    for i in employeeList:
        deleteemp = leaveMapping.objects.filter(employeeId=i,company_code=com_code)
        deleteemp.delete()
        
        Position=1
        for manager in ManagerIDlist:
            leaveMapping.objects.create(employeeId=i,managerId=manager,position=Position,company_code=com_code)
            Position=Position+1
   
    usersdata =Users.objects.filter(company_code=com_code,is_blocked=False)
    userserializer = UserSerializer(usersdata,many=True)
    return Response ({"data":userserializer.data,"response":{"n" : 1,"msg" : "Mapping Added successfully","status" : "success"}})

@api_view(['GET'])
def leavemappingJoinQuery(request):
    id = request.query_params.get('managerId', None)
    if id is not None:
        userobjects = leaveMapping.objects.filter(employeeId=id)
        Serializer = leaveMappingserializer(userobjects, many=True)
        for i in Serializer.data:
            managerStr = Users.objects.filter(id=i['managerId']).first()
            i['ManagerStr'] = managerStr.Firstname + " " + managerStr.Lastname
        return JsonResponse({'n': 1, 'Msg': 'Data fetched successfully', 'Status': 'Success','data':Serializer.data}, safe=False)
        # users = Users.objects.filter(id__in =Subquery(userobjects.values('id')) )
    else:
        return Response({'n': 0, 'Msg': 'Manager ID value is None', 'Status': 'Failed'})

@api_view(['GET'])
def leavemappinglistAPI(request):
    company_code = request.user.company_code
    d6=[]
    managerlist=[]
    employeeId=[]
    MainList=[]
    allemp = leaveMapping.objects.filter(company_code=company_code)
    serializer = leaveMappingserializer(allemp,many=True)
    for i in serializer.data:
        if i['employeeId'] not in employeeId:
            employeeId.append(i['employeeId'])
       
    for p in employeeId:
        d={}
        data = Users.objects.filter(id=int(p)).first()
        if data is not None:
            d['id'] = data.id
            d['empname'] = data.Firstname +" "+ data.Lastname
            
            manager = leaveMapping.objects.filter(employeeId=str(p)) 
            mserilizer=leaveMappingserializer(manager,many=True)
            
            managerlist=[]
            for c in mserilizer.data:
                managerdata = Users.objects.filter(id=int(c['managerId'])).first()
                if managerdata is not None:
                    managerstr = str(managerdata.Firstname) +" "+ str(managerdata.Lastname)
                else:
                    managerstr = ""
                managerRank = {
                    'manager':managerstr,
                    'position':c['position']
                }
                managerlist.append(managerRank)
            d['managerList']=managerlist
            MainList.append(d)
    return Response ({"data":MainList,"response":{"n" : 1,"msg" : "SUCCESS","status" : "success"}})

@api_view(['GET'])
def mapped_managers(request, format=None):
    if request.method == 'GET':
        mapping = leaveMapping.objects.filter(company_code = request.user.company_code,employeeId = request.user.id).order_by('position')
        if mapping.exists():
            serializer = leaveMappingserializer(mapping, many=True)
            dummy_id_list=[]
            manager_list=[]
            for i in serializer.data:
                manager=Users.objects.filter(id=i['managerId']).first()
                manager_serializer=UserSerializer(manager)
                i['Firstname']=manager_serializer.data['Firstname']
                i['Lastname']=manager_serializer.data['Lastname']
                if i['managerId'] not in dummy_id_list:
                    dummy_id_list.append(i['managerId'])
                    manager_list.append(i)
            return Response({"data":manager_list,"n":1,"Msg":"List fetched successfully","Status":"Success"})
        else:
            return Response({"data":'',"n":0,"Msg":"List not found","Status":"Failed"})
       
@api_view(['GET'])
def leave_calculation(request, format=None):

    if request.method == 'GET':
        company_code = request.user.company_code
        employeeId = request.user.id
        company_leave_rule = Leaverule.objects.filter(company_code=company_code).first()
        if company_leave_rule:
            serializer = leaveruleserializer(company_leave_rule)
            period_of_leaves=serializer.data['Periodof_L'].lower().replace(" ", "")
            Assigned_leaves=serializer.data['Assignedleaves']
            Max_leaves=serializer.data['maxleaves']
            employee_leave_obj = Leave.objects.filter(company_code = company_code,employeeId = employeeId,leave_status="Approved")
            employee_leave_serializer = leaveserializer(employee_leave_obj,many=True)
            leave_multiplyer=0
            total_leave_days=0
            remainingleaves=0


            today = date.today()
            year = today.year
            month = today.month


            monthly_start_date=date.fromisoformat(str(datetime.date(year, month, 1)))
            monthly_end_date=date.fromisoformat(str(last_day_of_month(datetime.date(year, month, 1))))

            if month <4:
                yearly_start_date=date.fromisoformat(str(year-1)+"-04-01")
                yearly_end_date=date.fromisoformat(str(year)+"-03-31")
            else:
                yearly_start_date=date.fromisoformat(str(year)+"-04-01")
                yearly_end_date=date.fromisoformat(str(year+1)+"-03-31")

            for i in employee_leave_serializer.data:
                leavetype=i['leavetype'].lower()
                if leavetype =="firsthalf":
                    leave_multiplyer=0.5
                elif leavetype =="secondhalf":
                    leave_multiplyer=0.5
                elif leavetype =="fullday":
                    leave_multiplyer=1
                lsd =date.fromisoformat(i['start_date'])         
                led =date.fromisoformat(i['end_date'])   
              
                
                if period_of_leaves =="yearly": 

                    if lsd>=yearly_start_date and lsd<=yearly_end_date:
                        start_date =  arrow.get(i['start_date'])
                        end_date =  arrow.get(i['end_date'])
                        delta =(start_date-end_date)
                        tdays=abs(delta.days)
                    
                        total_leave_days+=leave_multiplyer * tdays
                        remainingleaves=Assigned_leaves-total_leave_days

                if period_of_leaves =="monthly": 
                    if lsd>=monthly_start_date and lsd<=monthly_end_date:
                        start_date =  arrow.get(i['start_date'])
                        end_date =  arrow.get(i['end_date'])
                        delta =(start_date-end_date)
                        tdays=abs(delta.days)
                       
                        total_leave_days+=leave_multiplyer * tdays
                        remainingleaves=Assigned_leaves-total_leave_days
 
          

            return Response({"data":"","n":1,"Msg":"List fetched successfully","Status":"Success"})
        else:
            return Response({"data":'',"n":0,"Msg":"List not found","Status":"Failed"})

@csrf_exempt
@api_view(['GET'])
def leaverequestAPI(request):


    rejectedlist=[]
    loginuserID = request.user.id
    current_date = datetime.now().strftime('%Y-%m-%d')
    leavepending = leaveApproval.objects.filter(managerId=loginuserID,approvedBy=False,rejectedBy=False).distinct("leave_id")
    pendingSerializer = leaveapprovalserializer(leavepending,many=True)
    pending_leaves_data=pendingSerializer.data
    final_pending_list=[]
    
    for i in pending_leaves_data:
        leavedata = Leave.objects.filter(id=i['leave_id'],start_date__gte=current_date,Active=True,).exclude(leave_status="Withdraw").first()
        lSerializer = leaveserializer(leavedata)

        if leavedata is not None:
            i['start_date'] = lSerializer.data['start_date']
            i['attachment'] = lSerializer.data['attachment']
            i['WorkFromHome'] = lSerializer.data['WorkFromHome']
            i['leave_status'] = lSerializer.data['leave_status']
            i['dd_month_sdate'] = dd_month_year_format(str(lSerializer.data['start_date']))
            i['dd_month_edate'] = dd_month_year_format(str(lSerializer.data['end_date']))
            i['leavetype'] = lSerializer.data['leavetype']
            i['end_date'] = lSerializer.data['end_date']
            datedata = lSerializer.data['created_at']
            i['created_at']=datedata.split('T')[0]
            i['reason']= lSerializer.data['reason']
            i['ApplicationId'] = lSerializer.data['ApplicationId']
            i['currentuser'] = loginuserID
            userObj=Users.objects.filter(id=int(lSerializer.data['employeeId'])).first()
            i['applicant_name'] = userObj.Firstname +" "+ userObj.Lastname
            start_date =  arrow.get(i['start_date'])
            end_date =  arrow.get(i['end_date'])
            delta =(start_date-end_date)
            tdays=abs(delta.days)+1
            i['total_days']=tdays
            i['created_at']=dd_month_year_format(i['created_at'])
            i['start_date']=convertdate(i['start_date'])
            i['end_date']=convertdate(i['end_date'])
            mappingobj=leaveApproval.objects.filter(leave_id=i['leave_id']).distinct("managerId")
            elserializer1=leaveapprovalserializer(mappingobj,many=True)
            llist=[]
            pcount=1
            for t in elserializer1.data:
                leaveobjone=leaveApproval.objects.filter(managerId=t['managerId'],leave_id=i['leave_id']).first()
                userObj=Users.objects.filter(id=int(t['managerId'])).first()
                if userObj.Photo is not None and userObj.Photo !="":
                    managerpic = imageUrl +"/media/"+ str(userObj.Photo)
                else:
                    managerpic = imageUrl + "/static/assets/images/profile.png"
                if leaveobjone is not None:
                    if leaveobjone.approvedBy == True: 
                    
                        data1={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":leaveobjone.comment,
                            "managerid":leaveobjone.managerId,
                            "Photo":managerpic,
                            "status":"Approved",
                            "position":pcount,

                        }
                        llist.append(data1)
                    elif leaveobjone.rejectedBy == True:
                    
                        data1={
                            "position":pcount,

                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":leaveobjone.comment,
                            "managerid":leaveobjone.managerId,
                            "Photo":managerpic,
                            "status":"Rejected",
                            

                        }
                        llist.append(data1)
                    else:
                        data1={
                            "position":pcount,
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":"",
                            "managerid":leaveobjone.managerId,
                            "Photo":managerpic,
                            "status":"Pending",
                            

                        }
                        llist.append(data1)                        
                
                
                pcount+=1
            i['allleavedata']=llist   
            
            if lSerializer.data['leave_status'] !="Withdraw":
                final_pending_list.append(i)

 
 
    for p in final_pending_list:
        p['discard']=False 
        if p['allleavedata'] is not None and p['allleavedata'] !="":
            for j in p['allleavedata']:
                if j['rejectedBy'] == True:
                    p['discard']=True
            
    pendinglist=[] 

    for i in final_pending_list:
        if i['discard']!=True:
            if i not in pendinglist:
                pendinglist.append(i)


            
                
    approvedlist=[]
    leaveapprovel = leaveApproval.objects.filter(managerId=loginuserID,approvedBy=True).distinct("leave_id")
    approvelSerializer = leaveapprovalserializer(leaveapprovel,many=True)
    for i in approvelSerializer.data: 
        leavedata = Leave.objects.filter(Q(id=i['leave_id']),Q(Active=True),Q(leave_status="Approved")|Q(leave_status="Pending")).exclude(leave_status="Withdraw").first()
        lSerializer = leaveserializer(leavedata)

        if leavedata:
            i['attachment'] = lSerializer.data['attachment']

            i['dd_month_sdate'] = dd_month_year_format(str(lSerializer.data['start_date']))
            i['leave_status'] = lSerializer.data['leave_status']


            i['dd_month_edate'] = dd_month_year_format(str(lSerializer.data['end_date']))
            i['leavetype'] = lSerializer.data['leavetype']
            i['WorkFromHome'] = lSerializer.data['WorkFromHome']
            i['start_date'] = lSerializer.data['start_date']
            i['end_date'] = lSerializer.data['end_date']
            if current_date <= i['start_date']:
                i['allow_reject'] = True
                i['allow_approve'] = True
                
            else:
                i['allow_approve'] = False
                i['allow_reject'] = False
                
            datedata = lSerializer.data['created_at']
            i['created_at']=datedata.split('T')[0]
            i['reason']= lSerializer.data['reason']
            i['ApplicationId'] = lSerializer.data['ApplicationId']
            userObj=Users.objects.filter(id=int(lSerializer.data['employeeId'])).first()
            i['applicant_name'] = userObj.Firstname +" "+ userObj.Lastname

            start_date =  arrow.get(i['start_date'])
            end_date =  arrow.get(i['end_date'])
            delta =(start_date-end_date)
            tdays=abs(delta.days)+1
            i['total_days']=tdays
            i['created_at']=dd_month_year_format(i['created_at'])
            i['start_date']=convertdate(i['start_date'])
            i['end_date']=convertdate(i['end_date'])

            mappingobj=leaveApproval.objects.filter(leave_id=i['leave_id']).distinct("managerId")
            elserializer=leaveapprovalserializer(mappingobj,many=True)
            llist=[]
            acount=1

            for t in elserializer.data:
                leaveobjone=leaveApproval.objects.filter(managerId=t['managerId'],leave_id=i['leave_id']).first()
                userObj=Users.objects.filter(id=int(t['managerId'])).first()
                if userObj.Photo is not None and userObj.Photo !="":
                    managerpic = imageUrl +"/media/"+ str(userObj.Photo)
                else:
                    managerpic = imageUrl + "/static/assets/images/profile.png"
                if leaveobjone is not None:
                    if leaveobjone.approvedBy == True: 
                    
                        data={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":leaveobjone.comment,
                            "managerid":leaveobjone.managerId,
                            "Photo":managerpic,
                            "status":"Approved",
                            "position":acount,
                        }
                        llist.append(data)
                    elif leaveobjone.rejectedBy == True:
                    
                        data={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":leaveobjone.comment,
                            "managerid":leaveobjone.managerId,
                            
                            "Photo":managerpic,
                            "status":"Rejected",
                            "position":acount,
                        }
                        llist.append(data)
                    else:
                        data={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":"",
                            "managerid":leaveobjone.managerId,
                            "Photo":managerpic,
                            "status":"Pending",
                            "position":acount,
                        }
                        llist.append(data)

                acount+=1

            i['allleavedata']=llist

            approvedlist.append(i)


    rejectedleavelist=[]
    leavereject = leaveApproval.objects.filter(managerId=loginuserID).distinct("leave_id")
    rejectedSerializer = leaveapprovalserializer(leavereject,many=True)
    for i in rejectedSerializer.data:
        i['discard']=False
        # ApplicationId = i['ApplicationId']
        leaves_obj = leaveApproval.objects.filter(leave_id=i['leave_id'])
        leaves_obj_Serializer = leaveapprovalserializer(leaves_obj,many=True)
        for l in leaves_obj_Serializer.data:
            if l['rejectedBy']==True:
                i['discard']=True



        rejectedleavedata = Leave.objects.filter(id=i['leave_id'],Active=True,leave_status="Rejected").first()
        lSerializer = leaveserializer(rejectedleavedata)
        if rejectedleavedata is not None:

            i['dd_month_sdate'] = dd_month_year_format(str(lSerializer.data['start_date']))
            i['dd_month_edate'] = dd_month_year_format(str(lSerializer.data['end_date']))
            i['leave_status'] = lSerializer.data['leave_status']
            i['attachment'] = lSerializer.data['attachment']

            i['leavetype'] = lSerializer.data['leavetype']
            i['WorkFromHome'] = lSerializer.data['WorkFromHome']
            i['start_date'] = lSerializer.data['start_date'] 
            i['end_date'] = lSerializer.data['end_date']
            datedata = lSerializer.data['created_at']
            i['created_at']=datedata.split('T')[0]
            i['reason']= lSerializer.data['reason']
            i['ApplicationId'] = lSerializer.data['ApplicationId']
            userObj=Users.objects.filter(id=int(lSerializer.data['employeeId'])).first()
            i['applicant_name'] = userObj.Firstname +" "+ userObj.Lastname

            
            start_date =  arrow.get(i['start_date'])
            end_date =  arrow.get(i['end_date'])
            delta =(start_date-end_date)
            tdays=abs(delta.days)+1
            i['total_days']=tdays
            i['created_at']=dd_month_year_format(i['created_at'])
            i['start_date']=convertdate(i['start_date'])
            i['end_date']=convertdate(i['end_date'])
            if current_date <= i['start_date']:
                i['allow_reject'] = True
                i['allow_approve'] = True
                
            else:
                i['allow_approve'] = False
                i['allow_reject'] = False
            llist=[]
            mappingobj=leaveApproval.objects.filter(leave_id=i['leave_id']).distinct("managerId")
            rejectedleaveserializer=leaveapprovalserializer(mappingobj,many=True)
            rcount=1
            managercount=mappingobj.count()
            rejectedcount=0
            for t in rejectedleaveserializer.data:
                rejleaveobjone=leaveApproval.objects.filter(managerId=t['managerId'],leave_id=i['leave_id']).first()
                userObj=Users.objects.filter(id=int(t['managerId'])).first()
                if userObj.Photo is not None and userObj.Photo !="":
                    managerpic = imageUrl +"/media/"+ str(userObj.Photo)
                else:
                    managerpic = imageUrl + "/static/assets/images/profile.png"
                if rejleaveobjone is not None:
                    if rejleaveobjone.approvedBy == True: 
                    
                        data={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":rejleaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":rejleaveobjone.rejectedBy,
                            "comment":rejleaveobjone.comment,
                            "managerid":rejleaveobjone.managerId,
                            "Photo":managerpic,
                            "status":"Approved",
                            "position":rcount,
                        }
                        llist.append(data)
                    elif rejleaveobjone.rejectedBy == True:
                        rejectedcount+=1
                        data={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":rejleaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":rejleaveobjone.rejectedBy,
                            "comment":rejleaveobjone.comment,
                            "managerid":rejleaveobjone.managerId,
                            "Photo":managerpic,
                            "status":"Rejected",
                            "position":rcount,
                        }
                        llist.append(data)
                    else:
                        data={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":"",
                            "managerid":rejleaveobjone.managerId,
                            "Photo":managerpic,
                            "status":"Pending",
                            "position":rcount,
                        }
                        llist.append(data)


                rcount+=1
            if rejectedcount == managercount:
                i['universel_comment']="No response to  application"
            else:
                i['universel_comment']=""
                
                
            i['allleavedata']=llist
            rejectedleavelist.append(i)


    duplicatelist=[]           
    for i in rejectedleavelist:
        if i['discard']==True:
            if i['ApplicationId'] not in duplicatelist:
                duplicatelist.append(i['ApplicationId'])
                rejectedlist.append(i)
            
 
 
 
 
#  expired list
    leaveexpired = leaveApproval.objects.filter(managerId=loginuserID,approvedBy=False,rejectedBy=False).order_by('-id')
    expiredSerializer = leaveapprovalserializer(leaveexpired,many=True)
    expired_leaves_data=expiredSerializer.data
    
    final_expired_list=[]
    
    for i in expired_leaves_data:
        leavedata = Leave.objects.filter(id=i['leave_id'],start_date__lt=current_date,Active=True,).exclude(Q(leave_status="Rejected")|Q(leave_status="Withdraw")).first()
        lSerializer = leaveserializer(leavedata)

        if leavedata is not None:
            i['attachment'] = lSerializer.data['attachment']

            i['start_date'] = lSerializer.data['start_date']
            i['WorkFromHome'] = lSerializer.data['WorkFromHome']
            i['leave_status'] = lSerializer.data['leave_status']
            i['dd_month_sdate'] = dd_month_year_format(str(lSerializer.data['start_date']))
            i['dd_month_edate'] = dd_month_year_format(str(lSerializer.data['end_date']))
            i['leavetype'] = lSerializer.data['leavetype']
            i['end_date'] = lSerializer.data['end_date']
            datedata = lSerializer.data['created_at']
            i['created_at']=datedata.split('T')[0]
            i['reason']= lSerializer.data['reason']
            i['ApplicationId'] = lSerializer.data['ApplicationId']
            i['currentuser'] = loginuserID
            userObj=Users.objects.filter(id=int(lSerializer.data['employeeId'])).first()
            i['applicant_name'] = userObj.Firstname +" "+ userObj.Lastname
            start_date =  arrow.get(i['start_date'])
            end_date =  arrow.get(i['end_date'])
            delta =(start_date-end_date)
            tdays=abs(delta.days)+1
            i['total_days']=tdays
            i['created_at']=dd_month_year_format(i['created_at'])
            i['start_date']=convertdate(i['start_date'])
            i['end_date']=convertdate(i['end_date'])
            mappingobj=leaveApproval.objects.filter(leave_id=i['leave_id']).distinct("managerId")
            elserializer1=leaveapprovalserializer(mappingobj,many=True)
            expired_llist=[]
            pcount=1
            for t in elserializer1.data:
                leaveobjone=leaveApproval.objects.filter(managerId=t['managerId'],leave_id=i['leave_id']).first()
                userObj=Users.objects.filter(id=int(t['managerId'])).first()
                if userObj.Photo is not None and userObj.Photo !="":
                    managerpic = imageUrl +"/media/"+ str(userObj.Photo)
                else:
                    managerpic = imageUrl + "/static/assets/images/profile.png"
                if leaveobjone is not None:
                    if leaveobjone.approvedBy == True: 
                    
                        data1={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":leaveobjone.comment,
                            "managerid":leaveobjone.managerId,
                            "Photo":managerpic,
                            "status":"Approved",
                            "position":pcount,

                        }
                        expired_llist.append(data1)
                    elif leaveobjone.rejectedBy == True:
                    
                        data1={
                            "position":pcount,

                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":leaveobjone.comment,
                            "managerid":leaveobjone.managerId,
                            "Photo":managerpic,
                            "status":"Rejected",
                            

                        }
                        expired_llist.append(data1)
                    else:
                        data1={
                            "position":pcount,
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":"",
                            "managerid":leaveobjone.managerId,
                            "Photo":managerpic,
                            "status":"Pending",
                            

                        }
                        expired_llist.append(data1)                        
                
                
                pcount+=1
            i['allleavedata']=expired_llist   


 
 
 
 
 
            
  
            final_expired_list.append(i)
                

 
 
 
    for p in final_expired_list:
        p['discard']=False 
        if p['allleavedata'] is not None and p['allleavedata'] !="":
            for j in p['allleavedata']:
                if j['rejectedBy'] == True:
                    p['discard']=True
            
    expiredlist=[] 

    for i in final_expired_list:
        if i['discard']!=True:
            if i not in expiredlist:
                expiredlist.append(i)
 
 
 
#  withdrawn list

    leavewithdraw = leaveApproval.objects.filter(managerId=loginuserID).exclude(rejectedBy=True).distinct("leave_id")
    withdrawSerializer = leaveapprovalserializer(leavewithdraw,many=True)
    withdraw_leaves_data=withdrawSerializer.data
    final_withdraw_list=[]
    
    for i in withdraw_leaves_data:
        leavedata = Leave.objects.filter(id=i['leave_id'],Active=True,leave_status="Withdraw").first()
        lSerializer = leaveserializer(leavedata)

        if leavedata is not None:
            i['attachment'] = lSerializer.data['attachment']

            i['leave_status'] = lSerializer.data['leave_status']
            i['start_date'] = lSerializer.data['start_date']
            i['WorkFromHome'] = lSerializer.data['WorkFromHome']
            i['dd_month_sdate'] = dd_month_year_format(str(lSerializer.data['start_date']))
            i['dd_month_edate'] = dd_month_year_format(str(lSerializer.data['end_date']))
            i['leavetype'] = lSerializer.data['leavetype']
            i['end_date'] = lSerializer.data['end_date']
            datedata = lSerializer.data['created_at']
            i['created_at']=datedata.split('T')[0]
            i['reason']= lSerializer.data['reason']
            i['ApplicationId'] = lSerializer.data['ApplicationId']
            i['currentuser'] = loginuserID
            userObj=Users.objects.filter(id=int(lSerializer.data['employeeId'])).first()
            i['applicant_name'] = userObj.Firstname +" "+ userObj.Lastname
            start_date =  arrow.get(i['start_date'])
            end_date =  arrow.get(i['end_date'])
            delta =(start_date-end_date)
            tdays=abs(delta.days)+1
            i['total_days']=tdays
            i['created_at']=dd_month_year_format(i['created_at'])
            i['start_date']=convertdate(i['start_date'])
            i['end_date']=convertdate(i['end_date'])


            

            filterleaveobj=leaveApproval.objects.filter(leave_id=i['leave_id']).first()


            mappingobj=leaveApproval.objects.filter(leave_id=i['leave_id']).distinct("managerId")
            elserializer1=leaveapprovalserializer(mappingobj,many=True)
            llist=[]
            
            wcount=1
            for t in elserializer1.data:
                leaveobjone=leaveApproval.objects.filter(managerId=t['managerId'],leave_id=i['leave_id']).first()
                userObj=Users.objects.filter(id=int(t['managerId'])).first()
                if userObj.Photo is not None and userObj.Photo !="":
                    managerpic = imageUrl +"/media/"+ str(userObj.Photo)
                else:
                    managerpic = imageUrl + "/static/assets/images/profile.png"
                if leaveobjone is not None:
                    if leaveobjone.approvedBy == True: 
                    
                        data1={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":leaveobjone.comment,
                            "managerid":leaveobjone.managerId,
                            "Photo":managerpic,
                            "status":"Approved",
                            "position":wcount,

                        }
                        llist.append(data1)
                    elif leaveobjone.rejectedBy == True:
                    
                        data1={
                            "position":wcount,

                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":leaveobjone.comment,
                            "managerid":leaveobjone.managerId,
                            "Photo":managerpic,
                            "status":"Rejected",
                            

                        }
                        llist.append(data1)
                    else:
                        data1={
                            "position":wcount,

                            "applicationId":i['ApplicationId'],
                            "approvedBy":leaveobjone.approvedBy,
                            "name":userObj.Firstname+" "+userObj.Lastname,
                            "rejectedBy":leaveobjone.rejectedBy,
                            "comment":"",
                            "managerid":leaveobjone.managerId,
                            "Photo":managerpic,
                            "status":"Pending",
                            

                        }
                        llist.append(data1)                        

                
                wcount+=1
            i['allleavedata']=llist   
            
            if lSerializer.data['leave_status'] =="Withdraw":
                final_withdraw_list.append(i)

 
 
    for p in final_withdraw_list:
        p['discard']=False 
        if p['allleavedata'] is not None and p['allleavedata'] !="":
            for j in p['allleavedata']:
                if j['rejectedBy'] == True:
                    p['discard']=True
            
    withdrawlist=[] 

    for i in final_withdraw_list:
        if i['discard']!=True:
            if i not in withdrawlist:
                withdrawlist.append(i)


    return Response ({"expiredlist":expiredlist,"withdrawlist":withdrawlist,"pendinglist":pendinglist,"approvedlist":approvedlist,"rejectedlist":rejectedlist,"response":{"n" : 1,"msg" : "SUCCESS","status" : "success"}})
    
@api_view(['POST'])
def leaveactionAPI(request):
    data={}
    id = request.data.get('id')
    comment = request.data.get('comment')
    com_code = request.user.company_code
    employee_name=""
    Leave_Status=""
    acted_user = request.user.id  
    manageruserobj = Users.objects.filter(id=int(acted_user)).first()
    managername = manageruserobj.Firstname + " " + manageruserobj.Lastname 

    action_conditional_obj = leaveApproval.objects.filter(id=id).first()
    if action_conditional_obj is not None:
        emp_obj = Users.objects.filter(id=int(action_conditional_obj.employeeId)).first()
        employee_name = emp_obj.Firstname + " " + emp_obj.Lastname 
        action_conditional_leave_obj= Leave.objects.filter(id=action_conditional_obj.leave_id,Active=True).first()
        if action_conditional_leave_obj is not None:
            if action_conditional_leave_obj.leave_status != "Pending":
                return Response ({"data":{},"response":{"n" : 2,"msg" : "application journey is completed unable to perform actions","status" : "failed"}})
            elif action_conditional_obj.approvedBy == True or action_conditional_obj.rejectedBy ==True:    
                return Response ({"data":{},"response":{"n" : 2,"msg" : "you have already  perform actions","status" : "failed"}})
    if 'comment' in request.data.keys():

        if comment is not None and comment != "":
            data['rejectedBy'] = True
            data['comment']= comment
            leaveapp = leaveApproval.objects.filter(id=id).first()
            if leaveapp:
                aserializer = leaveapprovalserializer(leaveapp,data=data,partial=True)
                if aserializer.is_valid():
                    aserializer.save()
                    Leave_Status="Rejected"

                    leaveid = aserializer.data['leave_id']
                    appstatus = WorkFromHome(leaveid)
                    
                    Leave.objects.filter(id=leaveid,Active=True).update(leave_status="Rejected")
                    TaskNotification.objects.filter(leaveID=leaveid,To_manager=True,UserID=request.user.id).update(action_Taken=True)
                    TaskNotification.objects.filter(leaveID=leaveid,To_manager=True).update(action_Taken=True)

                    empid = aserializer.data['employeeId']
                    applctnid = aserializer.data['ApplicationId']
                
                    userobj = Users.objects.filter(id=int(empid),is_active=True).first()
                    username = userobj.Firstname + " " + userobj.Lastname 

                    Leaveobj = Leave.objects.filter(id=leaveid,Active=True).first()
                    
                    leave_dates_handled=date_handling(str(Leaveobj.start_date),str(Leaveobj.end_date))
                    
                    startd = str(Leaveobj.start_date) 
                    startdateday = startd.split("-")[2]
                    startdateyear = startd.split("-")[0]
                    startdatemonth = month_converter(int(startd.split("-")[1]))
                    startdate = str(startdateday) +" "+ startdatemonth +" "+str(startdateyear)

                    endd = str(Leaveobj.end_date)
                    enddateday = endd.split("-")[2]
                    enddateyear = endd.split("-")[0]
                    enddatemonth = month_converter(int(endd.split("-")[1]))
                    enddate = str(enddateday) +" "+ enddatemonth +" "+str(enddateyear)

                    if startd == endd:
                        notfmsg = "You have rejected " +str(appstatus)+ " Application of  <span class='actionuser'>" + username + " </span> dated <span class='notfleavedate'>" +startdate+"</span> "
                        firebasemsg="You have rejected " +str(appstatus)+ " Application of  " + username + " dated " +startdate
                    else:
                        notfmsg = "You have rejected " +str(appstatus)+ " Application of <span class='actionuser'>" + username + " </span> dated <span class='notfleavedate'>" +startdate+" to "+enddate+"</span> "
                        firebasemsg="You have rejected " +str(appstatus)+ " Application of  " + username + " dated " +startdate+" to "+enddate


                    fcmto_acted_manager=True
                    notf_type = "Leave"
                    logined_user_fcmtoken = manageruserobj.FirebaseID
                    if logined_user_fcmtoken is not None and logined_user_fcmtoken != "":
                        firebasenotification = ""
                        # firebasenotification = sendfirebasenotification(logined_user_fcmtoken,firebasemsg,notf_type,leaveid,fcmto_acted_manager)
                    else:
                        firebasenotification = ""
                    
                    






                    # notf for manager after action
                    empleaverejectobj=TaskNotification.objects.create(UserID_id=int(request.user.id),leaveID=leaveid,To_manager=True,company_code=com_code,NotificationTypeId_id=3,action_Taken=None,NotificationTitle="Leave",NotificationMsg=notfmsg)
                    dummyleaveid=empleaverejectobj.id
                    
                    # notf for employee after action
                    notificationmsg = "Your " +str(appstatus)+ " request with Application Id <span class='notfappid'>" +applctnid+"</span> has been <span class='rejectedmsg'> Rejected </span>"

                    TaskNotification.objects.create(UserID_id=empid,company_code=com_code,NotificationTypeId_id=3,NotificationTitle="Leave Rejected",NotificationMsg=notificationmsg,leaveID=leaveid)

                    # notf for admin after reject action
                    if startd == endd:
                        adminnotfmsg = str(appstatus)+ " request of " +username+ " </span> dated <span class='notfleavedate'> " +startdate+" </span></span> has been <span class='rejectedmsg'> Rejected </span>"
                    else:
                        adminnotfmsg = str(appstatus)+ " request of " +username+" </span> dated <span class='notfleavedate'>" +startdate+" to "+enddate+" </span></span> has been <span class='rejectedmsg'> Rejected </span>"

                    ad_mail = adminemail
                    adminobj = Users.objects.filter(email=ad_mail).first()
                    adminid = adminobj.id

                    TaskNotification.objects.create(UserID_id=int(adminid),leaveID=leaveid,company_code=com_code,NotificationTypeId_id=3,NotificationTitle="Leave",NotificationMsg=adminnotfmsg)
                    #firebase for employee(reject)
                    firebaseempnotfmsg =  "Your " +str(appstatus)+ " request with Application Id "+applctnid+" has been Rejected"

                    userfidobj = Users.objects.filter(id=int(empid)).first()
                    if userfidobj is not None:
                        firebasemsg =  firebaseempnotfmsg
                        fcmtoken = userfidobj.FirebaseID
                        notftype = "Leave"
                        fcmleaveid  = leaveid
                        fcmtomanager = False
                        
                        desktoptoken = userfidobj.desktopToken 
                        # desktopnotf = senddesktopnotf(desktoptoken,firebasemsg)
                        
                        if fcmtoken is not None and fcmtoken != "":
                            firebasenotification = ""
                            # firebasenotification = sendfirebasenotification(fcmtoken,firebasemsg,notftype,fcmleaveid,fcmtomanager)
                        else:
                            firebasenotification = ""


                    #notf for other managers
                    notfobjects = TaskNotification.objects.filter(leaveID=leaveid,To_manager=True).distinct('UserID').exclude(UserID=acted_user)
                    
                    notfser = TaskNotificationSerializer(notfobjects,many=True)
                    for p in notfser.data:
                        managerLeaveobj = Leave.objects.filter(id=p['leaveID'],Active=True).first()
                        if managerLeaveobj is not None:
                            mangeruserid = managerLeaveobj.employeeId

                            muserobj = Users.objects.filter(id=int(mangeruserid),is_active=True).first()
                            musername = muserobj.Firstname + " " + muserobj.Lastname 

                            sdt1 = str(managerLeaveobj.start_date) 
                            sdtday = sdt1.split("-")[2]
                            sdtyear = sdt1.split("-")[0]
                            sdmonth = month_converter(int(sdt1.split("-")[1]))
                            mstartdate = str(sdtday) +" "+ sdmonth +" "+str(sdtyear)


                            edt1 = str(managerLeaveobj.end_date)
                            edtday = edt1.split("-")[2]
                            edtyear = edt1.split("-")[0]
                            edmonth = month_converter(int(edt1.split("-")[1]))
                            menddate = str(edtday) +" "+ edmonth +" "+str(edtyear)


                            if sdt1 == edt1 :
                                managernotfmsg = "<span class='managernotfname'>" +managername + " </span> has rejected "+str(appstatus)+" Application of <span class='actionuser'>" +musername + "</span> dated  <span class='notfleavedate'>" + mstartdate +"</span>"

                                firebasemmnotfmsg = managername + " has rejected "+str(appstatus)+" Application of "+musername+" dated  " +mstartdate 

                            else:
                                managernotfmsg = "<span class='managernotfname'>" +managername + " </span> has rejected "+str(appstatus)+" Application of <span class='actionuser'>" +username + "</span> dated  <span class='notfleavedate'>" + mstartdate +" to " + menddate + "</span>"

                                firebasemmnotfmsg = managername + " has rejected "+str(appstatus)+" Application of "+musername+" dated " + mstartdate +" to " + menddate
                        
                    
                            TaskNotification.objects.create(UserID_id=p['UserID'],company_code=com_code,To_manager=True,action_Taken=None,leaveID=leaveid,NotificationTypeId_id=3,NotificationTitle="Leave",NotificationMsg=managernotfmsg,created_by = int(request.user.id))

                            userfffidobj = Users.objects.filter(id=int(p['UserID'])).first()
                            if userfffidobj is not None:
                                firebasemsg =  firebasemmnotfmsg
                                fcmtoken = userfffidobj.FirebaseID
                                notftype = "Leave"
                                fcmleaveid = leaveid
                                fcmtomanager = True
                                desktoptoken = userfffidobj.desktopToken 
                                # desktopnotf = senddesktopnotf(desktoptoken,firebasemsg)
                                if fcmtoken is not None and fcmtoken != "":
                                    firebasenotification = ""
                                    # firebasenotification = sendfirebasenotification(fcmtoken,firebasemsg,notftype,fcmleaveid,fcmtomanager)
                                else:
                                    firebasenotification = ""




                    username=username.title()
                    leave_obj1=Leave.objects.filter(id=leaveid).first()
                    Team_members=[adminemail,hremail]
                    leave_approval_object=leaveApproval.objects.filter(employeeId=empid,leave_id=leaveid,company_code=com_code).exclude(managerId=acted_user)
                    leave_approval_serializer = leaveapprovalserializer(leave_approval_object,many=True)
                    for manager in leave_approval_serializer.data:
                        m_obj = Users.objects.filter(id=int(manager['managerId']),is_active=True).first()
                        if m_obj is not None:
                            Team_members.append(m_obj.email)
                    data_dict = {
                                    "employeename":username,
                                    "actedmanager":managername,
                                    "startdate":startdate,
                                    "enddate":enddate,
                                    "reason":aserializer.data['comment'],
                                    "WFH":leave_obj1.WorkFromHome,
                                    "dates":leave_dates_handled,
                                    "leave_status":str(leave_obj1.leave_status),
                                    "applicationid":str(leave_obj1.ApplicationId),
                                    "Action_status":"Rejected",
                                    "leave_reason":str(leave_obj1.reason),
                    }


                    # Send a welcome email with a specific template
                    send_async_custom_template_email(
                        str(managername) +' has rejected  '+str(WorkFromHome(leaveid)) +' application',
                        data_dict,
                        EMAIL_HOST_USER,
                        Team_members,
                        "leaveactionmail.html"
                    )
                    send_async_custom_template_email(
                        str(managername) +' has rejected your '+str(WorkFromHome(leaveid)) +' application',
                        data_dict,
                        EMAIL_HOST_USER,
                        [str(userobj.email)],
                        "leaveactionmailtoemployee.html"
                    )
                    
                    

            
            
            else:
                return Response ({"data":{},"response":{"n" : 2,"msg" : "Someting went wrong,try again later","status" : "error"}})
        else:
            return Response ({"data":{},"response":{"n" : 2,"msg" : "please provide rejection reason","status" : "error"}})

    else:
        leaveapp = leaveApproval.objects.filter(id=id).first()
        data['approvedBy'] = True
        if leaveapp:
            aserializer = leaveapprovalserializer(leaveapp,data=data,partial=True)
            if aserializer.is_valid():
                Leave_Status="Approved"

                aserializer.save()
                empid = aserializer.data['employeeId']
                appid = aserializer.data['ApplicationId']
                leaveid = aserializer.data['leave_id']
                appstatus1 = WorkFromHome(leaveid)
                TaskNotification.objects.filter(action_Taken=False,leaveID=leaveid,To_manager=True,UserID=request.user.id).update(action_Taken=True)
                userobj = Users.objects.filter(id=int(empid),is_active=True).first()
                username1 = userobj.Firstname + " " + userobj.Lastname 
                Leaveobj = Leave.objects.filter(id=leaveid,Active=True).first()
                startd = str(Leaveobj.start_date) 
                startdateday = startd.split("-")[2]
                startdateyear = startd.split("-")[0]
                startdatemonth = month_converter(int(startd.split("-")[1]))
                startdate = str(startdateday) +" "+ startdatemonth +" "+str(startdateyear)
                endd = str(Leaveobj.end_date)
                enddateday = endd.split("-")[2]
                enddateyear = endd.split("-")[0]
                enddatemonth = month_converter(int(endd.split("-")[1]))
                enddate = str(enddateday) +" "+ enddatemonth +" "+str(enddateyear)
                if startd == endd:
                    notfmsg = "You have approved  " +str(appstatus1)+ " Application of <span class='actionuser'>" + username1 + "</span> dated <span class='notfleavedate'>" +startdate+" </span>"
                    firebasemsg = "You have approved  " +str(appstatus1)+ " Application of " + username1 + " dated " +startdate+" "
                else:
                    notfmsg = "You have approved  " +str(appstatus1)+ " Application of <span class='actionuser'>" + username1 + "</span> dated <span class='notfleavedate'>" +startdate+" to "+enddate+" </span>"
                    firebasemsg = "You have approved  " +str(appstatus1)+ " Application of " + username1 + " dated " +startdate+" to "+enddate+" "
                TaskNotification.objects.create(UserID_id=int(request.user.id),To_manager=True,company_code=com_code,action_Taken=None,NotificationTypeId_id=3,NotificationTitle="Leave",leaveID=leaveid,NotificationMsg=notfmsg)

                
                
                fcmto_acted_manager=True
                notf_type = "Leave"
                logined_user_fcmtoken = manageruserobj.FirebaseID
                if logined_user_fcmtoken is not None and logined_user_fcmtoken != "":
                    firebasenotification = ""
                    # firebasenotification = sendfirebasenotification(logined_user_fcmtoken,firebasemsg,notf_type,leaveid,fcmto_acted_manager)
                else:
                    firebasenotification = ""

                
                
                
                
                
                
                check_approval=leaveApproval.objects.filter(employeeId=empid,ApplicationId=appid,leave_id=leaveid)
                check_approvel_serializer=leaveapprovalserializer(check_approval,many=True)
                approved_status=[]
                for i in check_approvel_serializer.data:
                    approved_status.append(i['approvedBy'])
                lengthofstatus=len(approved_status)
                lengthofapprovedstatus=sum(approved_status)
                if lengthofstatus == lengthofapprovedstatus:
                    Leave.objects.filter(id=leaveapp.leave_id,Active=True).update(leave_status="Approved")
                    notificationmsg2 = "Your "+str(appstatus1)+" request with Application Id <span class='notfappid'>" +appid+"</span> has been <span class='approvedmsg'> Approved </span>"
                    firebasemsg1 = "Your "+str(appstatus1)+" request with Application Id " +appid+" has been Approved " 
                    TaskNotification.objects.create(UserID_id=empid,company_code=com_code,NotificationTypeId_id=3,NotificationTitle="Leave Approved",leaveID=leaveid,NotificationMsg=notificationmsg2)
                    firebasemsg =  firebasemsg1
                    fcmtoken = userobj.FirebaseID
                    notftype = "Leave"
                    fcmleaveid = leaveid
                    fcmtomanager = False
                    
                    desktoptoken = userobj.desktopToken 
                    # desktopnotf = senddesktopnotf(desktoptoken,firebasemsg)
                    
                    if fcmtoken is not None and fcmtoken != "":
                        firebasenotification = ""
                        # firebasenotification = sendfirebasenotification(fcmtoken,firebasemsg,notftype,fcmleaveid,fcmtomanager)
                    else:
                        firebasenotification = ""
                    # notf for admin after approve action
                    if startd == endd:
                        adminnotfmsg = str(appstatus1)+ " request of " +username1+ " </span> dated <span class='notfleavedate'> " +startdate+" </span>has been <span class='rejectedmsg'> Approved </span>"
                    else:
                        adminnotfmsg = str(appstatus1)+ "request of " +username1+ " </span> dated <span class='notfleavedate'> " +startdate+" to "+enddate+" </span>has been <span class='rejectedmsg'> Approved </span>"
                    ad_mail = adminemail
                    adminobj = Users.objects.filter(email=ad_mail).first()
                    adminid = adminobj.id
                    TaskNotification.objects.create(UserID_id=int(adminid),leaveID=leaveid,company_code=com_code,NotificationTypeId_id=3,NotificationTitle="Leave",NotificationMsg=adminnotfmsg)



                leave_dates_handled=date_handling(str(Leaveobj.start_date),str(Leaveobj.end_date))
                username=username1.title()
                leave_obj1=Leave.objects.filter(id=leaveid).first()

                if leave_obj1.leave_status == "Pending":
                    notificationmsg2 = str(managername) +" has approved your application  request with Application Id <span class='notfappid'>" +appid+"</span>  application is partially approved "
                    TaskNotification.objects.create(UserID_id=empid,company_code=com_code,NotificationTypeId_id=3,To_manager=False,NotificationTitle="Leave Partially Approved",leaveID=leaveid,NotificationMsg=notificationmsg2)

                Team_members=[adminemail,hremail]
                leave_approval_object=leaveApproval.objects.filter(employeeId=empid,leave_id=leaveid,company_code=com_code,approvedBy=False,rejectedBy=False).exclude(managerId=acted_user)
                if len(leave_approval_object) !=0:
                    leave_approval_serializer = leaveapprovalserializer(leave_approval_object,many=True)
                else:
                    leave_approval_object_all=leaveApproval.objects.filter(employeeId=empid,leave_id=leaveid,company_code=com_code).exclude(managerId=acted_user)
                    leave_approval_serializer = leaveapprovalserializer(leave_approval_object_all,many=True)
                for manager in leave_approval_serializer.data:
                    managerid = int(manager['managerId'])
                    m_obj = Users.objects.filter(id=int(manager['managerId']),is_active=True).first()
                    if m_obj is not None:
                        Team_members.append(m_obj.email)
                for m in leave_approval_serializer.data:
                    managerid = int(m['managerId'])
                    mm_obj = Users.objects.filter(id=managerid,is_active=True).first()
                    if mm_obj is not None:
                        managernotfmsg1 = "<span class='managernotfname'>" +managername + " </span> has Approved "+str(appstatus1)+" Application of <span class='actionuser'>" +username1 + "</span> dated  <span class='notfleavedate'>" + startdate +" to " + enddate + "</span>"
                        firebasemmnotfmsg1 = managername + " has Approved "+str(appstatus1)+" Application of "+username1+" dated " + startdate +" to " + enddate
                        TaskNotification.objects.create(UserID_id=managerid,company_code=com_code,To_manager=True,action_Taken=None,NotificationTypeId_id=3,NotificationTitle="Leave Approved",leaveID=leaveid,NotificationMsg=managernotfmsg1)
                        firebasemsg =  firebasemmnotfmsg1
                        fcmtoken = mm_obj.FirebaseID
                        notftype = "Leave"
                        fcmleaveid = leaveid
                        fcmtomanager = True
                        
                        desktoptoken = mm_obj.desktopToken 
                        # desktopnotf = senddesktopnotf(desktoptoken,firebasemsg)
                            
                        if fcmtoken is not None and fcmtoken != "":
                            firebasenotification = ""
                            # firebasenotification = sendfirebasenotification(fcmtoken,firebasemsg,notftype,fcmleaveid,fcmtomanager)
                        else:
                            firebasenotification = ""
                data_dict = {
                            "employeename":username,
                            "actedmanager":managername,
                            "startdate":startdate,
                            "enddate":enddate,
                            "dates":leave_dates_handled,
                            "leave_status":str(leave_obj1.leave_status),
                            "leave_reason":str(leave_obj1.reason),
                            "applicationid":str(leave_obj1.ApplicationId),
                            "Action_status":"Approved",
                            "WFH":leave_obj1.WorkFromHome
                        }
                send_async_custom_template_email(
                    str(managername) +' has approved '+str(WorkFromHome(leaveid)) +' application',
                    data_dict,
                    EMAIL_HOST_USER,
                    Team_members,
                    "leaveactionmail.html"
                )
                send_async_custom_template_email(
                    str(managername) +' has approved your '+str(WorkFromHome(leaveid)) +' application',
                    data_dict,
                    EMAIL_HOST_USER,
                    [str(userobj.email)],
                    "leaveactionmailtoemployee.html"
                )


        else:

            return Response ({"data":{},"response":{"n" : 2,"msg" : "Someting went wrong,try again later","status" : "error"}})

    return Response ({"data":{"status":Leave_Status,"employee_name":employee_name},"response":{"n" : 1,"msg" : "Successfully "+Leave_Status,"status" : "success"}})

@api_view(['POST'])
def reject_approved_application(request):
    data={}
    id = request.data.get('id')
    comment = request.data.get('comment')
    com_code = request.user.company_code
    employee_name=""
    Leave_Status=""
    acted_user = request.user.id  
    manageruserobj = Users.objects.filter(id=int(acted_user)).first()
    managername = manageruserobj.Firstname + " " + manageruserobj.Lastname 


    if comment is not None and comment != "":
        data['rejectedBy'] = True
        data['approvedBy'] = False
        data['comment']= comment
        leaveapp = leaveApproval.objects.filter(id=id).first()
        if leaveapp:
            aserializer = leaveapprovalserializer(leaveapp,data=data,partial=True)
            if aserializer.is_valid():
                aserializer.save()
                Leave_Status="Rejected"

                leaveid = aserializer.data['leave_id']
                appstatus = WorkFromHome(leaveid)
                
                Leave.objects.filter(id=leaveid,Active=True).update(leave_status="Rejected")
                TaskNotification.objects.filter(leaveID=leaveid,To_manager=True,UserID=request.user.id).update(action_Taken=True)
                TaskNotification.objects.filter(leaveID=leaveid,To_manager=True).update(action_Taken=True)

                empid = aserializer.data['employeeId']
                applctnid = aserializer.data['ApplicationId']
            
                userobj = Users.objects.filter(id=int(empid),is_active=True).first()
                username = userobj.Firstname + " " + userobj.Lastname 

                Leaveobj = Leave.objects.filter(id=leaveid,Active=True).first()
                
                leave_dates_handled=date_handling(str(Leaveobj.start_date),str(Leaveobj.end_date))
                
                startd = str(Leaveobj.start_date) 
                startdateday = startd.split("-")[2]
                startdateyear = startd.split("-")[0]
                startdatemonth = month_converter(int(startd.split("-")[1]))
                startdate = str(startdateday) +" "+ startdatemonth +" "+str(startdateyear)

                endd = str(Leaveobj.end_date)
                enddateday = endd.split("-")[2]
                enddateyear = endd.split("-")[0]
                enddatemonth = month_converter(int(endd.split("-")[1]))
                enddate = str(enddateday) +" "+ enddatemonth +" "+str(enddateyear)

                if startd == endd:
                    notfmsg = "You have rejected " +str(appstatus)+ " Application of  <span class='actionuser'>" + username + " </span> dated <span class='notfleavedate'>" +startdate+"</span> "
                else:
                    notfmsg = "You have rejected " +str(appstatus)+ " Application of <span class='actionuser'>" + username + " </span> dated <span class='notfleavedate'>" +startdate+" to "+enddate+"</span> "

                # notf for manager after action
                empleaverejectobj=TaskNotification.objects.create(UserID_id=int(request.user.id),leaveID=leaveid,To_manager=True,company_code=com_code,NotificationTypeId_id=3,action_Taken=None,NotificationTitle="Leave",NotificationMsg=notfmsg)
                
                # notf for employee after action
                notificationmsg = "Your " +str(appstatus)+ " request with Application Id <span class='notfappid'>" +applctnid+"</span> has been <span class='rejectedmsg'> Rejected </span>"

                TaskNotification.objects.create(UserID_id=empid,company_code=com_code,NotificationTypeId_id=3,NotificationTitle="Leave Rejected",NotificationMsg=notificationmsg,leaveID=leaveid)

                # notf for admin after reject action
                if startd == endd:
                    adminnotfmsg = str(appstatus)+ " request of " +username+ " </span> dated <span class='notfleavedate'> " +startdate+" </span></span> has been <span class='rejectedmsg'> Rejected </span>"
                else:
                    adminnotfmsg = str(appstatus)+ " request of " +username+" </span> dated <span class='notfleavedate'>" +startdate+" to "+enddate+" </span></span> has been <span class='rejectedmsg'> Rejected </span>"

                ad_mail = adminemail
                adminobj = Users.objects.filter(email=ad_mail).first()
                adminid = adminobj.id

                TaskNotification.objects.create(UserID_id=int(adminid),leaveID=leaveid,company_code=com_code,NotificationTypeId_id=3,NotificationTitle="Leave",NotificationMsg=adminnotfmsg)



                #firebase for employee(reject)
                firebaseempnotfmsg =  "Your " +str(appstatus)+ " request with Application Id "+applctnid+" has been Rejected"

                userfidobj = Users.objects.filter(id=int(empid)).first()
                if userfidobj is not None:
                    firebasemsg =  firebaseempnotfmsg
                    fcmtoken = userfidobj.FirebaseID
                    notftype = "Leave"
                    fcmleaveid  = leaveid
                    fcmtomanager = False
                    
                    desktoptoken = userfidobj.desktopToken 
                    # desktopnotf = senddesktopnotf(desktoptoken,firebasemsg)
                    
                    if fcmtoken is not None and fcmtoken != "":
                        firebasenotification = ""
                        # firebasenotification = sendfirebasenotification(fcmtoken,firebasemsg,notftype,fcmleaveid,fcmtomanager)
                    else:
                        firebasenotification = ""


                #notf for other managers
                notfobjects = TaskNotification.objects.filter(leaveID=leaveid,To_manager=True).distinct('UserID').exclude(UserID=acted_user)
                
                notfser = TaskNotificationSerializer(notfobjects,many=True)
                for p in notfser.data:
                    managerLeaveobj = Leave.objects.filter(id=p['leaveID'],Active=True).first()
                    if managerLeaveobj is not None:
                        mangeruserid = managerLeaveobj.employeeId

                        muserobj = Users.objects.filter(id=int(mangeruserid),is_active=True).first()
                        musername = muserobj.Firstname + " " + muserobj.Lastname 

                        sdt1 = str(managerLeaveobj.start_date) 
                        sdtday = sdt1.split("-")[2]
                        sdtyear = sdt1.split("-")[0]
                        sdmonth = month_converter(int(sdt1.split("-")[1]))
                        mstartdate = str(sdtday) +" "+ sdmonth +" "+str(sdtyear)


                        edt1 = str(managerLeaveobj.end_date)
                        edtday = edt1.split("-")[2]
                        edtyear = edt1.split("-")[0]
                        edmonth = month_converter(int(edt1.split("-")[1]))
                        menddate = str(edtday) +" "+ edmonth +" "+str(edtyear)


                        if sdt1 == edt1 :
                            managernotfmsg = "<span class='managernotfname'>" +managername + " </span> has rejected "+str(appstatus)+" Application of <span class='actionuser'>" +musername + "</span> dated  <span class='notfleavedate'>" + mstartdate +"</span>"

                            firebasemmnotfmsg = managername + " has rejected "+str(appstatus)+" Application of "+musername+" dated  " +mstartdate 

                        else:
                            managernotfmsg = "<span class='managernotfname'>" +managername + " </span> has rejected "+str(appstatus)+" Application of <span class='actionuser'>" +username + "</span> dated  <span class='notfleavedate'>" + mstartdate +" to " + menddate + "</span>"

                            firebasemmnotfmsg = managername + " has rejected "+str(appstatus)+" Application of "+musername+" dated " + mstartdate +" to " + menddate
                    
                
                        TaskNotification.objects.create(UserID_id=p['UserID'],company_code=com_code,To_manager=True,action_Taken=None,leaveID=leaveid,NotificationTypeId_id=3,NotificationTitle="Leave",NotificationMsg=managernotfmsg,created_by = int(request.user.id))

                        userfffidobj = Users.objects.filter(id=int(p['UserID'])).first()
                        if userfffidobj is not None:
                            firebasemsg =  firebasemmnotfmsg
                            fcmtoken = userfffidobj.FirebaseID
                            notftype = "Leave"
                            fcmleaveid = leaveid
                            fcmtomanager = True
                            desktoptoken = userfffidobj.desktopToken 
                            # desktopnotf = senddesktopnotf(desktoptoken,firebasemsg)
                    
                            if fcmtoken is not None and fcmtoken != "":
                                firebasenotification = ""
                                # firebasenotification = sendfirebasenotification(fcmtoken,firebasemsg,notftype,fcmleaveid,fcmtomanager)
                            else:
                                firebasenotification = ""




                username=username.title()
                leave_obj1=Leave.objects.filter(id=leaveid).first()
                Team_members=[adminemail,hremail]
                leave_approval_object=leaveApproval.objects.filter(employeeId=empid,leave_id=leaveid,company_code=com_code).exclude(managerId=acted_user)
                leave_approval_serializer = leaveapprovalserializer(leave_approval_object,many=True)
                for manager in leave_approval_serializer.data:
                    m_obj = Users.objects.filter(id=int(manager['managerId']),is_active=True).first()
                    if m_obj is not None:
                        Team_members.append(m_obj.email)
                data_dict = {
                                "employeename":username,
                                "actedmanager":managername,
                                "startdate":startdate,
                                "enddate":enddate,
                                "reason":aserializer.data['comment'],
                                "WFH":leave_obj1.WorkFromHome,
                                "dates":leave_dates_handled,
                                "leave_status":str(leave_obj1.leave_status),
                                "applicationid":str(leave_obj1.ApplicationId),
                                "Action_status":"Rejected",
                                "leave_reason":str(leave_obj1.reason),
                }


                # Send a welcome email with a specific template
                send_async_custom_template_email(
                    str(managername) +' has rejected  '+str(WorkFromHome(leaveid)) +' application',
                    data_dict,
                    EMAIL_HOST_USER,
                    Team_members,
                    "leaveactionmail.html"
                )
                send_async_custom_template_email(
                    str(managername) +' has rejected your '+str(WorkFromHome(leaveid)) +' application',
                    data_dict,
                    EMAIL_HOST_USER,
                    [str(userobj.email)],
                    "leaveactionmailtoemployee.html"
                )
                
                

            return Response ({"data":{"status":Leave_Status,"employee_name":employee_name},"response":{"n" : 1,"msg" : "Successfully rejected","status" : "success"}})
        
        else:
            return Response ({"data":{},"response":{"n" : 2,"msg" : "Someting went wrong,try again later","status" : "error"}})
    else:
        return Response ({"data":{},"response":{"n" : 2,"msg" : "Please provide valid rejection reason","status" : "error"}})

@api_view(['POST'])
def approve_rejected_application(request):
    data={}
    id = request.data.get('id')
    com_code = request.user.company_code
    employee_name=""
    Leave_Status=""
    acted_user = request.user.id  
    manageruserobj = Users.objects.filter(id=int(acted_user)).first()
    managername = manageruserobj.Firstname + " " + manageruserobj.Lastname 


    if id is not None and id != "":
        data['rejectedBy'] = False
        data['approvedBy'] = True
        data['comment'] = None
        leaveapp = leaveApproval.objects.filter(id=id).first()
        if leaveapp:
            aserializer = leaveapprovalserializer(leaveapp,data=data,partial=True)
            if aserializer.is_valid():
                aserializer.save()
                
                check_all_approvals=leaveApproval.objects.filter(leave_id=aserializer.data['leave_id'],rejectedBy=True).exclude(id=aserializer.data['id']).first()
                if check_all_approvals is not None:
                    Leave_Status="Rejected"
                else:
                    check_pendings=leaveApproval.objects.filter(leave_id=aserializer.data['leave_id'],approvedBy=False,rejectedBy=False).exclude(id=aserializer.data['id']).first()
                    if check_pendings is not None:
                        Leave_Status="Pending"
                    else:
                        Leave_Status="Approved"
                        
                    
                

                leaveid = aserializer.data['leave_id']
                appstatus = WorkFromHome(leaveid)
                
                Leave.objects.filter(id=leaveid,Active=True).update(leave_status=Leave_Status)
                TaskNotification.objects.filter(leaveID=leaveid,To_manager=True,UserID=request.user.id).update(action_Taken=True)
                TaskNotification.objects.filter(leaveID=leaveid,To_manager=True).update(action_Taken=True)

                empid = aserializer.data['employeeId']
                applctnid = aserializer.data['ApplicationId']
            
                userobj = Users.objects.filter(id=int(empid),is_active=True).first()
                username = userobj.Firstname + " " + userobj.Lastname 

                Leaveobj = Leave.objects.filter(id=leaveid,Active=True).first()
                
                leave_dates_handled=date_handling(str(Leaveobj.start_date),str(Leaveobj.end_date))
                
                startd = str(Leaveobj.start_date) 
                startdateday = startd.split("-")[2]
                startdateyear = startd.split("-")[0]
                startdatemonth = month_converter(int(startd.split("-")[1]))
                startdate = str(startdateday) +" "+ startdatemonth +" "+str(startdateyear)

                endd = str(Leaveobj.end_date)
                enddateday = endd.split("-")[2]
                enddateyear = endd.split("-")[0]
                enddatemonth = month_converter(int(endd.split("-")[1]))
                enddate = str(enddateday) +" "+ enddatemonth +" "+str(enddateyear)

                if startd == endd:
                    notfmsg = "You have approved " +str(appstatus)+ " Application of  <span class='actionuser'>" + username + " </span> dated <span class='notfleavedate'>" +startdate+"</span> "
                else:
                    notfmsg = "You have approved " +str(appstatus)+ " Application of <span class='actionuser'>" + username + " </span> dated <span class='notfleavedate'>" +startdate+" to "+enddate+"</span> "

                # notf for manager after action
                empleaverejectobj=TaskNotification.objects.create(UserID_id=int(request.user.id),leaveID=leaveid,To_manager=True,company_code=com_code,NotificationTypeId_id=3,action_Taken=None,NotificationTitle="Leave",NotificationMsg=notfmsg)
                
                # notf for employee after action
                notificationmsg = "Your " +str(appstatus)+ " request with Application Id <span class='notfappid'>" +applctnid+"</span> has been <span class='approvedmsg'> Approved </span>"

                TaskNotification.objects.create(UserID_id=empid,company_code=com_code,NotificationTypeId_id=3,NotificationTitle="Leave approved",NotificationMsg=notificationmsg,leaveID=leaveid)

                # notf for admin after reject action
                if startd == endd:
                    adminnotfmsg = str(appstatus)+ " request of " +username+ " </span> dated <span class='notfleavedate'> " +startdate+" </span></span> has been <span class='approvedmsg'> approved </span>"
                else:
                    adminnotfmsg = str(appstatus)+ " request of " +username+" </span> dated <span class='notfleavedate'>" +startdate+" to "+enddate+" </span></span> has been <span class='approvedmsg'> approved </span>"

                ad_mail = adminemail
                adminobj = Users.objects.filter(email=ad_mail).first()
                adminid = adminobj.id

                TaskNotification.objects.create(UserID_id=int(adminid),leaveID=leaveid,company_code=com_code,NotificationTypeId_id=3,NotificationTitle="Leave",NotificationMsg=adminnotfmsg)



                #firebase for employee(reject)
                firebaseempnotfmsg =  "Your " +str(appstatus)+ " request with Application Id "+applctnid+" has been approved"

                userfidobj = Users.objects.filter(id=int(empid)).first()
                if userfidobj is not None:
                    firebasemsg =  firebaseempnotfmsg
                    fcmtoken = userfidobj.FirebaseID
                    notftype = "Leave"
                    fcmleaveid  = leaveid
                    fcmtomanager = False
                    
                    desktoptoken = userfidobj.desktopToken 
                    # desktopnotf = senddesktopnotf(desktoptoken,firebasemsg)
                    
                    if fcmtoken is not None and fcmtoken != "":
                        firebasenotification = ""
                        # firebasenotification = sendfirebasenotification(fcmtoken,firebasemsg,notftype,fcmleaveid,fcmtomanager)
                    else:
                        firebasenotification = ""


                #notf for other managers
                notfobjects = TaskNotification.objects.filter(leaveID=leaveid,To_manager=True).distinct('UserID').exclude(UserID=acted_user)
                
                notfser = TaskNotificationSerializer(notfobjects,many=True)
                for p in notfser.data:
                    managerLeaveobj = Leave.objects.filter(id=p['leaveID'],Active=True).first()
                    if managerLeaveobj is not None:
                        mangeruserid = managerLeaveobj.employeeId

                        muserobj = Users.objects.filter(id=int(mangeruserid),is_active=True).first()
                        musername = muserobj.Firstname + " " + muserobj.Lastname 

                        sdt1 = str(managerLeaveobj.start_date) 
                        sdtday = sdt1.split("-")[2]
                        sdtyear = sdt1.split("-")[0]
                        sdmonth = month_converter(int(sdt1.split("-")[1]))
                        mstartdate = str(sdtday) +" "+ sdmonth +" "+str(sdtyear)


                        edt1 = str(managerLeaveobj.end_date)
                        edtday = edt1.split("-")[2]
                        edtyear = edt1.split("-")[0]
                        edmonth = month_converter(int(edt1.split("-")[1]))
                        menddate = str(edtday) +" "+ edmonth +" "+str(edtyear)


                        if sdt1 == edt1 :
                            managernotfmsg = "<span class='managernotfname'>" +managername + " </span> has approved "+str(appstatus)+" Application of <span class='actionuser'>" +musername + "</span> dated  <span class='notfleavedate'>" + mstartdate +"</span>"

                            firebasemmnotfmsg = managername + " has approved "+str(appstatus)+" Application of "+musername+" dated  " +mstartdate 

                        else:
                            managernotfmsg = "<span class='managernotfname'>" +managername + " </span> has approved "+str(appstatus)+" Application of <span class='actionuser'>" +username + "</span> dated  <span class='notfleavedate'>" + mstartdate +" to " + menddate + "</span>"

                            firebasemmnotfmsg = managername + " has approved "+str(appstatus)+" Application of "+musername+" dated " + mstartdate +" to " + menddate
                    
                
                        TaskNotification.objects.create(UserID_id=p['UserID'],company_code=com_code,To_manager=True,action_Taken=None,leaveID=leaveid,NotificationTypeId_id=3,NotificationTitle="Leave",NotificationMsg=managernotfmsg,created_by = int(request.user.id))

                        userfffidobj = Users.objects.filter(id=int(p['UserID'])).first()
                        if userfffidobj is not None:
                            firebasemsg =  firebasemmnotfmsg
                            fcmtoken = userfffidobj.FirebaseID
                            notftype = "Leave"
                            fcmleaveid = leaveid
                            fcmtomanager = True
                            desktoptoken = userfffidobj.desktopToken 
                            # desktopnotf = senddesktopnotf(desktoptoken,firebasemsg)
                    
                            if fcmtoken is not None and fcmtoken != "":
                                firebasenotification = ""
                                # firebasenotification = sendfirebasenotification(fcmtoken,firebasemsg,notftype,fcmleaveid,fcmtomanager)
                            else:
                                firebasenotification = ""




                username=username.title()
                leave_obj1=Leave.objects.filter(id=leaveid).first()
                Team_members=[adminemail,hremail]
                leave_approval_object=leaveApproval.objects.filter(employeeId=empid,leave_id=leaveid,company_code=com_code).exclude(managerId=acted_user)
                leave_approval_serializer = leaveapprovalserializer(leave_approval_object,many=True)
                for manager in leave_approval_serializer.data:
                    m_obj = Users.objects.filter(id=int(manager['managerId']),is_active=True).first()
                    if m_obj is not None:
                        Team_members.append(m_obj.email)
                data_dict = {
                                "employeename":username,
                                "actedmanager":managername,
                                "startdate":startdate,
                                "enddate":enddate,
                                "reason":aserializer.data['comment'],
                                "WFH":leave_obj1.WorkFromHome,
                                "dates":leave_dates_handled,
                                "leave_status":str(leave_obj1.leave_status),
                                "applicationid":str(leave_obj1.ApplicationId),
                                "Action_status":"Approved",
                                "leave_reason":str(leave_obj1.reason),
                }


                # Send a welcome email with a specific template
                send_async_custom_template_email(
                    str(managername) +' has approved  '+str(WorkFromHome(leaveid)) +' application',
                    data_dict,
                    EMAIL_HOST_USER,
                    Team_members,
                    "leaveactionmail.html"
                )
                send_async_custom_template_email(
                    str(managername) +' has approved your '+str(WorkFromHome(leaveid)) +' application',
                    data_dict,
                    EMAIL_HOST_USER,
                    [str(userobj.email)],
                    "leaveactionmailtoemployee.html"
                )
                
                return Response ({"data":{"status":Leave_Status,"employee_name":employee_name},"response":{"n" : 1,"msg" : "Successfully approved","status" : "success"}})
                
            else:
                return Response ({"data":{"status":Leave_Status,"employee_name":employee_name},"response":{"n" : 1,"msg" : "Sorry Serializer error","status" : "error"}})
                
        
        else:
            return Response ({"data":{},"response":{"n" : 2,"msg" : "Someting went wrong,try again later","status" : "error"}})
    else:
        return Response ({"data":{},"response":{"n" : 2,"msg" : "Application id not found","status" : "error"}})

@api_view(['POST'])
def AdminLeaveApprovedPostApi(request):
    data=request.data
    com_code = request.user.company_code
    adminObj=adminAttendanceRequest.objects.filter(employeeId=data['UserId'],date=data['date'],adminId=111).first()
    if adminObj is None:
        adminAttendanceRequest.objects.create(employeeId=data['UserId'],date=data['date'],comment=data['comment'],adminId=111,company_code=com_code)
        return Response ({"data":data,"response":{"n" : 1,"msg" : "SUCCESS","status" : "success"}})
    else:
        return Response ({"data":"","response":{"n" : 0,"msg" : "SUCCESS","status" : "success"}})

@api_view(['GET'])
def leave_user_emp_list(request):
    if request.method == 'GET':
        company_code = request.user.company_code 

        user = Users.objects.filter(is_active=True,company_code=company_code,)
        serializer = GetUserSerializerleavemapping(user,many=True)
        
        
        def searchmanagerfunction(id,position,company_code):
            filter_obj= leaveMapping.objects.filter(employeeId=id,position=position,company_code=company_code).first()
            if filter_obj is not None: 
                userobj=Users.objects.filter(id=filter_obj.managerId).first()
                if userobj is not None:
                    return {'userid':int(userobj.id) ,'username':str(userobj.Firstname) + " " + str(userobj.Lastname) }
                else:
                    return {'userid':None ,'username':None }
            else:
                return {'userid':None ,'username':None }

        for i in serializer.data:
            id_val=i['id']
            manager1=searchmanagerfunction(id_val,"1",company_code)
            manager2=searchmanagerfunction(id_val,"2",company_code)
            manager3=searchmanagerfunction(id_val,"3",company_code)
            i['managername1']=""
            i['managername2']=""
            i['managername3']=""
            i['managerid1']=""
            i['managerid2']=""
            i['managerid3']=""
            
            if manager1['username'] is not None and manager1['username'] !="":
                i['managername1']=manager1['username']
            if manager2['username'] is not None and manager2['username'] !="":
                i['managername2']=manager2['username']
            if manager3['username'] is not None and manager3['username'] !="":
                i['managername3']=manager3['username']

            if manager1['userid'] is not None and manager1['userid'] !="":
                i['managerid1']=manager1['userid']
            if manager2['userid'] is not None and manager2['userid'] !="":
                i['managerid2']=manager2['userid']
            if manager3['userid'] is not None and manager3['userid'] !="":
                i['managerid3']=manager3['userid']



        return Response({'n':1,'msg':'Employee list fetched successfully','status':'success','data':serializer.data})

@api_view(['POST'])
def Apply_Leave_Mapping(request):
    com_code = request.user.company_code
   
    New_dictionary =json.loads(request.POST.get("New_dictionary") )

    for i in New_dictionary:  
        if i['managerId']:
            if i['managerId'] != "":
                if str(i['managerId']) != str(i['employeeId']):
                    map_obj=leaveMapping.objects.filter(employeeId=i['employeeId'],position=i['position'],company_code=com_code).first()
                    if map_obj != None:
                        map_obj.position =i['position']
                        map_obj.managerId=i['managerId']
                        map_obj.save()
                    else:
                        leaveMapping.objects.create(employeeId=i['employeeId'],managerId=i['managerId'],position=i['position'],company_code=com_code)
            else:
                map_obj=leaveMapping.objects.filter(employeeId=i['employeeId'],position=i['position'],company_code=com_code).delete()



    return Response ({"data":"","response":{"n" : 1,"msg" : "Mapping applied successfully","status" : "success"}})

@api_view(['POST'])    
@permission_classes((AllowAny,))
def employee_yearly_total_days(request):
    month=request.POST.get("month")
    month=int(month)
    year=request.POST.get("year")
    
    empuserID=request.POST.get('UserId')
    if empuserID is None or empuserID =="":
        return Response ({
                          "Photo":'',
                          "Designation":'',
                          "Name":'',
                          "EmployeeId":'',
                          "avghour":0,
                          "daylongesthour":"00:00:00",
                          "year_total_leaves":0,
                          "total_6monthly_leave_count":0,                
                          "monthstr":'',
                          "year_total_days":0,
                          "monthly_late_count_data":{
                            "latecount":0,
                            "days":0 ,
                            "details":[]
                          },
                          "response":{"n" : 0,"msg" : "Please provide User id","status" : "failed"}})



    empid=Users.objects.filter(id=empuserID).first()

    

    if empid is None:
        return Response ({
            
            
                          "Photo":'',
                          "Designation":'',
                          "Name":'',
                          "EmployeeId":'',
                          "monthstr":'',  
                          "total_6monthly_leave_count":0,                
                          "avghour":0,
                          "monthly_late_count_data":{
                            "latecount":0,
                            "days":0 ,
                            "details":[]
                          },
                          


                          "response":{"n" : 0,"msg" : "No user found","status" : "failed"}})

    userdetails=UserSerializer(empid)
    employee_id=empid.employeeId
    year=int(year)


    def avghour(hourlist):
        if hourlist !=[]:
            datetime_list = [datetime.strptime(time_str, "%H:%M:%S") for time_str in hourlist]
            total_hours = sum([dt.hour for dt in datetime_list])          
            average_hour = total_hours / len(datetime_list)
            return(round(average_hour)) 
        else:
            return 0


    attendance_obj = attendance.objects.filter(Year=year, Month=month, employeeId=employee_id).distinct('date')



    if userdetails.data['Photo'] != "" and userdetails.data['Photo'] is not None:
        userimage = imageUrl + str(userdetails.data['Photo'])
    else:
        userimage = imageUrl + "/static/assets/images/profile.png"

    # Calculate the total time for each date
    total_time_list = []
    attendance_serializer=attendanceserializer(attendance_obj,many=True)
    
    for i in attendance_serializer.data:
        date = i['date']
        attendance_login_obj = attendance.objects.filter(employeeId=employee_id,date=date).order_by('time').first()
        attendance_logout_obj = attendance.objects.filter(employeeId=employee_id,date=date).order_by('time').last()
        format_str = "%H:%M:%S"
        
        start_datetime = datetime.strptime(str(attendance_login_obj.time), format_str)
        end_datetime = datetime.strptime(str(attendance_logout_obj.time), format_str)
        if start_datetime != end_datetime:

            time_diff = end_datetime - start_datetime

            total_time_list.append(str(time_diff))
            

    averagehours=avghour(total_time_list)

    
    
    
    

    
    attendance_obj_latecount = attendance.objects.filter(date__year=request.data.get("year"),date__month=request.data.get("month"),time__gte=time(10, 0, 0),time__lte=time(11, 0, 0),employeeId=employee_id).order_by('date','time').distinct('date','time')
    attendance_obj_latecount=attendance_obj_latecount.distinct('date')
    attendance_obj_early = attendance.objects.filter(date__year=request.data.get("year"),date__month=request.data.get("month"),time__gte=time(1, 0, 0),time__lt=time(10, 0, 0),employeeId=employee_id).distinct('date')
    exclude_date_list=AttendanceSerializerAttendanceWeekDate(attendance_obj_early,many=True)
    
    attendance_obj_latecount=attendance_obj_latecount.exclude(date__in=list(exclude_date_list.data))




    late_counts=0
    late_counts=attendance_obj_latecount.count()
    serializer_late_count=attendanceserializer(attendance_obj_latecount,many=True)
    late_marks_data=serializer_late_count.data


            
            
            
    late_days_count=assign_value_based_on_length(late_counts)

    total_6monthly_leave_count=0
    monthstr = "April-September"

    if 4 <= int(month) <= 9:
        monthstr = "April-September"
        total_6monthly_leave_count_obj = Leave.objects.filter(leave_status="Approved",employeeId=empuserID,WorkFromHome=False,start_date__month__gte=4,start_date__month__lte=9,start_date__year=year,Active=True).exclude(leave_status='Draft')
        if total_6monthly_leave_count_obj:
            total_6monthly_leave_count=total_6monthly_leave_count_obj.aggregate(Sum('number_of_days'))['number_of_days__sum']
        else:
            total_6monthly_leave_count=0
    else:
        monthstr = "October-March"
        total_6monthly_leave_count_obj  = Leave.objects.filter(leave_status="Approved",employeeId=empuserID,WorkFromHome=False,start_date__month__in=[10,11,12,1,2,3],start_date__year=year,Active=True).exclude(leave_status='Draft')
        if total_6monthly_leave_count_obj:
            total_6monthly_leave_count=total_6monthly_leave_count_obj.aggregate(Sum('number_of_days'))['number_of_days__sum']
        else:
            total_6monthly_leave_count=0





    
    return Response ({
                        
                        "Photo":userimage,
                        "Designation":userdetails.data['DesignationId'],
                        "Name":userdetails.data['Firstname']+" "+userdetails.data['Lastname'],
                        "EmployeeId":userdetails.data['employeeId'],
                        "avghour":averagehours,
                        "total_6monthly_leave_count":total_6monthly_leave_count,               
                        "monthstr":monthstr,
                        "monthly_late_count_data":{
                                "latecount":late_counts,
                                "days":late_days_count ,
                                "details":late_marks_data,
                            },
                            "response":{"n" : 1,"msg" : "SUCCESS","status" : "success"}})



@api_view(['POST'])
def dashboard_leave_card(request):
    employeeId = request.user.id 
    month = date.today()
    avaliable = 12
    starting_month = [4,5,6,7,8,9]
  
    current_month = month.month
   
    current_year = month.year
    if 4 <= current_month <= 9:
        monthstr = "April-September"
        pending_leave_countobj = Leave.objects.filter(employeeId=employeeId,WorkFromHome=False,start_date__month__gte=4,start_date__month__lte=9,start_date__year= current_year,Active=True,leave_status='Pending').count()
        approve_leave_countobj = Leave.objects.filter(employeeId=employeeId,WorkFromHome=False,start_date__month__gte=4,start_date__month__lte=9,start_date__year= current_year,Active=True,leave_status='Approved').count()
        rejected_leave_countobj = Leave.objects.filter(employeeId=employeeId,WorkFromHome=False,start_date__month__gte=4,start_date__month__lte=9,start_date__year= current_year,Active=True,leave_status='Rejected').count()
        total_count = Leave.objects.filter(employeeId=employeeId,WorkFromHome=False,start_date__month__gte=4,start_date__month__lte=9,start_date__year= current_year,Active=True).exclude(leave_status='Draft').count()
    else:
        monthstr = "October-March"
        pending_leave_countobj = Leave.objects.filter(employeeId=employeeId,WorkFromHome=False,start_date__month__in=[10,11,12,1,2,3],start_date__year= current_year,Active=True,leave_status='Pending').count()
        approve_leave_countobj = Leave.objects.filter(employeeId=employeeId,WorkFromHome=False,start_date__month__in=[10,11,12,1,2,3],start_date__year= current_year,Active=True,leave_status='Approved').count()
        rejected_leave_countobj = Leave.objects.filter(employeeId=employeeId,WorkFromHome=False,start_date__month__in=[10,11,12,1,2,3],start_date__year= current_year,Active=True,leave_status='Rejected').count()
    
        total_count  = Leave.objects.filter(employeeId=employeeId,WorkFromHome=False,start_date__month__in=[10,11,12,1,2,3],start_date__year= current_year,Active=True).exclude(leave_status='Draft').count()

    return Response({"response":{'n': 1,'msg':'Data saved successfully','Status': 'success','approved_leave_count':approve_leave_countobj,
                    'rejected_leave_count':rejected_leave_countobj,'monthstr':monthstr,'pending_leave_count':pending_leave_countobj,'total_leave_count':total_count}})

@api_view(['POST'])
def getattendancecount(request):

    month=request.data.get("month")
    year=request.data.get("year")
    ManagerID = request.query_params.get('userid', None)
    year=int(year)
    month=int(month)
    month_number=month

    list1=[]
    d2=[]
    holidaylist=[]
    company_code = request.user.company_code
    
    

    my_date = date.today()
    YYear, WWeek_num, dday_of_week = my_date.isocalendar()


    month = month_number
    year = year
    number_of_days = calendar.monthrange(year, month)[1]
    first_date = date(year, month, 1)
    last_date = date(year, month, number_of_days)
    delta = last_date - first_date
    dayslist=[(first_date + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(delta.days + 1)]

    year=year 
    month=month_number
    dt=date(year,month,1)   # first day of month
    first_w=dt.isoweekday()  # weekday of 1st day of the month
    if(first_w==7): # if it is Sunday return 0 
        first_w=0
    saturday2=14-first_w
    dt1=date(year,month,saturday2)
    list1.append(str(dt1))
    saturday4=28-first_w
    dt1=date(year,month,saturday4)
    list1.append(str(dt1))  
    weeklyofflist=list1
        
    # all sundays of the month
    def allsundays(month):
        d = date(year,month, 1) # day 1st of month
        d += timedelta(days = 6 - d.weekday())  # First Sunday
        while d.month == month:
            yield d
            d += timedelta(days = 7)

    for d in allsundays(month):
        d=str(d)
        weeklyofflist.append(d)


    totalteamcount=0
    presentteamcount=0
    onleaveteamcount=0
    onwfhteamcount=0
    team_with_attendanceid_list=[]
    team_without_attendanceid_list=[]
    leavelist=[]
    wfhlist=[]
    team_list=[]
    found_team_members=[]
    yettojoinlist=[]

    if ManagerID is not None:
        managerid=ManagerID
        mappingobj = UserToManager.objects.filter(ManagerID=managerid,company_code = company_code)
        totalteamcount=mappingobj.count()
        map_serializer = MappingSerializer(mappingobj, many=True)
        for j in map_serializer.data:
            team_list.append(j['UserID'])
            
        for team_member in team_list:
            team_member_obj=Users.objects.filter(id=team_member,is_active=True).order_by('id').first()
            if team_member_obj:
                if team_member_obj.employeeId is not None and team_member_obj.employeeId !="":
                    team_with_attendanceid_list.append(team_member_obj.employeeId)

    else:
        empobj=Users.objects.filter(is_active=True,company_code=company_code)
        empobjser=UserSerializer(empobj,many=True)
        totalteamcount=empobj.count()
        managerid="All"
        for j in empobjser.data:
            team_list.append(j['id'])
            if j['employeeId'] is not None and j['employeeId'] !="":
                team_with_attendanceid_list.append(j['employeeId'])








    leave_data=Leave.objects.filter(Active=True,company_code=company_code,employeeId__in=team_list,leave_status="Approved").order_by('id')
    leave_data_serializer = leaveserializer(leave_data, many=True)

    for l in leave_data_serializer.data:
        leave_userobj = Users.objects.filter(id=int(l['employeeId'])).first()

        date2 =date.fromisoformat(str(my_date))
        date1 =date.fromisoformat(l['start_date'])         
        date3 =date.fromisoformat(l['end_date']) 
        if date1 <= date2 and date2 <= date3:
            if l['WorkFromHome']:
                onwfhteamcount += 1

                wfhinfo={}
                wfhinfo['wfhcounter'] = onwfhteamcount
                wfhinfo['empname']= leave_userobj.Firstname +" "+ leave_userobj.Lastname
                wfhinfo['Leavetype']=l['leavetype']
                wfhlist.append(wfhinfo)
            else:
                onleaveteamcount += 1

                leaveinfo={}
                leaveinfo['leavecounter'] = onleaveteamcount
                leaveinfo['empname']= leave_userobj.Firstname +" "+ leave_userobj.Lastname
                leaveinfo['Leavetype']=l['leavetype']
                leavelist.append(leaveinfo)
                found_team_members.append(leave_userobj.id)

    present_count_obj=attendance.objects.filter(date=str(my_date),employeeId__in=team_with_attendanceid_list,company_code=company_code).order_by("time")
    present_count_obj_ser=attendanceserializer(present_count_obj,many=True)
    presentcounter = 0
    Presentlist=[]
    for emp in present_count_obj_ser.data:
        att_userobj = Users.objects.filter(employeeId=str(emp['employeeId'])).first()
        presentinfo={}
        presentcounter += 1
        presentinfo['presentcounter'] = presentcounter
        presentinfo['empname']= att_userobj.Firstname +" "+ att_userobj.Lastname
        presentinfo['checkintime']= str(emp['time'])
        Presentlist.append(presentinfo)
        found_team_members.append(att_userobj.id)

    presentteamcount=present_count_obj.count()




    not_found_team_members = []
    for emp in team_list:
        if emp not in found_team_members:
            not_found_team_members.append(emp)
            

    ytjcounter = 0 
    for u in not_found_team_members:
        yettojoinobj = {}
        userobj = Users.objects.filter(id=int(u)).first()
        ytjcounter += 1

        if userobj.employeeId is not None and  userobj.employeeId != "":
            yettojoinobj['ytjcounter'] = ytjcounter
            yettojoinobj['username'] =  userobj.Firstname +" "+ userobj.Lastname
            yettojoinobj['reason'] =  "Not marked attendance"
        else:
            yettojoinobj['ytjcounter'] = ytjcounter
            yettojoinobj['username'] =  userobj.Firstname +" "+ userobj.Lastname
            yettojoinobj['reason'] =  "Attendance Id not found"
        
        yettojoinlist.append(yettojoinobj)

    

    holidatlist = Holidays.objects.filter(Active=True).order_by('id')
    serializer = holidaysSerializer(holidatlist, many=True)
    for i in serializer.data:
        holiday=i['Date']
        holidaylist.append(holiday)






    for j in dayslist:
        team_attendance=TeamAttendance.objects.filter(managerid=managerid,date=j).first()
        if team_attendance:
            TeamAttendanceSerializer_day=TeamAttendanceSerializer(team_attendance)
            count_present=TeamAttendanceSerializer_day.data['present_count'],
            count_noattendenceid=TeamAttendanceSerializer_day.data['not_found_count'],
            count_total=TeamAttendanceSerializer_day.data['total_count'],
            count_leaves=TeamAttendanceSerializer_day.data['on_leave_count'],
        else:
            count_present=0,
            count_noattendenceid=0,
            count_total=0,
            count_leaves=0,    
            
        if j in weeklyofflist:
            d1={
                "att_status":"WO",
                "fulldate":j,
                "present":count_present,
                "noattendenceid":count_noattendenceid,
                "total":count_total,
                "leaves":count_leaves,
            }
            d2.append(d1)
        elif j in holidaylist:
            reason='Holiday'
            Holidays_obj=Holidays.objects.filter(Date=j,Active=True).first()
            if Holidays_obj:
                reason=str(Holidays_obj.Festival)
                
            d1={
                "att_status":"H",
                "fulldate":j,
                "present":count_present,
                "noattendenceid":count_noattendenceid,
                "total":count_total,
                "leaves":count_leaves,
                "reason":reason,

            }
            d2.append(d1)
        else:
            


                
            d1={
                "att_status":"P",
                "fulldate":j,
                "present":count_present,
                "noattendenceid":count_noattendenceid,
                "total":count_total,
                "leaves":count_leaves,
            }
            
            d2.append(d1)


                
                
    yettojoin=totalteamcount-(onleaveteamcount + presentteamcount)
    context = {
        "All_Mapped_Emp_Count":totalteamcount,
        "Total_count" : totalteamcount,
        "Present":presentteamcount,
        "Onleave":onleaveteamcount,
        "yettojoin":yettojoin,
        "WFH":onwfhteamcount,
        "Presentlist":Presentlist,
        "leavelist":leavelist,
        "wfhlist":wfhlist,
        "yettojoinlist":yettojoinlist,

    }
    return Response ({"data":d2,"Photo":'',"Designation":'----',"Name":"All Employee",
                      "EmployeeId":"----","context":context,
                      "response":{"n" : 1,"msg" : "SUCCESS","status" : "success"}})

@api_view(['POST'])
def get_daily_attendance_by_month(request):

    month=request.data.get("month")
    year=request.data.get("year")
    year=int(year)
    month=int(month)
    month_number=month

    list1=[]
    d2=[]
    holidaylist=[]
    company_code = request.user.company_code
    
    
    month = month_number
    year = year
    number_of_days = calendar.monthrange(year, month)[1]
    first_date = date(year, month, 1)
    last_date = date(year, month, number_of_days)
    delta = last_date - first_date
    dayslist=[(first_date + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(delta.days + 1)]

    year=year 
    month=month_number
    dt=date(year,month,1)   # first day of month
    first_w=dt.isoweekday()  # weekday of 1st day of the month
    if(first_w==7): # if it is Sunday return 0 
        first_w=0
    saturday2=14-first_w
    dt1=date(year,month,saturday2)
    list1.append(str(dt1))
    saturday4=28-first_w
    dt1=date(year,month,saturday4)
    list1.append(str(dt1))  
    weeklyofflist=list1
        
    # all sundays of the month
    def allsundays(month):
        d = date(year,month, 1) # day 1st of month
        d += timedelta(days = 6 - d.weekday())  # First Sunday
        while d.month == month:
            yield d
            d += timedelta(days = 7)

    for d in allsundays(month):
        d=str(d)
        weeklyofflist.append(d)



    team_with_attendanceid_list=[]

    team_list=[]



                    




    l1l2emplst_obj=EmployeeShiftDetails.objects.filter(is_active=True)
    l1l2serializers = L1L2Serializer(l1l2emplst_obj,many=True)
    team_list=list(l1l2serializers.data)
    
    Users_l1l2emplst_obj=Users.objects.filter(id__in=team_list,is_active=True).exclude(Q(employeeId__isnull=True)|Q(employeeId='None')|Q(employeeId=''))
    Users_l1l2serializers = UsersSerializeronlyattendance(Users_l1l2emplst_obj,many=True)
    team_with_attendanceid_list=list(Users_l1l2serializers.data)







    holidatlist = Holidays.objects.filter(Active=True).order_by('id')
    serializer = holidaysSerializer(holidatlist, many=True)
    for i in serializer.data:
        holiday=i['Date']
        holidaylist.append(holiday)










    for j in dayslist:

        count_noattendenceid=0
        count_total=len(team_with_attendanceid_list)
        count_leaves=0 
        count_present = attendance.objects.filter(employeeId__in=team_with_attendanceid_list,date=j).distinct('employeeId').count() 
        count_noattendenceid=int(count_total)-int(count_present)-int(count_leaves)
        
        leave_data=Leave.objects.filter(WorkFromHome=False,Active=True,company_code=company_code,employeeId__in=team_list,leave_status="Approved").order_by('id')
        leave_data_serializer = leaveserializer(leave_data, many=True)

        for l in leave_data_serializer.data:
            date2 =date.fromisoformat(str(j))
            date1 =date.fromisoformat(l['start_date'])         
            date3 =date.fromisoformat(l['end_date']) 
            if date1 <= date2 and date2 <= date3:
                count_leaves += 1



           
           
           
           
           
        if j in holidaylist:
            reason='Holiday'
            Holidays_obj=Holidays.objects.filter(Date=j,Active=True).first()
            if Holidays_obj:
                reason=str(Holidays_obj.Festival)
            d1={
                "att_status":"H",
                "fulldate":j,
                "present":count_present,
                "noattendenceid":count_noattendenceid,
                "total":count_total,
                "leaves":count_leaves,
                "reason":reason,
            }
            d2.append(d1)
        else:

            d1={
                "att_status":"P",
                "fulldate":j,
                "present":count_present,
                "noattendenceid":count_noattendenceid,
                "total":count_total,
                "leaves":count_leaves,
            }
            
            d2.append(d1)


                
                

    return Response ({"data":d2,"response":{"n" : 1,"msg" : "SUCCESS","status" : "success"}})

@api_view(['POST'])
def getattendancebydate(request):
    ManagerID = request.query_params.get('userid', None)

    company_code = request.user.company_code
    
    

    my_date = request.POST.get("my_date")
    totalteamcount=0
    presentteamcount=0
    onleaveteamcount=0
    onwfhteamcount=0
    team_with_attendanceid_list=[]
    leavelist=[]
    wfhlist=[]
    team_list=[]
    found_team_members=[]
    yettojoinlist=[]

    if ManagerID is not None:
        managerid=ManagerID
        mappingobj = UserToManager.objects.filter(ManagerID=managerid,company_code = company_code)
        totalteamcount=mappingobj.count()
        map_serializer = MappingSerializer(mappingobj, many=True)
        for j in map_serializer.data:
            team_list.append(j['UserID'])
            
        for team_member in team_list:
            team_member_obj=Users.objects.filter(id=team_member,is_active=True).order_by('id').first()
            if team_member_obj:

                if team_member_obj.employeeId is not None and team_member_obj.employeeId !="":
                    team_with_attendanceid_list.append(team_member_obj.employeeId)

                    
            
            
    else:
        empobj=Users.objects.filter(is_active=True,company_code = company_code)
        empobjser=UserSerializer(empobj,many=True)
        totalteamcount=empobj.count()
        managerid="All"
        for j in empobjser.data:
            team_list.append(j['id'])
            if j['employeeId'] is not None and j['employeeId'] !="":
                team_with_attendanceid_list.append(j['employeeId'])





    leave_data=Leave.objects.filter(Active=True,employeeId__in=team_list,leave_status="Approved").order_by('id')
    leave_data_serializer = leaveserializer(leave_data, many=True)

    for l in leave_data_serializer.data:
        leave_userobj = Users.objects.filter(id=int(l['employeeId'])).first()

        date2 =date.fromisoformat(str(my_date))
        date1 =date.fromisoformat(l['start_date'])         
        date3 =date.fromisoformat(l['end_date']) 
        if date1 <= date2 and date2 <= date3:
            if l['WorkFromHome']:
                onwfhteamcount += 1

                wfhinfo={}
                wfhinfo['wfhcounter'] = onwfhteamcount
                wfhinfo['empname']= leave_userobj.Firstname +" "+ leave_userobj.Lastname
                wfhinfo['Leavetype']=l['leavetype']
                wfhlist.append(wfhinfo)
            else:
                onleaveteamcount += 1

                leaveinfo={}
                leaveinfo['leavecounter'] = onleaveteamcount
                leaveinfo['empname']= leave_userobj.Firstname +" "+ leave_userobj.Lastname
                leaveinfo['Leavetype']=l['leavetype']
                leavelist.append(leaveinfo)
                found_team_members.append(leave_userobj.id)

    present_count_obj=attendance.objects.filter(date=str(my_date),employeeId__in=team_with_attendanceid_list).distinct("employeeId")
    present_count_obj_ser=attendanceserializer(present_count_obj,many=True)
    presentcounter = 0
    Presentlist=[]
    for emp in present_count_obj_ser.data:
        att_userobj = Users.objects.filter(employeeId=str(emp['employeeId'])).first()
        presentinfo={}
        presentcounter += 1
        presentinfo['presentcounter'] = presentcounter
        presentinfo['empname']= att_userobj.Firstname +" "+ att_userobj.Lastname
        presentinfo['checkintime']= str(emp['time'])
        Presentlist.append(presentinfo)
        found_team_members.append(att_userobj.id)

    presentteamcount=present_count_obj.count()




    not_found_team_members = []
    for emp in team_list:
        if emp not in found_team_members:
            not_found_team_members.append(emp)
            

    ytjcounter = 0 
    for u in not_found_team_members:
        yettojoinobj = {}
        userobj = Users.objects.filter(id=int(u)).first()
        ytjcounter += 1

        if userobj.employeeId is not None and  userobj.employeeId != "":
            yettojoinobj['ytjcounter'] = ytjcounter
            yettojoinobj['username'] =  userobj.Firstname +" "+ userobj.Lastname
            yettojoinobj['reason'] =  "Not marked attendance"
        else:
            yettojoinobj['ytjcounter'] = ytjcounter
            yettojoinobj['username'] =  userobj.Firstname +" "+ userobj.Lastname
            yettojoinobj['reason'] =  "Attendance Id not found"
        
        yettojoinlist.append(yettojoinobj)

    








                
                
    yettojoin=totalteamcount-(onleaveteamcount + presentteamcount)
    
    context = {
        "All_Mapped_Emp_Count":totalteamcount,
        "Total_count" : totalteamcount,
        "Present":presentteamcount,
        "Onleave":onleaveteamcount,
        "yettojoin":yettojoin,
        "WFH":onwfhteamcount,
        "Presentlist":Presentlist,
        "leavelist":leavelist,
        "wfhlist":wfhlist,
        "yettojoinlist":yettojoinlist,
    }
    return Response ({"context":context, "response":{"n" : 1,"msg" : "SUCCESS","status" : "success"}})

@api_view(['POST'])
def get_per_date_leaves_count(request):
    logined_user=request.user.id
    year = request.POST.get('year')
    month = request.POST.get('month')
    
    company_code = request.user.company_code

    def get_dates_between(start_date, end_date):
        # Convert the input strings to date objects
        start_date = date.fromisoformat(start_date)
        end_date = date.fromisoformat(end_date)

        # Initialize an empty list to store the dates
        all_dates = []

        # Loop through the date range and add each date to the list
        current_date = start_date
        while current_date <= end_date:
            all_dates.append(current_date.strftime('%Y-%m-%d'))
            current_date += timedelta(days=1)

        return all_dates

    def count_dates_in_list(dates_list):
        # Initialize a dictionary with default value 0 for each date
        date_counts = defaultdict(int)

        # Loop through the list and count occurrences of each date
        for date in dates_list:
            date_counts[date] += 1

        return dict(date_counts)
    
    def filter_dictionary_by_month_year(input_dict, target_month, target_year):
        filtered_dict = {}
        for date_str, count in input_dict.items():
            year, month, day = map(int, date_str.split('-'))
            if int(month)<10:
                month="0"+str(month)

            if str(year) == str(target_year) and str(month) == str(target_month):
                day=int(day)
                
                filtered_dict["day"+str(day)] = count
        return filtered_dict
    




    role_obj = Role.objects.filter(Q(RoleName__iexact="admin") | Q(RoleName__iexact="core team"),company_code=company_code,Active=True)
    role_serializer=RoleIdSerializer(role_obj,many=True)
    distinct_ids = {item['id'] for item in role_serializer.data}
    coreteam1=Users.objects.filter(RoleID__in=distinct_ids)


    check_core_member=coreteam1.filter(id=logined_user).first()
    if check_core_member is not None:
        employees_obj=Users.objects.filter(company_code=company_code)
        employee_ids_serializers= UsersSerializeronlyid(employees_obj,many=True)
        employees_id_list=list(employee_ids_serializers.data)
    else:
        employees_obj=leaveMapping.objects.filter(company_code=company_code,managerId=logined_user).order_by('employeeId').distinct('employeeId')
        employee_ids_serializers= leave_mapping_employee_id_serializer(employees_obj,many=True)
        employees_id_list=list(employee_ids_serializers.data)+[logined_user]







    leave_dates =[]
    wfh_dates =[]

    first_day = date(int(year), int(month), 1)
    last_day = first_day.replace(day=28) + timedelta(days=4)
    last_day = last_day - timedelta(days=last_day.day)
    
    leaves_in_month = Leave.objects.filter(
        Q(start_date__lte=last_day, end_date__gte=first_day) |
        Q(start_date__gte=first_day, end_date__lte=last_day) | 
        Q(start_date__lte=first_day, end_date__gte=last_day) & 
        (Q(leave_status='Pending') | Q(leave_status='Approved')) 
        ).exclude((Q(leave_status='Draft') |Q(leave_status='Withdraw') | Q(leave_status='Rejected')))
    
    
    leaves_in_month=leaves_in_month.filter(employeeId__in=employees_id_list)

    leaves_in_month_obj=leaves_in_month.filter(Active=True,company_code=company_code,WorkFromHome=False)

    wfh_in_month_obj=leaves_in_month.filter(Active=True,company_code=company_code,WorkFromHome=True)
    


    leave_serializer=leaveserializer(leaves_in_month_obj,many=True)
    wfh_serializer=leaveserializer(wfh_in_month_obj,many=True)
    
    for i in leave_serializer.data:
        leavedates=get_dates_between(i['start_date'],i['end_date'])
        leave_dates=leave_dates+leavedates
      
    for i in wfh_serializer.data:
        wfhdates=get_dates_between(i['start_date'],i['end_date'])
        wfh_dates=wfh_dates+wfhdates
        
    leave_date_counts_dict = count_dates_in_list(leave_dates)
    leave_filtered_data = filter_dictionary_by_month_year(leave_date_counts_dict,month,year)
    
    wfh_date_counts_dict = count_dates_in_list(wfh_dates)
    wfh_filtered_data = filter_dictionary_by_month_year(wfh_date_counts_dict,month,year)
        
            
    combined_dates = {}
    for day in leave_filtered_data:

        combined_dates[day] = {
            'leave': leave_filtered_data.get(day, 0),
            'workfromhome': wfh_filtered_data.get(day, 0)
        }




    for day in wfh_filtered_data:
        if day not in combined_dates:
            combined_dates[day] = {
                'leave': leave_filtered_data.get(day, 0),
                'workfromhome': wfh_filtered_data.get(day, 0)
            }

        
    return Response({'n':1,'msg':'successfully','data':combined_dates,'status':'success'})

@api_view(['POST'])
def get_late_mark_attendance(request):
    com_code = request.user.company_code
    
    dates_list=[]
    start_date=request.POST.get("start_date")
    end_date=request.POST.get("end_date")

    if start_date is not None and start_date !="" and end_date is not None and end_date !="":
        dates_list=get_date_range(start_date,end_date)

    if start_date is None or start_date =="" and end_date is None or end_date =="":
        current_date = datetime.now().date()
        current_date_str = current_date.strftime("%Y-%m-%d")

        dates_list=get_date_range(current_date_str,current_date_str)

    userlist=[]    
    empId=request.POST.get("empId")
    if empId is None or empId =="":
        
        userobjs = Users.objects.filter(is_active=True,company_code=com_code).exclude(employeeId__isnull=True)
        userobjs_serializer=UserSerializer(userobjs,many=True)
        for i in userobjs_serializer.data:
            userlist.append(i['employeeId'])
    else:
        userobjs = Users.objects.filter(id=empId,is_active=True,company_code=com_code).exclude(employeeId__isnull=True).first()
        userobjs_serializer=UserSerializer(userobjs)
        userlist.append(userobjs_serializer.data['employeeId'])
        
    if len(userlist) == 0:
        
        return Response ({"data":{},"response":{"n" : 0,"msg" : "1 No data Found","status" : "failed"}})
    
    lateemplist=[]
    for date in dates_list:
        for emp in userlist:
            
            userinfo={}
            userobj = Users.objects.filter(employeeId=emp).first()
            lateobj = attendance.objects.filter(employeeId=emp,date=date,company_code=com_code).order_by('time').first()
            if lateobj is not None:
                latestr = lateobj.time
                timestr = str(latestr)
                timehours = timestr.split(":")[0]
                timemints = timestr.split(":")[1]

                if int(timehours) == 10 and int(timemints) > 0:
                    userinfo['empname'] = userobj.Firstname +" "+ userobj.Lastname
                    userinfo['intime']= timestr
                    userinfo['date']= convertdate(str(date))
                    userinfo['Id']= emp

                    lateemplist.append(userinfo)
                elif int(timehours) > 10 :
                    userinfo['empname'] = userobj.Firstname +" "+ userobj.Lastname
                    userinfo['intime']= timestr
                    userinfo['date']= convertdate(str(date))
                    userinfo['Id']= emp
                    lateemplist.append(userinfo)
 


        
        
    if len(lateemplist) == 0:
        return Response ({"data":{},"response":{"n" : 0,"msg" : "2 No data Found","status" : "failed"}})
    

    context={
        'Employees':lateemplist,
    }
    
    return Response ({"data":context,"response":{"n" : 1,"msg" : "Success","status" : "success"}})
       
@api_view(['POST'])
def get_past_application(request):
    LeaveId = request.POST.get('LeaveId')
    leave_status = request.POST.get('leave_status')
    UserId = request.POST.get('UserId')
    
    if leave_status=="Pending":
        
        leave_application_obj=Leave.objects.filter(employeeId=UserId,Active=True,leave_status="Pending").exclude(Q(id=LeaveId)|Q(leave_status="Draft"))
    elif leave_status=="Rejected":
        
        leave_application_obj=Leave.objects.filter(employeeId=UserId,Active=True,leave_status="Rejected").exclude(Q(id=LeaveId)|Q(leave_status="Draft"))
    elif leave_status=="Approved": 
               
        leave_application_obj=Leave.objects.filter(employeeId=UserId,Active=True,leave_status="Approved").exclude(Q(id=LeaveId)|Q(leave_status="Draft"))
    elif leave_status=="Withdraw":
        
        leave_application_obj=Leave.objects.filter(employeeId=UserId,Active=True,leave_status="Withdraw").exclude(Q(id=LeaveId)|Q(leave_status="Draft"))
    else:
        leave_application_obj=Leave.objects.filter(employeeId=UserId,Active=True).exclude(Q(id=LeaveId)|Q(leave_status="Draft"))
        
        
    if leave_application_obj !=[]:
            
        leave_application_obj_ser = leaveserializer(leave_application_obj,many=True)
        for i in leave_application_obj_ser.data:
            
            lstdate = str(i['start_date'])
            lsmonth_name = calendar.month_abbr[int(lstdate.split('-')[1])]    
            lsdatestr = lstdate.split('-')[2]+" "+lsmonth_name
            i['strstartdate'] = lsdatestr
            
            ledate = str(i['end_date'])
            lemonth_name = calendar.month_abbr[int(ledate.split('-')[1])]    
            ledatestr = ledate.split('-')[2]+" "+lemonth_name
            i['strenddate'] = ledatestr
            
            i['dates']=date_handling2(str(i['start_date']),str(i['end_date']))
            i['total_days']=countdays(str(i['start_date']),str(i['end_date']))
            if i['WorkFromHome'] == True:
                i['Application_Type']= "Work From Home"
            else:
                i['Application_Type']= "Leave"
            i['start_date']=convertdate2(i['start_date'])
            i['end_date']=convertdate2(i['end_date'])
        
            mapping_obj=leaveApproval.objects.filter(leave_id=i['id'])
            leave_approvel_serializer=leaveapprovalserializer(mapping_obj,many=True)
            managers_status=[]
                
            for t in leave_approvel_serializer.data:
                leave_approvel_obj=leaveApproval.objects.filter(managerId=t['managerId'],leave_id=t['leave_id']).first()
                manager_obj=Users.objects.filter(id=int(t['managerId'])).first()
                manager_ser=UserSerializer(manager_obj)

                if manager_ser.data['Photo'] is not None and manager_ser.data['Photo'] != "":
                    manager_photo = imageUrl+ str(manager_ser.data['Photo'])
                else:
                    manager_photo = imageUrl + "/static/assets/images/profile.png"
                if leave_approvel_obj is not None:
                    if leave_approvel_obj.approvedBy == True: 
                    
                        data1={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leave_approvel_obj.approvedBy,
                            "name":(manager_obj.Firstname+" "+manager_obj.Lastname).title(),
                            "rejectedBy":leave_approvel_obj.rejectedBy,
                            "comment":"",
                            "managerid":leave_approvel_obj.managerId,
                            "Status":"Approved",
                            "icon":"✅",
                            "manager_photo":manager_photo,
                        }
                        managers_status.append(data1)
                    elif leave_approvel_obj.rejectedBy == True:
                    
                        data1={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leave_approvel_obj.approvedBy,
                            "name":(manager_obj.Firstname+" "+manager_obj.Lastname).title(),
                            "rejectedBy":leave_approvel_obj.rejectedBy,
                            "comment":leave_approvel_obj.comment,
                            "managerid":leave_approvel_obj.managerId,
                            "Status":"Rejected",
                            "icon":"❌",

                            "manager_photo":manager_photo,

                        }
                        managers_status.append(data1)
                    else:
                        data1={
                            "applicationId":i['ApplicationId'],
                            "approvedBy":leave_approvel_obj.approvedBy,
                            "name":(manager_obj.Firstname+" "+manager_obj.Lastname).title(),
                            "rejectedBy":leave_approvel_obj.rejectedBy,
                            "comment":"",
                            "managerid":leave_approvel_obj.managerId,
                            "Status":"Pending",
                            "manager_photo":manager_photo,
                            "icon":"⏳",
                        }
                        managers_status.append(data1)                        

            i['managers_status']=managers_status
            
        context={
            'Applications' :leave_application_obj_ser.data,
        }
        return Response({'n':1,'msg':'data found Successfully.','status':'success','data':context})
    else:
        return Response({'n':0,'msg':'Leave not found','status':'failed','data':{}})
   
@api_view(['POST'])
def employee_application_details(request):
    LeaveId = request.POST.get('LeaveId')
    managerId = request.user.id 
    currentLeaveobj =  Leave.objects.filter(id=LeaveId,Active=True).first()
    if currentLeaveobj is not None:
        currentleaveser = leaveserializer(currentLeaveobj)
        serialized_data = currentleaveser.data.copy()  # Create a copy of the serialized data
        if managerId is not None and managerId !='':
            manager_obj=leaveApproval.objects.filter(leave_id=LeaveId,managerId=managerId).first()
            if manager_obj is not None:
                
                serialized_data['leaveapprovelid']=manager_obj.id
                if manager_obj.rejectedBy == False and manager_obj.approvedBy == False:
                    serialized_data['Manager_Actionable']=True
                else:
                    serialized_data['Manager_Actionable']=False


                lstdate = str(serialized_data['start_date'])
                lsmonth_name = calendar.month_abbr[int(lstdate.split('-')[1])]    
                lsdatestr = lstdate.split('-')[2]+" "+lsmonth_name
                serialized_data['strstartdate'] = lsdatestr
                
                ledate = str(serialized_data['end_date'])
                lemonth_name = calendar.month_abbr[int(ledate.split('-')[1])]    
                ledatestr = ledate.split('-')[2]+" "+lemonth_name
                serialized_data['strenddate'] = ledatestr
                
                serialized_data['dates']=date_handling2(str(serialized_data['start_date']),str(serialized_data['end_date']))
                serialized_data['total_days']=countdays(str(serialized_data['start_date']),str(serialized_data['end_date']))
                user_obj=Users.objects.filter(id=currentLeaveobj.employeeId).first()
                user_ser=UserSerializer(user_obj)
                serialized_data['Name']=user_ser.data['Firstname'] +" "+user_ser.data['Lastname']
                serialized_data['Designation']=user_ser.data['DesignationId']
                serialized_data['UserId']=user_ser.data['id']
                if user_ser.data['Photo'] is not None and user_ser.data['Photo'] != "":
                    serialized_data['Photo'] = imageUrl+ str(user_ser.data['Photo'])
                else:
                    serialized_data['Photo'] = imageUrl + "/static/assets/images/profile.png"
                    
                if serialized_data['WorkFromHome'] == True:
                    serialized_data['Application_Type']= "Work From Home"
                else:
                    serialized_data['Application_Type']= "Leave"
                
                
                mapping_obj=leaveApproval.objects.filter(leave_id=LeaveId)
                leave_approvel_serializer=leaveapprovalserializer(mapping_obj,many=True)
                managers_status=[]
                Application_Approval =[]  
                
                for t in leave_approvel_serializer.data:
                    leave_approvel_obj=leaveApproval.objects.filter(managerId=t['managerId'],leave_id=LeaveId).first()
                    manager_obj=Users.objects.filter(id=int(t['managerId'])).first()
                    manager_ser=UserSerializer(manager_obj)

                    if manager_ser.data['Photo'] is not None and manager_ser.data['Photo'] != "":
                        manager_photo = imageUrl+ str(manager_ser.data['Photo'])
                    else:
                        manager_photo = imageUrl + "/static/assets/images/profile.png"
                        
                    manager_applications_obj=leaveApproval.objects.filter(managerId=t['managerId'])
                    manager_applications_obj_serializer=leaveapprovalserializer(manager_applications_obj,many=True)
                    total_count=manager_applications_obj.count()
                    pending_count=0
                    for application in manager_applications_obj_serializer.data:
                        application_obj=Leave.objects.filter(Active=True,id=application['leave_id'],).exclude(leave_status="Rejected").first()
                        if application_obj:
                            if application['approvedBy'] == False and application['rejectedBy'] == False:
                                pending_count+=1
                            
                    Avg_score = ((total_count - pending_count) / total_count) * 100

                            
                    if leave_approvel_obj is not None:
                        if leave_approvel_obj.approvedBy == True: 
                        
                            data1={
                                "applicationId":currentLeaveobj.ApplicationId,
                                "approvedBy":leave_approvel_obj.approvedBy,
                                "name":(manager_obj.Firstname+" "+manager_obj.Lastname).title(),
                                "rejectedBy":leave_approvel_obj.rejectedBy,
                                "comment":"",
                                "managerid":leave_approvel_obj.managerId,
                                "Status":"Approved",
                                "icon":"✅",
                                "manager_photo":manager_photo,
                            }
                            data2={
                                "name":(manager_obj.Firstname+" "+manager_obj.Lastname).title(),
                                "managerid":leave_approvel_obj.managerId,
                                "Avg_score":round(Avg_score),
                                "manager_photo":manager_photo,

                            }
                            managers_status.append(data1)
                            Application_Approval.append(data2)
                            
                        elif leave_approvel_obj.rejectedBy == True:
                        
                            data1={
                                "applicationId":currentLeaveobj.ApplicationId,
                                "approvedBy":leave_approvel_obj.approvedBy,
                                "name":(manager_obj.Firstname+" "+manager_obj.Lastname).title(),
                                "rejectedBy":leave_approvel_obj.rejectedBy,
                                "comment":leave_approvel_obj.comment,
                                "managerid":leave_approvel_obj.managerId,
                                "Status":"Rejected",
                                "icon":"❌",
                                "manager_photo":manager_photo,

                            }
                            data2={
                                "name":(manager_obj.Firstname+" "+manager_obj.Lastname).title(),
                                "managerid":leave_approvel_obj.managerId,
                                "Avg_score":round(Avg_score),
                                "manager_photo":manager_photo,

                            }
                            managers_status.append(data1)
                            Application_Approval.append(data2)
                        else:
                            data1={
                                "applicationId":currentLeaveobj.ApplicationId,
                                "approvedBy":leave_approvel_obj.approvedBy,
                                "name":(manager_obj.Firstname+" "+manager_obj.Lastname).title(),
                                "rejectedBy":leave_approvel_obj.rejectedBy,
                                "comment":"",
                                "managerid":leave_approvel_obj.managerId,
                                "Status":"Pending",
                                "icon":"⏳",
                                "manager_photo":manager_photo,

                            }
                            data2={
                                "name":(manager_obj.Firstname+" "+manager_obj.Lastname).title(),
                                "managerid":leave_approvel_obj.managerId,
                                "Avg_score":round(Avg_score),
                                "manager_photo":manager_photo,

                            }
                            managers_status.append(data1)
                            Application_Approval.append(data2)                    
                    else:
                        data1={
                                "applicationId":currentLeaveobj.ApplicationId,
                                "approvedBy":"",
                                "name":(manager_obj.Firstname+" "+manager_obj.Lastname).title(),
                                "rejectedBy":"",
                                "comment":"",
                                "Status":"NA",                       
                                "icon":"⏳",
                                "manager_photo":manager_photo,
                            }
                        data2={
                            "name":(manager_obj.Firstname+" "+manager_obj.Lastname).title(),
                            "managerid":leave_approvel_obj.managerId,
                            "Avg_score":round(Avg_score),
                            "manager_photo":manager_photo,
                        }
                        managers_status.append(data1)
                        Application_Approval.append(data2)
                        
                serialized_data['managers_status']=managers_status

                Rejected_application_obj=Leave.objects.filter(employeeId=currentLeaveobj.employeeId,Active=True,leave_status="Rejected")
                Pending_application_obj=Leave.objects.filter(employeeId=currentLeaveobj.employeeId,Active=True,leave_status="Pending")
                Withdraw_application_obj=Leave.objects.filter(employeeId=currentLeaveobj.employeeId,Active=True,leave_status="Withdraw")
                Approved_application_obj=Leave.objects.filter(employeeId=currentLeaveobj.employeeId,Active=True,leave_status="Approved")
                Total_application_obj=Leave.objects.filter(employeeId=currentLeaveobj.employeeId,Active=True).exclude(leave_status="Draft")
                percentagelist=[]
                TrackPro_Percentages_History=[]
                def average_percentage(numbers):
                    if not numbers:
                        return 0.0  # Return 0 if the list is empty to avoid division by zero
                    total_percentage = sum(numbers) / len(numbers)
                    return round(total_percentage)
                
                Percentage_Object=IntermediateTrackProResult.objects.filter(Employee=currentLeaveobj.employeeId).order_by('-Week','Year')
                if Percentage_Object.exists():
                    Ser=IntermediateGetTrackProResultSerializer(Percentage_Object,many=True)
                    for i in Ser.data:
                        percentagelist.append(i['TrackProPercent'])
                        TrackPro_Percentages_History_dict={}
                        TrackPro_Percentages_History_dict['Week']=i['Week']
                        TrackPro_Percentages_History_dict['Year']=i['Year']
                        TrackPro_Percentages_History_dict['TrackProPercent']=i['TrackProPercent']
                        TrackPro_Percentages_History.append(TrackPro_Percentages_History_dict)
                        
                past_applications_obj=Leave.objects.filter(employeeId=currentLeaveobj.employeeId,Active=True).exclude(Q(id=LeaveId)|Q(leave_status="Draft")).order_by('-start_date')
                past_application_obj_ser = leaveserializer(past_applications_obj,many=True)
                for i in past_application_obj_ser.data:
                    
                    lstdate = str(i['start_date'])
                    lsmonth_name = calendar.month_abbr[int(lstdate.split('-')[1])]    
                    lsdatestr = lstdate.split('-')[2]+" "+lsmonth_name
                    i['strstartdate'] = lsdatestr
                    
                    ledate = str(i['end_date'])
                    lemonth_name = calendar.month_abbr[int(ledate.split('-')[1])]    
                    ledatestr = ledate.split('-')[2]+" "+lemonth_name
                    i['strenddate'] = ledatestr
                        
                    
                    
                    i['dates']=date_handling2(str(i['start_date']),str(i['end_date']))
                    i['total_days']=countdays(str(i['start_date']),str(i['end_date']))
                    if i['WorkFromHome'] == True:
                        i['Application_Type']= "Work From Home"
                    else:
                        i['Application_Type']= "Leave"
                    i['start_date']=convertdate2(i['start_date'])
                    i['end_date']=convertdate2(i['end_date'])
                
                    mapping_obj=leaveApproval.objects.filter(leave_id=i['id'])
                    leave_approvel_serializer=leaveapprovalserializer(mapping_obj,many=True)
                    managers_status=[]
                        
                    for t in leave_approvel_serializer.data:
                        leave_approvel_obj=leaveApproval.objects.filter(managerId=t['managerId'],leave_id=t['leave_id']).first()
                        i_manager_obj=Users.objects.filter(id=int(t['managerId'])).first()
                        i_manager_ser=UserSerializer(i_manager_obj)
                        if i_manager_ser.data['Photo'] is not None and i_manager_ser.data['Photo'] != "":
                            i_manager_photo = imageUrl+ str(i_manager_ser.data['Photo'])
                        else:
                            i_manager_photo = imageUrl + "/static/assets/images/profile.png"
                            
                        if leave_approvel_obj is not None:
                            if leave_approvel_obj.approvedBy == True: 
                            
                                data1={
                                    "applicationId":i['ApplicationId'],
                                    "approvedBy":leave_approvel_obj.approvedBy,
                                    "name":(i_manager_obj.Firstname+" "+i_manager_obj.Lastname).title(),
                                    "rejectedBy":leave_approvel_obj.rejectedBy,
                                    "comment":"",
                                    "managerid":leave_approvel_obj.managerId,
                                    "Status":"Approved",
                                    "icon":"✅",
                                    "manager_photo":i_manager_photo,
                                }
                                managers_status.append(data1)
                            elif leave_approvel_obj.rejectedBy == True:
                            
                                data1={
                                    "applicationId":i['ApplicationId'],
                                    "approvedBy":leave_approvel_obj.approvedBy,
                                    "name":(i_manager_obj.Firstname+" "+i_manager_obj.Lastname).title(),
                                    "rejectedBy":leave_approvel_obj.rejectedBy,
                                    "comment":leave_approvel_obj.comment,
                                    "managerid":leave_approvel_obj.managerId,
                                    "Status":"Rejected",
                                    "icon":"❌",
                                    "manager_photo":i_manager_photo,
                                }
                                managers_status.append(data1)
                            else:
                                data1={
                                    "applicationId":i['ApplicationId'],
                                    "approvedBy":leave_approvel_obj.approvedBy,
                                    "name":(i_manager_obj.Firstname+" "+i_manager_obj.Lastname).title(),
                                    "rejectedBy":leave_approvel_obj.rejectedBy,
                                    "comment":"",
                                    "managerid":leave_approvel_obj.managerId,
                                    "Status":"Pending",
                                    "icon":"⏳",
                                    "manager_photo":i_manager_photo,
                                }
                                managers_status.append(data1)                        
                        else:
                            data1={
                                    "applicationId":i['ApplicationId'],
                                    "approvedBy":"",
                                    "name":(i_manager_obj.Firstname+" "+i_manager_obj.Lastname).title(),
                                    "rejectedBy":"",
                                    "comment":"",
                                    "Status":"NA",                       
                                    "icon":"⏳",
                                    "manager_photo":i_manager_photo,
                                }
                            managers_status.append(data1)
                    
                    i['managers_status']=managers_status
                    
                context={
                    "TrackPro_Percentages_History":TrackPro_Percentages_History,
                    "TrackPro_average_percentage":average_percentage(percentagelist),
                    'Past_Applications' :past_application_obj_ser.data,
                    'currentleavedata' :serialized_data,
                    'Application_Approval':Application_Approval,
                    'Rejected_application_count':Rejected_application_obj.count(),
                    'Pending_application_count':Pending_application_obj.count(),
                    'Withdraw_application_count':Withdraw_application_obj.count(),
                    'Approved_application_count':Approved_application_obj.count(),
                    "Total_application_count":Total_application_obj.count(),
                }
                return Response({'n':1,'msg':'data found Successfully.','status':'success','data':context})
            
            else:
                return Response({'n':0,'msg':'access denied ','status':'failed','data':{}})  
        else:
            return Response({'n':0,'msg':'please login manager id not found','status':'failed','data':{}})  
    else:
        return Response({'n':0,'msg':'Leave not found','status':'failed','data':{}})   
       
@api_view(['POST'])
def get_late_count_per_month(request):
    employeeId=request.POST.get("employeeId")
    Month=request.POST.get("Month")
    Year=request.POST.get("Year")
    attendance_obj=attendance.objects.filter(Month=Month,Year=Year,employeeId=employeeId,time__gte='10:00:00',time__lte='14:00:00').distinct('date')
    serializer=attendanceserializer(attendance_obj,many=True)
    days=assign_value_based_on_length(len(serializer.data))





    response_={
            'status':0,
            'Msg':'success',
            'data':{
                "latecount":len(serializer.data),
                "days":days,

            }
        }
    return Response(response_,status=400)

@api_view(['POST'])
def get_leaves_and_latemarks_by_date(request):
    start_date=request.POST.get("start_date")
    end_date=request.POST.get("end_date")
    empId=request.POST.get("empId")
    com_code = request.user.company_code
    dates_list=[]

    if start_date is not None and start_date !="" and end_date is not None and end_date !="":
        dates_list=get_date_range(start_date,end_date)

        
    if start_date is None or start_date =="" and end_date is None or end_date =="":
        current_date = datetime.now().date()
        current_date_str = current_date.strftime("%Y-%m-%d")
        dates_list=get_date_range(current_date_str,current_date_str)
        start_date=current_date_str
        end_date=current_date_str

            
    userlist=[]    
    if empId is not None or empId !="":
        userobjs = Users.objects.filter(id=empId,is_active=True,company_code=com_code).exclude(employeeId__isnull=True).first()
        userobjs_serializer=UserSerializer(userobjs)
        userlist.append(userobjs_serializer.data['employeeId'])
    

    
    if len(userlist) == 0:
        latemark_context={"data":{'latemarks':[]},"response":{"n" : 0,"msg" : "1 No data Found","status" : "failed"}}
    

    leave_obj = Leave.objects.filter(
    Q(start_date__lte=end_date, end_date__gte=start_date,Active=True,employeeId=empId,WorkFromHome=False,company_code=com_code) |
    Q(start_date__gte=start_date, end_date__lte=end_date,Active=True,employeeId=empId,WorkFromHome=False,company_code=com_code) | 
    Q(start_date__lte=start_date, end_date__gte=end_date,Active=True,employeeId=empId,WorkFromHome=False,company_code=com_code)          
    ).exclude(leave_status='Draft').order_by('start_date__month','start_date__year')
    leave_serializer=leaves_extra_feilds_serializer(leave_obj,many=True)
    # Create a dictionary to store leave requests by user, month, and year
    leave_requests_by_user_month_year = defaultdict(list)

    # Populate the dictionary
    for leave_request in leave_serializer.data:
        user = leave_request['employeeId']
        month = leave_request['start_date_month_name']
        year = leave_request['start_date_year']
        

        mapping_obj=leaveApproval.objects.filter(leave_id=leave_request['id'])
        leave_approvel_serializer=leaveapprovalserializer(mapping_obj,many=True)
        managers_status=[]
        
        for t in leave_approvel_serializer.data:
            leave_approvel_obj=leaveApproval.objects.filter(managerId=t['managerId'],leave_id=t['leave_id']).first()
            i_manager_obj=Users.objects.filter(id=int(t['managerId'])).first()
            i_manager_ser=UserSerializer(i_manager_obj)
            
            if i_manager_ser.data['Photo'] is not None and i_manager_ser.data['Photo'] != "":
                i_manager_photo = imageUrl+ str(i_manager_ser.data['Photo'])
            else:
                i_manager_photo = imageUrl + "/static/assets/images/profile.png"
                
            if leave_approvel_obj is not None:
                if leave_approvel_obj.rejectedBy == True:
                
                    data1={
                        "applicationId":leave_request['ApplicationId'],
                        "approvedBy":leave_approvel_obj.approvedBy,
                        "name":(i_manager_obj.Firstname+" "+i_manager_obj.Lastname).title(),
                        "rejectedBy":leave_approvel_obj.rejectedBy,
                        "comment":leave_approvel_obj.comment,
                        "managerid":leave_approvel_obj.managerId,
                        "Status":"Rejected",
                        "icon":"❌",
                        "manager_photo":i_manager_photo,
                    }
                    managers_status.append(data1)
                elif leave_approvel_obj.approvedBy == True : 
                
                    data1={
                        "applicationId":leave_request['ApplicationId'],
                        "approvedBy":leave_approvel_obj.approvedBy,
                        "name":(i_manager_obj.Firstname+" "+i_manager_obj.Lastname).title(),
                        "rejectedBy":leave_approvel_obj.rejectedBy,
                        "comment":"",
                        "managerid":leave_approvel_obj.managerId,
                        "Status":"Approved",
                        "icon":"✅",
                        "manager_photo":i_manager_photo,
                    }
                    managers_status.append(data1)
                else:
                    data1={
                        "applicationId":leave_request['ApplicationId'],
                        "approvedBy":leave_approvel_obj.approvedBy,
                        "name":(i_manager_obj.Firstname+" "+i_manager_obj.Lastname).title(),
                        "rejectedBy":leave_approvel_obj.rejectedBy,
                        "comment":"",
                        "managerid":leave_approvel_obj.managerId,
                        "Status":"Pending",
                        "icon":"⏳",
                        "manager_photo":i_manager_photo,
                    }
                    managers_status.append(data1)                        
        leave_request['managers_status']=managers_status
        key = (user, month, year)
        leave_requests_by_user_month_year[str(key)].append(leave_request)
        
        



    organized_leave_requests = defaultdict(list)
    for key, leave_requests in leave_requests_by_user_month_year.items():
        for leave in leave_requests:
            month_year_key = (leave['start_date_month_name'], leave['start_date_year'])
            organized_leave_requests[month_year_key].append(leave)

    # Generate HTML based on the organized dictionary
    html_output = ""
    for (month, year), leave_requests in organized_leave_requests.items():
        html_output += f"""
        <div class="accordion-item">
            <h2 class="accordion-header" id="heading{month}_{year}">
            <button class="accordion-button" type="button" data-bs-toggle="collapse" data-bs-target="#collapse{month}_{year}" aria-expanded="true" aria-controls="collapse{month}_{year}">
                {month} {year}
            </button>
            </h2>
            <div id="collapse{month}_{year}" class="accordion-collapse collapse" aria-labelledby="heading{month}_{year}" data-bs-parent="#accordionExample">
            <div class="accordion-body">
        """

        for index, leave in enumerate(leave_requests):
            if leave['WorkFromHome'] == True:
                leave['Application_Type']="WFH"
            else:
                leave['Application_Type']="Leave"
            managerdiv=''
            for j in leave['managers_status']:
                managerdiv+='<span>'+j['icon']+' '+j['name']+'</span>'

            if leave['number_of_days'] is None or leave['number_of_days'] =='':
                leave['number_of_days']=''

                
            html_output += f"""

            
                        <div class="Application_div overflow-auto"> 
                            <div class="px-3 py-2 d-flex justify-content-between">
                                <div><span class="application_titles">Application ID</span> <br> <span class="application_values">{leave['ApplicationId']}</span></div>
                                <div><span class="application_titles">From</span><br> <span class="application_values">{convertdate2(leave['start_date'])} </span></div>
                                <div><span class="application_titles">To</span><br> <span class="application_values">{convertdate2(leave['end_date'])}</span></div>
                                
                                <div><span class="application_titles">{str(leave['number_of_days'])} Days</span> <br><span class="application_values"> {leave['leavetype']} </span></div>
                                <div><span class="application_titles">
                                Status                                    
                                <i class="cursor-pointer fa fa-solid fa-info-circle infobtn" data-bs-toggle="modal" data-bs-target="#getreasonofleave{leave['id']}" ></i>

                                <div class="modal fade" id="getreasonofleave{leave['id']}" tabindex="-1" aria-labelledby="exampleModalLabel" aria-hidden="true">
                                    <div class="modal-dialog modal-dialog-centered w-25">
                                    <div class="modal-content pb-2">
                                        <div class="modal-header ps-3 " style="padding-top:15px !important; justify-content: start !important;">
                                        <h5 class="modal-title" id="exampleModalLabel"> Reason</h5>
                                        </div>
                                        <div class="modal-body">
                                            <div class="row px-3">
                            
                                                
                                            Reason - {leave['reason']}
                            
                                            </div>
                                        </div>
                                        <div class="modal-footer d-flex">\
                                            {managerdiv}
                                        </div>
                                    </div>
                                    </div>
                                </div>


                                
                                </span> <br> <span class="application_values Application{leave['leave_status']}">{leave['leave_status']}</span></div>
                           




                      


                            </div>


                        </div>




           
           
             """

        html_output += """
            </div>
            </div>
        </div>
        """



    if leave_requests_by_user_month_year == {}:
        leave_context={"data":{'leaves':'<h6 class="text-center">No data Found</h6>'},"response":{"n" : 0,"msg" : "No data Found","status" : "failed"}}
    else:

        leave_context={"data":{ 'leaves':html_output},"response":{"n" : 1,"msg" : "Success","status" : "success"}}


    lateemplist=[]
    for date in dates_list:
        shift_details=get_date_shift_details(str(date),userlist[0])
        userobj = Users.objects.filter(employeeId__in=userlist).first()
        userinfo={}
        if shift_details['latemark']:
            userinfo['empname'] = userobj.Firstname +" "+ userobj.Lastname
            userinfo['intime']= shift_details['latetime']
            userinfo['date']= convertdate(str(date))
            userinfo['Id']= userlist
            lateemplist.append(userinfo)




 




        
    if len(lateemplist) == 0:
        latemark_context={"data":{'latemarks':[]},"response":{"n" : 0,"msg" : "2 No data Found","status" : "failed"}}
    else:
        latemark_context={"data":{ 'latemarks':lateemplist},"response":{"n" : 1,"msg" : "Success","status" : "success"}}

    context={
        'latemarks':latemark_context,
        'leaves':leave_context,
    }
    
    return Response ({"data":context,"response":{"n" : 1,"msg" : "Success","status" : "success"}})

# ================================================================= unused apis ==================================================================

def last_day_of_month(any_day):
    import datetime

    next_month = any_day.replace(day=28) + datetime.timedelta(days=4)
    return next_month - datetime.timedelta(days=next_month.day)

@api_view(['GET'])
def mappingJoinQuery(request):
    id = request.query_params.get('ManagerID', None)
    if id is not None:
        userobjects = UserToManager.objects.filter(ManagerID=id)
        Serializer = MappingSerializer(userobjects, many=True)
        return JsonResponse({'n': 1, 'Msg': 'Data fetched successfully', 'Status': 'Success','data':Serializer.data}, safe=False)
        # users = Users.objects.filter(id__in =Subquery(userobjects.values('id')) )
    else:
        return Response({'n': 0, 'Msg': 'Manager ID value is None', 'Status': 'Failed'})

@api_view(['GET'])
def leaveListAPI(request):

    logined_user=request.user.id
    company_code=request.user.company_code
    if request.method == 'GET':

        
        role_obj = Role.objects.filter(Q(RoleName__iexact="admin") | Q(RoleName__iexact="core team"),company_code=company_code,Active=True)
        role_serializer=RoleIdSerializer(role_obj,many=True)
        distinct_ids = {item['id'] for item in role_serializer.data}

        coreteam1=Users.objects.filter(RoleID__in=distinct_ids)

        mapped_managers_obj=leaveMapping.objects.filter(company_code=company_code).distinct("managerId")
        managers_serializer=leaveMappingserializer(mapped_managers_obj,many=True)
        distinct_mids = {item['managerId'] for item in managers_serializer.data}
        coreteam2=Users.objects.filter(id__in=distinct_mids)

        combined_coreteam = coreteam1 | coreteam2
        logined_user_obj=combined_coreteam.filter(id=logined_user).first()



        leavelist=[]
        superlist=[]
        exist_checkinglist=[]

        leave_data= Leave.objects.filter(Active=True,company_code=company_code,leave_status="Approved",WorkFromHome=False).order_by('-id')
        leave_data_serializer = leaveserializer(leave_data, many=True)
        for i in leave_data_serializer.data:
            superlist.append(i)

        leave_datapending= Leave.objects.filter(Active=True,company_code=company_code,leave_status="Pending",WorkFromHome=False).order_by('-id')
        leave_datapending_serializer = leaveserializer(leave_datapending, many=True)
        for i in leave_datapending_serializer.data:
            superlist.append(i)


        wfh_data= Leave.objects.filter(Active=True,company_code=company_code,leave_status="Approved",WorkFromHome=True).order_by('-id')
        wfh_data_serializer = leaveserializer(wfh_data, many=True)
        for i in wfh_data_serializer.data:
            superlist.append(i)

        wfh_datapending= Leave.objects.filter(Active=True,company_code=company_code,leave_status="Pending",WorkFromHome=True).order_by('-id')
        wfh_datapending_serializer = leaveserializer(wfh_datapending, many=True)
        for i in wfh_datapending_serializer.data:
            superlist.append(i)

            
        for i in superlist:
            
            leave_appprovel_obj=leaveApproval.objects.filter(employeeId=i['employeeId'],leave_id=i['id'],company_code=company_code)
            leave_appprovel_serializer=leaveapprovalserializer(leave_appprovel_obj,many=True)

            employee=Users.objects.filter(id=i['employeeId']).first()
            employee_serializer=UserSerializer(employee)
            i['Firstname']=employee_serializer.data['Firstname']
            i['Lastname']=employee_serializer.data['Lastname']
            start_date =  arrow.get(i['start_date'])
            end_date =  arrow.get(i['end_date'])
            delta =(start_date-end_date)
            i['days']=abs(delta.days)+1
            pstart_date = i['start_date'] 
            pendingstart_date = pstart_date.split('-')
            pstart_datestr = '-'.join(reversed(pendingstart_date))
            i['start_date'] = pstart_datestr
            pend_date = i['end_date'] 
            psplitend_date = pend_date.split('-')
            psplitend_datestr = '-'.join(reversed(psplitend_date))
            i['end_date'] = psplitend_datestr
            
            created = i['created_at']
            datet = created.split('T')[0]
            i['created_at']= dateformat(datet)
            

                    
            i['start_date']=ddmonthyy(i['start_date'])
            i['end_date']=ddmonthyy(i['end_date'])
            
            if i['start_date']==i['end_date']:
                i['leavedates']="<p class='leave-date'>"+i['start_date'] +"</p><div class='leave-date'></div>"
            else:
                i['leavedates']= "<div class='leave-date'>"+i['start_date'] +" to </div> <div class='leave-date'>" +i['end_date']+"</div>"

            if i['leavetype']=="Fullday":

                if i['days'] > 1:
                    i['leaveduration']="Total " + str(i['days']) + " Days"
                else:
                    i['leaveduration']="Full Day"
            else:
                if i['leavetype'] =="SecondHalf":
                    i['leaveduration']="Second Half"
                if i['leavetype'] =="FirstHalf":
                    i['leaveduration']="First Half"


            if i['WorkFromHome']==True:
                if logined_user_obj is None:
                    i['reason']="This employee is working from home. "
                i['applicationtype']=" Work From Home"
            else: 
                if logined_user_obj is None:
                    i['reason']="This employee is on leave."
                i['applicationtype']="Leave"

                    
            dummy_idlist=[]   
            count=1         
            managerlist_unique=[]    

            for j in leave_appprovel_serializer.data:

                managers=Users.objects.filter(id=j['managerId']).first()
                manager_serializer=UserSerializer(managers)
                j['position']=count
                j['Photo']=manager_serializer.data['Photo']
                j['Firstname']=manager_serializer.data['Firstname'] 
                j['Lastname']=manager_serializer.data['Lastname']
                count+=1
                if j['managerId'] not in dummy_idlist:
                    dummy_idlist.append(j['managerId'])
                    managerlist_unique.append(j)
                
                
            i['managerList']=managerlist_unique

            todaydate = str(date.today())
            # est={}
            # est['empid']=i['employeeId']
            # est['WorkFromHome']=i['WorkFromHome']

            # exist_checkinglist.append(est)
            # def is_employee_entry_exists(data, empid, work_from_home):
            #     for entry in data:
            #         if entry['empid'] == empid and entry['WorkFromHome'] == work_from_home:
            #             return True
            #     return False

                
            delta1 = end_date - start_date   
            for da in range(delta1.days + 1): 
                day1 = start_date + timedelta(days=da) 
                new_date=day1.format('YYYY-MM-DD')
                if  todaydate == new_date:
                    leavelist.append(i)
                    
        return Response ({"data":leavelist,"response":{"n" : 1,"msg" : "SUCCESS","status" : "success"}})

@api_view(['POST'])
def get_filter_leaves(request):
    logined_user=request.user.id
    company_code=request.user.company_code
    if request.method == 'POST':


        role_obj = Role.objects.filter(Q(RoleName__iexact="admin") | Q(RoleName__iexact="core team"),company_code=company_code,Active=True)
        role_serializer=RoleIdSerializer(role_obj,many=True)
        distinct_ids = {item['id'] for item in role_serializer.data}
        coreteam1=Users.objects.filter(RoleID__in=distinct_ids)

        mapped_managers_obj=leaveMapping.objects.filter(company_code=company_code).distinct("managerId")
        managers_serializer=leaveMappingserializer(mapped_managers_obj,many=True)
        distinct_mids = {item['managerId'] for item in managers_serializer.data}
        coreteam2=Users.objects.filter(id__in=distinct_mids)
        combined_coreteam = coreteam1 | coreteam2
        logined_user_obj=combined_coreteam.filter(id=logined_user).first()


        newlist=[] 
        searchdate=request.POST.get('searchdate')    
        superlist=[]


        leave_data= Leave.objects.filter(Active=True,company_code=company_code,leave_status="Approved",WorkFromHome=False).order_by('-id')
        leave_data_serializer = leaveserializer(leave_data, many=True)
        for i in leave_data_serializer.data:
            superlist.append(i)


        leave_datapending= Leave.objects.filter(Active=True,company_code=company_code,leave_status="Pending",WorkFromHome=False).order_by('-id')
        leave_datapending_serializer = leaveserializer(leave_datapending, many=True)
        for i in leave_datapending_serializer.data:
            superlist.append(i)

        wfh_data= Leave.objects.filter(Active=True,company_code=company_code,leave_status="Approved",WorkFromHome=True).order_by('-id')
        wfh_data_serializer = leaveserializer(wfh_data, many=True)
        for i in wfh_data_serializer.data:
            superlist.append(i)


        wfh_datapending= Leave.objects.filter(Active=True,company_code=company_code,leave_status="Pending",WorkFromHome=True).order_by('-id')
        wfh_datapending_serializer = leaveserializer(wfh_datapending, many=True)
        for i in wfh_datapending_serializer.data:
            superlist.append(i)


        for i in superlist:

            leave_appprovel_obj=leaveApproval.objects.filter(employeeId=i['employeeId'],leave_id=i['id'],company_code=company_code)
            leave_appprovel_serializer=leaveapprovalserializer(leave_appprovel_obj,many=True)
            
            employee=Users.objects.filter(id=i['employeeId']).first()
            employee_serializer=UserSerializer(employee)
            i['Firstname']=employee_serializer.data['Firstname']
            i['Lastname']=employee_serializer.data['Lastname']
            start_date =  arrow.get(i['start_date'])
            end_date =  arrow.get(i['end_date'])
            delta =(start_date-end_date)
            i['days']=abs(delta.days)+1


            
            dummy_idlist=[]   
            count=1         
            managerlist_unique=[]    

            for j in leave_appprovel_serializer.data:

                managers=Users.objects.filter(id=j['managerId']).first()
                manager_serializer=UserSerializer(managers)
                j['position']=count
                j['Photo']=manager_serializer.data['Photo']
                j['Firstname']=manager_serializer.data['Firstname'] 
                j['Lastname']=manager_serializer.data['Lastname']
                count+=1
                if j['managerId'] not in dummy_idlist:
                    dummy_idlist.append(j['managerId'])
                    managerlist_unique.append(j)
                
                
            i['managerList']=managerlist_unique

  
            d2 =date.fromisoformat(searchdate)
            d1 =date.fromisoformat(i['start_date'])         
            d3 =date.fromisoformat(i['end_date'])    
            i['formated_start_date']=convert_to_dd_month(i['start_date'])
            i['formated_end_date']=convert_to_dd_month(i['end_date'])
            pstart_date = i['start_date'] 
            pendingstart_date = pstart_date.split('-')
            pstart_datestr = '-'.join(reversed(pendingstart_date))
            i['start_date'] = pstart_datestr

            pend_date = i['end_date'] 
            psplitend_date = pend_date.split('-')
            psplitend_datestr = '-'.join(reversed(psplitend_date))
            i['end_date'] = psplitend_datestr
            
            created = i['created_at']
            datet = created.split('T')[0]
        
        
            i['formated_created_at']= convert_to_dd_month(datet)

            
            
            i['created_at']= dateformat(datet)
            i['start_date']=ddmonthyy(i['start_date'])
            i['end_date']=ddmonthyy(i['end_date'])


                
                
            if i['start_date']==i['end_date']:
                i['leavedates']="<div class='leave-date'>"+i['start_date'] +"</div>  <div class='leave-date'></div>"
            else:
                i['leavedates']= "<div class='leave-date'>"+i['start_date'] +" to </div> <div class='leave-date'>" +i['end_date']+"</div>"

            if i['leavetype']=="Fullday":

                if i['days'] > 1:
                    i['leaveduration']="Total " + str(i['days']) + " Days"
                else:
                    i['leaveduration']="Full Day "
            else:
                if i['leavetype'] =="SecondHalf":
                    i['leaveduration']="Second Half "
                if i['leavetype'] =="FirstHalf":
                    i['leaveduration']="First Half "

            if i['WorkFromHome']==True:
                if logined_user_obj is None:
                    i['reason']="This employee is working from home. "
                i['applicationtype']=" Work From Home"
            else: 
                if logined_user_obj is None:
                    i['reason']="This employee is on leave. "
                i['applicationtype']="Leave"
               


            if d1 <= d2 and d2 <= d3:
                newlist.append(i)


        return Response ({"data":newlist,"response":{"n" : 1,"msg" : "SUCCESS","status" : "success"}})

# ================================================================= unused apis ==================================================================



# =================================================================start custom functions ==================================================================

def convert_to_12_hour_format(time_str):
    # Convert the input time string to a datetime object
    if time_str !='' and time_str is not None and time_str !="--:--":
        
        time_obj = datetime.strptime(time_str, "%H:%M:%S")

        # Format the time object as "hh:mm am/pm"
        formatted_time = time_obj.strftime("%I:%M %p")

        return formatted_time
    return '--:--'

def convert_to_dd_month(date_str):
    # Parse the input date string
    date_obj = datetime.strptime(date_str, '%Y-%m-%d')

    # Format the date as 'DD Month' (e.g., '12 Aug')
    formatted_date = date_obj.strftime('%d %b')

    return formatted_date

def convertdate(date):
    if date != "" and date is not None:
        datetime_obj = datetime.strptime(date, "%Y-%m-%d")
        day = datetime_obj.day
        month = datetime_obj.strftime("%B")
        year = datetime_obj.year

        # convert day to a string with appropriate suffix
        if 4 <= day <= 20 or 24 <= day <= 30:
            suffix = "th"
        else:
            suffix = ["st", "nd", "rd"][day % 10 - 1]

        # format the date in "Dth Month YYYY" format
        formatted_date = f"{day}{suffix} {month} {year}"

        return formatted_date
    return "---"
 
def checkholiday(date):
    holidaylist=[]
    #holiday list
    holidatlist = Holidays.objects.filter(Active=True).order_by('id')
    serializer = holidaysSerializer(holidatlist, many=True)
    for i in serializer.data:
        holiday=i['Date']
        holidaylist.append(holiday) 

    for i in holidaylist:
        if i == str(date):
            return True 

def is_leave_time_valid(leave_date_str):
    current_datetime = datetime.now()
    leave_date = datetime.strptime(leave_date_str, '%Y-%m-%d')
    leave_datetime = datetime(leave_date.year, leave_date.month, leave_date.day, 9, 30)
    if current_datetime > leave_datetime:
        return False
    else:
        return True

def WorkFromHome(leaveid):
    status=""
    leave_obj=Leave.objects.filter(id=leaveid).first()
    if leave_obj.WorkFromHome ==True:
        status="work from home"
    else:
        status="leave"
    return status

def wfhsorter(leaveid):
    status=""
    leave_obj=Leave.objects.filter(id=leaveid).first()
    if leave_obj.WorkFromHome ==True:
        status="WFH"
    else:
        status="Leave"
    return status

def countdays(date1,date2):
    datetime1 = datetime.strptime(date1, "%Y-%m-%d")
    datetime2 = datetime.strptime(date2, "%Y-%m-%d")
    timedelta = datetime2 - datetime1
    days = timedelta.days+1
    return days 

def get_date_range(start_date_str, end_date_str):
    start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
    end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
    
    date_range = []
    current_date = start_date
    
    while current_date <= end_date:
        date_range.append(current_date)
        current_date += timedelta(days=1)
    
    return date_range

def calculate_days(date1, date2, date_type):
    date1 = datetime.strptime(date1, '%Y-%m-%d')
    date2 = datetime.strptime(date2, '%Y-%m-%d')

    # Calculate the number of days between two dates
    days_difference = (date2 - date1).days
    # Adjust for half day or full day
    if date_type == 'FirstHalf':
        days_difference += 0.5
    elif date_type == 'SecondHalf':
        days_difference += 0.5
    elif date_type == 'Fullday':
        days_difference += 1
    else:
        raise ValueError("Invalid date type. Use 'halfday' or 'fullday'.")

    return days_difference

def dateformat(input_date):
    
    datetime_obj = datetime.strptime(input_date, "%Y-%m-%d")
    day = datetime_obj.day
    month = datetime_obj.strftime("%B")
    year = datetime_obj.year
    if 4 <= day <= 20 or 24 <= day <= 30:
        suffix = "th"
    else:
        suffix = ["st", "nd", "rd"][day % 10 - 1]
    formatted_date = f"{day}{suffix} {month} {year}"
    return formatted_date
        
def date_handling(date_str1, date_str2):
    date1 = datetime.strptime(date_str1, "%Y-%m-%d")
    date2 = datetime.strptime(date_str2, "%Y-%m-%d")

    if date1 == date2:
        return convertdate(date1.strftime("%Y-%m-%d"))  # Return single date

    # Return date range
    return f"{convertdate(date1.strftime('%Y-%m-%d'))} to {convertdate(date2.strftime('%Y-%m-%d'))}"

def convertdate2(input_date):
    try:
        # Parse the input date string into a datetime object
        date_obj = datetime.strptime(input_date, '%Y-%m-%d')

        # Format the date in the desired output format
        formatted_date = date_obj.strftime('%d %b %y')

        return formatted_date
    except ValueError:
        return " "
    
def date_handling2(date_str1, date_str2):
    date1 = datetime.strptime(date_str1, "%Y-%m-%d")
    date2 = datetime.strptime(date_str2, "%Y-%m-%d")

    if date1 == date2:
        return convertdate2(date1.strftime("%Y-%m-%d"))  # Return single date

    # Return date range
    return f"{convertdate2(date1.strftime('%Y-%m-%d'))} to {convertdate2(date2.strftime('%Y-%m-%d'))}"

def dd_month_format(input_date):
    # Convert the input date string to a datetime object
    date_obj = datetime.strptime(input_date, "%Y-%m-%d")

    # Extract day, month, and year
    day = date_obj.strftime("%d")
    month = date_obj.strftime("%b")

    # Determine the day suffix (e.g., "st", "nd", "rd", "th")
    if day.endswith(('11', '12', '13')):
        suffix = "th"
    else:
        last_digit = int(day[-1])
        if last_digit == 1:
            suffix = "st"
        elif last_digit == 2:
            suffix = "nd"
        elif last_digit == 3:
            suffix = "rd"
        else:
            suffix = "th"

    # Format the result
    formatted_date = f"{day}{suffix} {month.upper()}"
    return formatted_date

def dd_month_year_format(input_date):
    # Parse the input date string
    date_obj = datetime.strptime(input_date, '%Y-%m-%d')
    
    # Define the month abbreviations
    months = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN',
              'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']
    
    # Get the day with an appropriate suffix
    day = date_obj.day
    if 4 <= day <= 20 or 24 <= day <= 30:
        day_suffix = 'th'
    else:
        day_suffix = ['st', 'nd', 'rd'][day % 10 - 1]
    
    # Format the date in the desired format
    formatted_date = f"{day:02d}{day_suffix} {months[date_obj.month - 1]} {date_obj.year}"
    
    return formatted_date

def assign_value_based_on_length(length):
    if length >= 27:
        return 5
    if length >= 27:
        return 4.5
    elif length >= 24:
        return 4
    elif length >= 21:
        return 3.5
    elif length >= 18:
        return 3
    elif length >= 15:
        return 2.5
    elif length >= 12:
        return 2
    elif length >= 9:
        return 1.5
    elif length >= 6:
        return 1
    elif length >= 3:
        return 0.5
    else:   
        return 0
# =================================================================end custom functions ==================================================================

def month_converter(month):
    month_num = month  # month_num = 4 will work too
    month_name = datetime(1, int(month_num), 1).strftime("%B")
    return month_name

def excelreport(data):
   
    workbook = xlsxwriter.Workbook('static/report/Attendancereport.xlsx')
    wb=load_workbook('static/report/Attendancereport.xlsx')
    sheet=wb.worksheets[0]

    sheet.cell(row=1,column=4).value="Attendance List Report"
    
    sheet.cell(row=2,column=1).value="Sr No"
    sheet.cell(row=2,column=2).value="Employee Code"
    sheet.cell(row=2,column=3).value="Employee Name"
    sheet.cell(row=2,column=4).value='Designation'
    sheet.cell(row=2,column=5).value='Department'
    sheet.cell(row=2,column=6).value='On Time'
    sheet.cell(row=2,column=7).value='Late'
    sheet.cell(row=2,column=8).value='Leave'

    k = 3
    counter = 1
    for i in data:
        sheet.cell(row=k,column=1).value=counter
        sheet.cell(row=k,column=2).value=i['empcode']
        sheet.cell(row=k,column=3).value=i['empname']
        sheet.cell(row=k,column=4).value=i['Designation']
        sheet.cell(row=k,column=5).value=i['departmentid']
        sheet.cell(row=k,column=6).value=i['ontimecount']
        sheet.cell(row=k,column=7).value=i['Leavecount']
        sheet.cell(row=k,column=8).value=i['Latecount']
        k+=1
        counter+=1
    
    wb.save('static/report/Attendancereport.xlsx')

@api_view(['POST'])
def get_filter_applications(request):
    leavemanager = False
    logined_user=request.user.id
    company_code=request.user.company_code
    mapped_managers_obj = leaveMapping.objects.filter(company_code=company_code).distinct("managerId")
    managers_serializer=leaveMappingserializer(mapped_managers_obj,many=True)
    distinct_mids = {item['managerId'] for item in managers_serializer.data}

    if str(logined_user) in list(distinct_mids):
        leavemanager = True
    else:
        leavemanager = False

    role_obj = Role.objects.filter(Q(RoleName__iexact="admin") | Q(RoleName__iexact="core team"),company_code=company_code,Active=True)
    role_serializer=RoleIdSerializer(role_obj,many=True)
    distinct_ids = {item['id'] for item in role_serializer.data}
    coreteam1=Users.objects.filter(RoleID__in=distinct_ids)

    mapped_managers_obj=leaveMapping.objects.filter(company_code=company_code).distinct("managerId")
    managers_serializer=leaveMappingserializer(mapped_managers_obj,many=True)
    distinct_mids = {item['managerId'] for item in managers_serializer.data}
    coreteam2=Users.objects.filter(id__in=distinct_mids)
    combined_coreteam = coreteam1 | coreteam2
    logined_user_obj=combined_coreteam.filter(id=logined_user).first()



    check_core_member=coreteam1.filter(id=logined_user).first()
    if check_core_member is not None:
        employees_obj=Users.objects.filter(company_code=company_code)
        employee_ids_serializers= UsersSerializeronlyid(employees_obj,many=True)
        employees_id_list=list(employee_ids_serializers.data)
    else:
        employees_obj=leaveMapping.objects.filter(company_code=company_code,managerId=logined_user).order_by('employeeId').distinct('employeeId')
        employee_ids_serializers= leave_mapping_employee_id_serializer(employees_obj,many=True)
        employees_id_list=list(employee_ids_serializers.data)+[logined_user]



    newlist=[] 
    wfh_count=0
    leave_count=0
    searchdate=request.POST.get('searchdate')  

    application_obj= Leave.objects.filter(employeeId__in=employees_id_list,Active=True,company_code=company_code,start_date__lte=searchdate,end_date__gte=searchdate).exclude(Q(leave_status="Rejected")|Q(leave_status="Withdraw")|Q(leave_status="Draft")).order_by('start_date','leavetype')
    serializer=leaveserializer(application_obj,many=True)

    for i in serializer.data:

        employee=Users.objects.filter(id=i['employeeId']).first()
        employee_serializer=UserSerializer(employee)

        i['Firstname']=employee_serializer.data['Firstname']
        i['Lastname']=employee_serializer.data['Lastname']

        start_date =  arrow.get(i['start_date']) 
        end_date =  arrow.get(i['end_date'])
        delta =(start_date-end_date)
        i['days']=abs(delta.days)+1





        
        dummy_idlist=[]   
        count=1         
        managerlist_unique=[]    
        i['leaveapprovalid']=0
        i['action_taken']=False
        leave_appprovel_obj=leaveApproval.objects.filter(employeeId=i['employeeId'],leave_id=i['id'],company_code=company_code)
        leave_appprovel_serializer=leaveapprovalserializer(leave_appprovel_obj,many=True)
        for j in leave_appprovel_serializer.data:
            managers=Users.objects.filter(id=j['managerId'],is_active=True).first()
            if managers is not None:
                
                if managers.id == request.user.id:
                    if j['approvedBy'] == False and j['rejectedBy'] == False:
                        i['action_taken']=False
                    else:
                        i['action_taken']=True
                        
                    i['leaveapprovalid']=j['id']
                    
                manager_serializer=UserSerializer(managers)
                j['position']=count
                j['Photo']=manager_serializer.data['Photo']
                j['Firstname']=manager_serializer.data['Firstname'] 
                j['Lastname']=manager_serializer.data['Lastname']
                count+=1
                if j['managerId'] not in dummy_idlist:
                    dummy_idlist.append(j['managerId'])
                    managerlist_unique.append(j)
            
            
        i['managerList']=managerlist_unique

        if str(logined_user) in dummy_idlist :
            i['showarrow'] = True
        else:
            i['showarrow'] = False



        d2 =date.fromisoformat(searchdate)
        d1 =date.fromisoformat(i['start_date'])         
        d3 =date.fromisoformat(i['end_date'])    
        i['formated_start_date']=convert_to_dd_month(i['start_date'])
        i['formated_end_date']=convert_to_dd_month(i['end_date'])
        pstart_date = i['start_date'] 
        pendingstart_date = pstart_date.split('-')
        pstart_datestr = '-'.join(reversed(pendingstart_date))
        i['start_date'] = pstart_datestr

        pend_date = i['end_date'] 
        psplitend_date = pend_date.split('-')
        psplitend_datestr = '-'.join(reversed(psplitend_date))
        i['end_date'] = psplitend_datestr
        
        created = i['created_at']
        datet = created.split('T')[0]
    
    
        i['formated_created_at']= convert_to_dd_month(datet)

        
        
        i['created_at']= dateformat(datet)
        i['start_date']=ddmonthyy(i['start_date'])
        i['end_date']=ddmonthyy(i['end_date'])


            
            
        if i['start_date']==i['end_date']:
            i['leavedates']="<div class='leave-date'>"+i['start_date'] +"</div>  <div class='leave-date'></div>"
        else:
            i['leavedates']= "<div class='leave-date'>"+i['start_date'] +" to </div> <div class='leave-date'>" +i['end_date']+"</div>"

        if i['leavetype']=="Fullday":

            if i['days'] > 1:
                i['leaveduration']="Total " + str(i['days']) + " Days"
            else:
                i['leaveduration']="Full Day "
        else:
            if i['leavetype'] =="SecondHalf":
                i['leaveduration']="Second Half "
            if i['leavetype'] =="FirstHalf":
                i['leaveduration']="First Half "


            
        

        if d1 <= d2 and d2 <= d3:
            if i['WorkFromHome']==True:
                wfh_count=wfh_count+1
                
                if logined_user_obj is None:
                    i['reason']="This employee is working from home. "
                i['applicationtype']=" Work From Home"
            else: 
                leave_count=leave_count+1
                if logined_user_obj is None:
                    i['reason']="This employee is on leave. "
                i['applicationtype']="Leave"
            newlist.append(i)




    get_compoff_obj=ClaimedCompOffMaster.objects.filter(Q(status='Approved')|Q(status='Pending'))
    given_datetime = datetime.strptime(searchdate + ' ' + '05:30:00', '%Y-%m-%d %H:%M:%S')
    get_compoff_obj=get_compoff_obj.filter(user_id__in=employees_id_list,claim_date=given_datetime)
    compoff_serializer=claimed_compoff_serializers(get_compoff_obj,many=True)
    for compoff in compoff_serializer.data:
        compoff['leavedates']=compoff['claim_date']
        compoff['leave_status']=compoff['status']
        compoff['leaveduration']='FullDay'
        compoff['reason']='Comp-off Claimed for the date of '+compoff['date']
        if logined_user_obj is None:
            compoff['reason']="This employee is on compensatory off. "
        compoff['leavetype']='FullDay'
        compoff['created_at']=compoff['created_at']
        compoff['applicationtype']='Comp-Off'
        compoff['number_of_days']=1
        compoff['Firstname']=compoff['user_name']
        compoff['Lastname']=''
        compoff['days']=1
        compoff['action_taken']=False
        compoff_approval_obj=CompoffApproval.objects.filter(claimed_compoff_id=compoff['id']).distinct('manager_id')
        compoff_managers_serializer=compoff_approval_serializers(compoff_approval_obj,many=True)
        for manager in compoff_managers_serializer.data:
            manager['Photo']=manager['manager_base_image']
            manager['Firstname']=manager['manager_name'].split(' ')[0]
            manager['Lastname']=manager['manager_name'].split(' ')[1]
            if manager['status']==True:
                manager['approvedBy']=True
                manager['rejectedBy']=False
            elif manager['status']==False:
                manager['approvedBy']=False
                manager['rejectedBy']=True
            else:
                manager['approvedBy']=False
                manager['rejectedBy']=False

        compoff['managerList']=compoff_managers_serializer.data
        newlist.append(compoff)


    return Response ({"data":newlist,"leave_count":leave_count,"wfh_count":wfh_count,"leavemanager":leavemanager,"response":{"n" : 1,"msg" : "SUCCESS","status" : "success"}})

def attendance_scheduler(request):
    my_date = date.today()
    month = my_date.month
    year, week_num, day_of_week = my_date.isocalendar()
    curryear = year
    currentweek = week_num
    
    com_code = request.user.company_code
    empobjs=Users.objects.all().order_by('id')
    empserializer = UserSerializer(empobjs,many=True)

    checkincount = 0  
    checkouttime = 0
    overtime = 0

    for i in empserializer.data:
        Attendance_id = i['employeeId']
        if Attendance_id is not None:
            attinobjs = attendance.objects.filter(employeeId=Attendance_id,date=my_date,company_code=com_code).order_by('time').first()
            if attinobjs is not None:
                checkintime = attinobjs.time
                strcheckintime = str(checkintime)
                inhours = strcheckintime.split(":")[0]
                inmins = strcheckintime.split(":")[1]
                insecs = strcheckintime.split(":")[2]

                if int(inhours) < 10 and int(inmins) <= 60 and int(insecs) <= 60:
                    checkincount += 1
                elif int(inhours) == 10 and int(inmins) == 0 and int(insecs) == 0:
                    checkincount += 1
                else:
                    checkincount += 0
                
            attoutobjs = attendance.objects.filter(employeeId=Attendance_id,date=my_date,company_code=com_code).order_by('time').last()
            if attoutobjs is not None:
                checkouttime = attoutobjs.time
                strcheckouttime = str(checkouttime)
                outhours = strcheckouttime.split(":")[0]
                outmins = strcheckouttime.split(":")[1]
                outsecs = strcheckouttime.split(":")[2]

                if int(outhours) == 7 and 30 <= int(outmins) <= 40:
                   checkouttime += 1
                elif  int(outhours) == 7 and  int(outmins) > 40:
                    overtime += 1
                elif int(outhours) > 7 :
                    overtime += 1
                else:
                    overtime += 0
                    
    Attendancecount.objects.create(Date=my_date,Year=curryear,Week=currentweek,Month=month,CheckIn=checkincount,checkOut=checkouttime,overtime=overtime,company_code=com_code)


    return Response({'n':1,'msg':'data stored successfully','status':'success'})

@api_view(['POST'])
def get_existing_applcation_dates(request):

    year = request.POST.get('year')
    month = request.POST.get('month')
    
    company_code = request.user.company_code

    def get_dates_between(start_date, end_date):
        # Convert the input strings to date objects
        start_date = date.fromisoformat(start_date)
        end_date = date.fromisoformat(end_date)

        # Initialize an empty list to store the dates
        all_dates = []

        # Loop through the date range and add each date to the list
        current_date = start_date
        while current_date <= end_date:
            all_dates.append(current_date.strftime('%Y-%m-%d'))
            current_date += timedelta(days=1)

        return all_dates

    def count_dates_in_list(dates_list):
        # Initialize a dictionary with default value 0 for each date
        date_counts = defaultdict(int)

        # Loop through the list and count occurrences of each date
        for date in dates_list:
            date_counts[date] += 1

        return dict(date_counts)
    

    leave_dates =[]
    wfh_dates =[]

    first_day = date(int(year), int(month), 1)
    last_day = first_day.replace(day=28) + timedelta(days=4)
    last_day = last_day - timedelta(days=last_day.day)
    
    leaves_in_month = Leave.objects.filter(
        Q(start_date__lte=last_day, end_date__gte=first_day) |
        Q(start_date__gte=first_day, end_date__lte=last_day) | 
        Q(start_date__lte=first_day, end_date__gte=last_day) & 
        (Q(leave_status='Pending') | Q(leave_status='Approved')) 
        ).exclude((Q(leave_status='Draft')|Q(leave_status='Withdraw') | Q(leave_status='Rejected')))
    
    
    leaves_in_month_obj=leaves_in_month.filter(Active=True,company_code=company_code,WorkFromHome=False)
    
    wfh_in_month_obj=leaves_in_month.filter(Active=True,company_code=company_code,WorkFromHome=True)
    
    leave_serializer=leaveserializer(leaves_in_month_obj,many=True)
    wfh_serializer=leaveserializer(wfh_in_month_obj,many=True)
    
    for i in leave_serializer.data:
        leavedates=get_dates_between(i['start_date'],i['end_date'])
        leave_dates=leave_dates+leavedates
      
    for i in wfh_serializer.data:
        wfhdates=get_dates_between(i['start_date'],i['end_date'])
        wfh_dates=wfh_dates+wfhdates
    
            
    combined_dates = []

    for day in leave_dates:
        if str(day) not in combined_dates:
            combined_dates.append(day)


    for day in wfh_dates:
        if str(day) not in combined_dates:
            combined_dates.append(day)
            
    def order_and_format_dates(date_list):
        # Use the sorted function with a custom key to sort the dates
        sorted_dates = sorted(date_list)

        # Create a list to store the formatted dates
        formatted_dates = []

        for date_str in sorted_dates:
            # Parse the date string into a datetime object
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            
            # Format the datetime object as "01/09/2023" and append to the list
            formatted_date = date_obj.strftime("%d/%m/%Y")
            formatted_dates.append(formatted_date)

        return formatted_dates
        
    formatted_dates = order_and_format_dates(combined_dates)

        
        
    return Response({'n':1,'msg':'successfully','data':formatted_dates,'status':'success'})

# api at 8 am
@api_view(['POST'])
@permission_classes((AllowAny,))
def leavemodulecronejob830Am(request):
   
    current_date = date.today()
    
    
    leave_obj=Leave.objects.filter(Active=True,leave_status="Pending",start_date__gte=current_date)
    leave_serializer=leaveserializer(leave_obj,many=True)

    for leave in leave_serializer.data:
        dates=date_handling(leave['start_date'],leave['end_date'])
        start_date = datetime.strptime(leave['start_date'], '%Y-%m-%d').date()
        days_difference = (start_date - current_date).days
        created = leave['created_at'].split('T')[0]
        created = datetime.strptime(created, '%Y-%m-%d').date()
        created_days_difference = (created - current_date).days
        empname_obj= Users.objects.filter(id=int(leave['employeeId']),is_active=True).first()
        empname = empname_obj.Firstname + " "+ empname_obj.Lastname
        empname=empname.title()
        Approvel_obj=False
        Status=False

        if int(days_difference) == 0:
            Team_members=[adminemail,hremail]
            leave_approval_object=leaveApproval.objects.filter(employeeId=leave['employeeId'],leave_id=leave['id'],company_code=leave['company_code'],approvedBy=False,rejectedBy=False)
            leave_approval_serializer = leaveapprovalserializer(leave_approval_object,many=True)
            for manager in leave_approval_serializer.data:
                m_obj = Users.objects.filter(id=int(manager['managerId']),is_active=True).first()
                if m_obj is not None:
                    Team_members.append(m_obj.email)
            any_manager_leave_approval_obj=leaveApproval.objects.filter(employeeId=leave['employeeId'],leave_id=leave['id'],company_code=leave['company_code'],approvedBy=True,rejectedBy=False).first()
            if any_manager_leave_approval_obj is not None:
                Approvel_obj=True

            # sending to account/admin/hr departments and managers  who had not taken action on day of leave/wfh
            # employee has applied for leave/wfh if no action is taken application will be approved at 09:30'
            try:             
                dicti = {
                    "employeename":empname,
                    "dates":dates,
                    "startdate":convertdate(leave['start_date']),
                    "enddate":convertdate(leave['end_date']),
                    "type":str(wfhsorter(leave['id'])),
                    "reason":leave['reason'],
                    "applicationid":leave['ApplicationId'],
                    "day_count":0,
                    "Approvel_obj":Approvel_obj,
                    "Status":Status,
                    
                }
                
                message = get_template(
                    'leavecronjobreminder.html').render(dicti)
                msg = EmailMessage(
                    ' URGENT: Pending '+ str(WorkFromHome(leave['id']))+' Application for ' +str(empname) + ' - Action Required by 09:30 AM'  ,
                    message,
                    EMAIL_HOST_USER,
                    Team_members,
                )
                msg.content_subtype = "html"  
                msg.send()
            except Exception as e:
                print('exception occured fot mail', e)
                               
                               
            # sending to employee that his application is pending please reach out to respective managers for his application on day of leave
            try:             
                dicti = {
                    "employeename":empname,
                    "dates":dates,
                    "applicationid":str(leave['ApplicationId']),
                    "startdate":convertdate(leave['start_date']),
                    "enddate":convertdate(leave['end_date']),
                    "type":str(wfhsorter(leave['id'])),
                    "reason":leave['reason'],
                }
                
                message = get_template(
                    'leavenoresponse.html').render(dicti)
                msg = EmailMessage(
                    'Important: Pending Response on Your '+ str(WorkFromHome(leave['id']))+' application',
                    message,
                    EMAIL_HOST_USER,
                    [str(empname_obj.email)],
                )
                msg.content_subtype = "html"  
                msg.send()
            except Exception as e:
                print('exception occured fot mail', e)               
                                                              
        elif int(days_difference) == 1:
            # sending to hr departments and managers  who had not taken action on day before leave/wfh
            Team_members=[hremail]
            leave_approval_object=leaveApproval.objects.filter(employeeId=leave['employeeId'],leave_id=leave['id'],company_code=leave['company_code'],approvedBy=False,rejectedBy=False)
            leave_approval_serializer = leaveapprovalserializer(leave_approval_object,many=True)
            for manager in leave_approval_serializer.data:
                m_obj = Users.objects.filter(id=int(manager['managerId']),is_active=True).first()
                if m_obj is not None:
                    Team_members.append(m_obj.email)
            try:             
                dicti = {
                    "employeename":empname,
                    "dates":dates,
                    "startdate":convertdate(leave['start_date']),
                    "enddate":convertdate(leave['end_date']),
                    "type":str(wfhsorter(leave['id'])),
                    "reason":leave['reason'],
                    "day_count":1,
                    "Approvel_obj":Approvel_obj,
                    "Status":Status,
                    "applicationid":leave['ApplicationId'],


                }
                message = get_template(
                    'leavecronjobreminder.html').render(dicti)
                
                msg = EmailMessage(
                    'Pending '+ str(WorkFromHome(leave['id']))+' application of ' + str(empname),
                    message,
                    EMAIL_HOST_USER,
                    Team_members,
                )
                msg.content_subtype = "html"  # Main content is now text/html
                msg.send()
            except Exception as e:
                print('exception occured fot mail', e)
                                
        elif int(days_difference) == 2:
            # sending to hr departments and managers  who had not taken action on 2 day before leave/wfh

            Team_members=[hremail]
            leave_approval_object=leaveApproval.objects.filter(employeeId=leave['employeeId'],leave_id=leave['id'],company_code=leave['company_code'],approvedBy=False,rejectedBy=False)
            leave_approval_serializer = leaveapprovalserializer(leave_approval_object,many=True)
            for manager in leave_approval_serializer.data:
                m_obj = Users.objects.filter(id=int(manager['managerId']),is_active=True).first()
                if m_obj is not None:
                    Team_members.append(m_obj.email)                
            try:             
                dicti = {
                    "employeename":empname,
                    "dates":dates,
                    "startdate":convertdate(leave['start_date']),
                    "enddate":convertdate(leave['end_date']),
                    "type":str(wfhsorter(leave['id'])),
                    "reason":leave['reason'],
                    "day_count":2,
                    "Approvel_obj":Approvel_obj,
                    "Status":Status,
                    "applicationid":leave['ApplicationId'],


                }
                message = get_template(
                    'leavecronjobreminder.html').render(dicti)
                msg = EmailMessage(
                    'Pending '+ str(WorkFromHome(leave['id']))+' application of ' + str(empname),
                    message,
                    EMAIL_HOST_USER,
                    Team_members,
                )
                msg.content_subtype = "html"  # Main content is now text/html
                msg.send()
            except Exception as e:
                print('exception occured fot mail', e)
            
        elif int(days_difference) == 7:
            # sending to hr departments and managers  who had not taken action on 7 day before leave/wfh

            Team_members=[hremail]
            leave_approval_object=leaveApproval.objects.filter(employeeId=leave['employeeId'],leave_id=leave['id'],company_code=leave['company_code'],approvedBy=False,rejectedBy=False)
            leave_approval_serializer = leaveapprovalserializer(leave_approval_object,many=True)
            for manager in leave_approval_serializer.data:
                m_obj = Users.objects.filter(id=int(manager['managerId']),is_active=True).first()
                if m_obj is not None:
                    Team_members.append(m_obj.email)
            try:             
                dicti = {
                    "employeename":empname,
                    "dates":dates,
                    "startdate":convertdate(leave['start_date']),
                    "enddate":convertdate(leave['end_date']),
                    "type":str(wfhsorter(leave['id'])),
                    "reason":leave['reason'],
                    "day_count":7,
                    "Approvel_obj":Approvel_obj,
                    "Status":Status,
                    "applicationid":leave['ApplicationId'],

                }
                message = get_template(
                    'leavecronjobreminder.html').render(dicti)
                msg = EmailMessage(
                    'Pending '+ str(WorkFromHome(leave['id']))+' application of ' + str(empname),
                    message,
                    EMAIL_HOST_USER,
                    Team_members,
                )
                msg.content_subtype = "html"  # Main content is now text/html
                msg.send()
            except Exception as e:
                print('exception occured fot mail', e)
                        
        elif int(created_days_difference) == -1:
            # sending to managers  who had not taken action on day after leave/wfh

            Team_members=[]
            leave_approval_object=leaveApproval.objects.filter(employeeId=leave['employeeId'],leave_id=leave['id'],company_code=leave['company_code'],approvedBy=False,rejectedBy=False)
            leave_approval_serializer = leaveapprovalserializer(leave_approval_object,many=True)
            for manager in leave_approval_serializer.data:
                m_obj = Users.objects.filter(id=int(manager['managerId']),is_active=True).first()
                if m_obj is not None:
                    Team_members.append(m_obj.email)                
            try:             
                dicti = {
                    "employeename":empname,
                    "dates":dates,
                    "startdate":convertdate(leave['start_date']),
                    "enddate":convertdate(leave['end_date']),
                    "type":str(wfhsorter(leave['id'])),
                    "reason":leave['reason'],
                    "day_count":-1,
                    "Approvel_obj":Approvel_obj,
                    "Status":Status,
                    "applicationid":leave['ApplicationId'],



                }
                message = get_template(
                    'leavecronjobreminder.html').render(dicti)
                msg = EmailMessage(
                    'Pending '+ str(WorkFromHome(leave['id']))+' application of ' + str(empname),
                    message,
                    EMAIL_HOST_USER,
                    Team_members,
                )
                msg.content_subtype = "html"  # Main content is now text/html
                msg.send()
            except Exception as e:
                print('exception occured fot mail', e)
            
            
            
            
            
            


    return Response ({"data":'',"response":{"n" : 1,"msg" : "Success","status" : "success"}})

# api at 9.30 am
@api_view(['POST'])
@permission_classes((AllowAny,))
def leavemodulecronejob930Am(request):
    # num=int(request.POST.get("num"))
    current_date = date.today()
    current_date = current_date - timedelta()

    leave_obj=Leave.objects.filter(Active=True,start_date=current_date).exclude(Q(leave_status="Rejected")|Q(leave_status="Withdraw")|Q(leave_status="Draft"))
    leave_serializer=leaveserializer(leave_obj,many=True)

    for leave in leave_serializer.data:
        # sending to admin/hr departments and managers   on day of leave/wfh that emp is on leave
        dates=date_handling(leave['start_date'],leave['end_date'])

        start_date = datetime.strptime(leave['start_date'], '%Y-%m-%d').date()
        days_difference = (start_date - current_date).days
        created = leave['created_at'].split('T')[0]
        created = datetime.strptime(created, '%Y-%m-%d').date()
        created_days_difference = (created - current_date).days
        empname_obj= Users.objects.filter(id=int(leave['employeeId']),is_active=True).first()
        empname = empname_obj.Firstname + " "+ empname_obj.Lastname
        empname=empname.title()

        Team_members=[adminemail,hremail]
        leave_approval_object=leaveApproval.objects.filter(employeeId=leave['employeeId'],leave_id=leave['id'],company_code=leave['company_code'])
        leave_approval_serializer = leaveapprovalserializer(leave_approval_object,many=True)
        for manager in leave_approval_serializer.data:
            m_obj = Users.objects.filter(id=int(manager['managerId']),is_active=True).first()
            if m_obj is not None:
                Team_members.append(m_obj.email)            
            
        approved_leave_approval_object=leaveApproval.objects.filter(employeeId=leave['employeeId'],leave_id=leave['id'],company_code=leave['company_code'],approvedBy=True,rejectedBy=False)

        if leave_approval_object.count() == approved_leave_approval_object.count():
            # to the team
            try:             
                dicti = {
                    "employeename":empname,
                    "dates":dates,
                    "startdate":convertdate(leave['start_date']),
                    "enddate":convertdate(leave['end_date']),
                    "type":str(wfhsorter(leave['id'])),
                    "reason":leave['reason'],
                    "leave_status":"Approved",
                    "applicationid":leave['ApplicationId'],
                }
                
                message = get_template(
                    'leavecronjobapproved.html').render(dicti)
                msg = EmailMessage(
                    str(empname) +'  '+ str(WorkFromHome(leave['id'])) +' application has been approved',
                    message,
                    EMAIL_HOST_USER,
                    Team_members,
                )
                msg.content_subtype = "html"  # Main content is now text/html
                msg.send()
            except Exception as e:
                print('exception occured fot mail', e)
                
                
            # to the employee
            try:             
                dicti = {
                    "employeename":empname,
                    "dates":dates,
                    "startdate":convertdate(leave['start_date']),
                    "enddate":convertdate(leave['end_date']),
                    "type":str(wfhsorter(leave['id'])),
                    "reason":leave['reason'],
                    "leave_status":"Approved",
                    "applicationid":str(leave['ApplicationId']),
                }
                
                message = get_template(
                    'leaveshedularondayofleave.html').render(dicti)
                msg = EmailMessage(
                    'Your '+ str(WorkFromHome(leave['id'])) +' application has been approved',
                    message,
                    EMAIL_HOST_USER,
                    [str(empname_obj.email)],
                )
                msg.content_subtype = "html"  # Main content is now text/html
                msg.send()
            except Exception as e:
                print('exception occured fot mail', e)       
                
        elif approved_leave_approval_object.count() >=1:
            # leaveApproval.objects.filter(employeeId=leave['employeeId'],leave_id=leave['id'],company_code=leave['company_code']).update(approvedBy=True,rejectedBy=False)

            Leave.objects.filter(id=leave['id'],Active=True).update(leave_status="Approved")
            TaskNotification.objects.filter(leaveID=leave['id'],To_manager=True,action_Taken=False).update(action_Taken=True)


            # to the team
            try:             
                dicti = {
                    "employeename":empname,
                    "dates":dates,
                    "startdate":convertdate(leave['start_date']),
                    "enddate":convertdate(leave['end_date']),
                    "type":str(wfhsorter(leave['id'])),
                    "reason":leave['reason'],
                    "leave_status":"Approved",
                    "applicationid":leave['ApplicationId'],

                }
                
                message = get_template(
                    'leavecronjobapproved.html').render(dicti)
                msg = EmailMessage(
                    str(empname) +'  '+ str(WorkFromHome(leave['id'])) +' application has been Approved',
                    message,
                    EMAIL_HOST_USER,
                    Team_members,
                )
                msg.content_subtype = "html"  # Main content is now text/html
                msg.send()
            except Exception as e:
                print('exception occured fot mail', e)
                
                
                
            # to the employee
            try:             
                dicti = {
                    "employeename":empname,
                    "dates":dates,
                    "startdate":convertdate(leave['start_date']),
                    "enddate":convertdate(leave['end_date']),
                    "type":str(wfhsorter(leave['id'])),
                    "reason":leave['reason'],
                    "leave_status":"Approved",
                    "applicationid":str(leave['ApplicationId']),
                }
                
                message = get_template(
                    'leaveshedularondayofleave.html').render(dicti)
                msg = EmailMessage(
                    'Your '+ str(WorkFromHome(leave['id'])) +' application has been approved',
                    message,
                    EMAIL_HOST_USER,
                    [str(empname_obj.email)],
                )
                msg.content_subtype = "html"  # Main content is now text/html
                msg.send()
            except Exception as e:
                print('exception occured fot mail', e)

        # elif approved_leave_approval_object.count() == 0:
            # leaveApproval.objects.filter(employeeId=leave['employeeId'],leave_id=leave['id'],company_code=leave['company_code'],approvedBy=False).update(rejectedBy=True,comment="Not responded on time")
            # Leave.objects.filter(id=leave['id'],Active=True).update(leave_status="Rejected")
            
            # to the team
            # try:             
            #     dicti = {
            #         "employeename":empname,
            #         "dates":dates,
            #         "startdate":convertdate(leave['start_date']),
            #         "enddate":convertdate(leave['end_date']),
            #         "type":str(wfhsorter(leave['id'])),
            #         "reason":leave['reason'],
            #         "leave_status":"Rejected",
            #         "applicationid":leave['ApplicationId'],


            #     }
                
            #     message = get_template(
            #         'leavecronjobapproved.html').render(dicti)
            #     msg = EmailMessage(
            #         str(empname) +'  '+ str(WorkFromHome(leave['id'])) +' application has been rejected',
            #         message,
            #         EMAIL_HOST_USER,
            #         Team_members,
            #     )
            #     msg.content_subtype = "html"  # Main content is now text/html
            #     msg.send()
            # except Exception as e:
            #     print('e2xception occured fot mail', e)
                
            
            # # to the employeee
            # try:             
            #     dicti = {
            #         "employeename":empname,
            #         "dates":dates,
            #         "startdate":convertdate(leave['start_date']),
            #         "enddate":convertdate(leave['end_date']),
            #         "type":str(wfhsorter(leave['id'])),
            #         "reason":leave['reason'],
            #         "leave_status":"Rejected",
            #         "applicationid":str(leave['ApplicationId']),
                    
            #     }
                
            #     message = get_template(
            #         'leaveshedularondayofleave.html').render(dicti)
            #     msg = EmailMessage(
            #         'Your '+ str(WorkFromHome(leave['id'])) +' application has been rejected',
            #         message,
            #         EMAIL_HOST_USER,
            #         [str(empname_obj.email)],
            #     )
            #     msg.content_subtype = "html"  # Main content is now text/html
            #     msg.send()
            # except Exception as e:
            #     print('1exception occured fot mail', e)
                

    return Response ({"data":'',"response":{"n" : 1,"msg" : "Success","status" : "success"}})




@csrf_exempt
@api_view(['POST'])
@authentication_classes([])
@permission_classes([])
def team_attendance_scheduler(request):
    my_date = date.today()
    # my_date = request.POST.get('my_date')
    
    company_code =  "O001"
    on_leave_count=0
    on_wfh_count=0
    total_count=0
    present_count=0
    not_found_count=0
    without_attendanceid_total_count=0
    with_attendanceid_total_count=0
    
    on_leave_employees=[]
    on_wfh_employees=[]
    total_employees=[]
    present_employees=[]
    not_found_employees=[]
    without_attendanceid_total_employees=[]
    with_attendanceid_total_employees=[]
    
    
    with_attendanceid_total_count=Users.objects.filter(is_active=True,company_code=company_code).exclude(employeeId__isnull=True).order_by('id')
    with_attendanceid_total_count_ser=UserSerializer(with_attendanceid_total_count,many=True)
    for emp in with_attendanceid_total_count_ser.data:
        with_attendanceid_total_employees.append(emp['id'])
    with_attendanceid_total_count=with_attendanceid_total_count.count()

    without_attendanceid_total_count=Users.objects.filter(is_active=True,company_code=company_code).exclude(employeeId__isnull=False).order_by('id')
    without_attendanceid_total_count_ser=UserSerializer(without_attendanceid_total_count,many=True)
    for emp in without_attendanceid_total_count_ser.data:
        without_attendanceid_total_employees.append(emp['id'])
    without_attendanceid_total_count=without_attendanceid_total_count.count()

    leave_data=Leave.objects.filter(Active=True,company_code=company_code,leave_status="Approved").order_by('id')
    leave_data_serializer = leaveserializer(leave_data, many=True)
    
    for l in leave_data_serializer.data:
        d2 =date.fromisoformat(str(my_date))
        d1 =date.fromisoformat(l['start_date'])         
        d3 =date.fromisoformat(l['end_date']) 
        if d1 <= d2 and d2 <= d3:
            if l['WorkFromHome']:
                on_wfh_employees.append(l['employeeId'])
                on_wfh_count += 1
            else:
                on_leave_employees.append(l['employeeId'])
                on_leave_count += 1
                
                
    present_count=attendance.objects.filter(date=str(my_date),company_code=company_code)
    present_count_ser=attendanceserializer(present_count,many=True)
    for emp in present_count_ser.data:
        present_employees.append(emp['employeeId'])
    present_count=present_count.count()


    total_employees=without_attendanceid_total_employees + with_attendanceid_total_employees
    found_list=on_leave_employees + present_employees
    not_found_employees=list(set(total_employees) - set(found_list))

    
    
    total_count=without_attendanceid_total_count + with_attendanceid_total_count
    not_found_count=total_count - (on_leave_count + present_count)
    
    context={
        "managerid":"All",
        "team_hold_by":"All",
        "date":my_date,
        "present_count":present_count,
        "total_count":total_count,
        "on_leave_count":on_leave_count,
        "on_wfh_count":on_wfh_count,
        "not_found_count":not_found_count,
        "without_attendanceid_total_count":without_attendanceid_total_count,
        "with_attendanceid_total_count":with_attendanceid_total_count,
        "company_code":company_code,
        
        "on_leave_employees":on_leave_employees,
        "on_wfh_employees":on_wfh_employees,
        "total_employees":total_employees,
        "present_employees":present_employees,
        "not_found_employees":not_found_employees,
        "without_attendanceid_total_employees":without_attendanceid_total_employees,
        "with_attendanceid_total_employees":with_attendanceid_total_employees,
        
    }
    
    alredyexist_obj=TeamAttendance.objects.filter(managerid=context['managerid'],date=context['date']).order_by('time').first()
    if alredyexist_obj:
        TeamAttendance.objects.filter(managerid=context['managerid'],date=context['date']).update(
                                team_hold_by=context['team_hold_by'],
                                present_count=context['present_count'],
                                total_count=context['total_count'],
                                on_leave_count=context['on_leave_count'],
                                on_wfh_count=context['on_wfh_count'],
                                not_found_count=context['not_found_count'],
                                without_attendanceid_total_count=context['without_attendanceid_total_count'],
                                with_attendanceid_total_count=context['with_attendanceid_total_count'],
                                company_code=context['company_code'],
                                
                                
                                present_employees= ",".join(map(str, context['present_employees'])),
                                on_leave_employees= ",".join(map(str, context['on_leave_employees'])),
                                on_wfh_employees=",".join(map(str, context['on_wfh_employees'])),
                                not_found_employees=",".join(map(str, context['not_found_employees'])),
                                total_employees=",".join(map(str, context['total_employees'])),
                                without_attendanceid_employees=",".join(map(str, context['without_attendanceid_total_employees'])),
                                with_attendanceid_employees=",".join(map(str, context['with_attendanceid_total_employees'])),
                                
                                )
    else:
        TeamAttendance.objects.create(managerid=context['managerid'],
                                    team_hold_by=context['team_hold_by'],
                                    date=context['date'],
                                    present_count=context['present_count'],
                                    total_count=context['total_count'],
                                    on_leave_count=context['on_leave_count'],
                                    on_wfh_count=context['on_wfh_count'],
                                    not_found_count=context['not_found_count'],
                                    without_attendanceid_total_count=context['without_attendanceid_total_count'],
                                    with_attendanceid_total_count=context['with_attendanceid_total_count'],
                                    company_code=context['company_code'],
                                    
                                    present_employees= ",".join(map(str, context['present_employees'])),
                                    on_leave_employees= ",".join(map(str, context['on_leave_employees'])),
                                    on_wfh_employees=",".join(map(str, context['on_wfh_employees'])),
                                    not_found_employees=",".join(map(str, context['not_found_employees'])),
                                    total_employees=",".join(map(str, context['total_employees'])),
                                    without_attendanceid_employees=",".join(map(str, context['without_attendanceid_total_employees'])),
                                    with_attendanceid_employees=",".join(map(str, context['with_attendanceid_total_employees'])),
                                    
                                    )
    
    allemployees=Users.objects.filter(is_active=True,company_code=company_code).order_by('id')
    allemployees_serializer=UserSerializer(allemployees,many=True)
    for i in allemployees_serializer.data:
        i_on_leave_count=0
        i_on_wfh_count=0
        i_total_count=0
        i_present_count=0
        i_not_found_count=0
        
        i_without_attendanceid_total_count=0
        i_with_attendanceid_total_count=0
        i_Manager=i['Firstname'] + ' ' + i['Lastname']
        i_ManagerId=i['id']      
        
        
        i_on_leave_employees=[]
        i_on_wfh_employees=[]
        i_total_employees=[]
        i_present_employees=[]
        i_not_found_employees=[]
        i_without_attendanceid_total_employees=[]
        i_with_attendanceid_total_employees=[]
        
               
        mappingobj = UserToManager.objects.filter(ManagerID=i['id'],company_code = i['company_code'])
        team_list=[]
        attendance_id_list=[]
        if mappingobj.exists():

            
            map_serializer = MappingSerializer(mappingobj, many=True)
            for j in map_serializer.data:
                team_list.append(j['UserID'])
               
            for team_member in team_list:
                team_member_obj=Users.objects.filter(id=team_member,is_active=True,company_code=company_code).order_by('id').first()
                if team_member_obj:
                    if team_member_obj.employeeId is not None and team_member_obj.employeeId !="":
                        i_with_attendanceid_total_count += 1
                        i_without_attendanceid_total_employees.append(team_member_obj.id)
                        attendance_id_list.append(team_member_obj.employeeId)
                    else:
                        i_with_attendanceid_total_employees.append(team_member_obj.id)
                        i_without_attendanceid_total_count += 1
                    
            i_present_count=attendance.objects.filter(employeeId__in=attendance_id_list,date=str(my_date))
            i_present_count_ser=attendanceserializer(i_present_count,many=True)
            i_present_count=i_present_count.count()

            for emp in i_present_count_ser.data:
                i_present_employees.append(emp['employeeId'])
            
            i_leave_data=Leave.objects.filter(employeeId__in=team_list,Active=True,leave_status="Approved").order_by('id')
            i_leave_data_serializer = leaveserializer(i_leave_data, many=True)
            
            for l in i_leave_data_serializer.data:
                i_d2 =date.fromisoformat(str(my_date))
                i_d1 =date.fromisoformat(l['start_date'])         
                i_d3 =date.fromisoformat(l['end_date']) 
                if i_d1 <= i_d2 and i_d2 <= i_d3:
                    if l['WorkFromHome']:
                        i_on_wfh_count += 1
                        i_on_wfh_employees.append(l['employeeId'])
                    else:
                        i_on_leave_count += 1  
                        i_on_leave_employees.append(l['employeeId'])
                        
            i_total_count=len(team_list)
            i_total_employees=team_list
            i_found=i_on_leave_employees + i_present_employees
            i_not_found_employees=list(set(i_total_employees) - set(i_found))
            i_not_found_count=i_total_count - (i_on_leave_count + i_present_count)
            
            i_context={
                "managerid":i_ManagerId,
                "team_hold_by":i_Manager,
                "date":my_date,
                "present_count":i_present_count,
                "total_count":i_total_count,
                "on_leave_count":i_on_leave_count,
                "on_wfh_count":i_on_wfh_count,
                "not_found_count":i_not_found_count,
                "with_attendanceid_total_count":i_with_attendanceid_total_count,
                "without_attendanceid_total_count":i_without_attendanceid_total_count,
                "company_code":company_code,
                
                
                "on_leave_employees":i_on_leave_employees,
                "on_wfh_employees":i_on_wfh_employees,
                "total_employees":i_total_employees,
                "present_employees":i_present_employees,
                "not_found_employees":i_not_found_employees,
                "without_attendanceid_total_employees":i_without_attendanceid_total_employees,
                "with_attendanceid_total_employees":i_with_attendanceid_total_employees,
                
                
            }
    
            i_alredyexist_obj=TeamAttendance.objects.filter(managerid=i_context['managerid'],date=i_context['date']).order_by('time').first()
            if i_alredyexist_obj:
                TeamAttendance.objects.filter(managerid=i_context['managerid'],date=i_context['date']).update(
                            team_hold_by=i_context['team_hold_by'],
                            present_count=i_context['present_count'],
                            total_count=i_context['total_count'],
                            on_leave_count=i_context['on_leave_count'],
                            on_wfh_count=i_context['on_wfh_count'],
                            not_found_count=i_context['not_found_count'],
                            without_attendanceid_total_count=i_context['without_attendanceid_total_count'],
                            with_attendanceid_total_count=i_context['with_attendanceid_total_count'],
                            company_code=i_context['company_code'],
                           
                            present_employees= ",".join(map(str, i_context['present_employees'])),
                            on_leave_employees= ",".join(map(str, i_context['on_leave_employees'])),
                            on_wfh_employees=",".join(map(str, i_context['on_wfh_employees'])),
                            not_found_employees=",".join(map(str, i_context['not_found_employees'])),
                            total_employees=",".join(map(str, i_context['total_employees'])),
                            without_attendanceid_employees=",".join(map(str, i_context['without_attendanceid_total_employees'])),
                            with_attendanceid_employees=",".join(map(str, i_context['with_attendanceid_total_employees'])),
                            )
            else:
                    
                TeamAttendance.objects.create(managerid=i_context['managerid'],
                                            team_hold_by=i_context['team_hold_by'],
                                            date=i_context['date'],
                                            present_count=i_context['present_count'],
                                            total_count=i_context['total_count'],
                                            on_leave_count=i_context['on_leave_count'],
                                            on_wfh_count=i_context['on_wfh_count'],
                                            not_found_count=i_context['not_found_count'],
                                            without_attendanceid_total_count=i_context['without_attendanceid_total_count'],
                                            with_attendanceid_total_count=i_context['with_attendanceid_total_count'],
                                            company_code=i_context['company_code'],
                                            
                                            present_employees= ",".join(map(str, i_context['present_employees'])),
                                            on_leave_employees= ",".join(map(str, i_context['on_leave_employees'])),
                                            on_wfh_employees=",".join(map(str, i_context['on_wfh_employees'])),
                                            not_found_employees=",".join(map(str, i_context['not_found_employees'])),
                                            total_employees=",".join(map(str, i_context['total_employees'])),
                                            without_attendanceid_employees=",".join(map(str, i_context['without_attendanceid_total_employees'])),
                                            with_attendanceid_employees=",".join(map(str, i_context['with_attendanceid_total_employees'])),
                                            )
    
    
    
    
    my_date=str(my_date)


    return Response({'n':1,'msg':'data stored successfully of '  + my_date,'status':'success','data':{}})

@csrf_exempt
@api_view(['POST'])
@authentication_classes([])
@permission_classes([])
def team_attendance_scheduler_for_month_year(request):
    # my_date = date.today()
    year = request.POST.get('year')
    month = request.POST.get('month')
    

    def get_dates_for_month(year, month):
        # Calculate the first day of the given month
        first_day = datetime(year, month, 1)

        # Calculate the last day of the given month
        if month == 12:
            last_day = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            last_day = datetime(year, month + 1, 1) - timedelta(days=1)

        # Generate a list of dates for the month
        dates_list = [first_day + timedelta(days=i) for i in range((last_day - first_day).days + 1)]

        # Format the dates as 'YYYY-MM-DD' and return the list
        formatted_dates = [date.strftime('%Y-%m-%d') for date in dates_list]
        return formatted_dates

    date_list = get_dates_for_month(int(year), int(month))

    # Test the function
    for my_date in date_list:
        company_code =  "O001"
        on_leave_count=0
        on_wfh_count=0
        total_count=0
        present_count=0
        not_found_count=0
        without_attendanceid_total_count=0
        with_attendanceid_total_count=0
        
        on_leave_employees=[]
        on_wfh_employees=[]
        total_employees=[]
        present_employees=[]
        not_found_employees=[]
        without_attendanceid_total_employees=[]
        with_attendanceid_total_employees=[]
        
        
        with_attendanceid_total_count=Users.objects.filter(is_active=True,company_code=company_code).exclude(employeeId__isnull=True).order_by('id')
        with_attendanceid_total_count_ser=UserSerializer(with_attendanceid_total_count,many=True)
        for emp in with_attendanceid_total_count_ser.data:
            with_attendanceid_total_employees.append(emp['id'])
        with_attendanceid_total_count=with_attendanceid_total_count.count()

        without_attendanceid_total_count=Users.objects.filter(is_active=True,company_code=company_code).exclude(employeeId__isnull=False).order_by('id')
        without_attendanceid_total_count_ser=UserSerializer(without_attendanceid_total_count,many=True)
        for emp in without_attendanceid_total_count_ser.data:
            without_attendanceid_total_employees.append(emp['id'])
        without_attendanceid_total_count=without_attendanceid_total_count.count()

        leave_data=Leave.objects.filter(Active=True,company_code=company_code,leave_status="Approved").order_by('id')
        leave_data_serializer = leaveserializer(leave_data, many=True)
        
        for l in leave_data_serializer.data:
            d2 =date.fromisoformat(str(my_date))
            d1 =date.fromisoformat(l['start_date'])         
            d3 =date.fromisoformat(l['end_date']) 
            if d1 <= d2 and d2 <= d3:
                if l['WorkFromHome']:
                    on_wfh_employees.append(l['employeeId'])
                    on_wfh_count += 1
                else:
                    on_leave_employees.append(l['employeeId'])
                    on_leave_count += 1
                    
                    
        present_count=attendance.objects.filter(date=str(my_date),company_code=company_code)
        present_count_ser=attendanceserializer(present_count,many=True)
        for emp in present_count_ser.data:
            present_employees.append(emp['employeeId'])
        present_count=present_count.count()


        total_employees=without_attendanceid_total_employees + with_attendanceid_total_employees
        found_list=on_leave_employees + present_employees
        not_found_employees=list(set(total_employees) - set(found_list))

        
        
        total_count=without_attendanceid_total_count + with_attendanceid_total_count
        not_found_count=total_count - (on_leave_count + present_count)
        
        context={
            "managerid":"All",
            "team_hold_by":"All",
            "date":my_date,
            "present_count":present_count,
            "total_count":total_count,
            "on_leave_count":on_leave_count,
            "on_wfh_count":on_wfh_count,
            "not_found_count":not_found_count,
            "without_attendanceid_total_count":without_attendanceid_total_count,
            "with_attendanceid_total_count":with_attendanceid_total_count,
            "company_code":company_code,
            
            "on_leave_employees":on_leave_employees,
            "on_wfh_employees":on_wfh_employees,
            "total_employees":total_employees,
            "present_employees":present_employees,
            "not_found_employees":not_found_employees,
            "without_attendanceid_total_employees":without_attendanceid_total_employees,
            "with_attendanceid_total_employees":with_attendanceid_total_employees,
            
        }
        
        alredyexist_obj=TeamAttendance.objects.filter(managerid=context['managerid'],date=context['date']).order_by('time').first()
        if alredyexist_obj:
            TeamAttendance.objects.filter(managerid=context['managerid'],date=context['date']).update(
                                    team_hold_by=context['team_hold_by'],
                                    present_count=context['present_count'],
                                    total_count=context['total_count'],
                                    on_leave_count=context['on_leave_count'],
                                    on_wfh_count=context['on_wfh_count'],
                                    not_found_count=context['not_found_count'],
                                    without_attendanceid_total_count=context['without_attendanceid_total_count'],
                                    with_attendanceid_total_count=context['with_attendanceid_total_count'],
                                    company_code=context['company_code'],
                                    
                                    
                                    present_employees= ",".join(map(str, context['present_employees'])),
                                    on_leave_employees= ",".join(map(str, context['on_leave_employees'])),
                                    on_wfh_employees=",".join(map(str, context['on_wfh_employees'])),
                                    not_found_employees=",".join(map(str, context['not_found_employees'])),
                                    total_employees=",".join(map(str, context['total_employees'])),
                                    without_attendanceid_employees=",".join(map(str, context['without_attendanceid_total_employees'])),
                                    with_attendanceid_employees=",".join(map(str, context['with_attendanceid_total_employees'])),
                                    
                                    )
        else:
            TeamAttendance.objects.create(managerid=context['managerid'],
                                        team_hold_by=context['team_hold_by'],
                                        date=context['date'],
                                        present_count=context['present_count'],
                                        total_count=context['total_count'],
                                        on_leave_count=context['on_leave_count'],
                                        on_wfh_count=context['on_wfh_count'],
                                        not_found_count=context['not_found_count'],
                                        without_attendanceid_total_count=context['without_attendanceid_total_count'],
                                        with_attendanceid_total_count=context['with_attendanceid_total_count'],
                                        company_code=context['company_code'],
                                        
                                        present_employees= ",".join(map(str, context['present_employees'])),
                                        on_leave_employees= ",".join(map(str, context['on_leave_employees'])),
                                        on_wfh_employees=",".join(map(str, context['on_wfh_employees'])),
                                        not_found_employees=",".join(map(str, context['not_found_employees'])),
                                        total_employees=",".join(map(str, context['total_employees'])),
                                        without_attendanceid_employees=",".join(map(str, context['without_attendanceid_total_employees'])),
                                        with_attendanceid_employees=",".join(map(str, context['with_attendanceid_total_employees'])),
                                        
                                        )
        
        allemployees=Users.objects.filter(is_active=True,company_code=company_code).order_by('id')
        allemployees_serializer=UserSerializer(allemployees,many=True)
        for i in allemployees_serializer.data:
            i_on_leave_count=0
            i_on_wfh_count=0
            i_total_count=0
            i_present_count=0
            i_not_found_count=0
            
            i_without_attendanceid_total_count=0
            i_with_attendanceid_total_count=0
            i_Manager=i['Firstname'] + ' ' + i['Lastname']
            i_ManagerId=i['id']      
            
            
            i_on_leave_employees=[]
            i_on_wfh_employees=[]
            i_total_employees=[]
            i_present_employees=[]
            i_not_found_employees=[]
            i_without_attendanceid_total_employees=[]
            i_with_attendanceid_total_employees=[]
            
                
            mappingobj = UserToManager.objects.filter(ManagerID=i['id'],company_code = i['company_code'])
            team_list=[]
            attendance_id_list=[]
            if mappingobj.exists():

                
                map_serializer = MappingSerializer(mappingobj, many=True)
                for j in map_serializer.data:
                    team_list.append(j['UserID'])
                
                for team_member in team_list:
                    team_member_obj=Users.objects.filter(id=team_member,is_active=True,company_code=company_code).order_by('id').first()
                    if team_member_obj:

                        if team_member_obj.employeeId is not None and team_member_obj.employeeId !="":
                            i_with_attendanceid_total_count += 1
                            i_without_attendanceid_total_employees.append(team_member_obj.id)
                            attendance_id_list.append(team_member_obj.employeeId)
                        else:
                            i_with_attendanceid_total_employees.append(team_member_obj.id)
                            i_without_attendanceid_total_count += 1
                        
                i_present_count=attendance.objects.filter(employeeId__in=attendance_id_list,date=str(my_date))
                i_present_count_ser=attendanceserializer(i_present_count,many=True)
                i_present_count=i_present_count.count()

                for emp in i_present_count_ser.data:
                    i_present_employees.append(emp['employeeId'])
                
                i_leave_data=Leave.objects.filter(employeeId__in=team_list,Active=True,leave_status="Approved").order_by('id')
                i_leave_data_serializer = leaveserializer(i_leave_data, many=True)
                
                for l in i_leave_data_serializer.data:
                    i_d2 =date.fromisoformat(str(my_date))
                    i_d1 =date.fromisoformat(l['start_date'])         
                    i_d3 =date.fromisoformat(l['end_date']) 
                    if i_d1 <= i_d2 and i_d2 <= i_d3:
                        if l['WorkFromHome']:
                            i_on_wfh_count += 1
                            i_on_wfh_employees.append(l['employeeId'])
                        else:
                            i_on_leave_count += 1  
                            i_on_leave_employees.append(l['employeeId'])
                            
                i_total_count=len(team_list)
                i_total_employees=team_list
                i_found=i_on_leave_employees + i_present_employees
                i_not_found_employees=list(set(i_total_employees) - set(i_found))
                i_not_found_count=i_total_count - (i_on_leave_count + i_present_count)
                
                i_context={
                    "managerid":i_ManagerId,
                    "team_hold_by":i_Manager,
                    "date":my_date,
                    "present_count":i_present_count,
                    "total_count":i_total_count,
                    "on_leave_count":i_on_leave_count,
                    "on_wfh_count":i_on_wfh_count,
                    "not_found_count":i_not_found_count,
                    "with_attendanceid_total_count":i_with_attendanceid_total_count,
                    "without_attendanceid_total_count":i_without_attendanceid_total_count,
                    "company_code":company_code,
                    
                    
                    "on_leave_employees":i_on_leave_employees,
                    "on_wfh_employees":i_on_wfh_employees,
                    "total_employees":i_total_employees,
                    "present_employees":i_present_employees,
                    "not_found_employees":i_not_found_employees,
                    "without_attendanceid_total_employees":i_without_attendanceid_total_employees,
                    "with_attendanceid_total_employees":i_with_attendanceid_total_employees,
                    
                    
                }
        
                i_alredyexist_obj=TeamAttendance.objects.filter(managerid=i_context['managerid'],date=i_context['date']).order_by('time').first()
                if i_alredyexist_obj:
                    TeamAttendance.objects.filter(managerid=i_context['managerid'],date=i_context['date']).update(
                                team_hold_by=i_context['team_hold_by'],
                                present_count=i_context['present_count'],
                                total_count=i_context['total_count'],
                                on_leave_count=i_context['on_leave_count'],
                                on_wfh_count=i_context['on_wfh_count'],
                                not_found_count=i_context['not_found_count'],
                                without_attendanceid_total_count=i_context['without_attendanceid_total_count'],
                                with_attendanceid_total_count=i_context['with_attendanceid_total_count'],
                                company_code=i_context['company_code'],
                            
                                present_employees= ",".join(map(str, i_context['present_employees'])),
                                on_leave_employees= ",".join(map(str, i_context['on_leave_employees'])),
                                on_wfh_employees=",".join(map(str, i_context['on_wfh_employees'])),
                                not_found_employees=",".join(map(str, i_context['not_found_employees'])),
                                total_employees=",".join(map(str, i_context['total_employees'])),
                                without_attendanceid_employees=",".join(map(str, i_context['without_attendanceid_total_employees'])),
                                with_attendanceid_employees=",".join(map(str, i_context['with_attendanceid_total_employees'])),
                                )
                else:
                        
                    TeamAttendance.objects.create(managerid=i_context['managerid'],
                                                team_hold_by=i_context['team_hold_by'],
                                                date=i_context['date'],
                                                present_count=i_context['present_count'],
                                                total_count=i_context['total_count'],
                                                on_leave_count=i_context['on_leave_count'],
                                                on_wfh_count=i_context['on_wfh_count'],
                                                not_found_count=i_context['not_found_count'],
                                                without_attendanceid_total_count=i_context['without_attendanceid_total_count'],
                                                with_attendanceid_total_count=i_context['with_attendanceid_total_count'],
                                                company_code=i_context['company_code'],
                                                
                                                present_employees= ",".join(map(str, i_context['present_employees'])),
                                                on_leave_employees= ",".join(map(str, i_context['on_leave_employees'])),
                                                on_wfh_employees=",".join(map(str, i_context['on_wfh_employees'])),
                                                not_found_employees=",".join(map(str, i_context['not_found_employees'])),
                                                total_employees=",".join(map(str, i_context['total_employees'])),
                                                without_attendanceid_employees=",".join(map(str, i_context['without_attendanceid_total_employees'])),
                                                with_attendanceid_employees=",".join(map(str, i_context['with_attendanceid_total_employees'])),
                                                )
        
        
        
        
        my_date=str(my_date)


    return Response({'n':1,'msg':'data stored successfully of ','status':'success','data':date_list})

@csrf_exempt
@api_view(['POST'])
@authentication_classes([])
@permission_classes([])
def team_attendance_scheduler_by_date(request):
    my_date=request.POST.get("date")
    company_code =  "O001"
    on_leave_count=0
    on_wfh_count=0
    total_count=0
    present_count=0
    not_found_count=0
    without_attendanceid_total_count=0
    with_attendanceid_total_count=0
    
    on_leave_employees=[]
    on_wfh_employees=[]
    total_employees=[]
    present_employees=[]
    not_found_employees=[]
    without_attendanceid_total_employees=[]
    with_attendanceid_total_employees=[]
    
    
    with_attendanceid_total_count=Users.objects.filter(is_active=True,company_code=company_code).exclude(employeeId__isnull=True).order_by('id')
    with_attendanceid_total_count_ser=UserSerializer(with_attendanceid_total_count,many=True)
    for emp in with_attendanceid_total_count_ser.data:
        with_attendanceid_total_employees.append(emp['id'])
    with_attendanceid_total_count=with_attendanceid_total_count.count()

    without_attendanceid_total_count=Users.objects.filter(is_active=True,company_code=company_code).exclude(employeeId__isnull=False).order_by('id')
    without_attendanceid_total_count_ser=UserSerializer(without_attendanceid_total_count,many=True)
    for emp in without_attendanceid_total_count_ser.data:
        without_attendanceid_total_employees.append(emp['id'])
    without_attendanceid_total_count=without_attendanceid_total_count.count()

    leave_data=Leave.objects.filter(Active=True,company_code=company_code,leave_status="Approved").order_by('id')
    leave_data_serializer = leaveserializer(leave_data, many=True)
    
    for l in leave_data_serializer.data:
        d2 =date.fromisoformat(str(my_date))
        d1 =date.fromisoformat(l['start_date'])         
        d3 =date.fromisoformat(l['end_date']) 
        if d1 <= d2 and d2 <= d3:
            if l['WorkFromHome']:
                on_wfh_employees.append(l['employeeId'])
                on_wfh_count += 1
            else:
                on_leave_employees.append(l['employeeId'])
                on_leave_count += 1
                
                
    present_count=attendance.objects.filter(date=str(my_date),company_code=company_code)
    present_count_ser=attendanceserializer(present_count,many=True)
    for emp in present_count_ser.data:
        present_employees.append(emp['employeeId'])
    present_count=present_count.count()


    total_employees=without_attendanceid_total_employees + with_attendanceid_total_employees
    found_list=on_leave_employees + present_employees
    not_found_employees=list(set(total_employees) - set(found_list))

    
    
    total_count=without_attendanceid_total_count + with_attendanceid_total_count
    not_found_count=total_count - (on_leave_count + present_count)
    
    context={
        "managerid":"All",
        "team_hold_by":"All",
        "date":my_date,
        "present_count":present_count,
        "total_count":total_count,
        "on_leave_count":on_leave_count,
        "on_wfh_count":on_wfh_count,
        "not_found_count":not_found_count,
        "without_attendanceid_total_count":without_attendanceid_total_count,
        "with_attendanceid_total_count":with_attendanceid_total_count,
        "company_code":company_code,
        
        "on_leave_employees":on_leave_employees,
        "on_wfh_employees":on_wfh_employees,
        "total_employees":total_employees,
        "present_employees":present_employees,
        "not_found_employees":not_found_employees,
        "without_attendanceid_total_employees":without_attendanceid_total_employees,
        "with_attendanceid_total_employees":with_attendanceid_total_employees,
        
    }
    
    alredyexist_obj=TeamAttendance.objects.filter(managerid=context['managerid'],date=context['date']).order_by('time').first()
    if alredyexist_obj:
        TeamAttendance.objects.filter(managerid=context['managerid'],date=context['date']).update(
                                team_hold_by=context['team_hold_by'],
                                present_count=context['present_count'],
                                total_count=context['total_count'],
                                on_leave_count=context['on_leave_count'],
                                on_wfh_count=context['on_wfh_count'],
                                not_found_count=context['not_found_count'],
                                without_attendanceid_total_count=context['without_attendanceid_total_count'],
                                with_attendanceid_total_count=context['with_attendanceid_total_count'],
                                company_code=context['company_code'],
                                
                                
                                present_employees= ",".join(map(str, context['present_employees'])),
                                on_leave_employees= ",".join(map(str, context['on_leave_employees'])),
                                on_wfh_employees=",".join(map(str, context['on_wfh_employees'])),
                                not_found_employees=",".join(map(str, context['not_found_employees'])),
                                total_employees=",".join(map(str, context['total_employees'])),
                                without_attendanceid_employees=",".join(map(str, context['without_attendanceid_total_employees'])),
                                with_attendanceid_employees=",".join(map(str, context['with_attendanceid_total_employees'])),
                                
                                )
    else:
        TeamAttendance.objects.create(managerid=context['managerid'],
                                    team_hold_by=context['team_hold_by'],
                                    date=context['date'],
                                    present_count=context['present_count'],
                                    total_count=context['total_count'],
                                    on_leave_count=context['on_leave_count'],
                                    on_wfh_count=context['on_wfh_count'],
                                    not_found_count=context['not_found_count'],
                                    without_attendanceid_total_count=context['without_attendanceid_total_count'],
                                    with_attendanceid_total_count=context['with_attendanceid_total_count'],
                                    company_code=context['company_code'],
                                    
                                    present_employees= ",".join(map(str, context['present_employees'])),
                                    on_leave_employees= ",".join(map(str, context['on_leave_employees'])),
                                    on_wfh_employees=",".join(map(str, context['on_wfh_employees'])),
                                    not_found_employees=",".join(map(str, context['not_found_employees'])),
                                    total_employees=",".join(map(str, context['total_employees'])),
                                    without_attendanceid_employees=",".join(map(str, context['without_attendanceid_total_employees'])),
                                    with_attendanceid_employees=",".join(map(str, context['with_attendanceid_total_employees'])),
                                    
                                    )
    
    allemployees=Users.objects.filter(is_active=True,company_code=company_code).order_by('id')
    allemployees_serializer=UserSerializer(allemployees,many=True)
    for i in allemployees_serializer.data:
        i_on_leave_count=0
        i_on_wfh_count=0
        i_total_count=0
        i_present_count=0
        i_not_found_count=0
        
        i_without_attendanceid_total_count=0
        i_with_attendanceid_total_count=0
        i_Manager=i['Firstname'] + ' ' + i['Lastname']
        i_ManagerId=i['id']      
        
        
        i_on_leave_employees=[]
        i_on_wfh_employees=[]
        i_total_employees=[]
        i_present_employees=[]
        i_not_found_employees=[]
        i_without_attendanceid_total_employees=[]
        i_with_attendanceid_total_employees=[]
        
            
        mappingobj = UserToManager.objects.filter(ManagerID=i['id'],company_code = i['company_code'])
        team_list=[]
        attendance_id_list=[]
        if mappingobj.exists():

            
            map_serializer = MappingSerializer(mappingobj, many=True)
            for j in map_serializer.data:
                team_list.append(j['UserID'])
            
            for team_member in team_list:
                team_member_obj=Users.objects.filter(id=team_member,is_active=True,company_code=company_code).order_by('id').first()
                if team_member_obj:

                    if team_member_obj.employeeId is not None and team_member_obj.employeeId !="":
                        i_with_attendanceid_total_count += 1
                        i_without_attendanceid_total_employees.append(team_member_obj.id)
                        attendance_id_list.append(team_member_obj.employeeId)
                    else:
                        i_with_attendanceid_total_employees.append(team_member_obj.id)
                        i_without_attendanceid_total_count += 1
                    
            i_present_count=attendance.objects.filter(employeeId__in=attendance_id_list,date=str(my_date))
            i_present_count_ser=attendanceserializer(i_present_count,many=True)
            i_present_count=i_present_count.count()

            for emp in i_present_count_ser.data:
                i_present_employees.append(emp['employeeId'])
            
            i_leave_data=Leave.objects.filter(employeeId__in=team_list,Active=True,leave_status="Approved").order_by('id')
            i_leave_data_serializer = leaveserializer(i_leave_data, many=True)
            
            for l in i_leave_data_serializer.data:
                i_d2 =date.fromisoformat(str(my_date))
                i_d1 =date.fromisoformat(l['start_date'])         
                i_d3 =date.fromisoformat(l['end_date']) 
                if i_d1 <= i_d2 and i_d2 <= i_d3:
                    if l['WorkFromHome']:
                        i_on_wfh_count += 1
                        i_on_wfh_employees.append(l['employeeId'])
                    else:
                        i_on_leave_count += 1  
                        i_on_leave_employees.append(l['employeeId'])
                        
            i_total_count=len(team_list)
            i_total_employees=team_list
            i_found=i_on_leave_employees + i_present_employees
            i_not_found_employees=list(set(i_total_employees) - set(i_found))
            i_not_found_count=i_total_count - (i_on_leave_count + i_present_count)
            
            i_context={
                "managerid":i_ManagerId,
                "team_hold_by":i_Manager,
                "date":my_date,
                "present_count":i_present_count,
                "total_count":i_total_count,
                "on_leave_count":i_on_leave_count,
                "on_wfh_count":i_on_wfh_count,
                "not_found_count":i_not_found_count,
                "with_attendanceid_total_count":i_with_attendanceid_total_count,
                "without_attendanceid_total_count":i_without_attendanceid_total_count,
                "company_code":company_code,
                
                
                "on_leave_employees":i_on_leave_employees,
                "on_wfh_employees":i_on_wfh_employees,
                "total_employees":i_total_employees,
                "present_employees":i_present_employees,
                "not_found_employees":i_not_found_employees,
                "without_attendanceid_total_employees":i_without_attendanceid_total_employees,
                "with_attendanceid_total_employees":i_with_attendanceid_total_employees,
                
                
            }
    
            i_alredyexist_obj=TeamAttendance.objects.filter(managerid=i_context['managerid'],date=i_context['date']).order_by('time').first()
            if i_alredyexist_obj:
                TeamAttendance.objects.filter(managerid=i_context['managerid'],date=i_context['date']).update(
                            team_hold_by=i_context['team_hold_by'],
                            present_count=i_context['present_count'],
                            total_count=i_context['total_count'],
                            on_leave_count=i_context['on_leave_count'],
                            on_wfh_count=i_context['on_wfh_count'],
                            not_found_count=i_context['not_found_count'],
                            without_attendanceid_total_count=i_context['without_attendanceid_total_count'],
                            with_attendanceid_total_count=i_context['with_attendanceid_total_count'],
                            company_code=i_context['company_code'],
                        
                            present_employees= ",".join(map(str, i_context['present_employees'])),
                            on_leave_employees= ",".join(map(str, i_context['on_leave_employees'])),
                            on_wfh_employees=",".join(map(str, i_context['on_wfh_employees'])),
                            not_found_employees=",".join(map(str, i_context['not_found_employees'])),
                            total_employees=",".join(map(str, i_context['total_employees'])),
                            without_attendanceid_employees=",".join(map(str, i_context['without_attendanceid_total_employees'])),
                            with_attendanceid_employees=",".join(map(str, i_context['with_attendanceid_total_employees'])),
                            )
            else:
                    
                TeamAttendance.objects.create(managerid=i_context['managerid'],
                                            team_hold_by=i_context['team_hold_by'],
                                            date=i_context['date'],
                                            present_count=i_context['present_count'],
                                            total_count=i_context['total_count'],
                                            on_leave_count=i_context['on_leave_count'],
                                            on_wfh_count=i_context['on_wfh_count'],
                                            not_found_count=i_context['not_found_count'],
                                            without_attendanceid_total_count=i_context['without_attendanceid_total_count'],
                                            with_attendanceid_total_count=i_context['with_attendanceid_total_count'],
                                            company_code=i_context['company_code'],
                                            
                                            present_employees= ",".join(map(str, i_context['present_employees'])),
                                            on_leave_employees= ",".join(map(str, i_context['on_leave_employees'])),
                                            on_wfh_employees=",".join(map(str, i_context['on_wfh_employees'])),
                                            not_found_employees=",".join(map(str, i_context['not_found_employees'])),
                                            total_employees=",".join(map(str, i_context['total_employees'])),
                                            without_attendanceid_employees=",".join(map(str, i_context['without_attendanceid_total_employees'])),
                                            with_attendanceid_employees=",".join(map(str, i_context['with_attendanceid_total_employees'])),
                                            )
    
    
    
    
    my_date=str(my_date)


    return Response({'n':1,'msg':'data stored successfully of ','status':'success','data':my_date})

@api_view(['GET'])
@permission_classes((AllowAny,))
def previousdayleaves(request):
    today_date = datetime.now()
    # one_day_less = today_date - datetime.timedelta(days=1)
    today_date = today_date.strftime("%Y-%m-%d")
    leaveobj = Leave.objects.filter(Active=True,leave_status ='Pending',start_date__lte=today_date)
    leaveser = leaveserializer(leaveobj,many=True)
    leaves=[]
    for l in leaveser.data:
        leave_approval_object=leaveApproval.objects.filter(employeeId=l['employeeId'],leave_id=l['id'],company_code=l['company_code'],approvedBy=True,rejectedBy=False).exclude(rejectedBy=True).count()
        
        if leave_approval_object != 0 :
            leaves.append(l)
            
            Leave.objects.filter(id=l['id'],Active=True,leave_status = 'Pending').update(leave_status='Approved')
            TaskNotification.objects.filter(leaveID=l['id'],To_manager=True,action_Taken=False).update(action_Taken=True)

        
        else:
        
            Leave.objects.filter(id=l['id'],Active=True,leave_status ='Pending').update(leave_status='Rejected')
            TaskNotification.objects.filter(leaveID=l['id'],To_manager=True,action_Taken=False).update(action_Taken=True)
            leaveApproval.objects.filter(employeeId=l['employeeId'],leave_id=l['id'],company_code=l['company_code'],approvedBy=False).update(rejectedBy=True,comment="Not responded on time")

    return Response ({"data":leaves,"response":{"n" : 1,"msg" : "SUCCESS","status" : "success"}})

@api_view(['POST'])
def get_alloted_shift_header_details(request):
    company_code = request.user.company_code

    totalteamcount=0
    team_with_attendanceid_list=[]
    team_list=[]

          
    shift_obj=EmployeeShiftDetails.objects.filter(is_active=True)
    if shift_obj is not None:
        empobjser = EmployeeShiftDetailsSerializer(shift_obj,many=True)
        totalteamcount=shift_obj.count()
        for j in empobjser.data:
            team_list.append(j['employeeId'])
            userobj=Users.objects.filter(id=j['employeeId'],is_active=True,company_code=company_code).first()
            if userobj is not None:
                userobjser=UserSerializer(userobj)
                if userobjser.data['employeeId'] is not None and userobjser.data['employeeId'] !="":
                    team_with_attendanceid_list.append(userobjser.data['employeeId'])













                
                
    context = {
        "All_Mapped_Emp_Count":totalteamcount,
        "Total_count" : totalteamcount,

    }
    return Response ({"data":[],"Photo":'',"Designation":'----',"Name":"All Employee",
                      "EmployeeId":"----","context":context,
                      "response":{"n" : 1,"msg" : "SUCCESS","status" : "success"}})

@api_view(['POST'])
def change_working_status(request):
    company_code = request.user.company_code
    userid = request.user.id
    action_status=request.POST.get('status')
    task_notification_id=request.POST.get('task_notification_id')
    if action_status.lower() =='yes':
        TaskNotification.objects.filter(id=task_notification_id,UserID_id=userid,To_manager=False,leaveID=0,company_code=company_code,NotificationTypeId_id=7,NotificationTitle="Working Status").update(action_Taken=True)
        return Response ({"data":[],"response":{"n" : 1,"msg" : "Thanks for updating.","status" : "success"}})
    else:
        current_date = date.today()
        user_obj=Users.objects.filter(id=userid,is_active=True,company_code=company_code).first()
        if user_obj is not None:
            attendanceId=user_obj.employeeId
            if attendanceId !='' and attendanceId is not None and attendanceId !='None':
                current_time = datetime.now()
                current_week = current_date.isocalendar()[1]
                current_month = current_date.month
                current_year = current_date.year
                formatted_time = current_time.strftime("%H:%M:%S")
                
                TaskNotification.objects.filter(id=task_notification_id,UserID_id=userid,To_manager=False,leaveID=0,company_code=company_code,NotificationTypeId_id=7,NotificationTitle="Working Status").update(action_Taken=True)
                
                attendance_obj=attendance.objects.create(date=current_date,employeeId=attendanceId,time=formatted_time,company_code=company_code,Week=current_week,Year=current_year,Month=current_month,attendance_type="System Forced",checkout=True)
                return Response ({"data":[],"response":{"n" : 1,"msg" : "Thanks for updating.","status" : "success"}})
                 
            return Response ({"data":[],"response":{"n" : 1,"msg" : "User attendance id not found.","status" : "error"}})

        return Response ({"data":[],"response":{"n" : 1,"msg" : "User not found.","status" : "error"}})
  
@api_view(['POST'])
def mark_forced_system_checkout(request):
    company_code = request.user.company_code
    userid = request.user.id
    action_status=request.POST.get('status')
    if action_status.lower() =='no':

        current_date = date.today()
        user_obj=Users.objects.filter(id=userid,is_active=True,company_code=company_code).first()
        if user_obj is not None:
            attendanceId=user_obj.employeeId
            if attendanceId !='' and attendanceId is not None and attendanceId !='None':
                current_time = datetime.now()
                current_week = current_date.isocalendar()[1]
                current_month = current_date.month
                current_year = current_date.year
                formatted_time = current_time.strftime("%H:%M:%S")
                TaskNotification.objects.filter(UserID_id=userid,To_manager=False,leaveID=0,company_code=company_code,NotificationTypeId_id=7,NotificationTitle="Working Status").update(action_Taken=True)
                attendance_obj=attendance.objects.create(date=current_date,employeeId=attendanceId,time=formatted_time,company_code=company_code,Week=current_week,Year=current_year,Month=current_month)
                return Response ({"data":[],"response":{"n" : 1,"msg" : "We dont recive any response from you ,So we assume that you  have been checkout .","status" : "success"}})
                 
            return Response ({"data":[],"response":{"n" : 1,"msg" : "User attendance id not found.","status" : "error"}})

        return Response ({"data":[],"response":{"n" : 1,"msg" : "User not found.","status" : "error"}})

    else:
        return Response ({"data":[],"response":{"n" : 1,"msg" : "wrong action.","status" : "error"}})
        
def check_shift_status(shift_date, shift_start_time, shift_end_time):
    current_datetime = datetime.now()

    shift_start_datetime = datetime.strptime(f"{shift_date} {shift_start_time}", "%Y-%m-%d %H:%M:%S")
    shift_end_datetime = datetime.strptime(f"{shift_date} {shift_end_time}", "%Y-%m-%d %H:%M:%S")

    if current_datetime < shift_start_datetime:
        return "coming"
    elif shift_start_datetime <= current_datetime <= shift_end_datetime:
        return "ongoing"
    else:
        return "passed"
        
def is_in_shift(shift_date_str, shift_start_time_str, shift_end_time_str):
    # Convert shift date and time strings to datetime objects
    shift_date = datetime.strptime(shift_date_str, '%Y-%m-%d').date()
    shift_start_time = datetime.strptime(shift_start_time_str, '%H:%M').time()
    shift_end_time = datetime.strptime(shift_end_time_str, '%H:%M').time()

    # Get the current date and time
    current_datetime = datetime.now()

    # Extract the current date and time
    current_date = current_datetime.date()
    current_time = current_datetime.time()

    # Check if the current date is the shift date
    if current_date == shift_date:
        # Check if the current time is within the shift time range (across midnight)
        if shift_start_time <= current_time and current_time <= shift_end_time:
            return True
    return False

@api_view(['POST'])
@permission_classes((AllowAny,))
def warning_shedular_9hr_after_checkin(request):
    # company_code = request.user.company_code
    current_date = date.today()
    previous_date = current_date - timedelta(days=1)
    current_datetime = datetime.now()

    # check the current date shift 
    # find the login above the current shift 
    # if the login found check for logout of same shift till the shift end time
    # if not exits check its shift ending time and date  if its is passed  then check  total working hrs is full fill or not ,
    # if total workings hrs is fullfill then check whether the current date time is not in next day shift date time
    
    
    
    hrs24_diff=datetime.strptime("23:59", "%H:%M") - datetime.strptime("00:00", "%H:%M")
    

    users_obj=Users.objects.filter(is_active=True).exclude(Q(employeeId__isnull=True)|Q(employeeId='None')|Q(employeeId=''))
    usersserializer=UsersSerializer(users_obj,many=True)
    for user in usersserializer.data:
        if user['employeetype'] != '' and user['employeetype'] is not None:
            nextday = datetime.strptime(str(previous_date) , "%Y-%m-%d").date() + timedelta(days=1)            
            previousshift={}
            currentshift={}
            nextshift={}
            
            previousshift['shiftname']='General Shift'
            previousshift['intime']='09:30'
            previousshift['outtime']='18:30'
            previousshift['shiftdate']=str(previous_date)
            
            currentshift['shiftname']='General Shift'
            currentshift['intime']='09:30'
            currentshift['outtime']='18:30'
            currentshift['shiftdate']=str(current_date)
            
            nextshift['shiftname']='General Shift'
            nextshift['intime']='09:30'
            nextshift['outtime']='18:30'
            nextshift['shiftdate']=str(nextday)
            previous_alloted_shift_obj=ShiftAllotment.objects.filter(employeeId=user['id'],attendanceId=user['employeeId'],date=str(previous_date),is_active=True).first()
            if previous_alloted_shift_obj is not None:
                previous_ShiftAllotment_serializer=ShiftAllotmentSerializer(previous_alloted_shift_obj)
                if previous_ShiftAllotment_serializer.data['shiftId'] is not None and previous_ShiftAllotment_serializer.data['shiftId'] !='':
                    previous_shift_obj=ShiftMaster.objects.filter(id=previous_ShiftAllotment_serializer.data['shiftId'],is_active=True).first()
                    if previous_shift_obj is not None:
                        previous_shift_obj_serializer=ShiftMasterSerializer(previous_shift_obj)
                        previousshift['shiftname']=previous_shift_obj_serializer.data['shiftname']
                        previousshift['intime']=previous_shift_obj_serializer.data['intime']
                        previousshift['outtime']=previous_shift_obj_serializer.data['outtime']
                        previousshift['shiftdate']=str(previous_date)

            alloted_shift_obj=ShiftAllotment.objects.filter(employeeId=user['id'],attendanceId=user['employeeId'],date=str(current_date),is_active=True).first()
            if alloted_shift_obj is not None:
                ShiftAllotment_serializer=ShiftAllotmentSerializer(alloted_shift_obj)
                if ShiftAllotment_serializer.data['shiftId'] is not None and ShiftAllotment_serializer.data['shiftId'] !='':
                    shift_obj=ShiftMaster.objects.filter(id=ShiftAllotment_serializer.data['shiftId'],is_active=True).first()
                    if shift_obj is not None:
                        shift_obj_serializer=ShiftMasterSerializer(shift_obj)
                        currentshift['shiftname']=shift_obj_serializer.data['shiftname']
                        currentshift['intime']=shift_obj_serializer.data['intime']
                        currentshift['outtime']=shift_obj_serializer.data['outtime']
                        currentshift['shiftdate']=str(current_date)
                        
            next_alloted_shift_obj=ShiftAllotment.objects.filter(employeeId=user['id'],attendanceId=user['employeeId'],date=str(nextday),is_active=True).first()
            if next_alloted_shift_obj is not None:
                next_ShiftAllotment_serializer=ShiftAllotmentSerializer(next_alloted_shift_obj)
                if next_ShiftAllotment_serializer.data['shiftId'] is not None and next_ShiftAllotment_serializer.data['shiftId'] !='':
                    next_shift_obj=ShiftMaster.objects.filter(id=next_ShiftAllotment_serializer.data['shiftId'],is_active=True).first()
                    if next_shift_obj is not None:
                        next_shift_obj_serializer=ShiftMasterSerializer(next_shift_obj)
                        nextshift['shiftname']=next_shift_obj_serializer.data['shiftname']
                        nextshift['intime']=next_shift_obj_serializer.data['intime']
                        nextshift['outtime']=next_shift_obj_serializer.data['outtime']
                        nextshift['shiftdate']=str(nextday)
                    
                    
                    
                    
                    
                    
            
            previousshift_shift_starting_time =str(datetime.strptime(str(previousshift['intime']), "%H:%M")).split(" ")[1]
            previousshift_start_time = datetime.strptime(str(previousshift_shift_starting_time), "%H:%M:%S") - timedelta(hours=1)
            one_hrs_before_previousshift_shift_starting_time = previousshift_start_time.strftime("%H:%M:%S")
            previousshift_shift_ending_time =str(datetime.strptime(str(previousshift['outtime']), "%H:%M")).split(" ")[1]
            
            
            currentshift_shift_starting_time =str(datetime.strptime(str(currentshift['intime']), "%H:%M")).split(" ")[1]
            currentshift_start_time = datetime.strptime(str(currentshift_shift_starting_time), "%H:%M:%S") - timedelta(hours=1)
            one_hrs_before_currentshift_shift_starting_time = currentshift_start_time.strftime("%H:%M:%S")
            currentshift_shift_ending_time =str(datetime.strptime(str(currentshift['outtime']), "%H:%M")).split(" ")[1]
            
            
            nextshift_shift_starting_time =str(datetime.strptime(str(nextshift['intime']), "%H:%M")).split(" ")[1]
            nextshift_start_time = datetime.strptime(str(nextshift_shift_starting_time), "%H:%M:%S") - timedelta(hours=1)
            one_hrs_before_nextshift_shift_starting_time = nextshift_start_time.strftime("%H:%M:%S")
            nextshift_shift_ending_time =str(datetime.strptime(str(nextshift['outtime']), "%H:%M")).split(" ")[1]
            
            
            shiftstatus=check_shift_status(currentshift['shiftdate'],one_hrs_before_currentshift_shift_starting_time,currentshift_shift_ending_time)
            
            if shiftstatus != "passed":                
                login_obj=attendance.objects.filter(employeeId=user['employeeId'],date=str(currentshift['shiftdate']),time__gte=datetime.strptime(one_hrs_before_currentshift_shift_starting_time, "%H:%M:%S").time(),time__lte=datetime.strptime(currentshift_shift_ending_time, "%H:%M:%S").time()).order_by('time').first()
                if login_obj is not None:                    
                    logout_obj=attendance.objects.filter(employeeId=user['employeeId'],date=str(currentshift['shiftdate']),time__gte=datetime.strptime(one_hrs_before_currentshift_shift_starting_time, "%H:%M:%S").time()).exclude(id=login_obj.id).order_by('time').last()
                    if logout_obj is None:
                        logout_obj=attendance.objects.filter(employeeId=user['employeeId'],date=str(nextshift['shiftdate']),time__lte=datetime.strptime(one_hrs_before_nextshift_shift_starting_time, "%H:%M:%S").time()).exclude(id=login_obj.id).order_by('time').last()
                        if logout_obj is None:                            
                            current_shift_login_datetime = datetime.strptime(f"{str(login_obj.date)} {str(login_obj.time)}", "%Y-%m-%d %H:%M:%S")
                            time_difference_from_login_to_current = current_datetime - current_shift_login_datetime 
                            current_start_time = datetime.strptime(str(currentshift['intime']), "%H:%M")
                            current_end_time = datetime.strptime(str(currentshift['outtime']), "%H:%M")
                            if current_end_time < current_start_time:
                                current_end_time += timedelta(days=1)
                            current_shift_time_difference = current_end_time - current_start_time
                            if time_difference_from_login_to_current >=current_shift_time_difference:
                                if time_difference_from_login_to_current < hrs24_diff:

                                    notificationmsg='Hi <span class="actionuser">'+str(user['Firstname']) + ' ' +str(user['Lastname'])+'</span> are you still working.'
                                    TaskNotification.objects.create(UserID_id=user['id'],To_manager=False,leaveID=0,company_code=user['company_code'],NotificationTypeId_id=7,NotificationTitle="Working Status",NotificationMsg=notificationmsg)
                                    print('current working time is greater--- than shift time',time_difference_from_login_to_current ,'-', current_shift_time_difference)
        
                   
                if login_obj is None:
                    if is_in_shift(currentshift['shiftdate'], currentshift['intime'], currentshift['outtime']):
                        print("preseing in shift",currentshift)
                    else:
                        if previousshift['intime'] > previousshift['outtime']:
                            login_obj=attendance.objects.filter(Q(employeeId=user['employeeId'],date=str(previousshift['shiftdate']),time__gte=datetime.strptime(one_hrs_before_previousshift_shift_starting_time, "%H:%M:%S").time())|Q(employeeId=user['employeeId'],date=str(currentshift['shiftdate']),time__lte=datetime.strptime(previousshift_shift_ending_time, "%H:%M:%S").time())).order_by('time').first()
                        else:
                            login_obj=attendance.objects.filter(employeeId=user['employeeId'],date=str(previousshift['shiftdate']),time__gte=datetime.strptime(one_hrs_before_previousshift_shift_starting_time, "%H:%M:%S").time(),time__lte=datetime.strptime(previousshift_shift_ending_time, "%H:%M:%S").time()).order_by('time').first()
                            
                        if login_obj is not None:
                            logout_obj=attendance.objects.filter(employeeId=user['employeeId'],date=str(previousshift['shiftdate']),time__gte=datetime.strptime(one_hrs_before_previousshift_shift_starting_time, "%H:%M:%S").time()).exclude(id=login_obj.id).order_by('time').last()
                            if logout_obj is None:
                                logout_obj=attendance.objects.filter(employeeId=user['employeeId'],date=str(currentshift['shiftdate']),time__lte=datetime.strptime(one_hrs_before_currentshift_shift_starting_time, "%H:%M:%S").time()).exclude(id=login_obj.id).order_by('time').last()
                                if logout_obj is None:
                                    previous_shift_login_datetime = datetime.strptime(f"{str(login_obj.date)} {str(login_obj.time)}", "%Y-%m-%d %H:%M:%S")
                                    time_difference_from_login_to_previous = current_datetime - previous_shift_login_datetime 
                                    previous_start_time = datetime.strptime(str(previousshift['intime']), "%H:%M")
                                    previous_end_time = datetime.strptime(str(previousshift['outtime']), "%H:%M")
                                    if previous_end_time < previous_start_time:
                                        previous_end_time += timedelta(days=1)
                                        
                                    previous_shift_time_difference = previous_end_time - previous_start_time
                                    if time_difference_from_login_to_previous >=previous_shift_time_difference:
                                        if time_difference_from_login_to_previous < hrs24_diff:
                                            print('previous working time is greater--- than previous shift time',time_difference_from_login_to_previous ,'-', previous_shift_time_difference)
                                            notificationmsg='Hi <span class="actionuser">'+str(user['Firstname']) + ' ' +str(user['Lastname'])+'</span> are you still working.'
                                            TaskNotification.objects.create(UserID_id=user['id'],To_manager=False,leaveID=0,company_code=user['company_code'],NotificationTypeId_id=7,NotificationTitle="Working Status",NotificationMsg=notificationmsg)
                                        
                            
                   
            else:
                login_obj=attendance.objects.filter(employeeId=user['employeeId'],date=str(currentshift['shiftdate']),time__gte=datetime.strptime(one_hrs_before_currentshift_shift_starting_time, "%H:%M:%S").time(),time__lte=datetime.strptime(currentshift_shift_ending_time, "%H:%M:%S").time()).order_by('time').first()
                if login_obj is not None:
                    logout_obj=attendance.objects.filter(employeeId=user['employeeId'],date=str(currentshift['shiftdate']),time__gte=datetime.strptime(one_hrs_before_currentshift_shift_starting_time, "%H:%M:%S").time()).exclude(id=login_obj.id).order_by('time').last()
                    if logout_obj is None:
                        logout_obj=attendance.objects.filter(employeeId=user['employeeId'],date=str(nextshift['shiftdate']),time__lte=datetime.strptime(one_hrs_before_nextshift_shift_starting_time, "%H:%M:%S").time()).exclude(id=login_obj.id).order_by('time').last()
                        if logout_obj is None:                            
                            current_shift_login_datetime = datetime.strptime(f"{str(login_obj.date)} {str(login_obj.time)}", "%Y-%m-%d %H:%M:%S")
                            time_difference_from_login_to_current = current_datetime - current_shift_login_datetime 
                            current_start_time = datetime.strptime(str(currentshift['intime']), "%H:%M")
                            current_end_time = datetime.strptime(str(currentshift['outtime']), "%H:%M")
                            if current_end_time < current_start_time:
                                current_end_time += timedelta(days=1)

                            current_shift_time_difference = current_end_time - current_start_time
                            if time_difference_from_login_to_current >=current_shift_time_difference:
                                if time_difference_from_login_to_current <= hrs24_diff:
                                    print('current working time is greater--- than shift time',time_difference_from_login_to_current ,'-', current_shift_time_difference)
                                    notificationmsg='Hi <span class="actionuser">'+str(user['Firstname']) + ' ' +str(user['Lastname'])+'</span> are you still working.'
                                    TaskNotification.objects.create(UserID_id=user['id'],To_manager=False,leaveID=0,company_code=user['company_code'],NotificationTypeId_id=7,NotificationTitle="Working Status",NotificationMsg=notificationmsg)
                                
                if login_obj is None:
                    if is_in_shift(currentshift['shiftdate'], currentshift['intime'], currentshift['outtime']):
                        print("preseing in shift",currentshift)
                    else:
                        
                        if previousshift['intime'] > previousshift['outtime']:
                            login_obj=attendance.objects.filter(Q(employeeId=user['employeeId'],date=str(previousshift['shiftdate']),time__gte=datetime.strptime(one_hrs_before_previousshift_shift_starting_time, "%H:%M:%S").time())|Q(employeeId=user['employeeId'],date=str(currentshift['shiftdate']),time__lte=datetime.strptime(previousshift_shift_ending_time, "%H:%M:%S").time())).order_by('time').first()
                        else:
                            login_obj=attendance.objects.filter(employeeId=user['employeeId'],date=str(previousshift['shiftdate']),time__gte=datetime.strptime(one_hrs_before_previousshift_shift_starting_time, "%H:%M:%S").time(),time__lte=datetime.strptime(previousshift_shift_ending_time, "%H:%M:%S").time()).order_by('time').first()
                        if login_obj is not None:
                            logout_obj=attendance.objects.filter(employeeId=user['employeeId'],date=str(previousshift['shiftdate']),time__gte=datetime.strptime(one_hrs_before_previousshift_shift_starting_time, "%H:%M:%S").time()).exclude(id=login_obj.id).order_by('time').last()
                            if logout_obj is None:
                                logout_obj=attendance.objects.filter(employeeId=user['employeeId'],date=str(currentshift['shiftdate']),time__lte=datetime.strptime(one_hrs_before_currentshift_shift_starting_time, "%H:%M:%S").time()).exclude(id=login_obj.id).order_by('time').last()
                                if logout_obj is None:
                                    previous_shift_login_datetime = datetime.strptime(f"{str(login_obj.date)} {str(login_obj.time)}", "%Y-%m-%d %H:%M:%S")
                                    time_difference_from_login_to_previous = current_datetime - previous_shift_login_datetime 
                                    previous_start_time = datetime.strptime(str(previousshift['intime']), "%H:%M")
                                    previous_end_time = datetime.strptime(str(previousshift['outtime']), "%H:%M")
                                    if previous_end_time < previous_start_time:
                                        previous_end_time += timedelta(days=1)

                                    previous_shift_time_difference = previous_end_time - previous_start_time
                                    if time_difference_from_login_to_previous >= previous_shift_time_difference:
                                        if time_difference_from_login_to_previous < hrs24_diff:
                                            print('previous working time is greater--- than previous shift time',time_difference_from_login_to_previous ,'-', previous_shift_time_difference)
                                            notificationmsg='Hi <span class="actionuser">'+str(user['Firstname']) + ' ' +str(user['Lastname'])+'</span> are you still working.'
                                            TaskNotification.objects.create(UserID_id=user['id'],To_manager=False,leaveID=0,company_code=user['company_code'],NotificationTypeId_id=7,NotificationTitle="Working Status",NotificationMsg=notificationmsg)
       

    return Response ({"response":{"n" : 1,"msg" : "SUCCESS","status" : "success"}})
     
@api_view(['POST'])    
@permission_classes((AllowAny,))
def get_sift_emp_attendance_by_date(request):

    date = request.POST.get("my_date")
    month=date.split('-')[1]
    year=date.split('-')[0]
    
    l1l2emplst_obj=EmployeeShiftDetails.objects.filter(is_active=True)
    l1l2serializers = L1L2Serializer(l1l2emplst_obj,many=True)
    team_list=list(l1l2serializers.data)
    
    Users_l1l2emplst_obj=Users.objects.filter(id__in=team_list,is_active=True).exclude(Q(employeeId__isnull=True)|Q(employeeId='None')|Q(employeeId=''))
    Users_l1l2serializers = UsersSerializeronlyattendance(Users_l1l2emplst_obj,many=True)
    Userlist=list(Users_l1l2serializers.data)




    Presentlist=[]
    leavelist=[]
    wfhlist=[]
    yettojoinlist=[]
    
    
    for attendanceId in Userlist:
        user_obj=Users.objects.filter(employeeId=attendanceId,is_active=True).first()
        leave_objs = Leave.objects.filter(employeeId=user_obj.id,WorkFromHome=False,leave_status="Approved",start_date__month__gte=month,start_date__month__lte=month,start_date__year= year,Active=True).exclude(leave_status='Draft')
        leavesserializer=leaveserializer(leave_objs,many=True)
        wfh_objs = Leave.objects.filter(employeeId=user_obj.id,WorkFromHome=True,leave_status="Approved",start_date__month__gte=month,start_date__month__lte=month,start_date__year= year,Active=True).exclude(leave_status='Draft')
        wfhsserializer=leaveserializer(wfh_objs,many=True)
    
        current_date_shift_list=[]
        new_current_date_shift_list=[]
        return_dict={}
        
        shiftdate = datetime.strptime(date, '%Y-%m-%d').date()
        tomarrow_date = shiftdate + timedelta(days=1)
        yesterday_date = shiftdate - timedelta(days=1)
        current_shiftallotment_objs=ShiftAllotment.objects.filter(attendanceId=attendanceId,date=str(date),is_active=True)
        current_shiftallotment_serializers=ShiftAllotmentshiftIdSerializer(current_shiftallotment_objs,many=True)
        current_shiftId_list=list(current_shiftallotment_serializers.data)
        current_shift_obj=ShiftMaster.objects.filter(id__in=current_shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime')
        current_shift_serializer=ShiftMasterSerializer(current_shift_obj,many=True)
        current_shiftlist=current_shift_serializer.data
        
        tomarrow_shiftallotment_objs=ShiftAllotment.objects.filter(attendanceId=attendanceId,date=str(tomarrow_date),is_active=True)
        tomarrow_shiftallotment_serializers=ShiftAllotmentshiftIdSerializer(tomarrow_shiftallotment_objs,many=True)
        tomarrow_shiftId_list=list(tomarrow_shiftallotment_serializers.data)
        tomarrow_shift_obj=ShiftMaster.objects.filter(id__in=tomarrow_shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime')
        tomarrow_shift_serializer=ShiftMasterSerializer(tomarrow_shift_obj,many=True)
        tomarrow_shiftlist=tomarrow_shift_serializer.data

        count=1
        
        for shift in current_shiftlist:
            start_time = datetime.strptime(str(shiftdate) +' '+ shift['intime'], '%Y-%m-%d %H:%M')
            start_time_before_2hrs = start_time - timedelta(hours=2)
            
            if shift['intime'] > shift['outtime']:
                shift_end_date = shiftdate + timedelta(days=1)
            else:
                shift_end_date = shiftdate
                
            end_time = datetime.strptime(str(shift_end_date) +' '+ shift['outtime'], '%Y-%m-%d %H:%M')
            check_login_till=''
            if len(current_shiftlist) >= count+1:
            
                check_next_shift_in = datetime.strptime(str(shiftdate) +' '+ current_shiftlist[count]['intime'], '%Y-%m-%d %H:%M')
                check_login_till = check_next_shift_in - timedelta(hours=2)

                
            elif check_login_till=='':
                    
                if len(tomarrow_shiftlist) >= 1:
                    
                    check_next_shift_in = datetime.strptime(str(tomarrow_date) +' '+ tomarrow_shiftlist[0]['intime'], '%Y-%m-%d %H:%M')
                    check_login_till = check_next_shift_in - timedelta(hours=2)
                else:
                
                    if_no_shift_check_shift=ShiftAllotment.objects.filter(attendanceId=attendanceId,date=str(tomarrow_date),is_active=True)
                    tomerrow_shiftlist=ShiftAllotmentshiftIdSerializer(if_no_shift_check_shift,many=True)
                    next_date_shiftId_list=list(tomerrow_shiftlist.data)
                    check_weekly_off=ShiftMaster.objects.filter(id__in=next_date_shiftId_list,intime='00:00',outtime='00:00',is_active=True).order_by('intime').first()
                    if check_weekly_off is not None:
                        
                        if shift['outtime'] < shift['intime']:
                    
                            check_current_shift_out = datetime.strptime(str(tomarrow_date) +' '+ shift['outtime'], '%Y-%m-%d %H:%M')
                            check_current_shift_out_time = check_current_shift_out + timedelta(hours=2)
                            given_datetime = datetime.strptime(str(check_current_shift_out_time), '%Y-%m-%d %H:%M:%S')

                            truncated_datetime_str = given_datetime.strftime('%Y-%m-%d %H:%M')

                            
                            shift_end_date_time=str(truncated_datetime_str)
                        else:
                            shift_end_date_time=str(date) +' '+ '23:59'
                            
                        check_login_till = datetime.strptime(shift_end_date_time, '%Y-%m-%d %H:%M')
                    else:
                    
                        check_login_till = datetime.strptime(str(tomarrow_date) +' '+ '07:30', '%Y-%m-%d %H:%M')
                        
            shift_data={
                'shiftname':shift['shiftname'],
                'indatetime':str(start_time_before_2hrs),
                'outdatetime':str(check_login_till),
                'intime':shift['intime'],
                'outtime':shift['outtime'],
            }
            current_date_shift_list.append(shift_data)
            new_current_date_shift_list.append(shift_data)
            count+=1

        if len(current_shiftlist) == 0:       
        
            if_no_shift_check_shift=ShiftAllotment.objects.filter(attendanceId=attendanceId,date=str(date),is_active=True)
            todays_shiftlist=ShiftAllotmentshiftIdSerializer(if_no_shift_check_shift,many=True)
            toadys_shiftId_list=list(todays_shiftlist.data)
            check_weekly_off=ShiftMaster.objects.filter(id__in=toadys_shiftId_list,intime='00:00',outtime='00:00',is_active=True).order_by('intime').first()
            if check_weekly_off is not None:
                get_previous_day_shift=ShiftAllotment.objects.filter(attendanceId=attendanceId,date=str(yesterday_date),is_active=True)
                previous_day_shiftlist=ShiftAllotmentshiftIdSerializer(get_previous_day_shift,many=True)
                previous_shiftId_list=list(previous_day_shiftlist.data)
                get_previous_last_shift=ShiftMaster.objects.filter(id__in=previous_shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime').last()
                if get_previous_last_shift is not None:
                    if get_previous_last_shift.outtime < get_previous_last_shift.intime:
                        shift_end_date_time=str(date) +' '+ str(get_previous_last_shift.outtime)
                    else:
                        shift_end_date_time=str(yesterday_date) +' '+ '23:59'
                            
                    check_previous_shift_out = datetime.strptime(str(shift_end_date_time), '%Y-%m-%d %H:%M')
                    previous_shift_in_time = check_previous_shift_out + timedelta(hours=2)  
                else:
                    previous_shift_in_time= str(date)+' 00:00:00'
                get_net_day_shift=ShiftAllotment.objects.filter(attendanceId=attendanceId,date=str(tomarrow_date),is_active=True)
                nest_day_shiftlist=ShiftAllotmentshiftIdSerializer(get_net_day_shift,many=True)
                next_date_shiftId_list=list(nest_day_shiftlist.data)
                get_next_day_first_shift=ShiftMaster.objects.filter(id__in=next_date_shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime').first()
                if get_next_day_first_shift is not None:
                    check_next_shift_in = datetime.strptime(str(tomarrow_date) +' '+ str(get_next_day_first_shift.intime), '%Y-%m-%d %H:%M')
                    next_shift_in_time = check_next_shift_in - timedelta(hours=2)
                    shift_data={
                        'shiftname':str(check_weekly_off.shiftname),
                        'indatetime':str(previous_shift_in_time),
                        'outdatetime':str(next_shift_in_time),
                        'intime':check_weekly_off.intime,
                        'outtime':check_weekly_off.outtime,
                    }
                    current_date_shift_list.append(shift_data)
                    new_current_date_shift_list.append(shift_data)
                    
                else:
                    shift_data={
                        'shiftname':str(check_weekly_off.shiftname),
                        'indatetime':str(previous_shift_in_time),
                        'outdatetime':str(date)+' 23:59:59',
                        'intime':check_weekly_off.intime,
                        'outtime':check_weekly_off.outtime,
                    }
                    current_date_shift_list.append(shift_data)
                    new_current_date_shift_list.append(shift_data)

            else:
                shift_data={
                    'shiftname':'General',
                    'indatetime':str(date)+' 07:30:00',
                    'outdatetime':str(tomarrow_date)+' 07:30:00',
                    'intime':'09:30',
                    'outtime':'18:30',
                }
                current_date_shift_list.append(shift_data)
                new_current_date_shift_list.append(shift_data)

                
        current_date_first_in_datetime=''
        current_date_last_out_datetime=''
        extra_days=''
        total_working_hrs=''
        for s in new_current_date_shift_list:
            intime=''
            intimedate=''
            
            
            getallattendance = attendance.objects.filter(Q(employeeId=str(attendanceId),time__gte=s['indatetime'].split(' ')[1],date=s['indatetime'].split(' ')[0])|Q(employeeId=str(attendanceId),time__lte=str(s['outdatetime']).split(' ')[1],date=str(s['outdatetime']).split(' ')[0])).order_by('date','time')
            
            # getallattendance = getallattendance.filter(time__gte=currentshiftstarttime).order_by('date','time')
            
            attendance_serializer=attendanceserializer(getallattendance,many=True)
        
            
            sorted_data = sorted(attendance_serializer.data, key=lambda x: (x['date'], x['time']))
            
            mindatetime = datetime.strptime(s['indatetime'], '%Y-%m-%d %H:%M:%S')
            maxdatetime = datetime.strptime(s['outdatetime'], '%Y-%m-%d %H:%M:%S')
        
            sorted_data = [entry for entry in sorted_data if (mindatetime <= datetime.strptime(entry['date'] +' '+entry['time'],'%Y-%m-%d %H:%M:%S') <= maxdatetime)]
            if len(sorted_data) > 0:
                intimedate=sorted_data[0]['date']
                intime=str(sorted_data[0]['time'])
                
            if intimedate !='' and intimedate is not None:
                user_sdt = datetime.strptime(str(intimedate) + ' ' + str(intime), '%Y-%m-%d %H:%M:%S')
                shif_sdt = datetime.strptime(s['indatetime'].split(' ')[0] + ' ' + s['indatetime'].split(' ')[1], '%Y-%m-%d %H:%M:%S')
                if user_sdt < shif_sdt :
                    intimedate=''
                    intime=''

                    
            checkin_time = None
            total_working_time = 0
            attendance_log=[]
            attendance_history=[]

            for index, entry in enumerate(sorted_data):

                if entry['checkout']:
                    if entry['deviceId'] is not None and entry['deviceId'] !='':
                        attendance_type="Machine Checkout"
                        attendance_type_resaon=""
                    elif entry['Remote_Reason'] is not None and entry['Remote_Reason'] !='':
                        attendance_type="Remote Checkout"
                        attendance_type_resaon=entry['Remote_Reason']
                    else:
                        attendance_type="Online Checkout"
                        attendance_type_resaon=''                                

                    attendance_history.append({'Status':'Check-Out','datetime':entry['date']+' '+entry['time'],'attendance_type':attendance_type,'attendance_type_resaon':attendance_type_resaon})
                        
                    attendance_log.append({'checkout':entry['date']+' '+entry['time']})
                    if checkin_time:
                        checkout_datetime = datetime.strptime(entry['date'] + ' ' + entry['time'], '%Y-%m-%d %H:%M:%S')
                        checkin_datetime = datetime.strptime(checkin_time, '%Y-%m-%d %H:%M:%S')
                        working_time = checkout_datetime - checkin_datetime
                        total_working_time += working_time.total_seconds()
                        checkin_time = None
                        
                elif not entry['checkout']:
                    if entry['deviceId'] is not None and entry['deviceId'] !='':
                        attendance_type="Machine Checkin"
                        attendance_type_resaon=''
                    elif entry['Remote_Reason'] is not None and entry['Remote_Reason'] !='':
                        attendance_type="Remote Checkin"
                        attendance_type_resaon=entry['Remote_Reason']
                    else:
                        attendance_type="Online Checkin"
                        attendance_type_resaon=''  

                    checkin_time = entry['date'] + ' ' + entry['time']
                    attendance_log.append({'checkin':entry['date']+' '+entry['time']})
                    attendance_history.append({'Status':'Check-In','datetime':entry['date']+' '+entry['time'],'attendance_type':attendance_type,'attendance_type_resaon':attendance_type_resaon})
            # If the last entry is check-in, consider checkout time as current shift end  time
            if checkin_time and index == len(sorted_data) - 1:
                
                # check if the previous entry is also checkin or not if exist get the  difference betwnn  the current checkin  and  previous  checkin

                if int(int(index)-1) >=0:
                    if sorted_data[index-1]['checkout']==False:
                        previous_checkin_date_time=sorted_data[index-1]['date']+ ' '+sorted_data[index-1]['time']
                        checkout_datetime = datetime.strptime(previous_checkin_date_time, '%Y-%m-%d %H:%M:%S')
                        checkin_datetime = datetime.strptime(checkin_time, '%Y-%m-%d %H:%M:%S')
                        working_time = checkin_datetime-checkout_datetime 
                        total_working_time += working_time.total_seconds()


            # Convert total_working_time to hours, minutes, and seconds
            hours, remainder = divmod(total_working_time, 3600)
            minutes, seconds = divmod(remainder, 60)
            s['total_hrs']=str(add_leading_zero(int(hours))) +':' +str(add_leading_zero(int(minutes))) +':'+str(add_leading_zero(int(seconds)))

            if total_working_hrs == '':

                total_working_hrs=s['total_hrs']
            else:
                time_str_1 = total_working_hrs
                time_str_2 = s['total_hrs']



                # Convert time strings to timedelta objects and add them
                total_time_delta = timedelta(hours=int(time_str_1[:2]), minutes=int(time_str_1[3:5]), seconds=int(time_str_1[6:])) + \
                                timedelta(hours=int(time_str_2[:2]), minutes=int(time_str_2[3:5]), seconds=int(time_str_2[6:]))
                total_working_hrs=str(total_time_delta)
                
            s['attendance_log']=attendance_log
            s['attendance_history']=attendance_history
            if len(attendance_log) !=0:
                
                infirst_key = next(iter(attendance_log[0]))
                outfirst_key = next(iter(attendance_log[-1]))
                
                s['usersintime']=attendance_log[0][infirst_key]
                s['usersouttime']=attendance_log[-1][outfirst_key]
                if s['usersouttime'] == s['usersintime']:
                    s['usersouttime']= ''
                    
            else:
                s['usersintime']=''
                s['usersouttime']=''


            
            if current_date_first_in_datetime =='':
                if s['usersintime'] !='':
                    current_date_first_in_datetime=s['usersintime']
            elif s['usersintime'] !='':
                if datetime.strptime(current_date_first_in_datetime, '%Y-%m-%d %H:%M:%S') > datetime.strptime(s['usersintime'], '%Y-%m-%d %H:%M:%S'):
                    current_date_first_in_datetime=s['usersintime']

            if current_date_last_out_datetime =='':
                if s['usersouttime'] !='':
                    current_date_last_out_datetime=s['usersouttime']
            elif s['usersouttime'] !='':
                if datetime.strptime(current_date_last_out_datetime, '%Y-%m-%d %H:%M:%S') < datetime.strptime(s['usersouttime'], '%Y-%m-%d %H:%M:%S'):
                    current_date_last_out_datetime=s['usersouttime']

                
            if current_date_last_out_datetime != '' and current_date_first_in_datetime !='':
                if current_date_last_out_datetime.split(' ')[0] != current_date_first_in_datetime.split(' ')[0]:
                    current_date_last_out_datetime1 = datetime.strptime(str(current_date_last_out_datetime).split(' ')[0], '%Y-%m-%d')
                    current_date_first_in_datetime1 = datetime.strptime(str(current_date_first_in_datetime).split(' ')[0], '%Y-%m-%d')
                    extradaysdiff = current_date_last_out_datetime1 - current_date_first_in_datetime1
                    if extradaysdiff.days > 0:
                        extra_days='+'+str(extradaysdiff.days)
                
        return_dict['employeeId']=attendanceId
        return_dict['empname']=user_obj.Firstname +' '+ user_obj.Lastname
        return_dict['username']=user_obj.Firstname +' '+ user_obj.Lastname
        
        return_dict['reason']='Attendance not found'
        return_dict['fulldate']=date
        return_dict['shifts_list']=current_date_shift_list
        return_dict['shift']=new_current_date_shift_list[0]

        return_dict['Leavetype']=''


        
        if len(new_current_date_shift_list) > 1:
            shift_name=new_current_date_shift_list[0]['shiftname']
            return_dict['shift']['shiftname'] = str(shift_name) +  ' <span class="text-danger"> +1 </span>'
            new_current_date_shift_list[0]['swap']=''
        else:
            new_current_date_shift_list[0]['swap']=''
            
        if current_date_first_in_datetime !='':
            return_dict['checkintime']=str(current_date_first_in_datetime).split(' ')[1]
            checkin_attendance_type=attendance.objects.filter(time=str(current_date_first_in_datetime).split(' ')[1],date=str(current_date_first_in_datetime).split(' ')[0]).first()

            

                        
                        
            if checkin_attendance_type is not None:
                remote_map_location=''
                remote_map_name=''
                if checkin_attendance_type.remote_latitude is not None and checkin_attendance_type.remote_longitude is not None and checkin_attendance_type.remote_latitude !='' and checkin_attendance_type.remote_longitude !='':
                    remote_map_location=create_google_maps_url(checkin_attendance_type.remote_latitude,checkin_attendance_type.remote_longitude)
                    remote_map_name=get_location_name(checkin_attendance_type.remote_latitude,checkin_attendance_type.remote_longitude)
                
                if checkin_attendance_type.deviceId is not None and checkin_attendance_type.deviceId !='':
                    return_dict['check_in_attendance_type']='Machine CheckIn'
                    return_dict['check_in_attendance_type_resaon']=''  

                    if checkin_attendance_type.deviceId == 20:
                        remote_map_name='Zentro Pune Office'
                        remote_map_location=create_google_maps_url(18.566064883698555, 73.77574703366226)
                    elif checkin_attendance_type.deviceId == 19:
                        remote_map_name='Zentro Mumbai Office'
                        remote_map_location=create_google_maps_url(19.1764898841813, 72.95988765068624)

                    return_dict['check_in_remote_map_name']=remote_map_name
                    return_dict['check_in_remote_map_location']=remote_map_location
                    
                elif checkin_attendance_type.Remote_Reason is not None and checkin_attendance_type.Remote_Reason !='':
                    return_dict['check_in_attendance_type']="Remote CheckIn"
                    return_dict['check_in_attendance_type_resaon']=checkin_attendance_type.Remote_Reason
                    return_dict['check_in_remote_map_name']=remote_map_name
                    return_dict['check_in_remote_map_location']=remote_map_location
                    
                else:
                    return_dict['check_in_attendance_type']="Online CheckIn"
                    return_dict['check_in_attendance_type_resaon']='' 
                    if checkin_attendance_type.attendance_type !='' and  checkin_attendance_type.attendance_type is not None:
                        return_dict['check_in_attendance_type'] = checkin_attendance_type.attendance_type+' ' + 'CheckIn'
                     
                    return_dict['check_in_remote_map_name']=remote_map_name
                    return_dict['check_in_remote_map_location']=remote_map_location
                    
            return_dict['inTime_date']=convertdate2(str(str(current_date_first_in_datetime).split(' ')[0]))
        else:
            return_dict['checkintime']='--:--'
            return_dict['inTime_date']=''
            return_dict['check_in_attendance_type']=""
            return_dict['check_in_attendance_type_resaon']='' 
            
            
        if current_date_last_out_datetime !='':
            return_dict['checkouttime']=str(current_date_last_out_datetime).split(' ')[1]
            return_dict['outTime_date']=convertdate2(str(str(current_date_last_out_datetime).split(' ')[0]))

            checkout_attendance_type=attendance.objects.filter(time=str(current_date_last_out_datetime).split(' ')[1],date=str(current_date_last_out_datetime).split(' ')[0]).first()
            if checkout_attendance_type is not None:
                remote_map_location=''
                remote_map_name=''
                if checkout_attendance_type.remote_latitude is not None and checkout_attendance_type.remote_longitude is not None and checkout_attendance_type.remote_latitude !='' and checkout_attendance_type.remote_longitude !='':
                    remote_map_location=create_google_maps_url(checkout_attendance_type.remote_latitude,checkout_attendance_type.remote_longitude)
                    remote_map_name=get_location_name(checkout_attendance_type.remote_latitude,checkout_attendance_type.remote_longitude)
                
                
                
                if checkout_attendance_type.deviceId is not None and checkout_attendance_type.deviceId !='':
                    return_dict['check_out_attendance_type']='Machine CheckOut'
                    return_dict['check_out_attendance_type_resaon']=''
                      
                    if checkout_attendance_type.deviceId == 20:
                        remote_map_name='Zentro Pune Office'                        
                        remote_map_location=create_google_maps_url(18.566064883698555, 73.77574703366226)

                    elif checkout_attendance_type.deviceId == 19:
                        remote_map_name='Zentro Mumbai Office'
                        remote_map_location=create_google_maps_url(19.1764898841813, 72.95988765068624)

                    return_dict['checkout_remote_map_name']=remote_map_name
                    return_dict['checkout_remote_map_location']=remote_map_location
                    
                    
                elif checkout_attendance_type.Remote_Reason is not None and checkout_attendance_type.Remote_Reason !='':
                    return_dict['check_out_attendance_type']="Remote CheckOut"
                    return_dict['check_out_attendance_type_resaon']=checkout_attendance_type.Remote_Reason

                    return_dict['checkout_remote_map_name']=remote_map_name
                    return_dict['checkout_remote_map_location']=remote_map_location
                else:

                        
                    return_dict['check_out_attendance_type']="Online CheckOut"
                    return_dict['check_out_attendance_type_resaon']=''  
                    
                    if checkout_attendance_type.attendance_type !='' and  checkout_attendance_type.attendance_type is not None:
                        return_dict['check_out_attendance_type'] = checkout_attendance_type.attendance_type +' ' + 'CheckOut'
                        
                    return_dict['checkout_remote_map_name']=remote_map_name
                    return_dict['checkout_remote_map_location']=remote_map_location
            
        else:
            return_dict['checkouttime']='--:--'
            return_dict['outTime_date']=''
            return_dict['check_out_attendance_type']=""
            return_dict['check_out_attendance_type_resaon']='' 
            
            
        return_dict['extra_days']=extra_days
        return_dict['total_time']=total_working_hrs
        if return_dict['checkintime'] !='' and return_dict['checkintime'] !='--:--':
            return_dict['att_status']="P"
            

        else:
            return_dict['att_status']="A"

        foundleaves = [
            leave for leave in leavesserializer.data if datetime.strptime(leave['start_date'], '%Y-%m-%d').date() <= datetime.strptime(date, '%Y-%m-%d').date() <= datetime.strptime(leave['end_date'], '%Y-%m-%d').date()
        ]
        
        foundwfhs = [
            wfh for wfh in wfhsserializer.data if datetime.strptime(wfh['start_date'], '%Y-%m-%d').date() <= datetime.strptime(date, '%Y-%m-%d').date() <= datetime.strptime(wfh['end_date'], '%Y-%m-%d').date()
        ]
        
        if foundwfhs:
            return_dict['att_status']="WFH"
        if foundleaves:
            return_dict['att_status']="L"
            
            
            
        if return_dict['att_status']=='P':
            Presentlist.append(return_dict)
        elif return_dict['att_status']=='L':
            leavelist.append(return_dict)
        elif return_dict['att_status']=='WFH':
            wfhlist.append(return_dict)
            
        else:
            yettojoinlist.append(return_dict)  
            
            
            
    context= {
        "All_Mapped_Emp_Count": len(Userlist),
        "Total_count": len(Userlist),
        "Present": len(Presentlist),
        "Onleave": len(leavelist),
        "yettojoin": len(yettojoinlist),
        "WFH": len(wfhlist),
        "Presentlist":Presentlist,
        "leavelist": leavelist,
        "wfhlist":wfhlist,
        "yettojoinlist": yettojoinlist,
        }
    
    return Response ({"context":context, "response":{"n" : 1,"msg" : "SUCCESS","status" : "success"}})

@api_view(['POST'])    
@permission_classes((AllowAny,))
def get_all_emp_attendance_by_date(request):

    date = request.POST.get("my_date")
    month=date.split('-')[1]
    year=date.split('-')[0]
    

    
    Users_l1l2emplst_obj=Users.objects.filter(is_active=True).exclude(Q(employeeId__isnull=True)|Q(employeeId='None')|Q(employeeId=''))
    Users_l1l2serializers = UsersSerializeronlyattendance(Users_l1l2emplst_obj,many=True)
    Userlist=list(Users_l1l2serializers.data)




    Presentlist=[]
    leavelist=[]
    wfhlist=[]
    yettojoinlist=[]
    
    
    for attendanceId in Userlist:
        user_obj=Users.objects.filter(employeeId=attendanceId,is_active=True).first()
        leave_objs = Leave.objects.filter(employeeId=user_obj.id,WorkFromHome=False,leave_status="Approved",start_date__month__gte=month,start_date__month__lte=month,start_date__year= year,Active=True).exclude(leave_status='Draft')
        leavesserializer=leaveserializer(leave_objs,many=True)
        wfh_objs = Leave.objects.filter(employeeId=user_obj.id,WorkFromHome=True,leave_status="Approved",start_date__month__gte=month,start_date__month__lte=month,start_date__year= year,Active=True).exclude(leave_status='Draft')
        wfhsserializer=leaveserializer(wfh_objs,many=True)
    
        current_date_shift_list=[]
        new_current_date_shift_list=[]
        return_dict={}
        
        shiftdate = datetime.strptime(date, '%Y-%m-%d').date()
        tomarrow_date = shiftdate + timedelta(days=1)
        yesterday_date = shiftdate - timedelta(days=1)
        current_shiftallotment_objs=ShiftAllotment.objects.filter(attendanceId=attendanceId,date=str(date),is_active=True)
        current_shiftallotment_serializers=ShiftAllotmentshiftIdSerializer(current_shiftallotment_objs,many=True)
        current_shiftId_list=list(current_shiftallotment_serializers.data)
        current_shift_obj=ShiftMaster.objects.filter(id__in=current_shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime')
        current_shift_serializer=ShiftMasterSerializer(current_shift_obj,many=True)
        current_shiftlist=current_shift_serializer.data
        
        tomarrow_shiftallotment_objs=ShiftAllotment.objects.filter(attendanceId=attendanceId,date=str(tomarrow_date),is_active=True)
        tomarrow_shiftallotment_serializers=ShiftAllotmentshiftIdSerializer(tomarrow_shiftallotment_objs,many=True)
        tomarrow_shiftId_list=list(tomarrow_shiftallotment_serializers.data)
        tomarrow_shift_obj=ShiftMaster.objects.filter(id__in=tomarrow_shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime')
        tomarrow_shift_serializer=ShiftMasterSerializer(tomarrow_shift_obj,many=True)
        tomarrow_shiftlist=tomarrow_shift_serializer.data

        count=1
        
        for shift in current_shiftlist:
            start_time = datetime.strptime(str(shiftdate) +' '+ shift['intime'], '%Y-%m-%d %H:%M')
            start_time_before_2hrs = start_time - timedelta(hours=2)
            
            if shift['intime'] > shift['outtime']:
                shift_end_date = shiftdate + timedelta(days=1)
            else:
                shift_end_date = shiftdate
                
            end_time = datetime.strptime(str(shift_end_date) +' '+ shift['outtime'], '%Y-%m-%d %H:%M')
            check_login_till=''
            if len(current_shiftlist) >= count+1:
            
                check_next_shift_in = datetime.strptime(str(shiftdate) +' '+ current_shiftlist[count]['intime'], '%Y-%m-%d %H:%M')
                check_login_till = check_next_shift_in - timedelta(hours=2)

                
            elif check_login_till=='':
                    
                if len(tomarrow_shiftlist) >= 1:
                    
                    check_next_shift_in = datetime.strptime(str(tomarrow_date) +' '+ tomarrow_shiftlist[0]['intime'], '%Y-%m-%d %H:%M')
                    check_login_till = check_next_shift_in - timedelta(hours=2)
                else:
                
                    if_no_shift_check_shift=ShiftAllotment.objects.filter(attendanceId=attendanceId,date=str(tomarrow_date),is_active=True)
                    tomerrow_shiftlist=ShiftAllotmentshiftIdSerializer(if_no_shift_check_shift,many=True)
                    next_date_shiftId_list=list(tomerrow_shiftlist.data)
                    check_weekly_off=ShiftMaster.objects.filter(id__in=next_date_shiftId_list,intime='00:00',outtime='00:00',is_active=True).order_by('intime').first()
                    if check_weekly_off is not None:
                        
                        if shift['outtime'] < shift['intime']:
                    
                            check_current_shift_out = datetime.strptime(str(tomarrow_date) +' '+ shift['outtime'], '%Y-%m-%d %H:%M')
                            check_current_shift_out_time = check_current_shift_out + timedelta(hours=2)
                            given_datetime = datetime.strptime(str(check_current_shift_out_time), '%Y-%m-%d %H:%M:%S')

                            truncated_datetime_str = given_datetime.strftime('%Y-%m-%d %H:%M')

                            
                            shift_end_date_time=str(truncated_datetime_str)
                        else:
                            shift_end_date_time=str(date) +' '+ '23:59'
                            
                        check_login_till = datetime.strptime(shift_end_date_time, '%Y-%m-%d %H:%M')
                    else:
                    
                        check_login_till = datetime.strptime(str(tomarrow_date) +' '+ '07:30', '%Y-%m-%d %H:%M')
                        
            shift_data={
                'shiftname':shift['shiftname'],
                'indatetime':str(start_time_before_2hrs),
                'outdatetime':str(check_login_till),
                'intime':shift['intime'],
                'outtime':shift['outtime'],
            }
            current_date_shift_list.append(shift_data)
            new_current_date_shift_list.append(shift_data)
            count+=1

        if len(current_shiftlist) == 0:       
        
            if_no_shift_check_shift=ShiftAllotment.objects.filter(attendanceId=attendanceId,date=str(date),is_active=True)
            todays_shiftlist=ShiftAllotmentshiftIdSerializer(if_no_shift_check_shift,many=True)
            toadys_shiftId_list=list(todays_shiftlist.data)
            check_weekly_off=ShiftMaster.objects.filter(id__in=toadys_shiftId_list,intime='00:00',outtime='00:00',is_active=True).order_by('intime').first()
            if check_weekly_off is not None:
                get_previous_day_shift=ShiftAllotment.objects.filter(attendanceId=attendanceId,date=str(yesterday_date),is_active=True)
                previous_day_shiftlist=ShiftAllotmentshiftIdSerializer(get_previous_day_shift,many=True)
                previous_shiftId_list=list(previous_day_shiftlist.data)
                get_previous_last_shift=ShiftMaster.objects.filter(id__in=previous_shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime').last()
                if get_previous_last_shift is not None:
                    if get_previous_last_shift.outtime < get_previous_last_shift.intime:
                        shift_end_date_time=str(date) +' '+ str(get_previous_last_shift.outtime)
                    else:
                        shift_end_date_time=str(yesterday_date) +' '+ '23:59'
                            
                    check_previous_shift_out = datetime.strptime(str(shift_end_date_time), '%Y-%m-%d %H:%M')
                    previous_shift_in_time = check_previous_shift_out + timedelta(hours=2)  
                else:
                    previous_shift_in_time= str(date)+' 00:00:00'
                get_net_day_shift=ShiftAllotment.objects.filter(attendanceId=attendanceId,date=str(tomarrow_date),is_active=True)
                nest_day_shiftlist=ShiftAllotmentshiftIdSerializer(get_net_day_shift,many=True)
                next_date_shiftId_list=list(nest_day_shiftlist.data)
                get_next_day_first_shift=ShiftMaster.objects.filter(id__in=next_date_shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime').first()
                if get_next_day_first_shift is not None:
                    check_next_shift_in = datetime.strptime(str(tomarrow_date) +' '+ str(get_next_day_first_shift.intime), '%Y-%m-%d %H:%M')
                    next_shift_in_time = check_next_shift_in - timedelta(hours=2)
                    shift_data={
                        'shiftname':str(check_weekly_off.shiftname),
                        'indatetime':str(previous_shift_in_time),
                        'outdatetime':str(next_shift_in_time),
                        'intime':check_weekly_off.intime,
                        'outtime':check_weekly_off.outtime,
                    }
                    current_date_shift_list.append(shift_data)
                    new_current_date_shift_list.append(shift_data)
                    
                else:
                    shift_data={
                        'shiftname':str(check_weekly_off.shiftname),
                        'indatetime':str(previous_shift_in_time),
                        'outdatetime':str(date)+' 23:59:59',
                        'intime':check_weekly_off.intime,
                        'outtime':check_weekly_off.outtime,
                    }
                    current_date_shift_list.append(shift_data)
                    new_current_date_shift_list.append(shift_data)

            else:
                shift_data={
                    'shiftname':'General',
                    'indatetime':str(date)+' 07:30:00',
                    'outdatetime':str(tomarrow_date)+' 07:30:00',
                    'intime':'09:30',
                    'outtime':'18:30',
                }
                current_date_shift_list.append(shift_data)
                new_current_date_shift_list.append(shift_data)

                
        current_date_first_in_datetime=''
        current_date_last_out_datetime=''
        extra_days=''
        total_working_hrs=''
        for s in new_current_date_shift_list:
            intime=''
            intimedate=''
            
            
            getallattendance = attendance.objects.filter(Q(employeeId=str(attendanceId),time__gte=s['indatetime'].split(' ')[1],date=s['indatetime'].split(' ')[0])|Q(employeeId=str(attendanceId),time__lte=str(s['outdatetime']).split(' ')[1],date=str(s['outdatetime']).split(' ')[0])).order_by('date','time')
            
            # getallattendance = getallattendance.filter(time__gte=currentshiftstarttime).order_by('date','time')
            
            attendance_serializer=attendanceserializer(getallattendance,many=True)
        
            
            sorted_data = sorted(attendance_serializer.data, key=lambda x: (x['date'], x['time']))
            
            mindatetime = datetime.strptime(s['indatetime'], '%Y-%m-%d %H:%M:%S')
            maxdatetime = datetime.strptime(s['outdatetime'], '%Y-%m-%d %H:%M:%S')
        
            sorted_data = [entry for entry in sorted_data if (mindatetime <= datetime.strptime(entry['date'] +' '+entry['time'],'%Y-%m-%d %H:%M:%S') <= maxdatetime)]
            if len(sorted_data) > 0:
                intimedate=sorted_data[0]['date']
                intime=str(sorted_data[0]['time'])
                
            if intimedate !='' and intimedate is not None:
                user_sdt = datetime.strptime(str(intimedate) + ' ' + str(intime), '%Y-%m-%d %H:%M:%S')
                shif_sdt = datetime.strptime(s['indatetime'].split(' ')[0] + ' ' + s['indatetime'].split(' ')[1], '%Y-%m-%d %H:%M:%S')
                if user_sdt < shif_sdt :
                    intimedate=''
                    intime=''

                    
            checkin_time = None
            total_working_time = 0
            attendance_log=[]
            attendance_history=[]

            for index, entry in enumerate(sorted_data):

                if entry['checkout']:
                    if entry['deviceId'] is not None and entry['deviceId'] !='':
                        attendance_type="Machine Checkout"
                        attendance_type_resaon=""
                    elif entry['Remote_Reason'] is not None and entry['Remote_Reason'] !='':
                        attendance_type="Remote Checkout"
                        attendance_type_resaon=entry['Remote_Reason']
                    else:
                        attendance_type="Online Checkout"
                        attendance_type_resaon=''                                

                    attendance_history.append({'Status':'Check-Out','datetime':entry['date']+' '+entry['time'],'attendance_type':attendance_type,'attendance_type_resaon':attendance_type_resaon})
                        
                    attendance_log.append({'checkout':entry['date']+' '+entry['time']})
                    if checkin_time:
                        checkout_datetime = datetime.strptime(entry['date'] + ' ' + entry['time'], '%Y-%m-%d %H:%M:%S')
                        checkin_datetime = datetime.strptime(checkin_time, '%Y-%m-%d %H:%M:%S')
                        working_time = checkout_datetime - checkin_datetime
                        total_working_time += working_time.total_seconds()
                        checkin_time = None
                        
                elif not entry['checkout']:
                    if entry['deviceId'] is not None and entry['deviceId'] !='':
                        attendance_type="Machine Checkin"
                        attendance_type_resaon=''
                    elif entry['Remote_Reason'] is not None and entry['Remote_Reason'] !='':
                        attendance_type="Remote Checkin"
                        attendance_type_resaon=entry['Remote_Reason']
                    else:
                        attendance_type="Online Checkin"
                        attendance_type_resaon=''  

                    checkin_time = entry['date'] + ' ' + entry['time']
                    attendance_log.append({'checkin':entry['date']+' '+entry['time']})
                    attendance_history.append({'Status':'Check-In','datetime':entry['date']+' '+entry['time'],'attendance_type':attendance_type,'attendance_type_resaon':attendance_type_resaon})
            # If the last entry is check-in, consider checkout time as current shift end  time
            if checkin_time and index == len(sorted_data) - 1:
                
                # check if the previous entry is also checkin or not if exist get the  difference betwnn  the current checkin  and  previous  checkin

                if int(int(index)-1) >=0:
                    if sorted_data[index-1]['checkout']==False:
                        previous_checkin_date_time=sorted_data[index-1]['date']+ ' '+sorted_data[index-1]['time']
                        checkout_datetime = datetime.strptime(previous_checkin_date_time, '%Y-%m-%d %H:%M:%S')
                        checkin_datetime = datetime.strptime(checkin_time, '%Y-%m-%d %H:%M:%S')
                        working_time = checkin_datetime-checkout_datetime 
                        total_working_time += working_time.total_seconds()


            # Convert total_working_time to hours, minutes, and seconds
            hours, remainder = divmod(total_working_time, 3600)
            minutes, seconds = divmod(remainder, 60)
            s['total_hrs']=str(add_leading_zero(int(hours))) +':' +str(add_leading_zero(int(minutes))) +':'+str(add_leading_zero(int(seconds)))

            if total_working_hrs == '':

                total_working_hrs=s['total_hrs']
            else:
                time_str_1 = total_working_hrs
                time_str_2 = s['total_hrs']



                # Convert time strings to timedelta objects and add them
                total_time_delta = timedelta(hours=int(time_str_1[:2]), minutes=int(time_str_1[3:5]), seconds=int(time_str_1[6:])) + \
                                timedelta(hours=int(time_str_2[:2]), minutes=int(time_str_2[3:5]), seconds=int(time_str_2[6:]))
                total_working_hrs=str(total_time_delta)
                
            s['attendance_log']=attendance_log
            s['attendance_history']=attendance_history
            if len(attendance_log) !=0:
                
                infirst_key = next(iter(attendance_log[0]))
                outfirst_key = next(iter(attendance_log[-1]))
                
                s['usersintime']=attendance_log[0][infirst_key]
                s['usersouttime']=attendance_log[-1][outfirst_key]
                if s['usersouttime'] == s['usersintime']:
                    s['usersouttime']= ''
                    
            else:
                s['usersintime']=''
                s['usersouttime']=''


            
            if current_date_first_in_datetime =='':
                if s['usersintime'] !='':
                    current_date_first_in_datetime=s['usersintime']
            elif s['usersintime'] !='':
                if datetime.strptime(current_date_first_in_datetime, '%Y-%m-%d %H:%M:%S') > datetime.strptime(s['usersintime'], '%Y-%m-%d %H:%M:%S'):
                    current_date_first_in_datetime=s['usersintime']

            if current_date_last_out_datetime =='':
                if s['usersouttime'] !='':
                    current_date_last_out_datetime=s['usersouttime']
            elif s['usersouttime'] !='':
                if datetime.strptime(current_date_last_out_datetime, '%Y-%m-%d %H:%M:%S') < datetime.strptime(s['usersouttime'], '%Y-%m-%d %H:%M:%S'):
                    current_date_last_out_datetime=s['usersouttime']

                
            if current_date_last_out_datetime != '' and current_date_first_in_datetime !='':
                if current_date_last_out_datetime.split(' ')[0] != current_date_first_in_datetime.split(' ')[0]:
                    current_date_last_out_datetime1 = datetime.strptime(str(current_date_last_out_datetime).split(' ')[0], '%Y-%m-%d')
                    current_date_first_in_datetime1 = datetime.strptime(str(current_date_first_in_datetime).split(' ')[0], '%Y-%m-%d')
                    extradaysdiff = current_date_last_out_datetime1 - current_date_first_in_datetime1
                    if extradaysdiff.days > 0:
                        extra_days='+'+str(extradaysdiff.days)
                
        return_dict['employeeId']=attendanceId
        return_dict['empname']=user_obj.Firstname +' '+ user_obj.Lastname
        return_dict['username']=user_obj.Firstname +' '+ user_obj.Lastname
        
        return_dict['reason']='Attendance not found'
        return_dict['fulldate']=date
        return_dict['shifts_list']=current_date_shift_list
        return_dict['shift']=new_current_date_shift_list[0]

        return_dict['Leavetype']=''


        
        if len(new_current_date_shift_list) > 1:
            shift_name=new_current_date_shift_list[0]['shiftname']
            return_dict['shift']['shiftname'] = str(shift_name) +  ' <span class="text-danger"> +1 </span>'
            new_current_date_shift_list[0]['swap']=''
        else:
            new_current_date_shift_list[0]['swap']=''
            
        if current_date_first_in_datetime !='':
            return_dict['checkintime']=str(current_date_first_in_datetime).split(' ')[1]
            checkin_attendance_type=attendance.objects.filter(time=str(current_date_first_in_datetime).split(' ')[1],date=str(current_date_first_in_datetime).split(' ')[0]).first()

            

                        
                        
            if checkin_attendance_type is not None:
                remote_map_location=''
                remote_map_name=''
                if checkin_attendance_type.remote_latitude is not None and checkin_attendance_type.remote_longitude is not None and checkin_attendance_type.remote_latitude !='' and checkin_attendance_type.remote_longitude !='':
                    remote_map_location=create_google_maps_url(checkin_attendance_type.remote_latitude,checkin_attendance_type.remote_longitude)
                    remote_map_name=get_location_name(checkin_attendance_type.remote_latitude,checkin_attendance_type.remote_longitude)
                
                if checkin_attendance_type.deviceId is not None and checkin_attendance_type.deviceId !='':
                    return_dict['check_in_attendance_type']='Machine CheckIn'
                    return_dict['check_in_attendance_type_resaon']=''  

                    if checkin_attendance_type.deviceId == 20:
                        remote_map_name='Zentro Pune Office'
                        remote_map_location=create_google_maps_url(18.566064883698555, 73.77574703366226)
                    elif checkin_attendance_type.deviceId == 19:
                        remote_map_name='Zentro Mumbai Office'
                        remote_map_location=create_google_maps_url(19.1764898841813, 72.95988765068624)

                    return_dict['check_in_remote_map_name']=remote_map_name
                    return_dict['check_in_remote_map_location']=remote_map_location
                    
                elif checkin_attendance_type.Remote_Reason is not None and checkin_attendance_type.Remote_Reason !='':
                    return_dict['check_in_attendance_type']="Remote CheckIn"
                    return_dict['check_in_attendance_type_resaon']=checkin_attendance_type.Remote_Reason
                    return_dict['check_in_remote_map_name']=remote_map_name
                    return_dict['check_in_remote_map_location']=remote_map_location
                    
                else:
                    return_dict['check_in_attendance_type']="Online CheckIn"
                    return_dict['check_in_attendance_type_resaon']='' 
                    if checkin_attendance_type.attendance_type !='' and  checkin_attendance_type.attendance_type is not None:
                        return_dict['check_in_attendance_type'] = checkin_attendance_type.attendance_type+' ' + 'CheckIn'
                     
                    return_dict['check_in_remote_map_name']=remote_map_name
                    return_dict['check_in_remote_map_location']=remote_map_location
                    
            return_dict['inTime_date']=convertdate2(str(str(current_date_first_in_datetime).split(' ')[0]))
        else:
            return_dict['checkintime']='--:--'
            return_dict['inTime_date']=''
            return_dict['check_in_attendance_type']=""
            return_dict['check_in_attendance_type_resaon']='' 
            
            
        if current_date_last_out_datetime !='':
            return_dict['checkouttime']=str(current_date_last_out_datetime).split(' ')[1]
            return_dict['outTime_date']=convertdate2(str(str(current_date_last_out_datetime).split(' ')[0]))

            checkout_attendance_type=attendance.objects.filter(time=str(current_date_last_out_datetime).split(' ')[1],date=str(current_date_last_out_datetime).split(' ')[0]).first()
            if checkout_attendance_type is not None:
                remote_map_location=''
                remote_map_name=''
                if checkout_attendance_type.remote_latitude is not None and checkout_attendance_type.remote_longitude is not None and checkout_attendance_type.remote_latitude !='' and checkout_attendance_type.remote_longitude !='':
                    remote_map_location=create_google_maps_url(checkout_attendance_type.remote_latitude,checkout_attendance_type.remote_longitude)
                    remote_map_name=get_location_name(checkout_attendance_type.remote_latitude,checkout_attendance_type.remote_longitude)
                
                
                
                if checkout_attendance_type.deviceId is not None and checkout_attendance_type.deviceId !='':
                    return_dict['check_out_attendance_type']='Machine CheckOut'
                    return_dict['check_out_attendance_type_resaon']=''
                      
                    if checkout_attendance_type.deviceId == 20:
                        remote_map_name='Zentro Pune Office'                        
                        remote_map_location=create_google_maps_url(18.566064883698555, 73.77574703366226)

                    elif checkout_attendance_type.deviceId == 19:
                        remote_map_name='Zentro Mumbai Office'
                        remote_map_location=create_google_maps_url(19.1764898841813, 72.95988765068624)

                    return_dict['checkout_remote_map_name']=remote_map_name
                    return_dict['checkout_remote_map_location']=remote_map_location
                    
                    
                elif checkout_attendance_type.Remote_Reason is not None and checkout_attendance_type.Remote_Reason !='':
                    return_dict['check_out_attendance_type']="Remote CheckOut"
                    return_dict['check_out_attendance_type_resaon']=checkout_attendance_type.Remote_Reason

                    return_dict['checkout_remote_map_name']=remote_map_name
                    return_dict['checkout_remote_map_location']=remote_map_location
                else:

                        
                    return_dict['check_out_attendance_type']="Online CheckOut"
                    return_dict['check_out_attendance_type_resaon']=''  
                    
                    if checkout_attendance_type.attendance_type !='' and  checkout_attendance_type.attendance_type is not None:
                        return_dict['check_out_attendance_type'] = checkout_attendance_type.attendance_type +' ' + 'CheckOut'
                        
                    return_dict['checkout_remote_map_name']=remote_map_name
                    return_dict['checkout_remote_map_location']=remote_map_location
            
        else:
            return_dict['checkouttime']='--:--'
            return_dict['outTime_date']=''
            return_dict['check_out_attendance_type']=""
            return_dict['check_out_attendance_type_resaon']='' 
            
            
        return_dict['extra_days']=extra_days
        return_dict['total_time']=total_working_hrs
        if return_dict['checkintime'] !='' and return_dict['checkintime'] !='--:--':
            return_dict['att_status']="P"
            

        else:
            return_dict['att_status']="A"

        foundleaves = [
            leave for leave in leavesserializer.data if datetime.strptime(leave['start_date'], '%Y-%m-%d').date() <= datetime.strptime(date, '%Y-%m-%d').date() <= datetime.strptime(leave['end_date'], '%Y-%m-%d').date()
        ]
        
        foundwfhs = [
            wfh for wfh in wfhsserializer.data if datetime.strptime(wfh['start_date'], '%Y-%m-%d').date() <= datetime.strptime(date, '%Y-%m-%d').date() <= datetime.strptime(wfh['end_date'], '%Y-%m-%d').date()
        ]
        
        if foundwfhs:
            return_dict['att_status']="WFH"
        if foundleaves:
            return_dict['att_status']="L"
            
            
            
        if return_dict['att_status']=='P':
            Presentlist.append(return_dict)
        elif return_dict['att_status']=='L':
            leavelist.append(return_dict)
        elif return_dict['att_status']=='WFH':
            wfhlist.append(return_dict)
            
        else:
            yettojoinlist.append(return_dict)  
            
            
            
    context= {
        "All_Mapped_Emp_Count": len(Userlist),
        "Total_count": len(Userlist),
        "Present": len(Presentlist),
        "Onleave": len(leavelist),
        "yettojoin": len(yettojoinlist),
        "WFH": len(wfhlist),
        "Presentlist":Presentlist,
        "leavelist": leavelist,
        "wfhlist":wfhlist,
        "yettojoinlist": yettojoinlist,
        }
    
    return Response ({"context":context, "response":{"n" : 1,"msg" : "SUCCESS","status" : "success"}})

@api_view(['POST'])    
@permission_classes((AllowAny,))
def get_sift_emp_attendance_and_task_by_date(request):

    date = request.POST.get("my_date")
    shiftid = request.POST.get("shiftid")
    month=date.split('-')[1]
    year=date.split('-')[0]
    
    l1l2emplst_obj=EmployeeShiftDetails.objects.filter(is_active=True)
    l1l2serializers = L1L2Serializer(l1l2emplst_obj,many=True) 
    team_list=list(l1l2serializers.data)
    
    Users_l1l2emplst_obj=Users.objects.filter(id__in=team_list,is_active=True).exclude(Q(employeeId__isnull=True)|Q(employeeId='None')|Q(employeeId=''))
    Users_l1l2serializers = UsersSerializeronlyattendance(Users_l1l2emplst_obj,many=True)
    Userlist=list(Users_l1l2serializers.data)




    commonlist=[]
    
    for attendanceId in Userlist:
        user_obj=Users.objects.filter(employeeId=attendanceId,is_active=True).first()
        user_serializer=UsersSerializer(user_obj)
        leave_objs = Leave.objects.filter(employeeId=user_obj.id,WorkFromHome=False,leave_status="Approved",start_date__month__gte=month,start_date__month__lte=month,start_date__year= year,Active=True).exclude(leave_status='Draft')
        leavesserializer=leaveserializer(leave_objs,many=True)
        wfh_objs = Leave.objects.filter(employeeId=user_obj.id,WorkFromHome=True,leave_status="Approved",start_date__month__gte=month,start_date__month__lte=month,start_date__year= year,Active=True).exclude(leave_status='Draft')
        wfhsserializer=leaveserializer(wfh_objs,many=True)
    
        current_date_shift_list=[]
        new_current_date_shift_list=[]
        return_dict={}
        
        shiftdate = datetime.strptime(date, '%Y-%m-%d').date()
        tomarrow_date = shiftdate + timedelta(days=1)
        yesterday_date = shiftdate - timedelta(days=1)
        current_shiftallotment_objs=ShiftAllotment.objects.filter(attendanceId=attendanceId,date=str(date),is_active=True)
        current_shiftallotment_serializers=ShiftAllotmentshiftIdSerializer(current_shiftallotment_objs,many=True)
        current_shiftId_list=list(current_shiftallotment_serializers.data)
        current_shift_obj=ShiftMaster.objects.filter(id__in=current_shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime')
        current_shift_serializer=ShiftMasterSerializer(current_shift_obj,many=True)
        current_shiftlist=current_shift_serializer.data
        
        tomarrow_shiftallotment_objs=ShiftAllotment.objects.filter(attendanceId=attendanceId,date=str(tomarrow_date),is_active=True)
        tomarrow_shiftallotment_serializers=ShiftAllotmentshiftIdSerializer(tomarrow_shiftallotment_objs,many=True)
        tomarrow_shiftId_list=list(tomarrow_shiftallotment_serializers.data)
        tomarrow_shift_obj=ShiftMaster.objects.filter(id__in=tomarrow_shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime')
        tomarrow_shift_serializer=ShiftMasterSerializer(tomarrow_shift_obj,many=True)
        tomarrow_shiftlist=tomarrow_shift_serializer.data

        count=1
        
        for shift in current_shiftlist:
            start_time = datetime.strptime(str(shiftdate) +' '+ shift['intime'], '%Y-%m-%d %H:%M')
            start_time_before_2hrs = start_time - timedelta(hours=2)
            
            if shift['intime'] > shift['outtime']:
                shift_end_date = shiftdate + timedelta(days=1)
            else:
                shift_end_date = shiftdate
                
            end_time = datetime.strptime(str(shift_end_date) +' '+ shift['outtime'], '%Y-%m-%d %H:%M')
            check_login_till=''
            if len(current_shiftlist) >= count+1:
            
                check_next_shift_in = datetime.strptime(str(shiftdate) +' '+ current_shiftlist[count]['intime'], '%Y-%m-%d %H:%M')
                check_login_till = check_next_shift_in - timedelta(hours=2)

                
            elif check_login_till=='':
                    
                if len(tomarrow_shiftlist) >= 1:
                    
                    check_next_shift_in = datetime.strptime(str(tomarrow_date) +' '+ tomarrow_shiftlist[0]['intime'], '%Y-%m-%d %H:%M')
                    check_login_till = check_next_shift_in - timedelta(hours=2)
                else:
                
                    if_no_shift_check_shift=ShiftAllotment.objects.filter(attendanceId=attendanceId,date=str(tomarrow_date),is_active=True)
                    tomerrow_shiftlist=ShiftAllotmentshiftIdSerializer(if_no_shift_check_shift,many=True)
                    next_date_shiftId_list=list(tomerrow_shiftlist.data)
                    check_weekly_off=ShiftMaster.objects.filter(id__in=next_date_shiftId_list,intime='00:00',outtime='00:00',is_active=True).order_by('intime').first()
                    if check_weekly_off is not None:
                        
                        if shift['outtime'] < shift['intime']:
                    
                            check_current_shift_out = datetime.strptime(str(tomarrow_date) +' '+ shift['outtime'], '%Y-%m-%d %H:%M')
                            check_current_shift_out_time = check_current_shift_out + timedelta(hours=2)
                            given_datetime = datetime.strptime(str(check_current_shift_out_time), '%Y-%m-%d %H:%M:%S')

                            truncated_datetime_str = given_datetime.strftime('%Y-%m-%d %H:%M')

                            
                            shift_end_date_time=str(truncated_datetime_str)
                        else:
                            shift_end_date_time=str(date) +' '+ '23:59'
                            
                        check_login_till = datetime.strptime(shift_end_date_time, '%Y-%m-%d %H:%M')
                    else:
                    
                        check_login_till = datetime.strptime(str(tomarrow_date) +' '+ '07:30', '%Y-%m-%d %H:%M')
                        
            shift_data={
                'shiftname':shift['shiftname'],
                'indatetime':str(start_time_before_2hrs),
                'outdatetime':str(check_login_till),
                'intime':shift['intime'],
                'outtime':shift['outtime'],
            }
            current_date_shift_list.append(shift_data)
            new_current_date_shift_list.append(shift_data)
            count+=1

        if len(current_shiftlist) == 0:       
        
            if_no_shift_check_shift=ShiftAllotment.objects.filter(attendanceId=attendanceId,date=str(date),is_active=True)
            todays_shiftlist=ShiftAllotmentshiftIdSerializer(if_no_shift_check_shift,many=True)
            toadys_shiftId_list=list(todays_shiftlist.data)
            check_weekly_off=ShiftMaster.objects.filter(id__in=toadys_shiftId_list,intime='00:00',outtime='00:00',is_active=True).order_by('intime').first()
            if check_weekly_off is not None:
                get_previous_day_shift=ShiftAllotment.objects.filter(attendanceId=attendanceId,date=str(yesterday_date),is_active=True)
                previous_day_shiftlist=ShiftAllotmentshiftIdSerializer(get_previous_day_shift,many=True)
                previous_shiftId_list=list(previous_day_shiftlist.data)
                get_previous_last_shift=ShiftMaster.objects.filter(id__in=previous_shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime').last()
                if get_previous_last_shift is not None:
                    if get_previous_last_shift.outtime < get_previous_last_shift.intime:
                        shift_end_date_time=str(date) +' '+ str(get_previous_last_shift.outtime)
                    else:
                        shift_end_date_time=str(yesterday_date) +' '+ '23:59'
                            
                    check_previous_shift_out = datetime.strptime(str(shift_end_date_time), '%Y-%m-%d %H:%M')
                    previous_shift_in_time = check_previous_shift_out + timedelta(hours=2)  
                else:
                    previous_shift_in_time= str(date)+' 00:00:00'
                get_net_day_shift=ShiftAllotment.objects.filter(attendanceId=attendanceId,date=str(tomarrow_date),is_active=True)
                nest_day_shiftlist=ShiftAllotmentshiftIdSerializer(get_net_day_shift,many=True)
                next_date_shiftId_list=list(nest_day_shiftlist.data)
                get_next_day_first_shift=ShiftMaster.objects.filter(id__in=next_date_shiftId_list,is_active=True).exclude(intime='00:00',outtime='00:00').order_by('intime').first()
                if get_next_day_first_shift is not None:
                    check_next_shift_in = datetime.strptime(str(tomarrow_date) +' '+ str(get_next_day_first_shift.intime), '%Y-%m-%d %H:%M')
                    next_shift_in_time = check_next_shift_in - timedelta(hours=2)
                    shift_data={
                        'shiftname':str(check_weekly_off.shiftname),
                        'indatetime':str(previous_shift_in_time),
                        'outdatetime':str(next_shift_in_time),
                        'intime':check_weekly_off.intime,
                        'outtime':check_weekly_off.outtime,
                    }
                    current_date_shift_list.append(shift_data)
                    new_current_date_shift_list.append(shift_data)
                    
                else:
                    shift_data={
                        'shiftname':str(check_weekly_off.shiftname),
                        'indatetime':str(previous_shift_in_time),
                        'outdatetime':str(date)+' 23:59:59',
                        'intime':check_weekly_off.intime,
                        'outtime':check_weekly_off.outtime,
                    }
                    current_date_shift_list.append(shift_data)
                    new_current_date_shift_list.append(shift_data)

            else:
                shift_data={
                    'shiftname':'General',
                    'indatetime':str(date)+' 07:30:00',
                    'outdatetime':str(tomarrow_date)+' 07:30:00',
                    'intime':'09:30',
                    'outtime':'18:30',
                }
                current_date_shift_list.append(shift_data)
                new_current_date_shift_list.append(shift_data)

                
        current_date_first_in_datetime=''
        current_date_last_out_datetime=''
        extra_days=''
        total_working_hrs=''
        for s in new_current_date_shift_list:
            intime=''
            intimedate=''
            
            
            getallattendance = attendance.objects.filter(Q(employeeId=str(attendanceId),time__gte=s['indatetime'].split(' ')[1],date=s['indatetime'].split(' ')[0])|Q(employeeId=str(attendanceId),time__lte=str(s['outdatetime']).split(' ')[1],date=str(s['outdatetime']).split(' ')[0])).order_by('date','time')
            
            # getallattendance = getallattendance.filter(time__gte=currentshiftstarttime).order_by('date','time')
            
            attendance_serializer=attendanceserializer(getallattendance,many=True)
        
            
            sorted_data = sorted(attendance_serializer.data, key=lambda x: (x['date'], x['time']))
            
            mindatetime = datetime.strptime(s['indatetime'], '%Y-%m-%d %H:%M:%S')
            maxdatetime = datetime.strptime(s['outdatetime'], '%Y-%m-%d %H:%M:%S')
        
            sorted_data = [entry for entry in sorted_data if (mindatetime <= datetime.strptime(entry['date'] +' '+entry['time'],'%Y-%m-%d %H:%M:%S') <= maxdatetime)]
            if len(sorted_data) > 0:
                intimedate=sorted_data[0]['date']
                intime=str(sorted_data[0]['time'])
                
            if intimedate !='' and intimedate is not None:
                user_sdt = datetime.strptime(str(intimedate) + ' ' + str(intime), '%Y-%m-%d %H:%M:%S')
                shif_sdt = datetime.strptime(s['indatetime'].split(' ')[0] + ' ' + s['indatetime'].split(' ')[1], '%Y-%m-%d %H:%M:%S')
                if user_sdt < shif_sdt :
                    intimedate=''
                    intime=''

                    
            checkin_time = None
            total_working_time = 0
            attendance_log=[]
            attendance_history=[]

            for index, entry in enumerate(sorted_data):

                if entry['checkout']:
                    if entry['deviceId'] is not None and entry['deviceId'] !='':
                        attendance_type="Machine Checkout"
                        attendance_type_resaon=""
                    elif entry['Remote_Reason'] is not None and entry['Remote_Reason'] !='':
                        attendance_type="Remote Checkout"
                        attendance_type_resaon=entry['Remote_Reason']
                    else:
                        
                        
                        attendance_type="Online Checkout"
                        attendance_type_resaon=''                                

                    attendance_history.append({'Status':'Check-Out','datetime':entry['date']+' '+entry['time'],'attendance_type':attendance_type,'attendance_type_resaon':attendance_type_resaon})
                        
                    attendance_log.append({'checkout':entry['date']+' '+entry['time']})
                    if checkin_time:
                        checkout_datetime = datetime.strptime(entry['date'] + ' ' + entry['time'], '%Y-%m-%d %H:%M:%S')
                        checkin_datetime = datetime.strptime(checkin_time, '%Y-%m-%d %H:%M:%S')
                        working_time = checkout_datetime - checkin_datetime
                        total_working_time += working_time.total_seconds()
                        checkin_time = None
                        
                elif not entry['checkout']:
                    if entry['deviceId'] is not None and entry['deviceId'] !='':
                        attendance_type="Machine Checkin"
                        attendance_type_resaon=''
                    elif entry['Remote_Reason'] is not None and entry['Remote_Reason'] !='':
                        attendance_type="Remote Checkin"
                        attendance_type_resaon=entry['Remote_Reason']
                    else:
                        attendance_type="Online Checkin"
                        attendance_type_resaon=''  

                    checkin_time = entry['date'] + ' ' + entry['time']
                    attendance_log.append({'checkin':entry['date']+' '+entry['time']})
                    attendance_history.append({'Status':'Check-In','datetime':entry['date']+' '+entry['time'],'attendance_type':attendance_type,'attendance_type_resaon':attendance_type_resaon})
            # If the last entry is check-in, consider checkout time as current shift end  time
            if checkin_time and index == len(sorted_data) - 1:
                
                # check if the previous entry is also checkin or not if exist get the  difference betwnn  the current checkin  and  previous  checkin

                if int(int(index)-1) >=0:
                    if sorted_data[index-1]['checkout']==False:
                        previous_checkin_date_time=sorted_data[index-1]['date']+ ' '+sorted_data[index-1]['time']
                        checkout_datetime = datetime.strptime(previous_checkin_date_time, '%Y-%m-%d %H:%M:%S')
                        checkin_datetime = datetime.strptime(checkin_time, '%Y-%m-%d %H:%M:%S')
                        working_time = checkin_datetime-checkout_datetime 
                        total_working_time += working_time.total_seconds()


            # Convert total_working_time to hours, minutes, and seconds
            hours, remainder = divmod(total_working_time, 3600)
            minutes, seconds = divmod(remainder, 60)
            s['total_hrs']=str(add_leading_zero(int(hours))) +':' +str(add_leading_zero(int(minutes))) +':'+str(add_leading_zero(int(seconds)))

            if total_working_hrs == '':

                total_working_hrs=s['total_hrs']
            else:
                time_str_1 = total_working_hrs
                time_str_2 = s['total_hrs']



                # Convert time strings to timedelta objects and add them
                total_time_delta = timedelta(hours=int(time_str_1[:2]), minutes=int(time_str_1[3:5]), seconds=int(time_str_1[6:])) + \
                                timedelta(hours=int(time_str_2[:2]), minutes=int(time_str_2[3:5]), seconds=int(time_str_2[6:]))
                total_working_hrs=str(total_time_delta)
                
            s['attendance_log']=attendance_log
            s['attendance_history']=attendance_history
            if len(attendance_log) !=0:
                
                infirst_key = next(iter(attendance_log[0]))
                outfirst_key = next(iter(attendance_log[-1]))
                
                s['usersintime']=attendance_log[0][infirst_key]
                s['usersouttime']=attendance_log[-1][outfirst_key]
                if s['usersouttime'] == s['usersintime']:
                    s['usersouttime']= ''
                    
            else:
                s['usersintime']=''
                s['usersouttime']=''


            
            if current_date_first_in_datetime =='':
                if s['usersintime'] !='':
                    current_date_first_in_datetime=s['usersintime']
            elif s['usersintime'] !='':
                if datetime.strptime(current_date_first_in_datetime, '%Y-%m-%d %H:%M:%S') > datetime.strptime(s['usersintime'], '%Y-%m-%d %H:%M:%S'):
                    current_date_first_in_datetime=s['usersintime']

            if current_date_last_out_datetime =='':
                if s['usersouttime'] !='':
                    current_date_last_out_datetime=s['usersouttime']
            elif s['usersouttime'] !='':
                if datetime.strptime(current_date_last_out_datetime, '%Y-%m-%d %H:%M:%S') < datetime.strptime(s['usersouttime'], '%Y-%m-%d %H:%M:%S'):
                    current_date_last_out_datetime=s['usersouttime']

                
            if current_date_last_out_datetime != '' and current_date_first_in_datetime !='':
                if current_date_last_out_datetime.split(' ')[0] != current_date_first_in_datetime.split(' ')[0]:
                    current_date_last_out_datetime1 = datetime.strptime(str(current_date_last_out_datetime).split(' ')[0], '%Y-%m-%d')
                    current_date_first_in_datetime1 = datetime.strptime(str(current_date_first_in_datetime).split(' ')[0], '%Y-%m-%d')
                    extradaysdiff = current_date_last_out_datetime1 - current_date_first_in_datetime1
                    if extradaysdiff.days > 0:
                        extra_days='+'+str(extradaysdiff.days)
                
        return_dict['employeeId']=attendanceId
        return_dict['empname']=user_obj.Firstname +' '+ user_obj.Lastname
        return_dict['username']=user_obj.Firstname +' '+ user_obj.Lastname
        
        return_dict['reason']='Attendance not found'
        return_dict['fulldate']=date
        return_dict['shifts_list']=current_date_shift_list
        return_dict['shift']=new_current_date_shift_list[0]

        return_dict['Leavetype']=''


        
        if len(new_current_date_shift_list) > 1:
            shift_name=new_current_date_shift_list[0]['shiftname']
            return_dict['shift']['shiftname'] = str(shift_name) +  ' <span class="text-danger"> +1 </span>'
            new_current_date_shift_list[0]['swap']=''
        else:
            new_current_date_shift_list[0]['swap']=''
            
        if current_date_first_in_datetime !='':
            return_dict['checkintime']=str(current_date_first_in_datetime).split(' ')[1]
            checkin_attendance_type=attendance.objects.filter(time=str(current_date_first_in_datetime).split(' ')[1],date=str(current_date_first_in_datetime).split(' ')[0]).first()

            

                        
                        
            if checkin_attendance_type is not None:
                remote_map_location=''
                remote_map_name=''
                if checkin_attendance_type.remote_latitude is not None and checkin_attendance_type.remote_longitude is not None and checkin_attendance_type.remote_latitude !='' and checkin_attendance_type.remote_longitude !='':
                    remote_map_location=create_google_maps_url(checkin_attendance_type.remote_latitude,checkin_attendance_type.remote_longitude)
                    remote_map_name=get_location_name(checkin_attendance_type.remote_latitude,checkin_attendance_type.remote_longitude)
                
                if checkin_attendance_type.deviceId is not None and checkin_attendance_type.deviceId !='':
                    return_dict['check_in_attendance_type']='Machine CheckIn'
                    return_dict['check_in_attendance_type_resaon']=''  

                    if checkin_attendance_type.deviceId == 20:
                        remote_map_name='Zentro Pune Office'
                        remote_map_location=create_google_maps_url(18.566064883698555, 73.77574703366226)
                    elif checkin_attendance_type.deviceId == 19:
                        remote_map_name='Zentro Mumbai Office'
                        remote_map_location=create_google_maps_url(19.1764898841813, 72.95988765068624)

                    return_dict['check_in_remote_map_name']=remote_map_name
                    return_dict['check_in_remote_map_location']=remote_map_location
                    
                elif checkin_attendance_type.Remote_Reason is not None and checkin_attendance_type.Remote_Reason !='':
                    return_dict['check_in_attendance_type']="Remote CheckIn"
                    return_dict['check_in_attendance_type_resaon']=checkin_attendance_type.Remote_Reason
                   

                    return_dict['check_in_remote_map_name']=remote_map_name
                    return_dict['check_in_remote_map_location']=remote_map_location
                    
                else:
                    
                    
                    return_dict['check_in_attendance_type']="Online CheckIn"
                    return_dict['check_in_attendance_type_resaon']=''  
                    
                    if checkin_attendance_type.attendance_type !='' and  checkin_attendance_type.attendance_type is not None:
                        return_dict['check_in_attendance_type'] = checkin_attendance_type.attendance_type+' ' + 'CheckIn'
                        
                    return_dict['check_in_remote_map_name']=remote_map_name
                    return_dict['check_in_remote_map_location']=remote_map_location
                    
            return_dict['inTime_date']=convertdate2(str(str(current_date_first_in_datetime).split(' ')[0]))
        else:
            return_dict['checkintime']='--:--'
            return_dict['inTime_date']=''
            return_dict['check_in_attendance_type']=""
            return_dict['check_in_attendance_type_resaon']='' 
            
            
        if current_date_last_out_datetime !='':
            return_dict['checkouttime']=str(current_date_last_out_datetime).split(' ')[1]
            return_dict['outTime_date']=convertdate2(str(str(current_date_last_out_datetime).split(' ')[0]))

            checkout_attendance_type=attendance.objects.filter(time=str(current_date_last_out_datetime).split(' ')[1],date=str(current_date_last_out_datetime).split(' ')[0]).first()
            if checkout_attendance_type is not None:
                remote_map_location=''
                remote_map_name=''
                if checkout_attendance_type.remote_latitude is not None and checkout_attendance_type.remote_longitude is not None and checkout_attendance_type.remote_latitude !='' and checkout_attendance_type.remote_longitude !='':
                    remote_map_location=create_google_maps_url(checkout_attendance_type.remote_latitude,checkout_attendance_type.remote_longitude)
                    remote_map_name=get_location_name(checkout_attendance_type.remote_latitude,checkout_attendance_type.remote_longitude)
                
                
                
                if checkout_attendance_type.deviceId is not None and checkout_attendance_type.deviceId !='':
                    return_dict['check_out_attendance_type']='Machine CheckOut'
                    return_dict['check_out_attendance_type_resaon']=''
                      
                    if checkout_attendance_type.deviceId == 20:
                        remote_map_name='Zentro Pune Office'                        
                        remote_map_location=create_google_maps_url(18.566064883698555, 73.77574703366226)

                    elif checkout_attendance_type.deviceId == 19:
                        remote_map_name='Zentro Mumbai Office'
                        remote_map_location=create_google_maps_url(19.1764898841813, 72.95988765068624)

                    return_dict['checkout_remote_map_name']=remote_map_name
                    return_dict['checkout_remote_map_location']=remote_map_location
                    
                    
                elif checkout_attendance_type.Remote_Reason is not None and checkout_attendance_type.Remote_Reason !='':
                    return_dict['check_out_attendance_type']="Remote CheckOut"
                    return_dict['check_out_attendance_type_resaon']=checkout_attendance_type.Remote_Reason

                    return_dict['checkout_remote_map_name']=remote_map_name
                    return_dict['checkout_remote_map_location']=remote_map_location
                else:
                    return_dict['check_out_attendance_type']="Online CheckOut"
                    return_dict['check_out_attendance_type_resaon']=''  
                    if checkout_attendance_type.attendance_type !='' and  checkout_attendance_type.attendance_type is not None:
                        return_dict['check_out_attendance_type'] = checkout_attendance_type.attendance_type +' ' + 'CheckOut'
                    return_dict['checkout_remote_map_name']=remote_map_name
                    return_dict['checkout_remote_map_location']=remote_map_location
            
        else:
            return_dict['checkouttime']='--:--'
            return_dict['outTime_date']=''
            return_dict['check_out_attendance_type']=""
            return_dict['check_out_attendance_type_resaon']='' 
            
            
        return_dict['extra_days']=extra_days
        return_dict['total_time']=total_working_hrs
        if return_dict['checkintime'] !='' and return_dict['checkintime'] !='--:--':
            return_dict['att_status']="P"
            
            check_indatetime = datetime.strptime(return_dict['shift']['indatetime'], "%Y-%m-%d %H:%M:%S")
            check_usersintime = datetime.strptime(return_dict['shift']['usersintime'], "%Y-%m-%d %H:%M:%S")
            grace_period = timedelta(hours=float(2.5))
            allowed_time = check_indatetime + grace_period
            if check_usersintime >= allowed_time:
                return_dict['att_status']="late"
            
            
        else:
            return_dict['att_status']="A"
        return_dict['Designation']=user_serializer.data['DesignationId']
        return_dict['Task']=''

        foundleaves = [
            leave for leave in leavesserializer.data if datetime.strptime(leave['start_date'], '%Y-%m-%d').date() <= datetime.strptime(date, '%Y-%m-%d').date() <= datetime.strptime(leave['end_date'], '%Y-%m-%d').date()
        ]
        
        foundwfhs = [
            wfh for wfh in wfhsserializer.data if datetime.strptime(wfh['start_date'], '%Y-%m-%d').date() <= datetime.strptime(date, '%Y-%m-%d').date() <= datetime.strptime(wfh['end_date'], '%Y-%m-%d').date()
        ]
        
        if foundwfhs:
            return_dict['att_status']="WFH"
        if foundleaves:
            return_dict['att_status']="L"
            
            
        if return_dict['att_status']=='late':
            return_dict['status']="<span class='font-weight-500 text-warning'>Late Mark</span>"
        elif return_dict['att_status']=='P':
            return_dict['status']="<span class='font-weight-500 text-primary'>On Time</span>"
        elif return_dict['att_status']=='L':
            return_dict['status']="<span class='font-weight-500 text-secondary'>On Leave</span>"
        elif return_dict['att_status']=='WFH':
            return_dict['status']="<span class='font-weight-500 text-info'>WFH</span>"
        else:
            
            return_dict['status']="<span class='font-weight-500 text-danger'>Not Found</span>"
            
            
        return_dict['checkintime']=convert_to_12_hour_format(return_dict['checkintime']) 
        return_dict['checkouttime']=convert_to_12_hour_format(return_dict['checkouttime']) 
        commonlist.append(return_dict)            
            
            
    context= {
        "commonlist": commonlist,
        }
    
    return Response ({"context":context, "response":{"n" : 1,"msg" : "SUCCESS","status" : "success"}})

@api_view(['POST'])
@permission_classes((AllowAny,))
def send_mail(request):
    send_async_custom_template_email(
                    'First TrackProning: Performance and Productivity Concerns',
                    {"empname":"Mahesh"},
                    EMAIL_HOST_USER,
                    ['mahesh.kattale@onerooftech.com'],
                    "mails/senddemotemplate.html"
                )
    return Response("send")

@api_view(['POST'])
def granted_compoff_list(request):
    if request.method == 'POST':
        All_granted_compoff_obj=CompOffGrantedMaster.objects.filter(is_active=True).order_by('-created_at')
        Approved_granted_compoff_obj=CompOffGrantedMaster.objects.filter(status="Approved",is_active=True).order_by('-created_at')
        Rejected_granted_compoff_obj=CompOffGrantedMaster.objects.filter(status="Rejected",is_active=True).order_by('-created_at')
        Expired_granted_compoff_obj=CompOffGrantedMaster.objects.filter(status="Expired",is_active=True).order_by('-created_at')
        
        All_granted_compoff_serializer=CompOffGrantedMasterSerializer(All_granted_compoff_obj,many=True)
        Approved_granted_compoff_serializer=CompOffGrantedMasterSerializer(Approved_granted_compoff_obj,many=True)
        Rejected_granted_compoff_serializer=CompOffGrantedMasterSerializer(Rejected_granted_compoff_obj,many=True)
        Expired_granted_compoff_serializer=CompOffGrantedMasterSerializer(Expired_granted_compoff_obj,many=True)
        
        
        context={
                    'AllCompoff':All_granted_compoff_serializer.data,
                    'ApprovedCompoff':Approved_granted_compoff_serializer.data,
                    'RejectedCompoff':Rejected_granted_compoff_serializer.data,
                    'ExpiredCompoff':Expired_granted_compoff_serializer.data
                }
        
        return Response ({"context":context, "response":{"n" : 1,"msg" : "SUCCESS","status" : "success"}})
 
    
class pagination_get_late_mark_attendance(GenericAPIView):
    pagination_class = CustomPagination
    def post(self,request):
        
        com_code = request.user.company_code
        start_date=request.POST.get("start_date")
        end_date=request.POST.get("end_date")
        empId=request.POST.get("empId")
        
        late_obj = attendance.objects.filter(date__gte=start_date,date__lte=end_date,time__gte=time(10, 0, 0),time__lte=time(11, 0, 0),company_code=com_code).distinct('date','employeeId')
        if empId is not None and empId !='':
            late_obj=late_obj.filter(employeeId=empId)
            
        if late_obj.exists():
            page = self.paginate_queryset(late_obj)               
            serializer=attendance_user_serializer(page,many=True)
            return self.get_paginated_response(serializer.data)  
        else:
            return Response({"n": 0, "status": "failed", "count":0,"next":"","previous":"","msg": "No late marks found","data":[]})
            

@api_view(['POST'])
def latemarkexcelreport(request):
    print("abcd latemarkexcelreport")
    com_code = request.user.company_code
    start_date=request.POST.get("start_date")
    end_date=request.POST.get("end_date")
    empId=request.POST.get("empId")
    late_obj = attendance.objects.filter(date__gte=start_date,date__lte=end_date,time__gte=time(10, 0, 0),time__lte=time(11, 0, 0),company_code=com_code).distinct('date','employeeId')
    if empId is not None and empId !='':
        late_obj=late_obj.filter(employeeId=empId)

    serializer=attendance_user_serializer(late_obj,many=True)

    if os.path.exists("static/report/lateattendancereport.xlsx"):
        os.remove("static/report/lateattendancereport.xlsx")
        workbook = xlsxwriter.Workbook('static/report/lateattendancereport.xlsx')
        workbook.close()
    else:
        workbook = xlsxwriter.Workbook('static/report/lateattendancereport.xlsx')
        workbook.close()
    
    context={
        'emplist':serializer.data,
    }

    print("context",context)

    exportlatemarkreportdata(context)
    downloadurl = imageUrl + "/static/report/lateattendancereport.xlsx"

    return Response({'n':1,"status":"success","message":"Data found successdully","data":context,'downloadurl':downloadurl})
       

def exportlatemarkreportdata(data):

    workbook = xlsxwriter.Workbook('static/report/lateattendancereport.xlsx')
    wb=load_workbook('static/report/lateattendancereport.xlsx')
    sheet=wb.worksheets[0]
    
 
    sheet.cell(row=3,column=2).value="Date"
    sheet.cell(row=3,column=3).value="Name"
    sheet.cell(row=3,column=4).value='Time'
  

    k = 5
    counter = 1
    for p in data['emplist']:
        sheet.cell(row=k,column=2).value=p['date']
        sheet.cell(row=k,column=3).value=p['user_name']
        sheet.cell(row=k,column=4).value=p['time']
       
        k+=1
    s = k+1
    j = s+1
    
    
    counter+=1
    
    wb.save('static/report/lateattendancereport.xlsx')



            
    


@api_view(['POST'])
def get_manager_pending_applications_requests(request):
    current_date = datetime.today().date()
    user_id=request.user.id 
    applications_approvals = leaveApproval.objects.filter(managerId=user_id,approvedBy=False,rejectedBy=False).order_by("leave_id").distinct("leave_id")
    applications_ids_serializer = leave_ids_serializer(applications_approvals,many=True)
    applications_obj=Leave.objects.filter(id__in=list(applications_ids_serializer.data),Active=True,leave_status='Pending',start_date__gte=str(current_date)).order_by('-start_date')
    department_id=request.POST.get('department_value')

    name=request.POST.get('name')
    if name is not None and name !='':
        users_objs=Users.objects.filter(Q(Firstname__icontains = name,is_active=True)|Q(Lastname__icontains = name,is_active=True))
    else:
        users_objs=Users.objects.filter(is_active=True)

    if department_id is not None and department_id !='':
        users_objs=users_objs.filter(DepartmentID__in=[int(department_id)])
        
    users_ids=UsersSerializeronlyid(users_objs,many=True)

    applications_obj=applications_obj.filter(employeeId__in=list(users_ids.data))


    ApplicationType=request.POST.get('ApplicationType')
    if ApplicationType is not None and ApplicationType  != "":
        if ApplicationType =='Leave':
            applications_obj=applications_obj.filter(WorkFromHome=False)
        else:
            applications_obj=applications_obj.filter(WorkFromHome=True)

    month=request.POST.get('month')
    if month is not None and month  != "":
        applications_obj=applications_obj.filter(start_date__month=month)

    year=request.POST.get('year')
    if year is not None and year  != "":
        applications_obj=applications_obj.filter(start_date__year=year)




    if applications_obj.exists:
        applications_serailzer=applications_serializers(applications_obj,many=True)
        for application in applications_serailzer.data:
            approval_id=applications_approvals.filter(leave_id=application['id']).first()
            if approval_id is not None:
                application['approval_id']=approval_id.id
                if approval_id.approvedBy==False and approval_id.rejectedBy==False:
                    application['manager_action']=None
                    application['reverse_action']=None
                elif approval_id.approvedBy==True and approval_id.rejectedBy==False:
                    application['manager_action']=True
                    application['reverse_action']=True
                elif approval_id.approvedBy==False and approval_id.rejectedBy==True:
                    application['manager_action']=False
                    application['reverse_action']=False
                elif approval_id.approvedBy==True and approval_id.rejectedBy==True:
                    application['manager_action']=None
                    application['reverse_action']=None
                else:
                    application['manager_action']=None
                    application['reverse_action']=None
            else:
                application['approval_id']=''
                application['manager_action']=None
                application['reverse_action']=None


            date_obj = datetime.strptime(application['start_date'], "%d %b %Y").date()
            if date_obj < current_date:
                application['reverse_action']=None

            if application['attachment'] is not None and application['attachment'] !='':
                application['attachment']=imageUrl+application['attachment']
            else:
                application['attachment']=''
            if application['WorkFromHome']==True:
                application['ApplicationType']='Work From Home'
            else:
                application['ApplicationType']='Leave'
            if application['number_of_days'] is not None and application['number_of_days'] !='':
                application['total_days']=application['number_of_days']
            else:
                application['total_days']='NA'
                
            managers_obj = leaveApproval.objects.filter(leave_id=application['id']).distinct("managerId")
            managers_serializer = application_approval_serializers(managers_obj,many=True)
            application['managers']=managers_serializer.data

            for manager in application['managers']:
                if manager['approvedBy']==False and manager['rejectedBy']==False:
                    manager['status']=None
                elif manager['approvedBy']==True and manager['rejectedBy']==False:
                    manager['status']=True
                elif manager['approvedBy']==False and manager['rejectedBy']==True:
                    manager['status']=False
                elif manager['approvedBy']==True and manager['rejectedBy']==True:
                    manager['status']=None
                else:
                    manager['status']=None
        return Response ({"data":applications_serailzer.data, "response":{"n" : 1,"msg" : " Applications Found Successfully","status" : "success"}})
    else:
        return Response ({"data":[], "response":{"n" : 0,"msg" : "No Applications Found","status" : "success"}})

@api_view(['POST'])
def get_manager_approved_applications_requests(request):
    user_id=request.user.id 
    applications_approvals = leaveApproval.objects.filter(managerId=user_id,approvedBy=True,rejectedBy=False).order_by("leave_id").distinct("leave_id")
    applications_ids_serializer = leave_ids_serializer(applications_approvals,many=True)

    applications_obj=Leave.objects.filter(id__in=list(applications_ids_serializer.data),Active=True,).order_by('-start_date')
    applications_obj=applications_obj.filter(Q(leave_status='Approved')|Q(leave_status='Pending'))


    department_id=request.POST.get('department_value')

    name=request.POST.get('name')
    if name is not None and name !='':
        users_objs=Users.objects.filter(Q(Firstname__icontains = name,is_active=True)|Q(Lastname__icontains = name,is_active=True))
    else:
        users_objs=Users.objects.filter(is_active=True)

    if department_id is not None and department_id !='':
        users_objs=users_objs.filter(DepartmentID__in=[int(department_id)])
        
    users_ids=UsersSerializeronlyid(users_objs,many=True)

    applications_obj=applications_obj.filter(employeeId__in=list(users_ids.data))


    ApplicationType=request.POST.get('ApplicationType')
    if ApplicationType is not None and ApplicationType  != "":
        if ApplicationType =='Leave':
            applications_obj=applications_obj.filter(WorkFromHome=False)
        else:
            applications_obj=applications_obj.filter(WorkFromHome=True)

    month=request.POST.get('month')
    if month is not None and month  != "":
        applications_obj=applications_obj.filter(start_date__month=month)

    year=request.POST.get('year')
    if year is not None and year  != "":
        applications_obj=applications_obj.filter(start_date__year=year)



    if applications_obj.exists:
        applications_serailzer=applications_serializers(applications_obj,many=True)
        for application in applications_serailzer.data:
            approval_id=applications_approvals.filter(leave_id=application['id']).first()
            if approval_id is not None:
                application['approval_id']=approval_id.id
                if approval_id.approvedBy==False and approval_id.rejectedBy==False:
                    application['manager_action']=None
                    application['reverse_action']=None
                elif approval_id.approvedBy==True and approval_id.rejectedBy==False:
                    application['manager_action']=True
                    application['reverse_action']=True
                elif approval_id.approvedBy==False and approval_id.rejectedBy==True:
                    application['manager_action']=False
                    application['reverse_action']=False
                elif approval_id.approvedBy==True and approval_id.rejectedBy==True:
                    application['manager_action']=None
                    application['reverse_action']=None
                else:
                    application['manager_action']=None
                    application['reverse_action']=None
            else:
                application['approval_id']=''
                application['manager_action']=None
                application['reverse_action']=None


            date_obj = datetime.strptime(application['start_date'], "%d %b %Y").date()
            current_date = datetime.today().date()
            if date_obj < current_date:
                application['reverse_action']=None



            if application['attachment'] is not None and application['attachment'] !='':
                application['attachment']=imageUrl+application['attachment']
            else:
                application['attachment']=''
            if application['WorkFromHome']==True:
                application['ApplicationType']='Work From Home'
            else:
                application['ApplicationType']='Leave'
            if application['number_of_days'] is not None and application['number_of_days'] !='':

                application['total_days']=application['number_of_days']
            else:
                application['total_days']='NA'
                
            managers_obj = leaveApproval.objects.filter(leave_id=application['id']).distinct("managerId")
            managers_serializer = application_approval_serializers(managers_obj,many=True)
            application['managers']=managers_serializer.data

            for manager in application['managers']:
                if manager['approvedBy']==False and manager['rejectedBy']==False:
                    manager['status']=None
                elif manager['approvedBy']==True and manager['rejectedBy']==False:
                    manager['status']=True
                elif manager['approvedBy']==False and manager['rejectedBy']==True:
                    manager['status']=False
                elif manager['approvedBy']==True and manager['rejectedBy']==True:
                    manager['status']=None
                else:
                    manager['status']=None
        return Response ({"data":applications_serailzer.data, "response":{"n" : 1,"msg" : " Applications Found Successfully","status" : "success"}})
    else:
        return Response ({"data":[], "response":{"n" : 0,"msg" : "No Applications Found","status" : "success"}})

@api_view(['POST'])
def get_manager_rejected_applications_requests(request):
    user_id=request.user.id 
    applications_approvals = leaveApproval.objects.filter(managerId=user_id).order_by("leave_id").distinct("leave_id")
    applications_ids_serializer = leave_ids_serializer(applications_approvals,many=True)

    applications_obj=Leave.objects.filter(id__in=list(applications_ids_serializer.data),leave_status='Rejected',Active=True,).order_by('-start_date')

    department_id=request.POST.get('department_value')

    name=request.POST.get('name')
    if name is not None and name !='':
        users_objs=Users.objects.filter(Q(Firstname__icontains = name,is_active=True)|Q(Lastname__icontains = name,is_active=True))
    else:
        users_objs=Users.objects.filter(is_active=True)

    if department_id is not None and department_id !='':
        users_objs=users_objs.filter(DepartmentID__in=[int(department_id)])
        
    users_ids=UsersSerializeronlyid(users_objs,many=True)

    applications_obj=applications_obj.filter(employeeId__in=list(users_ids.data))


    ApplicationType=request.POST.get('ApplicationType')
    if ApplicationType is not None and ApplicationType  != "":
        if ApplicationType =='Leave':
            applications_obj=applications_obj.filter(WorkFromHome=False)
        else:
            applications_obj=applications_obj.filter(WorkFromHome=True)

    month=request.POST.get('month')
    if month is not None and month  != "":
        applications_obj=applications_obj.filter(start_date__month=month)

    year=request.POST.get('year')
    if year is not None and year  != "":
        applications_obj=applications_obj.filter(start_date__year=year)


    if applications_obj.exists:
        applications_serailzer=applications_serializers(applications_obj,many=True)
        for application in applications_serailzer.data:
            approval_id=applications_approvals.filter(leave_id=application['id']).first()
            if approval_id is not None:
                application['approval_id']=approval_id.id
                if approval_id.approvedBy==False and approval_id.rejectedBy==False:
                    application['manager_action']=None
                    application['reverse_action']=None
                elif approval_id.approvedBy==True and approval_id.rejectedBy==False:
                    application['manager_action']=True
                    application['reverse_action']=True
                elif approval_id.approvedBy==False and approval_id.rejectedBy==True:
                    application['manager_action']=False
                    application['reverse_action']=False
                elif approval_id.approvedBy==True and approval_id.rejectedBy==True:
                    application['manager_action']=None
                    application['reverse_action']=None
                else:
                    application['manager_action']=None
                    application['reverse_action']=None
            else:
                application['approval_id']=''
                application['manager_action']=None
                application['reverse_action']=None


            date_obj = datetime.strptime(application['start_date'], "%d %b %Y").date()
            current_date = datetime.today().date()
            if date_obj < current_date:
                application['reverse_action']=None

            if application['attachment'] is not None and application['attachment'] !='':
                application['attachment']=imageUrl+application['attachment']
            else:
                application['attachment']=''
            if application['WorkFromHome']==True:
                application['ApplicationType']='Work From Home'
            else:
                application['ApplicationType']='Leave'
            if application['number_of_days'] is not None and application['number_of_days'] !='':

                application['total_days']=application['number_of_days']
            else:
                application['total_days']='NA'

            managers_obj = leaveApproval.objects.filter(leave_id=application['id']).distinct("managerId")
            managers_serializer = application_approval_serializers(managers_obj,many=True)
            application['managers']=managers_serializer.data

            for manager in application['managers']:
                if manager['approvedBy']==False and manager['rejectedBy']==False:
                    manager['status']=None
                elif manager['approvedBy']==True and manager['rejectedBy']==False:
                    manager['status']=True
                elif manager['approvedBy']==False and manager['rejectedBy']==True:
                    manager['status']=False
                elif manager['approvedBy']==True and manager['rejectedBy']==True:
                    manager['status']=None
                else:
                    manager['status']=None


            rejected_managers_obj = leaveApproval.objects.filter(leave_id=application['id'],rejectedBy=True).first()
            if rejected_managers_obj is not None:
                application['reason']=rejected_managers_obj.comment
            else:
                application['reason']='Not Available'


        return Response ({"data":applications_serailzer.data, "response":{"n" : 1,"msg" : " Applications Found Successfully","status" : "success"}})
    else:
        return Response ({"data":[], "response":{"n" : 0,"msg" : "No Applications Found","status" : "success"}})

@api_view(['POST'])
def get_manager_withdraw_applications_requests(request):
    user_id=request.user.id 
    applications_approvals = leaveApproval.objects.filter(managerId=user_id).order_by("leave_id").distinct("leave_id")
    applications_ids_serializer = leave_ids_serializer(applications_approvals,many=True)

    applications_obj=Leave.objects.filter(id__in=list(applications_ids_serializer.data),Active=True,leave_status='Withdraw').order_by('-start_date')

    department_id=request.POST.get('department_value')

    name=request.POST.get('name')
    if name is not None and name !='':
        users_objs=Users.objects.filter(Q(Firstname__icontains = name,is_active=True)|Q(Lastname__icontains = name,is_active=True))
    else:
        users_objs=Users.objects.filter(is_active=True)

    if department_id is not None and department_id !='':
        users_objs=users_objs.filter(DepartmentID__in=[int(department_id)])
        
    users_ids=UsersSerializeronlyid(users_objs,many=True)

    applications_obj=applications_obj.filter(employeeId__in=list(users_ids.data))


    ApplicationType=request.POST.get('ApplicationType')
    if ApplicationType is not None and ApplicationType  != "":
        if ApplicationType =='Leave':
            applications_obj=applications_obj.filter(WorkFromHome=False)
        else:
            applications_obj=applications_obj.filter(WorkFromHome=True)

    month=request.POST.get('month')
    if month is not None and month  != "":
        applications_obj=applications_obj.filter(start_date__month=month)

    year=request.POST.get('year')
    if year is not None and year  != "":
        applications_obj=applications_obj.filter(start_date__year=year)

    if applications_obj.exists:
        applications_serailzer=applications_serializers(applications_obj,many=True)
        for application in applications_serailzer.data:
            approval_id=applications_approvals.filter(leave_id=application['id']).first()
            if approval_id is not None:
                application['approval_id']=approval_id.id
                if approval_id.approvedBy==False and approval_id.rejectedBy==False:
                    application['manager_action']=None
                    application['reverse_action']=None
                elif approval_id.approvedBy==True and approval_id.rejectedBy==False:
                    application['manager_action']=True
                    application['reverse_action']=True
                elif approval_id.approvedBy==False and approval_id.rejectedBy==True:
                    application['manager_action']=False
                    application['reverse_action']=False
                elif approval_id.approvedBy==True and approval_id.rejectedBy==True:
                    application['manager_action']=None
                    application['reverse_action']=None
                else:
                    application['manager_action']=None
                    application['reverse_action']=None
            else:
                application['approval_id']=''
                application['manager_action']=None
                application['reverse_action']=None


            date_obj = datetime.strptime(application['start_date'], "%d %b %Y").date()
            current_date = datetime.today().date()
            if date_obj < current_date:
                application['reverse_action']=None

            if application['attachment'] is not None and application['attachment'] !='':
                application['attachment']=imageUrl+application['attachment']
            else:
                application['attachment']=''
            if application['WorkFromHome']==True:
                application['ApplicationType']='Work From Home'
            else:
                application['ApplicationType']='Leave'
            if application['number_of_days'] is not None and application['number_of_days'] !='':

                application['total_days']=application['number_of_days']
            else:
                application['total_days']='NA'
                
            managers_obj = leaveApproval.objects.filter(leave_id=application['id']).distinct("managerId")
            managers_serializer = application_approval_serializers(managers_obj,many=True)
            application['managers']=managers_serializer.data

            for manager in application['managers']:
                if manager['approvedBy']==False and manager['rejectedBy']==False:
                    manager['status']=None
                elif manager['approvedBy']==True and manager['rejectedBy']==False:
                    manager['status']=True
                elif manager['approvedBy']==False and manager['rejectedBy']==True:
                    manager['status']=False
                elif manager['approvedBy']==True and manager['rejectedBy']==True:
                    manager['status']=None
                else:
                    manager['status']=None
        return Response ({"data":applications_serailzer.data, "response":{"n" : 1,"msg" : " Applications Found Successfully","status" : "success"}})
    else:
        return Response ({"data":[], "response":{"n" : 0,"msg" : "No Applications Found","status" : "success"}})

@api_view(['POST'])
def get_manager_expired_applications_requests(request):
    user_id=request.user.id 
    applications_approvals = leaveApproval.objects.filter(managerId=user_id,approvedBy=False,rejectedBy=False).order_by("leave_id").distinct("leave_id")
    applications_ids_serializer = leave_ids_serializer(applications_approvals,many=True)
    current_date = datetime.now().strftime('%Y-%m-%d')

    applications_obj=Leave.objects.filter(id__in=list(applications_ids_serializer.data),Active=True,start_date__lt=current_date).exclude(Q(leave_status="Rejected")|Q(leave_status="Withdraw")|Q(leave_status="Draft")).order_by('-start_date')


    department_id=request.POST.get('department_value')

    name=request.POST.get('name')
    if name is not None and name !='':
        users_objs=Users.objects.filter(Q(Firstname__icontains = name,is_active=True)|Q(Lastname__icontains = name,is_active=True))
    else:
        users_objs=Users.objects.filter(is_active=True)

    if department_id is not None and department_id !='':
        users_objs=users_objs.filter(DepartmentID__in=[int(department_id)])
        
    users_ids=UsersSerializeronlyid(users_objs,many=True)

    applications_obj=applications_obj.filter(employeeId__in=list(users_ids.data))


    ApplicationType=request.POST.get('ApplicationType')
    if ApplicationType is not None and ApplicationType  != "":
        if ApplicationType =='Leave':
            applications_obj=applications_obj.filter(WorkFromHome=False)
        else:
            applications_obj=applications_obj.filter(WorkFromHome=True)

    month=request.POST.get('month')
    if month is not None and month  != "":
        applications_obj=applications_obj.filter(start_date__month=month)

    year=request.POST.get('year')
    if year is not None and year  != "":
        applications_obj=applications_obj.filter(start_date__year=year)




    if applications_obj.exists:
        applications_serailzer=applications_serializers(applications_obj,many=True)
        for application in applications_serailzer.data:
            approval_id=applications_approvals.filter(leave_id=application['id']).first()
            if approval_id is not None:
                application['approval_id']=approval_id.id
                if approval_id.approvedBy==False and approval_id.rejectedBy==False:
                    application['manager_action']=None
                    application['reverse_action']=None
                elif approval_id.approvedBy==True and approval_id.rejectedBy==False:
                    application['manager_action']=True
                    application['reverse_action']=True
                elif approval_id.approvedBy==False and approval_id.rejectedBy==True:
                    application['manager_action']=False
                    application['reverse_action']=False
                elif approval_id.approvedBy==True and approval_id.rejectedBy==True:
                    application['manager_action']=None
                    application['reverse_action']=None
                else:
                    application['manager_action']=None
                    application['reverse_action']=None
            else:
                application['approval_id']=''
                application['manager_action']=None
                application['reverse_action']=None


            date_obj = datetime.strptime(application['start_date'], "%d %b %Y").date()
            current_date = datetime.today().date()
            if date_obj < current_date:
                application['reverse_action']=None



            if application['attachment'] is not None and application['attachment'] !='':
                application['attachment']=imageUrl+application['attachment']
            else:
                application['attachment']=''
            if application['WorkFromHome']==True:
                application['ApplicationType']='Work From Home'
            else:
                application['ApplicationType']='Leave'
            if application['number_of_days'] is not None and application['number_of_days'] !='':

                application['total_days']=application['number_of_days']
            else:
                application['total_days']='NA'
            managers_obj = leaveApproval.objects.filter(leave_id=application['id']).distinct("managerId")
            managers_serializer = application_approval_serializers(managers_obj,many=True)
            application['managers']=managers_serializer.data

            for manager in application['managers']:
                if manager['approvedBy']==False and manager['rejectedBy']==False:
                    manager['status']=None
                elif manager['approvedBy']==True and manager['rejectedBy']==False:
                    manager['status']=True
                elif manager['approvedBy']==False and manager['rejectedBy']==True:
                    manager['status']=False
                elif manager['approvedBy']==True and manager['rejectedBy']==True:
                    manager['status']=None
                else:
                    manager['status']=None
        return Response ({"data":applications_serailzer.data, "response":{"n" : 1,"msg" : " Applications Found Successfully","status" : "success"}})
    else:
        return Response ({"data":[], "response":{"n" : 0,"msg" : "No Applications Found","status" : "success"}})

@api_view(['POST'])
def get_manager_all_applications_requests(request):
    user_id=request.user.id 
    applications_approvals = leaveApproval.objects.filter(managerId=user_id).order_by("leave_id").distinct("leave_id")
    applications_ids_serializer = leave_ids_serializer(applications_approvals,many=True)

    applications_obj=Leave.objects.filter(id__in=list(applications_ids_serializer.data),Active=True).exclude(leave_status='Draft').order_by('-start_date')
    



    department_id=request.POST.get('department_value')

    name=request.POST.get('name')
    if name is not None and name !='':
        users_objs=Users.objects.filter(Q(Firstname__icontains = name,is_active=True)|Q(Lastname__icontains = name,is_active=True))
    else:
        users_objs=Users.objects.filter(is_active=True)

    if department_id is not None and department_id !='':
        users_objs=users_objs.filter(DepartmentID__in=[int(department_id)])

    users_ids=UsersSerializeronlyid(users_objs,many=True)

    applications_obj=applications_obj.filter(employeeId__in=list(users_ids.data))


    ApplicationType=request.POST.get('ApplicationType')
    if ApplicationType is not None and ApplicationType  != "":
        if ApplicationType =='Leave':
            applications_obj=applications_obj.filter(WorkFromHome=False)
        else:
            applications_obj=applications_obj.filter(WorkFromHome=True)

    month=request.POST.get('month')
    if month is not None and month  != "":
        applications_obj=applications_obj.filter(start_date__month=month)

    year=request.POST.get('year')
    if year is not None and year  != "":
        applications_obj=applications_obj.filter(start_date__year=year)








    if applications_obj.exists:
        applications_serailzer=applications_serializers(applications_obj,many=True)
        for application in applications_serailzer.data:
            approval_id=applications_approvals.filter(leave_id=application['id']).first()
            if approval_id is not None:
                application['approval_id']=approval_id.id
                if approval_id.approvedBy==False and approval_id.rejectedBy==False:
                    application['manager_action']=None
                    application['reverse_action']=None
                elif approval_id.approvedBy==True and approval_id.rejectedBy==False:
                    application['manager_action']=True
                    application['reverse_action']=True
                elif approval_id.approvedBy==False and approval_id.rejectedBy==True:
                    application['manager_action']=False
                    application['reverse_action']=False
                elif approval_id.approvedBy==True and approval_id.rejectedBy==True:
                    application['manager_action']=None
                    application['reverse_action']=None
                else:
                    application['manager_action']=None
                    application['reverse_action']=None
            else:
                application['approval_id']=''
                application['manager_action']=None
                application['reverse_action']=None


            date_obj = datetime.strptime(application['start_date'], "%d %b %Y").date()
            current_date = datetime.today().date()
            if date_obj < current_date:
                application['reverse_action']=None




            if application['attachment'] is not None and application['attachment'] !='':
                application['attachment']=imageUrl+application['attachment']
            else:
                application['attachment']=''
            if application['WorkFromHome']==True:
                application['ApplicationType']='Work From Home'
            else:
                application['ApplicationType']='Leave'

            if application['number_of_days'] is not None and application['number_of_days'] !='':
                application['total_days']=application['number_of_days']
            else:
                application['total_days']='NA'
                
            managers_obj = leaveApproval.objects.filter(leave_id=application['id']).distinct("managerId")
            managers_serializer = application_approval_serializers(managers_obj,many=True)
            application['managers']=managers_serializer.data

            for manager in application['managers']:
                if manager['approvedBy']==False and manager['rejectedBy']==False:
                    manager['status']=None
                elif manager['approvedBy']==True and manager['rejectedBy']==False:
                    manager['status']=True
                elif manager['approvedBy']==False and manager['rejectedBy']==True:
                    manager['status']=False
                elif manager['approvedBy']==True and manager['rejectedBy']==True:
                    manager['status']=None
                else:
                    manager['status']=None


        return Response ({"data":applications_serailzer.data, "response":{"n" : 1,"msg" : " Applications Found Successfully","status" : "success"}})
    else:
        return Response ({"data":[], "response":{"n" : 0,"msg" : "No Applications Found","status" : "success"}})

@api_view(['POST'])
def add_leave_type(request):
    if request.method == 'POST':
        data=request.data.copy()
        data['TypeName']=data['TypeName'].lower()
        data['is_active']=True
        data['company_code']=request.user.company_code

        already_exist_obj=LeaveTypeMaster.objects.filter(TypeName=data['TypeName'],is_active=True,company_code=data['company_code']).first()
        if already_exist_obj is not None:
            return Response ({"data":[], "response":{"n" : 0,"msg" : "Leave type name already exists","status" : "error"}})
        else:
            serializer=LeaveTypeMasterSerializer(data=data)
            if serializer.is_valid():
                serializer.save()
                return Response ({"data":[], "response":{"n" : 1,"msg" : "Leave type added successfully","status" : "success"}})
            else:
                first_key, first_value = next(iter(serializer.errors.items()))
                return Response ({"data":[],"response":{"n" : 0,"msg" : first_key[0] +' '+ first_value[0],"status" : "error"}})

@api_view(['POST'])
def update_leave_type(request):
    if request.method == 'POST':
        data=request.data.copy()
        data['TypeName']=data['TypeName'].lower()
        company_code=request.user.company_code
        if data['id'] is not None and data['id'] !='':
            update_leave_type_obj=LeaveTypeMaster.objects.filter(id=data['id'],is_active=True,company_code=company_code).first()
            if update_leave_type_obj is not None:

                already_exist_obj=LeaveTypeMaster.objects.filter(TypeName=data['TypeName'],is_active=True,company_code=company_code).exclude(id=data['id']).first()
                if already_exist_obj is not None:
                    return Response ({"data":[], "response":{"n" : 0,"msg" : "Leave type name already exists","status" : "error"}})
                else:
                    serializer=LeaveTypeMasterSerializer(update_leave_type_obj,data=data,partial=True)
                    if serializer.is_valid():
                        serializer.save()
                        return Response ({"data":[], "response":{"n" : 1,"msg" : "Leave type updated successfully","status" : "success"}})
                    else:
                        first_key, first_value = next(iter(serializer.errors.items()))
                        return Response ({"data":[],"response":{"n" : 0,"msg" : first_key[0] +' '+ first_value[0],"status" : "error"}})
            else:
                return Response ({"data":[], "response":{"n" : 0,"msg" : "leave typ not found ","status" : "error"}})
        else:
            return Response ({"data":[], "response":{"n" : 0,"msg" : "Please provide leave typ id ","status" : "error"}})

@api_view(['POST'])
def delete_leave_type(request):
    if request.method == 'POST':
        data=request.data.copy()
        company_code=request.user.company_code
        if data['id'] is not None and data['id'] !='':
            delete_leave_type_obj=LeaveTypeMaster.objects.filter(id=data['id'],is_active=True,company_code=company_code).first()
            if delete_leave_type_obj is not None:
                data['is_active']=False
                serializer=LeaveTypeMasterSerializer(delete_leave_type_obj,data=data,partial=True)
                if serializer.is_valid():
                    serializer.save()
                    return Response ({"data":[], "response":{"n" : 1,"msg" : "Leave type deleted successfully","status" : "success"}})
                else:
                    first_key, first_value = next(iter(serializer.errors.items()))
                    return Response ({"data":[],"response":{"n" : 0,"msg" : first_key[0] +' '+ first_value[0],"status" : "error"}})
            else:
                return Response ({"data":[], "response":{"n" : 0,"msg" : "leave typ not found ","status" : "error"}})
        else:
            return Response ({"data":[], "response":{"n" : 0,"msg" : "Please provide leave typ id ","status" : "error"}})

@api_view(['POST'])
def get_leave_type_by_id(request):
    if request.method == 'POST':
        data=request.data.copy()
        company_code=request.user.company_code
        if data['id'] is not None and data['id'] !='':
            leave_type_obj=LeaveTypeMaster.objects.filter(id=data['id'],is_active=True,company_code=company_code).first()
            if leave_type_obj is not None:
                serializer=LeaveTypeMasterSerializer(leave_type_obj)
                return Response ({"data":serializer.data, "response":{"n" : 1,"msg" : "Leave type found successfully","status" : "success"}})
            else:
                return Response ({"data":[], "response":{"n" : 0,"msg" : "leave typ not found ","status" : "error"}})
        else:
            return Response ({"data":[], "response":{"n" : 0,"msg" : "Please provide leave typ id ","status" : "error"}})

@api_view(['GET'])
def get_leave_type_list(request):
    if request.method == 'GET':
        company_code=request.user.company_code
        leave_type_obj=LeaveTypeMaster.objects.filter(is_active=True,company_code=company_code)
        if leave_type_obj is not None:
            serializer=LeaveTypeMasterSerializer(leave_type_obj,many=True)
            return Response ({"data":serializer.data, "response":{"n" : 1,"msg" : "Leave types found successfully","status" : "success"}})
        else:
            return Response ({"data":[], "response":{"n" : 0,"msg" : "leave typ not found ","status" : "error"}})




























