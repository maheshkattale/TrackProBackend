from Tasks.serializers import PostTaskMasterSerializer,ProjectTasksSerializer,GetTaskSerializer
from Tasks.models import ProjectTasks, Status, TaskMaster
from datetime import datetime
from Users.serializers import UserSerializer
from django.views.decorators.csrf import csrf_exempt
from Project.serializers import ProjectSerializer, ProjectTasksUpdateSerializer, ProjectUpdateSerializer, ProjectListSerializer,AllProjectListSerializer,ProjectTaskSerializer
from Project.models import ProjectMaster
from Users.models import Users
from rest_framework.decorators import api_view
from rest_framework.generics import GenericAPIView
from rest_framework.response import Response
from django.http import JsonResponse
from django.utils import timezone
import pytz
import calendar
from functools import reduce
from CheckTrackPro.views import stock_maindictlist
from datetime import datetime,timedelta
import datetime
from django.db.models import Q
from rest_framework import pagination
from rest_framework.pagination import PageNumberPagination
from Users.static_info import imageUrl
import os,psycopg2
from psycopg2.extras import RealDictCursor
from TrackProBackend.settings import *
import xlsxwriter
from openpyxl import load_workbook
import json

class CustomPagination(pagination.PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 50
    page_query_param = 'p'

    def get_paginated_response(self,data):
        response = Response({
            'n':1,
            'status':"success",
            'count':self.page.paginator.count,
            'next' : self.get_next_link(),
            'previous' : self.get_previous_link(),
            'data':data,
        })
        return response
# DISPLAY PROJECT LIST-------------------------------------------------------------------------------------


@api_view(['GET'])
def projectListAPI(request, format=None):
    """
    List all tasks, or create a new task.
    """
    if request.method == 'GET':
        project = ProjectMaster.objects.filter(company_code = request.user.company_code,Active=True).order_by('ProjectName')
        serializer = AllProjectListSerializer(project, many=True)
        return Response({'n':1,'Msg':'Projectlist fetched successfully','Status':'Success','data':serializer.data})


@api_view(['GET'])
def allProjectList(request, format=None):
    """
    List all tasks, or create a new task.
    """

    project = ProjectMaster.objects.filter(company_code = request.user.company_code,Active=True).order_by('ProjectName')
    serializer = ProjectSerializer(project, many=True)
    
    if project is not None:
        ProjectList=[]

        # ProjectList=serializer.data
        serializer = ProjectSerializer(project, many=True)

        for i in serializer.data:
            Projects = TaskMaster.objects.filter(Project=i['id'],Active=True).first()
            if Projects is not None:
                i['Used']=True
            else:
                i['Used']=False
            ProjectList.append(i)


        return Response({'n':1,'Msg':'Projectlist fetched successfully','Status':'Success','data':ProjectList})
    else:
        return Response({'n':0,'Msg':'No data found','Status':'Failed','data':''})


class paginationprojectlist(GenericAPIView):

    pagination_class = CustomPagination

    def post(self,request):
        projectname = request.POST.get('projectname')
        
        if projectname is None:
            projectname = ""

        project = ProjectMaster.objects.filter(ProjectName__icontains=projectname,company_code = request.user.company_code,Active=True).order_by('ProjectName')

        if project is not None:
            page4 = self.paginate_queryset(project)
            serializer = ProjectSerializer(page4, many=True)
            ProjectList=[]

            for i in serializer.data:
                Projects = TaskMaster.objects.filter(Project=i['id'],Active=True).first()
                if Projects is not None:
                    i['Used']=True
                else:
                    i['Used']=False
                ProjectList.append(i)

            return self.get_paginated_response(ProjectList)
            
        else:
            return Response({'n':0,'Msg':'No data found','Status':'Failed','data':''})
# Combined Project List Api--------------------------------------------------------------------------


@api_view(['POST'])
def combinedProjectListApi(request):
    """list all the projects and Ba and Co-ordinators"""
    ba = request.data.get('ba', None)
    coordinator = request.data.get('coordinator', None)
    company_code = request.user.company_code
    try:
        project = ProjectMaster.objects.filter(company_code=company_code)
        projserializer = ProjectSerializer(project, many=True)
        buisnessA = Users.objects.filter(
            is_active=True, is_admin=True, DepartmentID=ba)
        baserializer = UserSerializer(buisnessA, many=True)
        CoOD = Users.objects.filter(is_active=True, is_admin=True)
        coodserializer = UserSerializer(CoOD, many=True)

        return Response({"project": projserializer.data, "BA": baserializer.data, "Coordinator": coodserializer.data})
    except Exception as e:
        print(e)
# ADD PROJECT-------------------------------------------------------------------------------------


@api_view(['POST'])
def addProject(request):
    data = {"n": 0, "Msg": "", "Status": ""}
    user = request.user
    if user.is_admin == True:
        try:
            requestData = request.data.copy()
            ProjectName = requestData['ProjectName'].strip()
            requestData['ProjectName'] = ProjectName
            requestData['company_code'] = request.user.company_code
            
        except Exception as e:
            return Response({'Error': 'serializer not accepting data'})
        else:
            projectObject = ProjectMaster.objects.filter(ProjectName__in=[ProjectName.strip().capitalize(),ProjectName.strip(),ProjectName.title(),ProjectName.lower()],company_code=request.user.company_code,Active=True).first()
            if projectObject is None:
                serializer = ProjectUpdateSerializer(
                    data=requestData, context={'request': request})
                if serializer.is_valid():
                    serializer.validated_data['Active'] = True
                    serializer.validated_data['CreatedBy'] = user
                    u = serializer.save()
                    data['n'] = 1
                    data['Msg'] = 'project added succesfully'
                    data['Status'] = "Success"
                else:
                    data['n'] = 0
                    data['Msg'] = 'failed'
                    data['Status'] = "Failed"
            
                return Response(data=data)
            return Response({'n': 0, 'Msg': "Project already exists", 'Status': 'Failed'})
    else:
        return Response({'n': 0, 'Msg': "User has no permission to create", 'Status': 'Failed'})

# DELETE PROJECT-------------------------------------------------------------------------------------


@api_view(['POST'])
def projectDelete(request):
    data = {"n": 0, "Msg": "", "Status": ""}
    try:
        if request.query_params.get('projectID'):
            projectID = request.query_params.get('projectID')
        else:
            projectID = request.POST.get('projectID')
        if projectID is not None:
            Task_obj=TaskMaster.objects.filter(Project_id=projectID,Active=True).first()
            project = ProjectMaster.objects.filter(id=projectID, Active=True).first()
            if project is None:
                data['n'] = 0
                data['Msg'] = 'PROJECT DOES NOT EXIST'
                data['Status'] = "Failed"
            elif Task_obj is not None:
                data['n'] = 0
                data['Msg'] = "Unable to delete tasks are perform on this project"
                data['Status'] = "Failed"
            else:
                serializer = ProjectUpdateSerializer(project, request.data)
                if serializer.is_valid():
                    serializer.validated_data['Active'] = False
                    serializer.save()
                    data['n'] = 1
                    data['Msg'] = 'delete successfull'
                    data['Status'] = "Success"
                else:
                    data['n'] = 0
                    data['Msg'] = serializer.errors
                    data['Status'] = "Failed"
        else:
            data['n'] = 0
            data['Msg'] = 'project.id is none'
            data['Status'] = "Failed"
    except Exception as e:
        data['n'] = 0
        data['Msg'] = 'try method failed'
        data['Status'] = "Failed"
    return Response(data=data)
# UPDATE PROJECT-------------------------------------------------------------------------------------


@api_view(['POST'])
def projectUpdate(request):
    data = {'n': '', 'Msg': '', 'Status': ''}
    try:
        projectID = request.query_params.get('projectID')
        if projectID is None:
            data['n'] = 0
            data['Msg'] = 'project ID is none'
            data['Status'] = "Failed"
        else:
            project = ProjectMaster.objects.filter(id=projectID,company_code = request.user.company_code).first()
            if project is None:
                data['n'] = 0
                data['Msg'] = 'PROJECT DOES NOT EXIST'
                data['Status'] = "Failed"
            else:
                requestData = request.data.copy()
                ProjectName = requestData['ProjectName'].strip()
                requestData['ProjectName'] = ProjectName
                requestData['company_code'] = request.user.company_code
                projectObject = ProjectMaster.objects.filter(ProjectName__in=[ProjectName.strip().capitalize(),ProjectName.strip(),ProjectName.title(),ProjectName.lower()],company_code=request.user.company_code,Active=True).first()
                if project.ProjectName != requestData['ProjectName'] and projectObject is not None :
                    return Response({'n': 0, 'Msg': "Project already exists", 'Status': 'Failed'})
                serializer = ProjectUpdateSerializer(project, requestData)
                if serializer.is_valid():
                    serializer.validated_data['Active'] = True
                    serializer.validated_data['UpdatedBy'] = request.user
                    serializer.save()

                    data['n'] = 1
                    data['Msg'] = 'update successfull'
                    data['Status'] = "Success"
                else:
                    data = serializer.errors
        return Response(data=data)

    except Exception as e:
        data['n'] = 0
        data['Msg'] = 'try method failed'
        data['Status'] = "Failed"
        return Response(data=data)

# update coordinators-----------------------------------------------------------------------------


@api_view(['POST'])
def projectCoordinatorUpdate(request):
    data = {'n': '', 'Msg': '', 'Status': ''}
    try:
        projectID = request.query_params.get('projectID')
        if projectID is None:
            data['n'] = 0
            data['Msg'] = 'project ID is none'
            data['Status'] = "Failed"
        else:
            try:
                project = ProjectMaster.objects.get(id=projectID)
            except Exception as e:
                data['n'] = 0
                data['Msg'] = 'PROJECT DOES NOT EXIST'
                data['Status'] = "Failed"
            else:
                serializer = ProjectUpdateSerializer(project, request.data)
                if serializer.is_valid():
                    serializer.validated_data['UpdatedBy'] = request.user
                    serializer.save()

                    data['n'] = 1
                    data['Msg'] = 'update successfull'
                    data['Status'] = "Success"
                else:
                    data = serializer.errors
        return Response(data=data)

    except Exception as e:
        data['n'] = 0
        data['Msg'] = 'try method failed'
        data['Status'] = "Failed"
        return Response(data=data)

# get object in Project---------------------------------------------------------------------------------------------


@api_view(['GET'])
def getProject(request):
    if request.method == 'GET':
        id = request.query_params.get('projectID', None)
        if id is not None:
            i = ProjectMaster.objects.filter(id=id,company_code = request.user.company_code).first()
            if i is not None:
                serializer = ProjectUpdateSerializer(i)
                return Response({'n':1,'Msg':'Project fetched successfully','Status':'Success','data':serializer.data})
            return Response({'n':0,'Msg':'Project not found','Status':'Failed','data':''})
        return Response({'n':0,'Msg':'Please provide a project Id','Status':'Failed','data':''})
            


@api_view(['GET'])
def getProjectForUpdate(request):
    if request.method == 'GET':
        projectID = request.query_params.get('projectID', None)
        if projectID is not None:
            P = ProjectMaster.objects.filter(id=projectID)
            serializer = ProjectUpdateSerializer(P, many=True)
            return JsonResponse(serializer.data, safe=False)

# add project tasks--------------------------------------------------------------------------------


@api_view(['POST'])
def combinedProjectListApi(request):
    """list all the projects and Ba and Co-ordinators"""
    ba = request.data.get('ba', None)
    coordinator = request.data.get('coordinator', None)
    company_code = request.user.company_code
    try:
        project = ProjectMaster.objects.filter(company_code=company_code)
        projserializer = ProjectSerializer(project, many=True)
        buisnessA = Users.objects.filter(
            is_active=True, is_admin=True, DepartmentID=ba)
        baserializer = UserSerializer(buisnessA, many=True)
        CoOD = Users.objects.filter(is_active=True, is_admin=True)
        coodserializer = UserSerializer(CoOD, many=True)
        return Response({"project": projserializer.data, "BA": baserializer.data, "Coordinator": coodserializer.data})
    except Exception as e:
        print(e)

# Showlist of projecttasks-------------------------------------------------------------------------


@api_view(['GET'])
def projectTaskListAPI(request, format=None):
    """
    List all Project tasks, or create a new task.
    """
    if request.method == 'GET':
        company_code = request.user.company_code
        projecttask = ProjectTasks.objects.filter(company_code=company_code)
        serializer = ProjectTasksUpdateSerializer(projecttask, many=True)
        return Response(serializer.data)


# Add Project Tasks--------------------------------------------------------------------------------------
@api_view(['POST'])
def addProjectTask(request):
    user = request.user
    if user.is_admin == True:
        try:
            requestData = request.data.copy()
            requestData['company_code'] = request.user.company_code
            serializer = ProjectTasksUpdateSerializer(
                data=requestData, context={'request': request})
        except Exception as e:
            return Response({'Error': 'serializer not accepting data'})
        else:
            if serializer.is_valid():
                data = {}
                u = serializer.save()
                data['Status'] = 'task started successfully'
            else:
                data = serializer.errors
            return Response(data)
    else:
        return Response({'Error': 'User has no permission to create'})



# class projectreport(GenericAPIView):

#     pagination_class = CustomPagination

#     def post(self,request):
#         repeatYear=[]
#         uniqueYear=[] 
#         teamlist=[]
#         projectId = request.POST.get("projectName")
#         departname = request.POST.get("deptname")
#         if departname is not None and departname != "":
#             departname = [int(departname)]
#         else:
#             departname = []
#         employeename = request.POST.get("searchname")
#         pagenumber = request.data.get('p')
#         if projectId is not None and projectId != "":
#             projectname = ProjectMaster.objects.filter(id=projectId).first()

#             firsttaaskobject = TaskMaster.objects.filter(Project=projectId).order_by("id").first()
#             firsttaskser = GetTaskSerializer(firsttaaskobject)
#             if firsttaaskobject is not None:
#                 date = firsttaaskobject.AssignDate
#                 strdate = str(date)
#                 startmonth_name = calendar.month_abbr[int(strdate.split('-')[1])]  
#                 projectstart_date = strdate.split('-')[2] +" "+ startmonth_name +" "+strdate.split('-')[0]
#             else:
#                 projectstart_date="--"


#             lasttaskobject = TaskMaster.objects.filter(Project=projectId).order_by("-id").first()
#             lasttaskser =  PostTaskMasterSerializer(lasttaskobject)
#             if lasttaskobject is not None:
#                 date = lasttaskobject.AssignDate
#                 enddate = str(date)
#                 endmonth_name = calendar.month_abbr[int(enddate.split('-')[1])]  
#                 projectend_date = enddate.split('-')[2] +" "+endmonth_name+" "+enddate.split('-')[0]
#             else:
#                 projectend_date="--"

#             taskteamcount = TaskMaster.objects.filter(Project=projectId).distinct('AssignTo_id').count()

#             taskteamstr = TaskMaster.objects.filter(Project=projectId).distinct('AssignTo_id')
#             if taskteamstr is not None:
#                 teaskteamser =  GetTaskSerializer(taskteamstr,many=True)
#                 for t in teaskteamser.data:
#                     userObject=Users.objects.filter(id=t['AssignTo']).first()
#                     teammate = userObject.Firstname+" "+userObject.Lastname
#                     teamlist.append(teammate)

#             firsttotaltaskcount = TaskMaster.objects.filter(Project=projectId,AssignTo__in =  list( Users.objects.filter(Q(Firstname__icontains=employeename)|Q(Lastname__icontains=employeename)).values_list('id', flat=True)))

#             totaltaskcountobject = ""

#             if departname is not None and departname != []:
#                 totaltaskcountobject = firsttotaltaskcount.filter(AssignTo__in = list( Users.objects.filter(DepartmentID__in=departname).values_list('id', flat=True))).count()
#             else:
#                 totaltaskcountobject = firsttotaltaskcount.count()


#             totaltaskcount=totaltaskcountobject

#             taskObject=TaskMaster.objects.filter(Project=projectId,AssignTo__in =  list( Users.objects.filter(Q(Firstname__icontains=employeename)|Q(Lastname__icontains=employeename)).values_list('id', flat=True))).order_by('CreatedOn')

#             firstobject = ""
#             if departname is not None and departname != []:
#                 firstobject = taskObject.filter(AssignTo__in = list( Users.objects.filter(DepartmentID__in=departname).values_list('id', flat=True))).order_by('CreatedOn')
             

#             else:
#                 firstobject = taskObject

#             finalobject = firstobject
            
#             if finalobject.exists():
#                 page4 = self.paginate_queryset(finalobject)
#                 ser=PostTaskMasterSerializer(page4,many=True)
#                 projecthours=0


#                 for i in ser.data:
#                     endtasktime = TaskMaster.objects.filter(id=i['id']).first()
#                     if endtasktime is not None:
#                         if endtasktime.IsParent == True:
#                             childtaskobj = TaskMaster.objects.filter(ParentTaskId=i['id']).last()
#                             childtaskid = childtaskobj.id
#                             projectendtasktime = ProjectTasks.objects.filter(Task=childtaskid).order_by('-id').first()
#                             endtaskstring=projectendtasktime.EndDate
#                             if endtaskstring :
#                                 endtasktimestring = str(endtaskstring)
#                                 etime = endtasktimestring.split(' ')[0]
#                                 emonth_name = calendar.month_abbr[int(etime.split('-')[1])] 
#                                 userendtaskdate = etime.split('-')[2] +" "+emonth_name+" "+etime.split('-')[0]
#                                 i['endtaskdate']=userendtaskdate
#                             else:
#                                 i['endtaskdate']="--"
#                         elif endtasktime.IsParent == False:
                          
#                             if endtasktime.ParentTaskId is not None:
#                                 PARENTID = endtasktime.ParentTaskId
#                                 childtaskobj = TaskMaster.objects.filter(ParentTaskId=int(PARENTID)).order_by('-id').first()
#                                 childtaskid = childtaskobj.id
#                                 projectendtasktime = ProjectTasks.objects.filter(Task=childtaskid).order_by('-id').first()
#                                 endtaskstring=projectendtasktime.EndDate
#                                 if endtaskstring :
#                                     endtasktimestring = str(endtaskstring)
#                                     etime = endtasktimestring.split(' ')[0]
#                                     emonth_name = calendar.month_abbr[int(etime.split('-')[1])] 
#                                     userendtaskdate = etime.split('-')[2] +" "+emonth_name+" "+etime.split('-')[0]
#                                     i['endtaskdate']=userendtaskdate
#                                 else:
#                                     i['endtaskdate']="--"
#                             else:
#                                 projectendtasktime = ProjectTasks.objects.filter(Task=i['id']).order_by('-id').first()
#                                 if projectendtasktime is not None:
#                                     endtaskstring=projectendtasktime.EndDate
#                                     if endtaskstring :
#                                         endtasktimestring = str(endtaskstring)
#                                         etime = endtasktimestring.split(' ')[0]
#                                         emonth_name = calendar.month_abbr[int(etime.split('-')[1])] 
#                                         userendtaskdate = etime.split('-')[2] +" "+emonth_name+" "+etime.split('-')[0]
#                                         i['endtaskdate']=userendtaskdate
#                                     else:
#                                         i['endtaskdate']="--"
#                                 else:
#                                     i['endtaskdate']="--"

#                     strdate = str(i['AssignDate'])
#                     smonth_name = calendar.month_abbr[int(strdate.split('-')[1])]
#                     newdate = strdate.split('-')[2] +" "+smonth_name+" "+strdate.split('-')[0]
#                     i['AssignDate']=newdate


#                     userObject=Users.objects.filter(id=i['AssignTo']).first()
#                     i['CreatedBy']=userObject.Firstname+" "+userObject.Lastname
#                     repeatYear.append(i['Year'])

#                     if i['Zone'] == 1:
#                         i['grade'] = "<span data-bs-toggle='tooltip' data-bs-placement='bottom' title='Green'><img src='/static/Media/taskicons/activegreenpoints.svg' class='activeicons' alt='Paris'></span>"

#                     if i['Zone'] == 2:
#                         i['grade'] = "<span data-bs-toggle='tooltip' data-bs-placement='bottom' title='Yellow'><img src='/static/Media/taskicons/activeyellowpoints.svg' class='activeicons' alt='Paris'></span>"

#                     if i['Zone'] == 3:
#                         i['grade'] = "<span data-bs-toggle='tooltip' data-bs-placement='bottom' title='Red'><img src='/static/Media/taskicons/activeredpoints.svg' class='activeicons' alt='Paris' ></span>"

#                     if i['Zone'] == 4:
#                         i['grade'] = "<span data-bs-toggle='tooltip' data-bs-placement='bottom' title='Not Done'><img src='/static/Media/taskicons/activenotdonepoints.svg' class='activeicons' alt='Paris'></span>"

#                     if i['Zone'] == 5:
#                         i['grade'] = "<span data-bs-toggle='tooltip' data-bs-placement='bottom' title='Cancelled'><img src='/static/Media/taskicons/activecancelledpoints.svg' class='activeicons' alt='Paris'></span>"

#                     if i['Zone'] == 6:
#                         i['grade'] = "<span data-bs-toggle='tooltip' data-bs-placement='bottom' title='Rejected'><img src='/static/Media/taskicons/activerejectpoints.svg' class='activeicons' alt='Paris'></span>"

#                     if i['Zone'] == "" or i['Zone'] is None:
#                         i['grade'] = "--"

#                     currentzone = pytz.timezone("Asia/Kolkata") 
#                     currenttime = datetime.datetime.now(currentzone)
#                     newcurrenttime =currenttime.strftime("%Y-%m-%dT%H:%M:%S.%f%z")

#                     taskobj = TaskMaster.objects.filter(id=i['id']).first()
#                     taskidlist = []
#                     if taskobj.ParentTaskId is not None:
#                         parenttaskid  = taskobj.ParentTaskId
#                         taskidlist.append(parenttaskid)
#                         taskobject = TaskMaster.objects.filter(ParentTaskId=parenttaskid)
#                         taskser = PostTaskMasterSerializer(taskobject,many=True)
#                         for t in taskser.data:
#                             taskidlist.append(t['id'])
#                     else:
#                         taskidlist.append(i['id'])
#                         taskobject = TaskMaster.objects.filter(ParentTaskId=i['id'])
#                         taskser = PostTaskMasterSerializer(taskobject,many=True)
#                         for t in taskser.data:
#                             taskidlist.append(t['id'])


#                     projecttasktime = ProjectTasks.objects.filter(Task_id__in=taskidlist).order_by("id")
#                     if projecttasktime is not None:
#                         projectser = ProjectTasksSerializer(projecttasktime,many=True)
#                         totaltime=0
#                         for o in projectser.data:
#                             startstring = o['StartDate']
#                             starttime=startstring
#                             t1 = datetime.datetime.strptime(starttime, "%Y-%m-%dT%H:%M:%S.%f%z")

#                             endstring = o['EndDate']
#                             if endstring is not None:
#                                 endtime = o['EndDate']
#                             else:
#                                 endtime = str(newcurrenttime)
#                             t2=datetime.datetime.strptime(endtime, "%Y-%m-%dT%H:%M:%S.%f%z")
#                             tdelta=t2-t1
                        
                        
#                             if "day" in str(tdelta) or "days" in str(tdelta):
#                                 daystring = str(tdelta).split(",")[0]
#                                 noofdays = str(daystring).split(" ")[0]
#                                 daysmins = int(noofdays)*1440

#                                 thoursstr =  str(tdelta).split(",")[1]
#                                 thours = str(thoursstr).split(":")[0]
#                                 hrs = int(thours)*60
#                                 tmins = str(thoursstr).split(":")[1]
#                                 finalmins = int(hrs)+int(tmins)+int(daysmins)
#                             else:
#                                 thours = str(tdelta).split(":")[0]
#                                 hrs = int(thours)*60
#                                 tmins = str(tdelta).split(":")[1]
#                                 finalmins = int(hrs)+int(tmins)
#                             totaltime += finalmins
#                         projecthours += totaltime

#                         totalhours =totaltime
#                         hour = int (totalhours) // 60
#                         if (len(str(hour)) < 2):
#                             hours = "0"+str(hour)
#                         else:
#                             hours = str(hour)

#                         mins = int (totalhours) % 60
#                         if (len(str(mins)) < 2):
#                             minutes = "0"+str(mins)
#                         else:
#                             minutes = str(mins)

#                         hourstring = str(hours)+":"+str(minutes)+" "+"Hrs"

#                         i['taskhours'] = hourstring

#                     #  calculate day wise time strings
#                     protimelist = []
#                     seconddaylist = []
#                     finallist=[]
                
#                     for s in projectser.data:
#                         startstring = s['StartDate']
#                         enddatestr = s['EndDate']

#                         a = startstring.split('T')[0]
#                         revdate = a.split('-')[2]+"-"+ a.split('-')[1]+"-"+ a.split('-')[0]
#                         startdatestring = revdate

#                         if enddatestr is not None:
#                             b = enddatestr.split('T')[0]
#                             revenddate = b.split('-')[2]+"-"+ b.split('-')[1]+"-"+ b.split('-')[0]
#                             enddatestring = revenddate
#                         else:
#                             b = str(newcurrenttime).split('T')[0]
#                             revenddate = b.split('-')[2]+"-"+ b.split('-')[1]+"-"+ b.split('-')[0]
#                             enddatestring = revenddate

#                         # if task played and closed on same date
#                         if startdatestring == enddatestring:

#                             starttime=startstring
#                             t1 = datetime.datetime.strptime(starttime, "%Y-%m-%dT%H:%M:%S.%f%z")

#                             endstring = s['EndDate']
#                             if endstring is not None:
#                                 endtime = s['EndDate']
#                             else:
#                                 endtime = str(newcurrenttime)
#                             t2=datetime.datetime.strptime(endtime, "%Y-%m-%dT%H:%M:%S.%f%z")
#                             dtdelta=t2-t1
                        
                        
#                             if "day" in str(dtdelta) or "days" in str(dtdelta):
#                                 daystring = str(dtdelta).split(",")[0]
#                                 noofdays = str(daystring).split(" ")[0]
#                                 daysmins = int(noofdays)*1440

#                                 thoursstr =  str(dtdelta).split(",")[1]
#                                 thours = str(thoursstr).split(":")[0]
#                                 hrs = int(thours)*60
#                                 tmins = str(thoursstr).split(":")[1]
#                                 dfinalmins = int(hrs)+int(tmins)+int(daysmins)
#                             else:
#                                 dthours = str(dtdelta).split(":")[0]
#                                 dhrs = int(dthours)*60
#                                 dtmins = str(dtdelta).split(":")[1]
#                                 dfinalmins = int(dhrs)+int(dtmins)

#                             timeport ={
#                                 "date":startdatestring,
#                                 "time":dfinalmins
#                             }
#                             protimelist.append(timeport)
#                         #if task played and closed on different dates
#                         else:
                            
#                             # for startdate till 12 o'clock calculation
#                             starttime=startstring
#                             t1 = datetime.datetime.strptime(starttime, "%Y-%m-%dT%H:%M:%S.%f%z")

#                             startdateendtime = str(a)+"T"+"23:59:00.00000+05:30"
#                             strtdatetym = str(startdateendtime)
#                             t2=datetime.datetime.strptime(strtdatetym, "%Y-%m-%dT%H:%M:%S.%f%z")
#                             sdtdelta=t2-t1
                        
                        
#                             if "day" in str(sdtdelta) or "days" in str(sdtdelta):
#                                 daystring = str(sdtdelta).split(",")[0]
#                                 noofdays = str(daystring).split(" ")[0]
#                                 daysmins = int(noofdays)*1440

#                                 thoursstr =  str(sdtdelta).split(",")[1]
#                                 thours = str(thoursstr).split(":")[0]
#                                 hrs = int(thours)*60
#                                 tmins = str(thoursstr).split(":")[1]
#                                 frstdayfinalmins = int(hrs)+int(tmins)+int(daysmins)
#                             else:
#                                 dthours = str(sdtdelta).split(":")[0]
#                                 dhrs = int(dthours)*60
#                                 dtmins = str(sdtdelta).split(":")[1]
#                                 frstdayfinalmins = int(dhrs)+int(dtmins)

#                             secondtimeport ={
#                                 "date":startdatestring,
#                                 "time":frstdayfinalmins
#                             }
#                             seconddaylist.append(secondtimeport)
#                         #calculation of second day

#                             seconddaystrt = str(b)+"T"+"00:00:00.00000+05:30"
#                             seconddaystrtdatetym = str(seconddaystrt)
#                             t11=datetime.datetime.strptime(seconddaystrtdatetym, "%Y-%m-%dT%H:%M:%S.%f%z")

#                             endstring = s['EndDate']
#                             if endstring is not None:
#                                 endtime = s['EndDate']
#                             else:
#                                 endtime = str(newcurrenttime)

#                             seconddayenddt = endtime
#                             t12 = datetime.datetime.strptime(seconddayenddt, "%Y-%m-%dT%H:%M:%S.%f%z")

#                             seconddaytdelta=t12-t11

#                             if "day" in str(seconddaytdelta) or "days" in str(seconddaytdelta):
#                                 daystring = str(seconddaytdelta).split(",")[0]
#                                 noofdays = str(daystring).split(" ")[0]
#                                 daysmins = int(noofdays)*1440

#                                 thoursstr =  str(seconddaytdelta).split(",")[1]
#                                 thours = str(thoursstr).split(":")[0]
#                                 hrs = int(thours)*60
#                                 tmins = str(thoursstr).split(":")[1]
#                                 secondayfinalmins = int(hrs)+int(tmins)+int(daysmins)
#                             else:
#                                 dthours = str(seconddaytdelta).split(":")[0]
#                                 dhrs = int(dthours)*60
#                                 dtmins = str(seconddaytdelta).split(":")[1]
#                                 secondayfinalmins = int(dhrs)+int(dtmins)

#                             secondtimeport ={
#                                 "date":enddatestring,
#                                 "time":secondayfinalmins
#                             }
#                             seconddaylist.append(secondtimeport)

                        
#                         finallist = protimelist+seconddaylist


#                     # finaldatetimelist= reduce(lambda d1,d2: {k: d1.get(k,0)+d2.get(k,0)for k in set(d1)|set(d2)}, timelist)
                
#                     maindata = []
#                     for p in finallist:
                    
#                         v=stock_maindictlist("date", p['date'],"time",p['time'], maindata)
#                         if v == {}:
#                             timecalc={
#                                 "date":p['date'],
#                                 "time":p['time'],
#                             }
#                             maindata.append(timecalc)
                    
#                     for m in maindata:
#                         totalminutes = m['time']
#                         hour = int (totalminutes)//60
#                         if (len(str(hour)) < 2):
#                             hours = "0"+str(hour)
#                         else:
#                             hours = str(hour)

#                         mins = int (totalminutes) % 60
#                         if (len(str(mins)) < 2):
#                             minutes = "0"+str(mins)
#                         else:
#                             minutes = str(mins)

#                         hourstring = str(hours)+" Hrs "+str(minutes)+" mins"

#                         m['time'] = hourstring

#                     i['daywisetime']=maindata

#                     hovertaskidlist = []
#                     if taskobj.ParentTaskId is not None:
#                         hovertaskparentid = taskobj.ParentTaskId
#                         hovertaskidlist.append(hovertaskparentid)
#                         ttaskobject = TaskMaster.objects.filter(ParentTaskId=hovertaskparentid).exclude(id=i['id'])
#                         ttaskser = PostTaskMasterSerializer(ttaskobject,many=True)
#                         for t in ttaskser.data:
#                             hovertaskidlist.append(t['id'])

#                         hoverobjects = TaskMaster.objects.filter(id__in=hovertaskidlist)
#                         hovertaskser = PostTaskMasterSerializer(hoverobjects,many=True)
#                         for h in hovertaskser.data:

#                             userObject=Users.objects.filter(id=h['AssignTo']).first()
#                             h['CreatedBy']=userObject.Firstname+" "+userObject.Lastname

#                             chdate = str(h['AssignDate'])
#                             chmonth_name = calendar.month_abbr[int(chdate.split('-')[1])] 
#                             h['Changeddate'] = chdate.split('-')[2] +" "+chmonth_name+" "+chdate.split('-')[0]
                           

#                             hprojecttasktime = ProjectTasks.objects.filter(Task__in=taskidlist).order_by("id")
#                             if hprojecttasktime :
#                                 hprojectser = ProjectTasksSerializer(hprojecttasktime,many=True)
#                                 FMT = '%H:%M:%S.%f'
#                                 htotaltime=0
#                                 for o in hprojectser.data:
#                                     startstring = o['StartDate']
#                                     starttime=startstring
#                                     t1 = datetime.datetime.strptime(starttime, "%Y-%m-%dT%H:%M:%S.%f%z")

#                                     endstring = o['EndDate']
#                                     if endstring is not None:
#                                         endtime = o['EndDate']
#                                     else:
#                                         endtime = str(newcurrenttime)
#                                     t2=datetime.datetime.strptime(endtime, "%Y-%m-%dT%H:%M:%S.%f%z")
#                                     tdelta=t2-t1
                                
                                
#                                     if "day" in str(tdelta) or "days" in str(tdelta):
#                                         daystring = str(tdelta).split(",")[0]
#                                         noofdays = str(daystring).split(" ")[0]
#                                         daysmins = int(noofdays)*1440

#                                         thoursstr =  str(tdelta).split(",")[1]
#                                         thours = str(thoursstr).split(":")[0]
#                                         hrs = int(thours)*60
#                                         tmins = str(thoursstr).split(":")[1]
#                                         hfinalmins = int(hrs)+int(tmins)+int(daysmins)
#                                     else:
#                                         thours = str(tdelta).split(":")[0]
#                                         hrs = int(thours)*60
#                                         tmins = str(tdelta).split(":")[1]
#                                         hfinalmins = int(hrs)+int(tmins)
#                                     htotaltime += hfinalmins

#                                 htotalhours =htotaltime
#                                 hhour = int (htotalhours) // 60
#                                 if (len(str(hhour)) < 2):
#                                     rhhours = "0"+str(hhour)
#                                 else:
#                                     rhhours = str(hhour)

#                                 hmins = int (htotalhours) % 60
#                                 if (len(str(hmins)) < 2):
#                                     rminutes = "0"+str(hmins)
#                                 else:
#                                     rminutes = str(hmins)

#                                 rhourstring = str(rhhours)+":"+str(rminutes)+" "+"Hrs"

#                                 h['taskhoursstr'] = rhourstring

#                             remarkzone = h['Zone']
#                             if remarkzone == 1:
#                                 h['remarkstr'] = "<img src='/static/Media/taskicons/activegreenpoints.svg' class='activeicons' alt='green'>"
#                             elif remarkzone == 2:
#                                 h['remarkstr'] = "<img src='/static/Media/taskicons/activeyellowpoints.svg' class='activeicons' alt='yellow'>"
#                             elif remarkzone == 3:
#                                 h['remarkstr'] = "<img src='/static/Media/taskicons/activeredpoints.svg' class='activeicons' alt='red'>"
#                             elif remarkzone == 4:
#                                 h['remarkstr'] = "<img src='/static/Media/taskicons/activenotdonepoints.svg' class='activeicons' alt='notdone'>"
#                             elif remarkzone == 5:
#                                 h['remarkstr'] = "<img src='/static/Media/taskicons/activecancelledpoints.svg' class='activeicons' alt='notdone'>"
#                             elif remarkzone == 6:
#                                 h['remarkstr'] = "<img src='/static/Media/taskicons/activerejectpoints.svg' class='activeicons' alt='notdone'>"
#                             else:
#                                 h['remarkstr'] = "----"


#                             bonusstr = h['Bonus']
#                             if bonusstr == True:
#                                 h['bonusstr'] = "<img src='/static/Media/taskicons/activebonuspoints.svg' alt='bonus'>"
#                             else:
#                                 h['bonusstr'] = ""

#                         i['Hovertaskdata']  = hovertaskser.data
#                     else:
#                         i['Hovertaskdata'] = [] 
#                     if i['ParentTaskId'] is None:
#                         i['ParentTaskId'] = ""

#                 projecttotalminutes = projecthours
#                 phour = int (projecttotalminutes) // 60
#                 if (len(str(phour)) < 2):
#                     phours = "0"+str(phour)
#                 else:
#                     phours = str(phour)

#                 pmins = int (projecttotalminutes) % 60
#                 if (len(str(pmins)) < 2):
#                     pminutes = "0"+str(pmins)
#                 else:
#                     pminutes = str(pmins)

#                 phourstring = str(phours)+":"+str(pminutes)+" "+"Hrs"

#                 projecttotalhours = phourstring
#                 flmandays = phour/9
#                 mandays = round(flmandays,1)

#                 context ={
#                 "fisrttaskdetail":firsttaskser.data,
#                 "teammatecount" : taskteamcount,
#                 "totaltaskcount" : totaltaskcount,
#                 "lasttaskdetail":lasttaskser.data,
#                 "teamlist":teamlist,
#                 "projectstartdate":projectstart_date,
#                 "projectend_date":projectend_date,
#                 "projecttotalhours":projecttotalhours,
#                 "mandays":mandays,
#                 "uniqueYear":uniqueYear,
#                 "Taskdata":ser.data
#                 }
                
#                 return self.get_paginated_response(context)
#             return Response({'n':0,"status":"failed","message":"Data not found"})
#         return Response({'n':0,"status":"failed","message":"Please provide project id"})







class projectreport(GenericAPIView):

    pagination_class = CustomPagination

    def post(self,request):
        repeatYear=[]
        uniqueYear=[] 
        teamlist=[]
        projectId = request.POST.get("projectName")
        departname = request.POST.get("deptname")
        if departname is not None and departname != "":
            departname = [int(departname)]
        else:
            departname = []
        employeename = request.POST.get("searchname")
        pagenumber = request.data.get('p')
        if projectId is not None and projectId != "":
            projectname = ProjectMaster.objects.filter(id=projectId).first()

            firsttaaskobject = TaskMaster.objects.filter(Project=projectId).order_by("id").first()
            firsttaskser = GetTaskSerializer(firsttaaskobject)
            if firsttaaskobject is not None:
                date = firsttaaskobject.AssignDate
                strdate = str(date)
                startmonth_name = calendar.month_abbr[int(strdate.split('-')[1])]  
                projectstart_date = strdate.split('-')[2] +" "+ startmonth_name +" "+strdate.split('-')[0]
            else:
                projectstart_date="--"


            lasttaskobject = TaskMaster.objects.filter(Project=projectId).order_by("-id").first()
            lasttaskser =  PostTaskMasterSerializer(lasttaskobject)
            if lasttaskobject is not None:
                date = lasttaskobject.AssignDate
                enddate = str(date)
                endmonth_name = calendar.month_abbr[int(enddate.split('-')[1])]  
                projectend_date = enddate.split('-')[2] +" "+endmonth_name+" "+enddate.split('-')[0]
            else:
                projectend_date="--"

            taskteamcount = TaskMaster.objects.filter(Project=projectId).distinct('AssignTo_id').count()

            taskteamstr = TaskMaster.objects.filter(Project=projectId).distinct('AssignTo_id')
            if taskteamstr is not None:
                teaskteamser =  GetTaskSerializer(taskteamstr,many=True)
                for t in teaskteamser.data:
                    userObject=Users.objects.filter(id=t['AssignTo']).first()
                    teammate = userObject.Firstname+" "+userObject.Lastname
                    teamlist.append(teammate)

            firsttotaltaskcount = TaskMaster.objects.filter(Project=projectId,AssignTo__in =  list( Users.objects.filter(Q(Firstname__icontains=employeename)|Q(Lastname__icontains=employeename)).values_list('id', flat=True)))

            totaltaskcountobject = ""

            if departname is not None and departname != []:
                totaltaskcountobject = firsttotaltaskcount.filter(AssignTo__in = list( Users.objects.filter(DepartmentID__in=departname).values_list('id', flat=True))).count()
            else:
                totaltaskcountobject = firsttotaltaskcount.count()


            totaltaskcount=totaltaskcountobject

            taskObject=TaskMaster.objects.filter(Project=projectId,AssignTo__in =  list( Users.objects.filter(Q(Firstname__icontains=employeename)|Q(Lastname__icontains=employeename)).values_list('id', flat=True))).order_by('CreatedOn')

            firstobject = ""
            if departname is not None and departname != []:
                firstobject = taskObject.filter(AssignTo__in = list( Users.objects.filter(DepartmentID__in=departname).values_list('id', flat=True))).order_by('CreatedOn')
             

            else:
                firstobject = taskObject

            finalobject = firstobject
            
            if finalobject.exists():
                page4 = self.paginate_queryset(finalobject)
                ser=PostTaskMasterSerializer(page4,many=True)
                projecthours=0

                
                for i in ser.data:
                    endtasktime = TaskMaster.objects.filter(id=i['id']).first()
                    if endtasktime is not None:
                        if endtasktime.IsParent == True:
                            childtaskobj = TaskMaster.objects.filter(ParentTaskId=i['id']).last()
                            childtaskid = childtaskobj.id
                            projectendtasktime = ProjectTasks.objects.filter(Task=childtaskid).order_by('-id').first()
                            endtaskstring=projectendtasktime.EndDate
                            if endtaskstring :
                                endtasktimestring = str(endtaskstring)
                                etime = endtasktimestring.split(' ')[0]
                                emonth_name = calendar.month_abbr[int(etime.split('-')[1])] 
                                userendtaskdate = etime.split('-')[2] +" "+emonth_name+" "+etime.split('-')[0]
                                i['endtaskdate']=userendtaskdate
                            else:
                                i['endtaskdate']="--"
                        elif endtasktime.IsParent == False:
                          
                            if endtasktime.ParentTaskId is not None:
                                PARENTID = endtasktime.ParentTaskId
                                childtaskobj = TaskMaster.objects.filter(ParentTaskId=int(PARENTID)).order_by('-id').first()
                                childtaskid = childtaskobj.id
                                projectendtasktime = ProjectTasks.objects.filter(Task=childtaskid).order_by('-id').first()
                                endtaskstring=projectendtasktime.EndDate
                                if endtaskstring :
                                    endtasktimestring = str(endtaskstring)
                                    etime = endtasktimestring.split(' ')[0]
                                    emonth_name = calendar.month_abbr[int(etime.split('-')[1])] 
                                    userendtaskdate = etime.split('-')[2] +" "+emonth_name+" "+etime.split('-')[0]
                                    i['endtaskdate']=userendtaskdate
                                else:
                                    i['endtaskdate']="--"
                            else:
                                projectendtasktime = ProjectTasks.objects.filter(Task=i['id']).order_by('-id').first()
                                if projectendtasktime is not None:
                                    endtaskstring=projectendtasktime.EndDate
                                    if endtaskstring :
                                        endtasktimestring = str(endtaskstring)
                                        etime = endtasktimestring.split(' ')[0]
                                        emonth_name = calendar.month_abbr[int(etime.split('-')[1])] 
                                        userendtaskdate = etime.split('-')[2] +" "+emonth_name+" "+etime.split('-')[0]
                                        i['endtaskdate']=userendtaskdate
                                    else:
                                        i['endtaskdate']="--"
                                else:
                                    i['endtaskdate']="--"

                    strdate = str(i['AssignDate'])
                    smonth_name = calendar.month_abbr[int(strdate.split('-')[1])]
                    newdate = strdate.split('-')[2] +" "+smonth_name+" "+strdate.split('-')[0]
                    i['AssignDate']=newdate


                    userObject=Users.objects.filter(id=i['AssignTo']).first()
                    i['CreatedBy']=userObject.Firstname+" "+userObject.Lastname
                    repeatYear.append(i['Year'])

                    if i['Zone'] == 1:
                        i['grade'] = "<span data-bs-toggle='tooltip' data-bs-placement='bottom' title='Green'><img src='/static/Media/taskicons/activegreenpoints.svg' class='activeicons' alt='Paris'></span>"

                    if i['Zone'] == 2:
                        i['grade'] = "<span data-bs-toggle='tooltip' data-bs-placement='bottom' title='Yellow'><img src='/static/Media/taskicons/activeyellowpoints.svg' class='activeicons' alt='Paris'></span>"

                    if i['Zone'] == 3:
                        i['grade'] = "<span data-bs-toggle='tooltip' data-bs-placement='bottom' title='Red'><img src='/static/Media/taskicons/activeredpoints.svg' class='activeicons' alt='Paris' ></span>"

                    if i['Zone'] == 4:
                        i['grade'] = "<span data-bs-toggle='tooltip' data-bs-placement='bottom' title='Not Done'><img src='/static/Media/taskicons/activenotdonepoints.svg' class='activeicons' alt='Paris'></span>"

                    if i['Zone'] == 5:
                        i['grade'] = "<span data-bs-toggle='tooltip' data-bs-placement='bottom' title='Cancelled'><img src='/static/Media/taskicons/activecancelledpoints.svg' class='activeicons' alt='Paris'></span>"

                    if i['Zone'] == 6:
                        i['grade'] = "<span data-bs-toggle='tooltip' data-bs-placement='bottom' title='Rejected'><img src='/static/Media/taskicons/activerejectpoints.svg' class='activeicons' alt='Paris'></span>"

                    if i['Zone'] == "" or i['Zone'] is None:
                        i['grade'] = "--"

                    currentzone = pytz.timezone("Asia/Kolkata") 
                    currenttime = datetime.datetime.now(currentzone)
                    newcurrenttime =currenttime.strftime("%Y-%m-%dT%H:%M:%S.%f%z")

                    taskobj = TaskMaster.objects.filter(id=i['id']).first()
                    taskidlist = []
                    if taskobj.ParentTaskId is not None:
                        parenttaskid  = taskobj.ParentTaskId
                        taskidlist.append(parenttaskid)
                        taskobject = TaskMaster.objects.filter(ParentTaskId=parenttaskid)
                        taskser = PostTaskMasterSerializer(taskobject,many=True)
                        for t in taskser.data:
                            taskidlist.append(t['id'])
                    else:
                        taskidlist.append(i['id'])
                        taskobject = TaskMaster.objects.filter(ParentTaskId=i['id'])
                        taskser = PostTaskMasterSerializer(taskobject,many=True)
                        for t in taskser.data:
                            taskidlist.append(t['id'])


                    projecttasktime = ProjectTasks.objects.filter(Task_id__in=taskidlist).order_by("id")
                    if projecttasktime is not None:
                        projectser = ProjectTasksSerializer(projecttasktime,many=True)
                        totaltime=0
                        for o in projectser.data:
                            startstring = o['StartDate']
                            starttime=startstring
                            t1 = datetime.datetime.strptime(starttime, "%Y-%m-%dT%H:%M:%S.%f%z")

                            endstring = o['EndDate']
                            if endstring is not None:
                                endtime = o['EndDate']
                            else:
                                endtime = str(newcurrenttime)
                            t2=datetime.datetime.strptime(endtime, "%Y-%m-%dT%H:%M:%S.%f%z")
                            tdelta=t2-t1
                        
                        
                            if "day" in str(tdelta) or "days" in str(tdelta):
                                daystring = str(tdelta).split(",")[0]
                                noofdays = str(daystring).split(" ")[0]
                                daysmins = int(noofdays)*1440

                                thoursstr =  str(tdelta).split(",")[1]
                                thours = str(thoursstr).split(":")[0]
                                hrs = int(thours)*60
                                tmins = str(thoursstr).split(":")[1]
                                finalmins = int(hrs)+int(tmins)+int(daysmins)
                            else:
                                thours = str(tdelta).split(":")[0]
                                hrs = int(thours)*60
                                tmins = str(tdelta).split(":")[1]
                                finalmins = int(hrs)+int(tmins)
                            totaltime += finalmins
                        projecthours += totaltime

                        totalhours =totaltime
                        hour = int (totalhours) // 60
                        if (len(str(hour)) < 2):
                            hours = "0"+str(hour)
                        else:
                            hours = str(hour)

                        mins = int (totalhours) % 60
                        if (len(str(mins)) < 2):
                            minutes = "0"+str(mins)
                        else:
                            minutes = str(mins)

                        hourstring = str(hours)+":"+str(minutes)+" "+"Hrs"

                        i['taskhours'] = hourstring

                    #  calculate day wise time strings
                    protimelist = []
                    seconddaylist = []
                    finallist=[]
                
                    for s in projectser.data:
                        startstring = s['StartDate']
                        enddatestr = s['EndDate']

                        a = startstring.split('T')[0]
                        revdate = a.split('-')[2]+"-"+ a.split('-')[1]+"-"+ a.split('-')[0]
                        startdatestring = revdate

                        if enddatestr is not None:
                            b = enddatestr.split('T')[0]
                            revenddate = b.split('-')[2]+"-"+ b.split('-')[1]+"-"+ b.split('-')[0]
                            enddatestring = revenddate
                        else:
                            b = str(newcurrenttime).split('T')[0]
                            revenddate = b.split('-')[2]+"-"+ b.split('-')[1]+"-"+ b.split('-')[0]
                            enddatestring = revenddate

                        # if task played and closed on same date
                        if startdatestring == enddatestring:

                            starttime=startstring
                            t1 = datetime.datetime.strptime(starttime, "%Y-%m-%dT%H:%M:%S.%f%z")

                            endstring = s['EndDate']
                            if endstring is not None:
                                endtime = s['EndDate']
                            else:
                                endtime = str(newcurrenttime)
                            t2=datetime.datetime.strptime(endtime, "%Y-%m-%dT%H:%M:%S.%f%z")
                            dtdelta=t2-t1
                        
                        
                            if "day" in str(dtdelta) or "days" in str(dtdelta):
                                daystring = str(dtdelta).split(",")[0]
                                noofdays = str(daystring).split(" ")[0]
                                daysmins = int(noofdays)*1440

                                thoursstr =  str(dtdelta).split(",")[1]
                                thours = str(thoursstr).split(":")[0]
                                hrs = int(thours)*60
                                tmins = str(thoursstr).split(":")[1]
                                dfinalmins = int(hrs)+int(tmins)+int(daysmins)
                            else:
                                dthours = str(dtdelta).split(":")[0]
                                dhrs = int(dthours)*60
                                dtmins = str(dtdelta).split(":")[1]
                                dfinalmins = int(dhrs)+int(dtmins)

                            timeport ={
                                "date":startdatestring,
                                "time":dfinalmins
                            }
                            protimelist.append(timeport)
                        #if task played and closed on different dates
                        else:
                            
                            # for startdate till 12 o'clock calculation
                            starttime=startstring
                            t1 = datetime.datetime.strptime(starttime, "%Y-%m-%dT%H:%M:%S.%f%z")

                            startdateendtime = str(a)+"T"+"23:59:00.00000+05:30"
                            strtdatetym = str(startdateendtime)
                            t2=datetime.datetime.strptime(strtdatetym, "%Y-%m-%dT%H:%M:%S.%f%z")
                            sdtdelta=t2-t1
                        
                        
                            if "day" in str(sdtdelta) or "days" in str(sdtdelta):
                                daystring = str(sdtdelta).split(",")[0]
                                noofdays = str(daystring).split(" ")[0]
                                daysmins = int(noofdays)*1440

                                thoursstr =  str(sdtdelta).split(",")[1]
                                thours = str(thoursstr).split(":")[0]
                                hrs = int(thours)*60
                                tmins = str(thoursstr).split(":")[1]
                                frstdayfinalmins = int(hrs)+int(tmins)+int(daysmins)
                            else:
                                dthours = str(sdtdelta).split(":")[0]
                                dhrs = int(dthours)*60
                                dtmins = str(sdtdelta).split(":")[1]
                                frstdayfinalmins = int(dhrs)+int(dtmins)

                            secondtimeport ={
                                "date":startdatestring,
                                "time":frstdayfinalmins
                            }
                            seconddaylist.append(secondtimeport)
                        #calculation of second day

                            seconddaystrt = str(b)+"T"+"00:00:00.00000+05:30"
                            seconddaystrtdatetym = str(seconddaystrt)
                            t11=datetime.datetime.strptime(seconddaystrtdatetym, "%Y-%m-%dT%H:%M:%S.%f%z")

                            endstring = s['EndDate']
                            if endstring is not None:
                                endtime = s['EndDate']
                            else:
                                endtime = str(newcurrenttime)

                            seconddayenddt = endtime
                            t12 = datetime.datetime.strptime(seconddayenddt, "%Y-%m-%dT%H:%M:%S.%f%z")

                            seconddaytdelta=t12-t11

                            if "day" in str(seconddaytdelta) or "days" in str(seconddaytdelta):
                                daystring = str(seconddaytdelta).split(",")[0]
                                noofdays = str(daystring).split(" ")[0]
                                daysmins = int(noofdays)*1440

                                thoursstr =  str(seconddaytdelta).split(",")[1]
                                thours = str(thoursstr).split(":")[0]
                                hrs = int(thours)*60
                                tmins = str(thoursstr).split(":")[1]
                                secondayfinalmins = int(hrs)+int(tmins)+int(daysmins)
                            else:
                                dthours = str(seconddaytdelta).split(":")[0]
                                dhrs = int(dthours)*60
                                dtmins = str(seconddaytdelta).split(":")[1]
                                secondayfinalmins = int(dhrs)+int(dtmins)

                            secondtimeport ={
                                "date":enddatestring,
                                "time":secondayfinalmins
                            }
                            seconddaylist.append(secondtimeport)

                        
                        finallist = protimelist+seconddaylist


                    # finaldatetimelist= reduce(lambda d1,d2: {k: d1.get(k,0)+d2.get(k,0)for k in set(d1)|set(d2)}, timelist)
                
                    maindata = []
                    for p in finallist:
                    
                        v=stock_maindictlist("date", p['date'],"time",p['time'], maindata)
                        if v == {}:
                            timecalc={
                                "date":p['date'],
                                "time":p['time'],
                            }
                            maindata.append(timecalc)
                    
                    for m in maindata:
                        totalminutes = m['time']
                        hour = int (totalminutes)//60
                        if (len(str(hour)) < 2):
                            hours = "0"+str(hour)
                        else:
                            hours = str(hour)

                        mins = int (totalminutes) % 60
                        if (len(str(mins)) < 2):
                            minutes = "0"+str(mins)
                        else:
                            minutes = str(mins)

                        hourstring = str(hours)+" Hrs "+str(minutes)+" mins"

                        m['time'] = hourstring

                    i['daywisetime']=maindata

                    hovertaskidlist = []
                    if taskobj.ParentTaskId is not None:
                        hovertaskparentid = taskobj.ParentTaskId
                        hovertaskidlist.append(hovertaskparentid)
                        ttaskobject = TaskMaster.objects.filter(ParentTaskId=hovertaskparentid).exclude(id=i['id'])
                        ttaskser = PostTaskMasterSerializer(ttaskobject,many=True)
                        for t in ttaskser.data:
                            hovertaskidlist.append(t['id'])

                        hoverobjects = TaskMaster.objects.filter(id__in=hovertaskidlist)
                        hovertaskser = PostTaskMasterSerializer(hoverobjects,many=True)
                        for h in hovertaskser.data:

                            userObject=Users.objects.filter(id=h['AssignTo']).first()
                            h['CreatedBy']=userObject.Firstname+" "+userObject.Lastname

                            chdate = str(h['AssignDate'])
                            chmonth_name = calendar.month_abbr[int(chdate.split('-')[1])] 
                            h['Changeddate'] = chdate.split('-')[2] +" "+chmonth_name+" "+chdate.split('-')[0]
                           

                            hprojecttasktime = ProjectTasks.objects.filter(Task__in=taskidlist).order_by("id")
                            if hprojecttasktime :
                                hprojectser = ProjectTasksSerializer(hprojecttasktime,many=True)
                                FMT = '%H:%M:%S.%f'
                                htotaltime=0
                                for o in hprojectser.data:
                                    startstring = o['StartDate']
                                    starttime=startstring
                                    t1 = datetime.datetime.strptime(starttime, "%Y-%m-%dT%H:%M:%S.%f%z")

                                    endstring = o['EndDate']
                                    if endstring is not None:
                                        endtime = o['EndDate']
                                    else:
                                        endtime = str(newcurrenttime)
                                    t2=datetime.datetime.strptime(endtime, "%Y-%m-%dT%H:%M:%S.%f%z")
                                    tdelta=t2-t1
                                
                                
                                    if "day" in str(tdelta) or "days" in str(tdelta):
                                        daystring = str(tdelta).split(",")[0]
                                        noofdays = str(daystring).split(" ")[0]
                                        daysmins = int(noofdays)*1440

                                        thoursstr =  str(tdelta).split(",")[1]
                                        thours = str(thoursstr).split(":")[0]
                                        hrs = int(thours)*60
                                        tmins = str(thoursstr).split(":")[1]
                                        hfinalmins = int(hrs)+int(tmins)+int(daysmins)
                                    else:
                                        thours = str(tdelta).split(":")[0]
                                        hrs = int(thours)*60
                                        tmins = str(tdelta).split(":")[1]
                                        hfinalmins = int(hrs)+int(tmins)
                                    htotaltime += hfinalmins

                                htotalhours =htotaltime
                                hhour = int (htotalhours) // 60
                                if (len(str(hhour)) < 2):
                                    rhhours = "0"+str(hhour)
                                else:
                                    rhhours = str(hhour)

                                hmins = int (htotalhours) % 60
                                if (len(str(hmins)) < 2):
                                    rminutes = "0"+str(hmins)
                                else:
                                    rminutes = str(hmins)

                                rhourstring = str(rhhours)+":"+str(rminutes)+" "+"Hrs"

                                h['taskhoursstr'] = rhourstring

                            remarkzone = h['Zone']
                            if remarkzone == 1:
                                h['remarkstr'] = "<img src='/static/Media/taskicons/activegreenpoints.svg' class='activeicons' alt='green'>"
                            elif remarkzone == 2:
                                h['remarkstr'] = "<img src='/static/Media/taskicons/activeyellowpoints.svg' class='activeicons' alt='yellow'>"
                            elif remarkzone == 3:
                                h['remarkstr'] = "<img src='/static/Media/taskicons/activeredpoints.svg' class='activeicons' alt='red'>"
                            elif remarkzone == 4:
                                h['remarkstr'] = "<img src='/static/Media/taskicons/activenotdonepoints.svg' class='activeicons' alt='notdone'>"
                            elif remarkzone == 5:
                                h['remarkstr'] = "<img src='/static/Media/taskicons/activecancelledpoints.svg' class='activeicons' alt='notdone'>"
                            elif remarkzone == 6:
                                h['remarkstr'] = "<img src='/static/Media/taskicons/activerejectpoints.svg' class='activeicons' alt='notdone'>"
                            else:
                                h['remarkstr'] = "----"


                            bonusstr = h['Bonus']
                            if bonusstr == True:
                                h['bonusstr'] = "<img src='/static/Media/taskicons/activebonuspoints.svg' alt='bonus'>"
                            else:
                                h['bonusstr'] = ""

                        i['Hovertaskdata']  = hovertaskser.data
                    else:
                        i['Hovertaskdata'] = [] 
                    if i['ParentTaskId'] is None:
                        i['ParentTaskId'] = ""

                projecttotalminutes = projecthours
                phour = int (projecttotalminutes) // 60
                if (len(str(phour)) < 2):
                    phours = "0"+str(phour)
                else:
                    phours = str(phour)

                pmins = int (projecttotalminutes) % 60
                if (len(str(pmins)) < 2):
                    pminutes = "0"+str(pmins)
                else:
                    pminutes = str(pmins)

                phourstring = str(phours)+":"+str(pminutes)+" "+"Hrs"

                projecttotalhours = phourstring
                flmandays = phour/9
                mandays = round(flmandays,1)

                context ={
                "fisrttaskdetail":firsttaskser.data,
                "teammatecount" : taskteamcount,
                "totaltaskcount" : totaltaskcount,
                "lasttaskdetail":lasttaskser.data,
                "teamlist":teamlist,
                "projectstartdate":projectstart_date,
                "projectend_date":projectend_date,
                "projecttotalhours":projecttotalhours,
                "mandays":mandays,
                "uniqueYear":uniqueYear,
                "Taskdata":ser.data
                }
                
                return self.get_paginated_response(context)
            return Response({'data':{"projecttotalhours":'00:00 Hrs',"count":0},'n':0,"status":"successs","message":"Data not found"})
        return Response({'n':0,"status":"failed","message":"Please provide project id"})




class projectreporttotaltasktime(GenericAPIView):


    def post(self,request):
        projectId = request.POST.get("projectName")
        departname = request.POST.get("deptname")
        if departname is not None and departname != "":
            departname = [int(departname)]
        else:
            departname = []
        employeename = request.POST.get("searchname")
        if projectId is not None and projectId != "":


            taskObject=TaskMaster.objects.filter(Project=projectId,AssignTo__in =  list( Users.objects.filter(Q(Firstname__icontains=employeename)|Q(Lastname__icontains=employeename)).values_list('id', flat=True))).order_by('CreatedOn')

            firstobject = ""
            if departname is not None and departname != []:
                firstobject = taskObject.filter(AssignTo__in = list( Users.objects.filter(DepartmentID__in=departname).values_list('id', flat=True))).order_by('CreatedOn')
            else:
                firstobject = taskObject

            finalobject = firstobject
            
            if finalobject.exists():
                ser=PostTaskMasterSerializer(finalobject,many=True)
                projecthours=0

                
                for i in ser.data:

                    currentzone = pytz.timezone("Asia/Kolkata") 
                    currenttime = datetime.datetime.now(currentzone)
                    newcurrenttime =currenttime.strftime("%Y-%m-%dT%H:%M:%S.%f%z")

                    taskobj = TaskMaster.objects.filter(id=i['id']).first()
                    taskidlist = []
                    if taskobj.ParentTaskId is not None:
                        parenttaskid  = taskobj.ParentTaskId
                        taskidlist.append(parenttaskid)
                        taskobject = TaskMaster.objects.filter(ParentTaskId=parenttaskid)
                        taskser = PostTaskMasterSerializer(taskobject,many=True)
                        for t in taskser.data:
                            taskidlist.append(t['id'])
                    else:
                        taskidlist.append(i['id'])
                        taskobject = TaskMaster.objects.filter(ParentTaskId=i['id'])
                        taskser = PostTaskMasterSerializer(taskobject,many=True)
                        for t in taskser.data:
                            taskidlist.append(t['id'])


                    projecttasktime = ProjectTasks.objects.filter(Task_id__in=taskidlist).order_by("id")
                    if projecttasktime is not None:
                        projectser = ProjectTasksSerializer(projecttasktime,many=True)
                        totaltime=0
                        for o in projectser.data:
                            startstring = o['StartDate']
                            starttime=startstring
                            t1 = datetime.datetime.strptime(starttime, "%Y-%m-%dT%H:%M:%S.%f%z")

                            endstring = o['EndDate']
                            if endstring is not None:
                                endtime = o['EndDate']
                            else:
                                endtime = str(newcurrenttime)
                            t2=datetime.datetime.strptime(endtime, "%Y-%m-%dT%H:%M:%S.%f%z")
                            tdelta=t2-t1
                        
                        
                            if "day" in str(tdelta) or "days" in str(tdelta):
                                daystring = str(tdelta).split(",")[0]
                                noofdays = str(daystring).split(" ")[0]
                                daysmins = int(noofdays)*1440

                                thoursstr =  str(tdelta).split(",")[1]
                                thours = str(thoursstr).split(":")[0]
                                hrs = int(thours)*60
                                tmins = str(thoursstr).split(":")[1]
                                finalmins = int(hrs)+int(tmins)+int(daysmins)
                            else:
                                thours = str(tdelta).split(":")[0]
                                hrs = int(thours)*60
                                tmins = str(tdelta).split(":")[1]
                                finalmins = int(hrs)+int(tmins)
                            totaltime += finalmins
                        projecthours += totaltime

                projecttotalminutes = projecthours
                phour = int (projecttotalminutes) // 60
                if (len(str(phour)) < 2):
                    phours = "0"+str(phour)
                else:
                    phours = str(phour)

                pmins = int (projecttotalminutes) % 60
                if (len(str(pmins)) < 2):
                    pminutes = "0"+str(pmins)
                else:
                    pminutes = str(pmins)

                phourstring = str(phours)+":"+str(pminutes)+" "+"Hrs"

                projecttotalhours = phourstring


                context ={

                    "projecttotalhours":projecttotalhours,
                    "count":len(ser.data),

                }
                
                
                return Response({'data':context,'n':1,"status":"successs","message":"Data  found"})
            return Response({'data':{"projecttotalhours":'00:00 Hrs',"count":0},'n':1,"status":"successs","message":"Data  found"})
        return Response({'n':0,"status":"failed","message":"Please provide project id"})






class excelprojectreport(GenericAPIView):
    def post(self,request):
        repeatYear=[]
        uniqueYear=[] 
        teamlist=[]
        projectId = request.POST.get("projectName")
        departname = request.POST.get("deptname")
   
        if departname is not None and departname != "":
            departname = [int(departname)]
        else:
            departname = []
        employeename = request.POST.get("searchname")
        pagenumber = request.data.get('p')
        if projectId is not None and projectId != "":
            projectname = ProjectMaster.objects.filter(id=projectId).first()

            firsttaaskobject = TaskMaster.objects.filter(Project=projectId).order_by("id").first()
            firsttaskser = GetTaskSerializer(firsttaaskobject)
            if firsttaaskobject is not None:
                date = firsttaaskobject.AssignDate
                strdate = str(date)
                startmonth_name = calendar.month_abbr[int(strdate.split('-')[1])]  
                projectstart_date = strdate.split('-')[2] +" "+ startmonth_name +" "+strdate.split('-')[0]
            else:
                projectstart_date="--"


            lasttaskobject = TaskMaster.objects.filter(Project=projectId).order_by("-id").first()
            lasttaskser =  PostTaskMasterSerializer(lasttaskobject)
            if lasttaskobject is not None:
                date = lasttaskobject.AssignDate
                enddate = str(date)
                endmonth_name = calendar.month_abbr[int(enddate.split('-')[1])]  
                projectend_date = enddate.split('-')[2] +" "+endmonth_name+" "+enddate.split('-')[0]
            else:
                projectend_date="--"

            taskteamcount = TaskMaster.objects.filter(Project=projectId).distinct('AssignTo_id').count()

            taskteamstr = TaskMaster.objects.filter(Project=projectId).distinct('AssignTo_id')
            if taskteamstr is not None:
                teaskteamser =  GetTaskSerializer(taskteamstr,many=True)
                for t in teaskteamser.data:
                    userObject=Users.objects.filter(id=t['AssignTo']).first()
                    teammate = userObject.Firstname+" "+userObject.Lastname
                    teamlist.append(teammate)

            firsttotaltaskcount = TaskMaster.objects.filter(Project=projectId,AssignTo__in =  list( Users.objects.filter(Q(Firstname__icontains=employeename)|Q(Lastname__icontains=employeename)).values_list('id', flat=True)))

            totaltaskcountobject = ""

            if departname is not None and departname != []:
                totaltaskcountobject = firsttotaltaskcount.filter(AssignTo__in = list( Users.objects.filter(DepartmentID__in=departname).values_list('id', flat=True))).count()
            else:
                totaltaskcountobject = firsttotaltaskcount.count()


            totaltaskcount=totaltaskcountobject

            taskObject=TaskMaster.objects.filter(Project=projectId,AssignTo__in =  list( Users.objects.filter(Q(Firstname__icontains=employeename)|Q(Lastname__icontains=employeename)).values_list('id', flat=True))).order_by('CreatedOn')

            firstobject = ""
            if departname is not None and departname != []:
                firstobject = taskObject.filter(AssignTo__in = list( Users.objects.filter(DepartmentID__in=departname).values_list('id', flat=True))).order_by('CreatedOn')
             

            else:
                firstobject = taskObject

            finalobject = firstobject
            
            if finalobject.exists():          
                ser=PostTaskMasterSerializer(finalobject,many=True)
                projecthours=0
                for i in ser.data:
                    endtasktime = TaskMaster.objects.filter(id=i['id']).first()
                    if endtasktime is not None:
                        if endtasktime.IsParent == True:
                            childtaskobj = TaskMaster.objects.filter(ParentTaskId=i['id']).last()
                            childtaskid = childtaskobj.id
                            projectendtasktime = ProjectTasks.objects.filter(Task=childtaskid).order_by('-id').first()
                            endtaskstring=projectendtasktime.EndDate
                            if endtaskstring :
                                endtasktimestring = str(endtaskstring)
                                etime = endtasktimestring.split(' ')[0]
                                emonth_name = calendar.month_abbr[int(etime.split('-')[1])] 
                                userendtaskdate = etime.split('-')[2] +" "+emonth_name+" "+etime.split('-')[0]
                                i['endtaskdate']=userendtaskdate
                            else:
                                i['endtaskdate']="--"
                        elif endtasktime.IsParent == False:
                          
                            if endtasktime.ParentTaskId is not None:
                                PARENTID = endtasktime.ParentTaskId
                                childtaskobj = TaskMaster.objects.filter(ParentTaskId=int(PARENTID)).order_by('-id').first()
                                childtaskid = childtaskobj.id
                                projectendtasktime = ProjectTasks.objects.filter(Task=childtaskid).order_by('-id').first()
                                endtaskstring=projectendtasktime.EndDate
                                if endtaskstring :
                                    endtasktimestring = str(endtaskstring)
                                    etime = endtasktimestring.split(' ')[0]
                                    emonth_name = calendar.month_abbr[int(etime.split('-')[1])] 
                                    userendtaskdate = etime.split('-')[2] +" "+emonth_name+" "+etime.split('-')[0]
                                    i['endtaskdate']=userendtaskdate
                                else:
                                    i['endtaskdate']="--"
                            else:
                                projectendtasktime = ProjectTasks.objects.filter(Task=i['id']).order_by('-id').first()
                                if projectendtasktime is not None:
                                    endtaskstring=projectendtasktime.EndDate
                                    if endtaskstring :
                                        endtasktimestring = str(endtaskstring)
                                        etime = endtasktimestring.split(' ')[0]
                                        emonth_name = calendar.month_abbr[int(etime.split('-')[1])] 
                                        userendtaskdate = etime.split('-')[2] +" "+emonth_name+" "+etime.split('-')[0]
                                        i['endtaskdate']=userendtaskdate
                                    else:
                                        i['endtaskdate']="--"
                                else:
                                    i['endtaskdate']="--"

                    strdate = str(i['AssignDate'])
                    smonth_name = calendar.month_abbr[int(strdate.split('-')[1])]
                    newdate = strdate.split('-')[2] +" "+smonth_name+" "+strdate.split('-')[0]
                    i['AssignDate']=newdate


                    userObject=Users.objects.filter(id=i['AssignTo']).first()
                    i['CreatedBy']=userObject.Firstname+" "+userObject.Lastname
                    repeatYear.append(i['Year'])

                    if i['Zone'] == 1:
                        i['grade'] = "<span data-bs-toggle='tooltip' data-bs-placement='bottom' title='Green'><img src='/static/Media/taskicons/activegreenpoints.svg' class='activeicons' alt='Paris'></span>"

                    if i['Zone'] == 2:
                        i['grade'] = "<span data-bs-toggle='tooltip' data-bs-placement='bottom' title='Yellow'><img src='/static/Media/taskicons/activeyellowpoints.svg' class='activeicons' alt='Paris'></span>"

                    if i['Zone'] == 3:
                        i['grade'] = "<span data-bs-toggle='tooltip' data-bs-placement='bottom' title='Red'><img src='/static/Media/taskicons/activeredpoints.svg' class='activeicons' alt='Paris' ></span>"

                    if i['Zone'] == 4:
                        i['grade'] = "<span data-bs-toggle='tooltip' data-bs-placement='bottom' title='Not Done'><img src='/static/Media/taskicons/activenotdonepoints.svg' class='activeicons' alt='Paris'></span>"

                    if i['Zone'] == 5:
                        i['grade'] = "<span data-bs-toggle='tooltip' data-bs-placement='bottom' title='Cancelled'><img src='/static/Media/taskicons/activecancelledpoints.svg' class='activeicons' alt='Paris'></span>"

                    if i['Zone'] == 6:
                        i['grade'] = "<span data-bs-toggle='tooltip' data-bs-placement='bottom' title='Rejected'><img src='/static/Media/taskicons/activerejectpoints.svg' class='activeicons' alt='Paris'></span>"

                    if i['Zone'] == "" or i['Zone'] is None:
                        i['grade'] = "--"

                    currentzone = pytz.timezone("Asia/Kolkata") 
                    currenttime = datetime.datetime.now(currentzone)
                    newcurrenttime =currenttime.strftime("%Y-%m-%dT%H:%M:%S.%f%z")

                    taskobj = TaskMaster.objects.filter(id=i['id']).first()
                    taskidlist = []
                    if taskobj.ParentTaskId is not None:
                        parenttaskid  = taskobj.ParentTaskId
                        taskidlist.append(parenttaskid)
                        taskobject = TaskMaster.objects.filter(ParentTaskId=parenttaskid)
                        taskser = PostTaskMasterSerializer(taskobject,many=True)
                        for t in taskser.data:
                            taskidlist.append(t['id'])
                    else:
                        taskidlist.append(i['id'])
                        taskobject = TaskMaster.objects.filter(ParentTaskId=i['id'])
                        taskser = PostTaskMasterSerializer(taskobject,many=True)
                        for t in taskser.data:
                            taskidlist.append(t['id'])


                    projecttasktime = ProjectTasks.objects.filter(Task_id__in=taskidlist).order_by("id")
                    if projecttasktime is not None:
                        projectser = ProjectTasksSerializer(projecttasktime,many=True)
                        totaltime=0
                        for o in projectser.data:
                            startstring = o['StartDate']
                            starttime=startstring
                            t1 = datetime.datetime.strptime(starttime, "%Y-%m-%dT%H:%M:%S.%f%z")

                            endstring = o['EndDate']
                            if endstring is not None:
                                endtime = o['EndDate']
                            else:
                                endtime = str(newcurrenttime)
                            t2=datetime.datetime.strptime(endtime, "%Y-%m-%dT%H:%M:%S.%f%z")
                            tdelta=t2-t1
                        
                        
                            if "day" in str(tdelta) or "days" in str(tdelta):
                                daystring = str(tdelta).split(",")[0]
                                noofdays = str(daystring).split(" ")[0]
                                daysmins = int(noofdays)*1440

                                thoursstr =  str(tdelta).split(",")[1]
                                thours = str(thoursstr).split(":")[0]
                                hrs = int(thours)*60
                                tmins = str(thoursstr).split(":")[1]
                                finalmins = int(hrs)+int(tmins)+int(daysmins)
                            else:
                                thours = str(tdelta).split(":")[0]
                                hrs = int(thours)*60
                                tmins = str(tdelta).split(":")[1]
                                finalmins = int(hrs)+int(tmins)
                            totaltime += finalmins
                        projecthours += totaltime

                        totalhours =totaltime
                        hour = int (totalhours) // 60
                        if (len(str(hour)) < 2):
                            hours = "0"+str(hour)
                        else:
                            hours = str(hour)

                        mins = int (totalhours) % 60
                        if (len(str(mins)) < 2):
                            minutes = "0"+str(mins)
                        else:
                            minutes = str(mins)

                        hourstring = str(hours)+":"+str(minutes)+" "+"Hrs"

                        i['taskhours'] = hourstring

                    #  calculate day wise time strings
                    protimelist = []
                    seconddaylist = []
                    finallist=[]
                
                    for s in projectser.data:
                        startstring = s['StartDate']
                        enddatestr = s['EndDate']

                        a = startstring.split('T')[0]
                        revdate = a.split('-')[2]+"-"+ a.split('-')[1]+"-"+ a.split('-')[0]
                        startdatestring = revdate

                        if enddatestr is not None:
                            b = enddatestr.split('T')[0]
                            revenddate = b.split('-')[2]+"-"+ b.split('-')[1]+"-"+ b.split('-')[0]
                            enddatestring = revenddate
                        else:
                            b = str(newcurrenttime).split('T')[0]
                            revenddate = b.split('-')[2]+"-"+ b.split('-')[1]+"-"+ b.split('-')[0]
                            enddatestring = revenddate

                        # if task played and closed on same date
                        if startdatestring == enddatestring:

                            starttime=startstring
                            t1 = datetime.datetime.strptime(starttime, "%Y-%m-%dT%H:%M:%S.%f%z")

                            endstring = s['EndDate']
                            if endstring is not None:
                                endtime = s['EndDate']
                            else:
                                endtime = str(newcurrenttime)
                            t2=datetime.datetime.strptime(endtime, "%Y-%m-%dT%H:%M:%S.%f%z")
                            dtdelta=t2-t1
                        
                        
                            if "day" in str(dtdelta) or "days" in str(dtdelta):
                                daystring = str(dtdelta).split(",")[0]
                                noofdays = str(daystring).split(" ")[0]
                                daysmins = int(noofdays)*1440

                                thoursstr =  str(dtdelta).split(",")[1]
                                thours = str(thoursstr).split(":")[0]
                                hrs = int(thours)*60
                                tmins = str(thoursstr).split(":")[1]
                                dfinalmins = int(hrs)+int(tmins)+int(daysmins)
                            else:
                                dthours = str(dtdelta).split(":")[0]
                                dhrs = int(dthours)*60
                                dtmins = str(dtdelta).split(":")[1]
                                dfinalmins = int(dhrs)+int(dtmins)

                            timeport ={
                                "date":startdatestring,
                                "time":dfinalmins
                            }
                            protimelist.append(timeport)
                        #if task played and closed on different dates
                        else:
                            
                            # for startdate till 12 o'clock calculation
                            starttime=startstring
                            t1 = datetime.datetime.strptime(starttime, "%Y-%m-%dT%H:%M:%S.%f%z")

                            startdateendtime = str(a)+"T"+"23:59:00.00000+05:30"
                            strtdatetym = str(startdateendtime)
                            t2=datetime.datetime.strptime(strtdatetym, "%Y-%m-%dT%H:%M:%S.%f%z")
                            sdtdelta=t2-t1
                        
                        
                            if "day" in str(sdtdelta) or "days" in str(sdtdelta):
                                daystring = str(sdtdelta).split(",")[0]
                                noofdays = str(daystring).split(" ")[0]
                                daysmins = int(noofdays)*1440

                                thoursstr =  str(sdtdelta).split(",")[1]
                                thours = str(thoursstr).split(":")[0]
                                hrs = int(thours)*60
                                tmins = str(thoursstr).split(":")[1]
                                frstdayfinalmins = int(hrs)+int(tmins)+int(daysmins)
                            else:
                                dthours = str(sdtdelta).split(":")[0]
                                dhrs = int(dthours)*60
                                dtmins = str(sdtdelta).split(":")[1]
                                frstdayfinalmins = int(dhrs)+int(dtmins)

                            secondtimeport ={
                                "date":startdatestring,
                                "time":frstdayfinalmins
                            }
                            seconddaylist.append(secondtimeport)
                        #calculation of second day

                            seconddaystrt = str(b)+"T"+"00:00:00.00000+05:30"
                            seconddaystrtdatetym = str(seconddaystrt)
                            t11=datetime.datetime.strptime(seconddaystrtdatetym, "%Y-%m-%dT%H:%M:%S.%f%z")

                            endstring = s['EndDate']
                            if endstring is not None:
                                endtime = s['EndDate']
                            else:
                                endtime = str(newcurrenttime)

                            seconddayenddt = endtime
                            t12 = datetime.datetime.strptime(seconddayenddt, "%Y-%m-%dT%H:%M:%S.%f%z")

                            seconddaytdelta=t12-t11

                            if "day" in str(seconddaytdelta) or "days" in str(seconddaytdelta):
                                daystring = str(seconddaytdelta).split(",")[0]
                                noofdays = str(daystring).split(" ")[0]
                                daysmins = int(noofdays)*1440

                                thoursstr =  str(seconddaytdelta).split(",")[1]
                                thours = str(thoursstr).split(":")[0]
                                hrs = int(thours)*60
                                tmins = str(thoursstr).split(":")[1]
                                secondayfinalmins = int(hrs)+int(tmins)+int(daysmins)
                            else:
                                dthours = str(seconddaytdelta).split(":")[0]
                                dhrs = int(dthours)*60
                                dtmins = str(seconddaytdelta).split(":")[1]
                                secondayfinalmins = int(dhrs)+int(dtmins)

                            secondtimeport ={
                                "date":enddatestring,
                                "time":secondayfinalmins
                            }
                            seconddaylist.append(secondtimeport)

                        
                        finallist = protimelist+seconddaylist


                    # finaldatetimelist= reduce(lambda d1,d2: {k: d1.get(k,0)+d2.get(k,0)for k in set(d1)|set(d2)}, timelist)
                
                    maindata = []
                    for p in finallist:
                    
                        v=stock_maindictlist("date", p['date'],"time",p['time'], maindata)
                        if v == {}:
                            timecalc={
                                "date":p['date'],
                                "time":p['time'],
                            }
                            maindata.append(timecalc)
                    
                    for m in maindata:
                        totalminutes = m['time']
                        hour = int (totalminutes)//60
                        if (len(str(hour)) < 2):
                            hours = "0"+str(hour)
                        else:
                            hours = str(hour)

                        mins = int (totalminutes) % 60
                        if (len(str(mins)) < 2):
                            minutes = "0"+str(mins)
                        else:
                            minutes = str(mins)

                        hourstring = str(hours)+" Hrs "+str(minutes)+" mins"

                        m['time'] = hourstring

                    i['daywisetime']=maindata

                    hovertaskidlist = []
                    if taskobj.ParentTaskId is not None:
                        hovertaskparentid = taskobj.ParentTaskId
                        hovertaskidlist.append(hovertaskparentid)
                        ttaskobject = TaskMaster.objects.filter(ParentTaskId=hovertaskparentid).exclude(id=i['id'])
                        ttaskser = PostTaskMasterSerializer(ttaskobject,many=True)
                        for t in ttaskser.data:
                            hovertaskidlist.append(t['id'])

                        hoverobjects = TaskMaster.objects.filter(id__in=hovertaskidlist)
                        hovertaskser = PostTaskMasterSerializer(hoverobjects,many=True)
                        for h in hovertaskser.data:

                            userObject=Users.objects.filter(id=h['AssignTo']).first()
                            h['CreatedBy']=userObject.Firstname+" "+userObject.Lastname

                            chdate = str(h['AssignDate'])
                            chmonth_name = calendar.month_abbr[int(chdate.split('-')[1])] 
                            h['Changeddate'] = chdate.split('-')[2] +" "+chmonth_name+" "+chdate.split('-')[0]
                           

                            hprojecttasktime = ProjectTasks.objects.filter(Task__in=taskidlist).order_by("id")
                            if hprojecttasktime :
                                hprojectser = ProjectTasksSerializer(hprojecttasktime,many=True)
                                FMT = '%H:%M:%S.%f'
                                htotaltime=0
                                for o in hprojectser.data:
                                    startstring = o['StartDate']
                                    starttime=startstring
                                    t1 = datetime.datetime.strptime(starttime, "%Y-%m-%dT%H:%M:%S.%f%z")

                                    endstring = o['EndDate']
                                    if endstring is not None:
                                        endtime = o['EndDate']
                                    else:
                                        endtime = str(newcurrenttime)
                                    t2=datetime.datetime.strptime(endtime, "%Y-%m-%dT%H:%M:%S.%f%z")
                                    tdelta=t2-t1
                                
                                
                                    if "day" in str(tdelta) or "days" in str(tdelta):
                                        daystring = str(tdelta).split(",")[0]
                                        noofdays = str(daystring).split(" ")[0]
                                        daysmins = int(noofdays)*1440

                                        thoursstr =  str(tdelta).split(",")[1]
                                        thours = str(thoursstr).split(":")[0]
                                        hrs = int(thours)*60
                                        tmins = str(thoursstr).split(":")[1]
                                        hfinalmins = int(hrs)+int(tmins)+int(daysmins)
                                    else:
                                        thours = str(tdelta).split(":")[0]
                                        hrs = int(thours)*60
                                        tmins = str(tdelta).split(":")[1]
                                        hfinalmins = int(hrs)+int(tmins)
                                    htotaltime += hfinalmins

                                htotalhours =htotaltime
                                hhour = int (htotalhours) // 60
                                if (len(str(hhour)) < 2):
                                    rhhours = "0"+str(hhour)
                                else:
                                    rhhours = str(hhour)

                                hmins = int (htotalhours) % 60
                                if (len(str(hmins)) < 2):
                                    rminutes = "0"+str(hmins)
                                else:
                                    rminutes = str(hmins)

                                rhourstring = str(rhhours)+":"+str(rminutes)+" "+"Hrs"

                                h['taskhoursstr'] = rhourstring

                            remarkzone = h['Zone']
                            if remarkzone == 1:
                                h['remarkstr'] = "<img src='/static/Media/taskicons/activegreenpoints.svg' class='activeicons' alt='green'>"
                            elif remarkzone == 2:
                                h['remarkstr'] = "<img src='/static/Media/taskicons/activeyellowpoints.svg' class='activeicons' alt='yellow'>"
                            elif remarkzone == 3:
                                h['remarkstr'] = "<img src='/static/Media/taskicons/activeredpoints.svg' class='activeicons' alt='red'>"
                            elif remarkzone == 4:
                                h['remarkstr'] = "<img src='/static/Media/taskicons/activenotdonepoints.svg' class='activeicons' alt='notdone'>"
                            elif remarkzone == 5:
                                h['remarkstr'] = "<img src='/static/Media/taskicons/activecancelledpoints.svg' class='activeicons' alt='notdone'>"
                            elif remarkzone == 6:
                                h['remarkstr'] = "<img src='/static/Media/taskicons/activerejectpoints.svg' class='activeicons' alt='notdone'>"
                            else:
                                h['remarkstr'] = "----"


                            bonusstr = h['Bonus']
                            if bonusstr == True:
                                h['bonusstr'] = "<img src='/static/Media/taskicons/activebonuspoints.svg' alt='bonus'>"
                            else:
                                h['bonusstr'] = ""

                        i['Hovertaskdata']  = hovertaskser.data
                    else:
                        i['Hovertaskdata'] = [] 
                    if i['ParentTaskId'] is None:
                        i['ParentTaskId'] = ""

                projecttotalminutes =projecthours
                phour = int (projecttotalminutes) // 60
                if (len(str(phour)) < 2):
                    phours = "0"+str(phour)
                else:
                    phours = str(phour)

                pmins = int (projecttotalminutes) % 60
                if (len(str(pmins)) < 2):
                    pminutes = "0"+str(pmins)
                else:
                    pminutes = str(pmins)

                phourstring = str(phours)+":"+str(pminutes)+" "+"Hrs"

                projecttotalhours = phourstring
                flmandays = phour/9
                mandays = round(flmandays,1)

                context ={
                "fisrttaskdetail":firsttaskser.data,
                "teammatecount" : taskteamcount,
                "totaltaskcount" : totaltaskcount,
                "lasttaskdetail":lasttaskser.data,
                "teamlist":teamlist,
                "projectstartdate":projectstart_date,
                "projectend_date":projectend_date,
                "projecttotalhours":projecttotalhours,
                "mandays":mandays,
                "uniqueYear":uniqueYear,
                "Taskdata":ser.data
                }

                if os.path.exists("static/report/excelprojectreport.xlsx"):
                    os.remove("static/report/excelprojectreport.xlsx")
                    workbook = xlsxwriter.Workbook('static/report/excelprojectreport.xlsx')
                    workbook.close()
                else:
                    workbook = xlsxwriter.Workbook('static/report/excelprojectreport.xlsx')
                    workbook.close()
                exportprojectreportdata(context)
                downloadurl = imageUrl + "/static/report/excelprojectreport.xlsx"
                
                return Response({'n':1,"status":"success","message":"Data found successdully","data":context,'downloadurl':downloadurl})
            return Response({'n':0,"status":"failed","message":"Data not found"})
        return Response({'n':0,"status":"failed","message":"Please provide project id"})


def exportprojectreportdata(data):

    workbook = xlsxwriter.Workbook('static/report/excelprojectreport.xlsx')
    wb=load_workbook('static/report/excelprojectreport.xlsx')
    sheet=wb.worksheets[0]
    
 
    sheet.cell(row=3,column=2).value="Project: "
    sheet.cell(row=3,column=3).value="Start Date: "
    sheet.cell(row=3,column=4).value='End Date: '
    sheet.cell(row=3,column=5).value='Project Total Hours: '
    sheet.cell(row=3,column=6).value='No of Tasks: '

    k = 7
    counter = 1
  
    sheet.cell(row=4,column=2).value=data['fisrttaskdetail']['ProjectName']
    sheet.cell(row=4,column=3).value=data['projectstartdate']
    sheet.cell(row=4,column=4).value=data['projectend_date']
    sheet.cell(row=4,column=5).value=data['projecttotalhours'] + " (" + str(data['mandays']) + ") Mandays"
    sheet.cell(row=4,column=6).value=data['totaltaskcount']
    sheet.cell(row=5,column=2).value="Task Details"
    sheet.cell(row=5,column=3).value="Task Start Date"
    sheet.cell(row=5,column=4).value="Task End Date"
    sheet.cell(row=5,column=5).value="Employee Name"
    sheet.cell(row=5,column=6).value="Hours"
    for p in data['Taskdata']:
        sheet.cell(row=k,column=2).value=p['TaskTitle']
        sheet.cell(row=k,column=3).value=p['AssignDate']
        sheet.cell(row=k,column=4).value=p['endtaskdate']
        sheet.cell(row=k,column=5).value=p['CreatedBy']
        sheet.cell(row=k,column=6).value=p['taskhours']
        k+=1
    s = k+1
    j = s+1
    
    
    counter+=1
    
    wb.save('static/report/excelprojectreport.xlsx')





@api_view(['POST'])
def searchproject(request):
    searchproject = request.POST.get("project")
    company_code = request.user.company_code
    if searchproject is not None and searchproject != "":
        projectobject = ProjectMaster.objects.filter(Active=True,company_code=company_code,ProjectName__icontains = searchproject)
        if projectobject is not None:
            projectser = ProjectSerializer(projectobject,many=True)
            return Response({'n':1,'Msg':'Project fetched successfully','Status':'Success','data':projectser.data})
        return Response({'n':0,'Msg':'Project not found','Status':'Failed','data':''})
    allprojectobject = ProjectMaster.objects.filter(Active=True,company_code=company_code,ProjectName__icontains = searchproject)
    allprojectsere = ProjectSerializer(allprojectobject,many=True)
    return Response({'n':1,'Msg':'Project fetched successfully','Status':'Success','data':allprojectsere.data})

@api_view(['GET'])
def queryprojectlist(request):
    company_code = request.user.company_code
    try:
        conn = psycopg2.connect(database=env('DATABASE_NAME'), user= env('DATABASE_USER'),
                                password=env('DATABASE_PASSWORD'), host=env('DATABASE_HOST'), port=env('DATABASE_PORT'), cursor_factory=RealDictCursor)
        cur = conn.cursor()
    except Exception as e:
        return Response({"n": 0, "Msg": "Could not connect to database", "Status": "Failed"})
    else:
        query = """ SELECT id,"ProjectName" FROM public."Project_projectmaster" where "Active" = 'true' and "company_code" = '"""+company_code+"""' order by "ProjectName" """
        cur.execute(query)
        projectdata = cur.fetchall()
        return Response({'n':1,'Msg':'Project fetched successfully','Status':'Success','data':projectdata})

@api_view(['GET'])
def querylocationlist(request):
    company_code = request.user.company_code
    try:
        conn = psycopg2.connect(database=env('DATABASE_NAME'), user= env('DATABASE_USER'),
                                password=env('DATABASE_PASSWORD'), host=env('DATABASE_HOST'), port=env('DATABASE_PORT'), cursor_factory=RealDictCursor)
        cur = conn.cursor()
    except Exception as e:
        return Response({"n": 0, "Msg": "Could not connect to database", "Status": "Failed"})
    else:
        query = """ Select id,"LocationName" from public."Users_location" 
                    where "Active" = 'true' and "company_code" = '"""+company_code+"""' order by "LocationName" """
        cur.execute(query)
        locationdata = cur.fetchall()
        return Response({'n':1,'Msg':'Project fetched successfully','Status':'Success','data':locationdata})

@api_view(['GET'])
def querydepartmentlist(request):
    company_code = request.user.company_code
    try:
        conn = psycopg2.connect(database=env('DATABASE_NAME'), user= env('DATABASE_USER'),
                                password=env('DATABASE_PASSWORD'), host=env('DATABASE_HOST'), port=env('DATABASE_PORT'), cursor_factory=RealDictCursor)
        cur = conn.cursor()
    except Exception as e:
        return Response({"n": 0, "Msg": "Could not connect to database", "Status": "Failed"})
    else:
        query = """ select id,"DepartmentName" from public."Department_department" 
                where "Active" = 'True' and "company_code" = '"""+company_code+"""' order by "DepartmentName" """
        cur.execute(query)
        departmentdata = cur.fetchall()
        return Response({'n':1,'Msg':'Project fetched successfully','Status':'Success','data':departmentdata})
    
@api_view(['GET'])
def queryprojectmanagerlist(request):
    company_code = request.user.company_code
    try:
        conn = psycopg2.connect(database=env('DATABASE_NAME'), user= env('DATABASE_USER'),
                                password=env('DATABASE_PASSWORD'), host=env('DATABASE_HOST'), port=env('DATABASE_PORT'), cursor_factory=RealDictCursor)
        cur = conn.cursor()
    except Exception as e:
        return Response({"n": 0, "Msg": "Could not connect to database", "Status": "Failed"})
    else:
        query = """ select distinct pm."ProjectBA_id",u."Firstname",u."Lastname" from public."Project_projectmaster" as pm
                inner join public."Users_users" as u on u.id = pm."ProjectBA_id"
                where pm."Active" = 'True' and pm."company_code" = '"""+company_code+"""' order by u."Firstname" """
        cur.execute(query)
        projectmanagerdata = cur.fetchall()
        return Response({'n':1,'Msg':'Project fetched successfully','Status':'Success','data':projectmanagerdata})

@api_view(['GET'])
def queryuserdashboardlist(request):
    company_code = request.user.company_code
    try:
        conn = psycopg2.connect(database=env('DATABASE_NAME'), user= env('DATABASE_USER'),
                                password=env('DATABASE_PASSWORD'), host=env('DATABASE_HOST'), port=env('DATABASE_PORT'), cursor_factory=RealDictCursor)
        cur = conn.cursor()
    except Exception as e:
        return Response({"n": 0, "Msg": "Could not connect to database", "Status": "Failed"})
    else:
        query = """ select distinct us."company_code",
                array (select concat('{ 
                                    "id":',pm."ProjectBA_id",',
                                    "Firstname":"',u."Firstname",'",
                                    "Lastname":"',u."Lastname",'"}') 
                    from public."Project_projectmaster" as pm
                        inner join public."Users_users" as u on u.id = pm."ProjectBA_id"
                        where pm."Active" = 'True' and pm."company_code" = '"""+company_code+"""' 
                    order by u."Firstname") as projectmanagerdata,
                array ( select concat('{ 
                                    "id":',pr.id,',
                                    "ProjectName":"',pr."ProjectName",'"}') 
                    FROM public."Project_projectmaster" as pr
                    where pr."Active" = 'true' and pr."company_code" = '"""+company_code+"""' 
                    order by "ProjectName"
                    ) as project_data,
                array ( select concat('{ 
                                    "id":',ul.id,',
                                    "LocationName":"',ul."LocationName",'"}') 
                    from public."Users_location" as ul
                    where ul."Active" = 'true' and ul."company_code" = '"""+company_code+"""'  order by ul."LocationName"
                    ) as location_data,
                array ( select concat('{ 
                                    "id":',ud.id,',
                                    "DepartmentName":"',ud."DepartmentName",'"}') 
                    from public."Department_department" as ud
                        where ud."Active" = 'True' and ud."company_code" = '"""+company_code+"""'  order by ud."DepartmentName"
                    ) as department_data,
                array ( select concat('{ 
                                    "manager_id":',tm."ManagerID_id",',
                                    "manager_name":"',tm."ManagerIDStr",'"}') 
                    from public."Users_usertomanager" as tm
                        where tm."Active" = 'True' and tm."company_code" = '"""+company_code+"""'  order by tm."ManagerIDStr"
                    ) as taskmanager_data,
                array ( select concat('{ 
                                    "user_id":',ct."users_id",',
                                    "Firstname":"',uc."Firstname",'",
                                    "Lastname":"',uc."Lastname",'"}') 
                    from public."ClientMaster_createclient_Team" as ct
                    inner join public."Users_users" as uc on uc.id = ct."users_id"
                        order by uc."Firstname"
                    ) as clientside_data 
                from public."Users_users" as us where us."company_code" ='"""+company_code+"""'  """
        cur.execute(query)
        alldata = cur.fetchall()
        project_manager_list = []
        task_manager_list = []
        project_data_list = []
        location_data_list = []
        dept_data_list = []
        client_data_list = []
        for i in alldata:
            for p in i['projectmanagerdata']:
                project_manager_list.append(json.loads(p))   

            i['projectmanagerdata'] = list({v['id']:v for v in project_manager_list}.values())

            for p in i['taskmanager_data']:
                task_manager_list.append(json.loads(p))   

            i['taskmanager_data'] = list({v['manager_id']:v for v in task_manager_list}.values())

            for p in i['project_data']:
                project_data_list.append(json.loads(p))   
            
            i['project_data'] = project_data_list

            for p in i['location_data']:
                location_data_list.append(json.loads(p))   
            
            i['location_data'] = location_data_list

            for p in i['department_data']:
                dept_data_list.append(json.loads(p))   
            
            i['department_data'] = dept_data_list

            for p in i['clientside_data']:
                client_data_list.append(json.loads(p))   
            
            i['clientside_data'] = client_data_list
        

        return Response({'n':1,'Msg':'Details fetched successfully','Status':'Success','data':alldata[0]})



















